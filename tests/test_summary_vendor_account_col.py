"""Master quarterly summary: vendor arrays carry their identity in Account #.

Ford (2026-07-24): the Locus site id rode in the NAME column ("Johnson Farm &
Garden (3669325)") while Account # sat empty. The identifier belongs in the
identifier column: NEPOOL GIS id when set, else the monitor's site id.
"""
from __future__ import annotations

import secrets
import tempfile
from datetime import date
from pathlib import Path

import openpyxl

from api.db import SessionLocal
from api.models import (
    Array,
    Client,
    DailyGeneration,
    InverterConnection,
    Tenant,
)
from api.writers.gmp_raw_writer import build_quarterly_summary_workbook


def _seed() -> tuple[str, int, int]:
    tid = "ten_" + secrets.token_hex(5)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Sum Co", contact_email=f"{tid}@ex.test",
            tenant_key="sol_live_" + secrets.token_urlsafe(8),
            plan="standard", active=True, product="array_operator",
            generation_reports=True,
        ))
        c = Client(tenant_id=tid, name="Cl", active=True)
        db.add(c)
        db.flush()
        a1 = Array(tenant_id=tid, client_id=c.id, name="Monitor Only",
                   fuel_type="solar")
        a2 = Array(tenant_id=tid, client_id=c.id, name="Monitor With GIS",
                   fuel_type="solar", nepool_gis_id="55555")
        db.add_all([a1, a2])
        db.flush()
        for arr, sid in ((a1, 3669325), (a2, 777)):
            db.add(InverterConnection(
                array_id=arr.id, vendor="locus",
                config={"username": "u", "password": "p", "site_id": sid},
                status="ok",
            ))
            db.add(DailyGeneration(
                tenant_id=tid, array_id=arr.id, day=date(2026, 5, 10),
                kwh=100.0, source="locus",
            ))
        ids = (a1.id, a2.id)
        db.commit()
    return tid, *ids


def test_vendor_arrays_get_site_or_gis_id_in_account_column():
    tid, _a1, _a2 = _seed()
    out = Path(tempfile.mkdtemp()) / "sum.xlsx"
    build_quarterly_summary_workbook(tid, out, year=2026, quarter=2)
    ws = openpyxl.load_workbook(out)["Summary"]
    rows = {ws.cell(r, 1).value: ws.cell(r, 2).value
            for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value}
    # Site id in the Account # column — NOT baked into the name.
    assert rows.get("Monitor Only") == "3669325"
    # A NEPOOL GIS id outranks the monitor's site id.
    assert rows.get("Monitor With GIS") == "55555"
    assert not any("(" in str(k) for k in rows)  # no identifier-in-name rows
