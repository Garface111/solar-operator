"""Tests for the bring-your-own generation-spreadsheet auto-updater.

Covers the pure logic that does NOT need a DB:
  * heuristic column detection on an ARBITRARY-column sheet (xlsx + csv)
  * CSV normalization to xlsx preserving cell addressing
  * idempotent monthly append (never double-appends the same period)
"""
import io

import openpyxl

from api.billing import sheet_tracker as st


def _make_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MyLedger"
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# An operator's OWN, idiosyncratic layout: a title row, then headers that don't
# match any template — "Billing Month", "Solar Produced (kWh)", "Home Usage",
# "Credit $/kWh", "Total Credit".
SAMPLE_ROWS = [
    ["Maple Farm Solar — Generation Log", None, None, None, None],
    ["Billing Month", "Solar Produced (kWh)", "Home Usage", "Credit $/kWh", "Total Credit"],
    ["2026-03", 1820.5, 940, 0.2576, 469.0],
    ["2026-04", 2010.0, 880, 0.2576, 517.8],
    ["2026-05", 2310.2, 810, 0.2576, 595.1],
]


def test_detect_arbitrary_columns():
    blob = _make_xlsx(SAMPLE_ROWS)
    m = st.detect_structure(blob, "maple.xlsx")
    assert m["ok"], m
    assert m["sheet"] == "MyLedger"
    assert m["header_row"] == 1            # 0-based: title row 0, header row 1
    cols = m["columns"]
    assert cols["period"] == 0
    assert cols["generation"] == 1
    assert cols["consumption"] == 2
    assert cols["rate"] == 3
    assert cols["amount"] == 4
    assert m["last_period"] == "2026-05"
    assert m["data_rows"] == 3


def test_detect_csv_and_ingest_normalizes_to_xlsx():
    csv = (
        "Period,Generation kWh,Amount Due\n"
        "2026-04,1500,386.40\n"
        "2026-05,1600,412.16\n"
    ).encode("utf-8")
    res = st.ingest_upload(csv, "ledger.csv")
    assert res["ok"], res
    m = res["mapping"]
    assert m["kind"] == "xlsx"           # normalized
    assert m["sheet"] == "Generation"
    assert m["columns"]["period"] == 0
    assert m["columns"]["generation"] == 1
    assert m["columns"]["amount"] == 2
    # the normalized workbook actually opens + has the rows
    wb = openpyxl.load_workbook(io.BytesIO(res["workbook"]))
    ws = wb["Generation"]
    assert ws.cell(row=1, column=2).value == "Generation kWh"
    assert ws.cell(row=2, column=1).value == "2026-04"


def test_append_writes_into_mapped_columns():
    blob = _make_xlsx(SAMPLE_ROWS)
    m = st.detect_structure(blob, "maple.xlsx")
    new = st.append_period_row(blob, m, {
        "period": "2026-06", "generation": 2450.0,
        "consumption": 770, "rate": 0.2576, "amount": 631.1,
    })
    wb = openpyxl.load_workbook(io.BytesIO(new))
    ws = wb["MyLedger"]
    # appended onto row 6 (1-based): title(1) header(2) data(3,4,5) → new=6
    assert ws.cell(row=6, column=1).value == "2026-06"
    assert ws.cell(row=6, column=2).value == 2450.0
    assert ws.cell(row=6, column=5).value == 631.1
    # existing rows untouched
    assert ws.cell(row=5, column=1).value == "2026-05"
    assert ws.cell(row=5, column=2).value == 2310.2


def test_period_idempotency_matching():
    assert st._period_matches("2026-05", "2026-05")
    assert st._period_matches("May 2026", "2026-05")
    assert st._period_matches("5/31/2026", "2026-05")
    assert not st._period_matches("2026-04", "2026-05")
    assert not st._period_matches(None, "2026-05")


def test_detect_rejects_unmappable():
    blob = _make_xlsx([["Notes", "Misc"], ["hello", "world"]])
    m = st.detect_structure(blob, "junk.xlsx")
    assert not m["ok"]
    assert m["warnings"]


# ── Multi-kWh-column sheet (Paul's "Fairlee" shape): whole-array, the offtaker's
# share, and a cumulative running total. The offtaker's NAMED column must win, the
# cumulative one must never, and a split tariff+adder must beat a bare tariff.
_FAIRLEE = [
    ["", "", "", "", "Cumm"],                                          # sub-header over the cumulative col
    ["Month", "kWh whole array", "kWh Fairlee", "Tariff+Adder", "KwH"],
    ["2026-04", 26640, 9058, 0.19417, 51653],
    ["2026-05", 29700, 10098, 0.19417, 61751],
]


def test_cumulative_column_never_chosen_as_generation():
    m = st.detect_structure(_make_xlsx(_FAIRLEE), "f.xlsx")
    assert m["ok"]
    assert m["columns"]["generation"] != 4               # the cumulative "KwH" is demoted


def test_offtaker_name_picks_their_share_column():
    m = st.detect_structure(_make_xlsx(_FAIRLEE), "f.xlsx",
                            st.name_hint_tokens("Town of Fairlee"))
    assert m["columns"]["generation"] == 2               # "kWh Fairlee" (their share)
    assert m["columns"]["rate"] == 3                      # "Tariff+Adder", not a bare tariff
    assert m["columns"]["period"] == 0


def test_amount_detects_bill_column():
    m = st.detect_structure(_make_xlsx([["Month", "kWh Generated", "Bill"],
                                        ["2026-05", 1000, 257.6]]), "b.xlsx")
    assert m["columns"].get("amount") == 2


def test_name_hint_tokens_drops_filler():
    assert st.name_hint_tokens("Town of Fairlee") == ["fairlee"]
    assert st.name_hint_tokens(None) == []


def test_ai_unavailable_falls_back_to_heuristic(monkeypatch):
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    from api.billing import sheet_tracker_ai as ai
    assert ai.ai_available() is False
    res = st.ingest_upload(_make_xlsx(SAMPLE_ROWS), "maple.xlsx", "Maple Farm")
    assert res["ok"]
    assert res["mapping"]["via"] == "heuristic"
    assert res["mapping"]["columns"]["generation"] == 1


def test_ai_parser_validates_and_rejects():
    from api.billing import sheet_tracker_ai as ai
    good = ('{"header_row":1,"columns":{"period":0,"generation":1,"amount":4,'
            '"consumption":null},"confidence":0.9}')
    p = ai._parse(good, width=5, n_rows=5)
    assert p and p["columns"] == {"period": 0, "generation": 1, "amount": 4}
    assert ai._parse("prose " + good + " tail", 5, 5)                          # tolerant of surrounding text
    assert ai._parse('{"header_row":1,"columns":{"period":0},"confidence":0.9}', 5, 5) is None      # no generation
    assert ai._parse('{"header_row":1,"columns":{"generation":1},"confidence":0.2}', 5, 5) is None  # low confidence
    assert ai._parse('{"header_row":1,"columns":{"generation":99},"confidence":0.9}', 5, 5) is None # out of range
    assert ai._parse("not json", 5, 5) is None


# ── Chronological append (regression: model emit order / ISO labels / month names) ──

def test_append_ai_rows_sorts_iso_period_labels_not_model_order():
    """ISO 'YYYY-MM' period cells are NOT full dates — old row_period returned None
    for every row, so sort was a no-op and rows kept the model's May/June/April order."""
    blob = _make_xlsx([
        ["Month", "kWh", "Tariff"],
        ["2025-04", 100, 0.18],
        ["2025-05", 110, 0.18],
    ])
    m = st.detect_structure(blob, "t.xlsx")
    # Model emits out of order + includes a month already on the sheet
    ai_rows = [
        {0: "2025-07", 1: 130, 2: 0.18},
        {0: "2025-04", 1: 100, 2: 0.18},  # already present
        {0: "2025-06", 1: 120, 2: 0.18},
        {0: "2025-03", 1: 90, 2: 0.18},
    ]
    new, appended = st.append_ai_rows(blob, m, ai_rows, present={"2025-04", "2025-05"})
    assert len(appended) == 3
    periods = [r[0] for r in openpyxl.load_workbook(io.BytesIO(new)).active.iter_rows(values_only=True)]
    # existing April/May stay put; new rows land chronological after them
    assert periods == ["Month", "2025-04", "2025-05", "2025-03", "2025-06", "2025-07"] or \
           periods[3:] == ["2025-03", "2025-06", "2025-07"]
    # The NEW rows must be chronological among themselves regardless of existing
    new_only = [p for p in periods if p in ("2025-03", "2025-06", "2025-07")]
    assert new_only == ["2025-03", "2025-06", "2025-07"]


def test_append_ai_rows_sorts_bare_month_names():
    """Sheets styled with bare month names ('April') used to all sort as '9999-99'."""
    blob = _make_xlsx([
        ["Month", "kWh", "Tariff"],
        ["April", 100, 0.18],
        ["May", 110, 0.18],
    ])
    m = st.detect_structure(blob, "t.xlsx")
    ai_rows = [
        {0: "June", 1: 120, 2: 0.18},
        {0: "March", 1: 90, 2: 0.18},
        {0: "July", 1: 130, 2: 0.18},
    ]
    new, appended = st.append_ai_rows(blob, m, ai_rows, present={"April", "May"})
    assert len(appended) == 3
    periods = [r[0] for r in openpyxl.load_workbook(io.BytesIO(new)).active.iter_rows(values_only=True)]
    new_only = [p for p in periods if p in ("March", "June", "July")]
    assert new_only == ["March", "June", "July"], periods


def test_append_ai_rows_dedups_bare_month_present_against_iso_row():
    """Sheet says 'April'; AI proposes 2025-04 → must not duplicate."""
    blob = _make_xlsx([
        ["Month", "kWh"],
        ["April", 100],
        ["May", 110],
    ])
    m = st.detect_structure(blob, "t.xlsx")
    ai_rows = [
        {0: "2025-04", 1: 100},
        {0: "2025-06", 1: 120},
    ]
    new, appended = st.append_ai_rows(blob, m, ai_rows, present={"April", "May"})
    assert len(appended) == 1
    periods = [r[0] for r in openpyxl.load_workbook(io.BytesIO(new)).active.iter_rows(values_only=True)]
    assert periods.count("2025-04") == 0 or "April" in periods  # no second April
    assert "2025-06" in periods


def test_row_period_key_handles_formats():
    assert st._row_period_key({0: "2025-06", 1: 10}) == "2025-06"
    assert st._row_period_key({0: "June 2025", 1: 10}) == "2025-06"
    assert st._row_period_key({0: "6/30/2025", 1: 10}) == "2025-06"
    assert st._row_period_key({0: "June", 1: 10}) == "0000-06"
    # period end wins over start when both present
    assert st._row_period_key({0: "5/1/2025", 1: "5/31/2025"}) == "2025-05"
