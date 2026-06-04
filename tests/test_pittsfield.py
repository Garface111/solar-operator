"""
Regression tests for the Pittsfield GMCS detection bug.

Root cause: _detect_gmcs_shape required every A1 cell to match
"Array Name (NEPOOL-ID)".  Arrays without an assigned NEPOOL ID (e.g.
Pittsfield) have A1 = "Pittsfield" — no parenthetical — which caused
the whole workbook to fail GMCS detection and fall back to LLM/heuristic
extraction, silently dropping all arrays.

Fix: _detect_gmcs_shape only requires A1 to be non-empty; the row-5
header check ("Quarter" / "Generation (MWh)") is the definitive signal.
_extract_from_gmcs now includes sheets without NEPOOL IDs (nepool_gis_id=None).
"""
from __future__ import annotations

import io

import openpyxl

from api.ingest import _detect_gmcs_shape, _extract_from_gmcs


def _make_gmcs_wb(sheets: list[tuple[str, str | None]]) -> openpyxl.Workbook:
    """Build a minimal GMCS-format workbook.

    sheets: list of (array_name, nepool_gis_id_or_None)
    Each sheet gets a GMCS-style A1 header and the expected row-5 headers.
    """
    wb = openpyxl.Workbook()
    for i, (name, nepool_id) in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        a1_val = f"{name} ({nepool_id})" if nepool_id else name
        ws["A1"] = a1_val
        ws.cell(row=5, column=1, value="Quarter")
        ws.cell(row=5, column=2, value="Generation (MWh)")
    return wb


# ─── _detect_gmcs_shape ────────────────────────────────────────────────────

def test_detect_all_arrays_have_nepool():
    wb = _make_gmcs_wb([("Chester", "53984"), ("Londonderry", "98179")])
    assert _detect_gmcs_shape(wb) is True


def test_detect_pittsfield_no_nepool():
    """Single array with no NEPOOL ID must still be detected as GMCS."""
    wb = _make_gmcs_wb([("Pittsfield", None)])
    assert _detect_gmcs_shape(wb) is True


def test_detect_mixed_with_and_without_nepool():
    """The real-world Bruce case: 6 arrays with IDs + Pittsfield without one."""
    wb = _make_gmcs_wb([
        ("Chester", "53984"),
        ("Londonderry", "98179"),
        ("Tannery Brook", "46425"),
        ("Timberworks", "61959"),
        ("Waterford", "78671"),
        ("Pittsfield", None),         # <-- the bug trigger
        ("Starlake", None),
    ])
    assert _detect_gmcs_shape(wb) is True


def test_detect_rejects_non_gmcs():
    """A plain roster spreadsheet (no GMCS headers) must NOT be detected."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Operator Name"
    ws.cell(row=1, column=2, value="Array")
    ws.cell(row=2, column=1, value="Acme Solar")
    assert _detect_gmcs_shape(wb) is False


def test_detect_rejects_empty_a1():
    """Sheet with empty A1 cell is not GMCS — nothing to name the array."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = ""
    ws.cell(row=5, column=1, value="Quarter")
    ws.cell(row=5, column=2, value="Generation (MWh)")
    assert _detect_gmcs_shape(wb) is False


# ─── _extract_from_gmcs ────────────────────────────────────────────────────

def test_extract_includes_pittsfield_with_none_nepool():
    """Pittsfield (no NEPOOL ID) must appear in extracted rows with nepool_gis_id=None."""
    wb = _make_gmcs_wb([
        ("Chester", "53984"),
        ("Pittsfield", None),
    ])
    rows = _extract_from_gmcs(wb)
    names = {r["array_name"]: r["nepool_gis_id"] for r in rows}
    assert "Chester" in names
    assert names["Chester"] == "53984"
    assert "Pittsfield" in names
    assert names["Pittsfield"] is None


def test_extract_all_ids_present():
    wb = _make_gmcs_wb([
        ("Chester", "53984"),
        ("Londonderry", "98179"),
        ("Tannery Brook", "46425"),
    ])
    rows = _extract_from_gmcs(wb)
    assert len(rows) == 3
    ids = {r["array_name"]: r["nepool_gis_id"] for r in rows}
    assert ids["Chester"] == "53984"
    assert ids["Londonderry"] == "98179"
    assert ids["Tannery Brook"] == "46425"


def test_extract_skips_summary_sheet():
    wb = _make_gmcs_wb([("Chester", "53984")])
    summary = wb.create_sheet(title="Summary")
    summary["A1"] = "This is a summary"
    rows = _extract_from_gmcs(wb)
    assert len(rows) == 1
    assert rows[0]["array_name"] == "Chester"


def test_extract_operator_name_always_none():
    """operator_name is always None from GMCS extraction — set globally in preview UI."""
    wb = _make_gmcs_wb([("Chester", "53984"), ("Pittsfield", None)])
    rows = _extract_from_gmcs(wb)
    assert all(r["operator_name"] is None for r in rows)


def test_nepool_assign_gmcs_detection_with_pittsfield():
    """End-to-end: the nepool_assign._extract_pairs_from_gmcs path should
    return pairs for arrays WITH IDs even when Pittsfield (no ID) is present."""
    from api.nepool_assign import _extract_pairs_from_gmcs

    wb = _make_gmcs_wb([
        ("Chester", "53984"),
        ("Londonderry", "98179"),
        ("Pittsfield", None),  # no ID — should be silently skipped here
    ])
    pairs = _extract_pairs_from_gmcs(wb)
    pair_names = {p["array_name"] for p in pairs}
    assert "Chester" in pair_names
    assert "Londonderry" in pair_names
    # Pittsfield excluded from NEPOOL pairs (has no ID to assign)
    assert "Pittsfield" not in pair_names
    # IDs are correct
    assert any(p["nepool_gis_id"] == "53984" for p in pairs)
