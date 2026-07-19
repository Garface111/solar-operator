"""Generation reports are UTILITY data only — vendor twins must not appear.

Ford, 2026-07-19: "Waterford (Fronius)" was showing up in the Generation
Reports client roster next to the real "Waterford". The generation-reports
world is built from metered utility data; an array that exists only because an
inverter vendor reported it must be invisible there — roster, NEPOOL nudge,
workbook, and above all the $15/array/quarter billing unit.

The discriminator under test (api/report_arrays.not_vendor_only), verified on
prod ground truth:

    vendor-only  ==  ZERO live UtilityAccount rows  AND  >=1 Inverter row

The two shapes that a naive rule gets wrong are pinned here on purpose:

  * utility-backed array that ALSO has inverters (prod array 2333 "Waterford":
    2 accounts, 6 inverters, a real NEPOOL-GIS id) -> KEPT. A "has inverters ->
    hide" rule would delete a real reported array, and real money, from the
    workbook.
  * hand-added array with NO accounts and NO inverters -> KEPT. The operator
    just typed a name and is on their way to linking the utility account; hiding
    it makes that flow impossible.
"""
from __future__ import annotations

import secrets
from datetime import date, datetime, timedelta

import pytest

from api.db import SessionLocal
from api.models import (Array, Bill, Client, Inverter, Tenant, UtilityAccount)
from api.report_arrays import (is_vendor_only_array, not_vendor_only,
                               utility_backed_array_ids, vendor_only_array_ids)
from sqlalchemy import select


# ── fixture: one tenant, one client, the four array shapes ──────────────────

@pytest.fixture()
def world():
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Vendor Filter Co",
            contact_email=f"{tid}@vendorfilter.test",
            tenant_key="k_" + secrets.token_hex(8), plan="standard",
            active=True, product="array_operator", subscription_status="active",
            generation_reports=True,
        ))
        db.commit()
        c = Client(tenant_id=tid, name="Northeast Community", active=True)
        db.add(c); db.commit(); db.refresh(c)
        cid = c.id

        def mk_array(name, *, accounts=0, inverters=0, deleted_accounts=0):
            a = Array(tenant_id=tid, client_id=cid, name=name)
            db.add(a); db.commit(); db.refresh(a)
            for i in range(accounts):
                db.add(UtilityAccount(
                    tenant_id=tid, array_id=a.id, provider="gmp",
                    account_number=f"{a.id}{i}00001"))
            for i in range(deleted_accounts):
                db.add(UtilityAccount(
                    tenant_id=tid, array_id=a.id, provider="gmp",
                    account_number=f"{a.id}{i}99999",
                    deleted_at=datetime.utcnow()))
            for i in range(inverters):
                db.add(Inverter(
                    tenant_id=tid, array_id=a.id, vendor="fronius",
                    serial=f"SN-{a.id}-{i}"))
            db.commit()
            return a.id

        ids = {
            # the real one: meter + hardware (prod array 2333 shape)
            "utility_with_inverters": mk_array("Waterford", accounts=2, inverters=6),
            # the vendor twin (prod array 2735 shape)
            "vendor_only": mk_array("Waterford (Fronius)", inverters=12),
            # freshly hand-added, nothing linked yet
            "hand_added": mk_array("Brand New Site"),
            # plain metered array, no hardware monitored
            "utility_only": mk_array("Chester", accounts=1),
            # accounts existed but were removed -> no LIVE account + inverters
            "stale_accounts": mk_array("Retired Twin", deleted_accounts=2, inverters=3),
        }
    yield tid, cid, ids


# ── the discriminator itself ────────────────────────────────────────────────

def test_vendor_twin_is_excluded(world):
    _, cid, ids = world
    with SessionLocal() as db:
        keep = set(utility_backed_array_ids(db, client_id=cid))
        assert ids["vendor_only"] not in keep
        assert is_vendor_only_array(db, ids["vendor_only"]) is True


def test_utility_backed_array_with_inverters_is_kept(world):
    """THE regression a naive 'has inverters' rule breaks — prod array 2333."""
    _, cid, ids = world
    with SessionLocal() as db:
        keep = set(utility_backed_array_ids(db, client_id=cid))
        assert ids["utility_with_inverters"] in keep
        assert is_vendor_only_array(db, ids["utility_with_inverters"]) is False


def test_hand_added_array_with_nothing_linked_is_kept(world):
    """0 accounts + 0 inverters = the operator is mid-setup. Must stay visible."""
    _, cid, ids = world
    with SessionLocal() as db:
        keep = set(utility_backed_array_ids(db, client_id=cid))
        assert ids["hand_added"] in keep
        assert is_vendor_only_array(db, ids["hand_added"]) is False


def test_soft_deleted_accounts_do_not_count_as_utility_backing(world):
    _, cid, ids = world
    with SessionLocal() as db:
        assert is_vendor_only_array(db, ids["stale_accounts"]) is True


def test_vendor_only_and_utility_backed_partition_the_client(world):
    _, cid, ids = world
    with SessionLocal() as db:
        keep = set(utility_backed_array_ids(db, client_id=cid))
        drop = set(vendor_only_array_ids(db, client_id=cid))
        assert keep & drop == set()
        assert keep | drop == set(ids.values())
        assert drop == {ids["vendor_only"], ids["stale_accounts"]}


def test_criterion_composes_into_an_existing_select(world):
    tid, cid, ids = world
    with SessionLocal() as db:
        rows = db.execute(
            select(Array.id).where(
                Array.tenant_id == tid,
                Array.deleted_at.is_(None),
                not_vendor_only(),
            )
        ).all()
        got = {r[0] for r in rows}
    assert ids["vendor_only"] not in got
    assert ids["utility_with_inverters"] in got


def test_unknown_array_id_is_not_vendor_only(world):
    with SessionLocal() as db:
        assert is_vendor_only_array(db, 99_999_999) is False
        assert is_vendor_only_array(db, None) is False


# ── the surfaces ────────────────────────────────────────────────────────────

def _give_bills(array_id: int, kwh: float = 4000.0):
    """Put real metered generation on an array so it renders in a workbook."""
    with SessionLocal() as db:
        accts = db.execute(
            select(UtilityAccount).where(UtilityAccount.array_id == array_id,
                                         UtilityAccount.deleted_at.is_(None))
        ).scalars().all()
        acct = accts[0]
        end = date.today().replace(day=1) - timedelta(days=1)
        for k in range(14):
            m_end = (end.replace(day=1) - timedelta(days=30 * k))
            m_end = m_end.replace(day=28)
            db.add(Bill(
                tenant_id=acct.tenant_id, account_id=acct.id,
                period_start=m_end - timedelta(days=27), period_end=m_end,
                kwh_generated=kwh,
            ))
        db.commit()


def test_billing_unit_excludes_vendor_twins(world):
    """reported_array_ids is the $15/array/quarter unit — a twin must never bill."""
    _, cid, ids = world
    _give_bills(ids["utility_with_inverters"])
    from api.writers.gmcs_writer import reported_array_ids
    got = set(reported_array_ids(cid))
    assert ids["utility_with_inverters"] in got
    assert ids["vendor_only"] not in got
    assert ids["stale_accounts"] not in got


def test_workbook_omits_vendor_twins(world, tmp_path):
    from openpyxl import load_workbook
    from api.writers.gmcs_writer import build_workbook
    _, cid, ids = world
    _give_bills(ids["utility_with_inverters"])
    out = build_workbook(client_id=cid, out_path=tmp_path / "wb.xlsx")
    wb = load_workbook(out)
    titles = " | ".join(wb.sheetnames)
    wb.close()
    assert "Fronius" not in titles


def test_nepool_stats_do_not_count_vendor_twins(world):
    """The amber 'Add ID' nudge must not ask for a GIS id on a vendor twin."""
    from api.models import Array as A
    tid, cid, ids = world
    from sqlalchemy import func
    with SessionLocal() as db:
        counted = db.execute(
            select(func.count(A.id)).select_from(A)
            .join(Client, Client.id == A.client_id)
            .where(A.tenant_id == tid, A.deleted_at.is_(None),
                   A.nepool_gis_id.is_(None), not_vendor_only(),
                   Client.active.is_(True), Client.deleted_at.is_(None))
        ).scalar()
    # 5 arrays created, 2 are vendor-only -> 3 legitimately need an ID
    assert counted == 3
