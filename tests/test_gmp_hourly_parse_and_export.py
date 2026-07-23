"""GMP interval CSV → hourly aggregates + generation.xlsx hourly detail sheets.
"""
from __future__ import annotations

import secrets
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from api.adapters import gmp
from api.db import SessionLocal
from api.models import (
    Array,
    Client,
    GmpUsageRaw,
    Tenant,
    UtilityAccount,
)
from api.reports import gmp_daily_read
from api.writers.gmp_raw_writer import build_generation_workbook


def _csv_two_hours() -> str:
    """Four 15-min intervals at 10:00 and two at 11:00 on 2026-01-15."""
    rows = [
        "ServiceAgreement,IntervalStart,IntervalEnd,Quantity,UnitOfMeasure",
    ]
    for m in (0, 15, 30, 45):
        rows.append(
            f"SA1,2026-01-15 10:{m:02d}:00,2026-01-15 10:{m+15:02d}:00,1.5,kWh"
        )
    for m in (0, 15):
        rows.append(
            f"SA1,2026-01-15 11:{m:02d}:00,2026-01-15 11:{m+15:02d}:00,2.0,kWh"
        )
    # Second day, one hour
    rows.append("SA1,2026-01-16T14:00:00,2026-01-16T14:15:00,0.5,kWh")
    return "\n".join(rows) + "\n"


def test_parse_usage_csv_to_hourly_buckets():
    parsed = gmp.parse_usage_csv_to_hourly(_csv_two_hours())
    assert parsed["row_count"] == 7
    by = parsed["by_hour"]
    assert by[(date(2026, 1, 15), 10)]["kwh"] == 6.0  # 4 × 1.5
    assert by[(date(2026, 1, 15), 10)]["intervals"] == 4
    assert by[(date(2026, 1, 15), 11)]["kwh"] == 4.0  # 2 × 2.0
    assert by[(date(2026, 1, 15), 11)]["intervals"] == 2
    assert by[(date(2026, 1, 16), 14)]["kwh"] == 0.5
    # Daily aggregate still matches sum of hours for that day
    daily = gmp.parse_usage_csv_to_daily(_csv_two_hours())
    assert daily["by_day"][date(2026, 1, 15)]["kwh"] == 10.0
    assert daily["by_day"][date(2026, 1, 15)]["intervals"] == 6


def test_parse_interval_datetime_formats():
    assert gmp._parse_interval_datetime("2026-01-15 10:30:00").hour == 10
    assert gmp._parse_interval_datetime("2026-01-15T10:30:00.000Z").hour == 10
    assert gmp._parse_interval_datetime("01/15/2026 10:30").hour == 10
    assert gmp._parse_interval_date("2026-01-15 10:30:00") == date(2026, 1, 15)


def test_get_hourly_series_sums_meters_and_filters_window():
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Hourly Co", contact_email=f"{tid}@t.test",
            tenant_key="k_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        c = Client(tenant_id=tid, name="Client A", active=True)
        db.add(c)
        db.flush()
        arr = Array(tenant_id=tid, name="Barn", client_id=c.id, fuel_type="solar")
        db.add(arr)
        db.flush()
        ua1 = UtilityAccount(
            tenant_id=tid, provider="gmp", account_number="A1",
            array_id=arr.id,
        )
        ua2 = UtilityAccount(
            tenant_id=tid, provider="gmp", account_number="A2",
            array_id=arr.id,
        )
        db.add_all([ua1, ua2])
        db.flush()
        # Two meters, same hour — should sum.
        for ua, qty_line in (
            (ua1, "SA1,2026-01-15 10:00:00,2026-01-15 10:15:00,1.0,kWh"),
            (ua2, "SA2,2026-01-15 10:00:00,2026-01-15 10:15:00,2.5,kWh"),
        ):
            csv = (
                "ServiceAgreement,IntervalStart,IntervalEnd,Quantity,UnitOfMeasure\n"
                f"{qty_line}\n"
            )
            db.add(GmpUsageRaw(
                tenant_id=tid, account_id=ua.id, account_number=ua.account_number,
                window_start=date(2026, 1, 1), window_end=date(2026, 3, 31),
                fmt="csv", http_status=200, row_count=1, raw_csv=csv,
                interval_min=date(2026, 1, 15), interval_max=date(2026, 1, 15),
                fetched_at=datetime.utcnow(),
            ))
        # Outside window — ignored
        db.add(GmpUsageRaw(
            tenant_id=tid, account_id=ua1.id, account_number="A1",
            window_start=date(2025, 1, 1), window_end=date(2025, 3, 31),
            fmt="csv", http_status=200, row_count=1,
            raw_csv=(
                "ServiceAgreement,IntervalStart,IntervalEnd,Quantity,UnitOfMeasure\n"
                "SA1,2025-02-01 08:00:00,2025-02-01 08:15:00,99.0,kWh\n"
            ),
            interval_min=date(2025, 2, 1), interval_max=date(2025, 2, 1),
            fetched_at=datetime.utcnow(),
        ))
        db.commit()
        arr_id = arr.id

    series = gmp_daily_read.get_hourly_series(
        arr_id, start=date(2026, 1, 1), end=date(2026, 3, 31),
    )
    assert len(series) == 1
    assert series[0]["day"] == date(2026, 1, 15)
    assert series[0]["hour"] == 10
    assert series[0]["kwh"] == 3.5
    assert series[0]["meters"] == 2


def test_build_generation_workbook_hourly_detail(tmp_path: Path):
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Export Co", contact_email=f"{tid}@t.test",
            tenant_key="k_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        c = Client(tenant_id=tid, name="Farm", active=True)
        db.add(c)
        db.flush()
        arr = Array(tenant_id=tid, name="South Field", client_id=c.id, fuel_type="solar")
        db.add(arr)
        db.flush()
        ua = UtilityAccount(
            tenant_id=tid, provider="gmp", account_number="7001",
            array_id=arr.id,
        )
        db.add(ua)
        db.flush()
        db.add(GmpUsageRaw(
            tenant_id=tid, account_id=ua.id, account_number="7001",
            window_start=date(2026, 1, 1), window_end=date(2026, 3, 31),
            fmt="csv", http_status=200, row_count=6,
            raw_csv=_csv_two_hours(),
            interval_min=date(2026, 1, 15), interval_max=date(2026, 1, 16),
            fetched_at=datetime.utcnow(),
        ))
        db.commit()
        cid = c.id

    out = tmp_path / "gen.xlsx"
    build_generation_workbook(cid, out, year=2026, quarter=1)
    assert out.exists()
    wb = load_workbook(out)
    assert "Monthly Summary" in wb.sheetnames
    # Per-project detail sheet
    detail = [n for n in wb.sheetnames if n != "Monthly Summary"]
    assert detail, wb.sheetnames
    sh = wb[detail[0]]
    # Subtitle mentions Hourly
    assert "Hourly" in str(sh["A2"].value)
    # Headers
    # Find the header row under January
    headers = []
    for r in range(1, 20):
        if sh.cell(r, 1).value == "Date":
            headers = [sh.cell(r, c).value for c in range(1, 5)]
            break
    assert headers[0] == "Date"
    assert "Hour" in str(headers[1])
    assert "Generation" in str(headers[2])

    # Data rows include hour 10 with 6.0 kWh
    found_hour_10 = False
    for r in range(1, 80):
        if sh.cell(r, 1).value == "2026-01-15" and sh.cell(r, 2).value == 10:
            assert float(sh.cell(r, 3).value) == 6.0
            found_hour_10 = True
            break
    assert found_hour_10, "expected hourly row for 2026-01-15 hour 10"
