"""Master Array Data Pack — per-array mega spreadsheet from bills + daily gen."""
from __future__ import annotations

import io
import secrets
import zipfile
from datetime import date, datetime

from openpyxl import load_workbook

from api.db import SessionLocal
from api.models import Array, Bill, DailyGeneration, Tenant, UtilityAccount, now
from api.reports.array_master_pack import (
    array_master_filename,
    build_array_master_workbook,
    build_fleet_master_zip,
)


def _tenant() -> tuple[str, str]:
    tid = "ten_" + secrets.token_hex(6)
    key = "sol_test_" + secrets.token_hex(8)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Master Pack Co", contact_email=f"{key}@t.test",
            tenant_key=key, plan="standard", active=True, product="array_operator",
        ))
        db.commit()
    return tid, key


def _array_with_data(tid: str) -> int:
    with SessionLocal() as db:
        arr = Array(tenant_id=tid, name="Norwich Union Village", fuel_type="solar")
        db.add(arr)
        db.flush()
        ua = UtilityAccount(
            tenant_id=tid, array_id=arr.id, provider="gmp",
            account_number="1234567",
        )
        db.add(ua)
        db.flush()
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            period_start=datetime(2025, 1, 1),
            period_end=datetime(2025, 1, 31),
            billing_days=31,
            kwh_generated=1200,
            kwh_sent_to_grid=900,
            kwh_consumed=400,
            net_credit=180.0,
            solar_credit_usd=200.0,
            parse_status="parsed",
            pulled_at=now(),
        ))
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            period_start=datetime(2025, 6, 1),
            period_end=datetime(2025, 6, 30),
            billing_days=30,
            kwh_generated=1500,
            kwh_sent_to_grid=1100,
            parse_status="parsed",
            pulled_at=now(),
        ))
        for d, k in (
            (date(2026, 1, 10), 40.0),
            (date(2026, 1, 11), 42.0),
            (date(2026, 6, 1), 55.0),
        ):
            db.add(DailyGeneration(
                tenant_id=tid, array_id=arr.id, day=d, kwh=k,
                source="gmp_api", uploaded_at=now(),
            ))
        db.commit()
        return arr.id


def test_build_workbook_has_expected_sheets_and_rows():
    tid, _ = _tenant()
    aid = _array_with_data(tid)
    blob = build_array_master_workbook(tid, aid)
    assert blob and blob[:2] == b"PK"
    wb = load_workbook(io.BytesIO(blob))
    assert set(wb.sheetnames) >= {"Meta", "Bills", "Monthly", "Daily", "YoY"}

    bills = wb["Bills"]
    # header + 2 bills
    assert bills.cell(1, 1).value == "Period start"
    assert bills.cell(2, 6).value == 1200  # kWh generated first bill
    assert bills.cell(3, 6).value == 1500

    daily = wb["Daily"]
    assert daily.cell(1, 1).value == "Date"
    # 3 daily rows
    assert daily.cell(2, 2).value == 40.0
    assert daily.cell(4, 2).value == 55.0

    yoy = wb["YoY"]
    # At least calendar years present + trailing 12 mo label somewhere
    vals = [yoy.cell(r, 1).value for r in range(1, 12)]
    assert "Trailing 12 mo" in vals


def test_wrong_tenant_returns_none():
    tid, _ = _tenant()
    aid = _array_with_data(tid)
    other, _ = _tenant()
    assert build_array_master_workbook(other, aid) is None


def test_fleet_zip_contains_one_file_per_array():
    tid, _ = _tenant()
    a1 = _array_with_data(tid)
    with SessionLocal() as db:
        arr = Array(tenant_id=tid, name="Second Site", fuel_type="solar")
        db.add(arr)
        db.commit()
        a2 = arr.id
    blob, n = build_fleet_master_zip(tid)
    assert n == 2
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        names = zf.namelist()
    assert len(names) == 2
    assert any(str(a1) in n for n in names)
    assert any(str(a2) in n for n in names)


def test_download_endpoint_auth(client):
    tid, key = _tenant()
    aid = _array_with_data(tid)
    r = client.get(
        f"/v1/array-owners/arrays/{aid}/master-data.xlsx",
        headers={"Authorization": f"Bearer {key}"},
    )
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert r.content[:2] == b"PK"
    cd = r.headers.get("content-disposition", "")
    assert "master-data" in cd

    # zip
    r2 = client.get(
        "/v1/array-owners/master-data.zip",
        headers={"Authorization": f"Bearer {key}"},
    )
    assert r2.status_code == 200, r2.text
    assert r2.content[:2] == b"PK"
    assert r2.headers.get("x-array-count") == "1"


def test_filename_sanitizes():
    fn = array_master_filename("Norwich / Union Village!", 1356, when=date(2026, 7, 20))
    assert fn == "Norwich-Union-Village-1356-master-data-2026-07-20.xlsx"
    assert "/" not in fn
