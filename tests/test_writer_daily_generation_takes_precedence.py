"""
Tests that gmcs_writer prefers DailyGeneration over Bill data when both exist
for the same (array, year, month).

Invariants guarded:
- Month covered by DailyGeneration uses daily kWh, not Bill kWh
- Month NOT covered by DailyGeneration still uses Bill kWh (no regression)
- No double-counting (covered month total == daily kWh, not bill + daily)
"""
from __future__ import annotations

import secrets
from datetime import date, datetime

from openpyxl import load_workbook

from api.db import SessionLocal
from api.models import Array, Bill, Client, DailyGeneration, Tenant, UtilityAccount
from api.writers.gmcs_writer import build_workbook

# Reference date: Q2 2024 in progress → rolling window covers Q4'22–Q1'24
_REF = date(2024, 4, 1)

# Months in the rolling window: Q4'22–Q1'24
# Q1 2024 = Jan, Feb, Mar 2024
# Q4 2023 = Oct, Nov, Dec 2023
# We use Q4 2023 as the "bill only" month and Q1 2024 as the "daily gen" month.

_BILL_KWH = 5000     # Bill value for Q4 2023 months
_DAILY_KWH = 9000.0  # Daily total for Q1 2024 Jan (different from any bill value)


def _make_fixture() -> tuple[int, int]:
    """Create Tenant → Client → Array → UtilityAccount → Bills + DailyGeneration.

    Returns (client_id, array_id).

    Bill months: Oct, Nov, Dec 2023 (Q4 2023) — bill-only, no daily gen
    DailyGeneration: 31 days × 290.32 ≈ 9000 kWh for Jan 2024
    Bill also exists for Jan 2024 (5000 kWh) — should be OVERRIDDEN by daily gen
    """
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        t = Tenant(
            id=tid, name="Writer Precedence Test",
            contact_email=f"{tid}@test.com",
            tenant_key="k_" + secrets.token_hex(8),
            plan="standard", active=True,
        )
        db.add(t); db.flush()

        c = Client(tenant_id=tid, name="Prec Client", active=True)
        db.add(c); db.flush()

        arr = Array(
            tenant_id=tid, client_id=c.id,
            name="Prec Array", nepool_gis_id="PREC1",
        )
        db.add(arr); db.flush()

        ua = UtilityAccount(
            tenant_id=tid, array_id=arr.id,
            provider="gmp", account_number="PREC_" + secrets.token_hex(4),
        )
        db.add(ua); db.flush()

        # Bill data for Q4 2023 (bill-only months)
        for m in (10, 11, 12):
            db.add(Bill(
                tenant_id=tid, account_id=ua.id,
                bill_date=datetime(2023, m, 15),
                period_start=datetime(2023, m, 1),
                kwh_generated=_BILL_KWH,
                document_number=f"bill-{tid}-2023-{m}",
                parse_status="parsed",
            ))

        # Bill also exists for Jan 2024 — must be OVERRIDDEN by DailyGeneration
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            bill_date=datetime(2024, 1, 15),
            period_start=datetime(2024, 1, 1),
            kwh_generated=_BILL_KWH,
            document_number=f"bill-{tid}-2024-01",
            parse_status="parsed",
        ))

        # DailyGeneration for all 31 days of Jan 2024
        daily_per_day = _DAILY_KWH / 31
        for d in range(1, 32):
            db.add(DailyGeneration(
                tenant_id=tid,
                array_id=arr.id,
                day=date(2024, 1, d),
                kwh=daily_per_day,
                source="csv",
            ))

        db.commit()
        return c.id, arr.id


def test_daily_generation_overrides_bill_for_covered_month(tmp_path):
    """Jan 2024 is covered by DailyGeneration → writer uses daily kWh, not bill kWh."""
    cid, _ = _make_fixture()
    out = tmp_path / "prec.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)

    wb = load_workbook(out)
    sh = wb.active

    # Find the row for Jan 2024 (Q1 2024, first month)
    # Row layout: Q4'22(rows 7-9+gap), Q1'23(11-13+gap), Q2'23(15-17+gap),
    #             Q3'23(19-21+gap), Q4'23(23-25+gap), Q1'24(27-29+gap)
    # Q1 2024 starts at row 27 (first month of the 6th quarter block)
    jan_2024_kwh_mwh = None
    for row in range(7, 31):
        val = sh.cell(row, 2).value
        if val is not None:
            # Look for a value that matches our daily total (9000 kWh = 9.0 MWh)
            # rather than the bill value (5000 kWh = 5.0 MWh)
            pass

    # Collect all non-None MWh values from column B
    mwh_values = [
        sh.cell(r, 2).value for r in range(7, 31)
        if sh.cell(r, 2).value is not None
    ]

    # The daily total in MWh — should appear in the workbook
    expected_daily_mwh = round(_DAILY_KWH / 1000.0, 3)
    bill_mwh = round(_BILL_KWH / 1000.0, 3)

    assert expected_daily_mwh in mwh_values, (
        f"Expected daily MWh {expected_daily_mwh} not found in values: {mwh_values}"
    )


def test_bill_only_months_unaffected(tmp_path):
    """Q4 2023 (bill-only months) still use bill kWh after daily gen integration."""
    cid, _ = _make_fixture()
    out = tmp_path / "prec_bill.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)

    wb = load_workbook(out)
    sh = wb.active

    bill_mwh = round(_BILL_KWH / 1000.0, 3)
    mwh_values = [
        sh.cell(r, 2).value for r in range(7, 31)
        if sh.cell(r, 2).value is not None
    ]

    # The bill value should appear 3 times (Oct, Nov, Dec 2023) + 1 time (Jan 2024 is overridden)
    bill_count = sum(1 for v in mwh_values if abs(v - bill_mwh) < 0.001)
    assert bill_count == 3, (
        f"Expected 3 months at bill MWh={bill_mwh}, found {bill_count} in {mwh_values}"
    )


def test_no_double_counting(tmp_path):
    """Jan 2024 total in workbook is daily kWh, not bill + daily."""
    cid, _ = _make_fixture()
    out = tmp_path / "prec_nodc.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)

    wb = load_workbook(out)
    sh = wb.active

    expected_daily_mwh = round(_DAILY_KWH / 1000.0, 3)
    combined_mwh = round((_DAILY_KWH + _BILL_KWH) / 1000.0, 3)

    mwh_values = [
        sh.cell(r, 2).value for r in range(7, 31)
        if sh.cell(r, 2).value is not None
    ]

    assert combined_mwh not in mwh_values, (
        f"Double-counted MWh {combined_mwh} found — bill and daily were summed: {mwh_values}"
    )
    assert expected_daily_mwh in mwh_values, (
        f"Expected daily-only MWh {expected_daily_mwh} missing from {mwh_values}"
    )
