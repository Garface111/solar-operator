"""
Adversarial stress tests for the spreadsheet import pipeline (api/ingest.py).

CODE PATHS UNDER ATTACK AND WHY EACH ONE IS SUSPECT:

1. clean_text() (new) — previously absent; every string field went through
   str(v).strip() with no Unicode normalization. Smart quotes, NBSP, BOM, ZWS,
   em dashes, and b-string repr all pass through openpyxl str(c) unchanged, so
   the LLM and DB see them raw. Real operator spreadsheets carry all of these.

2. _coerce_nepool() (new) — previously _normalize() stored raw nepool_gis_id
   verbatim. "53984.0" (Excel float coercion), "53,984" (thousands sep), "#REF!"
   (broken formula), and "=SUM(…)" (formula text) all silently landed in the DB
   as garbage strings that would never match Bruce's live nepool_map lookups.

3. _normalize() boundary — single choke-point that all three ingest paths
   (GMCS extractor, LLM flatten, heuristic) pass through. Bad input here poisons
   everything downstream including the collision-detection logic.

4. _heuristic_extract() + row-0-as-header assumption — if the first visible row
   is a banner (merged title cells, date stamps) the parser silently loses the
   actual header, maps the wrong columns, and emits garbage or empty output.
   This is the no-LLM fallback so it must survive bad structure.

5. _xlsx_to_text() + openpyxl read_only=True — merged cell headers appear as
   the value in the top-left cell only; other cells in the merge return None →
   empty string → column shift in heuristic parser.

6. _file_to_text() routing — .xlsx named file with CSV content, .xls files, and
   tab-separated .csv all hit different code paths with different failure modes.

7. Security boundary — formula injection strings in text columns, malformed zip
   archives. The first must be stored as literal text; the second must not crash.

8. Domain logic gaps — NEPOOL+name blank combinations, intra-file duplicate
   NEPOOL IDs (only DB-level collision is checked, not within-upload), account
   number length (String(40) in DB, no pre-validate in ingest).
"""
from __future__ import annotations

import io
import secrets
import time
from unittest.mock import patch

import pytest
import openpyxl

from api.db import SessionLocal
from api.models import Tenant
from api.account import mint_session_for_tenant
from api.ingest import (
    clean_text,
    _coerce_nepool,
    _normalize,
    _heuristic_extract,
    _xlsx_to_text,
    _extract_from_gmcs,
    _detect_gmcs_shape,
    _csv_to_text,
)


# ── Fixtures / Helpers ─────────────────────────────────────────────────────────

@pytest.fixture()
def authed_stress(client):
    """Bare tenant for stress-test endpoint calls (no pre-seeded arrays)."""
    tenant_id = f"ten_{secrets.token_hex(8)}"
    with SessionLocal() as db:
        db.add(Tenant(
            id=tenant_id,
            name="Stress Test Op",
            contact_email=f"stress-{secrets.token_hex(4)}@example.com",
            tenant_key=f"sol_live_{secrets.token_urlsafe(18)}",
            plan="comped",
            active=True,
        ))
        db.commit()
    token = mint_session_for_tenant(tenant_id)
    return client, f"Bearer {token}"


def _xlsx_bytes_single(rows: list[list], sheet_name: str = "Roster") -> bytes:
    """Build an in-memory xlsx with a single sheet from a list of row lists."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_bytes_multi(sheets: list[tuple[str, list[list]]]) -> bytes:
    """Build an in-memory xlsx with multiple named sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title=title)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _preview(authed_stress, xlsx: bytes, filename: str = "roster.xlsx") -> dict:
    c, auth = authed_stress
    resp = c.post(
        "/v1/ingest/preview",
        files={"file": (filename, xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers={"Authorization": auth},
    )
    return resp


# ── Category 1: Character / Encoding Chaos ─────────────────────────────────────

class TestCleanText:
    """Unit tests for clean_text(). No DB, no xlsx — pure string normalisation."""

    def test_smart_quotes_replaced(self):
        """Curly/smart quotes U+2018–201D → plain ASCII equivalents.

        Operators copy-paste array names from PDF reports that use smart quotes.
        These survive openpyxl str(c) unchanged and previously landed in the DB
        verbatim, causing exact-match lookups to fail silently.
        """
        assert clean_text("‘Chester’") == "'Chester'"
        assert clean_text("“Londonderry Farm”") == '"Londonderry Farm"'

    def test_nbsp_becomes_regular_space(self):
        """Non-breaking space (U+00A0) → regular space, then stripped.

        Common when names are copy-pasted from browser tables or GMP portal HTML.
        """
        result = clean_text("Alice Moreau")
        assert result == "Alice Moreau"
        assert ' ' not in result

    def test_em_dash_becomes_hyphen(self):
        """Em dash (U+2014) and en dash (U+2013) → hyphen.

        Some operators record array ranges as 'Chester—East' or use em
        dashes in notes fields. Column parsers then fail to split on '-'.
        """
        assert clean_text("Chester—East") == "Chester-East"
        assert clean_text("Q1–2024") == "Q1-2024"

    def test_zero_width_space_removed(self):
        """Zero-width space (U+200B) stripped entirely.

        Invisible; frequently introduced when pasting from Google Sheets.
        Causes 'Chester' != 'Chester' string comparisons that are undebuggable.
        """
        result = clean_text("Chester​")
        assert result == "Chester"
        assert '​' not in result

    def test_bom_stripped(self):
        """BOM (U+FEFF) at start of cell value stripped.

        Created by CSV→XLSX conversions. The BOM appears as the first character
        of the first cell value, not just at the file level.
        """
        result = clean_text("﻿Chester")
        assert result == "Chester"
        assert '﻿' not in result

    def test_b_string_repr_stripped(self):
        """b'...' byte-string repr → inner string value.

        Happens when Python repr output gets copy-pasted into a spreadsheet cell,
        or when a CSV export serialises bytes objects instead of decoded strings.
        """
        assert clean_text("b'Chester'") == "Chester"
        assert clean_text('b"Tannery Brook"') == "Tannery Brook"

    def test_crlf_and_lf_normalised_to_space(self):
        """In-cell newlines (CRLF or bare LF) → single space.

        Multi-line cell content in Excel is common for address fields and notes.
        The heuristic splits on \\t so embedded newlines inside a cell value
        can accidentally look like a row boundary.
        """
        result = clean_text("line1\r\nline2")
        assert result == "line1 line2"
        result2 = clean_text("line1\nline2")
        assert result2 == "line1 line2"

    def test_multiple_chaos_chars_combined(self):
        """Multiple encoding issues in one string all resolved correctly.

        Represents a cell pasted from a PDF with smart quotes, BOM, and NBSP.
        """
        raw = "﻿‘Chester’ Farm​"
        result = clean_text(raw)
        assert result == "'Chester' Farm"
        assert all(c not in result for c in '﻿‘’ ​')


# ── Category 2: NEPOOL-ID Numeric Chaos ───────────────────────────────────────

class TestNepoolCoercion:
    """Unit tests for _coerce_nepool() / _normalize() NEPOOL handling.

    Every format variation is a concrete realistic scenario — not a made-up edge
    case. Bruce's GMP portal exports, NEPOOL exports, and Excel auto-formatting
    have produced every one of these in the wild.
    """

    def test_float_zero_frac(self):
        """53984.0 — Excel auto-converts an integer column to float on open."""
        result, err = _coerce_nepool("53984.0")
        assert result == "53984"
        assert err is None

    def test_leading_zero_string(self):
        """'053984' — Excel preserves leading zero when column is text type."""
        result, err = _coerce_nepool("053984")
        assert result == "53984"
        assert err is None

    def test_thousands_separator(self):
        """'53,984' — someone typed a thousands comma by hand."""
        result, err = _coerce_nepool("53,984")
        assert result == "53984"
        assert err is None

    def test_trailing_space(self):
        """'53984 ' — trailing space from sloppy copy-paste."""
        result, err = _coerce_nepool("53984 ")
        assert result == "53984"
        assert err is None

    def test_leading_apostrophe(self):
        """\"'53984\" — Excel's text-coercion prefix, stored in the cell value."""
        result, err = _coerce_nepool("'53984")
        assert result == "53984"
        assert err is None

    def test_six_digit_id(self):
        """100001 — NEPOOL IDs beyond 99999 (NEPOOL expansion already live)."""
        result, err = _coerce_nepool("100001")
        assert result == "100001"
        assert err is None

    def test_scientific_notation(self):
        """'5.3984E+04' — what Excel shows when the column is too narrow."""
        result, err = _coerce_nepool("5.3984e4")
        assert result == "53984"
        assert err is None

    def test_empty_string_is_none(self):
        """Empty string → None, no error. Blank cell = no ID supplied."""
        result, err = _coerce_nepool("")
        assert result is None
        assert err is None

    def test_dash_sentinel_is_none(self):
        """'-' → None, no error. Operators use this to mean 'TBD'."""
        result, err = _coerce_nepool("-")
        assert result is None
        assert err is None

    def test_na_sentinel_is_none(self):
        """'N/A' → None, no error. GMP portal export uses this."""
        result, err = _coerce_nepool("N/A")
        assert result is None
        assert err is None

    def test_tbd_sentinel_is_none(self):
        """'TBD' → None, no error. Common in draft rosters."""
        result, err = _coerce_nepool("TBD")
        assert result is None
        assert err is None

    def test_ref_error_rejected_with_message(self):
        """'#REF!' — Excel formula error in what should be a number column.

        Must be rejected with a structured error, not stored as '#REF!' in the
        DB where it would silently poison every NEPOOL lookup for that array.
        """
        result, err = _coerce_nepool("#REF!")
        assert result is None
        assert err is not None
        assert "#REF!" in err or "invalid" in err.lower()

    def test_formula_text_rejected_with_message(self):
        """'=SUM(A1:A5)' — literal formula text in a string-typed cell.

        Someone copy-pasted the formula text instead of its value. Must error,
        not land in the DB as '=SUM(A1:A5)'.
        """
        result, err = _coerce_nepool("=SUM(A1:A5)")
        assert result is None
        assert err is not None
        assert "invalid" in err.lower() or "fix" in err.lower()

    def test_two_ids_in_one_cell_flagged_not_silent(self):
        """'53984; 53985' — operator stuffed two NEPOOL IDs in one cell.

        Previously stored verbatim in the DB: '53984; 53985' matched nothing in
        the nepool_map, silently losing both IDs. Now surfaces a structured error
        so the operator knows they need to split the row.
        """
        rows = _normalize([{
            "operator_name": "Acme Op",
            "array_name": "Dual Array",
            "nepool_gis_id": "53984; 53985",
            "gmp_account_number": None,
            "notes": None,
        }])
        assert len(rows) == 1, "row must not be silently dropped"
        row = rows[0]
        assert row["nepool_gis_id"] is None
        assert "nepool_parse_error" in row, "must flag as parse error for operator review"

    def test_negative_id_rejected(self):
        """Negative NEPOOL ID — signals a column-shift parsing slip."""
        result, err = _coerce_nepool("-53984")
        assert result is None
        assert err is not None

    def test_normalize_propagates_parse_error(self):
        """_normalize() keeps the row and adds nepool_parse_error when coercion fails.

        Silent drop would be worse than a visible error: the operator would never
        know an array was lost from their import batch.
        """
        rows = _normalize([{
            "operator_name": "Op",
            "array_name": "Bad NEPOOL Array",
            "nepool_gis_id": "#VALUE!",
            "gmp_account_number": None,
            "notes": None,
        }])
        assert len(rows) == 1
        assert rows[0]["nepool_gis_id"] is None
        assert rows[0].get("nepool_parse_error") is not None

    def test_null_nepool_passthrough(self):
        """None NEPOOL → stays None, no error, row preserved."""
        rows = _normalize([{
            "operator_name": "Op",
            "array_name": "No NEPOOL Array",
            "nepool_gis_id": None,
            "gmp_account_number": None,
            "notes": None,
        }])
        assert len(rows) == 1
        assert rows[0]["nepool_gis_id"] is None
        assert "nepool_parse_error" not in rows[0]


# ── Category 3: Name / Account Collision + Weirdness ──────────────────────────

class TestNameCollisionWeirdness:
    """Test how _normalize handles tricky name and account-number scenarios."""

    def test_two_rows_same_array_name_different_nepool(self):
        """Homonym arrays: same name, legitimately different NEPOOL IDs.

        This is real — 'Chester (East)' and 'Chester' are two different arrays
        in Bruce's portfolio. Both must appear; neither should be deduplicated.
        """
        rows = _normalize([
            {"operator_name": "Op", "array_name": "Chester", "nepool_gis_id": "53984",
             "gmp_account_number": None, "notes": None},
            {"operator_name": "Op", "array_name": "Chester", "nepool_gis_id": "53985",
             "gmp_account_number": None, "notes": None},
        ])
        assert len(rows) == 2
        nepool_ids = {r["nepool_gis_id"] for r in rows}
        assert nepool_ids == {"53984", "53985"}

    def test_two_rows_same_nepool_different_names(self):
        """Duplicate NEPOOL, different names — operator typo or intentional split.

        Both rows must appear so the operator can review and decide, not be
        silently merged or deduplicated.
        """
        rows = _normalize([
            {"operator_name": "Op", "array_name": "Chester",
             "nepool_gis_id": "53984", "gmp_account_number": None, "notes": None},
            {"operator_name": "Op", "array_name": "Chester (Legacy)",
             "nepool_gis_id": "53984", "gmp_account_number": None, "notes": None},
        ])
        assert len(rows) == 2

    def test_blank_array_name_and_operator_row_dropped(self):
        """No array_name, no operator_name → row dropped even if NEPOOL is set.

        _normalize requires at least one of (operator_name, array_name) to keep
        a row. NEPOOL alone is not enough because we'd have nothing to name the
        Array record. Tests the 'drop if empty' boundary exactly.
        """
        rows = _normalize([{
            "operator_name": None,
            "array_name": "",
            "nepool_gis_id": "53984",
            "gmp_account_number": None,
            "notes": None,
        }])
        assert rows == [], "row with neither array_name nor operator_name must be dropped"

    def test_blank_row_except_notes_dropped(self):
        """Row with only notes populated (no name, no operator) → dropped.

        Arises when a GMCS sheet has footnotes below the data table that
        openpyxl reads as extra rows.
        """
        rows = _normalize([{
            "operator_name": None,
            "array_name": None,
            "nepool_gis_id": None,
            "gmp_account_number": None,
            "notes": "See page 2 for explanation",
        }])
        assert rows == []

    def test_encoding_cleaned_in_array_name_through_normalize(self):
        """Smart quotes in array_name are cleaned by _normalize via clean_text().

        Tests the integration path: raw LLM output with smart quotes → _normalize
        → clean array name in the DB row.
        """
        rows = _normalize([{
            "operator_name": "Op",
            "array_name": "“Chester” Farm",
            "nepool_gis_id": "53984",
            "gmp_account_number": None,
            "notes": None,
        }])
        assert len(rows) == 1
        assert rows[0]["array_name"] == '"Chester" Farm'

    def test_nbsp_in_operator_name_cleaned(self):
        """NBSP in operator name → regular space after clean_text.

        Fuzzy matching on 'Alice Moreau' vs 'Alice Moreau' would fail
        without normalisation.
        """
        rows = _normalize([{
            "operator_name": "Alice Moreau",
            "array_name": "Moreau Farm",
            "nepool_gis_id": "53984",
            "gmp_account_number": None,
            "notes": None,
        }])
        assert rows[0]["operator_name"] == "Alice Moreau"


# ── Category 4: Sheet / Workbook Structure Chaos ───────────────────────────────

class TestSheetStructureChaos:
    """Test how xlsx structure variations affect the text extraction layer."""

    def test_data_on_non_first_sheet_included_in_text(self):
        """_xlsx_to_text reads ALL sheets, not just sheet 0.

        Operators sometimes build workbooks where the first tab is a 'Summary'
        and the actual roster is on 'Q3 2024' (sheet 7). Previously only sheet 0
        was read, silently dropping every other tab.
        """
        filler_rows = [["Notes only, no roster data here"]]
        roster_rows = [
            ["Client", "Array Name", "NEPOOL ID"],
            ["Acme", "Acme Farm", "53984"],
        ]
        # 7 filler sheets + 1 roster sheet
        sheets = [(f"Filler{i}", filler_rows) for i in range(6)]
        sheets.append(("Q3 2024", roster_rows))
        data = _xlsx_bytes_multi(sheets)
        text = _xlsx_to_text(data)
        assert "Acme Farm" in text
        assert "53984" in text
        assert "Q3 2024" in text  # sheet separator header

    def test_hidden_rows_included_in_text(self):
        """Hidden rows are visible to _xlsx_to_text (read_only mode ignores hide state).

        Operators sometimes hide discontinued arrays. The parser currently sees
        them — this is documented behaviour, not a silent bug. The test pins the
        current behaviour so any future change is intentional.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Client", "Array Name", "NEPOOL ID"])
        ws.append(["Op", "Active Farm", "53984"])
        ws.append(["Op", "Discontinued Farm", "53985"])
        ws.row_dimensions[3].hidden = True
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        text = _xlsx_to_text(data)
        assert "Discontinued Farm" in text, (
            "hidden rows are currently included — this test pins that behaviour"
        )

    def test_wide_workbook_extra_columns_ignored_by_heuristic(self):
        """200 junk columns after the canonical ones don't cause quadratic blowup.

        Some NEPOOL-export files have 200+ columns of metadata to the right of
        the canonical fields. The heuristic parser maps by keyword and ignores
        unmapped indices — this test verifies no crash and correct extraction.
        """
        junk_headers = [f"junk_{i}" for i in range(200)]
        header = ["Array Name", "NEPOOL GIS ID"] + junk_headers
        data_row = ["Wide Farm", "53984"] + ["x"] * 200
        rows = _normalize(_heuristic_extract(
            "\t".join(header) + "\n" + "\t".join(data_row)
        ))
        assert len(rows) == 1
        assert rows[0]["array_name"] == "Wide Farm"
        assert rows[0]["nepool_gis_id"] == "53984"

    @pytest.mark.xfail(reason="defect-import-03: heuristic parser always treats row 0 "
                       "as header; a banner row at top causes actual header to be "
                       "treated as data, producing a spurious extra row", strict=True)
    def test_banner_row_at_top_breaks_heuristic(self):
        """xlsx with a title banner above the header row — heuristic misparses.

        Real GMCS exports start with 3 merged title rows before the column
        headers. The heuristic parser assumes row 0 is the header. When row 0 is
        a banner ('Solar Array Portfolio Q3 2024') with no keyword matches, the
        parser falls into positional-column mode for ALL subsequent rows,
        treating the real header row as a data row.

        Expected (correct): 1 data row (Acme Farm / 53984)
        Actual (broken): 2 rows — the real header row mistaken as data plus the
        actual data row, with column-name values ('Array Name') in the operator
        field.
        """
        text = "\n".join([
            "Solar Array Portfolio Q3 2024",  # banner — row 0, no keyword matches
            "Client\tArray Name\tNEPOOL GIS ID",  # real header — row 1, treated as data
            "Acme Corp\tAcme Farm\t53984",          # real data — row 2
        ])
        rows = _normalize(_heuristic_extract(text))
        # Should be exactly 1 row — only 'Acme Farm'
        assert len(rows) == 1
        assert rows[0]["array_name"] == "Acme Farm"

    @pytest.mark.xfail(reason="defect-import-04: merged cells in header row produce an "
                       "empty-string column that shifts subsequent column indices, "
                       "causing the heuristic to mismap NEPOOL ID to the wrong column",
                       strict=True)
    def test_merged_header_cells_cause_column_shift(self):
        """Merged 'Array Name' header spanning A:B shifts NEPOOL to wrong column.

        openpyxl read_only mode returns None for non-top-left cells in a merge.
        _xlsx_to_text converts None → ''. The heuristic header row sees:
          col 0 = 'array name'  (matches array_name)
          col 1 = ''            (unmatched — phantom column)
          col 2 = 'nepool id'   (matches nepool_gis_id at index 2)
        Data rows have no merge, so NEPOOL lands in col 1. col_field[2]=nepool
        but data col 1 is never mapped. NEPOOL is silently lost.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roster"
        # Merge A1:B1 — value in A1 only, B1 reads as None in read_only
        ws["A1"] = "Array Name"
        ws["C1"] = "NEPOOL GIS ID"
        ws.merge_cells("A1:B1")
        ws.append(["Acme Farm", "53984", ""])  # data: A2 array, B2 nepool, C2 empty
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        text = _xlsx_to_text(data)
        rows = _normalize(_heuristic_extract(text))
        assert len(rows) == 1
        assert rows[0]["nepool_gis_id"] == "53984"

    @pytest.mark.xfail(reason="defect-import-05: intra-file duplicate NEPOOL IDs are "
                       "not detected; only DB-level collision (nepool_map) is checked. "
                       "Two rows in the same upload with the same NEPOOL pass through "
                       "silently, creating ambiguity at commit time.", strict=True)
    def test_intrafile_duplicate_nepool_flagged(self, authed_stress):
        """Two rows in the same file with the same NEPOOL ID → should be flagged.

        Current behaviour: both rows appear in the preview with no warning.
        Correct behaviour: a warning (or per-row flag) should surface the
        intra-file duplicate so the operator resolves it before commit.

        Severity: HIGH — at commit time the second row will silently re-use the
        first row's Array record (find_or_create by name), which may or may not
        be the operator's intent.
        """
        xlsx = _xlsx_bytes_single([
            ["Client", "Array Name", "NEPOOL GIS ID"],
            ["Op A", "Array Alpha", "53984"],
            ["Op B", "Array Beta",  "53984"],  # same NEPOOL, different name
        ])
        hier = {"operators": [{
            "name": "TestOp",
            "clients": [
                {"name": "Op A", "logins": [{"utility": "gmp", "login_email": None,
                    "accounts": [{"account_number": None,
                        "arrays": [{"name": "Array Alpha", "nepool_gis_id": "53984",
                                    "notes": None, "confidence": 0.95}]}]}]},
                {"name": "Op B", "logins": [{"utility": "gmp", "login_email": None,
                    "accounts": [{"account_number": None,
                        "arrays": [{"name": "Array Beta", "nepool_gis_id": "53984",
                                    "notes": None, "confidence": 0.95}]}]}]},
            ],
        }]}
        with patch("api.ingest._llm_extract_hierarchical", return_value=hier):
            resp = _preview(authed_stress, xlsx)
        assert resp.status_code == 200
        body = resp.json()
        warning_kinds = [w["kind"] for w in body["warnings"]]
        # Expect a new warning kind like 'intrafile_nepool_duplicate'
        assert any("duplicate" in k or "intrafile" in k for k in warning_kinds), (
            "no warning for intra-file NEPOOL duplicate"
        )


# ── Category 5: CSV / XLSX Masquerade ──────────────────────────────────────────

class TestCsvMasquerade:
    """Test file-format boundary detection and routing."""

    def test_xlsx_named_but_csv_content_returns_400(self, authed_stress):
        """File named .xlsx but containing CSV bytes → clear HTTP 400.

        Operators sometimes save-as wrong in Excel. The code must detect the
        zipfile parse failure and return a human-readable error, not a 500 or
        a silent empty parse.
        """
        csv_content = b"Client,Array Name,NEPOOL ID\nOp,Chester,53984\n"
        resp = _preview(authed_stress, csv_content, filename="roster.xlsx")
        assert resp.status_code == 400
        body = resp.json()
        assert "detail" in body

    def test_xls_file_returns_400_with_clear_message(self, authed_stress):
        """Legacy .xls binary format → 400 with actionable error message.

        openpyxl cannot read .xls. The code must surface a clear error rather
        than a cryptic Python traceback.
        """
        # Minimal BIFF8 header bytes — definitely not a zip/xlsx
        biff_stub = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 512
        resp = _preview(authed_stress, biff_stub, filename="old_roster.xls")
        assert resp.status_code == 400
        body = resp.json()
        assert "xls" in body.get("detail", "").lower() or "legacy" in body.get("detail", "").lower()

    def test_csv_upload_routed_correctly(self, authed_stress):
        """A valid CSV file named .csv parses without error via heuristic path."""
        csv_content = b"Client,Array Name,NEPOOL ID\nAcme Corp,Acme Farm,53984\n"
        c, auth = authed_stress
        with patch("api.ingest._llm_extract_hierarchical", return_value=None):
            resp = c.post(
                "/v1/ingest/preview",
                files={"file": ("roster.csv", csv_content, "text/csv")},
                headers={"Authorization": auth},
            )
        assert resp.status_code == 200
        body = resp.json()
        assert body["count"] >= 1
        names = [r["array_name"] for r in body["arrays"]]
        assert "Acme Farm" in names

    def test_malformed_zip_returns_400(self, authed_stress):
        """Truncated/corrupted zip archive → HTTP 400, not Python KeyError or 500.

        openpyxl raises zipfile.BadZipFile or InvalidFileException. The endpoint
        must catch and re-raise as a user-readable 400.
        """
        garbage = b"PK\x03\x04" + b"\xff" * 100  # broken zip magic
        resp = _preview(authed_stress, garbage, filename="broken.xlsx")
        assert resp.status_code == 400


# ── Category 6: Size / Performance Edge Cases ──────────────────────────────────

class TestSizeEdgeCases:
    """Tests that would catch O(n²) blowup or memory issues on realistic file sizes."""

    def test_5000_row_workbook_parses_in_under_3_seconds(self):
        """5 000-row xlsx must not cause quadratic blowup in _xlsx_to_text.

        The heuristic / LLM text is already capped at MAX_TEXT_CHARS (60k) so
        actual parsing is fast; the bottleneck is the openpyxl iteration over
        5k rows. If this takes > 3s we have a structural problem.
        """
        rows: list[list] = [["Client", "Array Name", "NEPOOL GIS ID"]]
        for i in range(5000):
            rows.append([f"Op {i}", f"Array {i}", str(50000 + i)])
        data = _xlsx_bytes_single(rows)
        start = time.monotonic()
        text = _xlsx_to_text(data)
        elapsed = time.monotonic() - start
        assert elapsed < 3.0, f"_xlsx_to_text took {elapsed:.2f}s for 5k rows"
        assert "Array 0" in text  # sanity

    def test_embedded_image_does_not_crash_parser(self):
        """An xlsx with an embedded PNG must parse without error.

        openpyxl with read_only=True ignores drawing parts, so images should
        be silently skipped. This test guards against a regression if openpyxl
        handling changes.
        """
        # Build a minimal xlsx with a tiny embedded PNG via openpyxl
        from openpyxl.drawing.image import Image as XlImage
        import struct, zlib

        def _tiny_png() -> bytes:
            """Minimal 1×1 white PNG."""
            def chunk(tag, data):
                return struct.pack(">I", len(data)) + tag + data + struct.pack(">I", zlib.crc32(tag + data) & 0xFFFFFFFF)
            header = b'\x89PNG\r\n\x1a\n'
            ihdr = chunk(b'IHDR', struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0))
            idat = chunk(b'IDAT', zlib.compress(b'\x00\xff\xff\xff'))
            iend = chunk(b'IEND', b'')
            return header + ihdr + idat + iend

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Roster"
        ws.append(["Client", "Array Name", "NEPOOL ID"])
        ws.append(["Op", "Image Farm", "53984"])
        try:
            img = XlImage(io.BytesIO(_tiny_png()))
            ws.add_image(img, "E1")
        except Exception:
            pass  # if PIL not available, skip image; still a valid xlsx
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        # Must not raise
        text = _xlsx_to_text(data)
        assert "Image Farm" in text


# ── Category 7: Security / Trust Boundary ─────────────────────────────────────

class TestSecurity:
    """Formula injection and malformed-file security boundaries."""

    def test_excel_formula_injection_stored_as_literal_text(self):
        """=cmd|'/c calc'!A1 in array_name column → stored as literal string.

        openpyxl data_only=True returns the cached cell value, NOT the formula.
        For a text-typed cell containing this string (not a real formula), the
        value comes through as-is. _normalize must NOT strip the leading '=' —
        it's already literal text, not an executable formula in our context.

        If this string later appears in a GMCS report Excel file, it COULD be
        executed by the reader's Excel. That downstream risk is out of scope for
        this PR but is documented.
        """
        injection = "=cmd|'/c calc'!A1"
        rows = _normalize([{
            "operator_name": "Op",
            "array_name": injection,
            "nepool_gis_id": None,
            "gmp_account_number": None,
            "notes": None,
        }])
        assert len(rows) == 1
        assert rows[0]["array_name"] == injection, (
            "formula injection string must be stored verbatim, not evaluated or mangled"
        )

    def test_formula_injection_in_nepool_column_rejected(self):
        """=cmd|'...' in NEPOOL column → structured error, not stored.

        Unlike text fields where we preserve the literal string, a NEPOOL field
        must contain a 4–6 digit integer. Anything starting with '=' is clearly
        wrong and must be rejected with a visible error, never stored as-is.
        """
        rows = _normalize([{
            "operator_name": "Op",
            "array_name": "Injection Farm",
            "nepool_gis_id": "=cmd|'/c calc'!A1",
            "gmp_account_number": None,
            "notes": None,
        }])
        assert rows[0]["nepool_gis_id"] is None
        assert rows[0].get("nepool_parse_error") is not None

    def test_very_long_string_does_not_oom(self):
        """A 1 MB string in one cell must not OOM the parser.

        _xlsx_to_text truncates the full text at MAX_TEXT_CHARS (60k) before
        sending to LLM, so a huge cell value is truncated, not passed whole.
        This test guards against the slice happening AFTER a quadratic concat.
        """
        big_value = "A" * (1024 * 1024)  # 1 MB
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Big"
        ws.append(["Client", "Array Name", "NEPOOL ID"])
        ws.append(["Op", big_value, "53984"])
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        # Must complete without OOM or crash
        text = _xlsx_to_text(data)
        assert "53984" in text  # the NEPOOL column must still appear


# ── Category 8: Domain Logic ───────────────────────────────────────────────────

class TestDomainLogic:
    """Tests for the semantic rules around what rows should do to the DB."""

    def test_nepool_set_array_name_blank_row_dropped(self):
        """Row with NEPOOL filled but array_name blank → row dropped.

        Current documented behaviour: _normalize requires operator_name OR
        array_name to keep a row. NEPOOL alone is insufficient — we have no name
        to create an Array record with. This pins the current behaviour.
        """
        rows = _normalize([{
            "operator_name": None,
            "array_name": None,
            "nepool_gis_id": "53984",
            "gmp_account_number": None,
            "notes": None,
        }])
        assert rows == [], (
            "a row with only nepool_gis_id set and no name must be dropped"
        )

    def test_array_name_set_nepool_blank_row_preserved(self):
        """Row with array_name but no NEPOOL → kept with nepool_gis_id=None.

        Matches the 'we don't verify' stance from the Pittsfield fix: arrays
        without NEPOOL are valid and the operator can assign them later.
        """
        rows = _normalize([{
            "operator_name": "Op",
            "array_name": "Pittsfield",
            "nepool_gis_id": None,
            "gmp_account_number": None,
            "notes": None,
        }])
        assert len(rows) == 1
        assert rows[0]["array_name"] == "Pittsfield"
        assert rows[0]["nepool_gis_id"] is None

    def test_chester_disambiguation_all_three_surface(self):
        """Chester / Chester (East) / Chester Solar — all three must appear.

        A real scenario in Bruce's GMCS: three differently-named arrays all
        named 'Chester-something'. _normalize must not deduplicate by name.
        """
        raw_rows = [
            {"operator_name": "Bruce", "array_name": "Chester",
             "nepool_gis_id": "53984", "gmp_account_number": None, "notes": None},
            {"operator_name": "Bruce", "array_name": "Chester (East)",
             "nepool_gis_id": "53985", "gmp_account_number": None, "notes": None},
            {"operator_name": "Bruce", "array_name": "Chester Solar",
             "nepool_gis_id": "53986", "gmp_account_number": None, "notes": None},
        ]
        rows = _normalize(raw_rows)
        assert len(rows) == 3
        names = {r["array_name"] for r in rows}
        assert names == {"Chester", "Chester (East)", "Chester Solar"}

    @pytest.mark.xfail(reason="defect-import-06: account_number > 40 chars passes "
                       "through _normalize without validation; DB column is String(40) "
                       "and will raise OperationalError at commit time rather than "
                       "surfacing a clean parse error to the operator in preview",
                       strict=True)
    def test_account_number_too_long_flagged_at_parse_time(self):
        """account_number > 40 chars must be flagged in preview, not fail at commit.

        UtilityAccount.account_number is String(40). A 41-char value passes
        through _normalize today and only blows up when the DB INSERT executes.
        The operator sees a 500 error instead of a helpful 'account number too
        long' message in the import preview table.
        """
        long_acct = "A" * 41
        rows = _normalize([{
            "operator_name": "Op",
            "array_name": "Long Acct Farm",
            "nepool_gis_id": "53984",
            "gmp_account_number": long_acct,
            "notes": None,
        }])
        assert len(rows) == 1
        # Should have a parse error or truncation warning — currently doesn't
        assert "account_parse_error" in rows[0] or rows[0]["gmp_account_number"] != long_acct, (
            "account number longer than 40 chars must be flagged before commit"
        )

    def test_gmcs_shape_with_encoding_in_a1(self):
        """GMCS A1 cell with smart quotes in array name → clean_text normalises it.

        If the GMCS workbook was re-saved from Word or PDF, A1 may contain curly
        quotes around the array name. _extract_from_gmcs returns the raw cell
        value; _normalize (via clean_text) must clean it.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "“Chester” (53984)"  # curly quotes around name
        ws.cell(row=5, column=1, value="Quarter")
        ws.cell(row=5, column=2, value="Generation (MWh)")
        raw_rows = _extract_from_gmcs(wb)
        normalised = _normalize(raw_rows)
        assert len(normalised) == 1
        # After clean_text, curly quotes should become plain double quotes
        assert '“' not in normalised[0]["array_name"]
        assert '”' not in normalised[0]["array_name"]
        assert "53984" == normalised[0]["nepool_gis_id"]
