"""Tests for the "flawless" offtaker upload (Ford, 2026-07-01).

Covers the fuzzy array-name matcher, bulk-import v2 preview (CSV + XLSX, fuzzy
match, correctable dry-run shape), bulk-commit (idempotency), array-first manual
create (resolve the array's utility bill), and the xlsx template download.

BILLING-CRITICAL: a wrong array→offtaker match makes a wrong invoice, so these
assert that medium/none-confidence matches surface for review and are never
silently committed.
"""
from __future__ import annotations

import io
import secrets
from datetime import date, timedelta

from sqlalchemy import select

from api.account import mint_session_for_tenant
from api.db import SessionLocal
from api.models import (Tenant, Client, Array, UtilityAccount, Bill,
                        BillingReportSubscription)
from api.billing.offtaker_match import match_array


def _make_tenant() -> tuple[str, str]:
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Offtaker Upload Operator",
            contact_email=f"{tid}@operator.test",
            tenant_key="sol_live_" + secrets.token_urlsafe(12),
            plan="standard", active=True, product="array_operator",
        ))
        db.commit()
    return tid, f"Bearer {mint_session_for_tenant(tid)}"


def _make_array_with_bill(tid: str, name: str, acct_number: str,
                          with_bill: bool = True) -> tuple[int, int]:
    """An array + a GMP utility account, optionally with one settled Bill.
    Returns (array_id, utility_account_id)."""
    with SessionLocal() as db:
        c = Client(tenant_id=tid, name=f"{name} Client", active=True)
        db.add(c); db.flush()
        arr = Array(tenant_id=tid, name=name, client_id=c.id, fuel_type="solar")
        db.add(arr); db.flush()
        aid = arr.id
        acct = UtilityAccount(tenant_id=tid, array_id=aid, provider="gmp",
                              account_number=acct_number, nickname=name)
        db.add(acct); db.flush()
        uaid = acct.id
        if with_bill:
            pe = date.today().replace(day=1) - timedelta(days=1)
            db.add(Bill(tenant_id=tid, account_id=uaid,
                        bill_date=pe, period_start=pe.replace(day=1),
                        period_end=pe, kwh_generated=5000.0,
                        document_number="DOC-" + secrets.token_hex(3)))
        db.commit()
    return aid, uaid


# ─── the pure matcher (belt-and-suspenders over its own __main__ self-test) ──

def test_matcher_confidence_classes():
    arrays = [{"id": 1, "name": "Maple Street Solar (53984)"},
              {"id": 2, "name": "Route 7 Community Array"}]
    uas = [{"utility_account_id": 11, "array_id": 1, "array_name": "Maple Street Solar",
            "nickname": "Maple St", "provider": "gmp", "account_number": "111",
            "has_bill": True}]
    # exact (GIS id stripped)
    assert match_array("Maple Street Solar (53984)", arrays, uas)["confidence"] == "exact"
    # typo → not none, right array
    m = match_array("Maple Steet Solar", arrays, uas)
    assert m["confidence"] in ("high", "medium") and m["array_id"] == 1
    # unrelated → none, no array
    m = match_array("Zzz Nowhere Plant", arrays, uas)
    assert m["confidence"] == "none" and m["array_id"] is None


# ─── bulk-import v2 preview ──────────────────────────────────────────────────

def _bulk_import(client, auth, filename: str, content: bytes, ctype: str, **form):
    return client.post(
        "/v1/array-operator/billing/subscriptions/bulk-import",
        files={"file": (filename, content, ctype)},
        data=form, headers={"Authorization": auth})


def test_bulk_import_csv_fuzzy_match_and_shape(client):
    tid, auth = _make_tenant()
    aid1, ua1 = _make_array_with_bill(tid, "Maple Street Solar", "GMP-111", with_bill=True)
    aid2, ua2 = _make_array_with_bill(tid, "Route 7 Community Array", "GMP-222", with_bill=False)

    csv = (
        "Array,Offtaker,Share %,Email,Discount %,Notes\n"
        "Maple Street Solar,Jane Offtaker,25,jane@example.com,10,vip\n"          # exact + bill → ready
        "Route 7 Community,Town Elsewhere,15,,,\n"                                 # matched but NO bill → review
        "Zzz Nonexistent,Bad Row,50,,,\n"                                          # no match → review (none conf)
        "Maple Street Solar,,30,,,\n"                                              # missing name → blocked
    )
    r = _bulk_import(client, auth, "roster.csv", csv.encode(), "text/csv")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["dry_run"] is True
    rows = body["rows"]
    assert len(rows) == 4

    # Row 1: exact match to the array WITH a bill → ready.
    r1 = rows[0]
    assert r1["offtaker_name"] == "Jane Offtaker"
    assert r1["matched_array_id"] == aid1
    assert r1["matched_utility_account_id"] == ua1
    assert r1["confidence"] == "exact"
    assert abs(r1["allocation_pct"] - 0.25) < 1e-9
    assert abs(r1["discount_pct"] - 0.10) < 1e-9
    assert r1["email"] == "jane@example.com"
    assert r1["extra"].get("Notes") == "vip"   # unrecognized column preserved

    # Row 2: matched array but its account has NO settled bill → NOT ready.
    r2 = rows[1]
    assert r2["matched_array_id"] == aid2
    # confidence may be high (containment) but no bill → needs_review, not ready.

    # Row 3: garbage → no array bound.
    r3 = rows[2]
    assert r3["matched_array_id"] is None and r3["confidence"] == "none"

    # Row 4: missing offtaker name → blocked via `missing`.
    r4 = rows[3]
    assert "name" in r4["missing"]

    s = body["summary"]
    assert s["total"] == 4
    assert s["ready"] == 1           # only Jane
    assert s["blocked"] == 1         # the missing-name row
    assert s["needs_review"] == 2    # no-bill + no-match

    # Full pick-list returned for the frontend dropdowns.
    assert any(a["utility_account_id"] == ua1 for a in body["arrays"])


def test_bulk_import_xlsx_supported(client):
    from openpyxl import Workbook
    tid, auth = _make_tenant()
    aid, ua = _make_array_with_bill(tid, "Hilltop Farm", "GMP-777", with_bill=True)

    wb = Workbook()
    ws = wb.active
    ws.append(["Array", "Offtaker", "Share %"])
    ws.append(["Hilltop Farm", "Barn Co", 40])
    buf = io.BytesIO(); wb.save(buf)
    r = _bulk_import(
        client, auth, "roster.xlsx", buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert r.status_code == 200, r.text
    row = r.json()["rows"][0]
    assert row["matched_array_id"] == aid
    assert row["matched_utility_account_id"] == ua
    assert row["confidence"] == "exact"


def test_bulk_import_missing_required_column_422(client):
    tid, auth = _make_tenant()
    _make_array_with_bill(tid, "Some Array", "GMP-999")
    # No Array column → 422.
    csv = "Offtaker,Share %\nJane,25\n"
    r = _bulk_import(client, auth, "roster.csv", csv.encode(), "text/csv")
    assert r.status_code == 422
    assert "array" in r.json()["detail"].lower()


# ─── bulk-commit (decoupled write + idempotency) ─────────────────────────────

def test_bulk_commit_creates_and_is_idempotent(client):
    tid, auth = _make_tenant()
    aid, ua = _make_array_with_bill(tid, "Maple Street Solar", "GMP-111", with_bill=True)

    payload = {
        "rows": [
            {"offtaker_name": "Jane Offtaker", "array_id": aid,
             "utility_account_id": ua, "allocation_pct": 0.25,
             "email": "jane@example.com", "discount_pct": 0.10},
            {"offtaker_name": "Second Co", "array_id": aid,
             "utility_account_id": ua, "allocation_pct": 0.40},
        ],
        "cadence": "monthly", "delivery_mode": "approval",
    }
    r = client.post("/v1/array-operator/billing/subscriptions/bulk-commit",
                    json=payload, headers={"Authorization": auth})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 2
    assert body["skipped"] == []
    assert body["failed"] == []

    # The subs persisted, bound to the array + utility account.
    with SessionLocal() as db:
        subs = db.execute(select(BillingReportSubscription).where(
            BillingReportSubscription.tenant_id == tid)).scalars().all()
        assert len(subs) == 2
        for s in subs:
            assert s.utility_account_id == ua
            assert s.array_id == aid

    # Re-post the SAME batch → both skipped (idempotent, no duplicates).
    r2 = client.post("/v1/array-operator/billing/subscriptions/bulk-commit",
                     json=payload, headers={"Authorization": auth})
    assert r2.status_code == 200, r2.text
    b2 = r2.json()
    assert b2["created"] == 0
    assert len(b2["skipped"]) == 2
    with SessionLocal() as db:
        n = db.execute(select(BillingReportSubscription).where(
            BillingReportSubscription.tenant_id == tid)).scalars().all()
        assert len(n) == 2   # still 2, not 4


def test_bulk_commit_rejects_foreign_utility_account(client):
    tid, auth = _make_tenant()
    aid, ua = _make_array_with_bill(tid, "Mine", "GMP-111", with_bill=True)
    # Another tenant's account.
    other_tid, _ = _make_tenant()
    _, other_ua = _make_array_with_bill(other_tid, "Theirs", "GMP-222", with_bill=True)

    payload = {"rows": [
        {"offtaker_name": "X", "array_id": aid,
         "utility_account_id": other_ua, "allocation_pct": 0.5}]}
    r = client.post("/v1/array-operator/billing/subscriptions/bulk-commit",
                    json=payload, headers={"Authorization": auth})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] == 0
    assert len(body["failed"]) == 1
    assert "not found" in body["failed"][0]["error"]


# ─── array-first manual create ───────────────────────────────────────────────

def test_manual_create_array_first_resolves_utility_bill(client):
    """POST /subscriptions with array_id (no utility_account_id) resolves the
    array's utility bill and binds to it."""
    tid, auth = _make_tenant()
    aid, ua = _make_array_with_bill(tid, "Solo Array", "GMP-111", with_bill=True)

    r = client.post("/v1/array-operator/billing/subscriptions",
                    data={"customer_name": "Array First Co", "array_id": str(aid),
                          "allocation_pct": "0.5", "cadence": "monthly"},
                    headers={"Authorization": auth})
    assert r.status_code == 200, r.text
    sub = r.json()["subscription"]
    # Bound to the resolved utility account AND the array.
    assert sub["utility_account_id"] == ua
    assert sub["array_id"] == aid


def test_manual_create_array_multiple_bills_requires_override(client):
    """An array with two billed utility accounts is ambiguous → 400 asking for an
    explicit utility_account_id (never guess which bill)."""
    tid, auth = _make_tenant()
    with SessionLocal() as db:
        arr = Array(tenant_id=tid, name="Two Bills", fuel_type="solar")
        db.add(arr); db.flush(); aid = arr.id
        for n in ("A", "B"):
            acct = UtilityAccount(tenant_id=tid, array_id=aid, provider="gmp",
                                  account_number=f"GMP-{n}", nickname=f"Acct {n}")
            db.add(acct); db.flush()
            pe = date.today().replace(day=1) - timedelta(days=1)
            db.add(Bill(tenant_id=tid, account_id=acct.id, bill_date=pe,
                        period_start=pe.replace(day=1), period_end=pe,
                        kwh_generated=1000.0,
                        document_number="D-" + secrets.token_hex(3)))
        db.commit()

    r = client.post("/v1/array-operator/billing/subscriptions",
                    data={"customer_name": "Ambiguous Co", "array_id": str(aid),
                          "allocation_pct": "0.5"},
                    headers={"Authorization": auth})
    assert r.status_code == 400
    assert "multiple" in r.json()["detail"].lower()


# ─── format-agnostic detection (junk rows / weird headers / reordered) ───────

def test_bulk_import_messy_sheet_detected_by_content(client):
    """A roster with junk TITLE rows above the header, unhelpful headers ('Solar
    Site', 'Customer', '% of System'), and a NON-canonical column order still parses:
    the array column is found by CONTENT (its values are the tenant's real arrays)."""
    tid, auth = _make_tenant()
    aid1, ua1 = _make_array_with_bill(tid, "Maple Street Solar", "GMP-111", with_bill=True)
    aid2, ua2 = _make_array_with_bill(tid, "Route 7 Community Array", "GMP-222", with_bill=True)

    csv = (
        "Community Solar Roster — Q3 2026,,,\n"            # junk title row
        "Confidential — do not distribute,,,\n"           # junk subtitle
        ",,,\n"                                            # blank spacer
        "Solar Site,Customer,% of System,Contact\n"       # the REAL header (weird labels)
        "Maple Street Solar,Alice Cooper,25,alice@example.com\n"
        "Route 7 Community,Bob Dylan,15,bob@example.com\n"
    )
    r = _bulk_import(client, auth, "messy.csv", csv.encode(), "text/csv")
    assert r.status_code == 200, r.text
    body = r.json()

    # Detection block is surfaced for the frontend review UI.
    det = body["detection"]
    assert det["header_row"] == 3
    assert det["via"] in ("content", "mixed")
    cm = det["column_map"]
    assert cm["array_name"]["index"] == 0     # found by content, not header
    assert cm["offtaker_name"]["index"] == 1
    assert cm["allocation_pct"]["index"] == 2
    assert cm["email"]["index"] == 3

    # And the rows parsed correctly off that detected mapping.
    rows = body["rows"]
    assert len(rows) == 2
    assert rows[0]["offtaker_name"] == "Alice Cooper"
    assert rows[0]["matched_array_id"] == aid1
    assert rows[0]["confidence"] in ("exact", "high")
    assert abs(rows[0]["allocation_pct"] - 0.25) < 1e-9
    assert rows[0]["email"] == "alice@example.com"
    assert rows[1]["matched_array_id"] == aid2


def test_bulk_import_column_map_override(client):
    """An operator-confirmed column_map override parses by the given indices and
    SKIPS detection — this is how column corrections re-parse the sheet."""
    tid, auth = _make_tenant()
    aid, ua = _make_array_with_bill(tid, "Maple Street Solar", "GMP-111", with_bill=True)

    # Deliberately AMBIGUOUS headers so only an explicit override can parse it right.
    csv = (
        "colA,colB,colC\n"
        "Maple Street Solar,Jane Offtaker,25\n"
    )
    import json as _json
    override = _json.dumps({"array_name": 0, "offtaker_name": 1, "allocation_pct": 2})
    r = _bulk_import(client, auth, "blind.csv", csv.encode(), "text/csv",
                     column_map=override)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["detection"]["via"] == "override"
    row = body["rows"][0]
    assert row["offtaker_name"] == "Jane Offtaker"
    assert row["matched_array_id"] == aid
    assert abs(row["allocation_pct"] - 0.25) < 1e-9


# ─── template download ───────────────────────────────────────────────────────

def test_offtaker_template_download(client):
    r = client.get("/v1/array-operator/billing/offtaker-template.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "offtaker-template.xlsx" in r.headers.get("content-disposition", "")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert "Offtakers" in wb.sheetnames
    ws = wb["Offtakers"]
    # Ford 2026-07-09 richer roster format: the two account-number columns pin the
    # exact utility bill (master → array; offtaker's own → their sub-account).
    # Utility-agnostic headers (Ford multi-utility import) — not GMP-only labels.
    assert [c.value for c in ws[1]] == [
        "Master Array Account", "Master Utility Account Number",
        "Offtaker Utility Name", "Offtaker", "Share %", "Email",
        "Discount %", "Offtaker Account Number", "Budget Monthly ($)"]
