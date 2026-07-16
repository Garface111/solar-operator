"""Per-client raw utility generation export (Ford 2026-07-16).

GET /v1/account/clients/{id}/generation.xlsx?quarter= returns a workbook with a
Monthly Summary (projects x the quarter's months) across ALL the client's
utilities — GMP (bill + interval) and SmartHub co-ops (daily meter). Utility
generation only; never inverter telemetry.
"""
from __future__ import annotations
import io
import secrets
from datetime import date, datetime

from openpyxl import load_workbook

from api.account import mint_session_for_tenant
from api.db import SessionLocal
from api.models import Tenant, Client, Array, UtilityAccount, Bill, DailyGeneration


def _seed_client(gmp_kwh=30000):
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(id=tid, name="Export Co", contact_email=f"{tid}@t.test",
                      tenant_key="k_" + secrets.token_hex(8), plan="standard", active=True))
        c = Client(tenant_id=tid, name="Mixed Client", active=True)
        db.add(c); db.flush()
        # GMP project — generation in the bill's kwh_generated.
        gmp_arr = Array(tenant_id=tid, client_id=c.id, name="Londonderry", nepool_gis_id="GIS42")
        db.add(gmp_arr); db.flush()
        gmp_ua = UtilityAccount(tenant_id=tid, array_id=gmp_arr.id, provider="gmp",
                                account_number="G" + secrets.token_hex(3))
        db.add(gmp_ua); db.flush()
        for (y, m) in [(2026, 1), (2026, 2), (2026, 3)]:
            db.add(Bill(tenant_id=tid, account_id=gmp_ua.id, bill_date=datetime(y, m, 15),
                        period_start=datetime(y, m, 1), kwh_generated=gmp_kwh,
                        document_number=f"g-{tid}-{y}-{m}", parse_status="parsed"))
        # Co-op project — bills carry NO kwh_generated; generation is in the
        # SmartHub daily meter (DailyGeneration source='smarthub').
        vec_arr = Array(tenant_id=tid, client_id=c.id, name="Glover Coop")
        db.add(vec_arr); db.flush()
        db.add(UtilityAccount(tenant_id=tid, array_id=vec_arr.id, provider="vec",
                              account_number="V" + secrets.token_hex(3)))
        for dd in range(1, 32):  # January daily meter
            db.add(DailyGeneration(tenant_id=tid, array_id=vec_arr.id,
                                   day=date(2026, 1, dd), kwh=100.0, source="smarthub"))
        db.commit()
        return tid, c.id


def test_generation_export_covers_all_utilities(client):
    tid, cid = _seed_client(gmp_kwh=30000)
    auth = {"Authorization": f"Bearer {mint_session_for_tenant(tid)}"}
    r = client.get(f"/v1/account/clients/{cid}/generation.xlsx?quarter=Q1-2026", headers=auth)
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers["content-type"]

    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    assert "Monthly Summary" in wb.sheetnames
    sh = wb["Monthly Summary"]
    text = "\n".join(
        " ".join(str(sh.cell(row, col).value) for col in range(1, 8))
        for row in range(1, 16)
    )
    assert "Londonderry" in text        # GMP project (bill generation)
    assert "Glover Coop" in text        # co-op project (SmartHub meter) now INCLUDED
    assert "GMP" in text and "VEC" in text   # utility labels present
    assert "3100" in text or "90000" in text  # co-op Jan total 3100 and/or GMP quarter 90000
    wb.close()


def test_generation_directory_covers_all_clients(client):
    tid, _cid = _seed_client(gmp_kwh=30000)
    # a second client on the same tenant, with its own GMP project
    with SessionLocal() as db:
        c2 = Client(tenant_id=tid, name="Second Client", active=True)
        db.add(c2); db.flush()
        a2 = Array(tenant_id=tid, client_id=c2.id, name="Rooftop B")
        db.add(a2); db.flush()
        ua2 = UtilityAccount(tenant_id=tid, array_id=a2.id, provider="gmp",
                             account_number="G2" + secrets.token_hex(3))
        db.add(ua2); db.flush()
        db.add(Bill(tenant_id=tid, account_id=ua2.id, bill_date=datetime(2026, 2, 15),
                    period_start=datetime(2026, 2, 1), kwh_generated=5000,
                    document_number="g2-" + secrets.token_hex(4), parse_status="parsed"))
        db.commit()
    auth = {"Authorization": f"Bearer {mint_session_for_tenant(tid)}"}
    r = client.get("/v1/account/generation-directory.xlsx?quarter=Q1-2026", headers=auth)
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    sh = wb["Generation Summary"]
    text = "\n".join(" ".join(str(sh.cell(row, col).value) for col in range(1, 8))
                     for row in range(1, 20))
    assert "Mixed Client" in text and "Second Client" in text  # both clients present
    assert "Rooftop B" in text and "Londonderry" in text
    wb.close()


def test_generation_export_wrong_tenant_404(client):
    tid, cid = _seed_client()
    other = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(id=other, name="Other", contact_email=f"{other}@t.test",
                      tenant_key="k_" + secrets.token_hex(8), plan="standard", active=True))
        db.commit()
    auth = {"Authorization": f"Bearer {mint_session_for_tenant(other)}"}
    r = client.get(f"/v1/account/clients/{cid}/generation.xlsx?quarter=Q1-2026", headers=auth)
    assert r.status_code == 404
