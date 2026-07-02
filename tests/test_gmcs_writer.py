"""
Regression tests for api/writers/gmcs_writer.py — the "sacred" GMCS pixel format.

Invariants guarded here:
  - A1:C1 merged on every sheet
  - A1 value = "<Array Name> (<NEPOOL-GIS ID>)" or just "<Array Name>" when no ID
  - Row 5 header row: correct labels, bold, font size 14
  - RECs = int(MWh) — floor, not round (1999 kWh = 1.999 MWh → 1 REC)
  - Footnote text is byte-for-byte FOOTNOTE_TEXT
  - Arrays with excluded=True are never emitted
  - Column widths A–D all 24.0

Tests use synthetic operators; no Bruce-specific data.
"""
from __future__ import annotations

import math
import secrets
import tempfile
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from api.db import SessionLocal
from api.models import Array, Bill, Client, Tenant, UtilityAccount
from api.writers.gmcs_writer import FOOTNOTE_TEXT, build_workbook

# Reference date: Q2 2024 in progress → last complete quarter = Q1 2024.
# Rolling 6 quarters: Q4'22, Q1'23, Q2'23, Q3'23, Q4'23, Q1'24.
_REF = date(2024, 4, 1)


def _make_client_with_bills(
    kwh_per_month: int = 1999,
    months: list[tuple[int, int]] | None = None,
) -> tuple[str, int]:
    """Create Tenant → Client → Array (nepool_gis_id="GIS999") → UtilityAccount
    → Bills for `months` (default Q1 2024).  Returns (tenant_id, client_id)."""
    if months is None:
        months = [(2024, 1), (2024, 2), (2024, 3)]

    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        t = Tenant(
            id=tid, name="GMCS Regression Co",
            contact_email=f"{tid}@test.com",
            tenant_key="k_" + secrets.token_hex(8),
            plan="standard", active=True,
        )
        db.add(t); db.flush()

        c = Client(tenant_id=tid, name="Solar LLC", active=True)
        db.add(c); db.flush()

        arr = Array(
            tenant_id=tid, client_id=c.id,
            name="Hilltop Array", nepool_gis_id="GIS999",
        )
        db.add(arr); db.flush()

        ua = UtilityAccount(
            tenant_id=tid, array_id=arr.id,
            provider="gmp", account_number="ACC_" + secrets.token_hex(4),
        )
        db.add(ua); db.flush()

        for y, m in months:
            db.add(Bill(
                tenant_id=tid, account_id=ua.id,
                bill_date=datetime(y, m, 15),
                period_start=datetime(y, m, 1),
                kwh_generated=kwh_per_month,
                document_number=f"doc-{tid}-{y}-{m}",
                parse_status="parsed",
            ))

        db.commit()
        return tid, c.id


# ── A1:C1 merge ───────────────────────────────────────────────────────────────

def test_a1_c1_is_merged(tmp_path):
    _, cid = _make_client_with_bills()
    out = tmp_path / "merge.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    merge_ranges = {str(r) for r in wb.active.merged_cells.ranges}
    assert "A1:C1" in merge_ranges, f"A1:C1 missing; merges={merge_ranges}"


# ── A1 title format ───────────────────────────────────────────────────────────

def test_a1_title_includes_nepool_id(tmp_path):
    """Title = '<name> (<nepool_gis_id>)' when ID is set."""
    _, cid = _make_client_with_bills()
    out = tmp_path / "title_nepool.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.active["A1"].value == "Hilltop Array (GIS999)"


def test_a1_title_without_nepool_id(tmp_path):
    """Title = '<name>' alone when no NEPOOL-GIS ID."""
    tid = "ten_" + secrets.token_hex(6)
    out = tmp_path / "title_no_nepool.xlsx"
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="No NEPOOL Co", contact_email=f"{tid}@test.com",
            tenant_key="kn_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="No NEPOOL Client", active=True)
        db.add(c); db.flush()
        arr = Array(tenant_id=tid, client_id=c.id, name="Plain Array", nepool_gis_id=None)
        db.add(arr); db.flush()
        ua = UtilityAccount(
            tenant_id=tid, array_id=arr.id,
            provider="gmp", account_number="PLAIN_" + secrets.token_hex(4),
        )
        db.add(ua); db.flush()
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            bill_date=datetime(2024, 1, 15), period_start=datetime(2024, 1, 1),
            kwh_generated=1000, document_number="plain-doc-" + secrets.token_hex(4),
            parse_status="parsed",
        ))
        db.commit()
        cid = c.id

    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.active["A1"].value == "Plain Array"


# ── Row 5 header ──────────────────────────────────────────────────────────────

def test_row5_header_labels_and_font(tmp_path):
    _, cid = _make_client_with_bills()
    out = tmp_path / "header.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    labels = [sh.cell(5, col).value for col in range(1, 5)]
    assert labels == ["Quarter", "Generation (MWh)", "Reporting Amount", "RECs†"]
    assert sh.cell(5, 1).font.size == 14
    assert sh.cell(5, 1).font.bold is True


# ── RECs = floor(MWh) ────────────────────────────────────────────────────────

def test_recs_are_floor_of_mwh(tmp_path):
    """1999 kWh = 1.999 MWh → floor = 1 REC (not round = 2)."""
    _, cid = _make_client_with_bills(kwh_per_month=1999)
    out = tmp_path / "recs.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active

    found_data = False
    for row in range(7, 32):
        mwh = sh.cell(row, 2).value
        recs = sh.cell(row, 4).value
        if mwh is None:
            continue
        found_data = True
        assert recs == math.floor(mwh), (
            f"row {row}: mwh={mwh}, recs={recs}, expected floor={math.floor(mwh)}"
        )
        assert recs == 1  # 1999 kWh → int(1.999) = 1

    assert found_data, "no data rows found in workbook"


# ── Footnote verbatim ────────────────────────────────────────────────────────

def test_footnote_text_verbatim(tmp_path):
    """Footnote at row 31 must be byte-for-byte FOOTNOTE_TEXT."""
    _, cid = _make_client_with_bills()
    out = tmp_path / "footnote.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    # 6 quarters × (3 month rows + 1 gap) = 24 rows, starting at row 7 → ends at row 30
    # foot_row = 31 (pinned when row <= 31)
    assert sh.cell(31, 1).value == FOOTNOTE_TEXT, (
        f"Footnote mismatch at row 31.\n"
        f"Expected: {FOOTNOTE_TEXT!r}\n"
        f"Got:      {sh.cell(31, 1).value!r}"
    )


# ── Excluded arrays skipped ───────────────────────────────────────────────────

def test_excluded_array_not_in_workbook(tmp_path):
    """Array.excluded=True arrays must be absent from the workbook entirely."""
    tid = "ten_" + secrets.token_hex(6)
    out = tmp_path / "excluded.xlsx"
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Excluded Test Co", contact_email=f"{tid}@test.com",
            tenant_key="ke_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="Excl Client", active=True)
        db.add(c); db.flush()

        # Active array — should appear in workbook
        arr_active = Array(
            tenant_id=tid, client_id=c.id,
            name="Active Array", nepool_gis_id="ACT1", excluded=False,
        )
        db.add(arr_active); db.flush()
        ua_a = UtilityAccount(
            tenant_id=tid, array_id=arr_active.id,
            provider="gmp", account_number="ACT_" + secrets.token_hex(4),
        )
        db.add(ua_a); db.flush()
        db.add(Bill(
            tenant_id=tid, account_id=ua_a.id,
            bill_date=datetime(2024, 1, 15), period_start=datetime(2024, 1, 1),
            kwh_generated=2000, document_number="act-doc-" + secrets.token_hex(4),
            parse_status="parsed",
        ))

        # Excluded array — must NOT appear in workbook
        arr_excl = Array(
            tenant_id=tid, client_id=c.id,
            name="Excluded Array", nepool_gis_id="EXCL1", excluded=True,
        )
        db.add(arr_excl); db.flush()
        ua_e = UtilityAccount(
            tenant_id=tid, array_id=arr_excl.id,
            provider="gmp", account_number="EXCL_" + secrets.token_hex(4),
        )
        db.add(ua_e); db.flush()
        db.add(Bill(
            tenant_id=tid, account_id=ua_e.id,
            bill_date=datetime(2024, 1, 15), period_start=datetime(2024, 1, 1),
            kwh_generated=3000, document_number="excl-doc-" + secrets.token_hex(4),
            parse_status="parsed",
        ))
        db.commit()
        cid = c.id

    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Active Array"], f"unexpected sheets: {wb.sheetnames}"


# ── Non-producing arrays skipped ──────────────────────────────────────────────

def _add_array_with_account(db, tid: str, cid: int, name: str,
                            bill_months: list[tuple[int, int]] | None = None,
                            kwh: int = 2000) -> int:
    """Array + UtilityAccount (+ optional Bills). Returns array id."""
    arr = Array(tenant_id=tid, client_id=cid, name=name, nepool_gis_id=None)
    db.add(arr); db.flush()
    ua = UtilityAccount(
        tenant_id=tid, array_id=arr.id,
        provider="gmp", account_number="NP_" + secrets.token_hex(4),
    )
    db.add(ua); db.flush()
    for y, m in (bill_months or []):
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            bill_date=datetime(y, m, 15), period_start=datetime(y, m, 1),
            kwh_generated=kwh,
            document_number=f"np-{tid}-{name[:6]}-{y}-{m}",
            parse_status="parsed",
        ))
    return arr.id


def _make_mixed_client() -> int:
    """Client with one producing array (Q1-2024 bills), one array with an
    account but no bills at all, and one whose bills predate the window."""
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="NonProducing Test Co", contact_email=f"{tid}@test.com",
            tenant_key="knp_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="NP Client", active=True)
        db.add(c); db.flush()
        _add_array_with_account(db, tid, c.id, "Producing Array",
                                [(2024, 1), (2024, 2), (2024, 3)])
        _add_array_with_account(db, tid, c.id, "Never Producing Array", [])
        _add_array_with_account(db, tid, c.id, "Pre Window Array",
                                [(2021, 6), (2021, 7)])
        db.commit()
        return c.id


def test_nonproducing_arrays_get_no_sheet(tmp_path):
    """An array with zero generation in every window month is omitted — both
    the no-bills-at-all case and the bills-entirely-before-the-window case."""
    cid = _make_mixed_client()
    out = tmp_path / "nonproducing.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Producing Array"], f"sheets: {wb.sheetnames}"


def test_all_nonproducing_renders_no_data_stub(tmp_path):
    """When every array is non-producing the workbook is the honest stub, not
    a stack of blank per-array sheets."""
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="All Empty Co", contact_email=f"{tid}@test.com",
            tenant_key="kae_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="Empty Client", active=True)
        db.add(c); db.flush()
        _add_array_with_account(db, tid, c.id, "Dormant One", [])
        _add_array_with_account(db, tid, c.id, "Dormant Two", [(2021, 3)])
        db.commit()
        cid = c.id

    out = tmp_path / "all_empty.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.sheetnames == ["(no data)"], f"sheets: {wb.sheetnames}"


def test_daily_generation_only_array_keeps_sheet(tmp_path):
    """An array with NO bills but real DailyGeneration in the window must keep
    its sheet — even when a sibling array has bills (the old group list was
    built from bill-backed groups only and silently dropped it)."""
    from api.models import DailyGeneration

    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Daily Only Co", contact_email=f"{tid}@test.com",
            tenant_key="kdo_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="Daily Client", active=True)
        db.add(c); db.flush()
        _add_array_with_account(db, tid, c.id, "Billed Array",
                                [(2024, 1), (2024, 2), (2024, 3)])
        daily_arr_id = _add_array_with_account(db, tid, c.id, "Daily Array", [])
        for d in range(1, 11):
            db.add(DailyGeneration(
                tenant_id=tid, array_id=daily_arr_id,
                day=date(2024, 1, d), kwh=100.0, source="csv",
            ))
        db.commit()
        cid = c.id

    out = tmp_path / "daily_only.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert sorted(wb.sheetnames) == ["Billed Array", "Daily Array"], (
        f"sheets: {wb.sheetnames}"
    )
    # And the daily array's Jan-2024 month renders its real total (1000 kWh = 1 MWh).
    sh = wb["Daily Array"]
    mwh_values = [sh.cell(r, 2).value for r in range(7, 31)]
    assert 1.0 in mwh_values, f"expected 1.0 MWh in {mwh_values}"


# ── Column widths ─────────────────────────────────────────────────────────────

def test_column_widths_all_24(tmp_path):
    _, cid = _make_client_with_bills()
    out = tmp_path / "widths.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    for col in ("A", "B", "C", "D"):
        assert sh.column_dimensions[col].width == 24.0, (
            f"Column {col} width = {sh.column_dimensions[col].width}, expected 24.0"
        )
