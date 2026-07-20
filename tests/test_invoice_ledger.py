"""Default offtaker invoice ledger — generation + paid/collected spreadsheet.

Covers:
  * build_default_ledger columns + TOTAL collected
  * list_payment_rows money math (collected = amount − fee when paid)
  * ensure_default_ledger never overwrites a BYO (auto=False) sheet
  * GET /subscriptions/{id}/tracker always returns payments + auto ledger
  * GET /subscriptions/{id}/tracker/download streams xlsx with invoice columns
  * GET /subscriptions/{id}/payments roll-up
"""
from __future__ import annotations

import io
import secrets
from datetime import datetime
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from api.account import mint_session_for_tenant
from api.db import SessionLocal
from api.models import Tenant, BillingReportSubscription, OfftakerPayment


def _tenant(**kw) -> SimpleNamespace:
    tid = "ten_" + secrets.token_hex(5)
    defaults = dict(
        id=tid, name="Ledger Owner",
        contact_email=f"{tid}@owner.test",
        tenant_key="sol_live_" + secrets.token_urlsafe(10),
        plan="standard", active=True, product="array_operator",
    )
    defaults.update(kw)
    with SessionLocal() as db:
        t = Tenant(**defaults)
        db.add(t)
        db.commit()
        return SimpleNamespace(id=t.id)


def _sub(tenant_id: str, **kw) -> int:
    defaults = dict(
        tenant_id=tenant_id,
        customer_name="Town of Ledger",
        client_email="ledger@example.com",
        send_mode="to_client",
        allocation_pct=0.25,
        billing_model="percent_of_array",
        formats=["pdf"],
        enabled=True,
    )
    defaults.update(kw)
    with SessionLocal() as db:
        s = BillingReportSubscription(**defaults)
        db.add(s)
        db.commit()
        db.refresh(s)
        return s.id


def _pay(tenant_id: str, sid: int, **kw) -> int:
    defaults = dict(
        tenant_id=tenant_id,
        subscription_id=sid,
        invoice_number="2026-05",
        period_key="2026-05-31",
        amount_cents=10_000,
        fee_cents=50,
        status="paid",
        customer_name="Town of Ledger",
        paid_at=datetime(2026, 6, 5, 12, 0, 0),
    )
    defaults.update(kw)
    with SessionLocal() as db:
        p = OfftakerPayment(**defaults)
        db.add(p)
        db.commit()
        db.refresh(p)
        return p.id


def _auth(tid: str) -> dict:
    return {"Authorization": f"Bearer {mint_session_for_tenant(tid)}"}


# ─── unit: list + build ─────────────────────────────────────────────────────

def test_list_payment_rows_collected_only_when_paid():
    from api.billing.invoice_ledger import list_payment_rows
    t = _tenant()
    sid = _sub(t.id)
    _pay(t.id, sid, status="paid", amount_cents=10_000, fee_cents=50,
         period_key="2026-05-31", invoice_number="2026-05")
    _pay(t.id, sid, status="open", amount_cents=8000, fee_cents=40,
         period_key="2026-06-30", invoice_number="2026-06", paid_at=None)
    with SessionLocal() as db:
        sub = db.get(BillingReportSubscription, sid)
        rows = list_payment_rows(db, sub)
    assert len(rows) == 2
    # newest first
    assert rows[0]["period_label"] == "2026-06"
    assert rows[0]["status"] == "open"
    assert rows[0]["collected_usd"] is None
    assert rows[0]["amount_usd"] == 80.0
    assert rows[1]["status"] == "paid"
    assert rows[1]["collected_usd"] == 99.50  # 100 − 0.50
    assert rows[1]["fee_usd"] == 0.50
    assert rows[1]["paid_at"] is not None


def test_build_default_ledger_xlsx_has_collection_columns():
    from api.billing.invoice_ledger import build_default_ledger
    t = _tenant()
    sid = _sub(t.id)
    _pay(t.id, sid, status="paid", amount_cents=5190, fee_cents=26,
         period_key="2026-04-30", invoice_number="2026-04",
         paid_at=datetime(2026, 5, 2, 15, 0, 0))
    _pay(t.id, sid, status="open", amount_cents=6000, fee_cents=30,
         period_key="2026-05-31", invoice_number="2026-05", paid_at=None)
    with SessionLocal() as db:
        sub = db.get(BillingReportSubscription, sid)
        blob, mapping = build_default_ledger(db, sub)
    assert mapping["auto"] is True
    assert mapping["data_rows"] == 2
    assert "Collected $" in mapping["headers"]
    assert "Paid date" in mapping["headers"]
    wb = load_workbook(io.BytesIO(blob))
    ws = wb.active
    assert ws.title == "Invoice ledger"
    headers = [c.value for c in ws[1]]
    assert headers == [
        "Period", "Generation kWh", "Invoice $", "Status",
        "Paid date", "Collected $", "Platform fee $", "Invoice #",
    ]
    # oldest first in sheet
    assert ws.cell(2, 1).value == "2026-04"
    assert ws.cell(2, 4).value == "Paid"
    assert ws.cell(2, 5).value == "2026-05-02"
    assert float(ws.cell(2, 6).value) == pytest.approx(51.64, abs=0.01)  # 51.90 − 0.26
    assert ws.cell(3, 1).value == "2026-05"
    assert ws.cell(3, 4).value == "Awaiting payment"
    assert ws.cell(3, 6).value in ("", None)
    # TOTAL collected row
    last = ws.max_row
    assert "TOTAL" in str(ws.cell(last, 1).value or "").upper()
    assert float(ws.cell(last, 6).value) == pytest.approx(51.64, abs=0.01)


def test_ensure_default_preserves_byo_sheet():
    from api.billing.invoice_ledger import ensure_default_ledger, build_default_ledger
    t = _tenant()
    sid = _sub(t.id)
    _pay(t.id, sid)
    with SessionLocal() as db:
        sub = db.get(BillingReportSubscription, sid)
        # Seed a fake BYO sheet (auto=False)
        sub.tracker_workbook = b"PK\x03\x04fake-byo"
        sub.tracker_filename = "my_custom.xlsx"
        sub.tracker_map = {"ok": True, "auto": False, "headers": ["A", "B"], "columns": {"period": 0, "generation": 1}}
        db.add(sub)
        db.commit()
        db.refresh(sub)
        res = ensure_default_ledger(db, sub)
        assert res.get("preserved") is True
        assert res.get("auto") is False
        assert bytes(sub.tracker_workbook) == b"PK\x03\x04fake-byo"

        # Auto sheet gets rebuilt
        sub.tracker_map = {"ok": True, "auto": True}
        sub.tracker_workbook = b"old"
        db.add(sub)
        db.commit()
        db.refresh(sub)
        res2 = ensure_default_ledger(db, sub)
        assert res2.get("auto") is True
        assert res2.get("has_sheet") is True
        assert bytes(sub.tracker_workbook).startswith(b"PK")  # real xlsx


# ─── HTTP surface ───────────────────────────────────────────────────────────

def test_tracker_status_includes_payments_and_auto_ledger(client, monkeypatch):
    monkeypatch.setenv("SPREADSHEET_TRACKER_ENABLED", "true")
    t = _tenant()
    sid = _sub(t.id)
    _pay(t.id, sid, status="paid", amount_cents=10_000, fee_cents=50,
         period_key="2026-05-31", invoice_number="2026-05")
    _pay(t.id, sid, status="open", amount_cents=2000, fee_cents=10,
         period_key="2026-06-30", invoice_number="2026-06", paid_at=None)
    r = client.get(
        f"/v1/array-operator/billing/subscriptions/{sid}/tracker",
        headers=_auth(t.id),
    )
    assert r.status_code == 200, r.text
    tr = r.json()["tracker"]
    assert tr["enabled"] is True
    assert tr["has_sheet"] is True
    assert tr["auto"] is True
    assert tr["payments_paid"] == 1
    assert tr["payments_open"] == 1
    assert tr["collected_usd"] == 99.50
    assert len(tr["payments"]) == 2
    assert "Collected $" in (tr.get("headers") or [])


def test_tracker_download_default_ledger(client, monkeypatch):
    monkeypatch.setenv("SPREADSHEET_TRACKER_ENABLED", "true")
    t = _tenant()
    sid = _sub(t.id)
    _pay(t.id, sid, status="paid", amount_cents=5000, fee_cents=25)
    r = client.get(
        f"/v1/array-operator/billing/subscriptions/{sid}/tracker/download",
        headers=_auth(t.id),
    )
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in (r.headers.get("content-type") or "")
    wb = load_workbook(io.BytesIO(r.content))
    headers = [c.value for c in wb.active[1]]
    assert "Collected $" in headers
    assert "Invoice $" in headers


def test_subscription_payments_endpoint(client):
    t = _tenant()
    sid = _sub(t.id)
    _pay(t.id, sid, status="paid", amount_cents=10_000, fee_cents=50)
    r = client.get(
        f"/v1/array-operator/billing/subscriptions/{sid}/payments",
        headers=_auth(t.id),
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    assert body["count"] == 1
    assert body["collected_usd"] == 99.50
    assert body["paid_count"] == 1
    assert body["payments"][0]["status"] == "paid"


def test_tracker_status_works_even_when_flag_off(client, monkeypatch):
    """Invoice collection ledger is always available (not gated by BYO flag)."""
    monkeypatch.setenv("SPREADSHEET_TRACKER_ENABLED", "false")
    t = _tenant()
    sid = _sub(t.id)
    _pay(t.id, sid, status="paid", amount_cents=1000, fee_cents=5)
    r = client.get(
        f"/v1/array-operator/billing/subscriptions/{sid}/tracker",
        headers=_auth(t.id),
    )
    assert r.status_code == 200, r.text
    tr = r.json()["tracker"]
    assert tr["enabled"] is True
    assert tr["has_sheet"] is True
    assert tr["auto"] is True
    assert tr["collected_usd"] == 9.95
