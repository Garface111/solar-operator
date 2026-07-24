"""Quarterly summary spreadsheet — Name | Account # | months | Total.

Matches the Crown/GMP operator screenshot (Ford 2026-07-23).
"""
from __future__ import annotations

import secrets
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from api.account import mint_session_for_tenant
from api.db import SessionLocal
from api.models import (
    Array,
    Client,
    DailyGeneration,
    GmpDailyGeneration,
    Inverter,
    Tenant,
    UtilityAccount,
)
from api.writers.gmp_raw_writer import (
    build_all_clients_generation_workbook,
    build_quarterly_summary_workbook,
)


def _seed_tenant_with_gmp() -> tuple[str, int]:
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Summary Co", contact_email=f"{tid}@t.test",
            tenant_key="k_" + secrets.token_hex(8), plan="standard", active=True,
            generation_reports=True, product="array_operator",
        ))
        c = Client(tenant_id=tid, name="Client A", active=True)
        db.add(c)
        db.flush()
        arr = Array(
            tenant_id=tid, name="Owners Mill", client_id=c.id, fuel_type="solar",
        )
        db.add(arr)
        db.flush()
        ua = UtilityAccount(
            tenant_id=tid, provider="gmp", account_number="5232799117",
            array_id=arr.id,
        )
        db.add(ua)
        db.flush()
        # Q1 2026 — Jan 10, Feb 5, Mar 3 kWh across a few days
        for d, kwh in (
            (date(2026, 1, 10), 4.0),
            (date(2026, 1, 20), 3.0),  # Jan total 7
            (date(2026, 2, 5), 2.5),
            (date(2026, 3, 1), 1.25),
            (date(2026, 3, 15), 1.25),  # Mar total 2.5
        ):
            db.add(GmpDailyGeneration(
                tenant_id=tid, account_id=ua.id, account_number=ua.account_number,
                array_id=arr.id, day=d, kwh=kwh, interval_count=96, source="gmp_api",
                derived_at=datetime.utcnow(),
            ))
        # Second array, two accounts (multi-meter)
        arr2 = Array(
            tenant_id=tid, name="Starlake", client_id=c.id, fuel_type="solar",
        )
        db.add(arr2)
        db.flush()
        for acct, kwh in (("111", 10.0), ("222", 5.0)):
            ua2 = UtilityAccount(
                tenant_id=tid, provider="gmp", account_number=acct,
                array_id=arr2.id,
            )
            db.add(ua2)
            db.flush()
            db.add(GmpDailyGeneration(
                tenant_id=tid, account_id=ua2.id, account_number=acct,
                array_id=arr2.id, day=date(2026, 1, 5), kwh=kwh,
                interval_count=96, source="gmp_api",
                derived_at=datetime.utcnow(),
            ))
        db.commit()
        return tid, c.id


def test_build_quarterly_summary_format(tmp_path: Path):
    tid, _ = _seed_tenant_with_gmp()
    out = tmp_path / "summary.xlsx"
    build_quarterly_summary_workbook(tid, out, year=2026, quarter=1)
    assert out.exists()

    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary"]
    sh = wb["Summary"]
    headers = [sh.cell(1, c).value for c in range(1, 7)]
    assert headers == ["Name", "Account #", "January", "February", "March", "Total"]

    # Find Owners Mill row
    mill = None
    for r in range(2, 20):
        if sh.cell(r, 1).value == "Owners Mill":
            mill = r
            break
    assert mill is not None
    assert str(sh.cell(mill, 2).value) == "5232799117"
    assert float(sh.cell(mill, 3).value) == 7.0   # Jan
    assert float(sh.cell(mill, 4).value) == 2.5   # Feb
    assert float(sh.cell(mill, 5).value) == 2.5   # Mar
    assert float(sh.cell(mill, 6).value) == 12.0  # Total

    # Multi-meter array → one row per account
    starlake_accts = []
    for r in range(2, 20):
        if sh.cell(r, 1).value == "Starlake":
            starlake_accts.append(str(sh.cell(r, 2).value))
    assert sorted(starlake_accts) == ["111", "222"]


def _seed_vendor_only() -> tuple[str, int]:
    """Tenant with a Locus-monitored array (no utility account) + an empty
    Fronius twin (inverter, but no production)."""
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Vendor Co", contact_email=f"{tid}@t.test",
            tenant_key="k_" + secrets.token_hex(8), plan="standard", active=True,
            generation_reports=True, product="array_operator",
        ))
        c = Client(tenant_id=tid, name="Loc Client", active=True)
        db.add(c)
        db.flush()
        # Monitored vendor-only array: a Locus inverter + real daily production,
        # NO utility account. Must appear in the summary (blank account #).
        mon = Array(tenant_id=tid, name="Benson Site", client_id=c.id, fuel_type="solar")
        db.add(mon)
        db.flush()
        db.add(Inverter(
            tenant_id=tid, array_id=mon.id, vendor="locus",
            serial="LOC-1", nameplate_kw=10.0,
        ))
        for d, kwh in (
            (date(2026, 1, 10), 40.0),
            (date(2026, 2, 5), 25.0),
            (date(2026, 3, 3), 12.5),
        ):
            db.add(DailyGeneration(
                tenant_id=tid, array_id=mon.id, day=d, kwh=kwh, source="locus",
            ))
        # Empty vendor twin: an inverter but no production of its own → excluded,
        # so it can't double-report next to a real sibling.
        twin = Array(
            tenant_id=tid, name="Benson Site (Fronius)", client_id=c.id,
            fuel_type="solar",
        )
        db.add(twin)
        db.flush()
        db.add(Inverter(
            tenant_id=tid, array_id=twin.id, vendor="fronius",
            serial="FRO-1", nameplate_kw=10.0,
        ))
        db.commit()
        return tid, c.id


def test_quarterly_summary_includes_monitored_vendor_array(tmp_path: Path):
    tid, _ = _seed_vendor_only()
    out = tmp_path / "summary.xlsx"
    build_quarterly_summary_workbook(tid, out, year=2026, quarter=1)

    wb = load_workbook(out)
    sh = wb["Summary"]
    rows = {sh.cell(r, 1).value: r for r in range(2, 30) if sh.cell(r, 1).value}

    # The Locus-monitored array appears — blank account #, monitoring kWh.
    assert "Benson Site" in rows
    br = rows["Benson Site"]
    assert (sh.cell(br, 2).value or "") == ""
    assert float(sh.cell(br, 3).value) == 40.0   # Jan
    assert float(sh.cell(br, 4).value) == 25.0   # Feb
    assert float(sh.cell(br, 5).value) == 12.5   # Mar
    assert float(sh.cell(br, 6).value) == 77.5   # Total

    # The empty vendor twin (no production) is excluded.
    assert "Benson Site (Fronius)" not in rows


def test_generation_directory_includes_monitored_vendor(tmp_path: Path):
    """The all-clients generation directory (Download all generation) also shows
    monitored vendor arrays, labeled with the monitoring vendor as source."""
    tid, _ = _seed_vendor_only()
    out = tmp_path / "gen-dir.xlsx"
    build_all_clients_generation_workbook(tid, out, year=2026, quarter=1)

    wb = load_workbook(out)
    sh = wb["Generation Summary"]
    projects = {}
    for row in sh.iter_rows(min_row=5, values_only=True):
        if row and row[1]:  # a project row (col B = project name)
            projects[row[1]] = row[-1]  # project -> "utility · source"

    assert "Benson Site" in projects
    assert "monitor" in (projects["Benson Site"] or "")
    assert "LOCUS" in (projects["Benson Site"] or "")
    # Empty vendor twin still excluded here too.
    assert "Benson Site (Fronius)" not in projects


def test_quarterly_summary_endpoint(client):
    tid, _ = _seed_tenant_with_gmp()
    auth = {"Authorization": f"Bearer {mint_session_for_tenant(tid)}"}
    r = client.get(
        "/v1/account/quarterly-summary.xlsx?quarter=Q1-2026",
        headers=auth,
    )
    assert r.status_code == 200, r.text
    assert "spreadsheet" in r.headers.get("content-type", "")
    assert "quarterly-summary" in r.headers.get("content-disposition", "").lower()
