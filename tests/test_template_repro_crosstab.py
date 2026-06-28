"""Cross-tab template repro (Paul's templates pull from a data tab).

The invoice sheet must render without breaking references into other tabs. The
isolation step KEEPS the other sheets (hidden) so cross-tab formulas resolve on
the headless render — instead of deleting them (which turned `=+NFD!G6` into
#REF!) or nulling uncached cross-sheet refs (which silently dropped real data).
"""
import datetime
import io
import os

import openpyxl
import pytest

from api.billing.repro.template_repro import (
    _isolate_to_invoice_sheet,
    _autofit_columns,
    _prefill_volatile_crosstab,
    _resolve_referenced_cell_value,
    build_template_cell_map,
    offtaker_values_from_match,
    _fill_template_cells,
)


def _wb_with_crosstab():
    wb = openpyxl.Workbook()
    inv = wb.active
    inv.title = "Invoice"
    inv["A1"] = "Amount:"
    inv["B1"] = "=Data!A1"      # cross-tab ref, no cached value (fresh workbook)
    inv["B2"] = "=B1*2"         # intra-sheet formula
    data = wb.create_sheet("Data")
    data["A1"] = 500
    wb.create_sheet("Trend")["A1"] = 1
    return wb


def test_isolate_keeps_data_sheets_hidden_not_deleted():
    wb = _wb_with_crosstab()
    ws = wb["Invoice"]
    buf = io.BytesIO(); wb.save(buf)
    _isolate_to_invoice_sheet(wb, ws, buf.getvalue())

    # The data tabs are KEPT (so cross-tab formulas can still resolve), just hidden.
    assert "Data" in wb.sheetnames and "Trend" in wb.sheetnames
    assert wb["Data"].sheet_state == "hidden"
    assert wb["Trend"].sheet_state == "hidden"
    assert ws.sheet_state == "visible"
    # Active sheet is the visible invoice — never a hidden one.
    assert wb.active.title == "Invoice"


def test_isolate_does_not_null_uncached_cross_sheet_refs():
    """A cross-sheet formula with no Excel cache keeps its formula (resolves live
    against the hidden sheet) — it is NOT nulled, which was the data-loss bug."""
    wb = _wb_with_crosstab()
    ws = wb["Invoice"]
    buf = io.BytesIO(); wb.save(buf)
    _isolate_to_invoice_sheet(wb, ws, buf.getvalue())
    assert ws["B1"].value == "=Data!A1"   # preserved, not None


def test_autofit_widens_narrow_date_and_number_columns():
    """A date/number in a too-narrow column renders as '###' (it can't overflow into
    an empty neighbor the way text does). _autofit_columns must widen the column to
    fit — even when the cell is still a LIVE formula (=TODAY(), the Invoice/Due Date
    case) and even when an empty neighbor exists that text could have borrowed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions["B"].width = 4.0          # too narrow for a date
    ws.column_dimensions["D"].width = 3.0          # too narrow for a big number
    d = ws["B2"]; d.value = "=TODAY()"; d.number_format = "m/d/yyyy"   # live date formula
    n = ws["D2"]; n.value = 1234567.89; n.number_format = "#,##0.00"   # wide number value
    # C2 / E2 left EMPTY on purpose — a date/number can't borrow them.
    _autofit_columns(ws)
    assert ws.column_dimensions["B"].width > 8.0, ws.column_dimensions["B"].width
    assert ws.column_dimensions["D"].width > 8.0, ws.column_dimensions["D"].width


def test_autofit_gives_substituted_font_dates_a_generous_width():
    """The actual Paul bug: a date in 'Chalkboard' (a font the renderer lacks) is
    substituted with a WIDER font, so it overflows a column that looks wide enough by
    Excel's metric and renders '###'. _autofit_columns gives a non-standard-font date a
    generous allowance — comfortably more than the same date in a standard font."""
    from openpyxl.styles import Font

    def width_after(font_name):
        wb = openpyxl.Workbook(); ws = wb.active
        ws.column_dimensions["E"].width = 13.5           # Paul's actual column width
        c = ws["E2"]; c.value = "=TODAY()"; c.number_format = "mm-dd-yy"
        c.font = Font(name=font_name, size=16.0)
        _autofit_columns(ws)
        return ws.column_dimensions["E"].width

    chalk = width_after("Chalkboard")     # substituted → generous
    calibri = width_after("Calibri")      # standard → tuned
    assert chalk > 25.0, chalk            # wide enough for a 16pt substituted date
    assert chalk > calibri                # more than the standard-font allowance


def test_isolate_keeps_sheets_hidden_clears_print_area_and_strips_charts():
    """Every other tab is KEPT (hidden) so cross-tab refs still resolve (a template's
    'Total Array KwH'/'Net Rate' cells pull from the data tab — deleting it makes them
    #REF!). On each kept tab: the print area is cleared (so a hidden SECOND invoice sheet
    can't export a DUPLICATE page) and charts are STRIPPED (openpyxl re-saves chart XML the
    headless renderer REJECTS → the whole convert fails → the caller falls back to the lossy
    token-HTML invoice with the template's frozen sample numbers + blank dates)."""
    from openpyxl.chart import BarChart, Reference
    wb = openpyxl.Workbook()
    inv = wb.active; inv.title = "Invoice"; inv["A1"] = "real"
    dup = wb.create_sheet("Invoice Copy"); dup["A1"] = "dupe"; dup.print_area = "A1:B5"
    trend = wb.create_sheet("Trend"); trend["A1"] = 1; trend["A2"] = 2
    ch = BarChart(); ch.add_data(Reference(trend, min_col=1, min_row=1, max_row=2))
    trend.add_chart(ch, "C1")
    buf = io.BytesIO(); wb.save(buf)
    _isolate_to_invoice_sheet(wb, inv, buf.getvalue())
    assert wb["Invoice Copy"].sheet_state == "hidden"
    assert not wb["Invoice Copy"].print_area          # cleared → won't render a duplicate
    assert wb["Trend"].sheet_state == "hidden"
    assert len(wb["Trend"]._charts) == 0              # chart stripped → won't break the render
    assert inv.sheet_state == "visible"


# --- Volatile cross-tab pre-resolution (the Glover "KWH: 0 / 00:00:00" regression) ------
#
# The HCT invoice sheets pull Total Array KwH + the period dates from a Data tab with a
# VOLATILE array formula `{=INDIRECT(ADDRESS(N7,col,,,"Data"))}` that carries no Excel
# cache. Local LibreOffice recalcs volatile cells on load so they resolved here, but prod
# Gotenberg does NOT recalc a volatile formula into a HIDDEN sheet → it rendered 0 (kWh)
# and date-serial-0 "00:00:00" (the dates). The fix bakes those cells to literals BEFORE
# the data tab is hidden, so the render never needs a volatile recalc.

def _hct_shaped_wb():
    """A minimal workbook with the HCT shape: an Invoice sheet whose kWh + period-date
    cells are `INDIRECT(ADDRESS(N7,col,,,"Data"))` ARRAY formulas (no cache), and a Data
    tab whose kWh cell is itself an uncached arithmetic chain (=raw*$factor)."""
    from openpyxl.worksheet.formula import ArrayFormula
    wb = openpyxl.Workbook()
    inv = wb.active
    inv.title = "Template"
    inv["B7"] = 3                                       # N7-style row pointer → Data row 3
    inv["N7"] = 3
    # KWH cell: pulls Data col 5 (E3); From/To: Data cols 2/3 (B3/C3)
    inv["G19"] = ArrayFormula("G19", '=INDIRECT(ADDRESS(N7,5,,,"Data"))')
    inv["G19"].number_format = "0.0"
    inv["E16"] = ArrayFormula("E16", '=INDIRECT(ADDRESS(N7,2,,,"Data"))')
    inv["E16"].number_format = "mm-dd-yy"
    inv["G16"] = ArrayFormula("G16", '=INDIRECT(ADDRESS(N7,3,,,"Data"))')
    inv["G16"].number_format = "mm-dd-yy"
    data = wb.create_sheet("Data")
    data["B3"] = datetime.datetime(2026, 4, 24)        # From (literal datetime, cached)
    data["C3"] = datetime.datetime(2026, 5, 26)        # To
    data["D3"] = 18160                                  # raw kWh (literal)
    data["E3"] = "=+D3*0.95"                            # KWH = raw * factor (uncached formula)
    return wb


def test_prefill_resolves_indirect_address_chain():
    """The bounded resolver evaluates a shallow `=raw*factor` data-tab chain to the exact
    value (18160 * 0.95 = 17252), and reads literal cached dates straight through."""
    wb = _hct_shaped_wb()
    data = wb["Data"]
    assert _resolve_referenced_cell_value(data, data, "E3") == 17252.0
    assert _resolve_referenced_cell_value(data, data, "B3") == datetime.datetime(2026, 4, 24)


def test_prefill_volatile_crosstab_bakes_kwh_and_dates_to_literals():
    """After pre-resolution the invoice sheet's kWh + date cells are LITERAL values
    (non-zero kWh, real datetimes) — NOT formulas — so a render that does no volatile
    recalc still shows them. This is the exact Glover "Total Array KwH: 0 / 00:00:00" fix."""
    wb = _hct_shaped_wb()
    inv = wb["Template"]
    buf = io.BytesIO(); wb.save(buf)
    n = _prefill_volatile_crosstab(wb, inv, buf.getvalue())
    assert n == 3                                       # kWh + From + To all resolved
    assert inv["G19"].value == 17252.0                 # was INDIRECT(...) → would render 0
    assert inv["E16"].value == datetime.datetime(2026, 4, 24)   # was → "00:00:00"
    assert inv["G16"].value == datetime.datetime(2026, 5, 26)
    # none are formulas any more
    for coord in ("G19", "E16", "G16"):
        v = inv[coord].value
        assert not (isinstance(v, str) and v.startswith("=")), coord


def test_prefill_leaves_unresolvable_and_non_indirect_formulas_untouched():
    """Fail-safe: a cell that ISN'T the INDIRECT(ADDRESS(...)) shape, or whose chain can't
    be evaluated safely, is left exactly as-is — we never substitute a guessed number."""
    from openpyxl.worksheet.formula import ArrayFormula
    wb = openpyxl.Workbook()
    inv = wb.active; inv.title = "Template"; inv["N7"] = 3
    inv["A1"] = "=TODAY()"                              # plain live formula — not our shape
    inv["A2"] = "=SUM(Data!A1:A9)"                      # cross-tab but not INDIRECT(ADDRESS)
    inv["A3"] = ArrayFormula("A3", '=INDIRECT(ADDRESS(N7,5,,,"Data"))')
    data = wb.create_sheet("Data")
    data["E3"] = "=VLOOKUP(1,X:Y,2)"                    # unsafe chain → resolver bails
    buf = io.BytesIO(); wb.save(buf)
    n = _prefill_volatile_crosstab(wb, inv, buf.getvalue())
    assert n == 0
    assert inv["A1"].value == "=TODAY()"
    assert inv["A2"].value == "=SUM(Data!A1:A9)"
    assert getattr(inv["A3"].value, "text", inv["A3"].value) == '=INDIRECT(ADDRESS(N7,5,,,"Data"))'


# The real HCT template, when present locally (not in CI). Renders nothing — asserts on
# the FILLED xlsx cells so it needs no PDF parsing / renderer.
_REAL_TPL = "/mnt/c/Users/fordg/Downloads/Danville Big Buck Invoice - HCT - NEW (1) (2).xlsx"


@pytest.mark.skipif(not os.path.exists(_REAL_TPL), reason="real HCT template not present")
def test_real_hct_template_kwh_and_dates_nonzero_in_filled_xlsx():
    """End-to-end on Paul's real 'Valley Cares Template': after the fill+isolate pipeline,
    Total Array KwH (G19) is a non-zero number and the From/To dates (E16/G16) are real
    datetimes in the FILLED workbook — i.e. they would NOT render 0 / 00:00:00."""
    b = open(_REAL_TPL, "rb").read()

    class _Period:
        start = datetime.date(2026, 5, 21)
        end = datetime.date(2026, 6, 21)
        customer_kwh = 46.0

    class _Match:
        latest_period = _Period()
        computed_invoice = {"amount_owed": 8.36, "kwh": 46.0, "invoice_number": "GLOVER-001"}

    cm = build_template_cell_map(b)
    vals = offtaker_values_from_match(_Match())
    filled = _fill_template_cells(b, cm, vals, "Town of Glover")
    assert filled
    wb = openpyxl.load_workbook(io.BytesIO(filled))
    ws = wb[cm["sheet"]]
    kwh = ws["G19"].value
    assert isinstance(kwh, (int, float)) and kwh > 0, kwh                  # not 0
    assert isinstance(ws["E16"].value, datetime.datetime), ws["E16"].value  # not 00:00:00
    assert isinstance(ws["G16"].value, datetime.datetime), ws["G16"].value
