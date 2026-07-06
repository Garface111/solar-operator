"""
Regression tests for the NEPOOL-GIS reporting-quarter lag.

NEPOOL-GIS issues RECs ~2 quarters after the generation quarter ends (Q1
generation is uploaded the following July). So an automated REC report must
mirror the quarter the REC agent is *currently minting* — two quarters before
the in-progress quarter — not the most recently completed quarter.

These pin `default_reporting_reference_date` and its composition with the
rolling-window builder, plus the end-to-end default inside build_workbook.
Synthetic operators only; no Bruce-specific data.
"""
from __future__ import annotations

import secrets
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

from api.db import SessionLocal
from api.models import Array, Bill, Client, Tenant, UtilityAccount
import api.writers.gmcs_writer as gw
from api.writers.gmcs_writer import (
    build_workbook,
    default_reporting_reference_date,
    _rolling_quarters,
)


# (today, expected minting quarter that the window must terminate on)
_CASES = [
    (date(2026, 7, 6), (2026, 1)),    # Crown's real case: July → Q1 2026
    (date(2026, 7, 31), (2026, 1)),   # stable across the whole quarter
    (date(2026, 10, 15), (2026, 2)),  # October → Q2 2026
    (date(2027, 1, 10), (2026, 3)),   # January → Q3 (prior year)
    (date(2026, 4, 5), (2025, 4)),    # April → Q4 (prior year), crosses year
]


def test_default_reference_terminates_on_minting_quarter():
    for today, expected_q in _CASES:
        ref = default_reporting_reference_date(today)
        window = _rolling_quarters(ref, count=6)
        assert window[-1] == expected_q, (
            f"today={today}: window ends {window[-1]}, expected {expected_q}")


def test_minting_quarter_is_one_behind_last_complete():
    """The lagged default must be exactly one quarter earlier than the old
    'last complete quarter' behaviour — a guard against regressing to it."""
    today = date(2026, 7, 6)
    lagged = _rolling_quarters(default_reporting_reference_date(today), 6)[-1]
    last_complete = _rolling_quarters(today, 6)[-1]
    assert last_complete == (2026, 2)   # old behaviour
    assert lagged == (2026, 1)          # new behaviour


def _make_client_producing_in(months: list[tuple[int, int]]) -> int:
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Lag Test Co", contact_email=f"{tid}@test.com",
            tenant_key="klag_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="Lag Client", active=True)
        db.add(c); db.flush()
        arr = Array(tenant_id=tid, client_id=c.id, name="Lag Array",
                    nepool_gis_id="LAG1")
        db.add(arr); db.flush()
        ua = UtilityAccount(tenant_id=tid, array_id=arr.id, provider="gmp",
                            account_number="LAG_" + secrets.token_hex(4))
        db.add(ua); db.flush()
        for y, m in months:
            db.add(Bill(
                tenant_id=tid, account_id=ua.id,
                bill_date=datetime(y, m, 15), period_start=datetime(y, m, 1),
                kwh_generated=3000,
                document_number=f"lag-{y}-{m}-{secrets.token_hex(3)}",
                parse_status="parsed",
            ))
        db.commit()
        return c.id


def test_build_workbook_default_uses_minting_quarter(monkeypatch, tmp_path):
    """With no explicit reference_date and 'today' frozen to July 2026, the
    workbook window ends on Q1 2026: an array producing ONLY in Q1 2026 gets a
    sheet, while Q2 2026 (the last *complete* quarter) is outside the window."""
    class _FrozenDate(date):
        @classmethod
        def today(cls):
            return date(2026, 7, 6)
    monkeypatch.setattr(gw, "date", _FrozenDate)

    # Producing in Q1 2026 (in-window) — must appear.
    cid = _make_client_producing_in([(2026, 1), (2026, 2), (2026, 3)])
    out = tmp_path / "in_window.xlsx"
    build_workbook(client_id=cid, out_path=out)   # no reference_date → default
    assert load_workbook(out).sheetnames == ["Lag Array"]

    # Producing ONLY in Q2 2026 — outside the Q1-terminal window, so no sheet.
    cid2 = _make_client_producing_in([(2026, 4), (2026, 5), (2026, 6)])
    out2 = tmp_path / "out_of_window.xlsx"
    build_workbook(client_id=cid2, out_path=out2)
    assert load_workbook(out2).sheetnames == ["(no data)"]


def test_explicit_reference_date_bypasses_lag(tmp_path):
    """Passing an explicit reference_date must still terminate on the named
    quarter — the lag only applies to the automatic default."""
    # reference_date = first day after Q2 2026 → window ends Q2 2026.
    cid = _make_client_producing_in([(2026, 4), (2026, 5), (2026, 6)])
    out = tmp_path / "explicit.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=date(2026, 7, 1))
    assert load_workbook(out).sheetnames == ["Lag Array"]
