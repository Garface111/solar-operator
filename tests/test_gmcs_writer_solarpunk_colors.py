"""
Regression tests for the solarpunk color tokens in gmcs_writer.py.

Verifies:
  - Header row (row 5) fill = #064E3B (primary-700)
  - Cell borders = #E8E2D9 (cream-border, warmer hairline)
  - Row 6 bottom border = #E6B470 (wood-300 gold accent)
"""
from __future__ import annotations

import secrets
import tempfile
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from api.db import SessionLocal
from api.models import Array, Bill, Client, Tenant, UtilityAccount
from api.writers.gmcs_writer import build_workbook

_REF = date(2024, 4, 1)


def _make_client() -> tuple[str, int]:
    """Create a minimal Tenant → Client → Array → UtilityAccount → Bill."""
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Color Test Co", contact_email=f"{tid}@test.com",
            tenant_key="kc_" + secrets.token_hex(8), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="Color Client", active=True)
        db.add(c)
        db.flush()
        arr = Array(
            tenant_id=tid, client_id=c.id,
            name="Color Array", nepool_gis_id="COL1",
        )
        db.add(arr)
        db.flush()
        ua = UtilityAccount(
            tenant_id=tid, array_id=arr.id,
            provider="gmp", account_number="CA_" + secrets.token_hex(4),
        )
        db.add(ua)
        db.flush()
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            bill_date=datetime(2024, 1, 15), period_start=datetime(2024, 1, 1),
            kwh_generated=2500, document_number="col-doc-" + secrets.token_hex(4),
            parse_status="parsed",
        ))
        db.commit()
        return tid, c.id


def test_header_fill_is_primary700(tmp_path):
    """Header row cells must use fill fgColor #064E3B (primary-700)."""
    _, cid = _make_client()
    out = tmp_path / "colors.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    for col in range(1, 5):
        cell = sh.cell(5, col)
        rgb = cell.fill.fgColor.rgb
        assert rgb.upper().endswith("064E3B"), (
            f"Header cell (5,{col}) fill = {rgb!r}, expected ...064E3B"
        )


def test_header_border_is_cream(tmp_path):
    """Header row cell borders must be #E8E2D9 (cream-border hairline)."""
    _, cid = _make_client()
    out = tmp_path / "borders.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    for col in range(1, 5):
        cell = sh.cell(5, col)
        # Check at least one side; all four are set identically by BORDER.
        top_rgb = cell.border.top.color.rgb
        assert top_rgb.upper().endswith("E8E2D9"), (
            f"Header cell (5,{col}) top border = {top_rgb!r}, expected ...E8E2D9"
        )


def test_row6_gold_bottom_border(tmp_path):
    """Gap row 6 must have a medium gold bottom border #E6B470 (wood-300)."""
    _, cid = _make_client()
    out = tmp_path / "gold.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    for col in range(1, 5):
        cell = sh.cell(6, col)
        bottom = cell.border.bottom
        assert bottom.style is not None and bottom.style != "none", (
            f"Row 6 col {col} has no bottom border style"
        )
        rgb = bottom.color.rgb
        assert rgb.upper().endswith("E6B470"), (
            f"Row 6 col {col} bottom border = {rgb!r}, expected ...E6B470"
        )
