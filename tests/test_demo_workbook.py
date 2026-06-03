"""Tests for the generic demo workbook writer (api/writers/demo_writer.py).

Verifies it mirrors the gmcs_writer format: two named sheets, merged A1:C1
title, the exact shared footnote text, and RECs = floor(MWh) on every data row.
"""
from __future__ import annotations

import math
from datetime import date

import pytest
from openpyxl import load_workbook

from api.writers.demo_writer import build_demo_workbook
from api.writers.gmcs_writer import FOOTNOTE_TEXT


@pytest.fixture()
def demo_path(tmp_path):
    out = tmp_path / "sample.xlsx"
    # Fixed reference date so the quarter set is deterministic across runs.
    return build_demo_workbook(out, reference_date=date(2026, 6, 3))


def test_two_named_sheets(demo_path):
    wb = load_workbook(demo_path)
    assert wb.sheetnames == ["Demo Array A", "Demo Array B"]


def test_titles_and_a1_merge(demo_path):
    wb = load_workbook(demo_path)
    a = wb["Demo Array A"]
    b = wb["Demo Array B"]
    assert a["A1"].value == "Demo Array A (12345)"
    assert b["A1"].value == "Demo Array B (67890)"
    # A1:C1 must be merged on every sheet.
    for sh in (a, b):
        assert "A1:C1" in {str(r) for r in sh.merged_cells.ranges}


def test_header_row_5(demo_path):
    wb = load_workbook(demo_path)
    sh = wb["Demo Array A"]
    assert [sh.cell(5, c).value for c in range(1, 5)] == [
        "Quarter", "Generation (MWh)", "Reporting Amount", "RECs†",
    ]
    assert sh.cell(5, 1).font.size == 14


def test_footnote_matches_gmcs_writer(demo_path):
    wb = load_workbook(demo_path)
    for name in ("Demo Array A", "Demo Array B"):
        sh = wb[name]
        # Footnote pinned to row 31 (6 quarters never push past it).
        assert sh.cell(31, 1).value == FOOTNOTE_TEXT


def test_recs_are_floor_of_mwh(demo_path):
    wb = load_workbook(demo_path)
    found_any = False
    for name in ("Demo Array A", "Demo Array B"):
        sh = wb[name]
        for row in range(7, 31):
            mwh = sh.cell(row, 2).value
            recs = sh.cell(row, 4).value
            if mwh is None:
                continue
            found_any = True
            assert recs == math.floor(mwh)
            # Plausible Vermont community-solar range.
            assert 15.0 <= mwh <= 35.0
    assert found_any, "expected at least some data rows"


def test_column_widths(demo_path):
    wb = load_workbook(demo_path)
    sh = wb["Demo Array A"]
    for col in ("A", "B", "C", "D"):
        assert sh.column_dimensions[col].width == 24.0


def test_six_quarter_blocks(demo_path):
    """6 quarters × (3 month rows + 1 gap) starting at row 7."""
    wb = load_workbook(demo_path)
    sh = wb["Demo Array A"]
    labels = [sh.cell(r, 1).value for r in range(7, 31)]
    quarter_labels = [v for v in labels if v and str(v).startswith("Q")]
    assert len(quarter_labels) == 6
    # Ending at the prior complete quarter relative to 2026-06-03 (Q2'26 in
    # progress → last complete is Q1 2026).
    assert quarter_labels[-1] == "Q1 2026"
