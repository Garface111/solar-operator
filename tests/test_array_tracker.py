"""Operator/tenant-level generation-spreadsheet tracker — HTTP endpoint tests.

Covers the additive `/v1/array-operator/tracker` surface (api/array_tracker.py):
  * GET returns {enabled:false} when the feature flag is off (UI hides)
  * with the flag on: GET enabled:true/has_sheet:false before upload
  * POST detects an arbitrary-column sheet, stores it per-TENANT, returns the
    same shape as GET (has_sheet:true + detected columns/headers)
  * GET reflects the stored sheet
  * download streams an xlsx
  * DELETE detaches → {enabled:true, has_sheet:false}
  * cross-tenant isolation: tenant B sees its OWN (empty) tracker, not A's
  * demo tenant is refused on the mutating routes

The detection logic itself is covered by tests/test_sheet_tracker.py; this file
exercises the tenant-keyed storage + auth + flag-gating wiring.
"""
from __future__ import annotations

import io
import secrets
from datetime import date

import openpyxl
import pytest
from fastapi.testclient import TestClient

from api.db import SessionLocal
from api.models import Array, DailyGeneration, Tenant


def _make_tenant(**over) -> str:
    """Create an AO tenant and return a SIGNED dashboard session token (the
    Bearer the `/v1/array-operator/tracker` routes expect via
    account.tenant_from_session)."""
    from api.account import mint_session_for_tenant
    tid = "ten_" + secrets.token_hex(6)
    key = "sol_test_" + secrets.token_hex(8)
    fields = dict(
        id=tid, name="Op Test", contact_email=f"{key}@op.test",
        tenant_key=key, plan="comped", active=True, product="array_operator",
    )
    fields.update(over)
    with SessionLocal() as db:
        db.add(Tenant(**fields))
        db.commit()
    return mint_session_for_tenant(tid)


def _sheet_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MyLedger"
    rows = [
        ["Maple Farm Solar — Generation Log", None, None, None, None],
        ["Billing Month", "Solar Produced (kWh)", "Home Usage",
         "Credit $/kWh", "Total Credit"],
        ["2026-03", 1820.5, 940, 0.2576, 469.0],
        ["2026-04", 2010.0, 880, 0.2576, 517.8],
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def client():
    from api.app import app
    return TestClient(app)


def _auth(key):
    return {"Authorization": f"Bearer {key}"}


def _enable(monkeypatch):
    monkeypatch.setenv("SPREADSHEET_TRACKER_ENABLED", "true")


def test_flag_off_returns_disabled(client, monkeypatch):
    monkeypatch.delenv("SPREADSHEET_TRACKER_ENABLED", raising=False)
    key = _make_tenant()
    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    assert r.status_code == 200, r.text
    assert r.json()["tracker"] == {"enabled": False}
    # mutating routes 404 while the flag is off
    assert client.delete("/v1/array-operator/tracker",
                         headers=_auth(key)).status_code == 404


def test_full_lifecycle(client, monkeypatch):
    _enable(monkeypatch)
    key = _make_tenant()

    # before upload: enabled, the AUTO sheet is primary (always generated). This
    # tenant has no arrays yet so the auto grid is empty (data_rows 0) but
    # has_sheet is still true — we always produce the sheet.
    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    assert r.status_code == 200, r.text
    tr = r.json()["tracker"]
    assert tr["enabled"] is True
    assert tr["auto"] is True
    assert tr["has_sheet"] is True
    assert tr["data_rows"] == 0
    assert tr["headers"] == ["Period", "Total"]  # no array columns yet

    # upload an arbitrary-column sheet → OPTIONAL OVERRIDE takes precedence
    blob = _sheet_bytes()
    r = client.post(
        "/v1/array-operator/tracker", headers=_auth(key),
        files={"file": ("maple.xlsx", blob,
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    tr = r.json()["tracker"]
    assert tr["has_sheet"] is True
    assert tr["auto"] is False  # the upload override is now in effect
    assert tr["filename"] == "maple.xlsx"
    # detection ran — generation column found, last period read
    assert tr["columns"]["generation"] == 1
    assert tr["columns"]["period"] == 0
    assert tr["last_period"] == "2026-04"
    assert tr["updated_at"]

    # GET reflects the stored override sheet
    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    assert r.json()["tracker"]["has_sheet"] is True
    assert r.json()["tracker"]["auto"] is False

    # download streams an xlsx
    r = client.get("/v1/array-operator/tracker/download", headers=_auth(key))
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats")
    # it's a real workbook
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb.active.max_row >= 4

    # delete detaches the override → we FALL BACK to the auto sheet (which is
    # always available), not to "no sheet". The DELETE response itself reports
    # the override is gone (has_sheet:false), but a follow-up GET shows the auto
    # sheet has taken over again.
    r = client.delete("/v1/array-operator/tracker", headers=_auth(key))
    assert r.status_code == 200, r.text
    tr = r.json()["tracker"]
    assert tr["enabled"] is True
    assert tr["has_sheet"] is False

    # GET now reflects the auto sheet again
    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    tr = r.json()["tracker"]
    assert tr["auto"] is True
    assert tr["has_sheet"] is True

    # download still works (streams the freshly-generated auto sheet)
    r = client.get("/v1/array-operator/tracker/download", headers=_auth(key))
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")


def test_cross_tenant_isolation(client, monkeypatch):
    _enable(monkeypatch)
    key_a = _make_tenant()
    key_b = _make_tenant()
    # A uploads
    client.post(
        "/v1/array-operator/tracker", headers=_auth(key_a),
        files={"file": ("a.xlsx", _sheet_bytes(),
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet")})
    # B sees its OWN auto tracker (no override), never A's uploaded sheet.
    r = client.get("/v1/array-operator/tracker", headers=_auth(key_b))
    trb = r.json()["tracker"]
    assert trb["auto"] is True          # B has no override
    assert trb["filename"] != "a.xlsx"  # definitely not A's uploaded file
    # B's download is its own auto sheet, not A's bytes
    rb = client.get("/v1/array-operator/tracker/download", headers=_auth(key_b))
    assert rb.status_code == 200
    cd = rb.headers.get("content-disposition", "")
    assert "a.xlsx" not in cd


def test_demo_tenant_refused_on_upload(client, monkeypatch):
    _enable(monkeypatch)
    key = _make_tenant(is_demo=True)
    r = client.post(
        "/v1/array-operator/tracker", headers=_auth(key),
        files={"file": ("a.xlsx", _sheet_bytes(),
                        "application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet")})
    assert r.status_code == 403, r.text


def test_bad_upload_rejected(client, monkeypatch):
    _enable(monkeypatch)
    key = _make_tenant()
    # not an xlsx/csv → unreadable → 422 (detection couldn't find a gen column)
    r = client.post(
        "/v1/array-operator/tracker", headers=_auth(key),
        files={"file": ("junk.csv", b"hello world no columns here\n",
                        "text/csv")})
    assert r.status_code == 422, r.text


# ─── the AUTO-maintained per-array workbook ──────────────────────────────────

def _tid_of(key) -> str:
    """Decode the tenant id from a signed session token."""
    from api.account import tenant_from_session
    return tenant_from_session(f"Bearer {key}").id


def _add_array(tid: str, name: str) -> int:
    with SessionLocal() as db:
        a = Array(tenant_id=tid, name=name, fuel_type="solar")
        db.add(a)
        db.commit()
        db.refresh(a)
        return a.id


def _add_gen(tid: str, array_id: int, day: date, kwh: float, source="csv"):
    with SessionLocal() as db:
        db.add(DailyGeneration(tenant_id=tid, array_id=array_id, day=day,
                               kwh=kwh, source=source))
        db.commit()


def test_auto_workbook_column_per_array_row_per_month(client, monkeypatch):
    """The headline feature: a column per array, a row per month, each cell the
    array's MEASURED kWh for that month, with a trailing Total."""
    _enable(monkeypatch)
    key = _make_tenant()
    tid = _tid_of(key)
    a_north = _add_array(tid, "North Field")
    a_south = _add_array(tid, "South Field")

    # North: Mar (3×100=300), Apr (2×100=200). South: Mar (1×50=50) only.
    for d in (1, 2, 3):
        _add_gen(tid, a_north, date(2026, 3, d), 100.0)
    for d in (10, 11):
        _add_gen(tid, a_north, date(2026, 4, d), 100.0)
    _add_gen(tid, a_south, date(2026, 3, 5), 50.0)

    # GET reflects the auto sheet
    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    tr = r.json()["tracker"]
    assert tr["auto"] is True
    assert tr["has_sheet"] is True
    # header = Period, <names by name asc>, Total  → North then South
    assert tr["headers"] == ["Period", "North Field", "South Field", "Total"]
    assert tr["data_rows"] == 2           # March + April
    assert tr["last_period"] == "2026-04"

    # download the freshly-generated workbook and verify the grid
    r = client.get("/v1/array-operator/tracker/download", headers=_auth(key))
    assert r.status_code == 200, r.text
    assert "generation_by_array" in r.headers["content-disposition"]
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Period", "North Field", "South Field", "Total")
    # row order is chronological
    by_period = {row[0]: row for row in rows[1:]}
    assert by_period["2026-03"] == ("2026-03", 300.0, 50.0, 350.0)
    # South had no April generation → blank cell, Total = North only
    assert by_period["2026-04"] == ("2026-04", 200.0, None, 200.0)


def test_auto_workbook_excludes_bill_prorate_estimate(client, monkeypatch):
    """bill_prorate rows are an ESTIMATE and must never inflate a cell."""
    _enable(monkeypatch)
    key = _make_tenant()
    tid = _tid_of(key)
    a = _add_array(tid, "Solo Array")
    _add_gen(tid, a, date(2026, 5, 1), 120.0, source="csv")            # measured
    _add_gen(tid, a, date(2026, 5, 2), 999.0, source="bill_prorate")   # estimate
    _add_gen(tid, a, date(2026, 5, 3), 30.0, source="extension_pull")  # measured

    r = client.get("/v1/array-operator/tracker/download", headers=_auth(key))
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0] == ("Period", "Solo Array", "Total")
    # 120 + 30 = 150; the 999 estimate is excluded
    assert rows[1] == ("2026-05", 150.0, 150.0)


def test_auto_workbook_new_array_becomes_new_column(client, monkeypatch):
    """Adding an array later naturally produces a new column on next fetch."""
    _enable(monkeypatch)
    key = _make_tenant()
    tid = _tid_of(key)
    a1 = _add_array(tid, "Alpha")
    _add_gen(tid, a1, date(2026, 2, 1), 10.0)

    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    assert r.json()["tracker"]["headers"] == ["Period", "Alpha", "Total"]

    # add a second array + its generation
    a2 = _add_array(tid, "Beta")
    _add_gen(tid, a2, date(2026, 2, 2), 5.0)

    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    assert r.json()["tracker"]["headers"] == ["Period", "Alpha", "Beta", "Total"]
    r = client.get("/v1/array-operator/tracker/download", headers=_auth(key))
    rows = list(openpyxl.load_workbook(io.BytesIO(r.content)).active
                .iter_rows(values_only=True))
    assert rows[0] == ("Period", "Alpha", "Beta", "Total")
    assert rows[1] == ("2026-02", 10.0, 5.0, 15.0)


def test_auto_workbook_excludes_soft_deleted_array(client, monkeypatch):
    """Soft-deleted / excluded arrays don't get a column."""
    from datetime import datetime
    _enable(monkeypatch)
    key = _make_tenant()
    tid = _tid_of(key)
    live = _add_array(tid, "Live Array")
    dead = _add_array(tid, "Dead Array")
    _add_gen(tid, live, date(2026, 1, 1), 7.0)
    _add_gen(tid, dead, date(2026, 1, 1), 7.0)
    with SessionLocal() as db:
        d = db.get(Array, dead)
        d.deleted_at = datetime.utcnow()
        db.add(d)
        db.commit()

    r = client.get("/v1/array-operator/tracker", headers=_auth(key))
    assert r.json()["tracker"]["headers"] == ["Period", "Live Array", "Total"]
