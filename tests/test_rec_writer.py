"""
Tests for api/writers/rec_writer.py — the generic, fuel-aware REC writer used
for non-solar fuels (wind / hydro / digester / storage).

Guards:
  - sheet title reflects the fuel (e.g. "… · Wind"), not hard-coded solar wording
  - the RECs header column names the fuel ("Wind RECs†")
  - the footnote is the generic REC-attestation text, NOT the verbatim GMCS one
  - REC = int(MWh) floor (1999 kWh = 1.999 MWh → 1 REC), identical to solar
  - the registry dispatcher routes a wind client to rec_writer (solar untouched)
"""
from __future__ import annotations

import math
import secrets
from datetime import datetime, date

from openpyxl import load_workbook

from api.db import SessionLocal
from api.models import Array, Bill, Client, Tenant, UtilityAccount
from api.writers.rec_writer import build_workbook as rec_build_workbook
from api.writers.gmcs_writer import FOOTNOTE_TEXT as GMCS_FOOTNOTE
from api.writers import build_workbook as dispatch_build_workbook

# Q2 2024 in progress → last complete quarter = Q1 2024. Rolling 6 quarters
# include Q1 2024, where we place the bills.
_REF = date(2024, 4, 1)


def _make_wind_client(
    kwh_per_month: int = 1999,
    *,
    fuel_type: str = "wind",
    cert_registry: str | None = None,
    nepool_gis_id: str | None = "GIS777",
) -> tuple[str, int]:
    """Tenant → Client → wind Array → UtilityAccount → Q1-2024 Bills.

    Returns (tenant_id, client_id).
    """
    months = [(2024, 1), (2024, 2), (2024, 3)]
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        t = Tenant(
            id=tid, name="Ridgeline Wind Co",
            contact_email=f"{tid}@test.com",
            tenant_key="k_" + secrets.token_hex(8),
            plan="standard", active=True,
        )
        db.add(t); db.flush()

        c = Client(tenant_id=tid, name="Wind LLC", active=True)
        db.add(c); db.flush()

        arr = Array(
            tenant_id=tid, client_id=c.id,
            name="Ridgeline Wind", nepool_gis_id=nepool_gis_id,
            fuel_type=fuel_type, cert_registry=cert_registry,
        )
        db.add(arr); db.flush()

        ua = UtilityAccount(
            tenant_id=tid, array_id=arr.id,
            provider="gmp", account_number="WIND_" + secrets.token_hex(4),
        )
        db.add(ua); db.flush()

        for y, m in months:
            db.add(Bill(
                tenant_id=tid, account_id=ua.id,
                bill_date=datetime(y, m, 15),
                period_start=datetime(y, m, 1),
                kwh_generated=kwh_per_month,
                document_number=f"wind-{tid}-{y}-{m}",
                parse_status="parsed",
            ))

        db.commit()
        return tid, c.id


# ── sheet title reflects the fuel ─────────────────────────────────────────────

def test_title_reflects_fuel(tmp_path):
    _, cid = _make_wind_client()
    out = tmp_path / "wind_title.xlsx"
    rec_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    title = wb.active["A1"].value
    assert "Wind" in title, f"title should name the fuel: {title!r}"
    assert "GIS777" in title, f"title should carry the asset id: {title!r}"
    # A1:C1 merge preserved (same layout contract as the solar writer).
    merges = {str(r) for r in wb.active.merged_cells.ranges}
    assert "A1:C1" in merges


# ── header column names the fuel ──────────────────────────────────────────────

def test_header_recs_column_names_fuel(tmp_path):
    _, cid = _make_wind_client()
    out = tmp_path / "wind_header.xlsx"
    rec_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    labels = [sh.cell(5, col).value for col in range(1, 5)]
    assert labels[:3] == ["Quarter", "Generation (MWh)", "Reporting Amount"]
    assert labels[3] == "Wind RECs\u2020", f"RECs header = {labels[3]!r}"
    # header still bold size-14 (layout parity)
    assert sh.cell(5, 1).font.size == 14
    assert sh.cell(5, 1).font.bold is True


# ── footnote is generic, NOT the verbatim GMCS one ────────────────────────────

def test_footnote_is_generic_not_gmcs(tmp_path):
    _, cid = _make_wind_client(cert_registry="NEPOOL-GIS")
    out = tmp_path / "wind_footnote.xlsx"
    rec_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    foot = sh.cell(31, 1).value
    assert foot is not None
    assert foot != GMCS_FOOTNOTE, "must NOT reuse the verbatim GMCS footnote"
    assert "REC" in foot
    assert "wind" in foot.lower(), f"footnote should mention the fuel: {foot!r}"
    assert "NEPOOL-GIS" in foot, f"footnote should name the registry: {foot!r}"


# ── registry asset id is used, with custom registry name ──────────────────────

def test_custom_cert_registry_in_footnote(tmp_path):
    """A hydro array on the LIHI registry should show LIHI, not NEPOOL-GIS."""
    _, cid = _make_wind_client(fuel_type="hydro", cert_registry="LIHI")
    out = tmp_path / "hydro_footnote.xlsx"
    rec_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    assert "Hydro" in sh.cell(1, 1).value
    assert sh.cell(5, 4).value == "Hydro RECs\u2020"
    foot = sh.cell(31, 1).value
    assert "LIHI" in foot, f"footnote should name custom registry: {foot!r}"
    assert "NEPOOL-GIS" not in foot


# ── REC = floor(MWh) — identical math to solar ────────────────────────────────

def test_recs_are_floor_of_mwh(tmp_path):
    """1999 kWh = 1.999 MWh → floor = 1 REC (not round = 2)."""
    _, cid = _make_wind_client(kwh_per_month=1999)
    out = tmp_path / "wind_recs.xlsx"
    rec_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active

    found = False
    for row in range(7, 32):
        mwh = sh.cell(row, 2).value
        recs = sh.cell(row, 4).value
        if mwh is None:
            continue
        found = True
        assert recs == math.floor(mwh), (
            f"row {row}: mwh={mwh}, recs={recs}, "
            f"expected floor={math.floor(mwh)}"
        )
        assert recs == 1  # int(1.999)
    assert found, "no data rows found in wind workbook"


# ── column widths preserved (layout parity with solar) ────────────────────────

def test_column_widths_all_24(tmp_path):
    _, cid = _make_wind_client()
    out = tmp_path / "wind_widths.xlsx"
    rec_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    for col in ("A", "B", "C", "D"):
        assert sh.column_dimensions[col].width == 24.0


# ── registry dispatch: wind client routes to rec_writer (solar path intact) ───

def test_registry_dispatches_wind_to_rec_writer(tmp_path):
    """The public api.writers.build_workbook dispatcher must route a wind
    client through rec_writer — observable via the fuel-named RECs header and
    the generic (non-GMCS) footnote."""
    _, cid = _make_wind_client()
    out = tmp_path / "dispatch_wind.xlsx"
    dispatch_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    assert sh.cell(5, 4).value == "Wind RECs\u2020"
    assert sh.cell(31, 1).value != GMCS_FOOTNOTE


# \u2500\u2500 non-producing arrays skipped (same rule as gmcs_writer) \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500

def test_nonproducing_array_gets_no_sheet(tmp_path):
    """A wind array with no generation in the window is omitted from the
    workbook \u2014 identical skip rule to the solar writer."""
    tid, cid = _make_wind_client()
    with SessionLocal() as db:
        arr = Array(
            tenant_id=tid, client_id=cid,
            name="Idle Turbine", fuel_type="wind",
        )
        db.add(arr); db.flush()
        db.add(UtilityAccount(
            tenant_id=tid, array_id=arr.id,
            provider="gmp", account_number="IDLE_" + secrets.token_hex(4),
        ))
        db.commit()

    out = tmp_path / "idle_wind.xlsx"
    rec_build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Ridgeline Wind"], f"sheets: {wb.sheetnames}"
