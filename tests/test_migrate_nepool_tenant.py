"""scripts/migrate_nepool_tenant.py — the Phase-1 fold migration tool.

Synthetic fixture mirroring the real prod shape (dual-capture "sibling"
arrays): a NEPOOL tenant with 3 clients / 3 arrays migrated onto an AO tenant.

  A1 "North Field"  -> exact-name sibling on target        (link-name)
  A2 "South Field"  -> account-number sibling "S Field Two" (link-account + rename)
  A3 "Hydro Dam"    -> no match                             (move wholesale)
  C3 "Sample VEC"   -> zero-array client, moves as-is (flagged in the plan)

capture_client=True adds the Bruce-pair shape: a target capture client (like
prod 1557) holding TA1/TA2 + a plain capture array — exercising
--claim-linked-from and --deactivate-client.

Asserts: the mapping plan, NEPOOL fields carried, settings moved,
ReportDelivery history re-pointed, verify-mode workbook byte-equality on the
default writer dispatch + the post-state reports-iteration assertions, the
rollback proof (verify leaves the DB untouched), and that ambiguities /
unclaimed linked targets abort execute instead of guessing.
"""
from __future__ import annotations

import secrets
from datetime import datetime, date

import pytest

from api.db import SessionLocal
from api.models import (Tenant, Client, Array, UtilityAccount, Bill,
                        DailyGeneration, GmpDailyGeneration, ReportDelivery)
from scripts.migrate_nepool_tenant import (
    build_plan, run_verify, run_execute, MigrationBlocked, compare_workbooks,
)

# Q2-2024 reference -> rolling window Q4'22..Q1'24; all data sits in Jan 2024.
_REF = date(2024, 4, 1)


def _tenant(*, product: str, **kw) -> str:
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(id=tid, name=f"Mig {tid[-4:]}",
                      contact_email=f"{tid}@mig.test",
                      tenant_key="k_" + secrets.token_hex(8),
                      plan="standard", active=True, product=product, **kw))
        db.commit()
    return tid


def _bill_kwargs(kwh: int) -> dict:
    # Identical on source + target sibling so the workbooks agree.
    return dict(bill_date=datetime(2024, 1, 20),
                period_start=datetime(2024, 1, 1),
                period_end=datetime(2024, 1, 31),
                kwh_generated=kwh)


def _fixture(capture_client: bool = False) -> dict:
    """The synthetic two-tenant world. Returns ids keyed by name.

    capture_client=True reproduces the Bruce-pair prod shape: the TARGET has
    one capture-created client ("GMCS Capture", like prod client 1557) that
    holds ALL its arrays — the report siblings TA1/TA2 plus a plain capture
    array TA3 (no gis) — so LINK matches hit client-linked targets and need
    --claim-linked-from.
    """
    ids: dict = {}
    src = _tenant(product="nepool",
                  report_frequency="monthly", send_mode="to_both",
                  cc_on_reports=True,
                  email_subject_template="Sub {{quarter}}",
                  email_body_template="Hello {{client_name}} — custom body",
                  email_signoff="— Bruce",
                  send_from_email="bruce@src.test", send_from_name="Bruce")
    tgt = _tenant(product="array_operator",
                  send_from_email="ao@tgt.test")   # target already set -> kept
    ids["src"], ids["tgt"] = src, tgt

    with SessionLocal() as db:
        c1 = Client(tenant_id=src, name="Alpha Solar", active=True,
                    contact_email="alpha@client.test")
        c2 = Client(tenant_id=src, name="Beta Hydro", active=True,
                    contact_email="beta@client.test")
        # The "Sample VEC" shape: a zero-array client that moves as-is.
        c3 = Client(tenant_id=src, name="Sample VEC", active=True)
        db.add_all([c1, c2, c3])
        db.flush()
        ids["c1"], ids["c2"], ids["c3"] = c1.id, c2.id, c3.id

        # A1: exact-name sibling.
        a1 = Array(tenant_id=src, client_id=c1.id, name="North Field",
                   nepool_gis_id="111")
        # A2: sibling matched by utility account number, different name.
        a2 = Array(tenant_id=src, client_id=c1.id, name="South Field",
                   nepool_gis_id="222")
        # A3: unmapped — moves wholesale. Non-solar so the writer DISPATCH
        # (registry fuel resolution) is exercised through the oracle too.
        a3 = Array(tenant_id=src, client_id=c2.id, name="Hydro Dam",
                   nepool_gis_id="333", fuel_type="hydro", cert_registry="LIHI")
        db.add_all([a1, a2, a3])
        db.flush()
        ids["a1"], ids["a2"], ids["a3"] = a1.id, a2.id, a3.id

        shared_acct = "SHARED_" + secrets.token_hex(3)
        ids["shared_acct"] = shared_acct
        ua1 = UtilityAccount(tenant_id=src, array_id=a1.id, provider="gmp",
                             account_number="N1_" + secrets.token_hex(3))
        ua2 = UtilityAccount(tenant_id=src, array_id=a2.id, provider="gmp",
                             account_number=shared_acct)
        ua3 = UtilityAccount(tenant_id=src, array_id=a3.id, provider="gmp",
                             account_number="H3_" + secrets.token_hex(3))
        db.add_all([ua1, ua2, ua3])
        db.flush()
        ids["ua1"], ids["ua3"] = ua1.id, ua3.id
        ids["ua1_num"] = ua1.account_number

        db.add(Bill(tenant_id=src, account_id=ua1.id,
                    document_number="d-" + secrets.token_hex(3),
                    **_bill_kwargs(1500)))
        db.add(Bill(tenant_id=src, account_id=ua3.id,
                    document_number="d-" + secrets.token_hex(3),
                    **_bill_kwargs(900)))
        for d in (5, 6, 7):
            db.add(DailyGeneration(tenant_id=src, array_id=a2.id,
                                   day=date(2024, 1, d), kwh=100.0))

        # ── target siblings (dual capture: same numbers, its own rows) ──
        cap_id = None
        if capture_client:
            cap = Client(tenant_id=tgt, name="GMCS Capture", active=True)
            db.add(cap)
            db.flush()
            cap_id = cap.id
            ids["cap"] = cap_id
        ta1 = Array(tenant_id=tgt, client_id=cap_id, name="North Field")
        ta2 = Array(tenant_id=tgt, client_id=cap_id, name="S Field Two")
        db.add_all([ta1, ta2])
        if capture_client:
            # plain capture array with NO gis — stays behind on the capture
            # client; deactivation is allowed while it remains.
            ta3 = Array(tenant_id=tgt, client_id=cap_id, name="Plain Capture")
            db.add(ta3)
        db.flush()
        ids["ta1"], ids["ta2"] = ta1.id, ta2.id
        if capture_client:
            ids["ta3"] = ta3.id
        # dual capture: the twin holds the SAME GMP account number under its
        # own account row (prod: 4392604400 on both Bruce tenants).
        tua1 = UtilityAccount(tenant_id=tgt, array_id=ta1.id, provider="gmp",
                              account_number=ids["ua1_num"])
        tua2 = UtilityAccount(tenant_id=tgt, array_id=ta2.id, provider="gmp",
                              account_number=shared_acct)   # the sibling tell
        db.add_all([tua1, tua2])
        db.flush()
        ids["tua1"] = tua1.id
        db.add(Bill(tenant_id=tgt, account_id=tua1.id,
                    document_number="d-" + secrets.token_hex(3),
                    **_bill_kwargs(1500)))
        for d in (5, 6, 7):
            db.add(DailyGeneration(tenant_id=tgt, array_id=ta2.id,
                                   day=date(2024, 1, d), kwh=100.0))

        db.add(ReportDelivery(tenant_id=src, client_id=c1.id,
                              client_name="Alpha Solar", recipient="alpha@client.test",
                              cadence="monthly", status="sent",
                              sent_at=datetime(2026, 6, 1)))
        db.commit()
    return ids


# ── the mapping plan ───────────────────────────────────────────────────────────

def test_plan_maps_by_name_then_account_then_moves():
    ids = _fixture()
    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]))

    assert plan.blockers == []
    by_src = {a.source_id: a for a in plan.arrays}

    a1 = by_src[ids["a1"]]
    assert a1.action == "link-name" and a1.target_id == ids["ta1"]
    assert a1.rename_to is None                       # exact case already
    assert a1.fields["nepool_gis_id"] == "111"

    a2 = by_src[ids["a2"]]
    assert a2.action == "link-account" and a2.target_id == ids["ta2"]
    assert a2.rename_to == "South Field"              # report-name continuity
    assert a2.fields["nepool_gis_id"] == "222"

    a3 = by_src[ids["a3"]]
    assert a3.action == "move"                        # create-on-target default
    assert a3.fields["fuel_type"] == "hydro"
    assert a3.fields["cert_registry"] == "LIHI"

    assert {c["name"] for c in plan.clients} == {"Alpha Solar", "Beta Hydro",
                                                 "Sample VEC"}
    # zero-array client flagged so the operator sees it's benign
    by_name = {c["name"]: c for c in plan.clients}
    assert by_name["Sample VEC"]["arrays"] == 0
    assert by_name["Alpha Solar"]["arrays"] == 2
    assert plan.report_deliveries == 1

    settings = {s["field"]: s["action"] for s in plan.settings}
    assert settings["email_body_template"] == "copy"
    assert settings["send_mode"] == "copy"
    assert settings["cc_on_reports"] == "copy"
    assert settings["send_from_email"].startswith("keep-target")   # target set


# ── the byte-diff oracle + rollback proof ──────────────────────────────────────

def test_verify_workbooks_identical_and_rolls_back():
    ids = _fixture()
    results = run_verify(ids["src"], ids["tgt"], reference_date=_REF)

    for cid in (ids["c1"], ids["c2"], ids["c3"]):
        r = results[cid]
        assert r["level"] in ("identical", "content_identical"), r
    # C1 exercises the gmcs (solar) writer, C2 the rec (hydro) writer — both
    # through the real registry dispatch. C3 (zero arrays) renders the same
    # "(no data)" stub pre/post.

    # post-state: no capture client in this fixture -> active target clients
    # == exactly the migrated set.
    assert results["_post_state"]["ok"] is True, results["_post_state"]

    # ROLLBACK PROOF: verify must leave the world byte-for-byte as it found it.
    with SessionLocal() as db:
        assert db.get(Client, ids["c1"]).tenant_id == ids["src"]
        assert db.get(Client, ids["c2"]).tenant_id == ids["src"]
        assert db.get(Array, ids["a1"]).client_id == ids["c1"]     # not detached
        assert db.get(Array, ids["a3"]).tenant_id == ids["src"]    # not moved
        ta2 = db.get(Array, ids["ta2"])
        assert ta2.name == "S Field Two" and ta2.client_id is None
        assert ta2.nepool_gis_id is None
        tgt = db.get(Tenant, ids["tgt"])
        assert tgt.email_body_template is None                     # not copied
        assert tgt.generation_reports is False                     # flag NOT flipped
        rd = db.execute(
            __import__("sqlalchemy").select(ReportDelivery)
            .where(ReportDelivery.client_id == ids["c1"])).scalars().first()
        assert rd.tenant_id == ids["src"]


# ── execute mode ───────────────────────────────────────────────────────────────

def test_execute_moves_everything_and_carries_nepool_fields():
    ids = _fixture()
    plan = run_execute(ids["src"], ids["tgt"])
    assert plan.blockers == []

    with SessionLocal() as db:
        # clients re-pointed, ids stable
        assert db.get(Client, ids["c1"]).tenant_id == ids["tgt"]
        assert db.get(Client, ids["c2"]).tenant_id == ids["tgt"]

        # A1 -> TA1 linked, fields carried, source sibling detached
        ta1 = db.get(Array, ids["ta1"])
        assert ta1.client_id == ids["c1"]
        assert ta1.nepool_gis_id == "111"
        assert db.get(Array, ids["a1"]).client_id is None

        # A2 -> TA2 linked + renamed for report-name continuity
        ta2 = db.get(Array, ids["ta2"])
        assert ta2.client_id == ids["c1"]
        assert ta2.name == "South Field"
        assert ta2.nepool_gis_id == "222"
        assert db.get(Array, ids["a2"]).client_id is None

        # A3 moved wholesale: array + account + bill + daily rows on target
        a3 = db.get(Array, ids["a3"])
        assert a3.tenant_id == ids["tgt"] and a3.client_id == ids["c2"]
        assert a3.fuel_type == "hydro" and a3.cert_registry == "LIHI"
        ua3 = db.get(UtilityAccount, ids["ua3"])
        assert ua3.tenant_id == ids["tgt"]
        from sqlalchemy import select
        b3 = db.execute(select(Bill).where(Bill.account_id == ids["ua3"])).scalars().one()
        assert b3.tenant_id == ids["tgt"]
        dg = db.execute(select(DailyGeneration)
                        .where(DailyGeneration.array_id == ids["a3"])).scalars().all()
        assert dg == [] or all(r.tenant_id == ids["tgt"] for r in dg)

        # settings: hard copies land, soft copy keeps the target's own value
        tgt = db.get(Tenant, ids["tgt"])
        src = db.get(Tenant, ids["src"])
        # the reports-world marker: THIS is what makes the migrated AO tenant
        # eligible for scheduled sends/digests (api/report_eligibility)
        assert tgt.generation_reports is True
        from api.report_eligibility import tenant_reports_eligible
        assert tenant_reports_eligible(tgt) is True
        assert tgt.email_body_template == "Hello {{client_name}} — custom body"
        assert tgt.email_subject_template == "Sub {{quarter}}"
        assert tgt.email_signoff == "— Bruce"
        assert tgt.send_mode == "to_both" and tgt.cc_on_reports is True
        assert tgt.report_frequency == "monthly"
        assert tgt.send_from_email == "ao@tgt.test"      # kept, not clobbered
        assert src.email_body_template is not None        # source untouched

        # ReportDelivery history re-pointed
        rd = db.execute(select(ReportDelivery)
                        .where(ReportDelivery.client_id == ids["c1"])).scalars().first()
        assert rd.tenant_id == ids["tgt"]

    # And the migrated target state still renders the same workbooks: a fresh
    # post-commit build for C1 equals a reference build of the same data shape.
    from api.writers import build_workbook
    import tempfile, pathlib
    with tempfile.TemporaryDirectory() as td:
        p = build_workbook(client_id=ids["c1"],
                           out_path=pathlib.Path(td) / "c1.xlsx",
                           reference_date=_REF)
        from openpyxl import load_workbook
        wb = load_workbook(p)
        assert wb.sheetnames == ["North Field", "South Field"]
        assert wb["North Field"]["A1"].value == "North Field (111)"
        assert wb["South Field"]["A1"].value == "South Field (222)"


# ── claiming capture-created client links (the Bruce-pair shape) ───────────────

def test_linked_targets_require_explicit_claim():
    ids = _fixture(capture_client=True)
    with SessionLocal() as db:
        # no claim -> the guard fires exactly as on the Bruce prod dry-run
        plan = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]))
    assert any("--claim-linked-from" in c for c in plan.conflicts)
    with pytest.raises(MigrationBlocked):
        run_execute(ids["src"], ids["tgt"])

    with SessionLocal() as db:
        # claiming some OTHER client id does not unlock these targets
        plan2 = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]),
                           claim_from={ids["cap"] + 99999})
    assert any("already linked to client" in c for c in plan2.conflicts)


def test_claim_path_repoints_renames_and_carries_gis():
    ids = _fixture(capture_client=True)
    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]),
                          claim_from={ids["cap"]})
    assert plan.blockers == []
    by_src = {a.source_id: a for a in plan.arrays}
    assert by_src[ids["a1"]].claimed_from == ids["cap"]
    assert by_src[ids["a2"]].claimed_from == ids["cap"]
    assert by_src[ids["a2"]].rename_to == "South Field"

    run_execute(ids["src"], ids["tgt"], claim_from={ids["cap"]})
    with SessionLocal() as db:
        ta1 = db.get(Array, ids["ta1"])
        ta2 = db.get(Array, ids["ta2"])
        assert ta1.client_id == ids["c1"] and ta1.nepool_gis_id == "111"
        assert ta2.client_id == ids["c1"] and ta2.nepool_gis_id == "222"
        assert ta2.name == "South Field"                       # renamed
        assert db.get(Array, ids["a1"]).client_id is None      # source detached
        assert db.get(Array, ids["a2"]).client_id is None
        # the capture client keeps its plain array and (not deactivated) stays active
        cap = db.get(Client, ids["cap"])
        assert cap.active is True
        assert db.get(Array, ids["ta3"]).client_id == ids["cap"]


def test_deactivate_refused_while_report_arrays_remain():
    ids = _fixture(capture_client=True)
    with SessionLocal() as db:
        # make the leftover capture array a REPORT array (gis set) — the
        # capture client would still hold reports-world data after the claims
        db.get(Array, ids["ta3"]).nepool_gis_id = "999"
        db.commit()
    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]),
                          claim_from={ids["cap"]},
                          deactivate_clients={ids["cap"]})
    assert any("REPORT array" in c for c in plan.conflicts)
    with pytest.raises(MigrationBlocked):
        run_execute(ids["src"], ids["tgt"], claim_from={ids["cap"]},
                    deactivate_clients={ids["cap"]})
    with SessionLocal() as db:                                 # nothing moved
        assert db.get(Client, ids["c1"]).tenant_id == ids["src"]
        assert db.get(Client, ids["cap"]).active is True


def test_deactivate_wrong_tenant_client_is_a_conflict():
    ids = _fixture(capture_client=True)
    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]),
                          claim_from={ids["cap"]},
                          deactivate_clients={ids["c1"]})      # a SOURCE client
    assert any("not a live client on target" in c for c in plan.conflicts)


def test_verify_post_state_flags_still_active_capture_client():
    ids = _fixture(capture_client=True)

    # claim WITHOUT deactivating: workbooks still byte-agree, but the reports
    # iteration would still see the capture client -> post-state FAILS verify.
    results = run_verify(ids["src"], ids["tgt"], reference_date=_REF,
                         claim_from={ids["cap"]})
    for cid in (ids["c1"], ids["c2"]):
        assert results[cid]["level"] in ("identical", "content_identical")
    post = results["_post_state"]
    assert post["ok"] is False
    assert str(ids["cap"]) in post["detail"]

    # claim + deactivate: everything green.
    results2 = run_verify(ids["src"], ids["tgt"], reference_date=_REF,
                          claim_from={ids["cap"]},
                          deactivate_clients={ids["cap"]})
    for cid in (ids["c1"], ids["c2"], ids["c3"]):
        assert results2[cid]["level"] in ("identical", "content_identical")
    assert results2["_post_state"]["ok"] is True, results2["_post_state"]
    plan = results2["_plan"]
    assert plan.deactivate == [{"id": ids["cap"], "name": "GMCS Capture",
                                "remaining_arrays": 1}]

    # ROLLBACK PROOF for the claim+deactivate path too.
    with SessionLocal() as db:
        cap = db.get(Client, ids["cap"])
        assert cap.active is True                              # not deactivated
        assert db.get(Array, ids["ta1"]).client_id == ids["cap"]   # not claimed
        assert db.get(Client, ids["c1"]).tenant_id == ids["src"]

    # and EXECUTE actually retires it.
    run_execute(ids["src"], ids["tgt"], claim_from={ids["cap"]},
                deactivate_clients={ids["cap"]})
    with SessionLocal() as db:
        assert db.get(Client, ids["cap"]).active is False
        assert db.get(Array, ids["ta3"]).client_id == ids["cap"]  # arrays untouched


def test_plan_print_flags_claims_deactivation_and_zero_array_clients(capsys):
    from scripts.migrate_nepool_tenant import _print_plan
    ids = _fixture(capture_client=True)
    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]),
                          claim_from={ids["cap"]},
                          deactivate_clients={ids["cap"]})
    _print_plan(plan)
    out = capsys.readouterr().out
    assert f"CLAIM (from client {ids['cap']})" in out
    assert "(no arrays — moves as-is)" in out
    assert "Clients to DEACTIVATE on target" in out
    assert "1 plain capture array(s) remain" in out


# ── --carry-generation: the claimed twin adopts the NEPOOL-reported series ─────

def _diverge_target_twin(ids) -> None:
    """Writer-VISIBLE DailyGeneration divergence on the A2/TA2 pair: the
    source series is operator csv (a MEASURED source the writer reads), the
    twin's is bill_prorate (EXCLUDED by the writer) — so pre-carry the twin
    renders NO daily generation for that array at all."""
    from sqlalchemy import delete
    with SessionLocal() as db:
        db.execute(delete(DailyGeneration)
                   .where(DailyGeneration.array_id == ids["ta2"]))
        for d in (5, 6, 7, 8):
            db.add(DailyGeneration(tenant_id=ids["tgt"], array_id=ids["ta2"],
                                   day=date(2024, 1, d), kwh=55.0,
                                   source="bill_prorate"))
        db.commit()


def _seed_gmp_interval_divergence(ids) -> None:
    """THE Bruce prod byte-diff mechanism (probe 2026-07-16): the writer
    overlays GmpDailyGeneration — keyed by the tenant's OWN account row — for
    any month with near-full coverage. Bruce's source accounts hold 61–1200
    interval rows; the AO twins hold ZERO (Chester 24.879 interval-true vs
    23.255 flat-prorate). Seed 30 covered days x 52 kWh on the SOURCE account
    (1560 != the 1500 bill) and nothing on the twin's account."""
    with SessionLocal() as db:
        for d in range(1, 31):                       # 30 days >= 31-2 coverage
            db.add(GmpDailyGeneration(
                tenant_id=ids["src"], account_id=ids["ua1"],
                account_number=ids["ua1_num"], array_id=ids["a1"],
                day=date(2024, 1, d), kwh=52.0, interval_count=96))
        db.commit()


def test_carry_generation_replaces_target_series_and_leaves_moves_alone():
    ids = _fixture(capture_client=True)
    _diverge_target_twin(ids)
    _seed_gmp_interval_divergence(ids)

    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, ids["src"]), db.get(Tenant, ids["tgt"]),
                          claim_from={ids["cap"]}, deactivate_clients={ids["cap"]},
                          carry_generation=True)
    assert plan.blockers == []
    by_src = {a.source_id: a for a in plan.arrays}
    assert by_src[ids["a2"]].carry_generation is True
    assert ("generation series carried from source (3 rows replace 4 target rows)"
            in by_src[ids["a2"]].notes)
    assert by_src[ids["a1"]].carry_generation is True
    assert by_src[ids["a1"]].gmp_pairs == [(ids["ua1"], ids["tua1"])]
    assert ("GMP interval series carried (30 rows replace 0) across 1 matched "
            "account(s) — the workbook's actual overlay"
            in by_src[ids["a1"]].notes)
    assert by_src[ids["a3"]].carry_generation is False    # MOVE: untouched

    run_execute(ids["src"], ids["tgt"], claim_from={ids["cap"]},
                deactivate_clients={ids["cap"]}, carry_generation=True)

    from sqlalchemy import select
    with SessionLocal() as db:
        # claimed twin ends up with EXACTLY the source series
        rows = db.execute(select(DailyGeneration)
                          .where(DailyGeneration.array_id == ids["ta2"])
                          .order_by(DailyGeneration.day)).scalars().all()
        assert [(r.day, r.kwh) for r in rows] == [
            (date(2024, 1, 5), 100.0), (date(2024, 1, 6), 100.0),
            (date(2024, 1, 7), 100.0)]                     # sampled values
        assert all(r.tenant_id == ids["tgt"] for r in rows)
        # source array keeps none (rows re-pointed, not copied)
        assert db.execute(select(DailyGeneration)
                          .where(DailyGeneration.array_id == ids["a2"])
                          ).scalars().all() == []
        # GMP interval overlay re-pointed onto the twin's OWN account row
        g_rows = db.execute(select(GmpDailyGeneration)
                            .where(GmpDailyGeneration.account_id == ids["tua1"])
                            ).scalars().all()
        assert len(g_rows) == 30
        assert all(r.tenant_id == ids["tgt"] and r.array_id == ids["ta1"]
                   and r.kwh == 52.0 for r in g_rows)
        assert db.execute(select(GmpDailyGeneration)
                          .where(GmpDailyGeneration.account_id == ids["ua1"])
                          ).scalars().all() == []          # re-pointed, not copied
        # MOVE path unaffected: A3's series moved with the array as before
        a3_rows = db.execute(select(DailyGeneration)
                             .where(DailyGeneration.array_id == ids["a3"])
                             ).scalars().all()
        assert all(r.tenant_id == ids["tgt"] for r in a3_rows)
        # bills/accounts untouched: the twin keeps its own account
        accts = db.execute(select(UtilityAccount)
                           .where(UtilityAccount.array_id == ids["ta2"])
                           ).scalars().all()
        assert accts and all(a.tenant_id == ids["tgt"] for a in accts)


def test_verify_divergent_twin_fails_without_carry_and_greens_with_it(capsys):
    ids = _fixture(capture_client=True)
    _diverge_target_twin(ids)              # writer-visible daily divergence
    _seed_gmp_interval_divergence(ids)     # the actual Bruce prod mechanism

    # WITHOUT --carry-generation: the byte-diff oracle catches BOTH
    # divergences (the Bruce prod failure shape) — real cell differences.
    results = run_verify(ids["src"], ids["tgt"], reference_date=_REF,
                         claim_from={ids["cap"]},
                         deactivate_clients={ids["cap"]})
    assert results[ids["c1"]]["level"] == "different"

    # WITH --carry-generation: byte-parity by construction — the daily series
    # AND the GMP interval overlay both travel with the claim.
    results2 = run_verify(ids["src"], ids["tgt"], reference_date=_REF,
                          claim_from={ids["cap"]},
                          deactivate_clients={ids["cap"]},
                          carry_generation=True)
    for cid in (ids["c1"], ids["c2"], ids["c3"]):
        assert results2[cid]["level"] in ("identical", "content_identical"), results2[cid]
    assert results2["_post_state"]["ok"] is True

    # ROLLBACK PROOF: the carried series evaporated with the transaction.
    from sqlalchemy import select
    with SessionLocal() as db:
        twin = db.execute(select(DailyGeneration)
                          .where(DailyGeneration.array_id == ids["ta2"])
                          ).scalars().all()
        assert sorted(r.kwh for r in twin) == [55.0] * 4   # twin untouched
        src_rows = db.execute(select(DailyGeneration)
                              .where(DailyGeneration.array_id == ids["a2"])
                              ).scalars().all()
        assert len(src_rows) == 3                          # source untouched
        g_src = db.execute(select(GmpDailyGeneration)
                           .where(GmpDailyGeneration.account_id == ids["ua1"])
                           ).scalars().all()
        assert len(g_src) == 30                            # overlay untouched
        assert db.execute(select(GmpDailyGeneration)
                          .where(GmpDailyGeneration.account_id == ids["tua1"])
                          ).scalars().all() == []

    # and the plan flags the provenance change prominently
    from scripts.migrate_nepool_tenant import _print_plan
    _print_plan(results2["_plan"])
    out = capsys.readouterr().out
    assert "PROVENANCE CHANGE" in out
    assert "bill-prorate-only" in out
    assert "GMP 15-min interval overlay" in out
    assert "generation series carried from source (3 rows replace 4 target rows)" in out
    assert ("GMP interval series carried (30 rows replace 0) across 1 matched "
            "account(s)") in out


# ── never guess between two candidates ─────────────────────────────────────────

def test_ambiguous_name_match_blocks_execute():
    src = _tenant(product="nepool")
    tgt = _tenant(product="array_operator")
    with SessionLocal() as db:
        c = Client(tenant_id=src, name="Amb Client", active=True)
        db.add(c)
        db.flush()
        db.add(Array(tenant_id=src, client_id=c.id, name="dup house"))
        # TWO live target arrays that both match case-insensitively.
        db.add(Array(tenant_id=tgt, client_id=None, name="Dup House"))
        db.add(Array(tenant_id=tgt, client_id=None, name="DUP HOUSE"))
        db.commit()

    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, src), db.get(Tenant, tgt))
    assert plan.ambiguities and "will not guess" in plan.ambiguities[0]

    with pytest.raises(MigrationBlocked):
        run_execute(src, tgt)
    with pytest.raises(MigrationBlocked):
        run_verify(src, tgt, reference_date=_REF)

    # and nothing moved
    with SessionLocal() as db:
        from sqlalchemy import select
        cl = db.execute(select(Client).where(Client.tenant_id == src)).scalars().all()
        assert len(cl) == 1


def test_product_sanity_and_demo_are_blockers():
    src = _tenant(product="array_operator")        # wrong way around
    tgt = _tenant(product="nepool")
    with SessionLocal() as db:
        db.add(Client(tenant_id=src, name="X", active=True))
        db.commit()
    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, src), db.get(Tenant, tgt))
    assert any("expected 'nepool'" in c for c in plan.conflicts)
    assert any("expected 'array_operator'" in c for c in plan.conflicts)

    demo_src = _tenant(product="nepool", is_demo=True)
    ao = _tenant(product="array_operator")
    with SessionLocal() as db:
        db.add(Client(tenant_id=demo_src, name="Y", active=True))
        db.commit()
    with SessionLocal() as db:
        plan = build_plan(db, db.get(Tenant, demo_src), db.get(Tenant, ao))
    assert any("demo" in c for c in plan.conflicts)


def test_compare_workbooks_levels():
    from openpyxl import Workbook
    import io as _io

    def _wb_bytes(val, created=None):
        wb = Workbook()
        wb.active["A1"] = val
        if created is not None:
            wb.properties.created = created
            wb.properties.modified = created
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    same_t = datetime(2026, 1, 1, 12, 0, 0)
    a = _wb_bytes("x", created=same_t)
    b = _wb_bytes("x", created=same_t)
    assert compare_workbooks(a, b)[0] == "identical"

    c = _wb_bytes("x", created=datetime(2026, 1, 1, 12, 0, 5))
    assert compare_workbooks(a, c)[0] == "content_identical"   # timestamp-only

    d = _wb_bytes("y", created=same_t)
    level, detail = compare_workbooks(a, d)
    assert level == "different" and "row 1" in detail
