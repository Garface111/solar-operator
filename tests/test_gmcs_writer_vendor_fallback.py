"""Native multi-source generation reports.

Precedence: utility daily (VEC/GMP) > monitoring (Locus/AlsoEnergy/eGauge/…) > bill.
"""
from __future__ import annotations

import calendar
import secrets
from datetime import date

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from api.generation_sources import (
    MONITORING_REPORT_SOURCES,
    UTILITY_REAL_SOURCES,
    normalize_source,
)
from api.models import (
    Array,
    Base,
    Client,
    DailyGeneration,
    GmpDailyGeneration,
    Inverter,
    Tenant,
    UtilityAccount,
)
from api.writers import gmcs_writer


@pytest.fixture()
def db():
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    s = Session()
    import api.reports.gmp_daily_read as rd
    rd.SessionLocal = Session
    yield s
    s.close()


def _seed_utility(db):
    db.add(Tenant(id="ten_t", name="T", contact_email="t@x.com", tenant_key="t_key"))
    db.add(Array(id=1, tenant_id="ten_t", name="Chester"))
    db.add(UtilityAccount(id=10, tenant_id="ten_t", array_id=1, provider="gmp", account_number="A1"))
    db.commit()


def test_merge_precedence_utility_beats_monitoring_beats_bill():
    out = gmcs_writer._merge_report_months(
        vendor={(2025, 6): 100.0, (2025, 7): 200.0, (2025, 8): 300.0},
        bill={(2025, 6): 50.0, (2025, 7): 50.0},
        utility={(2025, 6): 10.0},
    )
    assert out[(2025, 6)] == 10.0   # utility
    assert out[(2025, 7)] == 200.0  # monitoring beats bill
    assert out[(2025, 8)] == 300.0  # monitoring only


def test_native_sources_include_screenshot_labels():
    """Sheet labels from Crown tracking → canonical sources in report tiers."""
    assert "locus" in MONITORING_REPORT_SOURCES
    assert "alsoenergy" in MONITORING_REPORT_SOURCES
    assert "egauge" in MONITORING_REPORT_SOURCES
    assert "metermate" in MONITORING_REPORT_SOURCES
    assert "langsend" in MONITORING_REPORT_SOURCES
    assert "smarthub" in UTILITY_REAL_SOURCES
    assert "utility_meter" in UTILITY_REAL_SOURCES  # VEC daily capture
    assert normalize_source("LOCUS") == "locus"
    assert normalize_source("ALSO") == "alsoenergy"
    assert normalize_source("VEC Website") == "smarthub"
    assert normalize_source("Meter Mate") == "metermate"
    assert normalize_source("Eguage - CAROLYN") == "egauge"
    assert normalize_source("Eguage") == "egauge"
    assert normalize_source("LangSendsData") == "langsend"
    assert normalize_source("VEC - Dave Lahar Data Request") == "dave_lahar"


def test_alsoenergy_is_in_monitoring_set():
    assert "alsoenergy" in gmcs_writer._MONITORING_REPORT_SOURCES
    assert "locus" in gmcs_writer._MONITORING_REPORT_SOURCES
    # back-compat alias
    assert "alsoenergy" in gmcs_writer._VENDOR_FALLBACK_SOURCES


@pytest.mark.parametrize("source", ["locus", "alsoenergy", "egauge", "metermate", "langsend"])
def test_monitoring_month_when_no_utility(db, source):
    _seed_utility(db)
    dim = calendar.monthrange(2025, 8)[1]
    for dd in range(1, dim + 1):
        db.add(DailyGeneration(
            tenant_id="ten_t", array_id=1, day=date(2025, 8, dd),
            kwh=10.0, source=source,
        ))
    db.commit()
    mon = gmcs_writer._monitoring_generation_by_month(
        db, 1, date(2025, 8, 1), date(2025, 8, 31)
    )
    util = gmcs_writer._daily_generation_by_month(
        db, 1, date(2025, 8, 1), date(2025, 8, 31)
    )
    assert (2025, 8) not in util
    assert mon[(2025, 8)] == pytest.approx(10.0 * dim)
    merged = gmcs_writer._merge_report_months(vendor=mon, bill={}, utility=util)
    assert merged[(2025, 8)] == pytest.approx(10.0 * dim)


def test_vec_utility_meter_is_utility_tier(db):
    """SmartHub/VEC daily capture (source=utility_meter) is settlement utility."""
    _seed_utility(db)
    dim = calendar.monthrange(2025, 5)[1]
    for dd in range(1, dim + 1):
        db.add(DailyGeneration(
            tenant_id="ten_t", array_id=1, day=date(2025, 5, dd),
            kwh=20.0, source="utility_meter",
        ))
    db.commit()
    util = gmcs_writer._daily_generation_by_month(
        db, 1, date(2025, 5, 1), date(2025, 5, 31)
    )
    mon = gmcs_writer._monitoring_generation_by_month(
        db, 1, date(2025, 5, 1), date(2025, 5, 31)
    )
    assert util[(2025, 5)] == pytest.approx(20.0 * dim)
    assert (2025, 5) not in mon


def test_smarthub_source_is_utility_tier(db):
    _seed_utility(db)
    dim = calendar.monthrange(2025, 4)[1]
    for dd in range(1, dim + 1):
        db.add(DailyGeneration(
            tenant_id="ten_t", array_id=1, day=date(2025, 4, dd),
            kwh=15.0, source="smarthub",
        ))
    db.commit()
    util = gmcs_writer._daily_generation_by_month(
        db, 1, date(2025, 4, 1), date(2025, 4, 30)
    )
    assert util[(2025, 4)] == pytest.approx(15.0 * dim)


def test_gmp_full_month_still_beats_locus(db):
    _seed_utility(db)
    dim = calendar.monthrange(2025, 6)[1]
    for dd in range(1, dim + 1):
        db.add(DailyGeneration(
            tenant_id="ten_t", array_id=1, day=date(2025, 6, dd),
            kwh=99.0, source="locus",
        ))
        db.add(GmpDailyGeneration(
            tenant_id="ten_t", account_id=10, account_number="A1",
            array_id=1, day=date(2025, 6, dd), kwh=20.0,
            interval_count=96, source="gmp_api",
        ))
    db.commit()
    util = gmcs_writer._daily_generation_by_month(
        db, 1, date(2025, 6, 1), date(2025, 6, 30)
    )
    mon = gmcs_writer._monitoring_generation_by_month(
        db, 1, date(2025, 6, 1), date(2025, 6, 30)
    )
    merged = gmcs_writer._merge_report_months(vendor=mon, bill={}, utility=util)
    assert merged[(2025, 6)] == pytest.approx(20.0 * dim)  # GMP, not Locus


def test_locus_beats_bill_for_same_month(db):
    """Native monitoring outranks bill pro-rate for month shape."""
    _seed_utility(db)
    dim = calendar.monthrange(2025, 7)[1]
    for dd in range(1, dim + 1):
        db.add(DailyGeneration(
            tenant_id="ten_t", array_id=1, day=date(2025, 7, dd),
            kwh=50.0, source="locus",
        ))
    db.commit()
    mon = gmcs_writer._monitoring_generation_by_month(
        db, 1, date(2025, 7, 1), date(2025, 7, 31)
    )
    util = gmcs_writer._daily_generation_by_month(
        db, 1, date(2025, 7, 1), date(2025, 7, 31)
    )
    bill = {(2025, 7): 1000.0}
    merged = gmcs_writer._merge_report_months(vendor=mon, bill=bill, utility=util)
    assert merged[(2025, 7)] == pytest.approx(50.0 * dim)


def test_inverter_only_client_workbook_uses_locus(tmp_path):
    """Locus-only client (no utility accounts) still gets a reportable sheet."""
    from api.db import SessionLocal
    import api.writers.gmcs_writer as w

    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Locus Only", contact_email=f"{tid}@t.com",
            tenant_key="k_" + secrets.token_hex(4), plan="standard", active=True,
        ))
        db.flush()
        c = Client(tenant_id=tid, name="Loc Client", active=True)
        db.add(c); db.flush()
        arr = Array(tenant_id=tid, client_id=c.id, name="Benson Site")
        db.add(arr); db.flush()
        db.add(Inverter(
            tenant_id=tid, array_id=arr.id, vendor="locus",
            serial="LOC-1", nameplate_kw=10.0,
        ))
        for m in (1, 2, 3):
            dim = calendar.monthrange(2024, m)[1]
            for d in range(1, dim + 1):
                db.add(DailyGeneration(
                    tenant_id=tid, array_id=arr.id, day=date(2024, m, d),
                    kwh=100.0, source="locus",
                ))
        db.commit()
        cid, aid = c.id, arr.id

    out = w.build_workbook(
        client_id=cid,
        reference_date=date(2024, 4, 1),
        quarters=6,
        out_path=tmp_path / "locus.xlsx",
    )
    wb = load_workbook(out)
    assert any("Benson" in (s or "") for s in wb.sheetnames)
    sh = wb[wb.sheetnames[0]]
    vals = []
    for row in sh.iter_rows(min_row=7, max_row=30, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:
                vals.append(float(cell.value))
    assert any(abs(v - 3.1) < 0.01 for v in vals), vals
    assert aid in w.reported_array_ids(cid, reference_date=date(2024, 4, 1), quarters=6)
    assert w.report_has_data(cid, reference_date=date(2024, 4, 1), quarters=6)
