"""All-time aggregated fleet report — tests.

Seeds a non-Bruce tenant with two arrays + DailyGeneration across two months
and two years, plus a Bill (for a month with no daily data), then asserts the
generated Excel totals match the seeded kWh, per-array rows exist, and the PDF
is non-empty with a %PDF header. Also covers the empty-fleet "no data" path.
"""
from __future__ import annotations

import io
import secrets
from datetime import date, datetime

from openpyxl import load_workbook

from api.db import SessionLocal
from api.models import Array, Bill, Client, DailyGeneration, Tenant, UtilityAccount
from api.reports.fleet_report import (
    aggregate_fleet,
    build_fleet_report,
    report_filename,
)


def _seed_fleet() -> tuple[str, dict]:
    """Create a tenant with 2 arrays + DailyGeneration across 2 months/years.

    Returns (tenant_id, expected) where expected carries the kWh we seeded.
    """
    tid = "ten_" + secrets.token_hex(6)
    expected = {
        "array_a_name": "Maple Field",
        "array_b_name": "Cedar Ridge",
        "by_year": {2023: 0.0, 2024: 0.0},
        "by_array": {},
    }
    with SessionLocal() as db:
        t = Tenant(
            id=tid, name="Fleet Report Test Co",
            company_name="Fleet Report Test Co",
            contact_email=f"{tid}@test.com",
            tenant_key="k_" + secrets.token_hex(8),
            plan="standard", active=True, product="array_operator",
        )
        db.add(t); db.flush()
        c = Client(tenant_id=tid, name="Default", active=True)
        db.add(c); db.flush()

        arr_a = Array(tenant_id=tid, client_id=c.id, name="Maple Field",
                      nepool_gis_id="MAPLE1")
        arr_b = Array(tenant_id=tid, client_id=c.id, name="Cedar Ridge",
                      nepool_gis_id="CEDAR9")
        db.add_all([arr_a, arr_b]); db.flush()

        # Array A: Jan 2023 (10 days × 100 = 1000) + Jan 2024 (10 days × 200 = 2000)
        a_total = 0.0
        for d in range(1, 11):
            db.add(DailyGeneration(tenant_id=tid, array_id=arr_a.id,
                                   day=date(2023, 1, d), kwh=100.0, source="csv"))
            a_total += 100.0
            expected["by_year"][2023] += 100.0
        for d in range(1, 11):
            db.add(DailyGeneration(tenant_id=tid, array_id=arr_a.id,
                                   day=date(2024, 1, d), kwh=200.0, source="csv"))
            a_total += 200.0
            expected["by_year"][2024] += 200.0

        # Array B: Feb 2024 (5 days × 50 = 250)
        b_total = 0.0
        for d in range(1, 6):
            db.add(DailyGeneration(tenant_id=tid, array_id=arr_b.id,
                                   day=date(2024, 2, d), kwh=50.0, source="csv"))
            b_total += 50.0
            expected["by_year"][2024] += 50.0

        expected["by_array"]["Maple Field"] = a_total
        expected["by_array"]["Cedar Ridge"] = b_total
        expected["total"] = a_total + b_total

        db.commit()
    return tid, expected


def _get_tenant(tid: str) -> Tenant:
    with SessionLocal() as db:
        return db.get(Tenant, tid)


def test_aggregate_totals_match_seeded():
    tid, exp = _seed_fleet()
    agg = aggregate_fleet(_get_tenant(tid))
    assert agg.total_generated_kwh == exp["total"]
    assert agg.by_year[2023] == exp["by_year"][2023]
    assert agg.by_year[2024] == exp["by_year"][2024]
    # per-array
    by_name = {r["name"]: r["kwh"] for r in agg.by_array}
    assert by_name["Maple Field"] == exp["by_array"]["Maple Field"]
    assert by_name["Cedar Ridge"] == exp["by_array"]["Cedar Ridge"]
    assert agg.array_count == 2


def test_xlsx_totals_and_arrays():
    tid, exp = _seed_fleet()
    blob = build_fleet_report(_get_tenant(tid), "xlsx")
    assert isinstance(blob, bytes) and len(blob) > 0

    wb = load_workbook(io.BytesIO(blob))
    sh = wb.active
    # Flatten all cell values to strings for content assertions.
    text_cells = [str(c.value) for row in sh.iter_rows() for c in row
                  if c.value is not None]
    blob_text = "\n".join(text_cells)

    # Per-array rows present.
    assert "Maple Field" in blob_text
    assert "Cedar Ridge" in blob_text
    assert "MAPLE1" in blob_text

    # Total generation kWh appears as a numeric cell equal to the seeded total.
    numeric = [c.value for row in sh.iter_rows() for c in row
               if isinstance(c.value, (int, float))]
    assert exp["total"] in numeric, (
        f"seeded total {exp['total']} not found among {numeric}")
    # Per-year totals present as numbers.
    assert exp["by_year"][2023] in numeric
    assert exp["by_year"][2024] in numeric
    # Per-array totals present.
    assert exp["by_array"]["Maple Field"] in numeric
    assert exp["by_array"]["Cedar Ridge"] in numeric


def test_pdf_non_empty_with_header():
    tid, _ = _seed_fleet()
    blob = build_fleet_report(_get_tenant(tid), "pdf")
    assert isinstance(blob, bytes)
    assert len(blob) > 500
    assert blob[:5] == b"%PDF-"


def test_bill_consumption_included():
    """A bill for a month with NO daily generation contributes generation +
    consumption to the fleet totals (no double counting)."""
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        t = Tenant(id=tid, name="Bill Fleet", company_name="Bill Fleet",
                   contact_email=f"{tid}@test.com",
                   tenant_key="k_" + secrets.token_hex(8),
                   plan="standard", active=True)
        db.add(t); db.flush()
        c = Client(tenant_id=tid, name="Default", active=True)
        db.add(c); db.flush()
        arr = Array(tenant_id=tid, client_id=c.id, name="Solo Array")
        db.add(arr); db.flush()
        ua = UtilityAccount(tenant_id=tid, array_id=arr.id, provider="gmp",
                            account_number="ACC_" + secrets.token_hex(4))
        db.add(ua); db.flush()
        # A bill fully inside one calendar month → all kWh to that month.
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            bill_date=datetime(2022, 6, 28),
            period_start=datetime(2022, 6, 1),
            period_end=datetime(2022, 6, 30),
            billing_days=30,
            kwh_generated=900, kwh_consumed=400,
            total_cost=-50.0, net_credit=50.0,
            document_number=f"bill-{tid}-2022-06", parse_status="parsed",
        ))
        db.commit()

    agg = aggregate_fleet(_get_tenant(tid))
    assert agg.total_generated_kwh == 900.0
    assert agg.total_consumed_kwh == 400.0
    assert agg.has_consumption is True
    assert agg.total_cost == -50.0
    assert agg.by_year[2022] == 900.0
    # XLSX + PDF still build.
    assert build_fleet_report(_get_tenant(tid), "xlsx")[:2] == b"PK"
    assert build_fleet_report(_get_tenant(tid), "pdf")[:5] == b"%PDF-"


def test_empty_fleet_is_valid_no_data():
    """An empty fleet yields a valid report saying 'no data yet' — not numbers."""
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        t = Tenant(id=tid, name="Empty Co", company_name="Empty Co",
                   contact_email=f"{tid}@test.com",
                   tenant_key="k_" + secrets.token_hex(8),
                   plan="standard", active=True)
        db.add(t); db.commit()

    agg = aggregate_fleet(_get_tenant(tid))
    assert agg.has_data is False
    assert agg.total_generated_kwh == 0.0

    xlsx = build_fleet_report(_get_tenant(tid), "xlsx")
    wb = load_workbook(io.BytesIO(xlsx))
    text = "\n".join(str(c.value) for row in wb.active.iter_rows()
                     for c in row if c.value is not None)
    assert "No fleet data" in text

    pdf = build_fleet_report(_get_tenant(tid), "pdf")
    assert pdf[:5] == b"%PDF-"


def test_report_filename():
    tid, _ = _seed_fleet()
    t = _get_tenant(tid)
    assert report_filename(t, "xlsx") == "FleetReport-Fleet-Report-Test-Co-AllTime.xlsx"
    assert report_filename(t, "pdf").endswith("-AllTime.pdf")
