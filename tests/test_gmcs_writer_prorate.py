"""
Tests that the GMCS writer uses pro-rate (not period_start) for cross-month bills,
and that the workbook structure is completely unchanged.

Uses a synthetic array with cross-month bills. reference_date=2025-08-01 puts
Q2 2025 (Apr/May/Jun) as the most recent complete quarter.

Rolling 6 quarters from 2025-08-01:
  Q1 2024, Q2 2024, Q3 2024, Q4 2024, Q1 2025, Q2 2025

Q2 2025 row layout (data starts at row 7, each quarter = 3 rows + 1 gap):
  Q1 2024: rows 7, 8, 9 (gap: 10)
  Q2 2024: rows 11, 12, 13 (gap: 14)
  Q3 2024: rows 15, 16, 17 (gap: 18)
  Q4 2024: rows 19, 20, 21 (gap: 22)
  Q1 2025: rows 23, 24, 25 (gap: 26)
  Q2 2025: rows 27 (Apr), 28 (May), 29 (Jun)
"""
from __future__ import annotations

import math
import secrets
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from api.db import SessionLocal
from api.models import Array, Bill, Client, Tenant, UtilityAccount
from api.writers.gmcs_writer import FOOTNOTE_TEXT, build_workbook

_REF = date(2025, 8, 1)

# Q2 2025 row mapping within the 6-quarter sheet
_APR_ROW = 27
_MAY_ROW = 28
_JUN_ROW = 29


def _make_cross_month_scenario() -> tuple[str, int]:
    """Tenant → Client → Array → UtilityAccount → 1 cross-month bill.

    Bill: 2025-04-11 → 2025-05-12, kwh=32000
      Total days: 32   April: 20   May: 12
      Pro-rate:  April = 32000 * 20/32 = 20000 kWh = 20.000 MWh
                 May   = 32000 * 12/32 = 12000 kWh = 12.000 MWh
      period_start would give: April = 32.000 MWh, May = 0 (wrong)

    Returns (tenant_id, client_id).
    """
    tid = "ten_" + secrets.token_hex(6)
    with SessionLocal() as db:
        db.add(Tenant(
            id=tid, name="Prorate Test Co",
            contact_email=f"{tid}@test.com",
            tenant_key="kp_" + secrets.token_hex(8),
            plan="standard", active=True,
        ))
        db.flush()

        c = Client(tenant_id=tid, name="Green Farms", active=True)
        db.add(c); db.flush()

        arr = Array(
            tenant_id=tid, client_id=c.id,
            name="Maple Ridge", nepool_gis_id="PR001",
        )
        db.add(arr); db.flush()

        ua = UtilityAccount(
            tenant_id=tid, array_id=arr.id,
            provider="gmp", account_number="PR_" + secrets.token_hex(4),
        )
        db.add(ua); db.flush()

        # Cross-month bill: April 11 → May 12, 2025
        db.add(Bill(
            tenant_id=tid, account_id=ua.id,
            bill_date=datetime(2025, 5, 15),
            period_start=datetime(2025, 4, 11),
            period_end=datetime(2025, 5, 12),
            kwh_generated=32000,
            document_number="cross-" + tid,
            parse_status="parsed",
        ))

        db.commit()
        return tid, c.id


# ── pro-rate values ───────────────────────────────────────────────────────────

def test_cross_month_bill_prorated_into_april(tmp_path):
    """April cell must show 20.000 MWh (32000 * 20/32 / 1000), not 32.000."""
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "prorate_apr.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    apr_mwh = sh.cell(_APR_ROW, 2).value
    assert apr_mwh is not None, "April cell is None — no data written"
    assert abs(apr_mwh - 20.0) < 0.01, (
        f"April MWh = {apr_mwh}, expected 20.000 (pro-rate). "
        f"If you see 32.000 the fix hasn't been applied."
    )


def test_cross_month_bill_prorated_into_may(tmp_path):
    """May cell must show 12.000 MWh (32000 * 12/32 / 1000), not None/0."""
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "prorate_may.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    may_mwh = sh.cell(_MAY_ROW, 2).value
    assert may_mwh is not None, (
        "May cell is None — pro-rate didn't split into May. "
        "period_start attribution would leave May empty."
    )
    assert abs(may_mwh - 12.0) < 0.01, (
        f"May MWh = {may_mwh}, expected 12.000 (pro-rate)."
    )


def test_prorate_values_sum_to_total_kwh(tmp_path):
    """April + May MWh together must equal 32.000 (kWh conservation)."""
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "prorate_sum.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    apr = sh.cell(_APR_ROW, 2).value or 0.0
    may = sh.cell(_MAY_ROW, 2).value or 0.0
    assert abs(apr + may - 32.0) < 0.01, (
        f"April ({apr}) + May ({may}) = {apr + may}, expected 32.000"
    )


# ── structure unchanged ───────────────────────────────────────────────────────

def test_structure_a1_merge_unchanged(tmp_path):
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "struct_merge.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    merge_ranges = {str(r) for r in wb.active.merged_cells.ranges}
    assert "A1:C1" in merge_ranges


def test_structure_a1_title_unchanged(tmp_path):
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "struct_title.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.active["A1"].value == "Maple Ridge (PR001)"


def test_structure_row5_header_unchanged(tmp_path):
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "struct_hdr.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    labels = [sh.cell(5, col).value for col in range(1, 5)]
    assert labels == ["Quarter", "Generation (MWh)", "Reporting Amount", "RECs†"]
    assert sh.cell(5, 1).font.size == 14
    assert sh.cell(5, 1).font.bold is True


def test_structure_recs_still_floor_of_mwh(tmp_path):
    """RECs in both April and May must be int(mwh) — floor."""
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "struct_recs.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    for row in (_APR_ROW, _MAY_ROW):
        mwh = sh.cell(row, 2).value
        recs = sh.cell(row, 4).value
        if mwh is not None:
            assert recs == math.floor(mwh), (
                f"row {row}: mwh={mwh}, recs={recs}, expected floor={math.floor(mwh)}"
            )


def test_structure_footnote_verbatim_at_row_31(tmp_path):
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "struct_foot.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    assert wb.active.cell(31, 1).value == FOOTNOTE_TEXT


def test_structure_column_widths_unchanged(tmp_path):
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "struct_widths.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    for col in ("A", "B", "C", "D"):
        assert sh.column_dimensions[col].width == 24.0


def test_june_cell_is_empty_for_single_cross_month_bill(tmp_path):
    """June must remain empty — the bill period ends May 12 so no June kWh."""
    _, cid = _make_cross_month_scenario()
    out = tmp_path / "struct_jun.xlsx"
    build_workbook(client_id=cid, out_path=out, reference_date=_REF)
    wb = load_workbook(out)
    sh = wb.active
    assert sh.cell(_JUN_ROW, 2).value is None, (
        f"June MWh should be None (no bill spans into June), got {sh.cell(_JUN_ROW, 2).value}"
    )
