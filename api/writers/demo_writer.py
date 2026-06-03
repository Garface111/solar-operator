"""
Demo workbook writer — a generic, customer-agnostic NEPOOL-GIS sample.

Produces the SAME pixel-perfect layout as ``gmcs_writer.build_workbook`` (one
sheet per array, A1:C1 merged title, row-5 header at size 14, rolling 6 quarters
of 3 month rows + a gap row, MWh in Excel "General" format, RECs = floor(MWh),
verbatim footnote, all columns width 24.0) — but populated with realistic FAKE
data so it can be shipped publicly (marketing download, onboarding welcome email)
without leaking any real customer's generation figures.

Two sheets:
  - "Demo Array A" titled "Demo Array A (12345)"
  - "Demo Array B" titled "Demo Array B (67890)"

Data is plausible Vermont community-solar generation (15-35 MWh/month, seasonal
— high in summer, low in winter) and fully deterministic (no RNG) so the file is
byte-stable across regenerations.
"""
from __future__ import annotations

import pathlib
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Reuse the gmcs_writer's quarter math + the canonical footnote so this demo
# stays in lockstep with the real report format.
from .gmcs_writer import (
    FOOTNOTE_TEXT,
    _quarter_months,
    _rolling_quarters,
)

# (sheet title, NEPOOL-GIS id, generation seed). The seed shifts the seasonal
# curve so the two arrays don't read as identical copies.
DEMO_ARRAYS: list[tuple[str, str, float]] = [
    ("Demo Array A", "12345", 0.0),
    ("Demo Array B", "67890", 6.5),
]

# Baseline monthly MWh by calendar month (1=Jan … 12=Dec). Tuned for a VT
# community-solar array: trough in Dec-Feb, peak in Jun-Aug, all within 15-35.
_MONTH_BASE_MWH = {
    1: 16.0, 2: 17.5, 3: 21.0, 4: 25.5, 5: 29.0, 6: 33.0,
    7: 34.0, 8: 32.0, 9: 28.0, 10: 24.0, 11: 19.5, 12: 15.5,
}


def _demo_mwh(year: int, month: int, seed: float) -> float:
    """Deterministic, plausible monthly generation in MWh (clamped 15-35).

    Combines the seasonal baseline with a small per-array seed offset and a
    tiny year/month wobble so consecutive years aren't carbon copies. No RNG —
    same inputs always yield the same value, keeping sample.xlsx byte-stable.
    """
    base = _MONTH_BASE_MWH[month]
    # Smooth, bounded wobble: a deterministic function of year+month.
    wobble = ((year * 12 + month) % 7) * 0.4 - 1.2  # roughly -1.2 … +1.2
    val = base + seed * 0.5 + wobble
    val = max(15.0, min(35.0, val))
    return round(val, 3)


def build_demo_workbook(out_path: pathlib.Path,
                        *, reference_date: Optional[date] = None) -> pathlib.Path:
    """Build the generic demo workbook and save it to ``out_path``.

    Mirrors gmcs_writer's formatting exactly. Shows 6 complete quarters ending
    at the prior complete quarter relative to ``reference_date`` (defaults to
    today), so the file always feels current.
    """
    out_path = pathlib.Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    ref = reference_date or date.today()
    qlist = _rolling_quarters(ref, count=6)

    wb = Workbook()
    default_sheet = wb.active

    # Styling constants — identical values to gmcs_writer.build_workbook.
    TITLE_FONT = Font(bold=True, size=14, color="1F4E2A")
    HDR_FONT = Font(bold=True, size=14, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="2E6B3A")
    QUARTER_FONT = Font(bold=True, size=11, color="1F4E2A")
    FOOTNOTE_FONT = Font(italic=True, size=9, color="666666")
    BORDER = Border(*[Side(style="thin", color="C8D4C4")] * 4)

    for idx, (name, nepool_id, seed) in enumerate(DEMO_ARRAYS):
        if idx == 0:
            sh = default_sheet
            sh.title = name
        else:
            sh = wb.create_sheet(title=name)

        # ── Title (A1 merged A1:C1) ──
        sh["A1"] = f"{name} ({nepool_id})"
        sh["A1"].font = TITLE_FONT
        sh["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells("A1:C1")

        # Row 5 header
        for col, label in enumerate(
            ["Quarter", "Generation (MWh)", "Reporting Amount", "RECs†"],
            start=1,
        ):
            c = sh.cell(5, col, label)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

        # Data: 6 quarter blocks × 3 month rows, gap row between blocks.
        row = 7
        for (qy, qq) in qlist:
            for i, (my, mm) in enumerate(_quarter_months(qy, qq)):
                if i == 0:
                    qc = sh.cell(row, 1, f"Q{qq} {qy}")
                    qc.font = QUARTER_FONT
                    qc.alignment = Alignment(horizontal="left", vertical="center")
                mwh = _demo_mwh(my, mm, seed)
                recs = int(mwh)  # floor of MWh
                gc = sh.cell(row, 2, mwh)
                gc.number_format = "General"
                gc.alignment = Alignment(horizontal="right")
                rc = sh.cell(row, 3, mwh)
                rc.number_format = "General"
                rc.alignment = Alignment(horizontal="right")
                ec = sh.cell(row, 4, recs)
                ec.number_format = "General"
                ec.alignment = Alignment(horizontal="right")
                row += 1
            row += 1  # gap row between quarter blocks

        # Footnote — verbatim, pinned to row 31 unless data ran past it.
        foot_row = 31 if row <= 31 else row
        fc = sh.cell(foot_row, 1, FOOTNOTE_TEXT)
        fc.font = FOOTNOTE_FONT
        fc.alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells(start_row=foot_row, start_column=1,
                       end_row=foot_row, end_column=4)

        # Column widths — uniform 24 across all four columns.
        for col_letter in ("A", "B", "C", "D"):
            sh.column_dimensions[col_letter].width = 24.0

    wb.save(out_path)
    return out_path
