"""Raw utility generation export (Ford 2026-07-16; hourly detail 2026-07-23).

A NEPOOL Operator can download a client's utility generation for a chosen
quarter, organized by month — the meter data behind the REC report, shaped for a
monthly-basis program to ingest. Covers ALL of the client's utilities:
  • GMP — the 15-minute interval meter, rolled up to HOURLY for the detail sheet
    (api.reports.gmp_daily_read.get_hourly_series from the raw sponge) + the
    bill's reported generation where no interval exists for the monthly grid.
  • SmartHub co-ops (VEC/WEC/…) — only daily RETURN-meter generation is captured
    (co-op bills carry no kwh_generated). Detail sheets fall back to one row per
    day with Hour blank when no GMP hourly exists — never fabricate hours.
Inverter/vendor telemetry (solaredge/fronius/…) is never included — this is
utility-measured generation, the REC basis.

Workbook:
  • "Monthly Summary" — projects × the quarter's 3 months grid + totals, with the
    utility + source per project, so every project shows.
  • a per-project HOURLY detail sheet wherever a meter feed exists (GMP
    interval→hour; SmartHub daily as day-grain rows).
"""
from __future__ import annotations

import calendar
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select

from ..bill_attribution import distribute_kwh_by_calendar_day
from ..db import SessionLocal
from ..generation_sources import (
    EXTENSION_SOURCES,
    MONITORING_REPORT_SOURCES,
    VENDOR_TELEMETRY_SOURCES,
)
from ..models import (
    Array, Bill, Client, DailyGeneration, GmpDailyGeneration, Tenant, UtilityAccount,
)
from ..report_arrays import is_vendor_only_array, not_vendor_only
from ..reports import gmp_daily_read
from .gmcs_writer import _daily_generation_by_month, _quarter_months, _sheet_name_for_array

# Non-utility DailyGeneration sources — never in a utility-generation export.
_METER_EXCLUDE = VENDOR_TELEMETRY_SOURCES | EXTENSION_SOURCES | {"bill_prorate"}

_GREEN = "064E3B"
_TITLE = Font(bold=True, size=14, color="1F4E2A")
_SUB = Font(italic=True, size=9, color="666666")
_HDR = Font(bold=True, size=11, color="FFFFFF")
_HDR_FILL = PatternFill("solid", fgColor=_GREEN)
_BOLD = Font(bold=True)
_MONTH = Font(bold=True, size=12, color="1F4E2A")
_CENTER = Alignment(horizontal="center")


def _accounts(db, array_id: int):
    return db.execute(
        select(UtilityAccount).where(UtilityAccount.array_id == array_id,
                                     UtilityAccount.deleted_at.is_(None))
    ).scalars().all()


def _providers_for_array(db, array_id: int) -> list[str]:
    return sorted({(ua.provider or "").upper() for ua in _accounts(db, array_id) if ua.provider})


def _bill_months(db, array_id: int) -> dict[tuple[int, int], float]:
    """Monthly generation from the utility BILLS (all of the array's utilities).
    Only GMP bills actually carry kwh_generated; co-op bills don't (their
    generation is in the meter feed), so they contribute 0 here — by design."""
    acct_ids = [ua.id for ua in _accounts(db, array_id)]
    out: dict[tuple[int, int], float] = defaultdict(float)
    if acct_ids:
        for b in db.execute(select(Bill).where(Bill.account_id.in_(acct_ids))).scalars().all():
            for key, kwh in distribute_kwh_by_calendar_day(b).items():
                out[key] += kwh
    return out


def _array_monthly(db, array_id: int, months, start, end):
    """Per (year,month) -> (kwh, source). The utility METER feed (GMP interval +
    SmartHub daily, vendor-excluded) wins where present; the GMP bill fills GMP
    months with no interval; 0 otherwise."""
    meter = _daily_generation_by_month(db, array_id, start, end)  # GMP interval + co-op daily
    bills = _bill_months(db, array_id)
    out = {}
    for key in months:
        if key in meter and meter[key]:
            out[key] = (meter[key], "meter")
        elif key in bills:
            out[key] = (bills[key], "bill")
        else:
            out[key] = (0.0, "")
    return out


def _hourly_utility_series(db, array_id: int, start, end):
    """Hourly (preferred) or daily-fallback utility generation for an array.

    GMP: re-derive hours from the raw 15-min sponge via get_hourly_series.
    SmartHub / other daily meter only: one row per day with hour=None (never
    invent 24 hourly buckets from a daily total).

    Returns list of {"day", "hour", "kwh", "intervals", "grain"} sorted.
    """
    hourly = gmp_daily_read.get_hourly_series(array_id, start=start, end=end, db=db)
    if hourly:
        return [
            {
                "day": r["day"],
                "hour": int(r["hour"]),
                "kwh": round(float(r["kwh"]), 4),
                "intervals": r.get("intervals"),
                "grain": "hour",
            }
            for r in hourly
        ]

    # Daily meter only (VEC/WEC SmartHub, or GMP with no raw sponge yet).
    rows = db.execute(
        select(DailyGeneration.day, DailyGeneration.kwh)
        .where(DailyGeneration.array_id == array_id,
               DailyGeneration.day >= start, DailyGeneration.day <= end,
               func.lower(DailyGeneration.source).notin_(_METER_EXCLUDE))
    ).all()
    by_day = {d: round(float(k or 0.0), 3) for d, k in rows}
    # Overlay GMP daily aggregates when present (higher fidelity than generic
    # DailyGeneration for GMP arrays that haven't got raw windows parsed yet).
    for r in gmp_daily_read.get_daily_series(array_id, start=start, end=end, db=db):
        by_day[r["day"]] = round(float(r["kwh"]), 3)
    return [
        {
            "day": d,
            "hour": None,
            "kwh": kwh,
            "intervals": None,
            "grain": "day",
        }
        for d, kwh in sorted(by_day.items())
    ]


def _write_summary(sh, client_name, rows, months, year, quarter):
    ncols = len(months) + 3
    end_col = chr(ord("A") + ncols - 1)
    sh["A1"] = f"{client_name} — utility generation, Q{quarter} {year}"
    sh["A1"].font = _TITLE
    sh.merge_cells(f"A1:{end_col}1")
    sh["A2"] = "Generation (kWh) by month · utility meter (GMP interval / co-op daily) or GMP bill"
    sh["A2"].font = _SUB
    sh.merge_cells(f"A2:{end_col}2")

    hdr = ["Project"] + [f"{calendar.month_name[m]} {y}" for (y, m) in months] + ["Quarter total", "Utility · source"]
    for col, label in enumerate(hdr, start=1):
        c = sh.cell(4, col, label)
        c.font, c.fill, c.alignment = _HDR, _HDR_FILL, _CENTER
    r = 5
    col_tot = defaultdict(float)
    for name, monthly, providers in rows:
        sh.cell(r, 1, name)
        qtot = 0.0
        srcs = set()
        for i, key in enumerate(months):
            kwh, src = monthly[key]
            sh.cell(r, 2 + i, round(kwh, 1))
            qtot += kwh
            col_tot[i] += kwh
            if src:
                srcs.add(src)
        sh.cell(r, 2 + len(months), round(qtot, 1)).font = _BOLD
        col_tot["q"] += qtot
        util = "/".join(providers) if providers else "—"
        sh.cell(r, 3 + len(months), f"{util} · {' + '.join(sorted(srcs)) or 'no data'}")
        r += 1
    sh.cell(r, 1, "All projects").font = _BOLD
    for i in range(len(months)):
        sh.cell(r, 2 + i, round(col_tot[i], 1)).font = _BOLD
    sh.cell(r, 2 + len(months), round(col_tot["q"], 1)).font = _BOLD
    sh.column_dimensions["A"].width = 26
    for i in range(len(months) + 2):
        sh.column_dimensions[chr(ord("B") + i)].width = 17


def _write_hourly(sh, arr, series, year, quarter):
    """Detail sheet: Date | Hour | Generation (kWh) | 15-min intervals.

    GMP rows have hour 0–23. Daily-only fallback rows leave Hour blank and note
    grain in the subtitle so we never invent 24 buckets from a day total.
    """
    by_key = {}
    grain = "hour"
    for r in series:
        if r.get("grain") == "day":
            grain = "day"
        by_key[(r["day"], r.get("hour"))] = r
    nid = getattr(arr, "nepool_gis_id", None)
    sh["A1"] = f"{arr.name}" + (f"  ·  NEPOOL-GIS {nid}" if nid else "")
    sh["A1"].font = _TITLE
    sh.merge_cells("A1:D1")
    if grain == "hour":
        sh["A2"] = (
            f"Hourly utility meter (from GMP 15-min intervals) · Q{quarter} {year}"
        )
    else:
        sh["A2"] = (
            f"Daily utility meter (hourly not available for this utility) · "
            f"Q{quarter} {year}"
        )
    sh["A2"].font = _SUB
    sh.merge_cells("A2:D2")
    row = 4
    for (y, m) in _quarter_months(year, quarter):
        sh.cell(row, 1, f"{calendar.month_name[m]} {y}").font = _MONTH
        row += 1
        for col, label in enumerate(
            ["Date", "Hour (0–23)", "Generation (kWh)", "15-min intervals"], start=1
        ):
            c = sh.cell(row, col, label)
            c.font, c.fill, c.alignment = _HDR, _HDR_FILL, _CENTER
        row += 1
        mt = 0.0
        if grain == "hour":
            for dd in range(1, calendar.monthrange(y, m)[1] + 1):
                d = date(y, m, dd)
                day_total = 0.0
                wrote_any = False
                for h in range(24):
                    cell = by_key.get((d, h))
                    if cell is None:
                        continue
                    wrote_any = True
                    sh.cell(row, 1, d.isoformat())
                    sh.cell(row, 2, h)
                    sh.cell(row, 3, cell["kwh"])
                    if cell.get("intervals") is not None:
                        sh.cell(row, 4, cell["intervals"])
                    day_total += cell["kwh"]
                    row += 1
                if wrote_any:
                    mt += day_total
        else:
            for dd in range(1, calendar.monthrange(y, m)[1] + 1):
                d = date(y, m, dd)
                cell = by_key.get((d, None))
                if cell is None:
                    continue
                sh.cell(row, 1, d.isoformat())
                # Hour left blank — daily total only.
                sh.cell(row, 3, cell["kwh"])
                mt += cell["kwh"]
                row += 1
        sh.cell(row, 1, f"{calendar.month_name[m]} total").font = _BOLD
        sh.cell(row, 3, round(mt, 4)).font = _BOLD
        row += 2
    sh.column_dimensions["A"].width = 14
    sh.column_dimensions["B"].width = 12
    sh.column_dimensions["C"].width = 18
    sh.column_dimensions["D"].width = 16


def build_generation_workbook(client_id: int, out_path, *, year: int, quarter: int) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    months = _quarter_months(year, quarter)
    start = date(months[0][0], months[0][1], 1)
    ly, lm = months[-1]
    end = date(ly, lm, calendar.monthrange(ly, lm)[1])

    with SessionLocal() as db:
        client = db.get(Client, client_id)
        if client is None:
            raise ValueError("client not found")
        arrays = db.execute(
            select(Array).where(Array.client_id == client_id,
                                Array.deleted_at.is_(None),
                                Array.excluded.is_(False),
                                not_vendor_only()).order_by(Array.name)
        ).scalars().all()

        summary_rows = []
        detail = []  # (array, series)
        for arr in arrays:
            monthly = _array_monthly(db, arr.id, months, start, end)
            if any(v[0] for v in monthly.values()):
                summary_rows.append((arr.name, monthly, _providers_for_array(db, arr.id)))
            series = _hourly_utility_series(db, arr.id, start, end)
            if series:
                detail.append((arr, series))

        wb = Workbook()
        summary = wb.active
        summary.title = "Monthly Summary"
        if summary_rows:
            _write_summary(summary, client.name, summary_rows, months, year, quarter)
        else:
            summary["A1"] = f"No utility generation for {client.name} in Q{quarter} {year}."
            summary["A1"].font = _TITLE

        used = {"Monthly Summary"}
        for arr, series in detail:
            sh = wb.create_sheet(title=_sheet_name_for_array(arr.name, used))
            _write_hourly(sh, arr, series, year, quarter)

    wb.save(str(out_path))
    return out_path


def _write_all_summary(sh, tenant_name, rows, months, year, quarter):
    """rows: [(client_name, project_name, monthly, providers)] across all clients."""
    ncols = len(months) + 4
    end_col = chr(ord("A") + ncols - 1)
    sh["A1"] = f"{tenant_name} — utility generation (all clients), Q{quarter} {year}"
    sh["A1"].font = _TITLE
    sh.merge_cells(f"A1:{end_col}1")
    sh["A2"] = "Generation (kWh) by month · utility meter (GMP interval / co-op daily) or GMP bill"
    sh["A2"].font = _SUB
    sh.merge_cells(f"A2:{end_col}2")

    hdr = (["Client", "Project"]
           + [f"{calendar.month_name[m]} {y}" for (y, m) in months]
           + ["Quarter total", "Utility · source"])
    for col, label in enumerate(hdr, start=1):
        c = sh.cell(4, col, label)
        c.font, c.fill, c.alignment = _HDR, _HDR_FILL, _CENTER
    r = 5
    grand = defaultdict(float)
    last_client = None
    for cname, pname, monthly, providers in rows:
        sh.cell(r, 1, cname if cname != last_client else "")  # show client once per group
        last_client = cname
        sh.cell(r, 2, pname)
        qtot = 0.0
        srcs = set()
        for i, key in enumerate(months):
            kwh, src = monthly[key]
            sh.cell(r, 3 + i, round(kwh, 1))
            qtot += kwh
            grand[i] += kwh
            if src:
                srcs.add(src)
        sh.cell(r, 3 + len(months), round(qtot, 1)).font = _BOLD
        grand["q"] += qtot
        util = "/".join(providers) if providers else "—"
        sh.cell(r, 4 + len(months), f"{util} · {' + '.join(sorted(srcs)) or 'no data'}")
        r += 1
    sh.cell(r, 1, "All clients").font = _BOLD
    for i in range(len(months)):
        sh.cell(r, 3 + i, round(grand[i], 1)).font = _BOLD
    sh.cell(r, 3 + len(months), round(grand["q"], 1)).font = _BOLD
    sh.column_dimensions["A"].width = 24
    sh.column_dimensions["B"].width = 24
    for i in range(len(months) + 2):
        sh.column_dimensions[chr(ord("C") + i)].width = 16


def build_all_clients_generation_workbook(tenant_id: str, out_path, *, year: int, quarter: int) -> Path:
    """One workbook covering EVERY client's utility generation for the quarter —
    a combined Generation Summary (client × project × months) plus an hourly
    meter detail sheet per project that has one. All-clients counterpart of the
    per-client export."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    months = _quarter_months(year, quarter)
    start = date(months[0][0], months[0][1], 1)
    ly, lm = months[-1]
    end = date(ly, lm, calendar.monthrange(ly, lm)[1])

    with SessionLocal() as db:
        tenant = db.get(Tenant, tenant_id)
        tenant_name = (tenant.company_name or tenant.name) if tenant else "Your fleet"
        clients = db.execute(
            select(Client).where(Client.tenant_id == tenant_id,
                                 Client.deleted_at.is_(None)).order_by(Client.name)
        ).scalars().all()

        all_rows = []   # (client_name, project_name, monthly, providers)
        detail = []     # (client_name, array, series)
        for c in clients:
            arrays = db.execute(
                select(Array).where(Array.client_id == c.id,
                                    Array.deleted_at.is_(None),
                                    Array.excluded.is_(False),
                                    not_vendor_only()).order_by(Array.name)
            ).scalars().all()
            for arr in arrays:
                monthly = _array_monthly(db, arr.id, months, start, end)
                if any(v[0] for v in monthly.values()):
                    all_rows.append((c.name, arr.name, monthly, _providers_for_array(db, arr.id)))
                series = _hourly_utility_series(db, arr.id, start, end)
                if series:
                    detail.append((c.name, arr, series))

        wb = Workbook()
        summary = wb.active
        summary.title = "Generation Summary"
        if all_rows:
            _write_all_summary(summary, tenant_name, all_rows, months, year, quarter)
        else:
            summary["A1"] = f"No utility generation across your clients in Q{quarter} {year}."
            summary["A1"].font = _TITLE

        used = {"Generation Summary"}
        for cname, arr, series in detail:
            sh = wb.create_sheet(title=_sheet_name_for_array(f"{cname[:9]}·{arr.name}", used))
            _write_hourly(sh, arr, series, year, quarter)

    wb.save(str(out_path))
    return out_path


# ── Quarterly Summary (Crown / GMP-style one-pager) ───────────────────────────
# Screenshot contract (Ford 2026-07-23): a single "Summary" sheet —
#   Name | Account # | {Month1} | {Month2} | {Month3} | Total
# One row per utility account under a report-eligible array. Month headers are
# bare month names (January, February, …) matching the operator spreadsheet.

_SUMMARY_HDR_FILL = PatternFill("solid", fgColor="C6EFCE")  # light green like the shot
_SUMMARY_HDR_FONT = Font(bold=True, size=11)
_SUMMARY_NUM = "0.000"


def _monitoring_months(
    db, array_id: int, months, start: date, end: date,
) -> dict[tuple[int, int], float]:
    """Monthly kWh from the array's MONITORING feed (Locus, AlsoEnergy, eGauge,
    Meter Mate, …) — the native production tier for arrays with no utility
    account. Utility daily still outranks this where both cover a month; callers
    merge utility over the top."""
    out: dict[tuple[int, int], float] = {k: 0.0 for k in months}
    rows = db.execute(
        select(DailyGeneration.day, DailyGeneration.kwh).where(
            DailyGeneration.array_id == array_id,
            DailyGeneration.day >= start,
            DailyGeneration.day <= end,
            func.lower(DailyGeneration.source).in_(MONITORING_REPORT_SOURCES),
        )
    ).all()
    for d, kwh in rows:
        key = (d.year, d.month)
        if key in out:
            out[key] += float(kwh or 0.0)
    return {k: round(v, 6) for k, v in out.items()}


def _has_monitoring_history(db, array_id: int) -> bool:
    """True if the array has any positive monitoring-tier DailyGeneration row.
    Mirrors gmcs_writer._array_has_monitoring_history so a vendor-only array
    (Locus/AlsoEnergy, no utility account) with real production is admitted to
    the summary; an empty vendor twin (Fronius twin, no production) stays out."""
    return db.execute(
        select(DailyGeneration.id).where(
            DailyGeneration.array_id == array_id,
            DailyGeneration.kwh > 0,
            func.lower(DailyGeneration.source).in_(MONITORING_REPORT_SOURCES),
        ).limit(1)
    ).first() is not None


def _account_monthly_kwh(
    db, ua: UtilityAccount, months, start: date, end: date, *, solo_array: bool,
) -> dict[tuple[int, int], float]:
    """Monthly kWh for ONE utility account across the quarter months.

    Priority:
      1. GmpDailyGeneration (per-meter GMP intervals) — true account level.
      2. Bills on this account with kwh_generated (calendar-day attribution).
      3. If this is the array's only live utility account, fall back to the
         array-level utility meter feed (SmartHub daily etc.).
    """
    out: dict[tuple[int, int], float] = {k: 0.0 for k in months}

    # 1) GMP per-account daily
    gmp_rows = db.execute(
        select(GmpDailyGeneration.day, GmpDailyGeneration.kwh).where(
            GmpDailyGeneration.account_id == ua.id,
            GmpDailyGeneration.day >= start,
            GmpDailyGeneration.day <= end,
        )
    ).all()
    if gmp_rows:
        for d, kwh in gmp_rows:
            key = (d.year, d.month)
            if key in out:
                out[key] += float(kwh or 0.0)
        return {k: round(v, 6) for k, v in out.items()}

    # 2) Bills on this account
    bills = db.execute(select(Bill).where(Bill.account_id == ua.id)).scalars().all()
    bill_hit = False
    for b in bills:
        for key, kwh in distribute_kwh_by_calendar_day(b).items():
            if key in out and kwh:
                out[key] += float(kwh)
                bill_hit = True
    if bill_hit:
        return {k: round(v, 6) for k, v in out.items()}

    # 3) Solo-account array → whole array meter feed
    if solo_array and ua.array_id is not None:
        meter = _daily_generation_by_month(db, ua.array_id, start, end)
        for key in months:
            if key in meter and meter[key]:
                out[key] = float(meter[key])
    return {k: round(v, 6) for k, v in out.items()}


def build_quarterly_summary_workbook(
    tenant_id: str, out_path, *, year: int, quarter: int,
) -> Path:
    """Fleet quarterly summary matching the operator spreadsheet screenshot.

    Sheet "Summary":
      Name | Account # | January | February | March | Total
    (month names follow the selected quarter — Q2 → April May June, etc.)

    One row per live UtilityAccount linked to a non-vendor-only, non-excluded
    array under any of the tenant's clients. Sorted by array name then account #.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    months = _quarter_months(year, quarter)
    start = date(months[0][0], months[0][1], 1)
    ly, lm = months[-1]
    end = date(ly, lm, calendar.monthrange(ly, lm)[1])
    month_labels = [calendar.month_name[m] for (_y, m) in months]

    with SessionLocal() as db:
        clients = db.execute(
            select(Client).where(
                Client.tenant_id == tenant_id,
                Client.deleted_at.is_(None),
                Client.active == True,  # noqa: E712
            )
        ).scalars().all()
        client_ids = [c.id for c in clients]

        arrays = []
        if client_ids:
            arrays = db.execute(
                select(Array).where(
                    Array.tenant_id == tenant_id,
                    Array.client_id.in_(client_ids),
                    Array.deleted_at.is_(None),
                    Array.excluded.is_(False),
                ).order_by(Array.name)
            ).scalars().all()
            # Native multi-source inclusion (mirrors gmcs_writer): utility-backed
            # arrays + bare onboarding stubs, PLUS vendor-only arrays that carry
            # real monitoring production (Locus / AlsoEnergy / eGauge / …). Empty
            # vendor twins (a Fronius twin with no production of its own) stay out
            # so they can't double-report next to a utility-backed sibling — and
            # so byte-pinned utility-only workbooks (Bruce) are unchanged.
            arrays = [
                a for a in arrays
                if not is_vendor_only_array(db, a.id)
                or _has_monitoring_history(db, a.id)
            ]

        # Pre-count live UAs per array for solo-account fallback.
        array_ids = [a.id for a in arrays]
        ua_count: dict[int, int] = defaultdict(int)
        all_uas: list[UtilityAccount] = []
        if array_ids:
            all_uas = list(db.execute(
                select(UtilityAccount).where(
                    UtilityAccount.array_id.in_(array_ids),
                    UtilityAccount.deleted_at.is_(None),
                ).order_by(UtilityAccount.account_number)
            ).scalars().all())
            for ua in all_uas:
                if ua.array_id is not None:
                    ua_count[ua.array_id] += 1

        uas_by_array: dict[int, list[UtilityAccount]] = defaultdict(list)
        for ua in all_uas:
            if ua.array_id is not None:
                uas_by_array[ua.array_id].append(ua)

        rows = []  # (name, account_number, monthly_dict)
        for arr in arrays:
            uas = uas_by_array.get(arr.id) or []
            if not uas:
                # No utility account — either a bare onboarding stub (all zeros,
                # so the operator sees the gap) or a monitoring-only array
                # (Locus / AlsoEnergy / eGauge). Use the array-level feed:
                # utility daily wins a month, else monitoring, else 0. Account #
                # is blank — these arrays settle off a monitor, not a meter.
                util = _daily_generation_by_month(db, arr.id, start, end)
                mon = _monitoring_months(db, arr.id, months, start, end)
                monthly = {
                    k: float(util.get(k) or mon.get(k) or 0.0) for k in months
                }
                rows.append((arr.name, "", monthly))
                continue
            solo = ua_count.get(arr.id, 0) == 1
            for ua in uas:
                monthly = _account_monthly_kwh(
                    db, ua, months, start, end, solo_array=solo,
                )
                rows.append((arr.name, ua.account_number or "", monthly))

        # Sort by name then account # (stable, spreadsheet-friendly).
        rows.sort(key=lambda r: (r[0].lower(), r[1]))

        wb = Workbook()
        sh = wb.active
        sh.title = "Summary"

        headers = ["Name", "Account #"] + month_labels + ["Total"]
        for col, label in enumerate(headers, start=1):
            cell = sh.cell(1, col, label)
            cell.font = _SUMMARY_HDR_FONT
            if col == 1:
                cell.fill = _SUMMARY_HDR_FILL

        r = 2
        for name, acct, monthly in rows:
            sh.cell(r, 1, name)
            # Account # as text so leading zeros / long IDs don't scientific-note.
            sh.cell(r, 2, str(acct) if acct else "")
            total = 0.0
            for i, key in enumerate(months):
                kwh = float(monthly.get(key) or 0.0)
                cell = sh.cell(r, 3 + i, round(kwh, 3))
                cell.number_format = _SUMMARY_NUM
                total += kwh
            tcell = sh.cell(r, 3 + len(months), round(total, 3))
            tcell.number_format = _SUMMARY_NUM
            tcell.font = _BOLD
            r += 1

        sh.column_dimensions["A"].width = 28
        sh.column_dimensions["B"].width = 14
        for i in range(len(months) + 1):
            sh.column_dimensions[chr(ord("C") + i)].width = 12

        # Freeze header row like a working spreadsheet.
        sh.freeze_panes = "A2"

    wb.save(str(out_path))
    return out_path
