"""
Generic non-solar REC writer.

Produces the SAME array×months×MWh→REC(floor) quarterly workbook as the sacred
``gmcs_writer`` (one sheet per array, rolling 6 complete quarters of 3 month
rows + a gap row, MWh in Excel "General" format, RECs = floor(MWh), uniform
column widths) — but with **fuel-correct labels** instead of the hard-coded
GMCS/NEPOOL-GIS solar wording:

  - the sheet title reflects the fuel (e.g. "Ridgeline Wind (GIS123) · Wind"),
  - the RECs header column names the fuel ("Wind RECs†"),
  - the certificate-registry id is read from the array (``cert_registry`` /
    ``nepool_gis_id``) rather than a literal "NEPOOL-GIS",
  - the footnote is a clean, generic REC-attestation sentence — *not* the
    verbatim GMCS footnote (which is pinned to Bruce's master and must never
    appear on a non-solar report).

This module deliberately re-implements the data-gathering + render loop rather
than refactoring ``gmcs_writer`` so the solar writer stays byte-for-byte
untouched. It reuses ``gmcs_writer``'s pure math helpers (quarter windows,
daily-generation aggregation, sheet-name sanitising) so the two stay in
lockstep on the arithmetic that matters: REC = int(MWh).
"""
from __future__ import annotations

import pathlib
from collections import defaultdict
from datetime import date, timedelta
from typing import Optional

from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ..bill_attribution import distribute_kwh_by_calendar_day
from ..db import SessionLocal, DATA_DIR
from ..models import Tenant, Client, UtilityAccount, Array, Bill, DailyGeneration

# Reuse the sacred writer's pure helpers so the quarter math + REC floor stay
# identical. None of these touch GMCS-specific formatting or wording.
from .gmcs_writer import (
    _rolling_quarters,
    _quarter_months,
    _sheet_name_for_array,
    _daily_generation_by_month,
    default_reporting_reference_date,
)

REPORTS_DIR = DATA_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True, parents=True)


# Human-facing fuel labels. Unknown fuels fall back to a Title-cased form of the
# raw value (or a neutral "REC" when entirely absent).
FUEL_LABELS = {
    "solar": "Solar",
    "wind": "Wind",
    "hydro": "Hydro",
    "digester": "Digester",
    "storage": "Storage",
}

# Default certificate registry. New-England RECs all clear through NEPOOL-GIS,
# but this is parameterised per-array (``cert_registry``) rather than hard-coded
# into the body of the report, so other registries can be wired in later.
DEFAULT_CERT_REGISTRY = "NEPOOL-GIS"


def _fuel_label(fuel: Optional[str]) -> str:
    f = (fuel or "solar").strip().lower()
    return FUEL_LABELS.get(f, (f or "REC").title())


def _array_fuel(arr: Optional[Array]) -> str:
    """Fuel type for an array, defensively defaulting to 'solar' so this module
    is correct whether or not the Array.fuel_type column has landed yet."""
    if arr is None:
        return "solar"
    return (getattr(arr, "fuel_type", None) or "solar").strip().lower()


def _array_registry(arr: Optional[Array]) -> str:
    """Certificate-registry name for an array (e.g. 'NEPOOL-GIS'). Read from the
    array if a cert_registry field exists, else the default."""
    if arr is None:
        return DEFAULT_CERT_REGISTRY
    return (getattr(arr, "cert_registry", None) or DEFAULT_CERT_REGISTRY).strip()


def _array_asset_id(arr: Optional[Array]) -> Optional[str]:
    """The registry asset id shown in the title — the cert_registry id, not a
    literal 'NEPOOL-GIS'. Falls back to the legacy nepool_gis_id column."""
    if arr is None:
        return None
    return (
        getattr(arr, "cert_registry_id", None)
        or getattr(arr, "nepool_gis_id", None)
    )


def _generic_footnote(registry: str, fuel_label: str) -> str:
    """A clean, generic REC-attestation footnote. Deliberately NOT the verbatim
    GMCS footnote (which is pinned to Bruce's solar master)."""
    fuel_lc = fuel_label.lower()
    return (
        f" † {registry} awards 1 REC for every whole MWh of {fuel_lc} "
        f"generation reported. Fractional MWh are tracked and an additional "
        f"REC is awarded once the running total exceeds 1 MWh."
    )


def build_workbook(tenant_id: Optional[str] = None,
                   year: Optional[int] = None,
                   out_path: Optional[pathlib.Path] = None,
                   *, quarters: int = 6,
                   reference_date: Optional[date] = None,
                   client_id: Optional[int] = None) -> pathlib.Path:
    """Generate a generic (non-solar) REC workbook for ONE client.

    Signature is identical to ``gmcs_writer.build_workbook`` so the writer
    registry can dispatch to either uniformly. REC math (int(MWh) floor),
    quarter windows, grouping and daily-generation precedence are identical to
    the solar writer; only the labels and footnote are fuel-aware.
    """
    if client_id is None and tenant_id is None:
        raise ValueError("build_workbook requires client_id or tenant_id")

    ref = reference_date if reference_date is not None \
        else default_reporting_reference_date(date.today())
    qlist = _rolling_quarters(ref, count=quarters)
    last_y, last_q = qlist[-1]
    if year is None:
        year = last_y

    with SessionLocal() as db:
        # Resolve client_id from tenant_id when legacy mode used.
        if client_id is None:
            client = db.execute(
                select(Client).where(Client.tenant_id == tenant_id)
                              .order_by(Client.id.asc())
            ).scalars().first()
            if client is None:
                raise ValueError(
                    f"Tenant {tenant_id} has no Client rows; run migrations.")
            client_id = client.id
        else:
            client = db.get(Client, client_id)
            if client is None:
                raise ValueError(f"unknown client {client_id}")

        tenant = db.get(Tenant, client.tenant_id)
        if not tenant:
            raise ValueError(f"unknown tenant {client.tenant_id}")

        if out_path is None:
            out_path = (REPORTS_DIR / tenant.id / f"client-{client.id}"
                        / f"{last_y}-Q{last_q}-REC-report.xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # Arrays scoped to THIS client only; skip excluded arrays (same rule as
        # the solar writer).
        arrays = db.execute(
            select(Array).where(
                Array.client_id == client.id,
                Array.excluded.is_(False),
            )
        ).scalars().all()
        arrays_by_id = {a.id: a for a in arrays}
        array_ids = list(arrays_by_id.keys())

        if array_ids:
            accounts = db.execute(
                select(UtilityAccount).where(UtilityAccount.array_id.in_(array_ids))
            ).scalars().all()
        else:
            accounts = []

        def group_for(acc: UtilityAccount) -> tuple[str, Optional[Array]]:
            if acc.array_id and acc.array_id in arrays_by_id:
                a = arrays_by_id[acc.array_id]
                return a.name, a
            return (acc.nickname or acc.account_number), None

        group_of: dict[int, str] = {}
        group_meta: dict[str, Optional[Array]] = {}
        for acc in accounts:
            name, a = group_for(acc)
            group_of[acc.id] = name
            group_meta.setdefault(name, a)

        if accounts:
            account_ids = [a.id for a in accounts]
            bills = db.execute(
                select(Bill).where(Bill.account_id.in_(account_ids))
            ).scalars().all()
        else:
            bills = []

        # Per-group kWh by (year, month) — pro-rated by calendar day (identical
        # to the solar writer).
        per_group: dict[str, dict[tuple[int, int], float]] = defaultdict(
            lambda: defaultdict(float))
        for b in bills:
            grp = group_of.get(b.account_id)
            if not grp:
                continue
            for (yy, mm), kwh in distribute_kwh_by_calendar_day(b).items():
                per_group[grp][(yy, mm)] += kwh

        groups = sorted(set(group_of.values()))

        # DailyGeneration precedence window.
        start_year, start_q = qlist[0]
        report_start = date(start_year, (start_q - 1) * 3 + 1, 1)
        end_year, end_q = qlist[-1]
        end_month = end_q * 3
        if end_month == 12:
            report_end = date(end_year, 12, 31)
        else:
            report_end = date(end_year, end_month + 1, 1) - timedelta(days=1)

        daily_gen_by_group: dict[str, dict[tuple[int, int], float]] = {}
        for grp_name, arr in group_meta.items():
            if arr is not None:
                dg = _daily_generation_by_month(db, arr.id, report_start, report_end)
                if dg:
                    daily_gen_by_group[grp_name] = dg

        # Snapshot the per-array fuel/registry/asset metadata while the session
        # is open (Array objects detach after the `with` block).
        group_fuel: dict[str, str] = {}
        group_registry: dict[str, str] = {}
        group_asset: dict[str, Optional[str]] = {}
        for grp_name, arr in group_meta.items():
            group_fuel[grp_name] = _array_fuel(arr)
            group_registry[grp_name] = _array_registry(arr)
            group_asset[grp_name] = _array_asset_id(arr)

    # Non-producing arrays get no sheet — identical rule to gmcs_writer: a
    # group with zero generation across every month of the reporting window
    # (bills AND daily data, daily-over-bill precedence) is omitted so the
    # registry upload only carries arrays with reportable generation.
    window_months = {m for (qy, qq) in qlist for m in _quarter_months(qy, qq)}
    groups = [
        grp for grp in groups
        if any(
            {**per_group.get(grp, {}), **daily_gen_by_group.get(grp, {})}
            .get(m, 0.0) > 0
            for m in window_months
        )
    ]

    # ── Build workbook ──────────────────────────────────────────────
    wb = Workbook()
    default_sheet = wb.active

    TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
    HDR_FONT = Font(bold=True, size=14, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    QUARTER_FONT = Font(bold=True, size=11, color="1E3A5F")
    FOOTNOTE_FONT = Font(italic=True, size=9, color="666666")
    BORDER = Border(*[Side(style="thin", color="D9E2EC")] * 4)
    ACCENT_SIDE = Side(style="medium", color="7FA8D9")

    used_names: set[str] = set()

    if not groups:
        groups = ["(no data)"]

    for grp_idx, grp in enumerate(groups):
        sheet_title = _sheet_name_for_array(grp, used_names)
        if grp_idx == 0:
            sh = default_sheet
            sh.title = sheet_title
        else:
            sh = wb.create_sheet(title=sheet_title)

        fuel = group_fuel.get(grp, "solar")
        fuel_label = _fuel_label(fuel)
        registry = group_registry.get(grp, DEFAULT_CERT_REGISTRY)
        asset_id = group_asset.get(grp)

        # ── Title (A1 merged A1:C1) — name (asset id) · Fuel ──
        if asset_id:
            title = f"{grp} ({asset_id}) \u00b7 {fuel_label}"
        else:
            title = f"{grp} \u00b7 {fuel_label}"
        sh["A1"] = title
        sh["A1"].font = TITLE_FONT
        sh["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells("A1:C1")

        # Row 5 header — RECs column names the fuel.
        header_labels = [
            "Quarter",
            "Generation (MWh)",
            "Reporting Amount",
            f"{fuel_label} RECs\u2020",
        ]
        for col, label in enumerate(header_labels, start=1):
            c = sh.cell(5, col, label)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

        for col in range(1, 5):
            sh.cell(6, col).border = Border(bottom=ACCENT_SIDE)

        # Data: 6 quarter blocks × 3 month rows; gap row between blocks.
        row = 7
        bill_months = per_group.get(grp, {})
        daily_months = daily_gen_by_group.get(grp, {})
        gen_by_month: dict[tuple[int, int], float] = {**bill_months, **daily_months}
        for (qy, qq) in qlist:
            for i, (my, mm) in enumerate(_quarter_months(qy, qq)):
                if i == 0:
                    qc = sh.cell(row, 1, f"Q{qq} {qy}")
                    qc.font = QUARTER_FONT
                    qc.alignment = Alignment(horizontal="left", vertical="center")
                kwh = gen_by_month.get((my, mm), 0.0)
                mwh = round(kwh / 1000.0, 3)
                recs = int(mwh)  # floor of MWh — IDENTICAL to gmcs_writer
                gc = sh.cell(row, 2, mwh if kwh else None)
                gc.number_format = "General"
                gc.alignment = Alignment(horizontal="right")
                rc = sh.cell(row, 3, mwh if kwh else None)
                rc.number_format = "General"
                rc.alignment = Alignment(horizontal="right")
                ec = sh.cell(row, 4, recs if kwh else None)
                ec.number_format = "General"
                ec.alignment = Alignment(horizontal="right")
                row += 1
            row += 1  # gap row between quarter blocks

        # Footnote — generic REC attestation (NOT the verbatim GMCS footnote).
        foot_row = 31 if row <= 31 else row
        fc = sh.cell(foot_row, 1, _generic_footnote(registry, fuel_label))
        fc.font = FOOTNOTE_FONT
        fc.alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells(start_row=foot_row, start_column=1,
                       end_row=foot_row, end_column=4)

        # Column widths — uniform 24, matching the solar layout.
        for col_letter in ("A", "B", "C", "D"):
            sh.column_dimensions[col_letter].width = 24.0

    wb.save(out_path)
    return out_path
