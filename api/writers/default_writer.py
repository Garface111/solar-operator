"""
Default writer — generates a clean arrays × months workbook from a tenant's
Bill rows. Used for "Make one for me" customers who didn't upload their
own template.

Layout:
  Sheet 1: <year> kWh Monthly
    Row 5: header row (Array | Region | Jan..Dec | YTD)
    Rows 6-N: one row per array (group of accounts), totals at bottom

Month attribution rule (general, applies to all default-writer tenants):
  - Default: bill represents the PRIOR month (use period_start.month)
  - This matches most utilities. Customers with a different rule (e.g.
    Bruce's Starlake same-month rule) should upload their own template
    and get a custom writer.

Account → Array grouping:
  - Default: each UtilityAccount.nickname IS the array name (1:1)
  - If account has array_id set (via the Arrays table), use that grouping
    instead. This lets future onboarding ask "do you sum Starlake N/S/C
    into one Starlake array?" and we honor it.
"""
from __future__ import annotations
import pathlib
from collections import defaultdict
from datetime import datetime
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..db import SessionLocal, DATA_DIR
from ..models import Tenant, UtilityAccount, Array, Bill


REPORTS_DIR = DATA_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True, parents=True)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _bill_target_month(bill: Bill) -> tuple[int, int] | None:
    """Return (year, month_1_to_12) this bill is attributed to, or None
    if no usable date is present."""
    # Prefer period_start (covers the actual generation window)
    src = bill.period_start or bill.bill_date
    if not src:
        return None
    return src.year, src.month


def build_workbook(tenant_id: str, year: int | None = None,
                   out_path: pathlib.Path | None = None) -> pathlib.Path:
    """Generate the default workbook for one tenant. Returns the saved path."""
    if year is None:
        year = datetime.utcnow().year
    if out_path is None:
        out_path = REPORTS_DIR / tenant_id / f"{year}-monthly-kwh.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with SessionLocal() as db:
        tenant = db.get(Tenant, tenant_id)
        if not tenant:
            raise ValueError(f"unknown tenant {tenant_id}")

        accounts = db.execute(
            select(UtilityAccount).where(UtilityAccount.tenant_id == tenant_id)
        ).scalars().all()

        # Pull arrays-table rows for grouping if any exist
        arrays_by_id = {a.id: a for a in db.execute(
            select(Array).where(Array.tenant_id == tenant_id)
        ).scalars().all()}

        # Build account → group label
        # Group key strategy:
        #   1. If account has array_id and that array exists → array.name
        #   2. Else fallback to account.nickname (1:1)
        def group_key(acc: UtilityAccount) -> str:
            if acc.array_id and acc.array_id in arrays_by_id:
                return arrays_by_id[acc.array_id].name
            return acc.nickname or acc.account_number

        group_of = {acc.id: group_key(acc) for acc in accounts}
        group_region = {}
        for acc in accounts:
            arr = arrays_by_id.get(acc.array_id) if acc.array_id else None
            group_region[group_of[acc.id]] = (arr.region or "") if arr else ""

        # Pull bills for this tenant
        bills = db.execute(
            select(Bill).where(Bill.tenant_id == tenant_id)
        ).scalars().all()

        # Aggregate: group → month → kWh
        per_group: dict[str, dict[int, float]] = defaultdict(lambda: defaultdict(float))
        per_group_bill_count: dict[str, int] = defaultdict(int)
        for b in bills:
            if b.kwh_generated is None or b.kwh_generated <= 0:
                continue
            target = _bill_target_month(b)
            if target is None or target[0] != year:
                continue
            grp = group_of.get(b.account_id)
            if not grp:
                continue
            per_group[grp][target[1]] += b.kwh_generated
            per_group_bill_count[grp] += 1

        # Stable order: alphabetical group names (deterministic across runs)
        groups = sorted(per_group.keys()) or sorted(set(group_of.values()))

    # ── Build workbook ────────────────────────────────────────────
    wb = Workbook()
    sh = wb.active
    sh.title = f"{year} kWh Monthly"

    HDR = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill("solid", fgColor="2E6B3A")
    TOTAL_FILL = PatternFill("solid", fgColor="EEF3EC")
    TOTAL_FONT = Font(bold=True)
    BORDER = Border(*[Side(style="thin", color="C8D4C4")] * 4)

    sh["A1"] = f"{tenant.name} — {year} Monthly kWh"
    sh["A1"].font = Font(bold=True, size=14, color="2E6B3A")
    sh.merge_cells("A1:O1")
    sh["A2"] = (f"Source: utility JSON bills · "
                f"Generated {datetime.now():%B %d, %Y %I:%M %p}")
    sh["A2"].font = Font(italic=True, size=10, color="666666")
    sh.merge_cells("A2:O2")
    sh["A3"] = "Month attribution: bill period start (PRIOR month convention)"
    sh["A3"].font = Font(italic=True, size=9, color="888888")
    sh.merge_cells("A3:O3")

    # Header row 5
    sh.cell(5, 1, "Array")
    sh.cell(5, 2, "Region")
    for i, m in enumerate(MONTHS):
        sh.cell(5, 3 + i, m)
    sh.cell(5, 15, "YTD")
    for col in range(1, 16):
        c = sh.cell(5, col)
        c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    # Data rows
    totals = [0.0] * 12
    for r_idx, grp in enumerate(groups):
        row = 6 + r_idx
        sh.cell(row, 1, grp).font = Font(bold=True)
        sh.cell(row, 2, group_region.get(grp, ""))
        ytd = 0.0
        for m in range(1, 13):
            v = per_group[grp].get(m, 0)
            if v:
                sh.cell(row, 2 + m, round(v))
                ytd += v
                totals[m - 1] += v
            cell = sh.cell(row, 2 + m)
            cell.number_format = "#,##0"
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
        c = sh.cell(row, 15, round(ytd))
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right")
        for cc in (1, 2):
            sh.cell(row, cc).border = BORDER

    trow = 6 + len(groups)
    sh.cell(trow, 1, "TOTAL").font = TOTAL_FONT
    for m in range(12):
        c = sh.cell(trow, 3 + m, round(totals[m]))
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL
        c.number_format = "#,##0"; c.border = BORDER
        c.alignment = Alignment(horizontal="right")
    yc = sh.cell(trow, 15, round(sum(totals)))
    yc.font = TOTAL_FONT; yc.fill = TOTAL_FILL
    yc.number_format = "#,##0"; yc.border = BORDER
    yc.alignment = Alignment(horizontal="right")
    for cc in (1, 2):
        sh.cell(trow, cc).fill = TOTAL_FILL
        sh.cell(trow, cc).border = BORDER

    sh.column_dimensions["A"].width = 20
    sh.column_dimensions["B"].width = 10
    for i in range(3, 16):
        sh.column_dimensions[get_column_letter(i)].width = 10

    # ── Per-account audit sheet ───────────────────────────────────
    sh2 = wb.create_sheet("Per-Account Detail")
    sh2["A1"] = "Per-account monthly kWh"
    sh2["A1"].font = Font(bold=True, size=12, color="2E6B3A")
    sh2.merge_cells("A1:R1")
    sh2.cell(3, 1, "Nickname"); sh2.cell(3, 2, "Account #")
    sh2.cell(3, 3, "Array"); sh2.cell(3, 4, "Bills")
    for i, m in enumerate(MONTHS):
        sh2.cell(3, 5 + i, m)
    sh2.cell(3, 17, "YTD")
    for col in range(1, 18):
        c = sh2.cell(3, col)
        c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    # Per-account aggregation
    with SessionLocal() as db:
        for r_idx, acc in enumerate(accounts):
            row = 4 + r_idx
            sh2.cell(row, 1, acc.nickname or "")
            sh2.cell(row, 2, acc.account_number)
            sh2.cell(row, 3, group_of[acc.id])

            acc_bills = db.execute(
                select(Bill).where(Bill.account_id == acc.id)
            ).scalars().all()

            months = defaultdict(float)
            count = 0
            for b in acc_bills:
                if b.kwh_generated is None or b.kwh_generated <= 0:
                    continue
                target = _bill_target_month(b)
                if target is None or target[0] != year:
                    continue
                months[target[1]] += b.kwh_generated
                count += 1

            sh2.cell(row, 4, count)
            ytd = 0
            for m in range(1, 13):
                v = months.get(m, 0)
                if v:
                    sh2.cell(row, 4 + m, round(v)).number_format = "#,##0"
                    ytd += v
            sh2.cell(row, 17, round(ytd)).number_format = "#,##0"
            for cc in range(1, 18):
                sh2.cell(row, cc).border = BORDER
                if cc >= 5:
                    sh2.cell(row, cc).alignment = Alignment(horizontal="right")

    for col, w in [(1, 20), (2, 14), (3, 16), (4, 7)]:
        sh2.column_dimensions[get_column_letter(col)].width = w
    for col in range(5, 18):
        sh2.column_dimensions[get_column_letter(col)].width = 9

    wb.save(out_path)
    return out_path
