"""
GMCS-format writer — mimics the Green Mountain Community Solar quarterly
NEPOOL-GIS-ready workbook used by Bruce.

Layout (one sheet per producing array — an array with zero generation across
every month of the reporting window gets NO sheet, so NEPOOL-GIS uploads only
carry arrays with reportable generation):
  A1:C1 (merged) — "<Array Name> (<optional ID>)"
  Row 5 — header: Quarter | Generation (MWh) | Reporting Amount | RECs†
  Rows 7-29 — 6 quarter blocks × 3 month rows each
    Each quarter:
      - first row holds quarter label (e.g. "Q3 2024") in col A
      - cols B,C = generation in MWh (kWh / 1000, 3 decimals)
      - col D = whole RECs (floor of MWh)
      - one blank row between quarters
  Row 31 — footnote: "† NEPOOL-GIS will award 1 REC per whole MWh of generation."

Default window: most recent 6 complete quarters ending at the quarter the
NEPOOL-GIS agent is currently minting — TWO quarters before the in-progress
quarter (NEPOOL-GIS issues RECs ~2 quarters after generation, so Q1 gen is
uploaded the following July). This mirrors what the REC agent submits, rather
than the most recently completed quarter. See default_reporting_reference_date.
An explicit reference_date / chosen quarter overrides this default.
"""
from __future__ import annotations
import pathlib
from collections import defaultdict
from datetime import datetime, date, timedelta
from typing import Optional

from sqlalchemy import func, select
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..bill_attribution import distribute_kwh_by_calendar_day
from ..db import SessionLocal, DATA_DIR
from ..generation_sources import EXTENSION_SOURCES, VENDOR_TELEMETRY_SOURCES
from ..models import Tenant, Client, UtilityAccount, Array, Bill, DailyGeneration
from ..report_arrays import not_vendor_only, utility_backed_array_ids


REPORTS_DIR = DATA_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True, parents=True)


# Footnote text — VERBATIM from Bruce's GMCS.xlsx. Single source of truth so
# the demo writer (api/writers/demo_writer.py) can reuse the exact wording.
# Never paraphrase — see CLAUDE.md "GMCS writer format rules".
FOOTNOTE_TEXT = (
    " † NEPOOL-GIS will award 1 REC for every MWH reported.  "
    "Additionally, NEPOOL-GIS will keep track of the decimal "
    "MWHs and award an additional REC when the total exceeds 1 MWH."
)


# ── helpers ──────────────────────────────────────────────────────────
def _quarter_of(month: int) -> int:
    return (month - 1) // 3 + 1


def _month_in_quarter(month: int) -> int:
    """Returns 1, 2, or 3 — position of month within its quarter."""
    return ((month - 1) % 3) + 1


def _quarter_months(year: int, q: int) -> list[tuple[int, int]]:
    start = (q - 1) * 3 + 1
    return [(year, start + i) for i in range(3)]


def _rolling_quarters(ref: date, count: int = 6) -> list[tuple[int, int]]:
    """Return list of (year, quarter) for the most recent `count` complete
    quarters relative to `ref`. Most-recent quarter LAST in the list
    (chronological order for spreadsheet display)."""
    # Determine current quarter then step back one to skip the in-progress one.
    cy, cq = ref.year, _quarter_of(ref.month)
    # last complete quarter:
    y, q = cy, cq - 1
    if q == 0:
        y, q = cy - 1, 4
    out: list[tuple[int, int]] = []
    for _ in range(count):
        out.append((y, q))
        q -= 1
        if q == 0:
            y, q = y - 1, 4
    out.reverse()  # oldest first
    return out


def default_reporting_reference_date(today: date) -> date:
    """Reference date whose 'last complete quarter' is the generation quarter the
    NEPOOL-GIS agent is currently minting — so our automated reports mirror what
    the REC agent (e.g. Crown) actually submits.

    NEPOOL-GIS issues RECs roughly TWO quarters after the generation quarter ends:
    Q1 (Jan–Mar) generation is uploaded the following July, Q2 in October, and so
    on. So the quarter a REC report must mirror is two quarters before the
    in-progress quarter — i.e. ONE quarter before the last *complete* quarter, not
    the last complete quarter itself. On 2026-07-06 (Q3 in progress) this resolves
    to Q1 2026, exactly what Crown uploads in July.

    Returns the first day of the quarter AFTER the minting quarter, so passing it
    through ``_rolling_quarters`` terminates the window on the minting quarter.

    Only used as the DEFAULT when no explicit reporting quarter is requested; an
    explicit ``reference_date`` (e.g. from a chosen quarter) bypasses this entirely.
    """
    cy, cq = today.year, _quarter_of(today.month)
    # minting quarter = two quarters back from the in-progress quarter
    y, q = cy, cq
    for _ in range(2):
        q -= 1
        if q == 0:
            y, q = y - 1, 4
    # first day of the quarter AFTER (y, q) makes (y, q) the last complete quarter
    nxt_month = q * 3 + 1
    if nxt_month > 12:
        return date(y + 1, 1, 1)
    return date(y, nxt_month, 1)


def _sheet_name_for_array(name: str, used: set[str]) -> str:
    """Excel sheet names: max 31 chars, no /\\?*[]"""
    bad = '/\\?*[]:'
    clean = "".join(c for c in (name or "Array") if c not in bad).strip()
    if not clean:
        clean = "Array"
    base = clean[:31]
    final = base
    i = 2
    while final in used:
        suffix = f" {i}"
        final = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(final)
    return final


# Inverter / extension telemetry — LAST resort for a month with no utility data.
# Precedence in build_workbook: utility daily (incl. GMP 15-min) > bill > vendor.
_VENDOR_FALLBACK_SOURCES = VENDOR_TELEMETRY_SOURCES | EXTENSION_SOURCES


def _daily_generation_by_month(
    db, array_id: int, start: date, end: date
) -> dict[tuple[int, int], float]:
    """Return {(year, month): kwh_sum} of UTILITY-side daily generation for
    an array in [start, end].

    Base source is DailyGeneration that is NOT inverter/extension telemetry and
    NOT the redundant ``bill_prorate`` smear (real bill kWh is applied
    separately via per_group). Real GMP 15-minute meter data
    (``GmpDailyGeneration`` via the reports read-contract) is the utility
    revenue meter Crown REC reports from to NEPOOL-GIS, so it WINS for any
    month it covers nearly fully — that is what makes our monthly numbers
    reconcile with Crown line-by-line instead of the flattened bill-proration
    estimate (bill proration keeps the quarter total but flattens the
    within-quarter monthly peaks, shifting per-month REC floors).

    Vendor telemetry (Locus / SolarEdge / …) is intentionally NOT here — see
    ``_vendor_generation_by_month`` + ``_merge_report_months`` for the
    fallback-only path (utility empty months only).

    Coverage guard: the GMP total only overrides a month when it has near-full
    daily coverage (>= days_in_month - 2), so a partial or stale-meter fragment
    (e.g. an account whose interval feed stopped years ago) can never undercount a
    month by silently replacing a full bill estimate with a few days of intervals.
    """
    import calendar as _cal
    from ..reports import gmp_daily_read

    # NEPOOL RECs settle on the UTILITY's measured generation (GMP bills +
    # GMP interval meter) first. Exclusions here (Ford 2026-07-16 — London_SE
    # was reporting SolarEdge, not GMP):
    #   • VENDOR/EXTENSION telemetry — wrong meter when utility data exists;
    #     re-introduced only as last-resort fallback via _vendor_generation_by_month.
    #   • bill_prorate — redundant smear of the same bill applied via per_group.
    _exclude = _VENDOR_FALLBACK_SOURCES | {"bill_prorate"}
    rows = db.execute(
        select(DailyGeneration.day, DailyGeneration.kwh)
        .where(
            DailyGeneration.array_id == array_id,
            DailyGeneration.day >= start,
            DailyGeneration.day <= end,
            func.lower(DailyGeneration.source).notin_(_exclude),
        )
    ).all()
    buckets: dict[tuple[int, int], float] = {}
    for day, kwh in rows:
        key = (day.year, day.month)
        buckets[key] = buckets.get(key, 0.0) + float(kwh)

    # Overlay authoritative GMP meter data where a month is (nearly) fully covered.
    for r in gmp_daily_read.get_monthly_totals(array_id, start=start, end=end, db=db):
        days_in_month = _cal.monthrange(r["year"], r["month"])[1]
        if r["days"] >= days_in_month - 2:
            buckets[(r["year"], r["month"])] = r["kwh"]

    return buckets


def _vendor_generation_by_month(
    db, array_id: int, start: date, end: date
) -> dict[tuple[int, int], float]:
    """Monthly kWh from inverter/extension telemetry only (Locus, SolarEdge, …).

    Used solely as a FALLBACK when a month has no utility daily and no bill
    kWh — never allowed to displace GMP interval or bill numbers.
    """
    if not _VENDOR_FALLBACK_SOURCES:
        return {}
    rows = db.execute(
        select(DailyGeneration.day, DailyGeneration.kwh)
        .where(
            DailyGeneration.array_id == array_id,
            DailyGeneration.day >= start,
            DailyGeneration.day <= end,
            func.lower(DailyGeneration.source).in_(_VENDOR_FALLBACK_SOURCES),
        )
    ).all()
    buckets: dict[tuple[int, int], float] = {}
    for day, kwh in rows:
        key = (day.year, day.month)
        buckets[key] = buckets.get(key, 0.0) + float(kwh)
    return buckets


def _merge_report_months(
    *,
    vendor: dict[tuple[int, int], float],
    bill: dict[tuple[int, int], float],
    utility: dict[tuple[int, int], float],
) -> dict[tuple[int, int], float]:
    """Precedence: utility daily (incl. GMP 15-min) > bill > vendor telemetry."""
    return {**vendor, **bill, **utility}


def _client_is_inverter_only(db, client_id: int) -> bool:
    """True when the client has no utility-backed arrays.

    Inverter-only fleets (e.g. Locus-only, GMP not connected yet) may use
    vendor telemetry as the sole generation source for reports. Mixed fleets
    keep excluding vendor-twin arrays so a Fronius twin cannot double-report
    alongside a GMP-backed sibling.
    """
    return not utility_backed_array_ids(
        db, client_id=client_id, include_excluded=False, include_deleted=False
    )


def _reportable_arrays_query(client_id: int, *, inverter_only_client: bool):
    """Array filter for generation reports.

    Utility-backed clients: ``not_vendor_only()`` (drops vendor twins).
    Inverter-only clients: every non-deleted, non-excluded array — vendor
    fallback supplies the kWh.
    """
    q = [
        Array.client_id == client_id,
        Array.excluded.is_(False),
        Array.deleted_at.is_(None),
    ]
    if not inverter_only_client:
        q.append(not_vendor_only())
    return q


# ── main builder ─────────────────────────────────────────────────────
def report_has_data(client_id: int, *, quarters: int = 6,
                    reference_date: Optional[date] = None) -> bool:
    """True iff the workbook for `client_id` would render at least one non-zero
    generation month in its reporting window.

    Mirrors build_workbook's data sourcing EXACTLY: same rolling-quarter window,
    same sources (utility daily + bill kWh + vendor fallback), so this never
    disagrees with what the rendered cells show. Used by the delivery layer to
    skip auto-sending a blank workbook (a client with arrays but no bills/daily
    data, or an empty onboarding stub) — exactly what an operator does by hand
    when they only send reports that have real numbers.

    Read-only. Cheap relative to building the whole workbook.
    """
    ref = reference_date if reference_date is not None \
        else default_reporting_reference_date(date.today())
    qlist = _rolling_quarters(ref, count=quarters)
    qmonths = set()
    for (qy, qq) in qlist:
        for (my, mm) in _quarter_months(qy, qq):
            qmonths.add((my, mm))
    start_year, start_q = qlist[0]
    report_start = date(start_year, (start_q - 1) * 3 + 1, 1)
    end_year, end_q = qlist[-1]
    end_month = end_q * 3
    report_end = (date(end_year, 12, 31) if end_month == 12
                  else date(end_year, end_month + 1, 1) - timedelta(days=1))

    with SessionLocal() as db:
        client = db.get(Client, client_id)
        if client is None:
            return False
        inv_only = _client_is_inverter_only(db, client.id)
        arrays = db.execute(
            select(Array).where(*_reportable_arrays_query(client.id, inverter_only_client=inv_only))
        ).scalars().all()
        array_ids = [a.id for a in arrays]
        if not array_ids:
            return False

        # 1) Utility daily + vendor fallback — any non-zero month in window.
        for arr_id in array_ids:
            util = _daily_generation_by_month(db, arr_id, report_start, report_end)
            vend = _vendor_generation_by_month(db, arr_id, report_start, report_end)
            if any(v > 0 for v in util.values()) or any(v > 0 for v in vend.values()):
                return True

        # 2) Bill kWh — attribute each bill across calendar days, keep only the
        #    months inside the report window (matches the renderer's per_group).
        accounts = db.execute(
            select(UtilityAccount).where(UtilityAccount.array_id.in_(array_ids))
        ).scalars().all()
        account_ids = [a.id for a in accounts]
        if account_ids:
            bills = db.execute(
                select(Bill).where(Bill.account_id.in_(account_ids))
            ).scalars().all()
            for b in bills:
                for (yy, mm), kwh in distribute_kwh_by_calendar_day(b).items():
                    if (yy, mm) in qmonths and kwh > 0:
                        return True
    return False


def reported_array_ids(client_id: int, *, quarters: int = 6,
                       reference_date: Optional[date] = None) -> list[int]:
    """The array ids that would ACTUALLY RENDER in this client's workbook.

    The per-array counterpart of `report_has_data`: same window, same sources
    (utility daily + bill + vendor fallback), same filters the builder applies
    — the client's arrays minus `excluded` (operator force-hide) minus
    soft-deleted (and vendor twins on mixed fleets), then minus the
    non-producing ones (an array whose every month in the window is zero gets
    no sheet, so it is not "reported").

    This is the BILLING unit for generation reports ($15 per array per quarter
    — THE FOLD, Ford Jul 2026): we bill exactly the arrays the operator
    actually reports, never an array that renders no sheet. Read-only.
    """
    ref = reference_date if reference_date is not None \
        else default_reporting_reference_date(date.today())
    qlist = _rolling_quarters(ref, count=quarters)
    qmonths = set()
    for (qy, qq) in qlist:
        for (my, mm) in _quarter_months(qy, qq):
            qmonths.add((my, mm))
    start_year, start_q = qlist[0]
    report_start = date(start_year, (start_q - 1) * 3 + 1, 1)
    end_year, end_q = qlist[-1]
    end_month = end_q * 3
    report_end = (date(end_year, 12, 31) if end_month == 12
                  else date(end_year, end_month + 1, 1) - timedelta(days=1))

    out: list[int] = []
    with SessionLocal() as db:
        client = db.get(Client, client_id)
        if client is None:
            return out
        inv_only = _client_is_inverter_only(db, client.id)
        arrays = db.execute(
            select(Array).where(*_reportable_arrays_query(client.id, inverter_only_client=inv_only))
        ).scalars().all()
        for arr in arrays:
            # 1) Utility daily + vendor fallback — any non-zero month in window.
            util = _daily_generation_by_month(db, arr.id, report_start, report_end)
            vend = _vendor_generation_by_month(db, arr.id, report_start, report_end)
            if any(v > 0 for v in util.values()) or any(v > 0 for v in vend.values()):
                out.append(arr.id)
                continue
            # 2) Bill kWh — attributed across calendar days, months in-window.
            accounts = db.execute(
                select(UtilityAccount).where(UtilityAccount.array_id == arr.id)
            ).scalars().all()
            account_ids = [a.id for a in accounts]
            if not account_ids:
                continue
            bills = db.execute(
                select(Bill).where(Bill.account_id.in_(account_ids))
            ).scalars().all()
            for b in bills:
                hit = False
                for (yy, mm), kwh in distribute_kwh_by_calendar_day(b).items():
                    if (yy, mm) in qmonths and kwh > 0:
                        out.append(arr.id)
                        hit = True
                        break
                if hit:
                    break
    return out


# ── main builder ─────────────────────────────────────────────────────
def build_workbook(tenant_id: Optional[str] = None,
                   year: Optional[int] = None,
                   out_path: Optional[pathlib.Path] = None,
                   *, quarters: int = 6,
                   reference_date: Optional[date] = None,
                   client_id: Optional[int] = None) -> pathlib.Path:
    """Generate the GMCS-format workbook for ONE client. Returns saved path.

    Calling conventions:
      build_workbook(client_id=N, ...)         -- preferred, post-Phase-1
      build_workbook(tenant_id="ten_…", ...)   -- legacy fallback: picks the
                                                  first Client under that
                                                  tenant; preserves callers
                                                  that haven't migrated yet.

    Exactly one of `client_id` or `tenant_id` must be provided.

    `year` is accepted for API back-compat with the legacy month-grid writer
    but the GMCS format is rolling-quarter based; it's only used to default
    the output filename.

    `quarters` is the number of trailing complete quarters to include
    (default 6 = 18 months, matching Bruce's GMCS workbook).
    """
    if client_id is None and tenant_id is None:
        raise ValueError("build_workbook requires client_id or tenant_id")

    ref = reference_date if reference_date is not None \
        else default_reporting_reference_date(date.today())
    qlist = _rolling_quarters(ref, count=quarters)
    last_y, last_q = qlist[-1]
    if year is None:
        year = last_y

    with SessionLocal() as db:
        # Resolve client_id from tenant_id when legacy mode used
        if client_id is None:
            client = db.execute(
                select(Client).where(Client.tenant_id == tenant_id)
                              .order_by(Client.id.asc())
            ).scalars().first()
            if client is None:
                raise ValueError(
                    f"Tenant {tenant_id} has no Client rows; run migrations.")
            client_id = client.id
        else:
            client = db.get(Client, client_id)
            if client is None:
                raise ValueError(f"unknown client {client_id}")

        tenant = db.get(Tenant, client.tenant_id)
        if not tenant:
            raise ValueError(f"unknown tenant {client.tenant_id}")

        if out_path is None:
            out_path = (REPORTS_DIR / tenant.id / f"client-{client.id}"
                        / f"{last_y}-Q{last_q}-GMCS-report.xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # Arrays scoped to THIS client only; skip excluded (below-REC-threshold)
        # AND soft-deleted arrays. Without the deleted_at filter, arrays the
        # operator removed still got a sheet in the report (Ford 2026-07-16 —
        # Bruce deleted arrays that kept showing up in the GMCS workbook).
        # Vendor twins stay out on mixed fleets; inverter-only clients keep
        # their arrays so Locus (etc.) can fall back into the report.
        inv_only = _client_is_inverter_only(db, client.id)
        arrays = db.execute(
            select(Array).where(
                *_reportable_arrays_query(client.id, inverter_only_client=inv_only)
            )
        ).scalars().all()
        arrays_by_id = {a.id: a for a in arrays}
        array_ids = list(arrays_by_id.keys())

        # Accounts under those arrays
        if array_ids:
            accounts = db.execute(
                select(UtilityAccount).where(UtilityAccount.array_id.in_(array_ids))
            ).scalars().all()
        else:
            accounts = []

        def group_for(acc: UtilityAccount) -> tuple[str, Optional[Array]]:
            if acc.array_id and acc.array_id in arrays_by_id:
                a = arrays_by_id[acc.array_id]
                return a.name, a
            return (acc.nickname or acc.account_number), None

        group_of: dict[int, str] = {}
        group_meta: dict[str, Optional[Array]] = {}
        for acc in accounts:
            name, a = group_for(acc)
            group_of[acc.id] = name
            group_meta.setdefault(name, a)
        # Arrays with no utility account (inverter-only fleets) still need a sheet.
        for arr in arrays:
            group_meta.setdefault(arr.name, arr)

        # Pull bills for those accounts only
        if accounts:
            account_ids = [a.id for a in accounts]
            bills = db.execute(
                select(Bill).where(Bill.account_id.in_(account_ids))
            ).scalars().all()
        else:
            bills = []

        # Per-group kWh by (year, month) — pro-rated by calendar day.
        per_group: dict[str, dict[tuple[int, int], float]] = defaultdict(
            lambda: defaultdict(float))
        for b in bills:
            grp = group_of.get(b.account_id)
            if not grp:
                continue
            for (year, month), kwh in distribute_kwh_by_calendar_day(b).items():
                per_group[grp][(year, month)] += kwh

        groups = sorted(group_meta.keys())

        # Query DailyGeneration for each array in the report window.
        # Results are stored outside the db session for use in workbook rendering.
        start_year, start_q = qlist[0]
        report_start = date(start_year, (start_q - 1) * 3 + 1, 1)
        end_year, end_q = qlist[-1]
        end_month = end_q * 3
        if end_month == 12:
            report_end = date(end_year, 12, 31)
        else:
            report_end = date(end_year, end_month + 1, 1) - timedelta(days=1)

        # Utility daily (GMP / utility meter) and vendor fallback (Locus/…)
        # kept separate so the merge can enforce utility > bill > vendor.
        daily_gen_by_group: dict[str, dict[tuple[int, int], float]] = {}
        vendor_gen_by_group: dict[str, dict[tuple[int, int], float]] = {}
        for grp_name, arr in group_meta.items():
            if arr is None:
                continue
            util = _daily_generation_by_month(db, arr.id, report_start, report_end)
            if util:
                daily_gen_by_group[grp_name] = util
            vend = _vendor_generation_by_month(db, arr.id, report_start, report_end)
            if vend:
                vendor_gen_by_group[grp_name] = vend

    # Non-producing arrays get no sheet: a group whose every month in the
    # reporting window is zero (utility daily + bills + vendor fallback, merged
    # with the same precedence the renderer uses) would render as an all-blank
    # sheet, and NEPOOL-GIS uploads must only carry arrays with reportable
    # generation. Operators can still force-hide an array via Array.excluded;
    # this filter is the automatic counterpart.
    window_months = {m for (qy, qq) in qlist for m in _quarter_months(qy, qq)}
    groups = [
        grp for grp in groups
        if any(
            _merge_report_months(
                vendor=vendor_gen_by_group.get(grp, {}),
                bill=per_group.get(grp, {}),
                utility=daily_gen_by_group.get(grp, {}),
            ).get(m, 0.0) > 0
            for m in window_months
        )
    ]

    # ── Build workbook ──────────────────────────────────────────────
    wb = Workbook()
    # remove default sheet at end
    default_sheet = wb.active

    TITLE_FONT = Font(bold=True, size=14, color="1F4E2A")
    HDR_FONT = Font(bold=True, size=14, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="064E3B")
    QUARTER_FONT = Font(bold=True, size=11, color="1F4E2A")
    FOOTNOTE_FONT = Font(italic=True, size=9, color="666666")
    BORDER = Border(*[Side(style="thin", color="E8E2D9")] * 4)
    GOLD_SIDE = Side(style="medium", color="E6B470")

    used_names: set[str] = set()

    if not groups:
        # produce an empty stub sheet so the file is never blank
        groups = ["(no data)"]

    for grp_idx, grp in enumerate(groups):
        sheet_title = _sheet_name_for_array(grp, used_names)
        if grp_idx == 0:
            sh = default_sheet
            sh.title = sheet_title
        else:
            sh = wb.create_sheet(title=sheet_title)

        # ── Title (A1 merged A1:C1) ──
        arr = group_meta.get(grp)
        nepool_id = getattr(arr, "nepool_gis_id", None) if arr else None
        title = f"{grp} ({nepool_id})" if nepool_id else grp
        sh["A1"] = title
        sh["A1"].font = TITLE_FONT
        sh["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells("A1:C1")

        # Row 5 header
        for col, label in enumerate(
            ["Quarter", "Generation (MWh)", "Reporting Amount", "RECs†"],
            start=1,
        ):
            c = sh.cell(5, col, label)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

        # Gold accent: single warm hairline under the header, carried into the
        # gap row (row 6) between the header and the first quarter block.
        for col in range(1, 5):
            sh.cell(6, col).border = Border(bottom=GOLD_SIDE)

        # Data: 6 quarter blocks × 3 month rows
        # First block starts at row 7; gap row between blocks (row 6, 10, 14, ...)
        row = 7
        # Precedence per (year, month): utility daily (GMP 15-min / meter) >
        # bill kWh > vendor telemetry (Locus/SolarEdge/…) as last-resort fill.
        gen_by_month = _merge_report_months(
            vendor=vendor_gen_by_group.get(grp, {}),
            bill=per_group.get(grp, {}),
            utility=daily_gen_by_group.get(grp, {}),
        )
        for (qy, qq) in qlist:
            for i, (my, mm) in enumerate(_quarter_months(qy, qq)):
                if i == 0:
                    qc = sh.cell(row, 1, f"Q{qq} {qy}")
                    qc.font = QUARTER_FONT
                    qc.alignment = Alignment(horizontal="left", vertical="center")
                kwh = gen_by_month.get((my, mm), 0.0)
                mwh = round(kwh / 1000.0, 3)
                recs = int(mwh)  # floor of MWh
                gc = sh.cell(row, 2, mwh if kwh else None)
                gc.number_format = "General"
                gc.alignment = Alignment(horizontal="right")
                rc = sh.cell(row, 3, mwh if kwh else None)
                rc.number_format = "General"
                rc.alignment = Alignment(horizontal="right")
                ec = sh.cell(row, 4, recs if kwh else None)
                ec.number_format = "General"
                ec.alignment = Alignment(horizontal="right")
                row += 1
            row += 1  # gap row between quarter blocks

        # Footnote — verbatim text from Bruce's GMCS.xlsx
        foot_row = row + (31 - row) if row <= 31 else row
        fc = sh.cell(foot_row, 1, FOOTNOTE_TEXT)
        fc.font = FOOTNOTE_FONT
        fc.alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells(start_row=foot_row, start_column=1,
                       end_row=foot_row, end_column=4)

        # Column widths — uniform 24 across all four columns for max readability.
        sh.column_dimensions["A"].width = 24.0
        sh.column_dimensions["B"].width = 24.0
        sh.column_dimensions["C"].width = 24.0
        sh.column_dimensions["D"].width = 24.0

    wb.save(out_path)
    return out_path


def build_directory_workbook(
    tenant_id: str,
    *,
    client_ids: Optional[list[int]] = None,
    year: Optional[int] = None,
    out_path: Optional[pathlib.Path] = None,
    quarters: int = 6,
    reference_date: Optional[date] = None,
) -> pathlib.Path:
    """NEPOOL-GIS upload directory: one sheet per producing array across ALL
    (or selected) clients of a tenant — same GMCS layout as the per-client
    reports. Sheet titles are ``Client — Array`` so collisions across clients
    stay unique. Used so the operator can bulk-upload the whole book to
    NEPOOL-GIS after client reports go out.
    """
    ref = reference_date if reference_date is not None \
        else default_reporting_reference_date(date.today())
    qlist = _rolling_quarters(ref, count=quarters)
    last_y, last_q = qlist[-1]
    if year is None:
        year = last_y

    with SessionLocal() as db:
        tenant = db.get(Tenant, tenant_id)
        if not tenant:
            raise ValueError(f"unknown tenant {tenant_id}")

        cq = select(Client).where(
            Client.tenant_id == tenant_id,
            Client.active == True,  # noqa: E712
            Client.deleted_at.is_(None),
        ).order_by(Client.name.asc())
        clients = db.execute(cq).scalars().all()
        if client_ids is not None:
            want = set(client_ids)
            clients = [c for c in clients if c.id in want]
        clients_by_id = {c.id: c for c in clients}
        if not clients:
            raise ValueError(f"Tenant {tenant_id} has no active clients for directory")

        if out_path is None:
            out_path = (REPORTS_DIR / tenant.id / "directory"
                        / f"{last_y}-Q{last_q}-NEPOOL-directory.xlsx")
        out_path.parent.mkdir(parents=True, exist_ok=True)

        # Per-client: utility fleets drop vendor twins; inverter-only clients keep them.
        from ..report_arrays import is_vendor_only_array
        inv_only_by_client = {
            cid: _client_is_inverter_only(db, cid) for cid in clients_by_id
        }
        arrays_all = db.execute(
            select(Array).where(
                Array.tenant_id == tenant_id,
                Array.client_id.in_(list(clients_by_id.keys())),
                Array.excluded.is_(False),
                Array.deleted_at.is_(None),
            ).order_by(Array.name.asc())
        ).scalars().all()
        arrays = []
        for a in arrays_all:
            if a.client_id is None:
                continue
            if inv_only_by_client.get(a.client_id) or not is_vendor_only_array(db, a.id):
                arrays.append(a)

        # Unique display group per array: "Client — Array"
        arrays_by_id: dict[int, Array] = {}
        group_meta: dict[str, Optional[Array]] = {}
        array_to_group: dict[int, str] = {}
        for a in arrays:
            c = clients_by_id.get(a.client_id) if a.client_id else None
            cname = (c.name if c else "Unassigned").strip() or "Client"
            aname = (a.name or f"Array {a.id}").strip()
            base = f"{cname} — {aname}"
            grp = base
            n = 2
            while grp in group_meta:
                grp = f"{base} ({n})"
                n += 1
            arrays_by_id[a.id] = a
            group_meta[grp] = a
            array_to_group[a.id] = grp

        array_ids = list(arrays_by_id.keys())
        if array_ids:
            accounts = db.execute(
                select(UtilityAccount).where(
                    UtilityAccount.array_id.in_(array_ids),
                    UtilityAccount.deleted_at.is_(None),
                )
            ).scalars().all()
        else:
            accounts = []

        group_of: dict[int, str] = {}
        for acc in accounts:
            if acc.array_id and acc.array_id in array_to_group:
                group_of[acc.id] = array_to_group[acc.array_id]

        if accounts:
            account_ids = [a.id for a in accounts]
            bills = db.execute(
                select(Bill).where(Bill.account_id.in_(account_ids))
            ).scalars().all()
        else:
            bills = []

        per_group: dict[str, dict[tuple[int, int], float]] = defaultdict(
            lambda: defaultdict(float))
        for b in bills:
            grp = group_of.get(b.account_id)
            if not grp:
                continue
            for (yy, mm), kwh in distribute_kwh_by_calendar_day(b).items():
                per_group[grp][(yy, mm)] += kwh

        groups = sorted(set(group_of.values()) | set(group_meta.keys()))

        start_year, start_q = qlist[0]
        report_start = date(start_year, (start_q - 1) * 3 + 1, 1)
        end_year, end_q = qlist[-1]
        end_month = end_q * 3
        if end_month == 12:
            report_end = date(end_year, 12, 31)
        else:
            report_end = date(end_year, end_month + 1, 1) - timedelta(days=1)

        daily_gen_by_group: dict[str, dict[tuple[int, int], float]] = {}
        vendor_gen_by_group: dict[str, dict[tuple[int, int], float]] = {}
        for grp_name, arr in group_meta.items():
            if arr is None:
                continue
            util = _daily_generation_by_month(db, arr.id, report_start, report_end)
            if util:
                daily_gen_by_group[grp_name] = util
            vend = _vendor_generation_by_month(db, arr.id, report_start, report_end)
            if vend:
                vendor_gen_by_group[grp_name] = vend

    window_months = {m for (qy, qq) in qlist for m in _quarter_months(qy, qq)}
    groups = [
        grp for grp in groups
        if any(
            _merge_report_months(
                vendor=vendor_gen_by_group.get(grp, {}),
                bill=per_group.get(grp, {}),
                utility=daily_gen_by_group.get(grp, {}),
            ).get(m, 0.0) > 0
            for m in window_months
        )
    ]

    wb = Workbook()
    default_sheet = wb.active

    TITLE_FONT = Font(bold=True, size=14, color="1F4E2A")
    HDR_FONT = Font(bold=True, size=14, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="064E3B")
    QUARTER_FONT = Font(bold=True, size=11, color="1F4E2A")
    FOOTNOTE_FONT = Font(italic=True, size=9, color="666666")
    BORDER = Border(*[Side(style="thin", color="E8E2D9")] * 4)
    GOLD_SIDE = Side(style="medium", color="E6B470")

    used_names: set[str] = set()
    if not groups:
        groups = ["(no data)"]

    for grp_idx, grp in enumerate(groups):
        sheet_title = _sheet_name_for_array(grp, used_names)
        if grp_idx == 0:
            sh = default_sheet
            sh.title = sheet_title
        else:
            sh = wb.create_sheet(title=sheet_title)

        arr = group_meta.get(grp)
        nepool_id = getattr(arr, "nepool_gis_id", None) if arr else None
        # A1 title keeps Client — Array (+ NEPOOL id) so ownership is clear
        # even after Excel truncates the sheet tab to 31 chars.
        title = f"{grp} ({nepool_id})" if nepool_id else grp
        sh["A1"] = title
        sh["A1"].font = TITLE_FONT
        sh["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells("A1:C1")

        for col, label in enumerate(
            ["Quarter", "Generation (MWh)", "Reporting Amount", "RECs†"],
            start=1,
        ):
            c = sh.cell(5, col, label)
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = BORDER

        for col in range(1, 5):
            sh.cell(6, col).border = Border(bottom=GOLD_SIDE)

        row = 7
        gen_by_month = _merge_report_months(
            vendor=vendor_gen_by_group.get(grp, {}),
            bill=per_group.get(grp, {}),
            utility=daily_gen_by_group.get(grp, {}),
        )
        for (qy, qq) in qlist:
            for i, (my, mm) in enumerate(_quarter_months(qy, qq)):
                if i == 0:
                    qc = sh.cell(row, 1, f"Q{qq} {qy}")
                    qc.font = QUARTER_FONT
                    qc.alignment = Alignment(horizontal="left", vertical="center")
                kwh = gen_by_month.get((my, mm), 0.0)
                mwh = round(kwh / 1000.0, 3)
                recs = int(mwh)
                gc = sh.cell(row, 2, mwh if kwh else None)
                gc.number_format = "General"
                gc.alignment = Alignment(horizontal="right")
                rc = sh.cell(row, 3, mwh if kwh else None)
                rc.number_format = "General"
                rc.alignment = Alignment(horizontal="right")
                ec = sh.cell(row, 4, recs if kwh else None)
                ec.number_format = "General"
                ec.alignment = Alignment(horizontal="right")
                row += 1
            row += 1

        foot_row = row + (31 - row) if row <= 31 else row
        fc = sh.cell(foot_row, 1, FOOTNOTE_TEXT)
        fc.font = FOOTNOTE_FONT
        fc.alignment = Alignment(horizontal="left", vertical="center")
        sh.merge_cells(start_row=foot_row, start_column=1,
                       end_row=foot_row, end_column=4)

        sh.column_dimensions["A"].width = 24.0
        sh.column_dimensions["B"].width = 24.0
        sh.column_dimensions["C"].width = 24.0
        sh.column_dimensions["D"].width = 24.0

    wb.save(out_path)
    return out_path
