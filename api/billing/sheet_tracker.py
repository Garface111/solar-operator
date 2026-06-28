"""Bring-your-own generation-spreadsheet auto-updater.

Ford's ask: an offtaker (BillingReportSubscription) operator should be able to
upload their *existing* generation-tracking spreadsheet — in whatever columns
they already use — and have Array Operator (a) figure out what each column means
("our magic"), then (b) keep appending a new row every month as a fresh GMP bill
lands, preserving their original layout. A "Download latest spreadsheet" button
on the invoice-generator page streams the current, kept-current file.

Design constraints honored here:
  * NO LLM at runtime (no ANTHROPIC_API_KEY on the box). Detection is a robust,
    deterministic, case-insensitive header-keyword heuristic. The detected
    mapping is surfaced to the operator so they can correct it.
  * Append is IDEMPOTENT — we never double-append the same billing period (we
    check the last data row's period before writing).
  * The user's existing rows/format/styling are preserved (openpyxl loads the
    real workbook and writes new cells into the SAME columns / sheet).
  * Whole feature is gated behind SPREADSHEET_TRACKER_ENABLED so a half-built
    state can never disturb the live invoice page.
  * New module → minimal collision surface with worker.py / routes.py.

The detected structure + the figures we append are sourced from the same
production invoice computation the rest of the page uses
(delivery.build_match(sub).computed_invoice), so the spreadsheet stays
consistent with the invoice the offtaker is billed from.
"""
from __future__ import annotations

import io
import logging
import os
import re
from datetime import date, datetime
from typing import Any, Optional

logger = logging.getLogger(__name__)

# ─── feature flag ────────────────────────────────────────────────────────────

def tracker_enabled() -> bool:
    """Master switch. OFF unless SPREADSHEET_TRACKER_ENABLED is truthy."""
    return os.getenv("SPREADSHEET_TRACKER_ENABLED", "").lower() in (
        "1", "true", "yes", "on")


# ─── heuristic column detection ("our magic") ────────────────────────────────
#
# Each logical field maps to a list of lowercase substrings; a header cell is
# assigned to the field whose keyword it contains. Order matters: more-specific
# fields are tested first so e.g. "consumption kwh" lands on consumption, not
# generation, and "amount due" lands on amount. Scoring prefers a keyword that
# appears earlier in the cell + a longer keyword match (more specific wins).

# (field, [(keyword, weight), ...]) — weight scales the base score so a
# unit-bearing or unambiguous token (e.g. "$/kwh" → rate) can out-rank a vaguer
# one (e.g. "credit" → amount) that happens to share the cell.
_FIELD_KEYWORDS: list[tuple[str, list[tuple[str, float]]]] = [
    # period / date column
    ("period", [("billing month", 3.0), ("billing date", 3.0),
                ("bill date", 3.0), ("invoice date", 3.0), ("statement date", 3.0),
                ("period", 2.5), ("month", 2.0), ("statement", 1.5),
                ("cycle", 1.5), ("date", 1.2)]),
    # consumption / usage (tested before generation so "usage kwh" isn't grabbed
    # by the bare "kwh" generation keyword)
    ("consumption", [("consumption", 3.0), ("consumed", 3.0), ("kwh used", 3.0),
                     ("usage", 2.5), ("used", 2.0), ("delivered", 1.5)]),
    # generation / production kWh
    ("generation", [("kwh generated", 3.5), ("solar kwh", 3.5),
                    ("generation", 3.0), ("generated", 3.0), ("production", 3.0),
                    ("produced", 3.0), ("solar", 1.5), ("output", 2.0),
                    ("kwh", 1.5), ("kw h", 1.5), ("energy", 1.2)]),
    # credit rate ($/kWh) — unit-bearing tokens are unambiguous and weighted high.
    # A COMBINED rate (tariff+adder) is the effective per-kWh credit, so it out-ranks
    # the bare "tariff" column (which omits the adder) when a sheet splits them.
    ("rate", [("tariff+adder", 5.0), ("rate+adder", 5.0), ("tariff + adder", 5.0),
              ("$/kwh", 4.0), ("per kwh", 4.0), ("/kwh", 4.0),
              ("credit rate", 4.0), ("net rate", 3.5), ("tariff", 3.0),
              ("rate", 2.0), ("price", 1.5)]),
    # amount / total $ — "bill" is the offtaker's actual paid amount on many sheets
    ("amount", [("amount due", 3.5), ("total due", 3.5), ("invoice total", 3.5),
                ("bill amount", 3.0), ("amount", 2.5), ("total credit", 2.5),
                ("bill", 2.2), ("total", 1.8), ("balance", 2.0), ("owed", 2.5),
                ("due", 1.5), ("credit", 1.2)]),
]

# A kWh column that is a CUMULATIVE / running-total (e.g. "Cumm KwH", "YTD kWh") is
# NOT the monthly generation we append to — demote it so a real per-period column wins.
_CUM_WORDS = ("cumm", "cumul", "ytd", "to date", "running", "lifetime", "year to date")

# Common, non-distinctive words to drop from an offtaker's name before using the rest
# as detection hints (so we match on "Fairlee", not "Town"/"of"/"Solar").
_NAME_STOP = {"town", "city", "village", "the", "and", "llc", "inc", "co", "company",
              "solar", "energy", "account", "customer", "school", "district",
              "department", "association", "homeowners"}

def _norm(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def _is_cumulative(s: str) -> bool:
    s = _norm(s)
    return any(w in s for w in _CUM_WORDS)


def name_hint_tokens(name: Optional[str]) -> list[str]:
    """Distinctive lowercase tokens from an offtaker's name, used to recognize THEIR
    own column in a multi-column sheet — e.g. 'Town of Fairlee' → ['fairlee'], which
    lets the detector pick 'kWh Fairlee' (their share) over a cumulative or whole-array
    kWh column. Drops short + non-distinctive words so we don't match on 'Town'/'Solar'."""
    if not name:
        return []
    toks = re.split(r"[^a-z0-9]+", _norm(name))
    return [t for t in toks if len(t) >= 4 and t not in _NAME_STOP]


def _open_grid(file_bytes: bytes, filename: str = "",
               hint_tokens: Optional[list[str]] = None) -> tuple[list[list], Optional[str], str]:
    """Open the upload into (grid, sheet_name, kind) the detector works on."""
    is_csv = filename.lower().endswith(".csv") or _looks_like_csv(file_bytes)
    if is_csv:
        return _read_csv_grid(file_bytes), None, "csv"
    grid, sheet = _read_xlsx_grid(file_bytes, hint_tokens)
    return grid, sheet, "xlsx"


def _period_tail(grid: list[list], header_row: int, columns: dict) -> tuple[int, Optional[str]]:
    """(data_rows, last_period) below a header — shared by the heuristic + the AI path."""
    data = [r for r in grid[header_row + 1:] if any(_norm(c) for c in r)]
    last_period = None
    pc = (columns or {}).get("period")
    if data and isinstance(pc, int):
        for r in reversed(data):
            if pc < len(r) and _norm(r[pc]):
                last_period = str(r[pc]).strip()
                break
    return len(data), last_period


_MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]


def _period_style(sample: Any) -> str:
    """Classify how this sheet writes its billing period, so an appended row matches it
    instead of always reading ISO. -> 'monthname' | 'monthname_year' | 'slash' | 'iso'."""
    s = _norm(sample)
    if not s:
        return "iso"
    if any(mn in s for mn in _MONTHS):                              # "April", "June 2026"
        return "monthname_year" if re.search(r"20\d{2}", s) else "monthname"
    if re.search(r"\d{1,2}[/.\-]\d{1,2}[/.\-]20\d{2}", s):          # 5/31/2026
        return "slash"
    return "iso"                                                    # 2026-05 (our default)


def _format_period(iso_label: Optional[str], style: Optional[str]) -> Optional[str]:
    """Render an ISO 'YYYY-MM' label in the sheet's own period style (see _period_style)."""
    ym = _ym(iso_label or "")
    if not ym or style in (None, "iso"):
        return iso_label
    year, month = ym
    name = _MONTH_NAMES[month - 1] if 1 <= month <= 12 else None
    if style == "monthname" and name:
        return name
    if style == "monthname_year" and name:
        return f"{name} {year}"
    if style == "slash":
        return f"{month:02d}/{year}"
    return iso_label


def _first_period_sample(file_bytes: bytes, mapping: dict) -> Optional[str]:
    """Period cell of the FIRST original data row (appends go to the bottom, so row 1 keeps
    the sheet's native style) — back-fills period_style for trackers detected before that
    field existed, so legacy sheets also append in-style. Best-effort; None on any problem."""
    try:
        cols = mapping.get("columns") or {}
        pc = cols.get("period")
        hr = int(mapping.get("header_row") or 0)
        if not isinstance(pc, int):
            return None
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = mapping.get("sheet")
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=hr + 2, max_row=hr + 40,
                                min_col=pc + 1, max_col=pc + 1, values_only=True):
            v = row[0]
            if v is not None and str(v).strip():
                return str(v).strip()
    except Exception:  # noqa: BLE001
        return None
    return None


def _score_cell(cell: str) -> dict[str, tuple[float, str]]:
    """For one header cell, return {field: (score, matched_keyword)} for every
    field whose keyword appears in the cell. Higher score = more confident.

    Score = keyword_weight × len(keyword) − position_penalty, with a big bonus
    when the cell *is* the keyword. The per-keyword weight lets an unambiguous
    unit-bearing token (e.g. '$/kwh' → rate) out-rank a vaguer token (e.g.
    'credit' → amount) that shares the same cell."""
    out: dict[str, tuple[float, str]] = {}
    if not cell:
        return out
    for field, kws in _FIELD_KEYWORDS:
        best: Optional[tuple[float, str]] = None
        for kw, weight in kws:
            idx = cell.find(kw)
            if idx < 0:
                continue
            score = weight * (len(kw) + 1.0) - idx * 0.5
            if cell == kw:
                score += 50
            if best is None or score > best[0]:
                best = (score, kw)
        if best is not None:
            out[field] = best
    return out


def detect_structure(file_bytes: bytes, filename: str = "",
                     hint_tokens: Optional[list[str]] = None) -> dict:
    """The detector. Open the uploaded XLSX/CSV, find the header row, and map
    each logical field to a column index. Returns a JSON-serializable mapping:

        {
          "ok": bool,
          "kind": "xlsx" | "csv",
          "sheet": <sheet name or None for csv>,
          "header_row": <0-based row index of the header within the sheet>,
          "columns": {"period": <col idx>, "generation": <col idx>, ...},
          "headers": [<raw header strings>],
          "data_rows": <count of data rows below the header>,
          "last_period": <string of the last data row's period cell, or None>,
          "warnings": [..],
        }

    Pure + offline. Never raises — returns ok=False with a warning instead.
    """
    try:
        grid, sheet, kind = _open_grid(file_bytes, filename, hint_tokens)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "warnings": [f"Could not open the file: {e}"]}

    if not grid:
        return {"ok": False, "warnings": ["The file appears to be empty."]}

    header_row, columns, headers = _find_header(grid, hint_tokens)
    if header_row is None or "generation" not in columns:
        return {
            "ok": False,
            "kind": kind,
            "sheet": sheet,
            "headers": headers or [],
            "warnings": [
                "Couldn't confidently find a generation/kWh column. "
                "You can map the columns manually."
            ],
        }

    data_rows, last_period = _period_tail(grid, header_row, columns)
    return {
        "ok": True,
        "kind": kind,
        "sheet": sheet,
        "header_row": header_row,
        "columns": columns,
        "headers": headers,
        "data_rows": data_rows,
        "last_period": last_period,
        "period_style": _period_style(last_period),
        "via": "heuristic",
        "warnings": [],
    }


def _looks_numeric(v) -> bool:
    """True if a cell holds a number (an actual data value), tolerating $ , and spaces."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(str(v).replace(",", "").replace("$", "").strip())
        return True
    except Exception:  # noqa: BLE001
        return False


def _find_header(grid: list[list],
                 hint_tokens: Optional[list[str]] = None) -> tuple[Optional[int], dict, list]:
    """Scan the first ~15 rows for the row that best looks like a header (the
    most fields matched, weighted by confidence). Returns
    (header_row_idx, {field: col_idx}, raw_headers).

    Two refinements on the generation column (the most ambiguous on real sheets):
      * a CUMULATIVE/running-total kWh column (this cell or the sub-header above it
        says 'cumm'/'ytd'/…) is demoted — it isn't the monthly figure we append.
      * a column whose header bears the OFFTAKER'S name (hint_tokens, e.g. 'Fairlee')
        is boosted — that's their generation share, the figure their invoice uses."""
    best_row = None
    best_cols: dict[str, int] = {}
    best_headers: list = []
    best_score = 0.0
    for ridx in range(min(len(grid), 15)):
        row = grid[ridx]
        # field -> (best_col_idx, score)
        claims: dict[str, tuple[int, float]] = {}
        for cidx, cell in enumerate(row):
            cell_n = _norm(cell)
            scores = _score_cell(cell_n)
            if "generation" in scores:
                sc, kw = scores["generation"]
                above = _norm(grid[ridx - 1][cidx]) if ridx > 0 and cidx < len(grid[ridx - 1]) else ""
                if _is_cumulative(cell_n) or _is_cumulative(above):
                    sc *= 0.05                       # cumulative/running-total → not the monthly column
                elif hint_tokens and any(t and t in cell_n for t in hint_tokens):
                    sc += 50.0                       # bears the offtaker's name → it IS their share
                # …but the active generation column must hold data in the RECENT rows (that's
                # where we append). An empty 'kWh <offtaker>' column next to a populated 'kWh
                # whole array' must not win on the name alone — demote it hard if the last rows
                # are blank, so the column that actually carries the data wins.
                _data = grid[ridx + 1:]
                _recent = _data[-15:] if len(_data) > 15 else _data
                if not any(cidx < len(rr) and _looks_numeric(rr[cidx]) for rr in _recent):
                    sc *= 0.02
                scores["generation"] = (sc, kw)
            for field, (sc, _kw) in scores.items():
                if field not in claims or sc > claims[field][1]:
                    claims[field] = (cidx, sc)
        # Resolve any column claimed by two fields → keep the higher score, demote
        # the loser to its next-best (handled implicitly: each field independently
        # picks its best column; collisions are rare and the score gap resolves it).
        cols = {f: ci for f, (ci, _s) in claims.items()}
        # de-dup columns: if two fields point at the same column, keep the
        # stronger-scoring field on it, drop the weaker.
        seen: dict[int, str] = {}
        for f in sorted(claims, key=lambda x: -claims[x][1]):
            ci = claims[f][0]
            if ci in seen:
                cols.pop(f, None)
            else:
                seen[ci] = f
        total = sum(claims[f][1] for f in cols)
        # require at least a generation-ish column to call it a header
        if "generation" in cols and total > best_score:
            best_score = total
            best_row = ridx
            best_cols = cols
            best_headers = [str(c).strip() if c is not None else "" for c in row]
    return best_row, best_cols, best_headers


def _add_month(d):
    """A date advanced ~one month (clamped to the month's length)."""
    import calendar
    import datetime as _dt
    y, m = d.year, d.month + 1
    if m > 12:
        m = 1; y += 1
    return _dt.date(y, m, min(d.day, calendar.monthrange(y, m)[1]))


def _shift_formula(f: str, drow: int) -> str:
    """Shift the RELATIVE (non-$) row refs in a formula by drow so a copied formula points at
    the new row; $-anchored rows are left alone. Best-effort — returns the original on trouble."""
    try:
        def repl(m):
            col, rowdollar, rownum = m.group(1), m.group(2), m.group(3)
            return m.group(0) if rowdollar == "$" else f"{col}{int(rownum) + drow}"
        return re.sub(r"(\$?[A-Z]{1,3})(\$?)(\d+)", repl, f)
    except Exception:  # noqa: BLE001
        return f


def _complete_row(ws, write_row: int, header_row: int, written: dict, gen_value,
                  period_start=None, period_end=None) -> None:
    """Fill the columns NOT explicitly mapped by CONTINUING the sheet's own pattern from the
    last data rows — the deterministic "just take the previous row into the next" completion:
      * date columns advance (exact billing dates when known, else by the period delta),
      * a column that's a fixed ratio of generation = the per-offtaker kWh share (new gen × ratio),
      * a constant column (e.g. Adder 0) carries forward,
      * a sequential +1 column (invoice #) increments,
      * formula columns (Value, Savings) copy with their row refs shifted to the new row.
    Never fabricates — leaves a column blank when there's no confident pattern. `written` is a
    {field: 0-based col} dict of the columns already filled."""
    import datetime as _dt
    prev = write_row - 1
    if prev < header_row + 2:
        return
    def _d(v):
        return v.date() if isinstance(v, _dt.datetime) else v
    rws = list(range(max(header_row + 2, prev - 2), prev + 1))   # up to 3 prior rows, oldest→newest
    max_col = ws.max_column
    gen_col = written.get("generation")
    done = {c for c in written.values() if isinstance(c, int)}   # 0-based cols already filled
    # Date columns: leftmost = period start, next = period end → use the EXACT billing dates.
    date_cols = [col for col in range(1, max_col + 1)
                 if (col - 1) not in done and isinstance(ws.cell(row=prev, column=col).value, (_dt.date, _dt.datetime))]
    if period_start and len(date_cols) >= 1:
        ws.cell(row=write_row, column=date_cols[0]).value = period_start; done.add(date_cols[0] - 1)
    if period_end and len(date_cols) >= 2:
        ws.cell(row=write_row, column=date_cols[1]).value = period_end; done.add(date_cols[1] - 1)
    for col in range(1, max_col + 1):
        if (col - 1) in done:
            continue
        vals = [ws.cell(row=r, column=col).value for r in rws]
        last = vals[-1]
        if last is None or (isinstance(last, str) and not last.strip()):
            continue
        cell = ws.cell(row=write_row, column=col)
        if isinstance(last, str) and last.startswith("="):
            cell.value = _shift_formula(last, write_row - prev)            # formula → shift to new row
        elif isinstance(last, (_dt.date, _dt.datetime)):
            if len(vals) >= 2 and isinstance(vals[-2], (_dt.date, _dt.datetime)):
                cell.value = _d(last) + (_d(last) - _d(vals[-2]))          # advance by the period delta
            else:
                cell.value = _add_month(_d(last))
        elif isinstance(last, (int, float)):
            nums = [v for v in vals if isinstance(v, (int, float))]
            if len(nums) == len(vals):
                if len(set(nums)) == 1:
                    cell.value = nums[-1]                                  # constant (e.g. Adder 0)
                elif len(nums) >= 2 and all(nums[i + 1] - nums[i] == 1 for i in range(len(nums) - 1)):
                    cell.value = nums[-1] + 1                              # sequential (invoice #)
                elif gen_col is not None and gen_value:
                    gprev = ws.cell(row=prev, column=gen_col + 1).value
                    if isinstance(gprev, (int, float)) and gprev:
                        ratio = last / gprev
                        if 0 < ratio <= 1.0001:                            # per-offtaker share = gen × ratio
                            cell.value = round(float(gen_value) * ratio, 2)
        elif isinstance(last, str) and len({str(v) for v in vals}) == 1:
            cell.value = last                                             # text constant → copy


# ─── append a new period row ─────────────────────────────────────────────────

def append_period_row(file_bytes: bytes, mapping: dict, row_values: dict) -> bytes:
    """Append ONE new data row to the workbook, writing each value into the
    column the detector mapped, preserving everything else. XLSX only (the
    stored source-of-truth is converted to xlsx on upload — see ingest_upload).

    `mapping` is a detect_structure() result. `row_values` is
    {field: value} for fields in {period, generation, consumption, rate, amount}.
    Returns the new workbook bytes. Raises on a structural problem (caller guards).
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes))
    sheet = mapping.get("sheet")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    header_row = int(mapping.get("header_row") or 0)
    cols = mapping.get("columns") or {}

    # Find the first empty data row at/after header_row+2 (openpyxl is 1-based;
    # header_row is 0-based). Walk down past existing rows in the mapped columns.
    probe_cols = [c for c in cols.values() if isinstance(c, int)]
    write_row = header_row + 2  # 1-based row just below the header
    max_row = ws.max_row
    r = header_row + 2
    while r <= max_row:
        if any((ws.cell(row=r, column=c + 1).value not in (None, ""))
               for c in probe_cols):
            write_row = r + 1
        r += 1

    written: dict = {}
    for field, col_idx in cols.items():
        if field in row_values and isinstance(col_idx, int):
            val = row_values[field]
            if val is None:
                continue
            ws.cell(row=write_row, column=col_idx + 1, value=val)
            written[field] = col_idx
    # Complete the rest of the row by continuing the sheet's own pattern (dates, the per-offtaker
    # share, constants like Adder, invoice #, formula columns) — best-effort, never raises.
    try:
        import datetime as _dt
        def _pd(s):
            try:
                return _dt.date.fromisoformat(str(s)[:10])
            except Exception:  # noqa: BLE001
                return None
        _complete_row(ws, write_row, header_row, written, row_values.get("generation"),
                      period_start=_pd(row_values.get("_period_start")),
                      period_end=_pd(row_values.get("_period_end")))
    except Exception:  # noqa: BLE001
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def period_label_for(computed: dict) -> Optional[str]:
    """A stable, human period label for idempotency + the appended period cell:
    'YYYY-MM' from period_end (the billing month). Mirrors the invoice number."""
    pe = computed.get("period_end")
    if pe:
        try:
            d = date.fromisoformat(str(pe)[:10])
            return d.strftime("%Y-%m")
        except Exception:  # noqa: BLE001
            pass
    return computed.get("month") or computed.get("invoice_number")


def _period_matches(existing: Optional[str], new_label: Optional[str]) -> bool:
    """Idempotency check: does the workbook's last period already cover this one?
    Tolerant of format drift ('2026-05', 'May 2026', '5/31/2026', etc.) by
    comparing the (year, month) we can extract from each."""
    if not existing or not new_label:
        return False
    a = _ym(existing)
    b = _ym(new_label)
    if a and b:
        return a == b
    return _norm(existing) == _norm(new_label)


_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct",
     "nov", "dec"], start=1)}


def _ym(s: str) -> Optional[tuple[int, int]]:
    s = _norm(s)
    m = re.search(r"(\d{1,2})[-/.]\d{1,2}[-/.](20\d{2})", s)  # 5/31/2026 (MM/DD/YYYY)
    if m:
        return int(m.group(2)), int(m.group(1))
    m = re.search(r"(20\d{2})[-/.](\d{1,2})", s)          # 2026-05
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"(\d{1,2})[-/.](20\d{2})", s)           # 05/2026
    if m:
        return int(m.group(2)), int(m.group(1))
    for name, num in _MONTHS.items():
        if name in s:
            y = re.search(r"20\d{2}", s)
            if y:
                return int(y.group(0)), num
    return None


# ─── public: append the latest period to a subscription's stored sheet ───────

def _row_from_computed(computed: dict, cols: dict, label: str, style) -> Optional[dict]:
    """Build the {col: value} row for ONE period from its computed invoice, using the
    SAME figures the invoice itself uses (no recomputation, no fabrication). None when
    there's no generation to record (→ caller skips the period)."""
    # Generation column = the WHOLE-ARRAY figure (what the existing rows hold + what the GMP
    # bill states), not the offtaker's allocated share. Prefer array_kwh, then project total,
    # then the allocated kWh only as a last resort.
    gen = computed.get("array_kwh")
    if gen in (None, 0):
        gen = computed.get("project_total_kwh")
    if gen in (None, 0):
        gen = computed.get("kwh")
    if not gen:
        return None
    row: dict[str, Any] = {}
    if "period" in cols:
        # The sheet's OWN style ("June 2026" vs ISO "2026-06") so rows look native.
        row["period"] = _format_period(label, style)
    if "generation" in cols:
        row["generation"] = round(float(gen), 2)
    if "consumption" in cols and computed.get("consumption_kwh") is not None:
        row["consumption"] = round(float(computed["consumption_kwh"]), 2)
    if "rate" in cols and computed.get("effective_rate_per_kwh") is not None:
        row["rate"] = round(float(computed["effective_rate_per_kwh"]), 6)
    if "amount" in cols and computed.get("amount_owed") is not None:
        row["amount"] = round(float(computed["amount_owed"]), 2)
    # Exact billing dates for row completion to fill date columns (underscore keys are NOT mapped
    # write fields — they're consumed by _complete_row in append_period_row).
    if computed.get("period_start"):
        row["_period_start"] = computed["period_start"]
    if computed.get("period_end"):
        row["_period_end"] = computed["period_end"]
    return row


def _offtaker_billed_labels(db, sub) -> Optional[list]:
    """Every billed period ("YYYY-MM", oldest→newest) the offtaker SHOULD have a row for
    — its GMP bills with excess sent to grid. None when the offtaker isn't bound to a
    utility account (→ caller uses the single-latest fallback)."""
    acct_id = getattr(sub, "utility_account_id", None)
    if acct_id is None:
        return None
    from sqlalchemy import select
    from ..models import Bill
    bills = db.execute(
        select(Bill).where(
            Bill.account_id == acct_id,
            Bill.kwh_sent_to_grid.isnot(None), Bill.kwh_sent_to_grid > 0,
            Bill.period_end.isnot(None))
    ).scalars().all()
    labels = set()
    for b in bills:
        pe = b.period_end.date() if isinstance(b.period_end, datetime) else b.period_end
        if pe:
            labels.add(pe.strftime("%Y-%m"))
    return sorted(labels)


def _bare_month(s: str) -> Optional[int]:
    """Month number from a bare month name with NO year ('April' -> 4); None otherwise."""
    sv = _norm(s)
    if re.search(r"20\d{2}", sv):
        return None
    for name, num in _MONTHS.items():
        if name in sv:
            return num
    return None


def _sheet_data_row_count(file_bytes: bytes, mapping: dict) -> int:
    """How many data rows hold a PARSEABLE period (a real month/date). Junk or typo cells like
    'Aguust', and blanks, don't count. The sheet's rows align 1:1 with the offtaker's EARLIEST
    billed periods, so billed[count:] is exactly the recent periods missing from the tail — no
    fragile year parsing, immune to a single garbled cell that otherwise broke the anchor."""
    n = 0
    try:
        cols = mapping.get("columns") or {}
        pc = cols.get("period")
        hr = int(mapping.get("header_row") or 0)
        if not isinstance(pc, int):
            return 0
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = mapping.get("sheet")
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        for row in ws.iter_rows(min_row=hr + 2, min_col=pc + 1, max_col=pc + 1, values_only=True):
            v = row[0]
            if v is None or not str(v).strip():
                continue
            if _ym(str(v)) or _bare_month(str(v)):
                n += 1
    except Exception:  # noqa: BLE001
        return 0
    return n


def _sheet_context(file_bytes: bytes, mapping: dict, n_rows: int = 6):
    """(headers, recent_rows, present_labels) for the AI planner. recent_rows keep formulas as
    strings and dates as ISO; present_labels = each data row's period as 'YYYY-MM' (from a date
    column when available, else the raw period cell) so the model knows the full set already there."""
    import datetime as _dt
    from openpyxl import load_workbook
    cols = mapping.get("columns") or {}
    pc = cols.get("period"); hr = int(mapping.get("header_row") or 0)
    sheet = mapping.get("sheet")
    wb = load_workbook(io.BytesIO(file_bytes))   # NOT data_only → formula cells come back as strings
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    maxc = ws.max_column

    def cv(v):
        return v.strftime("%Y-%m-%d") if isinstance(v, (_dt.date, _dt.datetime)) else v
    headers = [cv(ws.cell(row=hr + 1, column=c).value) for c in range(1, maxc + 1)]
    data_rows = [r for r in range(hr + 2, ws.max_row + 1)
                 if isinstance(pc, int) and ws.cell(row=r, column=pc + 1).value not in (None, "")]
    last = data_rows[-1] if data_rows else None
    date_cols = [c for c in range(1, maxc + 1)
                 if last and isinstance(ws.cell(row=last, column=c).value, (_dt.date, _dt.datetime))]
    end_col = date_cols[-1] if date_cols else ((pc + 1) if isinstance(pc, int) else 1)

    def rper(r):
        v = ws.cell(row=r, column=end_col).value
        if isinstance(v, (_dt.date, _dt.datetime)):
            return v.strftime("%Y-%m")
        ym = _ym(str(v)) if v is not None else None   # text dates ("7/18/2025") + "YYYY-MM-DD"
        if ym:
            return f"{ym[0]:04d}-{ym[1]:02d}"
        pv = ws.cell(row=r, column=(pc + 1) if isinstance(pc, int) else 1).value
        ym2 = _ym(str(pv)) if pv is not None else None
        return f"{ym2[0]:04d}-{ym2[1]:02d}" if ym2 else str(pv)
    present = [rper(r) for r in data_rows]
    recent = [[cv(ws.cell(row=r, column=c).value) for c in range(1, maxc + 1)] for r in data_rows[-n_rows:]]
    return headers, recent, present


def _candidate_facts(sub, billed, n: int = 14) -> list:
    """Real per-period figures (computed canonically, NOT by the model) for the recent billed
    periods — the model arranges these into rows, it never invents the billing math."""
    import calendar
    from .delivery import build_match
    facts = []
    for lbl in (billed or [])[-n:]:
        try:
            ci = getattr(build_match(sub, period_label=lbl), "computed_invoice", None) or {}
        except Exception:  # noqa: BLE001
            continue
        try:
            mo = int(lbl[5:7])
        except Exception:  # noqa: BLE001
            mo = 0
        facts.append({"period": lbl, "month": calendar.month_name[mo] if 1 <= mo <= 12 else lbl,
                      "start": ci.get("period_start"), "end": ci.get("period_end"),
                      "whole_kwh": ci.get("array_kwh"), "share_kwh": ci.get("kwh"),
                      "rate": ci.get("effective_rate_per_kwh"), "amount": ci.get("amount_owed")})
    return facts


def append_ai_rows(file_bytes: bytes, mapping: dict, ai_rows: list,
                   present: Optional[set] = None) -> bytes:
    """Append fully-specified rows (each {col_index: value}) from the AI planner. ISO date strings
    become dates; '=' strings become live formulas; null is skipped. Each new cell inherits the
    number FORMAT of the last existing data row, so dates/numbers display EXACTLY like the sheet.
    A row whose period (YYYY-MM, from its own dates) is already in `present` is skipped — a
    deterministic guard so the model can never duplicate a month already on the sheet."""
    import datetime as _dt
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes))
    sheet = mapping.get("sheet")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    hr = int(mapping.get("header_row") or 0)
    cols = mapping.get("columns") or {}
    probe = [c for c in cols.values() if isinstance(c, int)]
    present = set(present or ())

    template = None   # last existing data row → its per-column number_format is the template
    _drows = []
    r = hr + 2
    while r <= ws.max_row:
        if any(ws.cell(row=r, column=c + 1).value not in (None, "") for c in probe):
            template = r
            _drows.append(r)
        r += 1
    # Columns holding the SAME value in every existing data row → carry that EXACT constant onto
    # every new row (the operator's flat tariff, a zero adder), overriding whatever the model placed.
    # Varying columns (generation, share, dates) and formulas are left to the model.
    constants = {}
    for col in range(1, ws.max_column + 1):
        nb = [ws.cell(row=rr, column=col).value for rr in _drows]
        nb = [v for v in nb if v not in (None, "")]
        if len(nb) >= 2 and len(set(nb)) == 1 and not (isinstance(nb[0], str) and str(nb[0]).startswith("=")):
            constants[col] = nb[0]

    def next_row():
        wr = hr + 2
        rr = hr + 2
        while rr <= ws.max_row:
            if any(ws.cell(row=rr, column=c + 1).value not in (None, "") for c in probe):
                wr = rr + 1
            rr += 1
        return wr

    def conv(v):
        if isinstance(v, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", v.strip()):
            try:
                return _dt.date.fromisoformat(v.strip())
            except Exception:  # noqa: BLE001
                return v
        return v

    def row_period(ai_row):
        ds = []
        for v in ai_row.values():
            if isinstance(v, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", v.strip()):
                try:
                    ds.append(_dt.date.fromisoformat(v.strip()))
                except Exception:  # noqa: BLE001
                    pass
        return max(ds).strftime("%Y-%m") if ds else None

    for ai_row in ai_rows:
        if not isinstance(ai_row, dict):
            continue
        if present and row_period(ai_row) in present:
            continue                                  # already on the sheet → never duplicate
        wr = next_row()
        for k, v in ai_row.items():
            try:
                ci = int(k)
            except (TypeError, ValueError):
                continue
            if v is None:
                continue
            cell = ws.cell(row=wr, column=ci + 1, value=conv(v))
            if template is not None:                  # mimic the column's date/number format exactly
                cell.number_format = ws.cell(row=template, column=ci + 1).number_format
        for col, cval in constants.items():           # carry the flat tariff / zero adder exactly
            cc = ws.cell(row=wr, column=col, value=cval)
            if template is not None:
                cc.number_format = ws.cell(row=template, column=col).number_format
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def update_subscription_sheet(db, sub) -> dict:
    """RECONCILE the offtaker's stored BYO generation spreadsheet with their billed
    history: append a row for EVERY billed period missing from the sheet — so deleting a
    row (or re-uploading an edited file) re-populates it — each valued with that period's
    own real, canonically-computed figures. Best-effort: never raises into the bill-pull
    path. Idempotent: a period already in the sheet is never duplicated.

    Returns: {"status":"appended","periods":[...],"count":N} | {"status":"skipped",...}
             | {"status":"error","error":...}
    """
    if not tracker_enabled():
        return {"status": "skipped", "reason": "feature-disabled"}
    try:
        bytes_ = getattr(sub, "tracker_workbook", None)
        mapping = getattr(sub, "tracker_map", None)
        if not bytes_ or not mapping or not mapping.get("ok"):
            return {"status": "skipped", "reason": "no-tracker-sheet"}

        from .delivery import build_match
        cols = mapping.get("columns") or {}
        # Derive the write-style from the sheet's ACTUAL first row — the STORED style can be
        # wrong (seen live: 'iso' on a month-name sheet, which would write "2026-05" among
        # "May"/"June" rows). Fall back to the stored style, then iso.
        style = _period_style(_first_period_sample(bytes(bytes_), mapping)) or mapping.get("period_style") or "iso"

        billed = _offtaker_billed_labels(db, sub)

        # ── AI path (when the model is funded): hand it the sheet + the real billed figures and let
        # it decide which months are missing, produce a COMPLETE row for each, and validate. The
        # heuristic path below is the fallback when the AI is off, errors, or low-confidence.
        from . import sheet_tracker_ai as _ai
        if billed and _ai.ai_available():
            plan = None
            try:
                headers, recent, present = _sheet_context(bytes(bytes_), mapping)
                facts = _candidate_facts(sub, billed)
                if facts:
                    plan = _ai.ai_plan_rows(headers, recent, facts, sheet=mapping.get("sheet"),
                                            offtaker=getattr(sub, "customer_name", None),
                                            present_labels=present)
            except Exception:  # noqa: BLE001 — never break the reconcile over the AI path
                plan = None
            if plan is not None:
                rows = plan.get("rows") or []
                ai_meta = {"sane": plan.get("sane"), "explanation": plan.get("explanation"), "via": "ai"}
                if not rows:
                    return {"status": "skipped", "reason": "ai-nothing-missing", "ai": ai_meta}
                sub.tracker_workbook = append_ai_rows(bytes(bytes_), mapping, rows, present=set(present))
                pcol = cols.get("period")
                added = [r.get(pcol) for r in rows if isinstance(pcol, int) and r.get(pcol) is not None]
                new_map = dict(mapping)
                new_map["last_period"] = billed[-1]
                new_map["data_rows"] = int(mapping.get("data_rows") or 0) + len(rows)
                sub.tracker_map = new_map
                sub.tracker_updated_at = datetime.utcnow()
                db.add(sub)
                return {"status": "appended", "via": "ai", "count": len(rows),
                        "periods": [str(a) for a in added], "ai": ai_meta}
            # plan is None (AI off/failed) → fall through to the deterministic reconcile.

        # ── Reconcile path (GMP-bound offtaker) ── The sheet's data rows are chronological and
        # align 1:1 with the offtaker's earliest billed periods, so the periods MISSING from the
        # tail = billed[<# parseable rows already in the sheet>:]. This re-adds the recently
        # deleted/just-landed months without parsing any row's year — immune to bare month names
        # and to a single garbled cell. (Assumes the sheet starts at the account's first bill;
        # the cap bounds any misalignment.)
        if billed:
            n_rows = _sheet_data_row_count(bytes(bytes_), mapping)
            missing = billed[n_rows:] if 0 <= n_rows < len(billed) else []
            # Safety cap: never dump a long history — append at most the most-recent MAX_BACKFILL.
            MAX_BACKFILL = 14
            if len(missing) > MAX_BACKFILL:
                missing = missing[-MAX_BACKFILL:]
            cur = bytes(bytes_)
            appended: list = []
            for lbl in missing:                     # oldest → newest, so rows land in order
                m = build_match(sub, period_label=lbl)
                row = _row_from_computed(getattr(m, "computed_invoice", None) or {}, cols, lbl, style)
                if row is None:
                    continue                        # no excess bill for that period → skip, never fabricate
                cur = append_period_row(cur, mapping, row)
                appended.append(lbl)
            if not appended:
                return {"status": "skipped", "reason": "period-already-present",
                        "period": billed[-1]}
            sub.tracker_workbook = cur
            new_map = dict(mapping)
            new_map["period_style"] = style
            new_map["last_period"] = billed[-1]
            new_map["data_rows"] = int(mapping.get("data_rows") or 0) + len(appended)
            sub.tracker_map = new_map
            sub.tracker_updated_at = datetime.utcnow()
            db.add(sub)
            return {"status": "appended", "periods": appended, "count": len(appended)}

        # ── Fallback (no utility account, e.g. workbook-driven offtaker): single latest ──
        match = build_match(sub)
        computed = getattr(match, "computed_invoice", None) or {}
        label = period_label_for(computed)
        if _period_matches(mapping.get("last_period"), label):
            return {"status": "skipped", "reason": "period-already-present", "period": label}
        row_values = _row_from_computed(computed, cols, label, style)
        if row_values is None:
            return {"status": "skipped", "reason": "no-generation-figure"}
        sub.tracker_workbook = append_period_row(bytes(bytes_), mapping, row_values)
        new_map = dict(mapping)
        new_map["period_style"] = style
        new_map["last_period"] = label
        new_map["data_rows"] = int(mapping.get("data_rows") or 0) + 1
        sub.tracker_map = new_map
        sub.tracker_updated_at = datetime.utcnow()
        db.add(sub)
        return {"status": "appended", "period": label,
                "amount": row_values.get("amount"), "generation": row_values.get("generation")}
    except Exception as e:  # noqa: BLE001 — never break a bill pull over the sheet
        logger.exception("sheet_tracker append failed for sub %s", getattr(sub, "id", "?"))
        return {"status": "error", "error": f"{type(e).__name__}: {e}"}


def update_all_for_account(db, tenant_id: str, utility_account_id: int) -> list[dict]:
    """After a NEW GMP bill lands for `utility_account_id`, append to every
    enabled offtaker subscription bound to that account that has a tracker sheet.
    Best-effort; returns a per-subscription status list. Caller commits."""
    if not tracker_enabled():
        return []
    from sqlalchemy import select
    from ..models import BillingReportSubscription
    out: list[dict] = []
    try:
        subs = db.execute(
            select(BillingReportSubscription).where(
                BillingReportSubscription.tenant_id == tenant_id,
                BillingReportSubscription.utility_account_id == utility_account_id,
                BillingReportSubscription.enabled == True,  # noqa: E712
                BillingReportSubscription.deleted_at.is_(None),
            )
        ).scalars().all()
    except Exception as e:  # noqa: BLE001
        logger.exception("sheet_tracker: could not load subs for account %s",
                         utility_account_id)
        return [{"status": "error", "error": f"{type(e).__name__}: {e}"}]
    for sub in subs:
        if not getattr(sub, "tracker_workbook", None):
            continue
        r = update_subscription_sheet(db, sub)
        r["subscription_id"] = sub.id
        out.append(r)
    return out


# ─── file readers ────────────────────────────────────────────────────────────

def _looks_like_csv(b: bytes) -> bool:
    if b[:2] == b"PK" or b[:4] == b"\xd0\xcf\x11\xe0":
        return False  # xlsx / xls
    sample = b[:4096]
    try:
        text = sample.decode("utf-8", errors="ignore")
    except Exception:  # noqa: BLE001
        return False
    return ("," in text or "\t" in text or ";" in text) and "\x00" not in text


def _read_csv_grid(b: bytes) -> list[list]:
    import csv
    text = b.decode("utf-8-sig", errors="replace")
    # sniff the delimiter (comma / tab / semicolon)
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",\t;").delimiter
    except Exception:  # noqa: BLE001
        for d in ("\t", ";", ","):
            if d in sample:
                delim = d
                break
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    return rows


def _read_xlsx_grid(b: bytes,
                    hint_tokens: Optional[list[str]] = None) -> tuple[list[list], Optional[str]]:
    """Return (grid, sheet_name) for the sheet most likely to hold the ledger:
    the one whose first ~15 rows best match our header heuristic, else active."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(b), read_only=True, data_only=True)
    best_grid: list[list] = []
    best_name: Optional[str] = None
    best_score = -1.0
    for ws in wb.worksheets:
        grid = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            grid.append(list(row))
            if i > 300:
                break
        _hr, cols, _h = _find_header(grid, hint_tokens)
        score = len(cols) + (5 if "generation" in cols else 0)
        # Prefer the tab that holds the real LEDGER, not a 'SAMPLE'/'Template' example: weight
        # by how many data rows sit under the header, and penalize obvious non-data tab names.
        # (Real case: a workbook with a 17-row 'SAMPLE' tab AND the true 55-row 'Data' tab — both
        # map the same columns, so without this tie-break the sample tab silently won.)
        pc = cols.get("period")
        if isinstance(pc, int):
            ndata = sum(1 for r in grid[_hr + 1:]
                        if pc < len(r) and r[pc] not in (None, "") and str(r[pc]).strip())
            score += min(ndata, 60) * 0.1
        nm = (ws.title or "").lower()
        if any(w in nm for w in ("sample", "template", "example", "true up", "trueup",
                                 "trend", "instruction", "readme", "notes", "summary")):
            score -= 4.0
        if score > best_score:
            best_score = score
            best_grid = grid
            best_name = ws.title
    wb.close()
    return best_grid, best_name


def ingest_upload(file_bytes: bytes, filename: str,
                  offtaker_name: Optional[str] = None) -> dict:
    """Process an uploaded BYO sheet at upload time: detect structure, and
    normalize the stored bytes to XLSX (so append_period_row always has a real
    workbook to write into — even when the operator uploaded a CSV). Returns:

        {"ok": bool, "mapping": {..}, "workbook": <xlsx bytes>, "warnings": [..]}

    offtaker_name (optional): the customer this sheet bills. Used two ways — the
    deterministic heuristic boosts a column bearing their name (name_hint_tokens),
    and, when an ANTHROPIC_API_KEY is configured, the AI mapper reads their name to
    pick their share column on an arbitrary layout. Detection is deterministic +
    offline by default; the AI is a best-effort upgrade that NEVER blocks upload.
    """
    hint_tokens = name_hint_tokens(offtaker_name)
    mapping = detect_structure(file_bytes, filename, hint_tokens)

    # ── AI upgrade ────────────────────────────────────────────────────────────
    # When a key is present, let the model map columns on this sheet's real layout
    # (handles any format, multiple kWh columns, buried headers). Best-effort: any
    # failure leaves the heuristic mapping in place. Result is still surfaced to the
    # operator (mapping chips + remap endpoint), so a wrong guess stays correctable.
    try:
        from .sheet_tracker_ai import ai_available, ai_map_columns
        if ai_available():
            grid, sheet, kind = _open_grid(file_bytes, filename, hint_tokens)
            ai = ai_map_columns(grid, sheet, offtaker_name)
            if ai and grid and 0 <= ai["header_row"] < len(grid):
                hr, cols = ai["header_row"], ai["columns"]
                headers = [str(c).strip() if c is not None else "" for c in grid[hr]]
                data_rows, last_period = _period_tail(grid, hr, cols)
                mapping = {"ok": True, "kind": kind, "sheet": sheet, "header_row": hr,
                           "columns": cols, "headers": headers, "data_rows": data_rows,
                           "last_period": last_period, "period_style": _period_style(last_period), "via": "ai",
                           "ai_confidence": ai.get("confidence"),
                           "ai_reasoning": ai.get("reasoning"), "warnings": []}
    except Exception:  # noqa: BLE001 — the AI path can NEVER break an upload
        pass

    if not mapping.get("ok"):
        return {"ok": False, "mapping": mapping,
                "warnings": mapping.get("warnings", [])}

    if mapping["kind"] == "csv":
        # Materialize the CSV into an xlsx, preserving rows/cells 1:1 so the
        # detected (header_row, columns) still address the same cells.
        from openpyxl import Workbook
        grid = _read_csv_grid(file_bytes)
        wb = Workbook()
        ws = wb.active
        ws.title = "Generation"
        for row in grid:
            ws.append([("" if c is None else c) for c in row])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()
        mapping = dict(mapping)
        mapping["kind"] = "xlsx"
        mapping["sheet"] = "Generation"
    else:
        xlsx_bytes = file_bytes

    return {"ok": True, "mapping": mapping, "workbook": xlsx_bytes,
            "warnings": mapping.get("warnings", [])}
