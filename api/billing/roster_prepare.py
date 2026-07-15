"""High-power roster workbook preparation for offtaker bulk import.

Turns arbitrary operator spreadsheets into a clean rectangular grid the
column detector + bulk-import parser can trust:

  * multi-sheet pick (real roster tab, not Instructions/Sample/Summary)
  * encoding resilience (utf-8-sig, utf-16, cp1252, latin-1)
  * delimiter sniff (comma / tab / semicolon)
  * cell normalization (Excel float accounts, whitespace, NBSP)
  * multi-row header merge
  * section-banner array names (site title rows → fill-forward array column)
  * drop empty trailing columns / total-only noise hints for the detector

Pure + deterministic. Never raises on bad files — returns ok=False + warnings.
"""
from __future__ import annotations

import csv
import io
import re
from typing import Any, Optional

from .matcher import _s
from .offtaker_match import match_array

_MAGIC_XLSX = b"PK\x03\x04"
_TOTAL_ROW_RE = re.compile(
    r"^\s*(total|totals|subtotal|grand\s*total|sum|overall|aggregate|balance)\b",
    re.I,
)
_PHONE_RE = re.compile(
    r"^\+?1?[\s\-.]?\(?\d{3}\)?[\s\-.]?\d{3}[\s\-.]?\d{4}$"
)
_BAD_SHEET = (
    "sample", "template", "example", "instruction", "readme", "notes",
    "summary", "cover", "index", "toc", "pivot", "chart", "graph",
    "true up", "trueup", "trend", "dashboard", "howto", "how to",
)
_GOOD_SHEET = (
    "offtaker", "subscriber", "member", "roster", "customer", "allocation",
    "share", "participant", "billing", "data", "list", "export", "members",
)
_ROSTER_HINT = re.compile(
    r"\b(offtaker|subscriber|member|customer|allocation|share\s*%|ownership|"
    r"email|account\s*(#|number|no)|participant|beneficiary|bill\s*to)\b",
    re.I,
)


def _cell_str(v: Any) -> str:
    """Normalize one cell to a clean string for roster work."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        # Excel account numbers often arrive as 12345.0
        if v == int(v) and abs(v) < 1e15:
            return str(int(v))
        # Avoid scientific noise for mid-range floats
        s = f"{v:.12g}"
        return s
    s = str(v).replace("\u00a0", " ").replace("\r", " ").strip()
    # Excel sometimes stringifies floats with trailing .0
    if re.fullmatch(r"-?\d+\.0+", s):
        return s.split(".", 1)[0]
    return s


def _row_strs(row: Any) -> list[str]:
    if row is None:
        return []
    return [_cell_str(c) for c in row]


def _nonempty_cells(row: list[str]) -> list[tuple[int, str]]:
    return [(i, c) for i, c in enumerate(row) if c]


def _looks_total(row: list[str]) -> bool:
    for _, c in _nonempty_cells(row):
        return bool(_TOTAL_ROW_RE.search(c))
    return False


def _decode_text(raw: bytes) -> str:
    """Try common spreadsheet export encodings in order of likelihood."""
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16", errors="replace")
    if raw.startswith(b"\xef\xbb\xbf"):
        return raw.decode("utf-8-sig", errors="replace")
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _read_csv_grid(raw: bytes) -> list[list[str]]:
    text = _decode_text(raw)
    # Drop UTF-16 BOM leftovers / nulls from some Windows exports
    text = text.replace("\x00", "")
    sample = text[:8192]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except Exception:  # noqa: BLE001
        for d in ("\t", ";", "|", ","):
            if d in sample:
                delim = d
                break
    rows = list(csv.reader(io.StringIO(text), delimiter=delim))
    return [_row_strs(r) for r in rows]


def _sheet_penalty(name: str) -> float:
    nm = (name or "").lower()
    pen = 0.0
    if any(w in nm for w in _BAD_SHEET):
        pen -= 8.0
    if any(w in nm for w in _GOOD_SHEET):
        pen += 4.0
    return pen


def _roster_sheet_score(grid: list[list[str]], sheet_name: str = "") -> float:
    """How roster-like is this sheet? Higher = better pick for offtaker import."""
    if not grid:
        return -1.0
    score = _sheet_penalty(sheet_name)
    scan = grid[:20]
    hint_hits = 0
    emailish = 0
    pctish = 0
    data_rows = 0
    for ridx, row in enumerate(scan):
        joined = " ".join(c for c in row if c)
        if _ROSTER_HINT.search(joined):
            hint_hits += 1
        nonempty = _nonempty_cells(row)
        if not nonempty:
            continue
        if ridx > 0 and not _looks_total(row) and len(nonempty) >= 2:
            data_rows += 1
        for _, c in nonempty:
            if "@" in c and "." in c:
                emailish += 1
            if "%" in c or re.fullmatch(r"0?\.\d+", c) or re.fullmatch(r"\d{1,3}([.,]\d+)?", c):
                # weak percent/share signal
                try:
                    n = float(c.replace("%", "").replace(",", "."))
                    if 0 < n <= 100:
                        pctish += 1
                except ValueError:
                    pass
    score += hint_hits * 3.0
    score += min(emailish, 40) * 0.4
    score += min(pctish, 40) * 0.25
    score += min(data_rows, 80) * 0.15
    # Prefer wider tables (rosters have several columns)
    width = max((len(r) for r in scan), default=0)
    score += min(width, 12) * 0.2
    return score


def _read_xlsx_all(raw: bytes) -> list[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out: list[tuple[str, list[list[str]]]] = []
    try:
        for ws in wb.worksheets:
            grid: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                grid.append(_row_strs(row))
                if i > 5000:
                    break
            out.append((ws.title or f"Sheet{len(out)+1}", grid))
    finally:
        wb.close()
    return out


def _trim_grid(grid: list[list[str]]) -> list[list[str]]:
    """Drop trailing empty rows/cols so detection isn't distracted."""
    # trailing empty rows
    while grid and not _nonempty_cells(grid[-1]):
        grid = grid[:-1]
    if not grid:
        return grid
    # max used col
    max_c = 0
    for row in grid:
        for i, c in enumerate(row):
            if c:
                max_c = max(max_c, i + 1)
    return [row[:max_c] + [""] * max(0, max_c - len(row)) for row in grid]


def _guess_header_row(grid: list[list[str]], max_scan: int = 15) -> Optional[int]:
    """Pick the row with the most roster-ish header keywords (no detector import)."""
    best_i: Optional[int] = None
    best_sc = -1.0
    for i, row in enumerate(grid[:max_scan]):
        joined = " ".join(c for c in row if c).lower()
        if not joined.strip():
            continue
        sc = 0.0
        for kw, w in (
            ("offtaker", 3), ("subscriber", 3), ("customer", 2), ("member", 2),
            ("email", 3), ("share", 2), ("allocation", 3), ("account", 2),
            ("array", 2), ("site", 1.5), ("percent", 2), ("ownership", 2.5),
            ("name", 1.5), ("discount", 1.5), ("rate", 1),
        ):
            if kw in joined:
                sc += w
        # Prefer non-numeric rows
        ne = _nonempty_cells(row)
        if ne:
            nums = sum(1 for _, c in ne if re.fullmatch(r"-?\d+([.,]\d+)?%?", c))
            sc -= nums * 1.2
        if sc > best_sc:
            best_sc = sc
            best_i = i
    return best_i if best_sc > 0 else None


def _merge_multirow_headers(grid: list[list[str]], header_row: int) -> tuple[list[list[str]], list[str]]:
    """If the row under the header is also label-like, merge into one header row.

    Common in utility exports: row3='Subscriber' / row4='Name' → 'Subscriber Name'.
    """
    if header_row is None or header_row < 0 or header_row >= len(grid):
        return grid, []
    headers = list(grid[header_row])
    if header_row + 1 >= len(grid):
        return grid, headers
    nxt = grid[header_row + 1]
    # Sub-header row: mostly short LABEL text, almost no emails/numbers-as-data.
    # Must NOT swallow the first data row (e.g. "Hilltop Farm | Barn Co | 40").
    nonempty = _nonempty_cells(nxt)
    if len(nonempty) < 2:
        return grid, headers
    numericish = 0
    emailish = 0
    longish = 0
    labelish = 0
    label_kw = (
        "name", "email", "share", "%", "account", "site", "array", "contact",
        "interest", "allocation", "phone", "mobile", "rate", "discount", "id",
    )
    for _, c in nonempty:
        cl = c.lower()
        if "@" in c:
            emailish += 1
        if re.fullmatch(r"-?\d+([.,]\d+)?%?", c):
            numericish += 1
        if len(c) > 28:
            longish += 1
        if any(k in cl for k in label_kw) or len(c.split()) <= 2 and not re.search(r"\d{3,}", c):
            # short token without long account digits — weak label signal
            if any(k in cl for k in label_kw):
                labelish += 1
    n = max(len(nonempty), 1)
    # Any real data signal → never merge
    if emailish >= 1 or numericish >= 1 or longish / n > 0.5:
        return grid, headers
    # Need explicit label keywords on the sub-row (not just free-text names)
    if labelish < 2 and labelish / n < 0.5:
        return grid, headers
    # Merge
    width = max(len(headers), len(nxt))
    merged = []
    for i in range(width):
        a = headers[i] if i < len(headers) else ""
        b = nxt[i] if i < len(nxt) else ""
        if a and b and b.lower() not in a.lower():
            merged.append(f"{a} {b}".strip())
        else:
            merged.append(a or b)
    new_grid = grid[:header_row] + [merged] + grid[header_row + 2:]
    return new_grid, merged


def _section_expand(
    grid: list[list[str]],
    header_row: int,
    arrays: list[dict],
    utility_accounts: list[dict],
) -> tuple[list[list[str]], Optional[int], list[str]]:
    """Expand section-banner array names into a fill-forward array column.

    Layouts like:
        Maple Street Solar
        Alice, 25%, alice@x.com
        Bob, 25%, bob@x.com
        Hilltop Farm
        Carol, 50%, carol@x.com

    When there is no usable array column, invent one at the end and fill it.
    Returns (grid, array_col_index_or_None, warnings).
    """
    warnings: list[str] = []
    if header_row is None or header_row < 0 or header_row >= len(grid):
        return grid, None, warnings
    if not arrays and not utility_accounts:
        return grid, None, warnings

    data = grid[header_row + 1:]
    # Find rows that are single-cell (or first-cell-only) array banners
    banners: list[tuple[int, str]] = []  # absolute row index, array name
    for i, row in enumerate(data):
        abs_i = header_row + 1 + i
        if _looks_total(row):
            continue
        ne = _nonempty_cells(row)
        if len(ne) != 1:
            # also allow first cell filled and rest empty-ish with only weak noise
            if not ne:
                continue
            first_i, first_c = ne[0]
            if first_i != 0:
                continue
            if len(ne) > 2:
                continue
            # second cell might be a note — still try
        else:
            first_c = ne[0][1]
        try:
            m = match_array(first_c, arrays, utility_accounts)
        except Exception:  # noqa: BLE001
            continue
        if m.get("confidence") in ("exact", "high", "medium") and m.get("array_name"):
            banners.append((abs_i, m.get("array_name") or first_c))

    if len(banners) < 1:
        return grid, None, warnings

    # Need at least one data row under a banner that has multiple fields
    useful = False
    banner_set = {b[0] for b in banners}
    for i, row in enumerate(data):
        abs_i = header_row + 1 + i
        if abs_i in banner_set or _looks_total(row):
            continue
        if len(_nonempty_cells(row)) >= 2:
            useful = True
            break
    if not useful:
        return grid, None, warnings

    # Append synthetic array column
    width = max((len(r) for r in grid), default=0)
    arr_col = width
    new_grid: list[list[str]] = []
    for ridx, row in enumerate(grid):
        padded = list(row) + [""] * max(0, width - len(row))
        padded.append("")  # array col
        new_grid.append(padded)
    # Header label for synthetic col
    new_grid[header_row][arr_col] = "Array (from section)"

    current = ""
    banner_rows = {b[0]: b[1] for b in banners}
    for ridx in range(header_row + 1, len(new_grid)):
        if ridx in banner_rows:
            current = banner_rows[ridx]
            # blank the banner row's synthetic value — it's not an offtaker
            new_grid[ridx][arr_col] = ""
            # mark banner by clearing? keep name in col0 for review; parser skips low-field rows
            continue
        if current and len(_nonempty_cells(new_grid[ridx][:arr_col])) >= 2:
            new_grid[ridx][arr_col] = current

    warnings.append(
        f"Detected {len(banners)} section header(s) naming arrays — filled Array column from those banners."
    )
    return new_grid, arr_col, warnings


def prepare_roster_grid(
    file_bytes: bytes,
    filename: str = "",
    arrays: Optional[list[dict]] = None,
    utility_accounts: Optional[list[dict]] = None,
    preferred_sheet: Optional[str] = None,
) -> dict:
    """Open + normalize a roster workbook into a detector-ready grid.

    Returns:
      {
        ok, sheet, sheets:[{name,score}], grid: [[str...]], kind: csv|xlsx,
        warnings: [...], section_array_col: int|None
      }
    """
    arrays = arrays or []
    utility_accounts = utility_accounts or []
    warnings: list[str] = []
    name = (filename or "").lower()
    raw = file_bytes or b""
    if not raw:
        return {
            "ok": False, "sheet": None, "sheets": [], "grid": [],
            "kind": "empty", "warnings": ["Empty file."], "section_array_col": None,
        }

    is_xlsx = raw[:4] == _MAGIC_XLSX or name.endswith(".xlsx") or name.endswith(".xlsm")
    kind = "xlsx" if is_xlsx else "csv"
    candidates: list[tuple[str, list[list[str]]]] = []

    try:
        if is_xlsx:
            candidates = _read_xlsx_all(raw)
        else:
            candidates = [("csv", _read_csv_grid(raw))]
    except Exception as e:  # noqa: BLE001
        return {
            "ok": False, "sheet": None, "sheets": [], "grid": [],
            "kind": kind, "warnings": [f"Could not open file: {e}"],
            "section_array_col": None,
        }

    if not candidates:
        return {
            "ok": False, "sheet": None, "sheets": [], "grid": [],
            "kind": kind, "warnings": ["No sheets found."], "section_array_col": None,
        }

    scored = []
    for sname, grid in candidates:
        g = _trim_grid(grid)
        sc = _roster_sheet_score(g, sname)
        if preferred_sheet and sname.lower() == preferred_sheet.lower():
            sc += 20.0
        scored.append((sc, sname, g))
    scored.sort(key=lambda t: (-t[0], t[1]))
    best_score, best_name, best_grid = scored[0]
    sheet_meta = [{"name": n, "score": round(sc, 2)} for sc, n, _g in scored]

    if best_score < 0.5 and len(scored) > 1:
        warnings.append(
            f"Picked sheet '{best_name}' with low roster confidence — confirm columns carefully."
        )

    # Light header guess for multi-row merge + section expand (detector re-finds).
    header_row = _guess_header_row(best_grid)
    if header_row is None:
        header_row = 0
        warnings.append("Couldn't confidently find a header row during prepare; using row 1.")

    best_grid, _merged_headers = _merge_multirow_headers(best_grid, header_row)
    header_row2 = _guess_header_row(best_grid)
    if header_row2 is not None:
        header_row = header_row2

    best_grid, section_col, sec_warns = _section_expand(
        best_grid, header_row, arrays, utility_accounts,
    )
    warnings.extend(sec_warns)
    best_grid = _trim_grid(best_grid)

    return {
        "ok": bool(best_grid),
        "sheet": best_name if kind == "xlsx" else None,
        "sheets": sheet_meta,
        "grid": best_grid,
        "kind": kind,
        "warnings": warnings,
        "section_array_col": section_col,
        "header_row_hint": header_row,
    }


def grid_to_string_rows(grid: list[list[str]]) -> list[list[str]]:
    """Ensure every cell is str (bulk-import expects this)."""
    return [[("" if c is None else str(c)) for c in row] for row in (grid or [])]


def is_phone_like(s: str) -> bool:
    t = re.sub(r"[\s().\-]", "", s or "")
    if _PHONE_RE.match(s or "") or _PHONE_RE.match(t):
        return True
    # bare 10-digit US
    if re.fullmatch(r"\d{10}", t):
        return True
    return False
