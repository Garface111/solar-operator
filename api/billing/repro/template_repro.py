"""
Reproduce an offtaker's invoice IN THE OPERATOR'S OWN TEMPLATE — the research-
validated hybrid (deep-research wf_f378d415, Jun 2026): never convert the template
to HTML (the lossy anti-pattern); CODE writes the offtaker's values into the real
.xlsx, the real engine (Gotenberg) renders it pixel-perfect, a fail-closed guard
confirms the amount before trust.

WHY DIRECT-CELL-WRITE (not a ledger-row fill): an operator template is reused
across offtakers, but the Template invoice sheet's display cells hold the
template's OWN sample customer's numbers/terms (a flat "Amount Due", a baked-in
"Billing Rate: 88.5%"). Filling a Data-ledger row and re-pointing INDIRECT() does
NOT change those — it recomputes with the TEMPLATE's terms. Empirically that
rendered the sample's $2,150 for a $3,167 offtaker (the guard caught it). So we
LOCATE the dynamic display cells (matcher._build_token_map — the same tokenizer
the HTML path uses) and write THIS offtaker's computed values straight into them,
then swap the sample customer's name for the offtaker's. Values are computed in
Python; the cell's own number format renders them.

reproduce_in_template is FAIL-CLOSED on both amount (verify guard) and identity
(the offtaker name must render, else refuse). Mapping is deterministic for the
HCT family (no LLM); ai_field_map is the novel-template fallback. NOT wired into
the live send path until verified.
"""
from __future__ import annotations

import datetime as _dt
import io
import logging
from typing import Optional

from .pipeline import ReproResult, reproduce_invoice

log = logging.getLogger(__name__)

# Token (from matcher._build_token_map) → key in the offtaker values dict.
_TOKEN_KEYS = ("amount_due", "kwh", "solar_value", "billed_value", "solar_savings",
               "period_start", "period_end", "invoice_number", "invoice_date", "due_date")


def offtaker_values_from_match(match, *, invoice_date: Optional[_dt.date] = None,
                               due_days: int = 28) -> dict:
    """Raw (un-formatted) values to write into the template's mapped cells, taken
    from the offtaker's own computed invoice. The cell's number format renders them."""
    ci = match.computed_invoice or {}
    lp = match.latest_period
    inv_d = invoice_date or _dt.date.today()
    vals = {
        "amount_due": ci.get("amount_owed"),
        "kwh": ci.get("kwh", getattr(lp, "customer_kwh", None)),
        "solar_value": ci.get("solar_value"),
        "billed_value": ci.get("billed_value"),
        "solar_savings": ci.get("solar_savings"),
        "period_start": getattr(lp, "start", None),
        "period_end": getattr(lp, "end", None),
        "invoice_number": ci.get("invoice_number"),
        "invoice_date": inv_d,
        "due_date": inv_d + _dt.timedelta(days=due_days),
    }
    return {k: v for k, v in vals.items() if v is not None}


def build_template_cell_map(template_bytes: bytes) -> Optional[dict]:
    """Cacheable mapping of the operator template's invoice sheet: which dynamic
    display cells hold which field, plus the template's sample customer name (so we
    can swap it). Deterministic via the matcher for the HCT family. Returns
    {sheet, cells:{(r,c):token}, sample_name} or None if the template can't be read."""
    from openpyxl import load_workbook
    from ..matcher import (find_invoice_sheet, _content_bounds, _build_token_map,
                           match_billing_workbook)
    try:
        wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
    except Exception as e:  # noqa: BLE001
        log.warning("build_template_cell_map: cannot open template: %s", e)
        return None
    try:
        ws = find_invoice_sheet(wb) or (wb.worksheets[0] if wb.worksheets else None)
        if ws is None:
            return None
        r0, r1, c0, c1 = _content_bounds(ws, 70, 16)
        covered: set = set()
        for m in ws.merged_cells.ranges:
            for r in range(m.min_row, m.max_row + 1):
                for c in range(m.min_col, m.max_col + 1):
                    if (r, c) != (m.min_row, m.min_col):
                        covered.add((r, c))
        tok = _build_token_map(ws, r0, r1, c0, c1, covered)
        cells = {rc: t.strip("{} ").split()[0] for rc, t in tok.items()}  # "{{ x }}" → "x"
    finally:
        try:
            wb.close()
        except Exception:
            pass
    sample_name = None
    try:
        sample_name = (match_billing_workbook(template_bytes, allow_llm=False).customer or {}).get("name")
    except Exception:  # noqa: BLE001
        pass
    return {"sheet": ws.title, "cells": cells, "sample_name": sample_name}


def _fill_template_cells(template_bytes: bytes, cell_map: dict, values: dict,
                         customer_name: str) -> Optional[bytes]:
    """Write `values` into the mapped display cells of the template's invoice sheet
    and swap the sample customer's name for `customer_name`. Returns new xlsx bytes
    (formulas/styles elsewhere untouched), or None on failure."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(template_bytes))          # keep formulas
    except Exception as e:  # noqa: BLE001
        log.warning("_fill_template_cells: cannot open template: %s", e)
        return None
    sheet = cell_map.get("sheet")
    ws = wb[sheet] if sheet in wb.sheetnames else (wb.worksheets[0] if wb.worksheets else None)
    if ws is None:
        return None
    # 1) Write this offtaker's value into each mapped display cell. Mapped cells are
    #    OFFTAKER-SPECIFIC (amount, kWh, dates…), so a field that's absent from values
    #    is BLANKED, never left showing the template sample's number (values.get→None).
    for (r, c), token in (cell_map.get("cells") or {}).items():
        ws.cell(row=r, column=c).value = values.get(token)
    # 2) Swap the sample customer's name for the offtaker's. Exact-match cells become
    #    the offtaker name unconditionally; a cell that merely CONTAINS the name (e.g.
    #    "Valley Village (Valley Cares, Inc)") is rewritten ONLY when the name is long
    #    enough AND dominates the cell — so a short/embedded token can't clobber an
    #    unrelated label (e.g. "Sun" inside "HCT Sun Enterprises").
    sample = (cell_map.get("sample_name") or "").strip()
    # Match the name AND a punctuation-stripped core ("Valley Cares, Inc." vs the
    # bill-to's "(Valley Cares, Inc)") so a trailing . or ) can't defeat the swap.
    variants = [v for v in (sample, sample.rstrip(" .,;:)(")) if len(v) >= 3]
    wrote_name = False
    if variants:
        for sh in wb.worksheets:
            for row in sh.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    val, stripped = cell.value, cell.value.strip()
                    if any(stripped == v for v in variants):           # whole cell IS the name
                        cell.value = customer_name
                        wrote_name = True
                        continue
                    for v in variants:                                  # embedded, name-dominant only
                        if len(v) >= 6 and v in val and len(v) / max(len(stripped), 1) >= 0.5:
                            val = val.replace(v, customer_name)
                    if val != cell.value:
                        cell.value = val
                        wrote_name = True
    # 3) Auto-fit columns so the renderer doesn't clip overflow text (LibreOffice
    #    font metrics ≠ Excel's; a neighbor value blocks the overflow).
    _autofit_columns(ws)
    # 4) Isolate to the invoice sheet: a multi-sheet upload (Data ledger, Trends, a
    #    SAMPLE tab…) must NOT render its other sheets into the invoice — and an
    #    offtaker must never receive the operator's raw ledger. Flatten the invoice
    #    sheet's formulas to the file's own cached values (so it stays self-contained
    #    once the sheets it referenced are gone), then drop every other sheet.
    _isolate_to_invoice_sheet(wb, ws, template_bytes)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _references_other_sheet(formula: str, others: list) -> bool:
    """True only if `formula` has a GENUINE cross-sheet reference to one of `others`
    — `Sheet!…` or quoted `'Sheet'!…` with Excel's apostrophe-doubling — not a mere
    title substring (which would false-blank an intra-sheet cell) and catching the
    quoted/apostrophe form (which a bare-title test would miss → #REF!)."""
    up = formula.upper()
    for t in others:
        if (t.upper() + "!") in up:                       # Data!A1
            return True
        if ("'" + t.replace("'", "''").upper() + "'!") in up:   # 'Bob''s Data'!A1
            return True
    return False


# Fonts the headless renderer (LibreOffice) ships with or maps 1:1. A cell font NOT
# in this set is likely substituted with a WIDER fallback, so a date/number sized for
# the original font overflows to '###'. We widen those more aggressively (see below).
_SAFE_FONTS = {
    "", "calibri", "calibri light", "arial", "arial narrow", "times new roman",
    "helvetica", "verdana", "tahoma", "courier new", "georgia", "cambria",
    "trebuchet ms", "century gothic", "garamond", "book antiqua",
    "liberation sans", "liberation serif", "liberation mono", "dejavu sans",
    "dejavu serif", "carlito", "caladea",
}


def _autofit_columns(ws, factor: float = 0.08, cap: float = 72.0) -> None:
    """Widen columns so text the renderer would CLIP becomes visible — Excel overflows
    a long label across empty neighbors, but LibreOffice's font metrics differ (esp.
    for a font it lacks, e.g. the template's 'Chalkboard'), and a value in a neighbor
    cell blocks the overflow and truncates the label (…'rate o' instead of '…rate of
    $2,100.00'). For each text cell, estimate its rendered width; if that exceeds its
    own column PLUS the empty columns it can overflow into (until a non-empty cell),
    widen its column by the shortfall. Never shrinks; capped so a long free-text line
    that legitimately overflows the page can't bloat a column. Also cures '#####'
    (a number too wide for its column)."""
    import re
    from openpyxl.utils import get_column_letter

    def _fmt_is_dateish(nf: str) -> bool:
        """True for a date/time number format (renders like 12/31/2026). Dates need a
        FIXED width reservation: unlike text, a too-narrow date shows '###', it does
        not overflow into the next cell — which is exactly the bug this cures."""
        if not nf or nf == "General":
            return False
        s = re.sub(r'\[[^\]]*\]|"[^"]*"', "", nf)     # drop [color]/[$-409] + quoted literals
        return bool(re.search(r"[yd]", s, re.I) or "m" in s.lower()) and not re.search(r"[#0]", s)

    mr = min(ws.max_row or 1, 80)
    mc = min(ws.max_column or 1, 20)

    def cw(ci):
        d = ws.column_dimensions.get(get_column_letter(ci))
        return d.width if d and d.width else 8.43

    def occupied(r, ci):
        v = ws.cell(row=r, column=ci).value
        return v is not None and str(v).strip() != ""

    def disp_len(v):
        if isinstance(v, bool):
            return 3
        if isinstance(v, (_dt.datetime, _dt.date)):
            return 12                                            # a rendered date, not str()'s 19
        if isinstance(v, (int, float)):
            return max(len(str(v)), len(f"{abs(v):,.2f}") + 2)   # account for $ , . formatting
        return len(str(v))

    need: dict = {}
    for r in range(1, mr + 1):
        for ci in range(1, mc + 1):
            cell = ws.cell(row=r, column=ci)
            v = cell.value
            if v is None or (isinstance(v, str) and not v.strip()):
                continue
            nf = cell.number_format or ""
            is_formula = isinstance(v, str) and v.startswith("=")
            is_date = _fmt_is_dateish(nf)
            # Dates + numbers can't spill into an empty neighbor the way text does —
            # too narrow, they render as '###'. So size their OWN column to fit and
            # don't credit neighbor width. Text keeps the overflow allowance. Formula
            # cells are sized off their number format (the value isn't computed yet).
            if is_date:
                chars, numeric_like = 12, True               # e.g. 12/31/2026
            elif is_formula:
                if re.search(r"[#0]", nf):
                    chars, numeric_like = 13, True            # a number-producing formula
                else:
                    continue                                  # text formula — can't size it
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                chars, numeric_like = disp_len(v), True
            else:
                chars, numeric_like = disp_len(v), False
            size = (cell.font.size or 11) if cell.font else 11
            if numeric_like:
                fname = ((cell.font.name if cell.font else "") or "").strip().lower()
                # A font the renderer lacks (e.g. the template's 'Chalkboard') is
                # substituted with a WIDER one, and a DATE can't spill into a neighbor
                # so it shows '###'. Give substitution-risk date cells a generous
                # allowance. Numbers fit far more easily (short, e.g. '789.64') and over-
                # widening them only spreads the layout — they keep the tuned factor.
                ff = factor * (2.2 if (is_date and fname not in _SAFE_FONTS) else 1.0)
            else:
                ff = factor
            est = chars * size * ff
            avail = cw(ci)
            if not numeric_like:
                cc = ci + 1
                while cc <= mc and not occupied(r, cc):
                    avail += cw(cc)
                    cc += 1
            if est > avail:
                need[ci] = max(need.get(ci, 0.0), cw(ci) + (est - avail))
    for ci, w in need.items():
        ws.column_dimensions[get_column_letter(ci)].width = min(w, cap)


def _isolate_to_invoice_sheet(wb, ws, original_bytes: bytes) -> None:
    """Render exactly ONE invoice (the `ws` sheet) — not the operator's whole
    workbook — WITHOUT breaking the cross-tab references some templates use.

    Paul's templates vary: the self-contained ones flatten cleanly, but others pull
    straight from a data tab (e.g. `=+NFD!G6`). The old approach DELETED every other
    sheet, which turned those cross-tab formulas into #REF! (or, when we pre-nulled
    them, silently dropped the data). Instead we KEEP the other sheets but HIDE them:
    LibreOffice excludes hidden sheets from the PDF, yet still resolves formulas that
    reference them — verified on the prod Gotenberg renderer. We still flatten cells
    that carry an Excel-cached result (frozen, consistent); any cell without a cache
    keeps its formula and resolves live against the hidden sheets."""
    from openpyxl import load_workbook
    inv_title = ws.title
    others = [s.title for s in wb.worksheets if s.title != inv_title]
    if not others:
        ws.sheet_state = "visible"
        return
    invd = None
    try:
        wbd = load_workbook(io.BytesIO(original_bytes), data_only=True)
        if inv_title in wbd.sheetnames:
            invd = wbd[inv_title]
    except Exception as e:  # noqa: BLE001
        log.warning("_isolate: data_only load failed (%s); keeping formulas", e)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            cached = invd[cell.coordinate].value if invd is not None else None
            if cached is not None:
                cell.value = cached
            # else: leave the formula — cross-sheet refs resolve against the hidden
            # sheets below; intra-sheet ones LibreOffice recalculates. (No more
            # nulling, which was the source of Paul's broken/blank invoice cells.)
    # Now decide what to do with the OTHER tabs. After the flatten above, check whether
    # the invoice sheet STILL has any formula that genuinely references another tab.
    #   • None do  → DROP every other tab. Nothing points at them, so keeping them only
    #     risks the render: openpyxl re-saves a tab's chart/drawing XML (e.g. a
    #     'TTM Trendline' chart sheet) into markup Gotenberg's LibreOffice can REJECT,
    #     which fails the whole convert → reproduce_*_template returns None → the caller
    #     falls back to rendering the RAW upload (the template's own sample customer,
    #     unswapped, with '###' date columns). That was the regression from KEEPING all
    #     tabs for cross-tab refs: previously they were deleted, so no chart XML ever
    #     reached the renderer.
    #   • Some do  → KEEP all other tabs (hidden) so those refs still resolve, but STRIP
    #     their charts/images (never rendered from a hidden sheet anyway) so the same
    #     chart XML can't break the save, and clear their print areas so a hidden second
    #     invoice sheet can't export a duplicate page.
    inv_refs_others = any(
        isinstance(c.value, str) and c.value.startswith("=")
        and _references_other_sheet(c.value, others)
        for row in ws.iter_rows() for c in row
    )
    for name in others:
        try:
            if not inv_refs_others:
                del wb[name]
                continue
            osh = wb[name]
            osh.sheet_state = "hidden"
            osh.print_area = None
            osh._charts = []
            osh._images = []
        except Exception:  # noqa: BLE001
            pass
    ws.sheet_state = "visible"
    try:
        wb.active = wb.sheetnames.index(inv_title)
    except Exception:  # noqa: BLE001
        pass


def _detect_sample_leak(pdf_bytes: bytes, sample_values: dict, offtaker_values: dict):
    """A template-SAMPLE offtaker-specific number we meant to replace must not survive
    on the render. Returns a description if one does (and it isn't also the offtaker's
    own value), else None. Only the MAPPED (offtaker-specific) fields are checked —
    shared array-level figures (net rate, array meter) aren't in this set, so a
    legitimately-shared number is never flagged. This catches the leak the Amount-Due
    guard is blind to: an unmapped/duplicate cell frozen to the sample's $/kWh."""
    from .verify import _pdf_lines, _parse_signed
    lines = _pdf_lines(pdf_bytes)
    if not lines:
        return None
    page = {round(n, 2) for n in _parse_signed("\n".join(lines))}
    off = {round(float(v), 2) for v in offtaker_values.values() if isinstance(v, (int, float))}
    for token, sv in (sample_values or {}).items():
        if not isinstance(sv, (int, float)):
            continue
        s = round(float(sv), 2)
        if abs(s) < 0.005:
            continue
        ov = offtaker_values.get(token)
        if isinstance(ov, (int, float)) and round(float(ov), 2) == s:
            continue                                   # offtaker's value coincides — fine
        if s in page and s not in off:
            return f"sample '{token}' value {sv} still on the render"
    return None


def reproduce_in_template(template_bytes: bytes, *, offtaker_match,
                          customer_name: str,
                          cell_map: Optional[dict] = None,
                          invoice_date: Optional[_dt.date] = None,
                          verify: bool = False) -> Optional[ReproResult]:
    """Reproduce ONE offtaker's invoice in the operator's template, pixel-perfect.

    Writes the offtaker's computed values into the template's mapped display cells
    + swaps in their name, renders, and gates fail-closed on BOTH the amount (verify
    guard) and identity (the name must render). Returns None (caller falls back to
    the standard invoice) when the template can't be mapped, has no Amount-Due cell,
    or the result can't be verified — never ships a wrong amount or wrong name."""
    from .verify import _pdf_lines

    cm = cell_map or build_template_cell_map(template_bytes)
    if not cm or "amount_due" not in set((cm.get("cells") or {}).values()):
        log.info("reproduce_in_template: no mapped Amount Due cell — refusing")
        return None
    values = offtaker_values_from_match(offtaker_match, invoice_date=invoice_date)
    expected = values.get("amount_due")

    def fill(_field_map_override):
        filled = _fill_template_cells(template_bytes, cm, values, customer_name)
        if filled is None:
            raise RuntimeError("template cell fill failed")
        return filled

    res = reproduce_invoice(fill, expected_amount=expected, verify=verify)

    # Identity guard: the offtaker's name MUST appear on the render.
    if res and res.pdf:
        lines = _pdf_lines(res.pdf)
        nm = (customer_name or "").strip().lower()
        if nm and (lines is None or not any(nm in ln.lower() for ln in lines)):
            log.warning("reproduce_in_template: name %r not on render — refusing", customer_name)
            return None
        # Sample-leak guard: no template-sample offtaker number may survive (the
        # amount guard only checks Amount Due — this catches a stale duplicate/second
        # $ or kWh cell). On a leak, fall back to the standard invoice.
        leak = _detect_sample_leak(res.pdf, _template_self_values(template_bytes, cm), values)
        if leak:
            log.warning("reproduce_in_template: %s — refusing (fall back to standard)", leak)
            return None
        # Auto-center: crop to content + uniform margins so the invoice isn't left
        # wherever the template's print area landed it. After the guards (text is
        # unchanged, so they still hold).
        from .render import center_pdf_to_content
        res.pdf = center_pdf_to_content(res.pdf)
    return res


def _template_self_values(template_bytes: bytes, cell_map: dict) -> dict:
    """The template's OWN (internally-coherent) values at each mapped cell, read
    data_only so formula cells yield their cached result. Used to echo a faithful,
    self-consistent sample through the reproduction pipeline for the settings preview."""
    from openpyxl import load_workbook
    out: dict = {}
    try:
        wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
    except Exception:  # noqa: BLE001
        return out
    sheet = cell_map.get("sheet")
    ws = wb[sheet] if sheet in wb.sheetnames else (wb.worksheets[0] if wb.worksheets else None)
    if ws is not None:
        for (r, c), token in (cell_map.get("cells") or {}).items():
            v = ws.cell(row=r, column=c).value
            if v is not None:
                out[token] = v
    try:
        wb.close()
    except Exception:
        pass
    return out


def reproduce_template_preview(template_bytes: bytes, *,
                               sample_name: str = "Sample Offtaker") -> Optional[bytes]:
    """OUR reproduction of the operator's template for the settings preview — the
    SAME direct-cell-write pipeline a real send uses, so the operator sees exactly
    what our engine produces (and can spot any infidelity), not their raw upload.

    Echoes the template's own coherent numbers (consistent with its rate labels) and
    swaps in a clearly-sample bill-to so the pane is visibly our reconstruction.
    Returns rendered PDF bytes, or None to fall back to a plain render."""
    from .render import render_office_to_pdf, renderer_available, center_pdf_to_content
    if not renderer_available():
        return None
    cm = build_template_cell_map(template_bytes)
    if not cm:
        return None
    values = _template_self_values(template_bytes, cm)
    filled = _fill_template_cells(template_bytes, cm, values, sample_name)
    if filled is None:
        return None
    try:
        return center_pdf_to_content(render_office_to_pdf(filled, "reproduction_preview.xlsx"))
    except Exception as e:  # noqa: BLE001
        log.warning("reproduce_template_preview render failed: %s", e)
        return None
