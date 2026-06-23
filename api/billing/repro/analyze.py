"""
AI analysis — the "thoughtful" step: figure out WHERE the period data goes.

The deterministic matcher (api/billing/matcher) recognizes the HCT workbook
family by header tokens and yields a field_map (month/start/end/array_kwh/…
→ column index) that invoice_writer fills. That covers the known family well.
When a NEW or oddly-shaped workbook scores low, this asks Claude to read the
sheet and return the same field_map shape — so the fill still targets the right
cells instead of failing.

Output is a STRICT JSON field_map (0-indexed column numbers) + a confidence and
per-field notes, validated against FIELD_MAP_SCHEMA. The AI never writes the
invoice; it only locates columns. Falls back to the heuristic map (returns None)
whenever no API key is configured or the call fails — never hard-fails.
"""
from __future__ import annotations

import logging
from typing import Optional

from .llm import call_json, llm_available

log = logging.getLogger(__name__)

# The roles invoice_writer understands (api/billing/invoice_writer._write_ledger_row).
_ROLES = ["month", "start", "end", "array_kwh", "customer_kwh",
          "tariff", "adder", "value", "bill", "savings"]

FIELD_MAP_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "data_sheet": {"type": "string"},
        "header_row": {"type": "integer"},
        "field_map": {
            "type": "object",
            "additionalProperties": False,
            "properties": {r: {"type": "integer"} for r in _ROLES},
        },
        "confidence": {"type": "number"},
        "notes": {"type": "string"},
    },
    "required": ["data_sheet", "field_map", "confidence"],
}

_SYSTEM = (
    "You map the DATA-LEDGER columns of a solar-billing spreadsheet so a program "
    "can append one month's row. You are given a plain-text dump of each sheet "
    "(rows are 0-indexed; cells are tab-separated with their column index). Return "
    "the data-ledger sheet name, its header row, and field_map: role -> 0-indexed "
    "column. Roles: month, start (period start date), end (period end date), "
    "array_kwh (whole-array kWh), customer_kwh (this offtaker's kWh, may be absent), "
    "tariff, adder, value, bill, savings. Omit a role you can't find. The ledger is "
    "the sheet with monthly rows of kWh — NOT the invoice 'Template' sheet. Be "
    "conservative: only map a column you're confident about."
)


def _sheets_to_text(file_bytes: bytes, max_rows: int = 25, max_cols: int = 20) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(__import__("io").BytesIO(file_bytes), read_only=True, data_only=True)
    out = []
    try:
        for ws in wb.worksheets:
            out.append(f"# SHEET: {ws.title}")
            for r, row in enumerate(ws.iter_rows(values_only=True)):
                if r >= max_rows:
                    break
                cells = [f"{c}:{('' if v is None else str(v))[:24]}"
                         for c, v in enumerate(row[:max_cols])]
                out.append(f"r{r}\t" + "\t".join(cells))
            out.append("")
    finally:
        wb.close()
    return "\n".join(out)


def ai_field_map(file_bytes: bytes) -> Optional[dict]:
    """Ask Claude for the ledger field_map. Returns the validated dict, or None
    when the AI step is unavailable/failed (caller keeps the heuristic map)."""
    if not llm_available():
        return None
    try:
        result = call_json(
            system=_SYSTEM,
            user_text=_sheets_to_text(file_bytes),
            schema=FIELD_MAP_SCHEMA,
            max_tokens=1024,
        )
    except Exception as e:  # noqa: BLE001
        log.warning("ai_field_map failed: %s", e)
        return None
    fm = {k: int(v) for k, v in (result.get("field_map") or {}).items() if k in _ROLES}
    if "month" not in fm:
        log.info("ai_field_map: no month column found; ignoring")
        return None
    result["field_map"] = fm
    return result
