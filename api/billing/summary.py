"""
Performance summary — the second half of the "both" report Ford asked for.

Where the invoice says "here's what you owe", the summary says "here's how your
array performed": production this period vs. the same period a year ago, the
trailing-twelve-month total, and lifetime generation + savings. All derived from
the workbook's own ledger (no telemetry needed); when the customer's array is
telemetry-connected the caller can pass `peer` health to enrich it.

  build_summary(match)        → plain dict of the computed metrics
  render_summary_xlsx(match)  → one-sheet .xlsx
  render_summary_pdf(match)   → one-page PDF (reportlab)
"""
from __future__ import annotations

import pathlib
from collections import defaultdict
from typing import Optional

from .matcher import BillingMatch, Period

# Stable month abbreviations (locale-independent — never use calendar's, which
# follows the process locale and would drift the JSON shape across machines).
_MONTH_ABBR = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _same_month_prior_year(periods: list[Period], latest: Period) -> Optional[Period]:
    if not latest.end:
        return None
    target_year = latest.end.year - 1
    best = None
    for p in periods:
        if p.end and p.end.year == target_year and p.end.month == latest.end.month:
            best = p
    return best


def build_summary(match: BillingMatch, peer: Optional[dict] = None) -> dict:
    periods = [p for p in match.periods if (p.customer_kwh or p.array_kwh)]
    latest = match.latest_period
    out: dict = {
        "customer_name": match.customer.get("name"),
        "lifetime_kwh": match.project_totals.get("total_customer_kwh"),
        "lifetime_savings": match.project_totals.get("total_savings"),
        "period_count": match.project_totals.get("period_count"),
        "this_period_kwh": (latest.customer_kwh if latest else None),
        "this_period_month": (latest.month if latest else None),
        "this_period_end": (latest.end.isoformat() if latest and latest.end else None),
    }
    # Year-over-year for the latest period.
    if latest:
        prior = _same_month_prior_year(periods, latest)
        if prior and prior.customer_kwh and latest.customer_kwh:
            delta = latest.customer_kwh - prior.customer_kwh
            out["yoy_prior_kwh"] = prior.customer_kwh
            out["yoy_delta_kwh"] = round(delta, 1)
            out["yoy_delta_pct"] = round(100 * delta / prior.customer_kwh, 1)
    # Trailing twelve months (last 12 dated periods).
    dated = sorted([p for p in periods if p.end], key=lambda p: p.end)  # type: ignore[arg-type]
    ttm = dated[-12:]
    out["ttm_kwh"] = round(sum(p.customer_kwh or 0 for p in ttm), 1)
    out["ttm_savings"] = round(sum(p.savings or 0 for p in ttm), 2)
    out["ttm_points"] = [
        {"month": p.month, "end": p.end.isoformat() if p.end else None,
         "kwh": p.customer_kwh}
        for p in ttm
    ]
    if peer:
        out["peer"] = peer
    return out


def _empty_trends(customer_name: Optional[str]) -> dict:
    """The CONTRACT-1 shape for a subscription with no usable history. Returned
    verbatim on thin/empty data so the endpoint is a 200, never a 500."""
    return {
        "customer_name": customer_name,
        "years": [],
        "monthly_by_year": {},
        "seasonal_yoy": [],
        "ttm_kwh": None,
        "ttm_savings": None,
        "lifetime_kwh": None,
        "summary_note": None,
    }


def build_trends(match: BillingMatch) -> dict:
    """Multi-year billing trends for the macro trends tab (CONTRACT 1).

    Pure + derived entirely from the workbook ledger — the same source as
    build_summary, no telemetry. Groups customer_kwh (and savings) by calendar
    (year, month) so the frontend can overlay one line per year and read
    seasonal year-over-year growth at a glance.

      monthly_by_year: per year, the months that have data → {month, kwh, savings}.
      seasonal_yoy:    per calendar month present, each year's kWh + the latest
                       year's % change vs the immediately prior year (null if
                       that prior year has no value for the month).
      ttm/lifetime:    reused verbatim from build_summary.

    Thin/empty workbook → empty collections + null scalars (never raises).
    """
    out = _empty_trends(match.customer.get("name"))

    # Only periods with a real calendar date AND generation can be placed on the
    # year×month grid. Undated rows can't be attributed to a season.
    dated = [p for p in match.periods if p.end and (p.customer_kwh or p.array_kwh)]
    if not dated:
        return out

    # Sum customer_kwh + savings per (year, month) — defensive against a ledger
    # that splits a calendar month across two rows.
    agg: dict[tuple[int, int], dict] = defaultdict(lambda: {"kwh": 0.0, "savings": 0.0})
    for p in dated:
        cell = agg[(p.end.year, p.end.month)]
        cell["kwh"] += p.customer_kwh or 0.0
        cell["savings"] += p.savings or 0.0

    years = sorted({y for (y, _m) in agg})
    out["years"] = years
    out["monthly_by_year"] = {
        str(y): [
            {"month": m,
             "kwh": round(agg[(y, m)]["kwh"], 1),
             "savings": round(agg[(y, m)]["savings"], 2)}
            for m in range(1, 13) if (y, m) in agg
        ]
        for y in years
    }

    seasonal: list[dict] = []
    for m in sorted({mo for (_y, mo) in agg}):
        by_year = {str(y): round(agg[(y, m)]["kwh"], 1) for y in years if (y, m) in agg}
        present = sorted(int(y) for y in by_year)
        latest_y = present[-1]
        prior_v = by_year.get(str(latest_y - 1))   # immediately prior calendar year
        latest_v = by_year[str(latest_y)]
        delta_pct = (round(100 * (latest_v - prior_v) / prior_v, 1)
                     if prior_v else None)
        seasonal.append({
            "month": m,
            "label": _MONTH_ABBR[m],
            "by_year": by_year,
            "latest_delta_pct": delta_pct,
        })
    out["seasonal_yoy"] = seasonal

    # Trailing-12-month + lifetime totals are already correct in build_summary;
    # reuse them rather than recomputing (single source of truth).
    summary = build_summary(match)
    out["ttm_kwh"] = summary.get("ttm_kwh")
    out["ttm_savings"] = summary.get("ttm_savings")
    out["lifetime_kwh"] = summary.get("lifetime_kwh")
    out["summary_note"] = (
        f"{len(years)} years of billing history on record."
        if len(years) != 1 else "1 year of billing history on record."
    )
    return out


def render_summary_xlsx(match: BillingMatch, out_path: pathlib.Path,
                        peer: Optional[dict] = None) -> pathlib.Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    s = build_summary(match, peer)
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    bold = Font(bold=True, size=13)
    ws["A1"] = f"Performance summary — {s.get('customer_name') or 'Array'}"
    ws["A1"].font = bold

    rows = [
        ("This period (kWh)", s.get("this_period_kwh")),
        ("This period", s.get("this_period_month")),
        ("Year-over-year change (kWh)", s.get("yoy_delta_kwh")),
        ("Year-over-year change (%)", s.get("yoy_delta_pct")),
        ("Trailing 12 months (kWh)", s.get("ttm_kwh")),
        ("Trailing 12 months savings", s.get("ttm_savings")),
        ("Lifetime generation (kWh)", s.get("lifetime_kwh")),
        ("Lifetime savings ($)", s.get("lifetime_savings")),
        ("Billing periods on record", s.get("period_count")),
    ]
    r = 3
    for label, val in rows:
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Trailing 12 months").font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=1, value="Month")
    ws.cell(row=r, column=2, value="kWh")
    r += 1
    for pt in s.get("ttm_points", []):
        ws.cell(row=r, column=1, value=pt.get("month") or pt.get("end"))
        ws.cell(row=r, column=2, value=pt.get("kwh"))
        r += 1

    out_path = pathlib.Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def render_summary_pdf(match: BillingMatch, out_path: pathlib.Path,
                       peer: Optional[dict] = None) -> pathlib.Path:
    """Slick, on-brand performance-summary PDF — same energy hero band + bar
    chart as the invoice, with stat cards instead of a payable table."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("reportlab is required for PDF summaries") from e

    from . import _pdf_brand as brand

    s = build_summary(match, peer)
    out_path = pathlib.Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    HERO_H = 1.55 * inch
    INKDK = colors.HexColor(brand.INKDK)
    MUTEDDK = colors.HexColor(brand.MUTEDDK)
    GREEN_DK = colors.HexColor(brand.GREEN_DK)
    cust = s.get("customer_name") or "Array"

    def fmt(v, money=False, pct=False):
        if v is None:
            return "—"
        if money:
            return f"${v:,.2f}"
        if pct:
            return f"{v:+.1f}%"
        return f"{v:,.0f}" if isinstance(v, (int, float)) else str(v)

    # Hero: lifetime generation as the headline figure.
    life = s.get("lifetime_kwh")
    decorate = brand.make_hero_decorator(
        title="Performance Summary", subtitle=cust,
        right_label="LIFETIME GENERATION",
        right_value=(f"{life:,.0f} kWh" if life is not None else ""),
        footer_right="Performance summary", hero_h=HERO_H, light=True)

    styles = getSampleStyleSheet()
    lbl = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=11,
                         textColor=INKDK, fontName="Helvetica-Bold")
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8.5,
                           textColor=MUTEDDK)

    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        topMargin=HERO_H + 0.35 * inch, bottomMargin=0.8 * inch,
        leftMargin=0.85 * inch, rightMargin=0.85 * inch)
    story = []

    # ---- stat cards (2×3 grid) -----------------------------------------------
    yoy = (f"{fmt(s.get('yoy_delta_kwh'))} kWh  ({fmt(s.get('yoy_delta_pct'), pct=True)})"
           if s.get("yoy_delta_kwh") is not None else "—")
    cards = [
        ("THIS PERIOD", f"{fmt(s.get('this_period_kwh'))} kWh",
         s.get("this_period_month") or ""),
        ("YEAR OVER YEAR", yoy, "vs. same month last year"),
        ("TRAILING 12 MONTHS", f"{fmt(s.get('ttm_kwh'))} kWh",
         f"{fmt(s.get('ttm_savings'), money=True)} saved"),
        ("LIFETIME GENERATION", f"{fmt(s.get('lifetime_kwh'))} kWh",
         f"{s.get('period_count') or '—'} periods on record"),
        ("LIFETIME SAVINGS", fmt(s.get("lifetime_savings"), money=True),
         "total solar savings"),
        ("PEER HEALTH",
         str((s.get("peer") or {}).get("status") or "—") if s.get("peer") else "—",
         "telemetry-measured" if s.get("peer") else "connect telemetry"),
    ]

    def _card(c):
        eyebrow, value, sub = c
        inner = Table([[eyebrow], [value], [sub]], colWidths=[3.05 * inch])
        inner.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (0, 0), 7.5),
            ("TEXTCOLOR", (0, 0), (0, 0), MUTEDDK),
            ("BOTTOMPADDING", (0, 0), (0, 0), 7),    # gap under the eyebrow
            ("FONTNAME", (0, 1), (0, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 1), (0, 1), 16),
            ("LEADING", (0, 1), (0, 1), 19),
            ("TEXTCOLOR", (0, 1), (0, 1), GREEN_DK),
            ("BOTTOMPADDING", (0, 1), (0, 1), 6),    # gap between value and sub
            ("FONTSIZE", (0, 2), (0, 2), 8),
            ("TEXTCOLOR", (0, 2), (0, 2), MUTEDDK),
            ("TOPPADDING", (0, 0), (0, 0), 0),
            ("TOPPADDING", (0, 2), (0, 2), 0),
            ("BOTTOMPADDING", (0, 2), (0, 2), 0),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ]))
        return inner

    grid = []
    for i in range(0, len(cards), 2):
        grid.append([_card(cards[i]), _card(cards[i + 1])])
    gt = Table(grid, colWidths=[3.3 * inch, 3.3 * inch])
    gt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(brand.PAPER2)),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(brand.LINE)),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor(brand.LINE)),
        ("TOPPADDING", (0, 0), (-1, -1), 13),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 13),
        ("LEFTPADDING", (0, 0), (-1, -1), 15),
        ("RIGHTPADDING", (0, 0), (-1, -1), 15),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(gt)
    story.append(Spacer(1, 22))

    # ---- trailing-12-month energy chart --------------------------------------
    story.append(Paragraph("Trailing 12 months of production", lbl))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        "Your monthly generation over the last year — measured from your billing "
        "ledger. Year-over-year compares the same billing month one year earlier.",
        small))
    story.append(Spacer(1, 8))
    pts = [(pt.get("month") or "", pt.get("kwh")) for pt in s.get("ttm_points", [])]
    story.append(brand.make_chart_flowable(
        pts, 6.6 * inch, 1.9 * inch,
        empty_msg="No production history on record yet.", light=True))

    doc.build(story, onFirstPage=decorate, onLaterPages=decorate)
    return out_path
