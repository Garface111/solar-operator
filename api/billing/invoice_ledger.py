"""Default offtaker generation + invoice ledger spreadsheet.

Produces a kept-current .xlsx for each offtaker that combines:
  • generation (kWh share) for each billed period
  • invoice amount sent
  • payment status / paid date / money collected (Stripe offtaker pay-links)

Used when the operator has NOT uploaded a BYO sheet — "Download spreadsheet"
always works. When they upload their own layout, sheet_tracker keeps that
format; we still surface payment history via the payments API + optional
extra columns if mapped.

Columns (default auto ledger):
  Period | Generation kWh | Invoice $ | Status | Paid date | Collected $ | Platform fee $ | Invoice #
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any, Optional

logger = logging.getLogger(__name__)

# Logical field → display header (order matters for the sheet)
DEFAULT_COLUMNS: list[tuple[str, str]] = [
    ("period", "Period"),
    ("generation", "Generation kWh"),
    ("amount", "Invoice $"),
    ("status", "Status"),
    ("paid_date", "Paid date"),
    ("collected", "Collected $"),
    ("fee", "Platform fee $"),
    ("invoice_number", "Invoice #"),
]


def _default_mapping() -> dict:
    cols = {field: i for i, (field, _) in enumerate(DEFAULT_COLUMNS)}
    headers = [h for _, h in DEFAULT_COLUMNS]
    return {
        "ok": True,
        "kind": "xlsx",
        "sheet": "Invoice ledger",
        "header_row": 0,
        "headers": headers,
        "columns": cols,
        "data_rows": 0,
        "last_period": None,
        "auto": True,          # system-built default (not a BYO upload)
        "warnings": [],
        "period_style": "YYYY-MM",
    }


def _money(cents: Optional[int]) -> Optional[float]:
    if cents is None:
        return None
    return round(int(cents) / 100.0, 2)


def _period_label_from_payment(p) -> str:
    """Prefer YYYY-MM from period_key / invoice_number."""
    pk = str(getattr(p, "period_key", None) or getattr(p, "invoice_number", None) or "")
    # ISO date → month
    if len(pk) >= 7 and pk[4] == "-":
        return pk[:7]
    return pk[:40] or datetime.utcnow().strftime("%Y-%m")


def _status_label(status: str) -> str:
    s = (status or "").lower()
    return {
        "open": "Awaiting payment",
        "paid": "Paid",
        "failed": "Failed",
        "expired": "Expired",
    }.get(s, status or "—")


def list_payment_rows(db, sub) -> list[dict]:
    """Payment history for one offtaker, newest first — drives UI + ledger."""
    from sqlalchemy import select
    from ..models import OfftakerPayment
    rows = db.execute(
        select(OfftakerPayment).where(
            OfftakerPayment.subscription_id == sub.id,
        ).order_by(OfftakerPayment.id.desc())
    ).scalars().all()
    out = []
    for p in rows:
        out.append({
            "id": p.id,
            "period_key": p.period_key,
            "period_label": _period_label_from_payment(p),
            "invoice_number": p.invoice_number,
            "status": p.status,
            "status_label": _status_label(p.status),
            "amount_usd": _money(p.amount_cents),
            "fee_usd": _money(p.fee_cents),
            "collected_usd": (
                _money(max(int(p.amount_cents or 0) - int(p.fee_cents or 0), 0))
                if p.status == "paid" else None
            ),
            "pay_url": p.pay_url if p.status == "open" else None,
            "paid_at": p.paid_at.isoformat() + "Z" if p.paid_at else None,
            "created_at": p.created_at.isoformat() + "Z" if p.created_at else None,
            "customer_name": p.customer_name,
        })
    return out


def _generation_for_period(db, sub, period_label: str) -> Optional[float]:
    """Best-effort offtaker kWh for a YYYY-MM period from utility bills / match."""
    try:
        from .delivery import build_match
        # build_match may take period_label on manual path
        m = build_match(sub, period_label=period_label if len(period_label) == 7 else None)
        ci = (m.computed_invoice if m else None) or {}
        kwh = ci.get("kwh")
        if kwh is not None:
            return round(float(kwh), 2)
    except Exception:  # noqa: BLE001
        pass
    # Fall back: utility bill for that month
    try:
        from sqlalchemy import select
        from ..models import Bill
        uaid = getattr(sub, "utility_account_id", None)
        if uaid is None:
            return None
        y, mo = map(int, period_label.split("-")[:2])
        bills = db.execute(
            select(Bill).where(Bill.account_id == uaid, Bill.period_end.isnot(None))
        ).scalars().all()
        for b in bills:
            pe = b.period_end.date() if isinstance(b.period_end, datetime) else b.period_end
            if pe and pe.year == y and pe.month == mo:
                # Prefer allocated share if we have allocation
                kwh = getattr(b, "kwh_generated", None) or getattr(b, "kwh_sent_to_grid", None)
                if kwh is None:
                    continue
                pct = getattr(sub, "allocation_pct", None)
                if pct is not None and float(pct) > 0:
                    # offtaker own bill often already allocated (pct≈1)
                    return round(float(kwh) * (float(pct) if float(pct) <= 1 else 1.0), 2)
                return round(float(kwh), 2)
    except Exception:  # noqa: BLE001
        return None
    return None


def build_default_ledger(db, sub) -> tuple[bytes, dict]:
    """Build (or rebuild) the default invoice+generation ledger from payments.

    Rows are oldest→newest for readability in Excel. Generation filled when we
    can resolve it; payment columns come from OfftakerPayment.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    payments = list_payment_rows(db, sub)
    # oldest first for spreadsheet chronology
    payments = list(reversed(payments))

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice ledger"
    headers = [h for _, h in DEFAULT_COLUMNS]
    header_font = Font(bold=True, color="0E1420")
    header_fill = PatternFill("solid", fgColor="D9E7FB")  # sky pastel blue
    thin = Border(
        left=Side(style="thin", color="B0C4DE"),
        right=Side(style="thin", color="B0C4DE"),
        top=Side(style="thin", color="B0C4DE"),
        bottom=Side(style="thin", color="B0C4DE"),
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    last_period = None
    for i, p in enumerate(payments):
        r = i + 2
        pl = p["period_label"]
        last_period = pl
        gen = _generation_for_period(db, sub, pl)
        paid_date = ""
        if p.get("paid_at"):
            try:
                paid_date = p["paid_at"][:10]
            except Exception:
                paid_date = p["paid_at"]
        is_paid = (p.get("status") or "").lower() == "paid"
        values = {
            "period": pl,
            "generation": gen if gen is not None else "",
            "amount": p.get("amount_usd") if p.get("amount_usd") is not None else "",
            "status": p.get("status_label") or p.get("status") or "",
            "paid_date": paid_date if is_paid else "",
            "collected": (p.get("collected_usd") if is_paid and p.get("collected_usd") is not None else ""),
            "fee": (p.get("fee_usd") if is_paid and p.get("fee_usd") is not None else ""),
            "invoice_number": p.get("invoice_number") or "",
        }
        for c, (field, _) in enumerate(DEFAULT_COLUMNS, 1):
            cell = ws.cell(row=r, column=c, value=values.get(field, ""))
            cell.border = thin
            if field in ("amount", "collected", "fee") and values.get(field) != "":
                cell.number_format = '"$"#,##0.00'
            if field == "generation" and values.get(field) != "":
                cell.number_format = "#,##0.00"

    # Column widths
    widths = [12, 16, 12, 18, 12, 12, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Totals row when we have paid amounts
    if payments:
        r = len(payments) + 2
        ws.cell(row=r, column=1, value="TOTAL collected").font = Font(bold=True)
        total_collected = sum(
            (p.get("collected_usd") or 0)
            for p in payments if (p.get("status") or "").lower() == "paid"
        )
        total_inv = sum((p.get("amount_usd") or 0) for p in payments)
        ws.cell(row=r, column=3, value=round(total_inv, 2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6, value=round(total_collected, 2)).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6).font = Font(bold=True, color="1976D2")

    buf = io.BytesIO()
    wb.save(buf)
    mapping = _default_mapping()
    mapping["data_rows"] = len(payments)
    mapping["last_period"] = last_period
    mapping["auto"] = True
    return buf.getvalue(), mapping


def ensure_default_ledger(db, sub) -> dict:
    """If this offtaker has no BYO tracker sheet, build/refresh the default ledger.

    Never overwrites a non-auto (operator-uploaded) sheet. Returns tracker status bits.
    """
    m = getattr(sub, "tracker_map", None) or {}
    has = bool(getattr(sub, "tracker_workbook", None))
    # Preserve BYO uploads (auto is not True)
    if has and m.get("ok") and not m.get("auto"):
        return {"has_sheet": True, "auto": False, "preserved": True}

    blob, mapping = build_default_ledger(db, sub)
    name = f"{(sub.customer_name or 'offtaker').replace(' ', '_')}_invoice_ledger.xlsx"
    sub.tracker_workbook = blob
    sub.tracker_filename = name[:300]
    sub.tracker_map = mapping
    sub.tracker_updated_at = datetime.utcnow()
    db.add(sub)
    return {
        "has_sheet": True,
        "auto": True,
        "data_rows": mapping.get("data_rows"),
        "filename": name,
    }


def sync_payment_into_ledger(db, payment) -> dict:
    """After a payment is created or marked paid, rebuild the default ledger for
    that offtaker (auto sheets only). Safe no-op for BYO sheets."""
    from ..models import BillingReportSubscription
    sub = db.get(BillingReportSubscription, payment.subscription_id)
    if sub is None:
        return {"ok": False, "error": "subscription missing"}
    m = getattr(sub, "tracker_map", None) or {}
    if getattr(sub, "tracker_workbook", None) and m.get("ok") and not m.get("auto"):
        return {"ok": True, "skipped": True, "reason": "byo_sheet"}
    try:
        ensure_default_ledger(db, sub)
        db.commit()
        return {"ok": True, "rebuilt": True}
    except Exception as e:  # noqa: BLE001
        logger.exception("ledger rebuild failed for sub=%s", payment.subscription_id)
        try:
            db.rollback()
        except Exception:
            pass
        return {"ok": False, "error": str(e)[:200]}
