"""
Invoice generators — rebuild the HCT "Template" invoice from a BillingMatch.

Two output formats, the operator picks per subscription:
  render_invoice_xlsx — an .xlsx that mirrors the workbook's own Template sheet.
  render_invoice_pdf  — a clean PDF the customer can pay against (reportlab).

Both consume a `BillingMatch` (api/billing/matcher.py) and a chosen `Period`
(defaults to the latest). `invoice_for_period` does the dollar math via the
shared `compute_invoice` so XLSX and PDF never drift.
"""
from __future__ import annotations

import pathlib
from datetime import date, timedelta
from typing import Optional

from .matcher import BillingMatch, Period, compute_invoice

DUE_DAYS = 28


def invoice_for_period(match: BillingMatch, period: Period,
                       invoice_date: date) -> dict:
    """Assemble the full invoice payload for one period (dates + dollar math)."""
    inv = compute_invoice(
        period.customer_kwh, period.tariff, period.adder,
        match.billing_rate, match.billing_model,
        match.template.get("fixed_amount"),
    )
    due = invoice_date + timedelta(days=DUE_DAYS)
    inv.update({
        "invoice_number": period.end.strftime("%Y-%m") if period.end else invoice_date.strftime("%Y-%m"),
        "invoice_date": invoice_date.isoformat(),
        "due_date": due.isoformat(),
        "period_start": period.start.isoformat() if period.start else None,
        "period_end": period.end.isoformat() if period.end else None,
        "month": period.month,
        "customer_name": match.customer.get("name"),
        "billing_model": match.billing_model,
        "allocation_pct": match.allocation_pct,
        # The ARRAY's total this period — for a GMP-credit offtaker this is the
        # EXCESS sent to grid (Bruce's model), not gross production; the member's
        # share = array/excess × allocation.
        "array_kwh": period.array_kwh,
        # Provenance of the credit rate so the renderer can label the basis (excess
        # vs production) honestly and flag a banked-month reference estimate.
        "net_rate_source": (match.computed_invoice or {}).get("net_rate_source"),
        "net_rate_note": (match.computed_invoice or {}).get("net_rate_note"),
        # Provenance of the PRODUCTION figure (gmp_api | daily_csv | utility_bill |
        # bill_prorate, possibly '+'-joined across arrays) so the renderer can tell
        # the customer how their generation was measured vs estimated — honest data.
        "kwh_source": (match.computed_invoice or {}).get("kwh_source"),
        "project_total_kwh": match.project_totals.get("total_customer_kwh"),
        "project_total_savings": match.project_totals.get("total_savings"),
        # Per-array breakdown for multi-array offtakers (one line per array). Lives
        # on project_totals/computed_invoice; surfaced here so the PDF can render it.
        "array_breakdown": match.project_totals.get("array_breakdown")
            or (match.computed_invoice or {}).get("array_breakdown"),
        # Budget billing: build_match put the operator's fixed final amount on
        # computed_invoice.amount_owed (flagged budget_override) while preserving the
        # CALCULATED credit. invoice_for_period recomputes the credit fresh, so surface
        # the budget SEPARATELY here → the renderer shows BOTH the solar credit value
        # and the budgeted amount, with AMOUNT DUE = the budget (the actual bill).
        "budget_override": bool((match.computed_invoice or {}).get("budget_override")),
        "budgeted_amount": ((match.computed_invoice or {}).get("amount_owed")
                            if (match.computed_invoice or {}).get("budget_override") else None),
        # Quarterly coverage: which months this invoice sums, plus a human note
        # marking the covered range (esp. a mid-quarter service start) — the
        # renderers print it so a quarter is never mistaken for one month.
        "billing_cadence": (match.computed_invoice or {}).get("billing_cadence"),
        "period_months": (match.computed_invoice or {}).get("period_months"),
        "period_note": (match.computed_invoice or {}).get("period_note"),
    })
    return inv


def _money(x: Optional[float]) -> str:
    return f"${(x or 0):,.2f}"


def _pct(x: Optional[float]) -> str:
    return f"{round((x or 0) * 100):g}%"


def _rate_str(x: Optional[float]) -> str:
    """Per-kWh rate with trailing zeros trimmed: 0.25760 -> '$0.2576/kWh'."""
    return f"${('%.5f' % (x or 0)).rstrip('0').rstrip('.')}/kWh"


# Honest, customer-facing provenance line for the production figure. kwh_source is
# one of gmp_api | daily_csv | utility_bill | bill_prorate (or '+'-joined across
# arrays for a multi-array offtaker). We name the WEAKEST source present so an
# estimate-dominated period is never presented as fully metered (Ford's #1 trust
# rule). Returns None when provenance is unknown (renders nothing).
def _kwh_source_note(kwh_source: Optional[str]) -> Optional[str]:
    if not kwh_source:
        return None
    parts = {p for p in str(kwh_source).split("+") if p}
    if not parts:
        return None
    # Weakest-first: a prorated estimate anywhere downgrades the whole label.
    if "bill_prorate" in parts:
        return ("Production for at least one array this period is estimated from your "
                "utility bill (daily meter reads weren't available), so these figures "
                "are an approximation.")
    if "utility_bill" in parts:
        return "Production this period is taken from your settled utility bill."
    if "daily_csv" in parts:
        return "Production this period is from daily meter reads."
    if "gmp_api" in parts:
        return "Production this period is verified by daily meter reads from your utility."
    return None


# ─── XLSX ───────────────────────────────────────────────────────────────────

def render_invoice_xlsx(match: BillingMatch, out_path: pathlib.Path,
                        period: Optional[Period] = None,
                        invoice_date: Optional[date] = None) -> pathlib.Path:
    """Write an .xlsx invoice mirroring the HCT Template sheet layout."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    period = period or match.latest_period
    if period is None:
        raise ValueError("no billing period to invoice")
    invoice_date = invoice_date or date.today()
    inv = invoice_for_period(match, period, invoice_date)
    tpl = match.template
    # Budget billing: when the operator set a fixed amount, AMOUNT DUE is that budget
    # (the actual bill); the calculated solar credit value still shows as its own line.
    _budget_on = bool(inv.get("budget_override")) and inv.get("budgeted_amount") is not None
    _final_due = inv["budgeted_amount"] if _budget_on else inv["amount_owed"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 24
    bold = Font(bold=True)
    big = Font(bold=True, size=14)
    right = Alignment(horizontal="right")

    def put(cell, value, fnt=None, align=None):
        ws[cell] = value
        if fnt:
            ws[cell].font = fnt
        if align:
            ws[cell].alignment = align

    put("B2", tpl.get("title") or "Solar Credit Invoice", big)
    put("B4", tpl.get("operator") or "Your solar array owner", bold)
    if tpl.get("phone"):
        put("C4", tpl["phone"])
    if tpl.get("attn"):
        put("C5", tpl["attn"])
    if match.customer.get("email") or tpl.get("email"):
        put("C6", "email: " + (match.customer.get("email") or tpl.get("email")))

    put("B8", inv["customer_name"] or "Customer", bold)
    put("B9", "Invoice Number:"); put("C9", inv["invoice_number"], align=right)
    put("B10", "Invoice Date:"); put("C10", inv["invoice_date"], align=right)
    put("B11", "Due Date (28 days):"); put("C11", inv["due_date"], align=right)
    put("B12", "Time Period Covered:")
    put("C12", f"{inv['period_start']} → {inv['period_end']}", align=right)

    put("B14", "How this period's solar credit is calculated", bold)
    rate = (inv["tariff"] or 0) + (inv["adder"] or 0)
    _gmpc = inv.get("net_rate_source") in ("gmp_bill_credit", "gmp_credit_reference")
    # The "Solar generation sent to grid" (array total) + "Your share of the array"
    # (allocation %) breakdown lines were removed per Ford — the offtaker's own kWh
    # ("Your share of the generation") is the only input shown. Rows shifted up.
    put("B15", "Your share of the generation (kWh):" if _gmpc else "Your share of production (kWh):")
    put("C15", round(inv["kwh"], 0), align=right)
    put("B16", "Solar credit rate:")
    put("C16", _rate_str(rate), align=right)
    put("B17", "Your contractual payment share:"); put("C17", _pct(inv["billing_rate"]), align=right)
    put("B18", "Solar credit value due:"); put("C18", _money(inv["amount_owed"]), align=right)
    # One note line at B19 (the layout reserves a single row before Amount Owed).
    # Combine the data-source provenance with the banked-credit note if both apply.
    _notes = []
    if inv.get("period_note"):
        _notes.append(str(inv["period_note"]))
    _src_note = _kwh_source_note(inv.get("kwh_source"))
    if _src_note:
        _notes.append(f"Data source: {_src_note}")
    if inv.get("net_rate_source") == "gmp_credit_reference":
        _notes.append("Note: credit banked this period — rate is a reference estimate (trued up annually).")
    if _budget_on:
        _notes.append("Amount owed is your fixed budgeted amount; the solar credit value above is the calculated reference.")
    if _notes:
        put("B19", "  ".join(_notes))

    put("B20", "Amount Owed:", big)
    put("C20", _money(_final_due), big, right)
    put("B21", tpl.get("payable_to") or f"Please make payment to {tpl.get('operator') or 'your solar array owner'}.")

    put("B24", "Project Total kWh generation:")
    put("C24", round(inv.get("project_total_kwh") or 0, 0), align=right)
    put("B25", "Project total financial savings:")
    put("C25", _money(inv.get("project_total_savings")), align=right)

    put("B29", "Solar credit invoice service provided by ArrayOperator.com  ·  admin@solaroperator.org")

    out_path = pathlib.Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


# ─── PDF ────────────────────────────────────────────────────────────────────
# Shared brand kit (palette, dark energy hero band, energy bar chart) — one
# source of truth across invoice / performance summary / quarterly report.
from . import _pdf_brand as brand


def render_invoice_pdf(match: BillingMatch, out_path: pathlib.Path,
                       period: Optional[Period] = None,
                       invoice_date: Optional[date] = None) -> pathlib.Path:
    """Write a slick, on-brand one-page PDF invoice (reportlab).

    Design: a dark 'energy' hero band (Array Operator green glow) up top, a clean
    payable invoice body in the middle, and a juicy monthly-energy bar chart at
    the bottom. Matches the site palette so the document feels like the product.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("reportlab is required for PDF invoices") from e

    period = period or match.latest_period
    if period is None:
        raise ValueError("no billing period to invoice")
    invoice_date = invoice_date or date.today()
    inv = invoice_for_period(match, period, invoice_date)
    tpl = match.template
    # Budget billing: when the operator set a fixed amount, AMOUNT DUE is that budget
    # (the actual bill); the calculated solar credit value still shows as its own line.
    _budget_on = bool(inv.get("budget_override")) and inv.get("budgeted_amount") is not None
    _final_due = inv["budgeted_amount"] if _budget_on else inv["amount_owed"]

    out_path = pathlib.Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    HERO_H = 1.55 * inch
    # Day-skin palette (light): slate ink, utility-blue chrome, emerald money.
    INKDK = colors.HexColor(brand.DAY_INK)
    MUTEDDK = colors.HexColor(brand.DAY_MUTED)
    GREEN_DK = colors.HexColor(brand.DAY_GREEN)   # the money / credit figure
    BLUE = colors.HexColor(brand.DAY_BLUE)
    LINE_C = colors.HexColor(brand.DAY_LINE)
    operator_name = tpl.get("operator") or "Your solar array owner"
    title = tpl.get("title") or "Solar Credit Invoice"

    # White-label: the offtaker invoice is the OPERATOR's document, sent to their
    # customer under their own name — so we suppress the Array Operator sun-glyph,
    # wordmark, and the "service provided by ArrayOperator.com" footer attribution.
    decorate = brand.make_hero_decorator(
        title=title, subtitle=operator_name,
        right_label="AMOUNT DUE", right_value=brand._money(_final_due),
        footer_left=operator_name,
        footer_right=f"Invoice {inv['invoice_number']}", hero_h=HERO_H, light=True,
        show_brand=False)

    styles = getSampleStyleSheet()
    lbl = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=11,
                         textColor=INKDK, fontName="Helvetica-Bold")
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8.5,
                           textColor=MUTEDDK, leading=12)
    # Wrapping style for the bill-to / period values: a long customer name must
    # WRAP within its column, not overflow into the period cell (plain table
    # strings don't wrap → they bleed across the column boundary).
    billval = ParagraphStyle("billval", parent=styles["Normal"], fontSize=13,
                             leading=15, fontName="Helvetica-Bold",
                             textColor=INKDK)

    doc = SimpleDocTemplate(
        str(out_path), pagesize=letter,
        topMargin=HERO_H + 0.3 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.85 * inch, rightMargin=0.85 * inch)
    story = []

    # ---- meta block (bill-to + dates) ----------------------------------------
    from xml.sax.saxutils import escape as _xesc
    meta = [
        ["BILL TO", "PERIOD"],
        [Paragraph(_xesc(inv["customer_name"] or "Customer"), billval),
         Paragraph(f"{_xesc(inv['period_start'] or '')}  →  "
                   f"{_xesc(inv['period_end'] or '')}", billval)],
        ["", ""],
        ["Invoice date", inv["invoice_date"]],
        ["Due date (28 days)", inv["due_date"]],
    ]
    mt = Table(meta, colWidths=[3.3 * inch, 3.3 * inch])
    mt.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("TEXTCOLOR", (0, 0), (-1, 0), MUTEDDK),
        ("VALIGN", (0, 1), (-1, 1), "TOP"),
        ("FONTSIZE", (0, 3), (-1, -1), 9.5),
        ("TEXTCOLOR", (0, 3), (0, -1), MUTEDDK),
        ("TEXTCOLOR", (1, 3), (1, -1), INKDK),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(mt)
    story.append(Spacer(1, 18))

    # ---- per-array breakdown (multi-array offtakers) -------------------------
    # When the offtaker owns a share of several arrays, show one line per array
    # (array · its production · this offtaker's % · their kWh) so the summed
    # "Energy produced" total below is auditable. Single-array invoices skip this.
    breakdown = inv.get("array_breakdown") or []
    if len(breakdown) > 1:
        story.append(Paragraph("Your share by array", lbl))
        story.append(Spacer(1, 8))
        brows = [["Array", "Array produced", "Your %", "Your kWh"]]
        for b in breakdown:
            brows.append([
                str(b.get("array_name") or "Array"),
                f"{float(b.get('array_kwh') or 0):,.0f} kWh",
                _pct(b.get("allocation_pct") or 0),
                f"{float(b.get('customer_kwh') or 0):,.0f} kWh",
            ])
        brows.append(["Total", "", "", f"{inv['kwh']:,.0f} kWh"])
        bt = Table(brows, colWidths=[2.7 * inch, 1.5 * inch, 0.9 * inch, 1.5 * inch])
        bt.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 0), (-1, -1), INKDK),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor(brand.LINE)),
            ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.HexColor(brand.LINE)),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(bt)
        story.append(Spacer(1, 16))

    # ---- the solar-credit calculation (the operator's "invoice guts") --------
    # array production → your share → your kWh → credit rate → payment share →
    # credit value due. Mirrors the line-by-line math the array owner spec'd.
    story.append(Paragraph("How this period's solar credit is calculated", lbl))
    story.append(Spacer(1, 8))
    rate = (inv["tariff"] or 0) + (inv["adder"] or 0)     # excess credit + solar incentive
    _gmpc = inv.get("net_rate_source") in ("gmp_bill_credit", "gmp_credit_reference")
    # Customer-facing label: GMP members are billed on the EXCESS sent to grid; we
    # show their SHARE of it as "Solar generation" per Ford (still the excess figure).
    # The array-total + share-% breakdown lines were removed per Ford — the offtaker's
    # own kWh ("Your share of the generation") is the only input they need to see.
    _shr_lbl = "Your share of the generation" if _gmpc else "Your share of production"
    rows = [
        [_shr_lbl, f"{inv['kwh']:,.0f} kWh"],
        ["Solar credit rate", _rate_str(rate)],
        ["Your contractual payment share", _pct(inv["billing_rate"])],
        ["Solar credit value due", _money(inv["amount_owed"])],
    ]
    if _budget_on:
        rows.append(["Budgeted amount (your fixed bill)", _money(inv["budgeted_amount"])])
    rt = Table(rows, colWidths=[4.4 * inch, 2.2 * inch])
    rt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (-1, -1), INKDK),
        ("TEXTCOLOR", (0, 0), (0, 0), MUTEDDK),             # the share-of-generation input row quieter
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, LINE_C),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),  # emphasize the value-due row
        ("FONTSIZE", (0, -1), (-1, -1), 11),
        ("TEXTCOLOR", (1, -1), (1, -1), GREEN_DK),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, BLUE),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(rt)
    story.append(Spacer(1, 6))
    if inv.get("period_note"):
        story.append(Paragraph(_xesc(str(inv["period_note"])), small))
        story.append(Spacer(1, 3))
    _src_note = _kwh_source_note(inv.get("kwh_source"))
    if _src_note:
        story.append(Paragraph(f"Data source: {_xesc(_src_note)}", small))
    if inv.get("net_rate_source") == "gmp_credit_reference":
        story.append(Spacer(1, 3))
        story.append(Paragraph(
            "Credit banked with the utility this period — rate is a reference estimate, trued up annually.", small))
    story.append(Spacer(1, 16))

    # ---- amount due banner (light, day skin) ---------------------------------
    owed = Table([[f"AMOUNT DUE  ·  payable to {_xesc(operator_name)}", _money(_final_due)]],
                 colWidths=[4.4 * inch, 2.2 * inch])
    owed.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (0, 0), 10.5),
        ("FONTNAME", (1, 0), (1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (1, 0), (1, 0), 18),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(brand.DAY_GREENBG)),
        ("TEXTCOLOR", (0, 0), (0, 0), INKDK),
        ("TEXTCOLOR", (1, 0), (1, 0), GREEN_DK),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBEFORE", (0, 0), (0, -1), 3, GREEN_DK),
        ("TOPPADDING", (0, 0), (-1, -1), 13),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 13),
        ("LEFTPADDING", (0, 0), (-1, -1), 16),
        ("RIGHTPADDING", (0, 0), (-1, -1), 16),
        ("ROUNDEDCORNERS", [7, 7, 7, 7]),
    ]))
    story.append(owed)
    story.append(Spacer(1, 12))

    # ---- payment terms + master-account-holder contact -----------------------
    pay_to = tpl.get("payable_to") or f"Please make payment to {operator_name}."
    bits = []
    if tpl.get("attn"):
        bits.append(_xesc(str(tpl["attn"])))
    if tpl.get("phone"):
        bits.append(_xesc(str(tpl["phone"])))
    cmail = tpl.get("email") or match.customer.get("email")
    if cmail:
        bits.append(_xesc(str(cmail)))
    contact_line = ("<br/>" + "  ·  ".join(bits)) if bits else ""
    terms = Table([
        [Paragraph("<b>Payment</b>", small), Paragraph("<b>Questions about this invoice?</b>", small)],
        [Paragraph(f"Due within 28 days. {_xesc(pay_to)}", small),
         Paragraph(f"Contact {_xesc(operator_name)}{contact_line}", small)],
    ], colWidths=[3.3 * inch, 3.3 * inch])
    terms.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEABOVE", (0, 0), (-1, 0), 0.4, LINE_C),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 1), (-1, 1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(terms)
    story.append(Spacer(1, 12))

    # The monthly production chart lives in the separate generation report, not on
    # the invoice — drop the second-page chart from offtaker invoices (Ford).
    doc.build(story, onFirstPage=decorate, onLaterPages=decorate)
    return out_path
