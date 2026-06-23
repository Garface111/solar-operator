"""
Billing-workbook matcher — the heart of "match any spreadsheet uploaded".

The Array Operator billing reports are driven by the HCT Sun Enterprises
workbook family. Every workbook in that family, regardless of customer or sheet
names, shares a recognizable shape:

  * a DATA LEDGER sheet (named per-customer: "Fairlee", "NFD", "Valley Cares
    Data", "SAMPLE") with:
      - a metadata row    — CUSTOMER | ADDRESS | ACCT # | METER # |
                            "Percent of solar net metering credits…" (allocation)
                            | "Price Factor (100% - discount)" (billing rate)
                            | … | email
      - a ledger header   — Month | Date start | Date End | kWh whole array |
                            kWh «Customer» | Tariff | Adder | (Tariff+Adder) |
                            Value | Bill | Savings
      - monthly rows below.
  * an invoice "Template" sheet — the bill the customer pays.
  * an "Annual True-Up" sheet.

Three billing models occur and are auto-detected:
  fixed_budget      — a flat "Fixed Monthly Budget Payment" each month (Fairlee).
  flat_rate         — a flat estimated rate + a September true-up (Valley Cares).
  percent_of_array  — customer is billed for their % share of the whole array's
                      generation at (tariff + adder) × billing-rate (Norwich/NFD).

`match_billing_workbook(file_bytes)` returns a `BillingMatch` with a confidence
score and a field map so the UI can show what was recognized and let the
operator correct it. When the schema scan is low-confidence it falls back to the
LLM column-mapper already used by api/ingest.py, so a genuinely novel sheet still
matches.

This module is pure (no DB, no network except the optional LLM fallback) so it's
trivially unit-testable against the real sample workbooks.
"""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass, field, asdict
from datetime import date, datetime
from html import escape as _html_escape
from typing import Any, Optional

logger = logging.getLogger(__name__)

# A header row needs at least this fraction of the canonical ledger tokens to be
# accepted as the data-ledger header. 0.45 comfortably matches all three sample
# families (which hit 0.8+) while rejecting unrelated sheets.
LEDGER_HEADER_THRESHOLD = 0.45
# Below this overall confidence we hand off to the LLM fallback.
LLM_FALLBACK_THRESHOLD = 0.5


# ─── data structures ────────────────────────────────────────────────────────

@dataclass
class Period:
    """One monthly billing row from the data ledger."""
    month: Optional[str] = None
    start: Optional[date] = None
    end: Optional[date] = None
    array_kwh: Optional[float] = None
    customer_kwh: Optional[float] = None
    tariff: Optional[float] = None
    adder: Optional[float] = None
    value: Optional[float] = None     # solar value = customer_kwh × (tariff+adder)
    bill: Optional[float] = None      # value × billing_rate
    savings: Optional[float] = None   # value − bill
    flat_rate: Optional[float] = None  # flat-rate column, when present

    def to_dict(self) -> dict:
        d = asdict(self)
        for k in ("start", "end"):
            if isinstance(d[k], (date, datetime)):
                d[k] = d[k].isoformat()
        return d


@dataclass
class BillingMatch:
    matched: bool
    confidence: float
    source: str  # "schema" | "llm" | "none"
    data_sheet: Optional[str] = None
    customer: dict = field(default_factory=dict)   # name, address, acct, meter, email
    allocation_pct: Optional[float] = None
    billing_rate: Optional[float] = None           # price factor (e.g. 0.9 = 90%)
    billing_model: str = "percent_of_array"
    periods: list[Period] = field(default_factory=list)
    latest_period: Optional[Period] = None
    template: dict = field(default_factory=dict)    # static invoice header lifted from Template sheet
    computed_invoice: dict = field(default_factory=dict)
    field_map: dict = field(default_factory=dict)
    project_totals: dict = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "matched": self.matched,
            "confidence": round(self.confidence, 3),
            "source": self.source,
            "data_sheet": self.data_sheet,
            "customer": self.customer,
            "allocation_pct": self.allocation_pct,
            "billing_rate": self.billing_rate,
            "billing_model": self.billing_model,
            "periods": [p.to_dict() for p in self.periods],
            "latest_period": self.latest_period.to_dict() if self.latest_period else None,
            "template": self.template,
            "computed_invoice": self.computed_invoice,
            "field_map": self.field_map,
            "project_totals": self.project_totals,
            "warnings": self.warnings,
        }


# ─── cell helpers ───────────────────────────────────────────────────────────

def _s(v: Any) -> str:
    """Normalize a cell to a clean lowercase string for token matching."""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().lower()


def _num(v: Any) -> Optional[float]:
    """Coerce a cell value to a float, tolerating $, %, commas, and stray text."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "").replace("\xa0", "").strip()
    pct = s.endswith("%")
    s = s.rstrip("%").strip()
    try:
        n = float(s)
        return n / 100.0 if pct else n
    except ValueError:
        return None


def _as_date(v: Any) -> Optional[date]:
    """Coerce a cell to a date. Handles datetime objects and common string forms."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _grid(ws, max_rows: int = 200, max_cols: int = 30) -> list[list[Any]]:
    """Read a worksheet into a dense 2-D python list (0-indexed)."""
    rows: list[list[Any]] = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx >= max_rows:
            break
        rows.append(list(row[:max_cols]) if row else [])
    return rows


# ─── ledger header detection ────────────────────────────────────────────────

# canonical ledger header token → matcher predicate on a lowercased cell.
_LEDGER_TOKENS: dict[str, Any] = {
    "month": lambda c: c == "month" or c.startswith("month"),
    "start": lambda c: "date start" in c or c == "start" or "start" in c,
    "end": lambda c: "date end" in c or c == "end" or ("end" in c and "date" in c),
    "array_kwh": lambda c: "whole array" in c or ("kwh" in c and "array" in c),
    "tariff": lambda c: "tariff" in c or "tarrif" in c,
    "adder": lambda c: "adder" in c,
    "value": lambda c: c == "value" or c.endswith(" value"),
    "bill": lambda c: c == "bill",
    "savings": lambda c: "saving" in c,
}


def _score_header_row(cells: list[Any]) -> tuple[float, dict[str, int]]:
    """Score one row as a candidate ledger header.

    Returns (fraction_of_tokens_found, {field: col_index}). The customer-kWh
    column is resolved separately (it's the kWh column that is NOT 'whole array').
    """
    lc = [_s(c) for c in cells]
    found: dict[str, int] = {}
    for token, pred in _LEDGER_TOKENS.items():
        for idx, c in enumerate(lc):
            if c and pred(c) and token not in found:
                found[token] = idx
                break
    # Customer kWh = a kWh column that isn't the "whole array" column.
    array_col = found.get("array_kwh")
    for idx, c in enumerate(lc):
        if "kwh" in c and idx != array_col and "customer_kwh" not in found:
            found["customer_kwh"] = idx
            break
    # Score over the canonical tokens only (customer_kwh is a bonus column, not
    # one of the scored tokens), capped at 1.0.
    scored = sum(1 for t in _LEDGER_TOKENS if t in found)
    score = min(1.0, scored / len(_LEDGER_TOKENS))
    return score, found


def _candidate_sheets(wb) -> list[tuple[str, int, dict[str, int], float]]:
    """Every sheet whose best header row clears the ledger threshold.

    Returns a list of (sheet_name, header_row_index, field_map, score). A
    workbook can legitimately contain several ledger-shaped sheets (a live
    customer ledger plus a stale "SAMPLE"/"-old" sheet); the caller picks the
    current one by recency.

    Scans the first 20 rows per sheet (up from 14) so workbooks with a tall
    title/logo block above the metadata row are still matched correctly.
    """
    cands: list[tuple[str, int, dict[str, int], float]] = []
    for ws in wb.worksheets:
        grid = _grid(ws, max_rows=20)
        best_row, best_score, best_fmap = -1, 0.0, {}
        for r_idx, cells in enumerate(grid[:18]):
            score, fmap = _score_header_row(cells)
            if score > best_score:
                best_row, best_score, best_fmap = r_idx, score, fmap
        if best_score >= LEDGER_HEADER_THRESHOLD:
            cands.append((ws.title, best_row, best_fmap, best_score))
    return cands


# ─── metadata extraction ────────────────────────────────────────────────────

_META_TOKENS: dict[str, Any] = {
    "name": lambda c: c == "customer" or c.startswith("customer"),
    "address": lambda c: c == "address" or c.startswith("address"),
    "acct": lambda c: "acct" in c or "account" in c,
    "meter": lambda c: "meter" in c,
    "allocation": lambda c: "percent of solar" in c or "net metering credit" in c,
    "billing_rate": lambda c: "price factor" in c or "discount" in c,
    "email": lambda c: c == "email" or "e-mail" in c,
}


def _extract_metadata(grid: list[list[Any]], header_row: int) -> dict:
    """Find the metadata label row above the ledger header and read the values
    from the row directly beneath it (the HCT layout: row1 labels, row2 values).
    """
    out: dict[str, Any] = {}
    label_row_idx = None
    label_cols: dict[str, int] = {}
    # The metadata block sits in the rows above the ledger header.
    for r in range(0, max(header_row, 1)):
        cells = grid[r] if r < len(grid) else []
        lc = [_s(c) for c in cells]
        hits: dict[str, int] = {}
        for tok, pred in _META_TOKENS.items():
            for idx, c in enumerate(lc):
                if c and pred(c) and tok not in hits:
                    hits[idx_key := tok] = idx
                    break
        if len(hits) >= 3:  # CUSTOMER/ADDRESS/… row
            label_row_idx, label_cols = r, hits
            break
    if label_row_idx is None:
        return out
    value_row = grid[label_row_idx + 1] if label_row_idx + 1 < len(grid) else []

    def val(col: int) -> Any:
        return value_row[col] if col < len(value_row) else None

    if "name" in label_cols:
        out["name"] = _clean(val(label_cols["name"]))
    if "address" in label_cols:
        out["address"] = _clean(val(label_cols["address"]))
    if "acct" in label_cols:
        out["acct"] = _short_id(val(label_cols["acct"]))
    if "meter" in label_cols:
        out["meter"] = _short_id(val(label_cols["meter"]))
    if "email" in label_cols:
        out["email"] = _clean(val(label_cols["email"]))
    if "allocation" in label_cols:
        out["allocation_pct"] = _num(val(label_cols["allocation"]))
    if "billing_rate" in label_cols:
        out["billing_rate"] = _num(val(label_cols["billing_rate"]))
    return out


def _clean(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).replace("\xa0", " ").strip()
    return s or None


def _short_id(v: Any) -> Optional[str]:
    """Clean an acct/meter cell, rejecting prose that landed in the column
    (some ledgers reuse the meter cell for free-text notes)."""
    s = _clean(v)
    if not s:
        return None
    # An account/meter id is short and word-light; a note is a sentence.
    if len(s) > 25 or s.count(" ") >= 4:
        return None
    return s


# ─── period parsing ─────────────────────────────────────────────────────────

def _parse_periods(grid: list[list[Any]], header_row: int, fmap: dict[str, int]) -> list[Period]:
    periods: list[Period] = []
    blanks = 0
    for r in range(header_row + 1, len(grid)):
        cells = grid[r]

        def cell(field_name: str) -> Any:
            col = fmap.get(field_name)
            if col is None or col >= len(cells):
                return None
            return cells[col]

        month = _clean(cell("month"))
        start = _as_date(cell("start"))
        end = _as_date(cell("end"))
        ckwh = _num(cell("customer_kwh"))
        akwh = _num(cell("array_kwh"))
        # A real data row has at least a month label or a date or kWh.
        if not (month or start or end or ckwh or akwh):
            blanks += 1
            if blanks >= 4:  # consecutive blank run → end of ledger
                break
            continue
        blanks = 0
        periods.append(Period(
            month=month, start=start, end=end,
            array_kwh=akwh, customer_kwh=ckwh,
            tariff=_num(cell("tariff")), adder=_num(cell("adder")),
            value=_num(cell("value")), bill=_num(cell("bill")),
            savings=_num(cell("savings")),
        ))
    return periods


def _latest_period(periods: list[Period]) -> Optional[Period]:
    """The most recent period that actually has generation data."""
    real = [p for p in periods if (p.customer_kwh or p.array_kwh)]
    if not real:
        return None
    dated = [p for p in real if p.end]
    if dated:
        return max(dated, key=lambda p: p.end)  # type: ignore[arg-type]
    return real[-1]


# ─── billing-model detection ────────────────────────────────────────────────

def _find_template_sheet(wb):
    for ws in wb.worksheets:
        if "template" in (ws.title or "").lower():
            return ws
    return None


# ─── invoice-sheet finder (operator's own invoice-template upload) ───────────
# "Find the correct page of a spreadsheet": locate the sheet that IS the
# customer's invoice (the bill they pay) wherever it sits in the workbook — by
# title hint + invoice-ish content, NOT by position or an exact "Template" name.

_INVOICE_TITLE_HINTS = {
    "template": 4, "invoice": 4, "bill": 2, "remit": 2, "statement": 2,
}
# Content tokens that mark a sheet as an actual invoice (not a data ledger).
_INVOICE_CONTENT_SCORES = {
    "amount due": 3, "amount owed": 3, "total due": 3, "payable to": 3,
    "remit": 2, "bill to": 2, "invoice #": 2, "invoice no": 2,
    "invoice number": 2, "invoice": 2, "due date": 1, "attn": 1, "statement": 1,
}


def _looks_like_ledger(blob: str) -> bool:
    """The monthly DATA ledger / true-up sheet — NOT the invoice the customer pays."""
    return ("tariff" in blob and "adder" in blob) or "whole array" in blob


def _score_invoice_sheet(title: str, grid: list[list[Any]]) -> float:
    blob = " \n ".join(_s(c) for row in grid for c in row if c is not None)
    score = 0.0
    tl = (title or "").lower()
    for hint, pts in _INVOICE_TITLE_HINTS.items():
        if hint not in tl:
            continue
        if hint == "bill" and "billing" in tl:
            continue   # a "Billing" sheet is the ledger, not the invoice
        score += pts
        break
    for tok, pts in _INVOICE_CONTENT_SCORES.items():
        if tok in blob:
            score += pts
    if _looks_like_ledger(blob):
        score -= 6           # data ledger / true-up is not the invoice
    return score


def find_invoice_sheet(wb):
    """The worksheet that IS the customer's invoice, scored across the whole
    workbook (handles "page 4", non-'Template' names). None if nothing scores."""
    best, best_score = None, 0.0
    for ws in wb.worksheets:
        grid = _grid(ws, max_rows=60)
        sc = _score_invoice_sheet(ws.title or "", grid)
        if sc > best_score:
            best, best_score = ws, sc
    return best if best_score >= 2 else None


# ── Excel-invoice → faithful, tokenized HTML template ────────────────────────
# A crude text grid mangled real invoices (lost column widths, bold, alignment,
# number formatting) AND baked the sample's static numbers in — so an enabled
# template rendered the SAMPLE's $ and kWh onto every offtaker. This converter
# (1) preserves layout (widths, bold, alignment, $/%/date formatting, merges) and
# (2) replaces clearly-labelled data cells with {{ tokens }} so each invoice fills
# with that offtaker's real numbers via build_token_context.

# Label substring (lowercased, on a cell) → token that replaces the row's value.
# Order matters — first contained substring wins, so the more specific
# "billing at the estimated" precedes "billing rate".
_TOKEN_LABELS = [
    ("amount due", "{{ amount_due }}"), ("amount owed", "{{ amount_due }}"),
    ("final amount", "{{ amount_due }}"),
    ("billing at the estimated", "{{ amount_due }}"),   # flat-rate "billed this period" cell
    ("invoice number", "{{ invoice_number }}"), ("invoice no", "{{ invoice_number }}"),
    ("invoice date", "{{ invoice_date }}"),
    ("due date", "{{ due_date }}"),
    ("solar value", "{{ solar_value }}"),
    ("billing rate", "{{ billed_value }}"),             # the $ this rate produces
    ("solar savings", "{{ solar_savings }}"),
]


def _excel_disp(value, number_format: str) -> str:
    """A cell's DISPLAYED string, honoring its number format ($ / % / date / int)
    — so we never dump a raw float like 3519.0629599999997 onto the invoice."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (datetime, date)):
        return value.strftime("%b %d, %Y")
    if isinstance(value, (int, float)):
        nf = number_format or ""
        if "%" in nf:
            return f"{value * 100:.0f}%"
        if "$" in nf or "accounting" in nf.lower():
            return f"${value:,.2f}"
        if abs(value - round(value)) < 1e-9:
            return f"{int(round(value)):,}"
        return f"{value:,.2f}"
    return str(value)


def _content_bounds(ws, max_rows: int, max_cols: int):
    """(r0, r1, c0, c1) of the invoice's real content. Honors the sheet's PRINT
    AREA when set (so scratch cells in far columns are excluded), else the full
    used range trimmed of trailing empties."""
    pa = getattr(ws, "print_area", None)
    if pa:
        try:
            from openpyxl.utils import range_boundaries
            ref = pa.split(",")[0].split("!")[-1].replace("$", "")
            c0, r0, c1, r1 = range_boundaries(ref)
            return r0, min(r1, r0 + max_rows - 1), c0, min(c1, c0 + max_cols - 1)
        except Exception:  # noqa: BLE001 — fall back to used-range trimming
            pass
    r0, c0 = 1, 1
    r1 = min(ws.max_row or 0, max_rows)
    c1 = min(ws.max_column or 0, max_cols)

    def row_has(r):
        return any(ws.cell(row=r, column=c).value not in (None, "") for c in range(c0, c1 + 1))

    def col_has(c):
        return any(ws.cell(row=r, column=c).value not in (None, "") for r in range(r0, r1 + 1))

    while r1 >= r0 and not row_has(r1):
        r1 -= 1
    while c1 >= c0 and not col_has(c1):
        c1 -= 1
    return r0, r1, c0, c1


def _build_token_map(ws, r0: int, r1: int, c0: int, c1: int, covered: set) -> dict:
    """Map (row, col) → token for clearly-labelled data cells, so the invoice
    fills per-period instead of showing the uploaded sample's frozen numbers."""
    tok: dict = {}

    def value_cell_right(r, c):
        for cc in range(c + 1, c1 + 1):
            if (r, cc) in covered:
                continue
            v = ws.cell(row=r, column=cc).value
            if v not in (None, ""):
                return cc
        return None

    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = _s(v)
            token = None
            if s.startswith("kwh"):                     # "KWH:" row
                token = "{{ kwh }}"
            else:
                for lab, t in _TOKEN_LABELS:
                    if lab in s:
                        token = t
                        break
            if token:
                vc = value_cell_right(r, c)
                if vc is not None:
                    tok[(r, vc)] = token
        # "Time Period Covered:  From … To …" — map the two dates in that row.
        rowtxt = " ".join(_s(ws.cell(row=r, column=c).value)
                          for c in range(c0, c1 + 1)
                          if ws.cell(row=r, column=c).value not in (None, ""))
        if "period covered" in rowtxt:
            dates = [c for c in range(c0, c1 + 1)
                     if isinstance(ws.cell(row=r, column=c).value, (date, datetime))]
            if len(dates) >= 2:
                tok[(r, dates[0])] = "{{ period_start }}"
                tok[(r, dates[1])] = "{{ period_end }}"
    return tok


# NOTE: contains literal % ("width:100%") and {{ tokens }} → assembled with
# .replace(), never %-format / str.format (both choke on those).
_EXCEL_TEMPLATE_SHELL = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
  @page { size: letter; margin: 0.6in; }
  body { font-family: Helvetica, Arial, sans-serif; color:#16202b; font-size:11pt; }
  table.xl { width:100%; border-collapse:collapse; table-layout:fixed; }
  table.xl td { padding:2px 5px; vertical-align:top; line-height:1.3; word-wrap:break-word; }
</style></head>
<body>
__INVOICE_TABLE__
</body></html>"""


def _sheet_to_invoice_html(ws, max_rows: int = 70, max_cols: int = 16) -> str:
    """Faithful HTML rendering of an invoice sheet: column widths, bold, alignment,
    number formats and merged cells preserved; labelled data cells tokenized."""
    r0, r1, c0, c1 = _content_bounds(ws, max_rows, max_cols)
    if r1 < r0 or c1 < c0:
        return ""

    # Merged cells → colspan/rowspan on the top-left, skip the covered cells.
    covered: set = set()
    span: dict = {}
    for m in ws.merged_cells.ranges:
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != (m.min_row, m.min_col):
                    covered.add((r, c))
        span[(m.min_row, m.min_col)] = (m.max_row - m.min_row + 1, m.max_col - m.min_col + 1)

    tok = _build_token_map(ws, r0, r1, c0, c1, covered)

    # Column widths → percentages so proportions survive (table-layout:fixed).
    from openpyxl.utils import get_column_letter
    raw_w = []
    for c in range(c0, c1 + 1):
        cd = ws.column_dimensions.get(get_column_letter(c))
        raw_w.append(cd.width if (cd and cd.width) else 8.43)
    total_w = sum(raw_w) or 1.0
    colgroup = "".join(f'<col style="width:{w / total_w * 100:.2f}%">' for w in raw_w)

    rows_html = []
    for r in range(r0, r1 + 1):
        tds = []
        for c in range(c0, c1 + 1):
            if (r, c) in covered:
                continue
            cell = ws.cell(row=r, column=c)
            if (r, c) in tok:
                inner = tok[(r, c)]                       # token braces, unescaped
            else:
                inner = _html_escape(_excel_disp(cell.value, cell.number_format))
            styles = []
            al = cell.alignment.horizontal if cell.alignment else None
            if al in ("right", "center", "left"):
                styles.append(f"text-align:{al}")
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                styles.append("text-align:right")
            f = cell.font
            if f and f.bold:
                styles.append("font-weight:700")
            if f and f.size:
                styles.append(f"font-size:{max(9, min(18, round(float(f.size) * 0.8)))}px")
            attrs = ""
            sp = span.get((r, c))
            if sp:
                if sp[0] > 1:
                    attrs += f' rowspan="{sp[0]}"'
                if sp[1] > 1:
                    attrs += f' colspan="{sp[1]}"'
            style = f' style="{";".join(styles)}"' if styles else ""
            tds.append(f"<td{attrs}{style}>{inner or '&nbsp;'}</td>")
        rows_html.append("<tr>" + "".join(tds) + "</tr>")

    table = f'<table class="xl"><colgroup>{colgroup}</colgroup>{"".join(rows_html)}</table>'
    return _EXCEL_TEMPLATE_SHELL.replace("__INVOICE_TABLE__", table)


def excel_to_template_html(file_bytes: bytes) -> tuple[Optional[str], Optional[str]]:
    """Find the invoice sheet in an uploaded workbook and convert it to faithful,
    tokenized HTML (the operator's invoice-template seed).

    Returns (sheet_name, html). (None, None) when the workbook can't be opened or
    has no sheets; (sheet_name, None) when the sheet is empty. data_only (not
    read_only) so we get both cached values AND cell styles/number formats.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:  # noqa: BLE001
        logger.warning("excel_to_template_html: cannot open workbook: %s", e)
        return None, None
    try:
        ws = find_invoice_sheet(wb) or (wb.worksheets[0] if wb.worksheets else None)
        if ws is None:
            return None, None
        html = _sheet_to_invoice_html(ws)
        return ws.title, (html or None)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _detect_billing_model(wb, periods: list[Period]) -> tuple[str, dict]:
    """Return (model, template_dict). Template dict carries static invoice header
    text + the flat/fixed amount when applicable."""
    template: dict[str, Any] = {}
    tws = _find_template_sheet(wb)
    blob = ""
    if tws is not None:
        tgrid = _grid(tws, max_rows=50)
        # Lift static invoice header fields + the flat/fixed amount.
        template = _lift_template_header(tgrid)
        blob = " \n ".join(_s(c) for row in tgrid for c in row if c is not None)

    if "fixed monthly budget" in blob:
        return "fixed_budget", template
    if "billing at the estimated rate" in blob or "flat rate" in blob:
        return "flat_rate", template
    # A populated flat-rate column in the ledger also signals a flat-rate plan.
    if any(p.flat_rate for p in periods):
        return "flat_rate", template
    if "% of total array" in blob or "of total array" in blob:
        return "percent_of_array", template
    return "percent_of_array", template


_MONEY_RE = re.compile(r"\$?\s*([0-9][0-9,]*(?:\.[0-9]+)?)")


def _lift_template_header(tgrid: list[list[Any]]) -> dict:
    """Pull the static invoice header (contact block, attn, payable-to, the
    fixed/flat amount) out of a Template sheet so regenerated invoices keep the
    operator's exact wording."""
    out: dict[str, Any] = {}
    lines = [[_clean(c) for c in row] for row in tgrid]
    flat: list[str] = []
    for row in lines:
        for c in row:
            if c:
                flat.append(c)
    joined_lc = " \n ".join(c.lower() for c in flat)

    # Operator contact block — the first "<Company>, LLC" style line.
    for c in flat:
        if c and ("llc" in c.lower() or "enterprises" in c.lower()) and "payable" not in c.lower():
            out.setdefault("operator", c)
            break
    for c in flat:
        cl = c.lower()
        if cl.startswith("attn") or "attn:" in cl:
            out["attn"] = c
        if "payable to" in cl:
            out["payable_to"] = c
        if cl.startswith("ph:") or cl.startswith("ph "):
            out["phone"] = c
        if cl.startswith("email:"):
            out["email"] = c.split(":", 1)[1].strip()

    # Fixed/flat amount: pick by label specificity. "Amount Due/Owed" beats the
    # "estimated rate" line; both beat "fixed monthly budget". Take the FIRST
    # number after the label (the value column) — the rightmost can be a stray
    # scratch cell (e.g. Valley Cares Template M29=1841.34 next to E29=2150).
    priority = {"amount due": 3, "amount owed": 3, "final amount": 3,
                "fixed monthly budget": 2, "billing at the estimated rate": 1}
    best_amt: Optional[float] = None
    best_pri = 0
    for row in tgrid:
        for ci, c in enumerate(row):
            cl = _s(c)
            for label, pri in priority.items():
                if label in cl:
                    amt = _row_first_number(row, ci + 1)
                    if amt is not None and pri >= best_pri:
                        best_amt, best_pri = amt, pri
                    break
    if best_amt is not None:
        out["fixed_amount"] = best_amt
    if "title" not in out:
        out["title"] = "Invoice - Solar Power Generation"
    return out


def _row_first_number(row: list[Any], after_col: int) -> Optional[float]:
    """The first numeric value in a row at/after a column (the value column)."""
    for ci in range(after_col, len(row)):
        n = _num(row[ci])
        if n is not None:
            return n
    return None


# ─── invoice computation ────────────────────────────────────────────────────

def compute_invoice(match_customer_kwh: Optional[float], tariff: Optional[float],
                    adder: Optional[float], billing_rate: Optional[float],
                    billing_model: str, fixed_amount: Optional[float]) -> dict:
    """Compute the canonical invoice numbers for one period.

    Mirrors the HCT 'Template' sheet math:
      net_value       = kwh × tariff
      incentive_value = kwh × adder
      solar_value     = kwh × (tariff + adder)
      billed_value    = solar_value × billing_rate
      solar_savings   = solar_value − billed_value
      amount_owed     = fixed/flat amount when set, else billed_value
    """
    kwh = match_customer_kwh or 0.0
    tariff = tariff or 0.0
    adder = adder or 0.0
    rate = billing_rate if billing_rate is not None else 0.9
    net_value = kwh * tariff
    incentive_value = kwh * adder
    solar_value = kwh * (tariff + adder)
    billed_value = solar_value * rate
    solar_savings = solar_value - billed_value
    if billing_model in ("fixed_budget", "flat_rate") and fixed_amount is not None:
        amount_owed = fixed_amount
    else:
        amount_owed = billed_value
    return {
        "kwh": round(kwh, 2),
        "tariff": tariff,
        "adder": adder,
        "billing_rate": rate,
        "net_value": round(net_value, 2),
        "incentive_value": round(incentive_value, 2),
        "solar_value": round(solar_value, 2),
        "billed_value": round(billed_value, 2),
        "solar_savings": round(solar_savings, 2),
        "amount_owed": round(amount_owed, 2),
    }


def _project_totals(periods: list[Period]) -> dict:
    return {
        "total_array_kwh": round(sum(p.array_kwh or 0 for p in periods), 1),
        "total_customer_kwh": round(sum(p.customer_kwh or 0 for p in periods), 1),
        "total_solar_value": round(sum(p.value or 0 for p in periods), 2),
        "total_savings": round(sum(p.savings or 0 for p in periods), 2),
        "period_count": len([p for p in periods if (p.customer_kwh or p.array_kwh)]),
    }


# ─── public entry point ─────────────────────────────────────────────────────

def match_billing_workbook(file_bytes: bytes, *, allow_llm: bool = True) -> BillingMatch:
    """Match an uploaded billing workbook against the HCT family schema.

    Pure + offline except for the optional LLM fallback (controlled by
    `allow_llm`; tests pass allow_llm=False for determinism).
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return BillingMatch(matched=False, confidence=0.0, source="none",
                            warnings=[f"Could not open workbook: {e}"])

    try:
        candidates = _candidate_sheets(wb)
        if not candidates:
            wb.close()
            if allow_llm:
                return _llm_fallback(file_bytes)
            return BillingMatch(matched=False, confidence=0.0, source="none",
                                warnings=["No billing-ledger sheet recognized."])

        # Pick the CURRENT ledger when several exist: parse each candidate and
        # choose by most-recent period end, then period count, then header score.
        scored: list[tuple] = []
        for (name, hrow, fmap_c, score_c) in candidates:
            grid_c = _grid(wb[name], max_rows=200)
            periods_c = _parse_periods(grid_c, hrow, fmap_c)
            latest_c = _latest_period(periods_c)
            recency = latest_c.end if (latest_c and latest_c.end) else date.min
            scored.append((recency, len(periods_c), score_c,
                           name, hrow, fmap_c, score_c, grid_c, periods_c))
        scored.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
        chosen = scored[0]
        (sheet_name, header_row, fmap, header_score, grid, periods) = (
            chosen[3], chosen[4], chosen[5], chosen[6], chosen[7], chosen[8])

        meta = _extract_metadata(grid, header_row)
        billing_model, template = _detect_billing_model(wb, periods)
        wb.close()

        allocation_pct = meta.get("allocation_pct")
        billing_rate = meta.get("billing_rate")
        latest = _latest_period(periods)

        customer = {
            "name": meta.get("name"),
            "address": meta.get("address"),
            "acct": meta.get("acct"),
            "meter": meta.get("meter"),
            "email": meta.get("email") or template.get("email"),
        }

        computed = {}
        if latest is not None:
            computed = compute_invoice(
                latest.customer_kwh, latest.tariff, latest.adder,
                billing_rate, billing_model, template.get("fixed_amount"),
            )
            computed["invoice_number"] = (
                latest.end.strftime("%Y-%m") if latest.end else None
            )
            computed["period_start"] = latest.start.isoformat() if latest.start else None
            computed["period_end"] = latest.end.isoformat() if latest.end else None
            computed["month"] = latest.month

        # Confidence: blend header score with how much metadata we recovered.
        meta_score = sum(1 for k in ("name", "allocation_pct", "billing_rate")
                         if meta.get(k) is not None) / 3.0
        confidence = round(min(1.0, 0.6 * header_score + 0.4 * meta_score), 3)

        warnings: list[str] = []
        if not customer["name"]:
            warnings.append("Customer name not found — confirm it before sending.")
        if billing_rate is None:
            warnings.append("Billing rate (price factor) not found — defaulting to 90%.")
        if latest is None:
            warnings.append("No monthly data rows with generation were found.")

        match = BillingMatch(
            matched=confidence >= LLM_FALLBACK_THRESHOLD and latest is not None,
            confidence=confidence,
            source="schema",
            data_sheet=sheet_name,
            customer=customer,
            allocation_pct=allocation_pct,
            billing_rate=billing_rate,
            billing_model=billing_model,
            periods=periods,
            latest_period=latest,
            template=template,
            computed_invoice=computed,
            field_map={k: v for k, v in fmap.items()},
            project_totals=_project_totals(periods),
            warnings=warnings,
        )

        if not match.matched and allow_llm:
            llm = _llm_fallback(file_bytes)
            if llm.matched:
                return llm
        return match
    except Exception as e:  # noqa: BLE001
        logger.exception("billing matcher failed")
        try:
            wb.close()
        except Exception:
            pass
        if allow_llm:
            return _llm_fallback(file_bytes)
        return BillingMatch(matched=False, confidence=0.0, source="none",
                            warnings=[f"Matcher error: {e}"])


# ─── LLM fallback ───────────────────────────────────────────────────────────

_LLM_PROMPT = (
    "You are reading a solar net-metering BILLING spreadsheet. Extract the "
    "customer and their latest monthly billing period into EXACTLY this JSON:\n"
    '{"customer":{"name":"","address":null,"acct":null,"meter":null,"email":null},'
    '"allocation_pct":null,"billing_rate":null,'
    '"billing_model":"percent_of_array|fixed_budget|flat_rate",'
    '"latest_period":{"month":null,"period_start":"YYYY-MM-DD","period_end":"YYYY-MM-DD",'
    '"customer_kwh":0,"array_kwh":0,"tariff":0,"adder":0}}\n'
    "Rules: billing_rate is a fraction (90% → 0.9). allocation_pct is a fraction. "
    "Pick the period with the most recent end date that has kWh. "
    "Return ONLY JSON, no prose."
)


def _llm_fallback(file_bytes: bytes) -> BillingMatch:
    """Best-effort schema mapping via the LLM extractor used by api/ingest.py.
    Returns an unmatched BillingMatch if no LLM key is configured or it fails."""
    try:
        import json
        import os
        from ..ingest import _xlsx_to_text, _extract_json_block
        import httpx

        anthropic_key = os.getenv("ANTHROPIC_API_KEY")
        openai_key = os.getenv("OPENAI_API_KEY")
        if not (anthropic_key or openai_key):
            return BillingMatch(matched=False, confidence=0.0, source="none",
                                warnings=["Schema not recognized and no LLM key configured."])

        text = _xlsx_to_text(file_bytes)[:60_000]
        content = ""
        if anthropic_key:
            resp = httpx.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": anthropic_key,
                         "anthropic-version": "2023-06-01",
                         "content-type": "application/json"},
                json={"model": os.getenv("INGEST_LLM_MODEL", "claude-sonnet-4-5"),
                      "max_tokens": 2048,
                      "messages": [
                          {"role": "user", "content": f"{_LLM_PROMPT}\n\n{text}"},
                          {"role": "assistant", "content": "{"}]},
                timeout=60.0,
            )
            resp.raise_for_status()
            body = resp.json()
            content = "{" + "".join(b.get("text", "") for b in body.get("content", []))
        else:
            resp = httpx.post(
                "https://api.openai.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {openai_key}",
                         "content-type": "application/json"},
                json={"model": os.getenv("INGEST_OPENAI_MODEL", "gpt-4o-mini"),
                      "response_format": {"type": "json_object"},
                      "messages": [{"role": "user", "content": f"{_LLM_PROMPT}\n\n{text}"}]},
                timeout=60.0,
            )
            resp.raise_for_status()
            content = resp.json()["choices"][0]["message"]["content"]

        data = _extract_json_block(content)
        lp = data.get("latest_period") or {}
        latest = Period(
            month=lp.get("month"),
            start=_as_date(lp.get("period_start")),
            end=_as_date(lp.get("period_end")),
            array_kwh=_num(lp.get("array_kwh")),
            customer_kwh=_num(lp.get("customer_kwh")),
            tariff=_num(lp.get("tariff")),
            adder=_num(lp.get("adder")),
        )
        billing_rate = _num(data.get("billing_rate"))
        model = data.get("billing_model") or "percent_of_array"
        computed = compute_invoice(latest.customer_kwh, latest.tariff, latest.adder,
                                   billing_rate, model, None)
        computed["invoice_number"] = latest.end.strftime("%Y-%m") if latest.end else None
        computed["period_start"] = latest.start.isoformat() if latest.start else None
        computed["period_end"] = latest.end.isoformat() if latest.end else None
        return BillingMatch(
            matched=bool((data.get("customer") or {}).get("name")),
            confidence=0.6, source="llm",
            customer=data.get("customer") or {},
            allocation_pct=_num(data.get("allocation_pct")),
            billing_rate=billing_rate,
            billing_model=model,
            periods=[latest] if (latest.customer_kwh or latest.array_kwh) else [],
            latest_period=latest if (latest.customer_kwh or latest.array_kwh) else None,
            computed_invoice=computed,
            warnings=["Matched via AI fallback — review the fields carefully."],
        )
    except Exception as e:  # noqa: BLE001
        logger.warning("LLM billing fallback failed: %s", e)
        return BillingMatch(matched=False, confidence=0.0, source="none",
                            warnings=[f"AI fallback failed: {e}"])
