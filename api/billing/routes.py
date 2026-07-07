"""
Array Operator — automatic billing-report endpoints.

Mounted under /v1/array-operator/billing. Authed as the calling tenant via the
same session bearer used everywhere else (account.tenant_from_session).

  POST   /match                          multipart .xlsx → match preview (saves nothing)
  GET    /subscriptions                  list this tenant's report schedules
  POST   /subscriptions                  multipart: create/replace a schedule (+workbook)
  PATCH  /subscriptions/{id}             edit cadence / slider / formats / emails
  DELETE /subscriptions/{id}             soft-delete a schedule
  POST   /subscriptions/{id}/send-now    test/manual send (test=1 forces to_me)
  GET    /subscriptions/{id}/preview     stream invoice|summary as pdf|xlsx
  GET    /subscriptions/{id}/trends      multi-year billing trends (JSON, CONTRACT 1)

THREE SPREADSHEET SYSTEMS — keep them straight (they're easy to conflate):
  1. api/writers/gmcs_writer.py        WRITES the NEPOOL-GIS *GMCS filing*
                                       workbook — the NEPOOL Operator product's
                                       quarterly output. NOT a customer invoice.
  2. api/billing/matcher.py            READS an uploaded *customer billing*
                                       workbook (HCT family) → BillingMatch.
                                       Powers POST /match and onboarding.
  3. api/billing/invoice_writer.py     WRITES the customer's invoice back into
                                       THEIR OWN uploaded format (loads the
                                       stored original, populates the period,
                                       preserves styling/formulas/Template).
                                       Served by GET /preview?kind=invoice&fmt=xlsx
                                       for workbook subs.
Manual (typed-in) customers have no uploaded workbook, so their invoice is the
standard generated one (api/billing/invoice.py).
"""
from __future__ import annotations

import csv
import io
import json
import logging
import re as _re
from datetime import datetime, timedelta, date
from typing import Optional

from fastapi import APIRouter, Header, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from sqlalchemy import select, func, or_
from sqlalchemy.orm import selectinload

from ..db import SessionLocal
from ..models import BillingReportSubscription, Client, ReportDraft
from ..account import tenant_from_session, require_not_demo
from .matcher import match_billing_workbook
from .delivery import deliver_subscription, build_match, generate_files, next_send_at

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/v1/array-operator/billing", tags=["array-operator-billing"])

MAX_UPLOAD_BYTES = 8 * 1024 * 1024  # 8 MB — these workbooks are tens of KB
MAX_PDF_BYTES = 12 * 1024 * 1024    # GMP invoice PDFs
VALID_MODES = {"to_me", "to_client", "to_both"}
VALID_CADENCE = {"monthly", "quarterly"}
VALID_FORMATS = {"pdf", "xlsx"}
VALID_DELIVERY = {"approval", "auto"}

# File-type magic bytes (checked before extension so mis-named files are caught)
_MAGIC_PDF  = b"%PDF"
_MAGIC_XLSX = b"PK\x03\x04"   # ZIP / OpenXML (.xlsx, .xlsm, .docx, …)
_MAGIC_XLS  = b"\xd0\xcf\x11\xe0"  # OLE2 compound doc (.xls, .doc, …)


def _resolved_pricing_fields(s, pricing_ctx=None) -> dict:
    """The pricing actually applied to this customer (auto-resolved net rate +
    discount + provenance), for the UI card. Best-effort; never raises.
    pricing_ctx (delivery.build_pricing_ctx) batches the DB work for list
    callers — without it each row costs its own session + queries (17s at 800
    offtakers)."""
    try:
        from .delivery import resolve_discount_pricing
        p = resolve_discount_pricing(s, ctx=pricing_ctx)
        return {
            "resolved_net_rate": round(p["net_rate"], 5),
            "resolved_discount_pct": round(p["discount_pct"], 5),
            "resolved_effective_rate": p["effective_rate"],
            "resolved_net_source": p["net_source"],
            "resolved_net_note": p.get("net_rate_note"),
        }
    except Exception:  # noqa: BLE001
        return {}


def _sub_dict(s: BillingReportSubscription, pricing_ctx=None) -> dict:
    return {
        "id": s.id,
        "customer_name": s.customer_name,
        "client_id": s.client_id,
        "array_id": getattr(s, "array_id", None),
        "utility_account_id": getattr(s, "utility_account_id", None),
        "allocation_pct": getattr(s, "allocation_pct", None),
        "array_allocations": getattr(s, "array_allocations", None),
        # GMP allocation share for the bill-accuracy cross-check — DISTINCT from
        # allocation_pct (the billing multiplier). Surfaced so the setup/edit form
        # can pre-fill it. NULL → the check falls back to allocation_pct.
        "array_share_pct": getattr(s, "array_share_pct", None),
        # Per-offtaker cross-check variance threshold (percentage points, Bruce
        # 2026-07-07). NULL → the fleet default SHARE_VARIANCE_THRESHOLD_PCT; the
        # setup/edit form pre-fills this and shows the effective default when blank.
        "crosscheck_threshold_pct": getattr(s, "crosscheck_threshold_pct", None),
        "billing_model": s.billing_model,
        "rate_per_kwh": getattr(s, "rate_per_kwh", None),
        "discount_pct": getattr(s, "discount_pct", None),
        "net_rate_per_kwh": getattr(s, "net_rate_per_kwh", None),
        # The pricing actually applied (auto-resolved net rate + discount, with
        # provenance) so the card can SHOW the auto rate instead of a blank box.
        **_resolved_pricing_fields(s, pricing_ctx),
        "auto_attach_gmp": getattr(s, "auto_attach_gmp", False),
        "cadence": s.cadence,
        "annual_trueup": s.annual_trueup,
        "delivery_mode": getattr(s, "delivery_mode", "approval") or "approval",
        "send_mode": s.send_mode,
        "client_email": s.client_email,
        "cc_emails": s.cc_emails,
        "operator_email": s.operator_email,
        "formats": s.formats or ["pdf"],
        "include_summary": s.include_summary,
        "enabled": s.enabled,
        "source_filename": s.source_filename,
        "last_sent_at": s.last_sent_at.isoformat() if s.last_sent_at else None,
        "next_send_at": s.next_send_at.isoformat() if s.next_send_at else None,
        # The exactly-once period stamp — the redesign's provider progress bars
        # compare it against the pipeline's last period for an exact sent count.
        "last_sent_period_end": getattr(s, "last_sent_period_end", None),
        "last_invoice_number": s.last_invoice_number,
        "invoice_number_start": getattr(s, "invoice_number_start", None),
        "invoice_number_next": getattr(s, "invoice_number_next", None),
        "budget_amount_usd": getattr(s, "budget_amount_usd", None),
        # A trimmed preview of the parsed workbook for the UI card.
        "preview": (s.parsed_map or {}).get("computed_invoice") if s.parsed_map else None,
    }


async def _read_capped(file: UploadFile, max_bytes: int) -> bytes:
    """Read an upload into memory but NEVER buffer more than max_bytes (+1 chunk):
    stop and 413 as soon as the body exceeds the cap, instead of loading the whole
    (possibly multi-GB) body into one bytes object and checking size afterwards —
    which let an authenticated caller OOM the shared worker (#33)."""
    declared = getattr(file, "size", None)
    if declared is not None and declared > max_bytes:
        raise HTTPException(413, f"File too large (max {max_bytes // (1024 * 1024)} MB)")
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise HTTPException(413, f"File too large (max {max_bytes // (1024 * 1024)} MB)")
        chunks.append(chunk)
    return b"".join(chunks)


def _reject_zip_bomb(data: bytes, max_uncompressed: int = 200 * 1024 * 1024) -> None:
    """An .xlsx is a ZIP; a tiny upload can inflate to gigabytes and OOM the shared
    worker when openpyxl parses it (#34). Reject BEFORE parsing when the zip's
    declared uncompressed size exceeds a sane bound. Best-effort: a non-zip /
    unreadable central directory just passes (the real parse will reject it)."""
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            total = sum(getattr(i, "file_size", 0) or 0 for i in z.infolist())
    except Exception:  # noqa: BLE001 — not a zip / unreadable → let the parser judge
        return
    if total > max_uncompressed:
        raise HTTPException(
            413, "That spreadsheet expands to too much data to process safely.")


def _xlsx_formula_values_missing(data: bytes) -> bool:
    """True when the workbook has formula cells with NO cached values (e.g. exported
    by Google Sheets / LibreOffice) — so data_only reads come back blank and a seeded
    invoice template would show an empty Amount Due (#11). Best-effort, read-only,
    bounded; False on any error (don't block a normal upload on a probe hiccup)."""
    try:
        import openpyxl
        wb_f = openpyxl.load_workbook(io.BytesIO(data), data_only=False, read_only=True)
        wb_v = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        scanned = 0
        for ws_f in wb_f.worksheets:
            ws_v = wb_v[ws_f.title]
            for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows()):
                for cf, cv in zip(row_f, row_v):
                    scanned += 1
                    if scanned > 50000:      # bound the scan; huge sheets pass
                        return False
                    val = cf.value
                    if isinstance(val, str) and val.startswith("=") and cv.value is None:
                        return True
        return False
    except Exception:  # noqa: BLE001
        return False


async def _read_upload(file: UploadFile) -> bytes:
    data = await _read_capped(file, MAX_UPLOAD_BYTES)
    if not data:
        raise HTTPException(400, "The uploaded file is empty")
    # Detect actual file type by magic bytes before trusting the extension.
    if data[:4] == _MAGIC_PDF:
        raise HTTPException(
            400, "That's a PDF — please upload the Excel (.xlsx) workbook "
                 "(the HCT Sun spreadsheet, not a printed copy)")
    name = (file.filename or "").lower()
    is_excel = data[:4] in (_MAGIC_XLSX[:4], _MAGIC_XLS)
    if not (name.endswith(".xlsx") or name.endswith(".xls") or is_excel):
        raise HTTPException(400, "Upload an .xlsx billing workbook")
    _reject_zip_bomb(data)   # #34: refuse a decompression bomb before openpyxl
    return data


# ─── match preview ──────────────────────────────────────────────────────────

@router.post("/match")
async def billing_match(file: UploadFile = File(...),
                        authorization: Optional[str] = Header(default=None)):
    """Parse an uploaded billing workbook and return what we recognized. Saves
    nothing — this powers the upload→preview→confirm step in the UI."""
    tenant_from_session(authorization)  # auth only
    data = await _read_upload(file)
    match = match_billing_workbook(data)
    return {"ok": True, "filename": file.filename, "match": match.to_dict()}


# ─── GMP utility-bill selector (offtaker binding) ───────────────────────────

@router.get("/utility-accounts")
def list_utility_accounts(authorization: Optional[str] = Header(default=None)):
    """List this tenant's GMP and VEC/SmartHub utility accounts so the add-offtaker
    UI can let the operator SELECT the utility bill that connects to an offtaker.

    GMP offtaker invoices are computed EXCLUSIVELY from the chosen account's utility
    PAPER BILLS (Bill.kwh_generated per billing period) — never vendor/inverter
    data. VEC/SmartHub accounts carry no EXCESS+credit breakdown, so a VEC offtaker
    is priced as allocation_pct × the array's MEASURED generation × an operator-
    entered net rate (delivery enforces that rate; see build_manual_match).

    This endpoint surfaces each account with the array it feeds and a summary of the
    bills we hold (count + latest period + that period's kWh). The bill summary is
    GMP-shaped (Bill.kwh_generated, always null for SmartHub) so a VEC account shows
    empty bill stats — that's acceptable; the account still appears so it can be
    bound. The `provider` field lets the frontend label GMP vs VEC correctly.
    """
    from ..models import UtilityAccount, Bill, Array
    from ..adapters.smarthub import ALL_SMARTHUB_PROVIDERS

    t = tenant_from_session(authorization)
    out = []
    with SessionLocal() as db:
        accts = db.execute(
            select(UtilityAccount).where(
                UtilityAccount.tenant_id == t.id,
                or_(UtilityAccount.provider == "gmp",
                    UtilityAccount.provider.in_(sorted(ALL_SMARTHUB_PROVIDERS))),
                UtilityAccount.deleted_at.is_(None),
            ).order_by(UtilityAccount.nickname, UtilityAccount.account_number)
        ).scalars().all()
        # Batched bill summary — the per-account latest-bill + count pair was
        # ~2,500 queries at 828 accounts (5.3s of the Reports tab's list-bundle;
        # caught at Anna scale). Two grouped queries + one arrays map instead.
        acct_ids = [a.id for a in accts]
        counts: dict[int, int] = {}
        latest_by_acct: dict[int, tuple] = {}   # account_id -> (period_end, kwh)
        if acct_ids:
            for aid, n in db.execute(
                    select(Bill.account_id, func.count(Bill.id))
                    .where(Bill.account_id.in_(acct_ids),
                           Bill.kwh_generated.isnot(None))
                    .group_by(Bill.account_id)):
                counts[aid] = int(n or 0)
            for aid, pe, kwh in db.execute(
                    select(Bill.account_id, Bill.period_end, Bill.kwh_generated)
                    .where(Bill.account_id.in_(acct_ids),
                           Bill.kwh_generated.isnot(None),
                           Bill.period_end.isnot(None))
                    .order_by(Bill.account_id, Bill.period_end.desc())):
                if aid not in latest_by_acct:
                    latest_by_acct[aid] = (pe, kwh)
        arr_ids = {a.array_id for a in accts if a.array_id}
        arr_names = ({aid: nm for aid, nm in db.execute(
                          select(Array.id, Array.name).where(Array.id.in_(arr_ids)))}
                     if arr_ids else {})
        for a in accts:
            pe, kwh = latest_by_acct.get(a.id, (None, None))
            out.append({
                "utility_account_id": a.id,
                "account_number": a.account_number,
                "nickname": a.nickname,
                "provider": a.provider,
                "array_id": a.array_id,
                "array_name": arr_names.get(a.array_id),
                "bill_count": counts.get(a.id, 0),
                "has_bill": pe is not None,
                "latest_period_end": pe.date().isoformat() if pe else None,
                "latest_period_label": pe.strftime("%Y-%m") if pe else None,
                "latest_kwh_generated": (int(kwh) if kwh is not None else None),
            })
    return {"ok": True, "utility_accounts": out}


@router.post("/utility-accounts/{utility_account_id}/vec-bill")
async def upload_vec_bill(utility_account_id: int,
                          file: UploadFile = File(...),
                          authorization: Optional[str] = Header(default=None)):
    """Upload a VEC/SmartHub bill PDF for a bound utility account → parse it into a
    settled net-meter Bill (kwh_sent_to_grid + solar_credit_usd + the bill's own
    credit rate). Once a parsed VEC bill exists, the offtaker invoice auto-prices
    from the bill exactly like a GMP offtaker — excess kWh × the bill's own rate, no
    operator-entered rate needed (see delivery.build_manual_match).

    Tenant-scoped: the account must belong to the caller's tenant. Caps at 12 MB and
    rejects anything that isn't a PDF (by %PDF magic). On a parse/guard failure
    returns HTTP 422 with the reason.
    """
    from ..adapters.vec_bill import ingest_vec_bill_pdf

    t = tenant_from_session(authorization)
    require_not_demo(t)
    raw = await _read_capped(file, MAX_PDF_BYTES)
    if not raw:
        raise HTTPException(400, "The uploaded file is empty")
    if raw[:4] != _MAGIC_PDF:
        raise HTTPException(415, "Upload a PDF bill.")
    with SessionLocal() as db:
        res = ingest_vec_bill_pdf(db, t.id, utility_account_id, raw)
        if not res.get("ok"):
            raise HTTPException(422, res.get("reason") or "Couldn't read that bill.")
        # A manual upload is its own bill-land path (no extension, no server pull) —
        # keep this account's offtaker generation-spreadsheet rows current too.
        from .sheet_tracker import maybe_append_for_account
        maybe_append_for_account(db, t.id, utility_account_id)
        db.commit()
        return res


# ─── subscriptions CRUD ─────────────────────────────────────────────────────

def _attach_template_fit_warnings(db, tenant_id, rows, sub_dicts) -> None:
    """Set d["template_fit_warning"] on each serialized offtaker whose EFFECTIVE
    invoice template mismatches its billing model. Cheap: one query for per-offtaker
    templates + one for the tenant default; the xlsx budget-cell check is cached by
    content hash, so a shared tenant template is parsed once for the whole list."""
    from ..models import OfftakerInvoiceTemplate, OfftakerSubscriptionTemplate
    from .repro.template_repro import template_has_budget_cell
    tenant_tpl = db.execute(select(OfftakerInvoiceTemplate).where(
        OfftakerInvoiceTemplate.tenant_id == tenant_id)).scalars().first()
    tenant_bytes = (bytes(tenant_tpl.file_bytes) if tenant_tpl and tenant_tpl.enabled
                    and tenant_tpl.file_bytes else None)
    sub_ids = [r.id for r in rows]
    per = {}
    if sub_ids:
        for st in db.execute(select(OfftakerSubscriptionTemplate).where(
                OfftakerSubscriptionTemplate.subscription_id.in_(sub_ids))).scalars():
            per[st.subscription_id] = st
    by_id = {r.id: r for r in rows}
    for d in sub_dicts:
        sub = by_id.get(d.get("id"))
        if sub is None:
            continue
        st = per.get(sub.id)
        eff_bytes = (bytes(st.file_bytes) if st and st.enabled and st.file_bytes
                     else tenant_bytes)
        if not eff_bytes:
            continue                       # no enabled template → standard invoice for all
        on_budget = getattr(sub, "budget_amount_usd", None) is not None
        has_budget = template_has_budget_cell(eff_bytes)
        if has_budget and not on_budget:
            d["template_fit_warning"] = (
                "This offtaker isn\u2019t on a fixed budget, but the assigned template "
                "has a \u201cFixed Monthly Budget Payment\u201d line \u2014 the invoice "
                "can\u2019t render in it and falls back to the standard format. Use a "
                "template without a budget line, or set a fixed budget for this offtaker.")
        elif on_budget and not has_budget:
            d["template_fit_warning"] = (
                "This offtaker is on a fixed budget, but the assigned template has no "
                "budget line \u2014 the budget amount won\u2019t show on their invoice.")


@router.get("/subscriptions")
def list_subscriptions(authorization: Optional[str] = Header(default=None)):
    from ..models import Tenant, UtilityAccount
    from .delivery import build_pricing_ctx
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        rows = db.execute(
            select(BillingReportSubscription)
            .where(BillingReportSubscription.tenant_id == t.id,
                   BillingReportSubscription.deleted_at.is_(None))
            .order_by(BillingReportSubscription.created_at.desc())
        ).scalars().all()
        # One batched pricing context instead of a session + 3 queries per row
        # (at 800 offtakers the per-row form took 17s of the tab's first paint).
        ctx = build_pricing_ctx(db, db.get(Tenant, t.id))
        subs = [_sub_dict(s, ctx) for s in rows]
        # Enrich each sub with the LINKED utility account's display name so the
        # card can say which GMP bill feeds the offtaker (the UI only had the id).
        ua_ids = {d["utility_account_id"] for d in subs if d.get("utility_account_id")}
        if ua_ids:
            amap = {a.id: (a.nickname or (f"GMP {a.account_number}" if a.account_number else None))
                    for a in db.execute(
                        select(UtilityAccount).where(UtilityAccount.id.in_(ua_ids))
                    ).scalars().all()}
            for d in subs:
                d["utility_account_name"] = amap.get(d.get("utility_account_id"))
        # The fleet-default cross-check variance threshold (Bruce 2026-07-07): the
        # setup/edit form shows "flags beyond X%" from this when the per-offtaker
        # override is blank. Lazy import keeps reconcile_bills out of the module
        # import cycle (it imports delivery lazily itself).
        # Template ↔ billing-model fit: flag offtakers whose EFFECTIVE invoice
        # template doesn't match their budget setting, so a silent fall-back to the
        # standard invoice (a non-budget offtaker on a fixed-budget template) is
        # visible up front instead of a mysterious "wrong-looking invoice".
        try:
            _attach_template_fit_warnings(db, t.id, rows, subs)
        except Exception:  # noqa: BLE001 — a warning must never break the list
            logger.exception("template-fit warning enrichment failed")
        from .reconcile_bills import SHARE_VARIANCE_THRESHOLD_PCT
        return {"ok": True, "subscriptions": subs,
                "crosscheck_threshold_default_pct": SHARE_VARIANCE_THRESHOLD_PCT}


@router.get("/list-bundle")
def list_bundle(authorization: Optional[str] = Header(default=None)):
    """One round-trip for the Reports / offtaker-list view.

    The tab needs three things to render the offtaker cards and their edit
    dropdowns: the subscriptions, the tenant's arrays, and the GMP utility
    accounts. The frontend used to fetch these as three parallel calls — and one
    of them was the *full fleet-tree* (per-inverter peer analysis), which is by
    far the heaviest. This folds all three into a single cheap call: the
    subscriptions + utility-accounts payloads are produced by the very same
    handlers (so the shapes stay byte-identical), and arrays are a LIGHT direct
    query (id / name / client) instead of the fleet-tree, since the offtaker
    dropdown only needs to name the arrays, not analyze them.
    """
    from ..models import Array

    t = tenant_from_session(authorization)
    arrays = []
    with SessionLocal() as db:
        rows = db.execute(
            select(Array)
            # Eager-load the client so the `a.client.name` read below doesn't fire
            # a lazy SELECT per array (one batched IN query instead of N). This
            # runs on every Reports-tab view.
            .options(selectinload(Array.client))
            .where(Array.tenant_id == t.id, Array.deleted_at.is_(None))
            .order_by(Array.name)
        ).scalars().all()
        arrays = [
            {
                "id": a.id,
                "name": a.name,
                "client_name": (a.client.name if (a.client_id and a.client) else None),
            }
            for a in rows
        ]
    # Reuse the existing endpoints' logic verbatim so the shapes never drift.
    subs = list_subscriptions(authorization)
    uacc = list_utility_accounts(authorization)
    return {
        "ok": True,
        "subscriptions": subs.get("subscriptions", []),
        # Fleet-default cross-check variance threshold (Bruce 2026-07-07): the
        # setup/edit form shows "flags beyond X%" when the per-offtaker override
        # is blank. Passed through from list_subscriptions so both load paths have it.
        "crosscheck_threshold_default_pct": subs.get("crosscheck_threshold_default_pct"),
        "arrays": arrays,
        "utility_accounts": uacc.get("utility_accounts", []),
    }


def _arrays_and_accounts_for_tenant(t) -> tuple[list[dict], list[dict]]:
    """Shared pick-list for the offtaker-matcher flows (bulk-import v2 + commit).

    Returns (arrays, utility_accounts) in the shapes offtaker_match.match_array
    expects AND the frontend needs to build correction dropdowns:
      arrays           = [{id, name}, ...]
      utility_accounts = [{utility_account_id, array_id, array_name, nickname,
                           provider, account_number, has_bill, utility_label}, ...]

    The utility_accounts payload is the SAME source as GET /utility-accounts (so
    the shapes never drift) — we call that handler and enrich each row with a
    precomputed `utility_label` for the review UI. Arrays are the light id/name
    query used by list-bundle. Tenant-scoped throughout.
    """
    from ..models import Array
    from .offtaker_match import _utility_label

    with SessionLocal() as db:
        rows = db.execute(
            select(Array)
            .where(Array.tenant_id == t.id, Array.deleted_at.is_(None))
            .order_by(Array.name)
        ).scalars().all()
        arrays = [{"id": a.id, "name": a.name} for a in rows]

    # Reuse the utility-accounts handler verbatim (it re-derives auth from the
    # tenant is not possible — it needs the header — so we rebuild the same query
    # here tenant-scoped rather than re-authing). Keep it byte-shaped like the
    # /utility-accounts payload plus account_number (already returned there).
    from ..models import UtilityAccount, Bill, Array as _Array
    from ..adapters.smarthub import ALL_SMARTHUB_PROVIDERS

    uaccts: list[dict] = []
    with SessionLocal() as db:
        accts = db.execute(
            select(UtilityAccount).where(
                UtilityAccount.tenant_id == t.id,
                or_(UtilityAccount.provider == "gmp",
                    UtilityAccount.provider.in_(sorted(ALL_SMARTHUB_PROVIDERS))),
                UtilityAccount.deleted_at.is_(None),
            ).order_by(UtilityAccount.nickname, UtilityAccount.account_number)
        ).scalars().all()
        for a in accts:
            latest = db.execute(
                select(Bill)
                .where(Bill.account_id == a.id,
                       Bill.kwh_generated.isnot(None),
                       Bill.period_end.isnot(None))
                .order_by(Bill.period_end.desc())
            ).scalars().first()
            arr = db.get(_Array, a.array_id) if a.array_id else None
            row = {
                "utility_account_id": a.id,
                "account_number": a.account_number,
                "nickname": a.nickname,
                "provider": a.provider,
                "array_id": a.array_id,
                "array_name": arr.name if arr else None,
                "has_bill": latest is not None,
            }
            row["utility_label"] = _utility_label(row)
            uaccts.append(row)
    return arrays, uaccts


def _parse_formats(raw: Optional[str]) -> list[str]:
    if not raw:
        return ["pdf"]
    try:
        vals = json.loads(raw) if raw.strip().startswith("[") else raw.split(",")
    except json.JSONDecodeError:
        vals = raw.split(",")
    fmts = [v.strip().lower() for v in vals if v.strip().lower() in VALID_FORMATS]
    return fmts or ["pdf"]


# Sanity ceiling for a $/kWh rate. VT solar value is ~$0.18/kWh; anything above
# this is almost certainly a units mistake (e.g. cents typed as dollars).
MAX_RATE_PER_KWH = 5.0


def _validate_rate(rate):
    """Coerce/validate an optional $/kWh rate. None passes through (no rate set).
    Rejects negatives and absurd values so a fat-fingered rate can't silently
    produce a wild invoice."""
    if rate is None:
        return None
    try:
        r = float(rate)
    except (TypeError, ValueError):
        raise HTTPException(400, "rate_per_kwh must be a number ($/kWh)")
    if r < 0 or r > MAX_RATE_PER_KWH:
        raise HTTPException(400, f"rate_per_kwh must be between 0 and {MAX_RATE_PER_KWH} $/kWh")
    return r


def _validate_discount(pct):
    """Validate an optional discount fraction in [0, 1). None passes through.
    Accepts a fraction (0.10 = 10% off). Rejects ≥1 (would zero/inverse the bill)."""
    if pct is None:
        return None
    try:
        d = float(pct)
    except (TypeError, ValueError):
        raise HTTPException(400, "discount_pct must be a number (fraction, e.g. 0.10 for 10%)")
    if not (0 <= d < 1):
        raise HTTPException(400, "discount_pct must be a fraction in [0, 1) — e.g. 0.10 for 10% off")
    return d


def _validate_array_share(pct):
    """Validate the optional array_share_pct — the offtaker's GMP allocation share
    of the array's group excess, used by the bill-accuracy cross-check (DISTINCT
    from allocation_pct, the billing multiplier). A fraction in (0, 1]; None passes
    through (the check falls back to allocation_pct). Mirrors the PATCH-path rule."""
    if pct is None:
        return None
    try:
        p = float(pct)
    except (TypeError, ValueError):
        raise HTTPException(400, "array_share_pct must be a number between 0 and 1")
    if not (0.0 < p <= 1.0):
        raise HTTPException(400, "array_share_pct must be a fraction in (0, 1] — e.g. 0.25 for 25%")
    return p


def _validate_crosscheck_threshold(pct):
    """Validate the optional crosscheck_threshold_pct — the per-offtaker variance
    threshold (PERCENTAGE POINTS) the bill-accuracy cross-check flags beyond
    (Bruce 2026-07-07, the knob that replaced the manual cross-check-share entry).
    A positive number; None passes through (the check uses the fleet default).
    Mirrors the PATCH-path rule."""
    if pct is None:
        return None
    try:
        p = float(pct)
    except (TypeError, ValueError):
        raise HTTPException(400, "crosscheck_threshold_pct must be a positive number")
    if not (0.0 < p <= 100.0):
        raise HTTPException(400, "crosscheck_threshold_pct must be a percentage-point value in (0, 100] — e.g. 0.1")
    return p


def _validate_invoice_start(v):
    """Validate the optional invoice_number_start (a whole number seed for
    sequential invoice numbering). None passes through. Mirrors the PATCH rule."""
    if v is None:
        return None
    try:
        n = int(v)
    except (TypeError, ValueError):
        raise HTTPException(400, "invoice_number_start must be a whole number")
    if not (0 <= n <= 9_999_999):
        raise HTTPException(400, "invoice_number_start must be between 0 and 9999999")
    return n


def _validate_budget(v):
    """Validate the optional budget_amount_usd (a fixed dollar total that overrides
    the calculated amount). None/blank passes through. Mirrors the PATCH rule."""
    if v is None or v == "":
        return None
    try:
        amt = float(v)
    except (TypeError, ValueError):
        raise HTTPException(400, "budget_amount_usd must be a number")
    if amt < 0:
        raise HTTPException(400, "budget_amount_usd can't be negative")
    return amt


def _sync_invoicing_quantity(tenant_id: str) -> None:
    """After an offtaker (BillingReportSubscription) is added or removed, keep the
    AO invoicing Stripe quantity in sync with the offtaker count. No-op unless the
    tenant is on the per-offtaker invoicing plan; best-effort so a Stripe hiccup
    never fails the offtaker mutation itself."""
    try:
        from ..stripe_helpers import reconcile_offtaker_quantity
        reconcile_offtaker_quantity(tenant_id)
    except Exception:  # noqa: BLE001 — the reconcile alerts on real failures
        pass


@router.post("/subscriptions")
async def create_subscription(
    file: Optional[UploadFile] = File(default=None),
    customer_name: Optional[str] = Form(default=None),
    array_id: Optional[int] = Form(default=None),
    utility_account_id: Optional[int] = Form(default=None),
    allocation_pct: Optional[float] = Form(default=None),
    array_allocations: Optional[str] = Form(default=None),
    array_share_pct: Optional[float] = Form(default=None),
    crosscheck_threshold_pct: Optional[float] = Form(default=None),
    invoice_number_start: Optional[int] = Form(default=None),
    budget_amount_usd: Optional[float] = Form(default=None),
    rate_per_kwh: Optional[float] = Form(default=None),
    discount_pct: Optional[float] = Form(default=None),
    net_rate_per_kwh: Optional[float] = Form(default=None),
    cadence: str = Form(default="monthly"),
    send_mode: str = Form(default="to_me"),
    delivery_mode: str = Form(default="approval"),
    client_email: Optional[str] = Form(default=None),
    cc_emails: Optional[str] = Form(default=None),
    operator_email: Optional[str] = Form(default=None),
    formats: Optional[str] = Form(default=None),
    include_summary: bool = Form(default=False),  # AO summary opt-in (Ford 2026-06-24)
    annual_trueup: bool = Form(default=False),
    enabled: bool = Form(default=True),
    authorization: Optional[str] = Header(default=None),
):
    """Create a report schedule for one customer. Two paths:

      * UPLOAD — a billing .xlsx is attached; the matcher recognizes it and the
        stored workbook bytes are the per-cycle source of truth.
      * MANUAL — no file; the operator typed the customer in (name + array_id +
        allocation_pct). No workbook is stored; allocation_pct × the array's
        period generation drives delivery/draft. (Paul Bozuwa's demo path.)
    """
    t = tenant_from_session(authorization)
    require_not_demo(t)

    if cadence not in VALID_CADENCE:
        raise HTTPException(400, "cadence must be monthly or quarterly")
    if send_mode not in VALID_MODES:
        raise HTTPException(400, "send_mode must be to_me, to_client, or to_both")
    if delivery_mode not in VALID_DELIVERY:
        raise HTTPException(400, "delivery_mode must be approval or auto")

    # array_share_pct + invoice_number_start are optional on create (they were
    # PATCH-only until now); validate here so a bad value fails before any write.
    array_share_val = _validate_array_share(array_share_pct)
    threshold_val = _validate_crosscheck_threshold(crosscheck_threshold_pct)
    invoice_start_val = _validate_invoice_start(invoice_number_start)

    if file is None:
        return await _create_manual_subscription(
            t, customer_name=customer_name, array_id=array_id,
            utility_account_id=utility_account_id,
            allocation_pct=allocation_pct, array_allocations=array_allocations,
            array_share_pct=array_share_val,
            crosscheck_threshold_pct=threshold_val,
            invoice_number_start=invoice_start_val,
            rate_per_kwh=rate_per_kwh,
            discount_pct=discount_pct, net_rate_per_kwh=net_rate_per_kwh,
            cadence=cadence, send_mode=send_mode,
            delivery_mode=delivery_mode, client_email=client_email,
            cc_emails=cc_emails, operator_email=operator_email, formats=formats,
            include_summary=include_summary, annual_trueup=annual_trueup,
            budget_amount_usd=budget_amount_usd, enabled=enabled)

    data = await _read_upload(file)
    match = match_billing_workbook(data)
    if not match.matched:
        raise HTTPException(
            422, "Couldn't recognize this as a billing workbook — "
                 "check the file or fill the fields in manually.")

    name = customer_name or match.customer.get("name") or "Customer"
    client_email = client_email or match.customer.get("email")

    with SessionLocal() as db:
        # Link or create the Client "underneath" the operator.
        client = db.execute(
            select(Client).where(Client.tenant_id == t.id, Client.name == name,
                                 Client.deleted_at.is_(None))
        ).scalar_one_or_none()
        if client is None:
            client = Client(tenant_id=t.id, name=name, contact_email=client_email,
                            active=True)
            db.add(client)
            db.flush()

        sub = BillingReportSubscription(
            tenant_id=t.id,
            client_id=client.id,
            customer_name=name,
            source_workbook=data,
            source_filename=file.filename,
            parsed_map=match.to_dict(),
            billing_model=match.billing_model,
            rate_per_kwh=rate_per_kwh,
            discount_pct=_validate_discount(discount_pct),
            net_rate_per_kwh=_validate_rate(net_rate_per_kwh),
            array_share_pct=array_share_val,
            crosscheck_threshold_pct=threshold_val,
            invoice_number_start=invoice_start_val,
            invoice_number_next=invoice_start_val,
            cadence=cadence,
            annual_trueup=annual_trueup,
            delivery_mode=delivery_mode,
            send_mode=send_mode,
            client_email=client_email,
            cc_emails=cc_emails,
            operator_email=operator_email or t.contact_email,
            formats=_parse_formats(formats),
            include_summary=include_summary,
            enabled=enabled,
            next_send_at=next_send_at(cadence),
        )
        db.add(sub)
        db.commit()
        _sync_invoicing_quantity(t.id)
        return {"ok": True, "subscription": _sub_dict(sub)}


async def _create_manual_subscription(
    t, *, customer_name, array_id, allocation_pct, array_allocations=None,
    utility_account_id=None, array_share_pct=None, crosscheck_threshold_pct=None,
    invoice_number_start=None, budget_amount_usd=None,
    rate_per_kwh, discount_pct,
    net_rate_per_kwh, cadence,
    send_mode, delivery_mode, client_email, cc_emails, operator_email, formats,
    include_summary, annual_trueup, enabled,
):
    """Create a workbook-less subscription from typed fields.

    THREE shapes (checked in priority order):
      * OFFTAKER ↔ UTILITY BILL — utility_account_id + allocation_pct. The
        offtaker's invoice is computed EXCLUSIVELY from that GMP account's utility
        PAPER BILLS (Bill.kwh_generated per period) — never vendor/inverter data,
        never the hourly interval data, and with NO fallback. This is the path
        Ford specified for offtaker reports.
      * single array  — array_id + allocation_pct (legacy, unchanged).
      * multi array    — array_allocations: JSON list of {array_id, allocation_pct}.
        The offtaker owns a share of several arrays; delivery sums each array's
        (period kWh × pct) into one combined invoice.
    """
    import json as _json
    from ..models import Array, UtilityAccount

    name = (customer_name or "").strip()
    if not name:
        raise HTTPException(400, "customer_name is required for a manual customer")
    budget_val = _validate_budget(budget_amount_usd)

    # ── ARRAY-FIRST resolution (Ford, 2026-07-01) ────────────────────────────
    # The offtaker binding model is ARRAY-FIRST with a utility-bill override: the
    # operator names the ARRAY, and we resolve WHICH utility bill invoices it FROM.
    # When array_id is given WITHOUT an explicit utility_account_id, look up the
    # array's linked GMP/VEC utility account(s) and pick the one to bill from:
    #   * prefer an account that actually HAS a settled bill;
    #   * if exactly one billable account, use it;
    #   * if MULTIPLE billable accounts and none is unambiguous, require an
    #     explicit utility_account_id override (400) — never guess which bill.
    # If resolved, we set utility_account_id and fall through to the OFFTAKER↔
    # UTILITY BILL path (a single, audited billing binding). If the array has NO
    # connected utility bill at all, we DON'T error — we fall through to the
    # legacy array-generation path below (allocation_pct × the array's measured
    # generation), which is a supported billing mode for generation-only arrays.
    # (An explicit utility_account_id always wins and skips this resolution.)
    if utility_account_id is None and array_id is not None and not array_allocations:
        from ..models import UtilityAccount as _UA, Array as _Arr, Bill
        from ..adapters import is_smarthub_provider as _is_sh
        with SessionLocal() as _db:
            _arr = _db.get(_Arr, array_id)
            if _arr is None or _arr.tenant_id != t.id or _arr.deleted_at is not None:
                raise HTTPException(404, f"Array {array_id} not found")
            _accts = _db.execute(
                select(_UA).where(
                    _UA.array_id == array_id,
                    _UA.tenant_id == t.id,
                    _UA.deleted_at.is_(None),
                )
            ).scalars().all()
            # Only GMP / VEC-SmartHub accounts can bill an offtaker invoice.
            _billable = [a for a in _accts
                         if (a.provider or "").lower() == "gmp"
                         or _is_sh((a.provider or "").lower())]
            if _billable:
                # Prefer accounts with a settled bill; if that narrows to exactly
                # one, use it. Else if a single billable account overall, use it.
                # Anything still ambiguous demands an explicit override.
                _with_bill = []
                for a in _billable:
                    _has = _db.execute(
                        select(func.count(Bill.id)).where(
                            Bill.account_id == a.id, Bill.kwh_generated.isnot(None))
                    ).scalar() or 0
                    if _has:
                        _with_bill.append(a)
                if len(_with_bill) == 1:
                    utility_account_id = _with_bill[0].id
                elif len(_with_bill) == 0 and len(_billable) == 1:
                    utility_account_id = _billable[0].id
                elif len(_billable) == 1:
                    utility_account_id = _billable[0].id
                else:
                    raise HTTPException(
                        400, "This array has multiple connected utility bills — "
                             "pass utility_account_id to choose which one invoices "
                             "this offtaker.")
            # else: no billable account → fall through to legacy generation path.

    # ── OFFTAKER ↔ UTILITY BILL path (highest priority) ──────────────────────
    if utility_account_id is not None:
        if allocation_pct is None:
            raise HTTPException(
                400, "allocation_pct is required when binding an offtaker to a utility bill")
        try:
            pct = float(allocation_pct)
        except (TypeError, ValueError):
            raise HTTPException(400, "allocation_pct must be a number between 0 and 1")
        if not (0.0 < pct <= 1.0):
            raise HTTPException(400, "allocation_pct must be a fraction in (0, 1] "
                                     "(e.g. 0.25 for 25%)")
        rate_val = _validate_rate(rate_per_kwh)
        disc_val = _validate_discount(discount_pct)
        net_val = _validate_rate(net_rate_per_kwh)
        share_val = _validate_array_share(array_share_pct)
        threshold_val = _validate_crosscheck_threshold(crosscheck_threshold_pct)
        inv_start_val = _validate_invoice_start(invoice_number_start)
        with SessionLocal() as db:
            acct = db.get(UtilityAccount, utility_account_id)
            if (acct is None or acct.tenant_id != t.id
                    or acct.deleted_at is not None):
                raise HTTPException(404, f"Utility account {utility_account_id} not found")
            from ..adapters import is_smarthub_provider
            _prov = (acct.provider or "").lower()
            if _prov != "gmp" and not is_smarthub_provider(_prov):
                raise HTTPException(
                    400, "Offtaker reports bind to a GMP or VEC/SmartHub utility "
                         "account (utility-bill data only).")
            # Self-heal the account ↔ array link (Bruce, 2026-07-03): a freshly
            # captured GMP account lands with array_id = NULL until it's matched
            # to an array, which left the add-offtaker bill picker empty even
            # though every bill was downloaded. When the operator explicitly
            # binds this account while creating an offtaker for a named array,
            # record that link so the group resolves silently ("Invoices
            # from: …") on the next add. Never re-links an account already tied
            # to a different array — an explicit prior binding wins.
            if acct.array_id is None and array_id is not None:
                _heal_arr = db.get(Array, array_id)
                if (_heal_arr is not None and _heal_arr.tenant_id == t.id
                        and _heal_arr.deleted_at is None):
                    acct.array_id = array_id
            # ── Sub-meter auto-routing (Ford 2026-07-07) ─────────────────────
            # If this offtaker bills off its OWN sub-account (distinct from the
            # array's HOST meter), that sub-account's excess ALREADY reflects the
            # offtaker's metered share of the net-meter group. So the operator's
            # ONE entered share is the GROUP share (array_share_pct, which real_math
            # bills as share x group excess), and allocation_pct is pinned to 1.0
            # (bill 100% of the sub-meter). We must NEVER also multiply the
            # sub-account's own excess by the share again -- that double-count is
            # the wrong-bill-audit bug. Percent-of-array offtakers (account IS the
            # host) are untouched: allocation_pct stays their share of the host.
            _host_id = None
            if acct.array_id is not None:
                _host_id = db.execute(
                    select(UtilityAccount.id).where(
                        UtilityAccount.array_id == acct.array_id,
                        UtilityAccount.deleted_at.is_(None))
                    .order_by(UtilityAccount.id)).scalars().first()
            _is_submeter = _host_id is not None and _host_id != utility_account_id
            if _is_submeter:
                if share_val is None:
                    share_val = pct          # the ONE entered share = the group share
                pct = 1.0                    # always bill 100% of the sub-meter
            # Quarterly cadence for a GMP (bill-priced) offtaker aggregates the
            # FULL quarter — delivery sums all three monthly bills and HOLDS the
            # invoice until every month's bill has landed (never bills short).
            # VEC/SmartHub model-A offtakers price a single month of measured
            # generation (no parsed bill), so quarterly stays blocked for them
            # until that path aggregates too — rather than under-bill 2 of 3
            # months.
            if (cadence or "monthly") == "quarterly" and is_smarthub_provider(_prov):
                raise HTTPException(
                    400, "Quarterly billing isn't available for VEC/SmartHub "
                         "offtakers yet — their invoices price a single month of "
                         "measured generation, so a quarterly invoice would bill "
                         "only one of the three months. Use monthly cadence.")
            # #7: offtaker allocations on ONE utility meter can't sum past 100%, or
            # the same excess is billed to two people. Sum the live offtakers already
            # bound to this account and reject if adding this share crosses 100%
            # (0.5pp rounding epsilon). Runs for both single-create and each
            # bulk-commit row (earlier rows are already committed, so the running
            # total stays correct across a batch).
            _existing_alloc = float(db.execute(
                select(func.coalesce(func.sum(BillingReportSubscription.allocation_pct), 0.0))
                .where(
                    BillingReportSubscription.tenant_id == t.id,
                    BillingReportSubscription.utility_account_id == utility_account_id,
                    BillingReportSubscription.deleted_at.is_(None),
                )
            ).scalar() or 0.0)
            if _existing_alloc + pct > 1.0 + 0.005:
                raise HTTPException(
                    409,
                    f"This would over-allocate the meter to {(_existing_alloc + pct) * 100:.0f}% "
                    f"— offtakers sharing one utility account can't sum past 100% or the "
                    f"meter's excess is billed twice. It's already at {_existing_alloc * 100:.0f}%.")
            client = db.execute(
                select(Client).where(Client.tenant_id == t.id, Client.name == name,
                                     Client.deleted_at.is_(None))
            ).scalar_one_or_none()
            if client is None:
                client = Client(tenant_id=t.id, name=name, contact_email=client_email,
                                active=True)
                db.add(client)
                db.flush()
            sub = BillingReportSubscription(
                tenant_id=t.id,
                client_id=client.id,
                customer_name=name,
                utility_account_id=utility_account_id,
                # Keep array_id populated (from the account's array) for list views,
                # but delivery uses the utility-bill path because utility_account_id
                # is set — it never reads vendor/array data for this offtaker.
                array_id=acct.array_id,
                allocation_pct=pct,
                array_allocations=None,
                array_share_pct=share_val,
                crosscheck_threshold_pct=threshold_val,
                invoice_number_start=inv_start_val,
                invoice_number_next=inv_start_val,
                budget_amount_usd=budget_val,
                rate_per_kwh=rate_val,
                discount_pct=disc_val,
                net_rate_per_kwh=net_val,
                source_workbook=None,
                source_filename=None,
                parsed_map=None,
                billing_model="percent_of_array",
                cadence=cadence,
                annual_trueup=annual_trueup,
                delivery_mode=delivery_mode,
                send_mode=send_mode,
                client_email=client_email,
                cc_emails=cc_emails,
                operator_email=operator_email or t.contact_email,
                formats=_parse_formats(formats),
                include_summary=include_summary,
                enabled=enabled,
                next_send_at=next_send_at(cadence),
            )
            db.add(sub)
            db.commit()
            _sync_invoicing_quantity(t.id)
            return {"ok": True, "subscription": _sub_dict(sub)}

    # Parse the optional multi-array allocations list.
    allocs: list[dict] = []
    if array_allocations:
        try:
            parsed = _json.loads(array_allocations)
        except (ValueError, TypeError):
            raise HTTPException(400, "array_allocations must be valid JSON")
        if not isinstance(parsed, list):
            raise HTTPException(400, "array_allocations must be a JSON list")
        for r in parsed:
            if not isinstance(r, dict):
                raise HTTPException(400, "each allocation must be an object")
            try:
                aid = int(r.get("array_id"))
            except (TypeError, ValueError):
                raise HTTPException(400, "each allocation needs a numeric array_id")
            try:
                p = float(r.get("allocation_pct"))
            except (TypeError, ValueError):
                raise HTTPException(400, "each allocation needs a numeric allocation_pct")
            if not (0.0 < p <= 1.0):
                raise HTTPException(400, "allocation_pct must be a fraction in (0, 1] "
                                         "(e.g. 0.25 for 25%)")
            allocs.append({"array_id": aid, "allocation_pct": p})
        if not allocs:
            raise HTTPException(400, "array_allocations had no usable rows")

    if not allocs:
        # Legacy single-array path.
        if array_id is None:
            raise HTTPException(400, "array_id is required for a manual customer")
        if allocation_pct is None:
            raise HTTPException(400, "allocation_pct is required for a manual customer")
        try:
            pct = float(allocation_pct)
        except (TypeError, ValueError):
            raise HTTPException(400, "allocation_pct must be a number between 0 and 1")
        if not (0.0 < pct <= 1.0):
            raise HTTPException(400, "allocation_pct must be a fraction in (0, 1] "
                                     "(e.g. 0.25 for 25%)")
    else:
        pct = None
    rate_val = _validate_rate(rate_per_kwh)
    disc_val = _validate_discount(discount_pct)
    net_val = _validate_rate(net_rate_per_kwh)
    share_val = _validate_array_share(array_share_pct)
    threshold_val = _validate_crosscheck_threshold(crosscheck_threshold_pct)
    inv_start_val = _validate_invoice_start(invoice_number_start)

    with SessionLocal() as db:
        # Validate every referenced array belongs to this tenant.
        aids_to_check = [a["array_id"] for a in allocs] if allocs else [array_id]
        for aid in aids_to_check:
            arr = db.get(Array, aid)
            if arr is None or arr.tenant_id != t.id or arr.deleted_at is not None:
                raise HTTPException(404, f"Array {aid} not found")

        client = db.execute(
            select(Client).where(Client.tenant_id == t.id, Client.name == name,
                                 Client.deleted_at.is_(None))
        ).scalar_one_or_none()
        if client is None:
            client = Client(tenant_id=t.id, name=name, contact_email=client_email,
                            active=True)
            db.add(client)
            db.flush()

        sub = BillingReportSubscription(
            tenant_id=t.id,
            client_id=client.id,
            customer_name=name,
            # Keep a single array_id/allocation_pct for back-compat + list views;
            # when multi-array, store the first as the representative and the full
            # list in array_allocations (which delivery prefers when present).
            array_id=(allocs[0]["array_id"] if allocs else array_id),
            allocation_pct=(allocs[0]["allocation_pct"] if allocs else pct),
            array_allocations=(allocs or None),
            array_share_pct=share_val,
            crosscheck_threshold_pct=threshold_val,
            invoice_number_start=inv_start_val,
            invoice_number_next=inv_start_val,
            budget_amount_usd=budget_val,
            rate_per_kwh=rate_val,
            discount_pct=disc_val,
            net_rate_per_kwh=net_val,
            source_workbook=None,
            source_filename=None,
            parsed_map=None,
            billing_model="percent_of_array",
            cadence=cadence,
            annual_trueup=annual_trueup,
            delivery_mode=delivery_mode,
            send_mode=send_mode,
            client_email=client_email,
            cc_emails=cc_emails,
            operator_email=operator_email or t.contact_email,
            formats=_parse_formats(formats),
            include_summary=include_summary,
            enabled=enabled,
            next_send_at=next_send_at(cadence),
        )
        db.add(sub)
        db.commit()
        _sync_invoicing_quantity(t.id)
        return {"ok": True, "subscription": _sub_dict(sub)}


class SubscriptionPatch(BaseModel):
    customer_name: Optional[str] = None
    cadence: Optional[str] = None
    delivery_mode: Optional[str] = None
    send_mode: Optional[str] = None
    client_email: Optional[str] = None
    cc_emails: Optional[str] = None
    operator_email: Optional[str] = None
    formats: Optional[list[str]] = None
    include_summary: Optional[bool] = None
    annual_trueup: Optional[bool] = None
    enabled: Optional[bool] = None
    # Redesigned Reports tab: inline-edit a manual customer's allocation % and
    # (re)assign the array. allocation_pct is a fraction in (0, 1]. array_id ties
    # the manual customer to the array whose generation is split.
    allocation_pct: Optional[float] = None
    array_id: Optional[int] = None
    # GMP allocation share for the bill-accuracy cross-check (fraction in (0, 1]).
    # Distinct from allocation_pct. Explicit null clears it (check falls back to
    # allocation_pct). Uses model_fields_set to tell "clear" from "omitted".
    array_share_pct: Optional[float] = None
    # Per-offtaker cross-check VARIANCE THRESHOLD (percentage points, Bruce
    # 2026-07-07): how far GMP's implied share may drift from the entered billing
    # share before the check flags. Explicit null clears it (falls back to the
    # fleet default SHARE_VARIANCE_THRESHOLD_PCT). model_fields_set tells
    # "clear" from "omitted".
    crosscheck_threshold_pct: Optional[float] = None
    # Re-bind the offtaker to a different GMP utility bill (the billing source).
    # Validated to a GMP account the operator owns; array_id is refreshed from it.
    utility_account_id: Optional[int] = None
    # Per-customer billing rate override ($/kWh). Send a number to set it, or
    # explicit null to CLEAR it (fall back to the operator's global rate). We
    # use model_fields_set in the handler to tell "null to clear" from "omitted".
    rate_per_kwh: Optional[float] = None
    # Discount-model overrides. Send a number to set, explicit null to clear
    # (falls back to the operator global, then the 10%/VT default).
    discount_pct: Optional[float] = None
    net_rate_per_kwh: Optional[float] = None
    # Per-customer 'auto-attach the captured GMP bill PDF' toggle.
    auto_attach_gmp: Optional[bool] = None
    # Sequential invoice numbering: the operator's starting invoice number. Setting
    # it (re)seeds the running counter so the next invoice uses it and each real send
    # adds 1. Explicit null clears it (back to the period-date number).
    invoice_number_start: Optional[int] = None
    # Budget billing: a fixed final amount that overrides the calculated Amount Due
    # (line items still show). Explicit null clears it (back to the calculated total).
    budget_amount_usd: Optional[float] = None


@router.patch("/subscriptions/{sub_id}")
def patch_subscription(sub_id: int, body: SubscriptionPatch,
                       authorization: Optional[str] = Header(default=None)):
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        if body.cadence is not None:
            if body.cadence not in VALID_CADENCE:
                raise HTTPException(400, "cadence must be monthly or quarterly")
            # Quarterly for a GMP (bill-priced) offtaker aggregates the full
            # quarter (all three bills summed; delivery holds until complete).
            # VEC/SmartHub model-A offtakers still price a single month of
            # measured generation → quarterly stays blocked for them.
            if (body.cadence == "quarterly"
                    and getattr(sub, "utility_account_id", None) is not None):
                from ..models import UtilityAccount
                from ..adapters import is_smarthub_provider
                _acct = db.get(UtilityAccount, sub.utility_account_id)
                if _acct is not None and is_smarthub_provider(
                        (_acct.provider or "").lower()):
                    raise HTTPException(
                        400, "Quarterly billing isn't available for VEC/SmartHub "
                             "offtakers yet — their invoices price a single month "
                             "of measured generation, so a quarterly invoice would "
                             "bill only one of the three months. Keep this "
                             "offtaker on monthly cadence.")
            sub.cadence = body.cadence
            sub.next_send_at = next_send_at(body.cadence)
        if body.send_mode is not None:
            if body.send_mode not in VALID_MODES:
                raise HTTPException(400, "invalid send_mode")
            sub.send_mode = body.send_mode
        if body.delivery_mode is not None:
            if body.delivery_mode not in VALID_DELIVERY:
                raise HTTPException(400, "delivery_mode must be approval or auto")
            sub.delivery_mode = body.delivery_mode
        if body.customer_name is not None:
            sub.customer_name = body.customer_name
        if body.client_email is not None:
            sub.client_email = body.client_email
        if body.cc_emails is not None:
            sub.cc_emails = body.cc_emails
        if body.operator_email is not None:
            sub.operator_email = body.operator_email
        if body.formats is not None:
            fmts = [f for f in body.formats if f in VALID_FORMATS]
            sub.formats = fmts or ["pdf"]
        if body.include_summary is not None:
            sub.include_summary = body.include_summary
        if body.annual_trueup is not None:
            sub.annual_trueup = body.annual_trueup
        if body.enabled is not None:
            sub.enabled = body.enabled
        if body.allocation_pct is not None:
            try:
                pct = float(body.allocation_pct)
            except (TypeError, ValueError):
                raise HTTPException(400, "allocation_pct must be a number in (0, 1]")
            if not (0.0 < pct <= 1.0):
                raise HTTPException(400, "allocation_pct must be a fraction in (0, 1] "
                                         "(e.g. 0.95 for 95%)")
            sub.allocation_pct = pct
        if "array_share_pct" in body.model_fields_set:
            # null clears it (check falls back to allocation_pct); a number validates.
            sub.array_share_pct = _validate_array_share(body.array_share_pct)
        if "crosscheck_threshold_pct" in body.model_fields_set:
            # null clears it (check falls back to the fleet default); a number validates.
            sub.crosscheck_threshold_pct = _validate_crosscheck_threshold(body.crosscheck_threshold_pct)
        if body.array_id is not None:
            from ..models import Array
            arr = db.get(Array, body.array_id)
            if arr is None or arr.tenant_id != t.id or arr.deleted_at is not None:
                raise HTTPException(404, "Array not found")
            sub.array_id = body.array_id
        if body.utility_account_id is not None:
            # Re-bind the offtaker's billing source to a different GMP or VEC/
            # SmartHub utility bill. Mirror creation: validate ownership +
            # provider, refresh array_id from it.
            from ..models import UtilityAccount
            from ..adapters import is_smarthub_provider
            acct = db.get(UtilityAccount, body.utility_account_id)
            if acct is None or acct.tenant_id != t.id or acct.deleted_at is not None:
                raise HTTPException(404, f"Utility account {body.utility_account_id} not found")
            _prov = (acct.provider or "").lower()
            if _prov != "gmp" and not is_smarthub_provider(_prov):
                raise HTTPException(
                    400, "Offtaker invoices bind to a GMP or VEC/SmartHub utility "
                         "account.")
            sub.utility_account_id = body.utility_account_id
            sub.array_id = acct.array_id
        if "rate_per_kwh" in body.model_fields_set:
            # null clears the override; a number sets it (validated).
            sub.rate_per_kwh = _validate_rate(body.rate_per_kwh)
        if "discount_pct" in body.model_fields_set:
            sub.discount_pct = _validate_discount(body.discount_pct)
        if "net_rate_per_kwh" in body.model_fields_set:
            sub.net_rate_per_kwh = _validate_rate(body.net_rate_per_kwh)
        if body.auto_attach_gmp is not None:
            sub.auto_attach_gmp = body.auto_attach_gmp
        if "invoice_number_start" in body.model_fields_set:
            v = body.invoice_number_start
            if v is None:
                sub.invoice_number_start = None
                sub.invoice_number_next = None
            else:
                try:
                    n = int(v)
                except (TypeError, ValueError):
                    raise HTTPException(400, "invoice_number_start must be a whole number")
                if not (0 <= n <= 9_999_999):
                    raise HTTPException(400, "invoice_number_start must be between 0 and 9999999")
                # Only (re)seed the running counter when the START actually changes,
                # so saving unrelated edits never resets an in-progress sequence.
                if sub.invoice_number_start != n or sub.invoice_number_next is None:
                    sub.invoice_number_start = n
                    sub.invoice_number_next = n
        if "budget_amount_usd" in body.model_fields_set:
            v = body.budget_amount_usd
            if v is None:
                sub.budget_amount_usd = None
            else:
                try:
                    amt = float(v)
                except (TypeError, ValueError):
                    raise HTTPException(400, "budget_amount_usd must be a number")
                if amt < 0:
                    raise HTTPException(400, "budget_amount_usd can't be negative")
                sub.budget_amount_usd = amt
        # ── Sub-meter invariant (Ford 2026-07-07) ────────────────────────────
        # An offtaker on its OWN sub-account (distinct from the array HOST meter)
        # bills 100% of that meter -> allocation_pct == 1.0; its share of the
        # net-meter GROUP lives in array_share_pct (which real_math bills as
        # share x group excess). Enforce whenever the share/account/array was
        # touched, so the entered share is never ALSO multiplied by allocation_pct
        # (the double-count that made the bill audit wrong). Percent-of-array
        # offtakers (account IS the host) are untouched.
        if (body.allocation_pct is not None
                or body.utility_account_id is not None
                or body.array_id is not None):
            _hid = None
            if sub.utility_account_id is not None and sub.array_id is not None:
                from ..models import UtilityAccount
                _hid = db.execute(
                    select(UtilityAccount.id).where(
                        UtilityAccount.array_id == sub.array_id,
                        UtilityAccount.deleted_at.is_(None))
                    .order_by(UtilityAccount.id)).scalars().first()
            if _hid is not None and _hid != sub.utility_account_id:
                # A sub-metered offtaker's ONE share edit arrives as allocation_pct;
                # it IS the group share, so route it to array_share_pct (real_math)
                # every time -- unless the operator set array_share_pct explicitly in
                # this same patch. allocation_pct is pinned to 1.0 (100%% of the
                # sub-meter) so it can never re-multiply the sub-account excess.
                if (body.allocation_pct is not None
                        and "array_share_pct" not in body.model_fields_set):
                    sub.array_share_pct = _validate_array_share(float(body.allocation_pct))
                sub.allocation_pct = 1.0
        db.commit()
        return {"ok": True, "subscription": _sub_dict(sub)}


# ── BULK OFFTAKER IMPORT v2 (Ford, 2026-07-01 — the "flawless" upload) ────────
# A roster (.csv OR .xlsx) → many offtakers at once. v2 is ARRAY-FIRST: the
# operator names each offtaker's ARRAY (human name) and we FUZZY-MATCH it to the
# tenant's arrays (offtaker_match.match_array), returning a confidence class +
# correctable alternatives so a wrong array→offtaker match can never slip through
# to a wrong invoice. Deterministic header detection (no LLM). "Scrape as much as
# possible": unrecognized columns are kept per-row in `extra` so nothing is lost.
# The dry-run PREVIEW never writes; the separate bulk-commit endpoint writes only
# the reviewed/corrected rows the frontend sends back.
_BULK_HEADER_ALIASES = {
    "name": {"name", "offtakername", "customername", "customer", "offtaker",
              "tenantname", "clientname", "fullname"},
    # ARRAY NAME (bulk-import v2, Ford 2026-07-01) — the array this offtaker draws
    # from. Fuzzy-matched to the tenant's arrays; now a REQUIRED column.
    "array": {"array", "arrayname", "arrayid", "site", "sitename", "arraysite",
              "project", "projectname", "solararray", "generator"},
    "email": {"email", "clientemail", "offtakeremail", "customeremail", "contactemail",
              "emailaddress"},
    "percent": {"percent", "pct", "share", "allocation", "allocationpct",
                "percentage", "offtakerpct", "sharepct"},
    "account_number": {"accountnumber", "accountno", "account", "acctnumber", "acctno",
                        "gmpaccount", "utilityaccount", "meternumber", "meterno",
                        "accountnum", "acct"},
    "discount": {"discount", "discountpct", "discountpercent"},
    # Optional per-offtaker net rate ($/kWh) scraped when present.
    "rate": {"rate", "rateperkwh", "creditrate", "netrate", "netrateperkwh",
             "price", "priceperkwh", "usdperkwh"},
}


def _bulk_norm_header(s: str) -> str:
    # Drop everything but a-z0-9. The '%' sign is treated as noise so headers
    # like "Share %" / "Discount %" normalize to "share" / "discount" and match
    # their aliases (a trailing % glued on would otherwise break exact matching).
    return _re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _bulk_classify_columns(header_row: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, h in enumerate(header_row):
        nh = _bulk_norm_header(h)
        if not nh:
            continue
        for field, aliases in _BULK_HEADER_ALIASES.items():
            if field in mapping:
                continue
            if nh in aliases:
                mapping[field] = idx
                break
    return mapping


async def _read_roster_upload(file: UploadFile) -> bytes:
    """Read a roster upload with the same size/empty guards as the billing-
    workbook path but WITHOUT the xlsx-magic-bytes gate (that gate rejects CSVs).
    Accepts both .csv (plain text) and .xlsx (OpenXML). Returns raw bytes; the
    caller sniffs the type."""
    data = await _read_capped(file, MAX_UPLOAD_BYTES)
    if not data:
        raise HTTPException(400, "The uploaded file is empty")
    _reject_zip_bomb(data)   # #34: refuse a decompression bomb before openpyxl
    return data


def _roster_rows(raw: bytes, filename: str) -> list[list[str]]:
    """Parse a roster's bytes into a list of string rows (header + data), from
    EITHER .xlsx or .csv. xlsx is detected by the ZIP/OpenXML magic (PK\\x03\\x04);
    everything else is treated as delimited text. Every cell is stringified +
    stripped so downstream parsing is uniform regardless of source format."""
    name = (filename or "").lower()
    is_xlsx = raw[:4] == _MAGIC_XLSX or name.endswith(".xlsx")
    if is_xlsx:
        # Reuse openpyxl (already a dependency; the matcher/invoice_writer use it).
        from openpyxl import load_workbook
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception:  # noqa: BLE001
            raise HTTPException(
                422, "Couldn't read that .xlsx — re-save it from Excel/Google "
                     "Sheets, or export as CSV.")
        ws = wb.active
        rows: list[list[str]] = []
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c).strip() for c in r])
        wb.close()
        return rows
    # CSV / delimited text.
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("latin-1")
        except Exception:
            raise HTTPException(
                422, "Couldn't read that file — upload a .csv or .xlsx "
                     "(Excel/Google Sheets: File → Download).")
    return [[(c or "").strip() for c in row] for row in csv.reader(io.StringIO(text))]


def _bulk_pct(raw: str) -> tuple[Optional[float], Optional[str]]:
    """Parse a percent cell to a fraction in (0,1]. Accepts "25", "25%", "0.25".
    Returns (value, None) or (None, error_message)."""
    if not raw:
        return None, "missing percent"
    try:
        v = float(raw.replace("%", "").strip())
    except ValueError:
        return None, f'"{raw}" isn\'t a number'
    # Percent-first (the template says "Accepts 25, 25%, or 0.25"): a value >= 1 is
    # a whole-number percent, only a value < 1 is already a fraction. Crucially "1"
    # is 1%, NOT 100% — the old `> 1.0` test read "1" as the fraction 1.0 = 100%, a
    # 100x over-billing multiplier for an offtaker meant to get 1% (#8). A genuine
    # 100% share is entered as "100".
    if v >= 1.0:
        v = v / 100.0  # "25"/"25%" -> 0.25, "1" -> 0.01, "100" -> 1.0; never "2500%"
    if not (0.0 < v <= 1.0):
        return None, "percent must be between 0 and 100"
    return v, None


def _bulk_discount(raw: str) -> tuple[Optional[float], Optional[str]]:
    """Parse a discount cell to a fraction in [0,1). None when blank."""
    if not raw:
        return None, None
    try:
        v = float(raw.replace("%", "").strip())
    except ValueError:
        return None, f'"{raw}" isn\'t a valid discount'
    # Same percent-first convention as _bulk_pct: "10"/"10%" -> 0.10, "1" -> 1%
    # (not 100%), "0.1" -> 10%. A 100% discount is invalid and still rejected.
    if v >= 1.0:
        v = v / 100.0
    if not (0.0 <= v < 1.0):
        return None, "discount must be between 0 and 100 (and under 100)"
    return v, None


def _bulk_rate(raw: str) -> tuple[Optional[float], Optional[str]]:
    """Parse an optional $/kWh net-rate cell. None when blank."""
    if not raw:
        return None, None
    try:
        v = float(raw.replace("$", "").strip())
    except ValueError:
        return None, f'"{raw}" isn\'t a valid rate ($/kWh)'
    if v < 0 or v > MAX_RATE_PER_KWH:
        return None, f"rate must be between 0 and {MAX_RATE_PER_KWH} $/kWh"
    return v, None


# Maps the format-agnostic detector's field names → this endpoint's internal
# colmap keys, so a detected mapping (or an operator override) plugs straight in.
_DETECTOR_FIELD_TO_BULK = {
    "offtaker_name": "name",
    "array_name": "array",
    "allocation_pct": "percent",
    "email": "email",
    "discount_pct": "discount",
    "net_rate": "rate",
    "account_number": "account_number",
}


@router.post("/subscriptions/bulk-import")
async def bulk_import_offtakers(
    file: UploadFile = File(...),
    dry_run: bool = Form(default=True),
    cadence: str = Form(default="monthly"),
    delivery_mode: str = Form(default="approval"),
    column_map: Optional[str] = Form(default=None),
    authorization: Optional[str] = Header(default=None),
):
    """Bulk-import v2 — parse a roster (.csv/.xlsx) and return a per-row, fuzzy-
    matched, CORRECTABLE preview (dry_run defaults True — a PREVIEW, never writes).

    The frontend renders the preview table with a per-row array dropdown (built
    from the returned `arrays` pick-list + each row's `alternatives`), the operator
    reviews/corrects, then posts the reviewed rows to POST .../bulk-commit. This
    endpoint NEVER auto-writes — even with dry_run=false it returns the same
    preview and points the caller at bulk-commit (the two paths are decoupled so a
    correction round-trip can't force a re-parse). A wrong array→offtaker match
    makes a wrong invoice, so medium/none-confidence matches surface for review and
    are never silently committed.

    Required columns: offtaker `name`, `array` (name), `percent`.
    Optional/scraped: `email`, `discount`, `rate`/`credit_rate`, `account_number`,
    plus any UNRECOGNIZED columns preserved per-row in `extra`.

    Column detection is layered so simple sheets stay fast and messy ones still work:
      1. `column_map` override — a JSON `{field: column_index}` the operator confirmed
         in the review UI. When present we parse by it and SKIP detection entirely.
      2. Clean alias match — the deterministic `_BULK_HEADER_ALIASES` fast path; used
         as-is when the header row exactly names all three required columns.
      3. Format-agnostic detector — `roster_detector.detect_roster_columns`, which
         reads any layout (unknown headers, junk title rows, reordered columns) by
         combining header keywords with CONTENT sniffing (the array column is found by
         fuzzy-matching cell values to the tenant's real arrays). Preferred whenever the
         alias path misses a required field.
    The response carries a `detection` block (sheet, header_row, headers, column_map,
    unmapped_columns, preview, via, warnings) so the frontend can show a column-mapping
    review before the operator confirms.
    """
    t = tenant_from_session(authorization)
    require_not_demo(t)
    if cadence not in VALID_CADENCE:
        raise HTTPException(400, "cadence must be monthly or quarterly")
    if delivery_mode not in VALID_DELIVERY:
        raise HTTPException(400, "delivery_mode must be approval or auto")

    import json as _json

    from .offtaker_match import match_array
    from .roster_detector import detect_roster_columns

    raw = await _read_roster_upload(file)
    rows = _roster_rows(raw, file.filename or "")
    if not rows:
        raise HTTPException(422, "That file is empty.")

    # Pick-list for matching + the frontend correction dropdowns. Needed by the
    # detector too (content array-matching resolves against these real arrays).
    arrays, uaccts = _arrays_and_accounts_for_tenant(t)

    # ── Resolve the column mapping (override → alias fast-path → detector). ───────
    detection: Optional[dict] = None
    header_row_idx = 0

    # (1) Operator-confirmed override: parse by it, skip detection.
    override_map: Optional[dict] = None
    if column_map:
        try:
            parsed = _json.loads(column_map)
        except Exception:  # noqa: BLE001
            raise HTTPException(400, "column_map must be valid JSON")
        if not isinstance(parsed, dict):
            raise HTTPException(400, "column_map must be a JSON object {field: index}")
        override_map = {}
        for field, idx in parsed.items():
            bulk_key = _DETECTOR_FIELD_TO_BULK.get(field, field)
            if idx is None:
                continue
            try:
                override_map[bulk_key] = int(idx)
            except (TypeError, ValueError):
                raise HTTPException(400, f'column_map["{field}"] must be an integer index')

    if override_map is not None:
        colmap = override_map
    else:
        # (2) Clean alias match on the first row.
        header = rows[0]
        colmap = _bulk_classify_columns(header)
        alias_ok = all(f in colmap for f in ("name", "array", "percent"))
        if not alias_ok:
            # (3) Format-agnostic detector — reads junk rows / weird headers / any order.
            detection = detect_roster_columns(raw, file.filename or "", arrays, uaccts)
            header_row_idx = detection.get("header_row") or 0
            det_map = detection.get("column_map") or {}
            colmap = {}
            for field, info in det_map.items():
                if info and isinstance(info.get("index"), int):
                    colmap[_DETECTOR_FIELD_TO_BULK.get(field, field)] = info["index"]

    missing_cols = [f for f in ("name", "array", "percent") if f not in colmap]
    if missing_cols:
        raise HTTPException(
            422, "Couldn't find a column for: " + ", ".join(missing_cols) + ". "
                 "Include a header row with Offtaker (name), Array, and Share % "
                 "(download the template for the exact layout), or map the columns "
                 "manually and re-upload.")

    # The header is at row header_row_idx (0 for the alias/override fast paths); the
    # data rows are everything below it.
    header = rows[header_row_idx] if header_row_idx < len(rows) else rows[0]
    data_rows = rows[header_row_idx + 1:]
    ua_by_number = {(u.get("account_number") or "").strip().lower(): u
                    for u in uaccts if u.get("account_number")}
    ua_by_id = {u["utility_account_id"]: u for u in uaccts}

    # Which header indices are "recognized" — everything else feeds `extra`.
    recognized_idx = set(colmap.values())

    def _cell(row: list[str], field: str) -> str:
        idx = colmap.get(field)
        if idx is None or idx >= len(row):
            return ""
        return (row[idx] or "").strip()

    results: list[dict] = []
    # 1-based sheet row of the first data row (header is at header_row_idx, 0-based).
    for i, row in enumerate(data_rows, start=header_row_idx + 2):
        if not any((c or "").strip() for c in row):
            continue  # silently skip a wholly-blank trailing line
        errors: list[str] = []
        missing: list[str] = []

        name = _cell(row, "name")
        if not name:
            missing.append("name")
        array_raw = _cell(row, "array")
        if not array_raw:
            missing.append("array")

        pct, pct_err = _bulk_pct(_cell(row, "percent"))
        if pct_err == "missing percent":
            missing.append("percent")
        elif pct_err:
            errors.append(pct_err)

        email = _cell(row, "email") or None
        discount, disc_err = _bulk_discount(_cell(row, "discount"))
        if disc_err:
            errors.append(disc_err)
        rate, rate_err = _bulk_rate(_cell(row, "rate"))
        if rate_err:
            errors.append(rate_err)

        # Preserve unrecognized columns verbatim ("scrape as much as possible").
        extra: dict[str, str] = {}
        for idx, cell in enumerate(row):
            if idx in recognized_idx:
                continue
            key = header[idx].strip() if idx < len(header) else f"col{idx}"
            val = (cell or "").strip()
            if key and val:
                extra[key] = val

        # ── Fuzzy array match (the heart of v2). ─────────────────────────────
        m = match_array(array_raw, arrays, uaccts) if array_raw else {
            "array_id": None, "array_name": None, "utility_account_id": None,
            "utility_label": None, "provider": None, "confidence": "none",
            "alternatives": [], "flags": ["no_match"],
        }

        # An explicit account_number column, when present, OVERRIDES the fuzzy
        # utility-account choice (the operator was precise; honor it) — but only
        # to a utility account belonging to this tenant.
        acct_num = _cell(row, "account_number")
        matched_ua_id = m.get("utility_account_id")
        matched_ua_label = m.get("utility_label")
        provider = m.get("provider")
        if acct_num:
            ua = ua_by_number.get(acct_num.strip().lower())
            if ua is None:
                errors.append(f'no connected utility account matches "{acct_num}"')
            else:
                matched_ua_id = ua["utility_account_id"]
                matched_ua_label = ua.get("utility_label")
                provider = ua.get("provider")

        # Bill availability of the finally-chosen account decides ready-ness.
        chosen_ua = ua_by_id.get(matched_ua_id) if matched_ua_id else None
        has_bill = bool(chosen_ua and chosen_ua.get("has_bill"))
        confidence = m.get("confidence", "none")

        results.append({
            "row": i,
            "offtaker_name": name or None,
            "array_name_raw": array_raw or None,
            "matched_array_id": m.get("array_id"),
            "matched_array_name": m.get("array_name"),
            "matched_utility_account_id": matched_ua_id,
            "matched_utility_label": matched_ua_label,
            "provider": provider,
            "confidence": confidence,
            "alternatives": m.get("alternatives", []),
            "allocation_pct": pct,
            "email": email,
            "discount_pct": discount,
            "extra": extra,
            "missing": missing,
            "errors": errors,
        })

    if not results:
        raise HTTPException(422, "No offtaker rows found below the header.")

    def _status(r: dict) -> str:
        if r["missing"]:
            return "blocked"
        if r["errors"]:
            return "blocked"
        # ready needs: no errors, high/exact confidence, AND a settled bill to
        # price from. Everything matched-but-uncertain (or no bill) is review.
        chosen = ua_by_id.get(r["matched_utility_account_id"]) if r["matched_utility_account_id"] else None
        has_bill = bool(chosen and chosen.get("has_bill"))
        if r["confidence"] in ("exact", "high") and has_bill:
            return "ready"
        return "needs_review"

    summary = {"total": len(results), "ready": 0, "needs_review": 0, "blocked": 0}
    for r in results:
        summary[_status(r)] += 1

    resp = {
        "ok": True,
        "dry_run": True,  # this endpoint is always a preview; commit is separate
        "rows": results,
        "summary": summary,
        # The FULL pick-list so the frontend can build correction dropdowns.
        "arrays": [
            {
                "array_id": u["array_id"],
                "array_name": u.get("array_name"),
                "utility_account_id": u["utility_account_id"],
                "utility_label": u.get("utility_label"),
                "provider": u.get("provider"),
                "has_bill": u.get("has_bill"),
            }
            for u in uaccts
        ],
    }
    # Surface the column-mapping detection so the frontend can show a review UI. On the
    # clean-alias / override fast paths there's no detector run, so synthesize a minimal
    # block from the resolved colmap (mapping the internal keys back to detector fields).
    if detection is not None:
        resp["detection"] = {
            "sheet": detection.get("sheet"),
            "header_row": detection.get("header_row"),
            "headers": detection.get("headers", []),
            "column_map": detection.get("column_map", {}),
            "unmapped_columns": detection.get("unmapped_columns", []),
            "preview": detection.get("preview", []),
            "data_rows": detection.get("data_rows"),
            "via": detection.get("via"),
            "warnings": detection.get("warnings", []),
        }
    else:
        _bulk_to_field = {v: k for k, v in _DETECTOR_FIELD_TO_BULK.items()}
        det_cm: dict = {f: None for f in _DETECTOR_FIELD_TO_BULK}
        for bulk_key, idx in colmap.items():
            field = _bulk_to_field.get(bulk_key, bulk_key)
            det_cm[field] = {
                "index": idx,
                "header": header[idx].strip() if idx < len(header) else "",
                "confidence": "high",  # exact alias match or operator-confirmed
            }
        resp["detection"] = {
            "sheet": None,
            "header_row": header_row_idx,
            "headers": [h.strip() for h in header],
            "column_map": det_cm,
            "unmapped_columns": [
                {"index": idx, "header": header[idx].strip() if header[idx] else "",
                 "sample": []}
                for idx in range(len(header)) if idx not in set(colmap.values())
            ],
            "preview": [row for row in data_rows[:5] if any((c or "").strip() for c in row)],
            "data_rows": len(results),
            "via": "override" if override_map is not None else "alias",
            "warnings": [],
        }
    return resp


class _BulkCommitRow(BaseModel):
    offtaker_name: str
    array_id: Optional[int] = None
    utility_account_id: int
    allocation_pct: float
    email: Optional[str] = None
    discount_pct: Optional[float] = None


class BulkCommitBody(BaseModel):
    rows: list[_BulkCommitRow]
    cadence: str = "monthly"
    delivery_mode: str = "approval"


@router.post("/subscriptions/bulk-commit")
async def bulk_commit_offtakers(body: BulkCommitBody,
                                authorization: Optional[str] = Header(default=None)):
    """Commit reviewed/corrected offtaker rows (decoupled from re-parsing).

    The frontend sends this AFTER the operator reviews the bulk-import preview and
    fixes any matches. Each row is created via _create_manual_subscription bound to
    the row's array_id + utility_account_id (both validated as belonging to the
    tenant). Idempotent: a row whose (tenant, customer_name, utility_account_id)
    already matches a LIVE subscription is SKIPPED (never duplicated), so re-running
    a partially-committed batch is safe.

    Returns {ok, created, skipped, failed:[{offtaker_name, error}]}.
    """
    t = tenant_from_session(authorization)
    require_not_demo(t)
    if body.cadence not in VALID_CADENCE:
        raise HTTPException(400, "cadence must be monthly or quarterly")
    if body.delivery_mode not in VALID_DELIVERY:
        raise HTTPException(400, "delivery_mode must be approval or auto")
    if not body.rows:
        raise HTTPException(400, "No rows to commit.")

    from ..models import UtilityAccount

    # Pre-validate every utility account belongs to the tenant + is a billing
    # provider, and snapshot existing (name, ua) subs for idempotency — one query.
    with SessionLocal() as db:
        owned_ua = {
            a.id: a for a in db.execute(
                select(UtilityAccount).where(
                    UtilityAccount.tenant_id == t.id,
                    UtilityAccount.deleted_at.is_(None))
            ).scalars().all()
        }
        existing = db.execute(
            select(BillingReportSubscription.customer_name,
                   BillingReportSubscription.utility_account_id,
                   BillingReportSubscription.allocation_pct,
                   BillingReportSubscription.client_email,
                   BillingReportSubscription.discount_pct).where(
                BillingReportSubscription.tenant_id == t.id,
                BillingReportSubscription.deleted_at.is_(None))
        ).all()
    # Key -> the live sub's money-driving values, so we can tell a true no-op
    # (same values → safe skip) from a real conflict (already live with DIFFERENT
    # allocation/email/discount → must NOT silently skip and leave the stale value,
    # nor silently overwrite live billing; surface it for the operator to resolve).
    existing_vals = {((n or "").strip().lower(), ua): (al, em, di)
                     for (n, ua, al, em, di) in existing}
    existing_keys = set(existing_vals.keys())

    created: list[dict] = []
    skipped: list[dict] = []
    failed: list[dict] = []
    for r in body.rows:
        name = (r.offtaker_name or "").strip()
        if not name:
            failed.append({"offtaker_name": r.offtaker_name, "error": "missing offtaker_name"})
            continue
        ua = owned_ua.get(r.utility_account_id)
        if ua is None:
            failed.append({"offtaker_name": name,
                           "error": f"utility account {r.utility_account_id} not found"})
            continue
        # Idempotency vs conflict (#16). A row that already exists live is a safe
        # no-op ONLY if its money-driving values match; if allocation/email/discount
        # DIFFER, skipping would silently keep the stale value — surface it as a
        # conflict for the operator to resolve (edit or remove the live offtaker)
        # rather than framing it as "already exists".
        _key = (name.lower(), r.utility_account_id)
        if _key in existing_vals:
            _al, _em, _di = existing_vals[_key]
            def _num_eq(a, b):
                if a is None and b is None:
                    return True
                if a is None or b is None:
                    return False
                return abs(float(a) - float(b)) < 1e-9
            _same = (_num_eq(_al, r.allocation_pct)
                     and (_em or None) == (r.email or None)
                     and _num_eq(_di, r.discount_pct))
            if _same:
                skipped.append({"offtaker_name": name,
                                "utility_account_id": r.utility_account_id,
                                "reason": "already exists (identical) — no change"})
            else:
                failed.append({"offtaker_name": name,
                               "error": (f"'{name}' already exists on this utility account "
                                         f"with different values (live allocation "
                                         f"{(_al or 0) * 100:.0f}%). Import won't overwrite live "
                                         f"billing — edit or remove the existing offtaker first.")})
            continue
        # _create_manual_subscription is async but performs only sync DB work; the
        # OFFTAKER↔UTILITY-BILL path validates array_id + utility_account_id
        # ownership + provider itself and persists the binding.
        try:
            out = await _create_manual_subscription(
                t, customer_name=name, array_id=r.array_id,
                allocation_pct=r.allocation_pct,
                utility_account_id=r.utility_account_id, rate_per_kwh=None,
                discount_pct=r.discount_pct, net_rate_per_kwh=None,
                cadence=body.cadence,
                send_mode=("to_client" if r.email else "to_me"),
                delivery_mode=body.delivery_mode, client_email=r.email,
                cc_emails=None, operator_email=None, formats=None,
                include_summary=False, annual_trueup=False, enabled=True,
            )
        except HTTPException as e:
            failed.append({"offtaker_name": name, "error": str(e.detail)})
            continue
        created.append({"offtaker_name": name,
                        "subscription_id": out["subscription"]["id"]})
        existing_keys.add((name.lower(), r.utility_account_id))

    return {"ok": True, "created": len(created), "created_rows": created,
            "skipped": skipped, "failed": failed}


@router.get("/offtaker-template.xlsx")
def offtaker_template_xlsx():
    """A blank, ready-to-fill offtaker-roster .xlsx for the bulk import.

    Header row: Array | Offtaker | Share % | Email | Discount % — matching the
    bulk-import v2 column aliases — plus a few realistic example rows and a second
    "Instructions" sheet. No auth: it's a static blank template (no tenant data).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Offtakers"
    headers = ["Array", "Offtaker", "Share %", "Email", "Discount %"]
    ws.append(headers)
    # Realistic examples (share as whole-number percents — the importer accepts
    # "25", "25%" and "0.25" alike).
    examples = [
        ["Maple Street Solar", "Jane Offtaker", 25, "jane@example.com", 10],
        ["Maple Street Solar", "Green Grocer LLC", 40, "ap@greengrocer.com", 0],
        ["Route 7 Community Array", "Town of Elsewhere", 15, "clerk@elsewhere.gov", 5],
    ]
    for row in examples:
        ws.append(row)
    # Widen the columns so the template is legible on open.
    for col, width in zip("ABCDE", (26, 24, 10, 28, 12)):
        ws.column_dimensions[col].width = width

    info = wb.create_sheet("Instructions")
    notes = [
        ["How to fill this in"],
        [""],
        ["Array", "The name of the solar array. We'll fuzzy-match it to your "
                  "arrays, so it doesn't have to be exact — you can correct any "
                  "match before importing."],
        ["Offtaker", "The customer's name (who receives the invoice)."],
        ["Share %", "This offtaker's share of that array (e.g. 25 for 25%). "
                    "Accepts 25, 25%, or 0.25."],
        ["Email", "Optional. The customer's email (leave blank to keep sends to "
                  "yourself for now)."],
        ["Discount %", "Optional. A discount off the credit rate (e.g. 10 for 10%)."],
        [""],
        ["Tip", "You can add extra columns (account number, notes, etc.) — we "
                "keep them with each row and show them during review."],
    ]
    for row in notes:
        info.append(row)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    data = buf.getvalue()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="offtaker-template.xlsx"'})


class GlobalRatePatch(BaseModel):
    default_billing_rate_per_kwh: Optional[float] = None
    # Discount model: the operator's global default net rate + discount.
    default_net_rate_per_kwh: Optional[float] = None
    default_discount_pct: Optional[float] = None


# ── Long-sweep executor ──────────────────────────────────────────────────────
# reconcile/audit rebuild every offtaker's invoice figures — ~63s at 800
# offtakers, which CROSSES the Railway edge gateway timeout (504 in the
# browser; caught live on ten_anna_800, 2026-07-03). Each sweep now computes
# ONCE per tenant in a background thread; the route answers instantly from
# the cache, or {pending:true} while computing (the frontend polls). Also
# collapses concurrent identical sweeps — which were piling up on the DB —
# into a single run. Math untouched: the thread calls the same functions.
import threading as _threading
import time as _time

_SWEEP_TTL_S = 600            # a computed sweep stays fresh for 10 minutes
_sweeps: dict = {}            # (tenant_id, kind) -> {status, result, at}
_sweeps_lock = _threading.Lock()


def _sweep_result(tenant_id: str, kind: str, compute) -> dict:
    key = (tenant_id, kind)
    now = _time.time()
    with _sweeps_lock:
        ent = _sweeps.get(key)
        if ent:
            if ent["status"] == "done" and now - ent["at"] <= _SWEEP_TTL_S:
                return {"ready": True, "result": ent["result"]}
            if ent["status"] == "running":
                return {"ready": False}
        _sweeps[key] = {"status": "running", "result": None, "at": now}

    def _run():
        try:
            with SessionLocal() as db:   # the thread OWNS its session (pool-leak rule)
                res = compute(db)
            with _sweeps_lock:
                _sweeps[key] = {"status": "done", "result": res, "at": _time.time()}
        except Exception:  # noqa: BLE001
            logger.exception("background %s sweep failed for %s", kind, tenant_id)
            with _sweeps_lock:
                _sweeps.pop(key, None)   # next request retries fresh

    _threading.Thread(target=_run, daemon=True, name=f"sweep-{kind}").start()
    return {"ready": False}


@router.get("/reconcile-bills")
def reconcile_bills_route(authorization: Optional[str] = Header(default=None)):
    """Compare each offtaker invoice's produced-kWh against the captured GMP bill
    for the same array + period — a READ-ONLY trust check before sending.

    Per array: our_kwh (what the invoice uses) vs gmp_kwh (the utility's metered
    generation), with a match|mismatch|no_bill verdict. 'no_bill' = no GMP bill
    is linked to that array yet (awaiting capture) — reported honestly, never
    fabricated.

    Answers from the background sweep cache; {ok:false, pending:true} while the
    sweep is computing — poll again in a few seconds."""
    from .reconcile_bills import reconcile_tenant
    t = tenant_from_session(authorization)
    tid = t.id
    s = _sweep_result(tid, "reconcile", lambda db: reconcile_tenant(db, tid))
    if not s["ready"]:
        return {"ok": False, "pending": True}
    return s["result"]


@router.get("/audit-by-array")
def audit_by_array_route(authorization: Optional[str] = Header(default=None)):
    """The bill-audit sandbox (Anna/Bruce): the fleet organized as GMP allocates
    it — per utility, each array's master bill on top with its offtakers'
    should-be vs GMP-credited math underneath, flagged where GMP got it wrong.
    Read-only; reuses the same allocation cross-check as /reconcile-bills.

    Served DIRECTLY (no background sweep): audit_by_array now computes the
    allocation-only cross-check (no per-offtaker invoice rebuild), so it returns
    in a few seconds even at 800 offtakers — fast enough to answer inline
    without the poll gap the sweep imposes (Ford 2026-07-04: "make it load
    faster")."""
    from .reconcile_bills import audit_by_array
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        return audit_by_array(db, t.id)


@router.get("/export-periods")
def invoice_export_periods(authorization: Optional[str] = Header(default=None)):
    """The settled billing periods the operator can target for a FLEET-WIDE
    invoice export (Bruce 2026-07-07) — the union of every enabled offtaker's
    billable periods, newest first. Feeds the billing-cycle picker on the export
    popover. Read-only; empty `periods` = keep the implicit 'latest bill per
    offtaker' default (nothing to choose between). Each entry
    {label, pretty, count} where `count` is how many offtakers have a settled
    bill for that period, so the operator sees how much a chosen cycle covers."""
    from .delivery import settled_periods_for_sub
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        subs = db.execute(
            select(BillingReportSubscription).where(
                BillingReportSubscription.tenant_id == t.id,
                BillingReportSubscription.deleted_at.is_(None),
                BillingReportSubscription.enabled == True)  # noqa: E712
        ).scalars().all()
        agg: dict[str, dict] = {}
        for sub in subs:
            try:
                for p in settled_periods_for_sub(sub):
                    lbl = p.get("label")
                    if not lbl:
                        continue
                    row = agg.setdefault(lbl, {"label": lbl, "pretty": p.get("pretty") or lbl, "count": 0})
                    row["count"] += 1
            except Exception:  # noqa: BLE001
                continue        # one unreadable offtaker never sinks the list
    # Newest-first by label (YYYY-MM / YYYY-Qn both sort correctly as strings).
    periods = sorted(agg.values(), key=lambda r: r["label"], reverse=True)
    return {"ok": True, "periods": periods}


@router.get("/invoice-export.csv")
def invoice_export_csv(authorization: Optional[str] = Header(default=None),
                       account_code: str = Query(default=""),
                       format: str = Query(default="xero"),
                       tax_type: str = Query(default=""),
                       item_name: str = Query(default="Solar Credit"),
                       invoice_date: Optional[str] = Query(default=None),
                       period: Optional[str] = Query(default=None),
                       memo: str = Query(default="")):
    """Batch invoice-export for QuickBooks Online (CSV), QuickBooks Desktop (IIF),
    or Xero (CSV) — Anna/Bruce's ask #3 (IIF + period + memo added 2026-07-07).

    `format=xero` emits Xero's Sales-Invoice import columns; `format=quickbooks`
    emits QuickBooks Online's invoice-import columns; `format=iif` (aka
    quickbooks-desktop/qbd) emits QuickBooks Desktop's native .IIF transaction
    file. Only offtakers with a real billable invoice are included — never a
    fabricated $0 row. `account_code` feeds Xero's AccountCode AND the IIF income
    account; `tax_type` (Xero TaxType) and `item_name` (QuickBooks Online
    Product/Service) are operator-set per their chart of accounts.

    `invoice_date` (YYYY-MM-DD) sets every invoice's date (default today);
    `period` (YYYY-MM or YYYY-Qn) targets that settled bill period per offtaker
    instead of each one's latest (offtakers without a settled bill for it are
    skipped, never fabricated); `memo` overrides the per-line description.
    """
    from .qb_export import build_invoice_register, normalize_format
    t = tenant_from_session(authorization)
    fmt = normalize_format(format)
    # Parse the operator-entered invoice date (fall back to today on anything
    # unparseable so a stray value never 500s the export).
    inv_dt: Optional[date] = None
    if invoice_date:
        try:
            inv_dt = date.fromisoformat(invoice_date[:10])
        except ValueError:
            inv_dt = None
    target_period = (period or "").strip() or None
    # build_invoice_register rebuilds every offtaker's invoice (~52s at 800),
    # which 504'd at the Railway edge during the CSV download (caught live on
    # ten_anna_800). Compute it in the background sweep + cache; while it runs,
    # answer 202 {pending:true} — the frontend polls, then downloads the file.
    # The cache key includes every input so a different date/period/memo/format
    # never serves a stale file.
    kind = (f"register:{fmt}:{account_code}:{tax_type}:{item_name}:"
            f"{inv_dt.isoformat() if inv_dt else ''}:{target_period or ''}:{memo}")
    s = _sweep_result(t.id, kind, lambda db: build_invoice_register(
        db, t.id, account_code=account_code, fmt=fmt,
        tax_type=tax_type, item_name=item_name, invoice_date=inv_dt,
        period=target_period, memo=memo))
    if not s["ready"]:
        return Response(content='{"ok":false,"pending":true}',
                        media_type="application/json", status_code=202)
    text, count = s["result"]
    label = {"quickbooks": "quickbooks", "iif": "quickbooks-desktop"}.get(fmt, "xero")
    # IIF is a tab-delimited plain-text file (not CSV) with a .iif extension so
    # QuickBooks Desktop's File → Utilities → Import recognizes it.
    if fmt == "iif":
        ext, media = "iif", "text/plain"
    else:
        ext, media = "csv", "text/csv"
    stamp = target_period or date.today().isoformat()
    fname = f"offtaker-invoices-{label}-{stamp}.{ext}"
    return Response(
        content=text, media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{fname}"',
                 "X-Invoice-Count": str(count), "X-Export-Format": label})


@router.get("/invoice-archive")
def invoice_archive_manifest(authorization: Optional[str] = Header(default=None)):
    """Browsable invoice-archive manifest (Anna/Bruce's ask #2): months → arrays →
    offtakers, with what's available for each (invoice / offtaker bill / array
    bill). Read-only; built on demand from the same source the live invoices use.

    Same background-sweep pattern as reconcile/audit — list_archive rebuilds a
    match per offtaker (~60s at 800), which 504'd at the edge. {ok:false,
    pending:true} while the sweep runs; the frontend polls."""
    from .invoice_archive import list_archive
    t = tenant_from_session(authorization)
    tid = t.id
    s = _sweep_result(tid, "archive", lambda db: list_archive(db, tid))
    if not s["ready"]:
        return {"ok": False, "pending": True}
    return s["result"]


@router.get("/invoice-archive.zip")
def invoice_archive_zip(authorization: Optional[str] = Header(default=None),
                        month: Optional[str] = Query(default=None)):
    """Download a month's invoice archive as a .zip laid out per Bruce:
    <month>/<array>/{invoice, each offtaker bill, the array's own bill}. Defaults
    to the latest month present. Missing files are omitted, never fabricated."""
    from .invoice_archive import build_archive_zip
    t = tenant_from_session(authorization)
    # Validate `month` before it reaches the Content-Disposition filename and the
    # zip arcnames (#37) — the sibling download endpoints sanitize, this one didn't.
    if month is not None and not _re.match(r"^\d{4}-\d{2}$", month):
        raise HTTPException(400, "month must be in YYYY-MM format")
    # Rendering a whole month's invoices + bills is the heaviest download here
    # (hundreds of PDFs at 800 offtakers) — background-compute it via the sweep
    # so the request returns instantly (202 {pending} to poll) instead of
    # holding the connection past the edge timeout.
    s = _sweep_result(t.id, f"zip:{month or 'latest'}",
                      lambda db: build_archive_zip(db, t.id, month=month))
    if not s["ready"]:
        return Response(content='{"ok":false,"pending":true}',
                        media_type="application/json", status_code=202)
    data, fname, count = s["result"]
    return Response(
        content=data, media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"',
                 "X-File-Count": str(count)})


@router.get("/gmp-expected-rate")
def gmp_expected_rate(authorization: Optional[str] = Header(default=None),
                      year: int = Query(...), month: int = Query(...),
                      commission_date: Optional[str] = Query(default=None),
                      commission_year: Optional[int] = Query(default=None),
                      age_years: Optional[int] = Query(default=None),
                      regime: Optional[str] = Query(default=None)):
    """Expected GMP $/kWh from the published rate schedule (Anna/Bruce's ask #4).

    An array uses GMP Rate #1 for its first 11 years from its commissioning
    DATE and the Blended Statewide Rate from the 11-year anniversary on — the
    boundary is day-accurate (Bruce's C4 ask: a bare year can misclassify an
    array near the mark; GMP itself has called an array 11 two years early).
    `commission_date` (YYYY-MM-DD) is preferred; `commission_year` is the
    legacy year-only path (read as Jan 1 of that year). Age picks the regime,
    year+month picks the cell. Feeds the setup page's passive 'expected
    billing rate' hint ONLY. Per Bruce (2026-07) this is deliberately NOT an
    active rate-vs-bill cross-check: the schedule needs hand-maintenance and
    can slip, while the bill's own scraped credit rate is the billing truth —
    so no flag ever raises from this, and it never overrides a bill's own
    billed rate.
    """
    from ..rate_schedule_gmp import expected_gmp_rate as _er
    tenant_from_session(authorization)  # gate to signed-in operators
    cd = None
    if commission_date:
        try:
            cd = date.fromisoformat(commission_date[:10])
        except ValueError:
            raise HTTPException(400, "commission_date must be YYYY-MM-DD")
    out = _er(year, month, commission_date=cd, commission_year=commission_year,
              age_years=age_years, regime=regime)
    if out is None:
        raise HTTPException(503, "GMP rate schedule unavailable")
    return out


@router.get("/global-rate")
def get_global_rate(authorization: Optional[str] = Header(default=None)):
    """The operator's global billing defaults. The discount model:
    invoice = kWh × net_rate × (1 − discount). When unset, net rate falls back
    to the VT default and discount to the built-in 10%."""
    from .delivery import MANUAL_TARIFF, DEFAULT_DISCOUNT
    t = tenant_from_session(authorization)
    net = getattr(t, "default_net_rate_per_kwh", None)
    disc = getattr(t, "default_discount_pct", None)
    return {
        "ok": True,
        # legacy flat rate (kept for back-compat)
        "default_billing_rate_per_kwh": getattr(t, "default_billing_rate_per_kwh", None),
        # discount model + the effective defaults actually applied
        "default_net_rate_per_kwh": net,
        "default_discount_pct": disc,
        "effective_net_rate_per_kwh": net if net is not None else MANUAL_TARIFF,
        "effective_discount_pct": disc if disc is not None else DEFAULT_DISCOUNT,
    }


@router.put("/global-rate")
def set_global_rate(body: GlobalRatePatch,
                    authorization: Optional[str] = Header(default=None)):
    """Set (or clear, via null) the operator's global billing defaults — net
    rate and/or discount %. Every customer without a per-customer override uses
    these. Only the fields present in the request body are changed."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        from ..models import Tenant
        tt = db.get(Tenant, t.id)
        if tt is None:
            raise HTTPException(404, "Account not found")
        if "default_billing_rate_per_kwh" in body.model_fields_set:
            tt.default_billing_rate_per_kwh = _validate_rate(body.default_billing_rate_per_kwh)
        if "default_net_rate_per_kwh" in body.model_fields_set:
            tt.default_net_rate_per_kwh = _validate_rate(body.default_net_rate_per_kwh)
        if "default_discount_pct" in body.model_fields_set:
            tt.default_discount_pct = _validate_discount(body.default_discount_pct)
        db.commit()
        return {"ok": True,
                "default_net_rate_per_kwh": tt.default_net_rate_per_kwh,
                "default_discount_pct": tt.default_discount_pct,
                "default_billing_rate_per_kwh": tt.default_billing_rate_per_kwh}


# ─── First-run setup wizard ──────────────────────────────────────────────────

@router.get("/setup-state")
def setup_state(authorization: Optional[str] = Header(default=None)):
    """One call that powers the first-run Reports setup wizard. Returns the
    owner's arrays (with the age/utility/location we need for auto-rates and
    what's still MISSING), whether any customers exist yet, and the global
    rate/discount defaults. The UI shows the wizard when has_customers is False.
    """
    from ..models import Array, UtilityAccount
    from .delivery import DEFAULT_DISCOUNT
    from ..rate_schedule import resolve_net_rate, array_age_bucket, AGE_THRESHOLD_YEARS
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        arrays = db.execute(
            select(Array).where(Array.tenant_id == t.id, Array.deleted_at.is_(None))
            .order_by(Array.name)
        ).scalars().all()
        out_arrays = []
        for a in arrays:
            acct = db.execute(
                select(UtilityAccount).where(UtilityAccount.array_id == a.id)
            ).scalars().first()
            provider = acct.provider if acct else None
            fc = a.first_connect_date
            # The auto rate this array's customers would get today.
            rr = resolve_net_rate(db, provider=provider, region=a.region,
                                  first_connect_date=fc, period_end=None)
            out_arrays.append({
                "array_id": a.id,
                "name": a.name,
                "region": a.region,
                "provider": provider,
                "first_connect_date": fc.date().isoformat() if fc else None,
                "install_year": fc.year if fc else None,
                "age_known": fc is not None,
                "age_bucket": array_age_bucket(fc) if fc else None,
                "auto_net_rate": round(rr.rate, 5),
                "auto_net_source": rr.source,
                "auto_net_note": rr.note,
            })
        n_customers = db.execute(
            select(func.count(BillingReportSubscription.id)).where(
                BillingReportSubscription.tenant_id == t.id,
                BillingReportSubscription.deleted_at.is_(None))
        ).scalar() or 0
        return {
            "ok": True,
            "has_customers": n_customers > 0,
            "customer_count": int(n_customers),
            "age_threshold_years": AGE_THRESHOLD_YEARS,
            "arrays": out_arrays,
            "global": {
                "default_net_rate_per_kwh": getattr(t, "default_net_rate_per_kwh", None),
                "default_discount_pct": getattr(t, "default_discount_pct", None),
                "effective_discount_pct": (getattr(t, "default_discount_pct", None)
                                           if getattr(t, "default_discount_pct", None) is not None
                                           else DEFAULT_DISCOUNT),
            },
        }


class ArrayAgeBody(BaseModel):
    # Either a full ISO commissioning date (the day-accurate path the UI now
    # sends — Bruce's C4 ask) or a bare install year (legacy; read as Jan 1).
    install_year: Optional[int] = None
    first_connect_date: Optional[str] = None
    region: Optional[str] = None   # north | central | south (optional location)


@router.patch("/arrays/{array_id}")
def set_array_setup(array_id: int, body: ArrayAgeBody,
                    authorization: Optional[str] = Header(default=None)):
    """Set an array's commissioning (in-service) date — feeds the GMP rate
    regime, day-accurate at the 11-year boundary — and optional region.
    Tenant-scoped. Both paths are validated to 1990-01-01..today."""
    from datetime import date as _date
    from ..models import Array
    t = tenant_from_session(authorization)
    require_not_demo(t)
    fc = None
    if body.first_connect_date:
        try:
            fc = datetime.fromisoformat(body.first_connect_date[:10])
        except ValueError:
            raise HTTPException(400, "first_connect_date must be YYYY-MM-DD")
        if fc.date() < _date(1990, 1, 1) or fc.date() > _date.today():
            raise HTTPException(400, "first_connect_date must be between 1990-01-01 and today")
    elif body.install_year is not None:
        yr = int(body.install_year)
        if yr < 1990 or yr > _date.today().year:
            raise HTTPException(400, f"install_year must be 1990–{_date.today().year}")
        fc = datetime(yr, 1, 1)
    with SessionLocal() as db:
        arr = db.get(Array, array_id)
        if arr is None or arr.tenant_id != t.id or arr.deleted_at is not None:
            raise HTTPException(404, "Array not found")
        if fc is not None:
            arr.first_connect_date = fc
        if "region" in body.model_fields_set and body.region is not None:
            arr.region = body.region.strip().lower() or None
        db.commit()
        return {"ok": True, "array_id": arr.id,
                "first_connect_date": arr.first_connect_date.date().isoformat()
                if arr.first_connect_date else None,
                "region": arr.region}


@router.delete("/subscriptions/{sub_id}")
def delete_subscription(sub_id: int, authorization: Optional[str] = Header(default=None)):
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        sub.deleted_at = datetime.utcnow()
        sub.enabled = False
        # Dismiss any pending drafts so a deleted offtaker leaves nothing behind in
        # the approval inbox (they used to orphan there — the "(sample)" leftovers).
        for d in db.execute(select(ReportDraft).where(
                ReportDraft.subscription_id == sub.id,
                ReportDraft.status == "pending")).scalars().all():
            d.status = "dismissed"
            d.dismissed_at = datetime.utcnow()
        db.commit()
        _sync_invoicing_quantity(t.id)
        return {"ok": True}


@router.post("/subscriptions/{sub_id}/send-now")
def send_now(sub_id: int, test: bool = Query(default=True),
             authorization: Optional[str] = Header(default=None)):
    """Manually deliver this subscription now. Defaults to a TEST send (to the
    operator) so the slider can be exercised safely before going live."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        result = deliver_subscription(db, sub, t, triggered_by="manual", is_test=test)
        status = 200 if result.get("ok") else 422
        if not result.get("ok"):
            raise HTTPException(status, result.get("error", "send failed"))
        return {"ok": True, "result": result}


def _content_disposition_attachment(filename: str) -> str:
    """Content-Disposition value safe for latin-1 HTTP-header encoding.

    HTTP header values are latin-1; a non-latin-1 char in the filename (an em-dash
    from a customer name, smart quotes, an emoji) raises UnicodeEncodeError in the
    ASGI layer and 500s the whole download (seen live on subscription previews).
    Emit an ASCII-only ``filename=`` (RFC 6266) plus an RFC 5987 ``filename*`` that
    carries the real UTF-8 name for modern clients.
    """
    from urllib.parse import quote
    safe = filename or "download"
    ascii_name = safe.encode("ascii", "replace").decode("ascii")
    for ch in ('"', "\\", "\r", "\n", "?"):
        ascii_name = ascii_name.replace(ch, "_")
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(safe, safe='')}"


@router.get("/subscriptions/{sub_id}/preview")
def preview(sub_id: int, kind: str = Query(default="invoice"),
            fmt: str = Query(default="pdf"),
            variant: Optional[str] = Query(default=None),
            authorization: Optional[str] = Header(default=None)):
    """Render the current invoice or summary on demand for download.

    variant (invoice PDF only): "default" forces OUR standard format; "template"
    forces the operator's template even when it's toggled off (a preview-only
    compare for the approval-inbox buttons); None = exactly what gets sent."""
    t = tenant_from_session(authorization)
    if fmt not in VALID_FORMATS:
        raise HTTPException(400, "fmt must be pdf or xlsx")
    if kind not in ("invoice", "summary"):
        raise HTTPException(400, "kind must be invoice or summary")
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        match = build_match(sub)
        if not match.matched:
            raise HTTPException(422, "stored workbook is no longer recognizable")
        import tempfile, pathlib
        with tempfile.TemporaryDirectory(prefix="ao-prev-") as tmp:
            tmpd = pathlib.Path(tmp)
            if kind == "invoice":
                # Invoice in the customer's OWN uploaded format: for workbook
                # subs we load their stored .xlsx and populate it (preserving
                # all styling/formulas/the Template sheet). Manual customers (no
                # source_workbook) fall back to the standard generated invoice.
                if fmt == "xlsx" and getattr(sub, "source_workbook", None):
                    from .invoice_writer import (
                        populate_invoice_workbook, InvoiceWriterError)
                    try:
                        blob = populate_invoice_workbook(sub)
                    except InvoiceWriterError as e:
                        raise HTTPException(422, str(e))
                    media = ("application/vnd.openxmlformats-officedocument"
                             ".spreadsheetml.sheet")
                    fname = f"{sub.customer_name.replace(' ', '_')}_invoice.xlsx"
                    return StreamingResponse(
                        io.BytesIO(blob), media_type=media,
                        headers={"Content-Disposition": _content_disposition_attachment(fname)})
                from . import invoice as inv
                p = None
                if fmt == "pdf":
                    # Mirror the SEND chain so the preview is exactly what gets
                    # delivered: own-workbook repro → operator-template repro →
                    # operator token-HTML template → standard. `variant` lets the
                    # approval-inbox compare buttons force one side (see docstring).
                    from .delivery import (_render_from_repro,
                                           _render_from_operator_template_repro,
                                           _render_from_operator_template)
                    pp = tmpd / "p.pdf"
                    if variant == "default":
                        pass                          # force OUR standard format
                    elif variant == "template":       # force the template (even if off)
                        if (_render_from_repro(match, sub, pp)
                                or _render_from_operator_template_repro(match, sub, pp, force=True)
                                or _render_from_operator_template(match, sub, pp, force=True)):
                            p = pp
                    elif (_render_from_repro(match, sub, pp)
                            or _render_from_operator_template_repro(match, sub, pp)
                            or _render_from_operator_template(match, sub, pp)):
                        p = pp
                if p is None:
                    p = (inv.render_invoice_pdf(match, tmpd / "p.pdf") if fmt == "pdf"
                         else inv.render_invoice_xlsx(match, tmpd / "p.xlsx"))
            else:
                from . import summary as summ
                p = (summ.render_summary_pdf(match, tmpd / "p.pdf") if fmt == "pdf"
                     else summ.render_summary_xlsx(match, tmpd / "p.xlsx"))
            blob = p.read_bytes()
    media = ("application/pdf" if fmt == "pdf"
             else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fname = f"{sub.customer_name.replace(' ', '_')}_{kind}.{fmt}"
    return StreamingResponse(io.BytesIO(blob), media_type=media,
                             headers={"Content-Disposition": _content_disposition_attachment(fname)})


# ─── Bring-your-own generation spreadsheet tracker ───────────────────────────
# The operator uploads their OWN running generation sheet (whatever columns they
# use); we detect its structure and keep appending a monthly row as fresh GMP
# bills land. A "Download latest spreadsheet" button streams the kept-current
# file. Whole feature gated behind SPREADSHEET_TRACKER_ENABLED.

_XLSX_MEDIA = ("application/vnd.openxmlformats-officedocument"
               ".spreadsheetml.sheet")


def _tracker_status_dict(sub) -> dict:
    """The tracker card state for the offtaker editor. Honest about whether a
    sheet is attached, what we detected, and when we last appended."""
    m = getattr(sub, "tracker_map", None) or {}
    has = bool(getattr(sub, "tracker_workbook", None)) and bool(m.get("ok"))
    up = getattr(sub, "tracker_updated_at", None)
    return {
        "enabled": True,
        "has_sheet": has,
        "filename": getattr(sub, "tracker_filename", None),
        "columns": m.get("columns") if has else None,
        "headers": m.get("headers") if has else None,
        "header_row": m.get("header_row") if has else None,
        "sheet": m.get("sheet") if has else None,
        "data_rows": m.get("data_rows") if has else None,
        "last_period": m.get("last_period") if has else None,
        "updated_at": up.isoformat() + "Z" if up else None,
        "warnings": m.get("warnings") or [],
    }


def _friendly_period(lbl: str) -> str:
    """'2026-05' -> 'May 2026' for human-readable upload feedback; passthrough on anything odd."""
    try:
        import calendar
        y, m = str(lbl).split("-")[:2]
        return f"{calendar.month_name[int(m)]} {y}"
    except Exception:  # noqa: BLE001
        return str(lbl)


@router.get("/subscriptions/{sub_id}/tracker")
def tracker_status(sub_id: int, authorization: Optional[str] = Header(default=None)):
    """Tracker state for one offtaker (drives the card). 404 only on a missing
    sub; returns {enabled:false} when the feature flag is off so the UI hides."""
    from .sheet_tracker import tracker_enabled
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        if not tracker_enabled():
            return {"ok": True, "tracker": {"enabled": False}}
        return {"ok": True, "tracker": _tracker_status_dict(sub)}


@router.post("/subscriptions/{sub_id}/tracker")
async def tracker_upload(sub_id: int,
                         file: UploadFile = File(...),
                         authorization: Optional[str] = Header(default=None)):
    """Upload the offtaker's existing generation spreadsheet (XLSX or CSV). We
    detect its structure ('our magic'), normalize to xlsx, and store it as the
    running ledger we keep current. Returns the detected mapping for review."""
    from .sheet_tracker import tracker_enabled, ingest_upload
    t = tenant_from_session(authorization)
    require_not_demo(t)
    if not tracker_enabled():
        raise HTTPException(404, "Spreadsheet tracker is not enabled.")
    raw = await _read_capped(file, MAX_UPLOAD_BYTES)
    if not raw:
        raise HTTPException(400, "Empty file.")
    name = file.filename or "generation.xlsx"
    is_x = raw[:4] in (_MAGIC_XLSX, _MAGIC_XLS)
    is_csv = name.lower().endswith(".csv") or (not is_x)
    if not is_x and not is_csv:
        raise HTTPException(415, "Upload an .xlsx or .csv generation sheet.")
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        # Do NOT pass the offtaker's name as a column hint here: on sheets that carry BOTH a
        # 'kWh whole array' column (populated) and an empty 'kWh <offtaker>' column, the name
        # boost dragged generation onto the empty named column, so appended rows landed blank.
        # The generation column is detected by keyword + data-presence, which is correct here.
        res = ingest_upload(raw, name, None)
        if not res.get("ok"):
            warn = "; ".join(res.get("warnings") or []) or "Couldn't read that sheet."
            raise HTTPException(422, warn)
        sub.tracker_workbook = res["workbook"]
        sub.tracker_filename = name
        sub.tracker_map = res["mapping"]
        sub.tracker_updated_at = datetime.utcnow()
        db.add(sub)
        # Transparency: immediately reconcile the freshly-uploaded sheet against the offtaker's
        # GMP bills, so the operator SEES that we processed it and produced an updated
        # spreadsheet — not a silent store. Best-effort: a reconcile hiccup never fails upload.
        from .sheet_tracker import update_subscription_sheet
        try:
            recon = update_subscription_sheet(db, sub) or {}
        except Exception:  # noqa: BLE001
            recon = {"status": "error"}
        db.commit()
        db.refresh(sub)
        added = recon.get("periods") if isinstance(recon.get("periods"), list) else (
            [recon["period"]] if recon.get("status") == "appended" and recon.get("period") else [])
        return {"ok": True, "tracker": _tracker_status_dict(sub),
                "processed": {
                    "sheet": (res.get("mapping") or {}).get("sheet"),
                    "status": recon.get("status"),
                    "added": [_friendly_period(p) for p in added],
                    "added_count": len(added),
                    "normalized": (res.get("mapping") or {}).get("normalized") or 0,
                    "ai": recon.get("ai"),   # {sane, explanation, via} when the AI planner ran
                }}


@router.patch("/subscriptions/{sub_id}/tracker")
def tracker_remap(sub_id: int, body: dict,
                  authorization: Optional[str] = Header(default=None)):
    """Operator correction of the detected column mapping (nice-to-have). Body:
    {"columns": {"period": <idx>, "generation": <idx>, ...}}. Validates indices
    against the stored headers and updates the mapping in place."""
    from .sheet_tracker import tracker_enabled
    t = tenant_from_session(authorization)
    require_not_demo(t)   # mutates + commits — never on the shared demo tenant (#36)
    if not tracker_enabled():
        raise HTTPException(404, "Spreadsheet tracker is not enabled.")
    cols = (body or {}).get("columns")
    if not isinstance(cols, dict):
        raise HTTPException(400, "columns object required")
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        m = dict(getattr(sub, "tracker_map", None) or {})
        if not m.get("ok"):
            raise HTTPException(409, "No tracker sheet to remap.")
        ncol = len(m.get("headers") or [])
        clean = {}
        for k in ("period", "generation", "consumption", "rate", "amount"):
            if k in cols and cols[k] is not None:
                ci = int(cols[k])
                if ci < 0 or (ncol and ci >= ncol):
                    raise HTTPException(400, f"{k} column out of range")
                clean[k] = ci
        if "generation" not in clean:
            raise HTTPException(400, "A generation/kWh column is required.")
        m["columns"] = clean
        sub.tracker_map = m
        sub.tracker_updated_at = datetime.utcnow()
        db.add(sub)
        db.commit()
        db.refresh(sub)
        return {"ok": True, "tracker": _tracker_status_dict(sub)}


@router.delete("/subscriptions/{sub_id}/tracker")
def tracker_remove(sub_id: int, authorization: Optional[str] = Header(default=None)):
    """Detach the BYO sheet from this offtaker."""
    from .sheet_tracker import tracker_enabled
    t = tenant_from_session(authorization)
    require_not_demo(t)   # mutates + commits — never on the shared demo tenant (#36)
    if not tracker_enabled():
        raise HTTPException(404, "Spreadsheet tracker is not enabled.")
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        sub.tracker_workbook = None
        sub.tracker_filename = None
        sub.tracker_map = None
        sub.tracker_updated_at = datetime.utcnow()
        db.add(sub)
        db.commit()
        return {"ok": True, "tracker": _tracker_status_dict(sub)}


@router.get("/subscriptions/{sub_id}/tracker/download")
def tracker_download(sub_id: int, authorization: Optional[str] = Header(default=None)):
    """Stream the current, kept-current generation spreadsheet. Before streaming
    we opportunistically append the latest period (idempotent) so 'Download
    latest' always reflects the freshest computed bill even between worker runs."""
    from .sheet_tracker import tracker_enabled, update_subscription_sheet
    t = tenant_from_session(authorization)
    require_not_demo(t)   # opportunistic append below commits — never on demo (#36)
    if not tracker_enabled():
        raise HTTPException(404, "Spreadsheet tracker is not enabled.")
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        if not getattr(sub, "tracker_workbook", None):
            raise HTTPException(404, "No spreadsheet uploaded yet.")
        # Keep-current on demand: append the latest period if it isn't there.
        res = update_subscription_sheet(db, sub)
        if res.get("status") == "appended":
            db.add(sub)
            db.commit()
            db.refresh(sub)
        blob = bytes(sub.tracker_workbook)
        base = (getattr(sub, "tracker_filename", None) or "generation.xlsx")
        if base.lower().endswith(".csv"):
            base = base[:-4] + ".xlsx"
        elif not base.lower().endswith(".xlsx"):
            base = base + ".xlsx"
    return StreamingResponse(
        io.BytesIO(blob), media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{base}"'})


@router.get("/subscriptions/{sub_id}/preview-math")
def preview_math(sub_id: int, period: Optional[str] = Query(default=None),
                 authorization: Optional[str] = Header(default=None)):
    """Compute (without persisting a draft) the auditable billing math for a
    subscription's billing period: the array's period generation, the customer's
    allocation %, the resulting customer-share kWh, the $/kWh rate, and the
    dollar amount. Powers the run-table rows so every customer shows real
    numbers eagerly — no draft required.

    `period` (Bruce 2026-07-07, C4) previews a SPECIFIC settled bill period
    ('YYYY-MM' | 'YYYY-Qn') so the preview matches the period the operator will
    draft; omit it for the latest bill (the default).

    Never fabricates: when the array has no generation for the period yet,
    `has_data` is false and the kWh/amount fields are null so the UI can show a
    muted 'No generation data yet' instead of a bogus number.
    """
    t = tenant_from_session(authorization)
    target_period = (period or "").strip() or None
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
    try:
        match = build_match(sub, period_label=target_period)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(422, f"workbook unreadable: {e}")
    if not match.matched or not match.latest_period:
        return {
            "subscription_id": sub.id,
            "source": match.source,
            "has_data": False,
            "allocation_pct": match.allocation_pct,
            "array_total_kwh": None,
            "customer_kwh": None,
            "amount_usd": None,
            "rate": None,
            "period_start": None,
            "period_end": None,
        }
    ci = match.computed_invoice or {}
    array_total = ci.get("project_total_kwh") or ci.get("array_kwh")
    cust_kwh = ci.get("kwh")
    pct = match.allocation_pct
    if array_total is None and cust_kwh is not None and pct:
        array_total = round(cust_kwh / pct, 1)
    amount = ci.get("amount_owed")
    rate = None
    if cust_kwh and amount is not None and cust_kwh > 0:
        rate = amount / cust_kwh
    # "Has data" means the array actually produced generation this period — a
    # zero array total means we have nothing real to show yet.
    has_data = bool(array_total)
    return {
        "subscription_id": sub.id,
        "source": match.source,
        "has_data": has_data,
        # allocation_pct is the share ACTUALLY multiplied against
        # array_total_kwh: build_match keeps the (base, share, billed-kWh)
        # triple on ONE basis, so allocation_pct × array_total_kwh ==
        # customer_kwh always holds — the operator never sees "100%" beside a
        # kWh figure that is 99.4% of the displayed total (mixed bases).
        "allocation_pct": pct,
        "array_total_kwh": array_total if has_data else None,
        "customer_kwh": cust_kwh if has_data else None,
        "amount_usd": amount if has_data else None,
        "rate": rate if has_data else None,
        "rate_source": ci.get("rate_source"),
        # Which basis the pair above uses ('real_math' = share × the array's
        # group excess; 'gmp_credited' = pct × the offtaker's own bill excess),
        # plus both raw figures for the side-by-side audit.
        "billing_basis": ci.get("billing_basis"),
        "gmp_credited_kwh": (ci.get("gmp_credited_kwh") if has_data else None),
        "own_bill_excess_kwh": (ci.get("own_bill_excess_kwh") if has_data else None),
        "array_group_excess_kwh": (ci.get("array_group_excess_kwh") if has_data else None),
        # Discount model: the savings story the customer sees.
        "net_rate_per_kwh": ci.get("net_rate_per_kwh"),
        "discount_pct": ci.get("discount_pct"),
        "effective_rate_per_kwh": ci.get("effective_rate_per_kwh"),
        "net_rate_source": ci.get("net_rate_source"),
        "net_rate_note": ci.get("net_rate_note"),
        "discount_source": ci.get("discount_source"),
        "solar_savings_usd": (ci.get("solar_savings") if has_data else None),
        "kwh_source": ci.get("kwh_source"),
        # Period identity: month is 'YYYY-MM' (or 'YYYY-Qn' for a quarterly
        # offtaker); period_months lists the months a quarterly invoice sums.
        "month": ci.get("month"),
        "billing_cadence": ci.get("billing_cadence"),
        "period_months": ci.get("period_months"),
        "period_note": ci.get("period_note"),
        "period_start": ci.get("period_start"),
        "period_end": ci.get("period_end"),
    }


@router.get("/subscriptions/{sub_id}/daily-series")
def subscription_daily_series(
    sub_id: int,
    period: Optional[str] = Query(default=None, description="YYYY-MM month, or YYYY-Qn quarter; default = latest month with data"),
    authorization: Optional[str] = Header(default=None),
):
    """Real DAILY generation for an offtaker's array over a billing period —
    powers the daily-generation bar graph in reports. Points are the array's
    measured DailyGeneration rows, scaled by the offtaker's allocation_pct so the
    bars show THAT offtaker's daily share. Never fabricates: when the array has no
    daily rows for the window, returns points:[] + has_data:false so the UI shows
    an honest empty state instead of invented bars.
    """
    from datetime import date as _date
    from ..models import DailyGeneration, Array
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        array_id = sub.array_id
        alloc = sub.allocation_pct if sub.allocation_pct is not None else 1.0
        if array_id is None:
            return {"subscription_id": sub_id, "has_data": False, "points": [],
                    "reason": "no_array", "allocation_pct": alloc}

        # Resolve the window. Explicit period wins; else the latest month that has
        # any daily rows for this array.
        start = end = None
        label = None
        if period:
            p = period.strip().upper()
            try:
                if "Q" in p:
                    yr, q = p.split("-Q") if "-Q" in p else (p[:4], p[-1])
                    yr = int(yr); q = int(q)
                    sm = 3 * (q - 1) + 1
                    start = _date(yr, sm, 1)
                    end = (_date(yr + (sm + 2) // 12, ((sm + 2) % 12) + 1, 1)
                           if sm + 3 > 12 else _date(yr, sm + 3, 1)) - timedelta(days=1)
                    label = f"Q{q} {yr}"
                else:
                    yr, mo = p.split("-"); yr = int(yr); mo = int(mo)
                    start = _date(yr, mo, 1)
                    end = (_date(yr + 1, 1, 1) if mo == 12 else _date(yr, mo + 1, 1)) - timedelta(days=1)
                    label = start.strftime("%B %Y")
            except (ValueError, IndexError):
                raise HTTPException(400, "period must be YYYY-MM or YYYY-Qn")
        if start is None:
            latest = db.execute(
                select(DailyGeneration.day)
                .where(DailyGeneration.array_id == array_id)
                .order_by(DailyGeneration.day.desc()).limit(1)
            ).scalar_one_or_none()
            if latest is None:
                return {"subscription_id": sub_id, "has_data": False, "points": [],
                        "reason": "no_daily_data", "allocation_pct": alloc}
            start = latest.replace(day=1)
            end = (_date(latest.year + 1, 1, 1) if latest.month == 12
                   else _date(latest.year, latest.month + 1, 1)) - timedelta(days=1)
            label = start.strftime("%B %Y")

        rows = db.execute(
            select(DailyGeneration.day, DailyGeneration.kwh, DailyGeneration.source)
            .where(DailyGeneration.array_id == array_id,
                   DailyGeneration.day >= start, DailyGeneration.day <= end)
            .order_by(DailyGeneration.day.asc())
        ).all()
        arr = db.get(Array, array_id)
        # Surface data provenance so the UI can mark estimated days (split out of a
        # bill via bill_prorate) distinctly from measured rows — never render an
        # estimate as if it were measured.
        points = [{"day": d.isoformat(),
                   "array_kwh": round(k or 0.0, 1),
                   "kwh": round((k or 0.0) * alloc, 1),
                   "source": src or "csv",
                   "is_estimated": (src == "bill_prorate")} for (d, k, src) in rows]
        total = round(sum(p["kwh"] for p in points), 1)
        return {
            "subscription_id": sub_id,
            "has_data": bool(points),
            "period_label": label,
            "period_start": start.isoformat(),
            "period_end": end.isoformat(),
            "allocation_pct": alloc,
            "array_name": arr.name if arr else None,
            "total_kwh": total,
            "points": points,
        }


@router.get("/subscriptions/{sub_id}/trends")
def subscription_trends(sub_id: int,
                        authorization: Optional[str] = Header(default=None)):
    """Multi-year billing trends for a subscription (CONTRACT 1). Tenant-scoped:
    404 if the sub isn't owned. Thin/unreadable workbook → 200 with empty
    collections + null scalars; never 500 on real data."""
    t = tenant_from_session(authorization)
    from . import summary as summ
    with SessionLocal() as db:
        sub, client = _resolve_trends_target(db, t.id, sub_id)
        if sub is None:
            # A valid CLIENT exists but has no billing workbook yet → honest
            # empty state, not a 404. (The reports UI is client-centric and
            # links trends by client id; a client without an uploaded workbook
            # simply has no history to chart.)
            return summ._empty_trends(client.name if client else None)
        try:
            match = build_match(sub)
            trends = summ.build_trends(match)
        except Exception:  # noqa: BLE001 — thin/missing/corrupt workbook
            logger.warning("trends build failed for sub %s", sub.id, exc_info=True)
            trends = summ._empty_trends(sub.customer_name)
        if not trends.get("customer_name"):
            trends["customer_name"] = sub.customer_name
        return trends


def _resolve_trends_target(db, tenant_id: str, ident: int):
    """Resolve the trends route's {id} to (subscription, client).

    The reports UI is client-centric and links by client id, but trends are
    derived from a BillingReportSubscription's stored workbook. Accept EITHER:
      1. a BillingReportSubscription.id (direct), or
      2. a Client.id → that client's newest non-deleted subscription.
    Returns (sub, client). Raises 404 only when `ident` matches neither a sub
    nor a client owned by this tenant. (sub may be None when a real client has
    no subscription yet — caller renders the empty trends state.)
    """
    sub = db.execute(
        select(BillingReportSubscription).where(
            BillingReportSubscription.id == ident,
            BillingReportSubscription.tenant_id == tenant_id,
            BillingReportSubscription.deleted_at.is_(None))
    ).scalar_one_or_none()
    if sub is not None:
        return sub, None

    client = db.execute(
        select(Client).where(
            Client.id == ident,
            Client.tenant_id == tenant_id,
            Client.deleted_at.is_(None))
    ).scalar_one_or_none()
    if client is None:
        raise HTTPException(404, "Subscription not found")

    csub = db.execute(
        select(BillingReportSubscription).where(
            BillingReportSubscription.client_id == client.id,
            BillingReportSubscription.tenant_id == tenant_id,
            BillingReportSubscription.deleted_at.is_(None))
        .order_by(BillingReportSubscription.id.desc())
    ).scalars().first()
    return csub, client


def _get_owned(db, tenant_id: str, sub_id: int) -> BillingReportSubscription:
    sub = db.execute(
        select(BillingReportSubscription).where(
            BillingReportSubscription.id == sub_id,
            BillingReportSubscription.tenant_id == tenant_id,
            BillingReportSubscription.deleted_at.is_(None))
    ).scalar_one_or_none()
    if sub is None:
        raise HTTPException(404, "Subscription not found")
    return sub


# ─── DRAFT → APPROVE → SEND inbox (Paul Bozuwa's core workflow) ───────────────
# "I get an email drafted... the customer invoice PDF, the GMP invoice PDF...
#  I go over it and approve it or modify it and then send." Nothing reaches a
#  real customer until the operator clicks Approve & send.

def _calc_credit_for_budget(sub):
    """The CALCULATED solar credit value (pre-budget-override) for a budget-billed
    offtaker, so the email preview can show it alongside the budgeted amount as two
    distinct rows. None when there's no budget. Best-effort — a slow/unreadable
    workbook just yields None (the preview falls back to a single row)."""
    if sub is None or getattr(sub, "budget_amount_usd", None) is None:
        return None
    try:
        ci = build_match(sub).computed_invoice or {}
        return ci.get("solar_credit_value")
    except Exception:  # noqa: BLE001 — never break draft serialization on a parse hiccup
        return None


def _draft_letter_default(d: ReportDraft, sub, email_fields: dict) -> Optional[dict]:
    """{'letter': plain-text mass-template letter, 'subject': rendered subject}
    for THIS draft — what the send uses when the operator hasn't written a
    note, so the card's 'Edit email' box + envelope prefill with exactly what
    would go out. Rendered from the draft's own snapshot figures (no
    build_match). email_fields comes from delivery._offtaker_email_fields —
    fetch it ONCE per request and share it across a draft loop (a per-draft
    fetch would be the next N+1)."""
    if not email_fields:
        return None
    try:
        from ..email_templates import (DEFAULT_OFFTAKER_BODY_TEMPLATE,
                                       DEFAULT_OFFTAKER_SUBJECT_TEMPLATE,
                                       build_offtaker_context, render_merge,
                                       html_to_text)
        from .delivery import _attachments_line
        kwh_s = (f"{d.customer_kwh:,.0f} kWh" if d.customer_kwh is not None
                 else "your production")
        amount_s = (f"${d.amount_usd:,.2f}" if d.amount_usd is not None
                    else "the amount due")
        ctx = build_offtaker_context(
            offtaker_name=((getattr(sub, "customer_name", None) or d.customer_name)
                           if sub else d.customer_name),
            tenant_name=email_fields.get("tenant_name") or "your solar provider",
            tenant_email=email_fields.get("tenant_email", ""),
            period=(d.period_label or "the latest period"),
            kwh=kwh_s, amount=amount_s,
            invoice_number=str(d.invoice_number or ""),
            attachments_line=_attachments_line(
                getattr(sub, "auto_attach_gmp", True) is not False if sub else True,
                getattr(sub, "include_summary", False) is True if sub else False),
            signoff_template=email_fields.get("signoff_t"),
            tenant_signoff_name=email_fields.get("signoff_name"),
        )
        body_t = (email_fields.get("body_t") or "").strip() or DEFAULT_OFFTAKER_BODY_TEMPLATE
        subj_t = (email_fields.get("subject_t") or "").strip() or DEFAULT_OFFTAKER_SUBJECT_TEMPLATE
        subject = render_merge(subj_t, ctx)
        if not ctx["invoice_number"]:
            subject = subject.replace(" ()", "")
        return {"letter": html_to_text(render_merge(body_t, ctx)),
                "subject": subject}
    except Exception:  # noqa: BLE001 — the prefill is a nicety, never break a draft
        logger.exception("draft letter-default render failed")
        return None


def _rate_meta_from_ci(ci: Optional[dict]) -> dict:
    """The Solar-credit-rate provenance the editable rate field needs, pulled
    from a draft's computed invoice: the bill-derived DEFAULT (value + honest
    source/note, banked-reference vs the bill's own rate) and the rate the
    invoice ACTUALLY resolved to (source 'customer' when a per-offtaker override
    won). Empty for non-bill offtakers / no settled bill. Cheap dict reads — the
    caller already computed `ci`, so this adds no DB work."""
    ci = ci or {}
    return {
        "default_net_rate_per_kwh": ci.get("default_net_rate_per_kwh"),
        "default_net_rate_source": ci.get("default_net_rate_source"),
        "default_net_rate_note": ci.get("default_net_rate_note"),
        "resolved_net_rate_per_kwh": ci.get("net_rate_per_kwh"),
        "resolved_net_rate_source": ci.get("net_rate_source"),
    }


def _draft_dict(d: ReportDraft, sub=None, gmp_auto_status=None, operator_name=None,
                email_fields: Optional[dict] = None, attach_provider=None,
                rate_meta: Optional[dict] = None) -> dict:
    _letter = (_draft_letter_default(d, sub, email_fields)
               if email_fields else None) or {}
    rate_meta = rate_meta or {}
    return {
        # The mass-template letter + subject rendered for this draft — the
        # operator's per-draft note overrides the letter; None when
        # email_fields wasn't supplied (payloads that don't render the email).
        "email_letter_default": _letter.get("letter"),
        "email_subject_default": _letter.get("subject"),
        "id": d.id,
        "subscription_id": d.subscription_id,
        # Display name follows the LIVE subscription (renaming the offtaker
        # anywhere updates the draft card + email preview), not the draft's frozen
        # snapshot. Falls back to the draft's stored name if the sub is gone.
        "customer_name": ((getattr(sub, "customer_name", None) or d.customer_name) if sub else d.customer_name),
        "status": d.status,
        "period_label": d.period_label,
        "array_total_kwh": d.array_total_kwh,
        "allocation_pct": d.allocation_pct,
        "customer_kwh": d.customer_kwh,
        "amount_usd": d.amount_usd,
        "invoice_number": d.invoice_number,
        "has_gmp_pdf": d.gmp_invoice_pdf is not None,
        "gmp_filename": d.gmp_filename,
        "note": d.note,
        # Email-envelope fields so the inbox can render a FAITHFUL reproduction of
        # the message the offtaker receives (from/to/subject + which attachments).
        "client_email": getattr(sub, "client_email", None) if sub else None,
        "send_mode": getattr(sub, "send_mode", None) if sub else None,
        "include_summary": (getattr(sub, "include_summary", True) if sub else True),
        "operator_name": operator_name,
        # Auto-attach state (resolved from the subscription when provided):
        #   auto_attach_gmp   — is the toggle on for this customer
        #   gmp_auto_status   — "ready" (a captured bill PDF will attach) |
        #                       "pending" (toggle on, GMP account exists, no PDF
        #                       captured yet) | "no_gmp" (array has no GMP account)
        #                       | None (toggle off / not resolvable)
        "auto_attach_gmp": (getattr(sub, "auto_attach_gmp", False) if sub else None),
        "gmp_auto_status": gmp_auto_status,
        # The bound utility account's provider code (gmp/vec/wec/…), so the UI
        # labels the auto-attach checkbox + status by provider ("Auto-attach the
        # VEC bill" vs "…GMP bill") instead of hardcoding GMP. None when unbound.
        "attach_provider": attach_provider,
        # Editable offtaker details, surfaced so the approval inbox can edit the
        # offtaker inline (live-update) without leaving the draft. These mirror the
        # SubscriptionPatch fields; the money-affecting ones (allocation/discount/
        # rate/utility bill) recompute the draft via generate_draft on change.
        "cadence": (getattr(sub, "cadence", None) if sub else None),
        "cc_emails": (getattr(sub, "cc_emails", None) if sub else None),
        "discount_pct": (getattr(sub, "discount_pct", None) if sub else None),
        # The operator's per-customer OVERRIDE ($/kWh) — None when they haven't set
        # one (blank field = use the default). Distinct from the default below.
        "net_rate_per_kwh": (getattr(sub, "net_rate_per_kwh", None) if sub else None),
        # The bill-derived DEFAULT rate + its HONEST source/note (from this draft's
        # computed invoice), so the editable Solar-credit-rate field can show
        # "default: $X — from your GMP bill | comparable-months reference (banked)".
        # None in the list payload (computed only on the single-draft endpoints the
        # editor calls, to avoid an N+1 across the whole inbox).
        "default_net_rate_per_kwh": rate_meta.get("default_net_rate_per_kwh"),
        "default_net_rate_source": rate_meta.get("default_net_rate_source"),
        "default_net_rate_note": rate_meta.get("default_net_rate_note"),
        # The rate the invoice ACTUALLY used + its source ("customer" when the
        # override won) — lets the editor confirm an override is in force.
        "resolved_net_rate_per_kwh": rate_meta.get("resolved_net_rate_per_kwh"),
        "resolved_net_rate_source": rate_meta.get("resolved_net_rate_source"),
        "utility_account_id": (getattr(sub, "utility_account_id", None) if sub else None),
        "budget_amount_usd": (getattr(sub, "budget_amount_usd", None) if sub else None),
        # Calculated solar credit value (pre-budget-override). When a budget is set the
        # email shows BOTH: this value + the budgeted amount (amount_usd). None otherwise.
        "solar_credit_value": _calc_credit_for_budget(sub),
        "has_workbook": ((getattr(sub, "source_workbook", None) is not None) if sub else False),
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "sent_at": d.sent_at.isoformat() if d.sent_at else None,
    }


def _resolve_gmp_auto_status(db, sub, _prov_cache=None, _status_cache=None) -> Optional[str]:
    """Honest auto-attach status for the draft card (never implies a PDF exists
    when it doesn't). Provider-aware: a VEC/SmartHub-bound offtaker's status reflects
    its VEC bill (persisted Bill.pdf_bytes), a GMP-bound one its GMP bill. Kept named
    `gmp_auto_status` on the wire for back-compat; the frontend labels it by
    attach_provider. Statuses: ready | pending | no_gmp | None (toggle off).

    `_prov_cache`/`_status_cache` (Ford 2026-07-07 perf): the status is identical for
    every draft on the SAME account, so the draft-inbox loop (805 drafts on ~27
    accounts) passes these dicts to memoize the per-account provider + bill-PDF lookup
    — collapsing the per-draft query N+1 (~3s → ~0.3s). Omitted → queries as before."""
    if sub is None or not getattr(sub, "auto_attach_gmp", False):
        return None
    uaid = getattr(sub, "utility_account_id", None)
    array_id = getattr(sub, "array_id", None)
    ckey = uaid if uaid is not None else (("arr", array_id) if array_id is not None else "none")
    if _status_cache is not None and ckey in _status_cache:
        return _status_cache[ckey]
    result = "pending"
    try:
        from ..reports import gmp_bill_pdf_read as gbp
        prov = _bound_provider(db, sub, _prov_cache)
        if uaid is not None and prov and _is_smarthub_provider_code(prov):
            # VEC/SmartHub-bound → check the persisted VEC bill PDF for that account.
            found = gbp.get_vec_bill_pdf_for_account(uaid, db=db)
            result = "ready" if (found and found.get("bytes")) else "pending"
        elif uaid is not None:
            # Prefer the offtaker's BOUND utility account — the exact bill the invoice
            # is computed from, so the attached PDF matches the invoice's source.
            found = gbp.get_bill_pdf_for_account(uaid, db=db)
            result = "ready" if (found and found.get("bytes")) else "pending"
        elif array_id is None:
            result = "no_gmp"
        elif not gbp.has_capturable_gmp_account(array_id, db=db):
            result = "no_gmp"
        else:
            found = gbp.get_bill_pdf_for_period(array_id, db=db)
            result = "ready" if (found and found.get("bytes")) else "pending"
    except Exception:  # noqa: BLE001
        result = "pending"
    if _status_cache is not None:
        _status_cache[ckey] = result
    return result


def _bound_provider(db, sub, _cache=None) -> Optional[str]:
    """Lowercase provider code of the subscription's bound utility account, or None.
    Drives provider-aware auto-attach label + status. Best-effort/fail-safe.
    `_cache` (uaid→provider) memoizes the lookup across a draft-inbox loop."""
    uaid = getattr(sub, "utility_account_id", None) if sub else None
    if uaid is None:
        return None
    if _cache is not None and uaid in _cache:
        return _cache[uaid]
    try:
        from ..models import UtilityAccount
        ua = db.get(UtilityAccount, uaid)
        val = (ua.provider or "").lower() or None if ua else None
    except Exception:  # noqa: BLE001
        val = None
    if _cache is not None:
        _cache[uaid] = val
    return val


def _is_smarthub_provider_code(code: Optional[str]) -> bool:
    if not code:
        return False
    try:
        from ..adapters.smarthub import is_smarthub_provider
        return is_smarthub_provider(code)
    except Exception:  # noqa: BLE001
        return False


def _get_owned_draft(db, tenant_id: str, draft_id: int) -> ReportDraft:
    d = db.get(ReportDraft, draft_id)
    if d is None or d.tenant_id != tenant_id:
        raise HTTPException(404, "Draft not found")
    return d


@router.get("/drafts")
def list_drafts(status: str = Query(default="pending"),
                authorization: Optional[str] = Header(default=None)):
    """The approval inbox: drafts awaiting the operator's review. status=pending
    (default) | sent | dismissed | all."""
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        q = select(ReportDraft).where(ReportDraft.tenant_id == t.id)
        if status != "all":
            q = q.where(ReportDraft.status == status)
        rows = db.execute(q.order_by(ReportDraft.created_at.desc())).scalars().all()
        # One tenant fetch for the mass-template fields, shared by every draft
        # in the loop (per-draft fetches would be the next 800-scale N+1).
        from .delivery import _offtaker_email_fields
        email_fields = _offtaker_email_fields(t.id)
        # Batch the per-draft lookups (Ford 2026-07-07: /drafts was ~3s at 805 drafts,
        # an N+1 of sub-fetch + provider + bill-PDF-status PER draft). One query for
        # all the subs, and per-ACCOUNT provider/status caches shared across the loop
        # (805 drafts sit on ~27 accounts) → ~3s to ~0.3s. Same output, fewer queries.
        _sub_ids = [d.subscription_id for d in rows if d.subscription_id]
        subs_by_id = {}
        if _sub_ids:
            subs_by_id = {s.id: s for s in db.execute(
                select(BillingReportSubscription).where(
                    BillingReportSubscription.id.in_(set(_sub_ids)))
            ).scalars().all()}
        _prov_cache: dict = {}
        _status_cache: dict = {}
        operator_name = getattr(t, "name", None)
        out = []
        for d in rows:
            sub = subs_by_id.get(d.subscription_id) if d.subscription_id else None
            out.append(_draft_dict(d, sub=sub,
                                   gmp_auto_status=_resolve_gmp_auto_status(db, sub, _prov_cache, _status_cache),
                                   operator_name=operator_name,
                                   email_fields=email_fields,
                                   attach_provider=_bound_provider(db, sub, _prov_cache)))
        return {"drafts": out}


@router.get("/subscriptions/{sub_id}/draft-versions")
def draft_versions(sub_id: int, authorization: Optional[str] = Header(default=None)):
    """All draft VERSIONS for one offtaker — one per billing period, LATEST FIRST — so the
    operator can look back at older invoices in the approval inbox. The newest is the live
    pending draft; earlier periods are superseded/sent. Read-only."""
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        drafts = db.execute(
            select(ReportDraft).where(ReportDraft.subscription_id == sub.id)
            .order_by(ReportDraft.created_at.desc())
        ).scalars().all()
        # Dedupe to the most-recent draft per billing period; sort latest period first.
        seen, out = set(), []
        for d in drafts:
            key = d.period_label or ("#" + str(d.id))
            if key in seen:
                continue
            seen.add(key)
            out.append(d)
        out.sort(key=lambda d: (d.period_label or ""), reverse=True)
        return {"versions": [_draft_dict(d, sub) for d in out]}


@router.get("/subscriptions/{sub_id}/bill-periods")
def subscription_bill_periods(sub_id: int,
                              authorization: Optional[str] = Header(default=None)):
    """The billable periods the operator can DRAFT for this offtaker (Bruce
    2026-07-07, Comment 4) — every settled utility bill with excess, newest
    first, plus which one is the implicit latest default. Feeds the period
    selector on the draft flow. Read-only; empty `periods` = keep the implicit
    latest-bill default (nothing to choose between)."""
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        cadence = getattr(sub, "cadence", None) or "monthly"
        from .delivery import settled_periods_for_sub
        periods = settled_periods_for_sub(sub)
        return {"ok": True, "cadence": cadence, "periods": periods}


@router.post("/subscriptions/{sub_id}/draft")
def generate_draft(sub_id: int, period: Optional[str] = Query(default=None),
                   authorization: Optional[str] = Header(default=None)):
    """Build a pending draft for this subscription's billing period, from its
    stored workbook. This is what the (operator-built) GMP-detection backend will
    call when a new GMP invoice lands; the operator can also trigger it manually.
    Reuses an existing pending draft for the same period (idempotent).

    `period` (Bruce 2026-07-07, Comment 4) targets a SPECIFIC settled bill period
    instead of the implicit latest — 'YYYY-MM' (monthly) or 'YYYY-Qn' (quarterly),
    from GET /subscriptions/{id}/bill-periods. Omit it to draft the latest bill
    (the long-standing default). A chosen historical period NEVER supersedes the
    other pending drafts (only a latest/default draft does), so the operator can
    hold several periods side by side in the approval inbox."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    # An explicit period pins a specific historical bill; None = the latest bill.
    target_period = (period or "").strip() or None
    with SessionLocal() as db:
        sub = _get_owned(db, t.id, sub_id)
        # ALWAYS pull the freshest GMP bill for this offtaker's bound account first, so
        # the invoice reflects the latest statement the moment it's generated — don't wait
        # for the 6h scheduler. Best-effort: a lapsed session (or a vendor hiccup) just
        # falls back to the last captured bill, so a refresh failure never blocks a draft.
        # READ COMMITTED → build_match's fresh bill query below sees the committed pull.
        uaid = getattr(sub, "utility_account_id", None)
        if uaid:
            try:
                from ..worker import pull_account_bills
                pull_account_bills(t.id, uaid)
            except Exception:  # noqa: BLE001
                pass
        try:
            match = build_match(sub, period_label=target_period)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(422, f"workbook unreadable: {e}")
        if not match.matched or not match.latest_period:
            if target_period:
                raise HTTPException(
                    422, f"no settled bill for {target_period} on this offtaker's "
                         "account — pick another billing period.")
            raise HTTPException(422, "no current billing period in the stored workbook")

        ci = match.computed_invoice or {}
        inv_no = ci.get("invoice_number")
        period_label = None
        if ci.get("period_start") or ci.get("period_end"):
            period_label = f"{ci.get('period_start') or '—'} → {ci.get('period_end') or '—'}"

        # Idempotent + self-healing. A draft generated BEFORE the customer's
        # first utility bill landed has invoice_number=None and no period; once
        # the bill arrives its invoice_number/period drift to e.g. "2026-05", so
        # the old strict `invoice_number == inv_no` match MISSED that placeholder
        # and spawned a SECOND pending draft — a stale $0 duplicate in the
        # approval inbox (hit live on Paul Bozuwa / HCT Sun, draft 8, hand-
        # refreshed via a prod script). Instead, refresh THE pending draft that
        # belongs to this period — same invoice number, same period label, or the
        # not-yet-periodised placeholder (invoice_number NULL) — and fold any
        # extra duplicates into one so the invariant "one pending draft per
        # (subscription, period)" holds going forward.
        pendings = db.execute(
            select(ReportDraft).where(
                ReportDraft.subscription_id == sub.id,
                ReportDraft.status == "pending",
            ).order_by(ReportDraft.created_at.asc())
        ).scalars().all()

        def _same_period(dr: ReportDraft) -> bool:
            if inv_no is not None and dr.invoice_number == inv_no:
                return True            # exact period key — true idempotency
            if dr.invoice_number is None:
                return True            # pre-bill placeholder — adopt it now
            if period_label is not None and dr.period_label == period_label:
                return True            # invoice number reformatted, same period
            return False

        matches = [dr for dr in pendings if _same_period(dr)]
        existing = matches[0] if matches else None
        for dup in matches[1:]:        # collapse pre-existing duplicates
            dup.status = "dismissed"
            dup.dismissed_at = datetime.utcnow()
        # SUPERSEDE older-period drafts — ONLY when drafting the implicit latest bill.
        # The default draft uses the LATEST bill, so any OTHER pending draft is a stale
        # earlier period that must not linger in front of it (Paul Bozuwa: a May $3,167
        # draft sat in front of the new June one). But when the operator EXPLICITLY picks
        # a historical period (Bruce, Comment 4), leave the others alone — they may be
        # deliberately holding several periods for review side by side.
        if not target_period:
            for dr in pendings:
                if dr not in matches:
                    dr.status = "dismissed"
                    dr.dismissed_at = datetime.utcnow()

        d = existing or ReportDraft(
            tenant_id=t.id, subscription_id=sub.id,
            customer_name=sub.customer_name, status="pending")
        d.period_label = period_label
        # The period's TOTAL array generation is Paul's anchor number ("what GMP
        # reports"). The computed invoice carries the customer's SHARE (kwh) and
        # the allocation %, so the array total = share / pct. Fall back to any
        # explicit array field, else null.
        cust_kwh = ci.get("kwh")
        pct = match.allocation_pct
        array_total = ci.get("project_total_kwh") or ci.get("array_kwh")
        if array_total is None and cust_kwh is not None and pct:
            array_total = round(cust_kwh / pct, 1)
        d.array_total_kwh = array_total
        d.allocation_pct = pct
        d.customer_kwh = cust_kwh
        d.amount_usd = ci.get("amount_owed")
        d.invoice_number = inv_no
        if existing is None:
            db.add(d)
        db.commit()
        # Bruce's automatic cross-check (2026-07): at the moment an invoice is
        # generated, verify GMP's allocation against the operator's entered share
        # in the background and surface the verdict WITH the draft — the check
        # "pops up when an invoice is generated", no button. Single-subscription
        # only (~tens of ms); the batch surface stays /reconcile-bills. Fail-soft
        # by contract: a cross-check that can't run is null, never a blocker.
        try:
            from .reconcile_bills import generation_crosscheck
            crosscheck = generation_crosscheck(db, sub)
        except Exception:  # noqa: BLE001 — never block a draft on the cross-check
            crosscheck = None
        # Pass `sub` so the recompute response carries the SUBSCRIPTION-derived fields
        # (budget_amount_usd + the CALCULATED solar_credit_value) — without it the
        # frontend's post-edit refresh got nulls, so the "How we calculated" panel lost
        # the budget split and back-derived a fake rate (budget ÷ kWh) from amount_usd,
        # and the email preview collapsed to one row. Mirror the /drafts list overlay.
        from .delivery import _offtaker_email_fields
        return {"ok": True, "draft": _draft_dict(
            d, sub=sub,
            gmp_auto_status=_resolve_gmp_auto_status(db, sub),
            operator_name=getattr(t, "name", None),
            email_fields=_offtaker_email_fields(t.id),
            attach_provider=_bound_provider(db, sub),
            rate_meta=_rate_meta_from_ci(ci)),
            "crosscheck": crosscheck}


@router.post("/drafts/{draft_id}/gmp-invoice")
async def attach_gmp_invoice(draft_id: int, file: UploadFile = File(...),
                             authorization: Optional[str] = Header(default=None)):
    """Attach the period's GMP utility-invoice PDF to a draft. Paul sends this
    alongside the customer invoice 'to prove we're not just making this up.'"""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    data = await _read_capped(file, MAX_PDF_BYTES)
    if not data:
        raise HTTPException(400, "empty file")
    head = data[:5]
    if head[:4] != b"%PDF":
        raise HTTPException(400, "that doesn't look like a PDF")
    with SessionLocal() as db:
        d = _get_owned_draft(db, t.id, draft_id)
        if d.status != "pending":
            raise HTTPException(409, "draft already resolved")
        d.gmp_invoice_pdf = data
        d.gmp_filename = (file.filename or "GMP_invoice.pdf")[:300]
        db.commit()
        return {"ok": True, "draft": _draft_dict(d)}


@router.get("/drafts/{draft_id}/gmp-bill")
def get_draft_gmp_bill(draft_id: int, authorization: Optional[str] = Header(default=None)):
    """Stream the GMP utility-bill PDF that rides with this draft — the manually
    attached one if present, else the auto-captured bill for the period (the same
    PDF the send path attaches). 404 when none is captured yet (the 'pending' chip)."""
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        d = _get_owned_draft(db, t.id, draft_id)
        if d.gmp_invoice_pdf:
            fn = "".join(c for c in (d.gmp_filename or "gmp_bill.pdf")
                         if c.isalnum() or c in "._- ") or "gmp_bill.pdf"
            return StreamingResponse(io.BytesIO(d.gmp_invoice_pdf), media_type="application/pdf",
                headers={"Content-Disposition": "inline; filename=" + fn})
        sub = db.get(BillingReportSubscription, d.subscription_id) if d.subscription_id else None
        if sub is not None and getattr(sub, "auto_attach_gmp", False):
            # Resolve the SAME PDF the send path attaches: the BOUND utility account's
            # bill for THIS draft's period — so the download matches the invoice exactly
            # (incl. older versions in the version dropdown), not just the global newest
            # across the array. Falls back to the array-keyed lookup only for legacy
            # array-based subs with no bound account.
            from ..reports import gmp_bill_pdf_read as gbp
            from .delivery import _parse_iso_date
            ps = pe = None
            if d.period_label and "→" in d.period_label:
                parts = [p.strip() for p in d.period_label.split("→")]
                if len(parts) == 2:
                    ps, pe = _parse_iso_date(parts[0]), _parse_iso_date(parts[1])
            uaid = getattr(sub, "utility_account_id", None)
            found = None
            try:
                if uaid is not None:
                    # Provider-aware: get_bill_pdf_for_account returns None for a
                    # non-GMP account, so fall back to the SmartHub/VEC sibling — a VEC
                    # offtaker's auto-captured bill lives in the same Bill.pdf_bytes but
                    # only the VEC reader surfaces it (Ford 2026-07-07: Glover VEC bill
                    # 404'd from this route as "pending" though it was captured).
                    found = (gbp.get_bill_pdf_for_account(uaid, ps, pe, db=db)
                             or gbp.get_vec_bill_pdf_for_account(uaid, ps, pe, db=db))
                elif getattr(sub, "array_id", None):
                    found = gbp.get_bill_pdf_for_period(sub.array_id, ps, pe, db=db)
            except Exception:
                found = None
            if found and found.get("bytes"):
                _fn = "".join(c for c in (found.get("filename") or "utility_bill.pdf")
                              if c.isalnum() or c in "._- ") or "utility_bill.pdf"
                return StreamingResponse(io.BytesIO(found["bytes"]),
                    media_type=found.get("content_type") or "application/pdf",
                    headers={"Content-Disposition": "inline; filename=" + _fn})
        raise HTTPException(404, "No utility bill captured for this period yet")


class DraftPatch(BaseModel):
    note: Optional[str] = None


@router.patch("/drafts/{draft_id}")
def patch_draft(draft_id: int, body: DraftPatch,
                authorization: Optional[str] = Header(default=None)):
    """Modify a draft before sending (Paul: 'approve it or MODIFY it and send')."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        d = _get_owned_draft(db, t.id, draft_id)
        if d.status != "pending":
            raise HTTPException(409, "draft already resolved")
        if body.note is not None:
            d.note = body.note[:2000]
        db.commit()
        return {"ok": True, "draft": _draft_dict(d)}


@router.post("/drafts/{draft_id}/approve")
def approve_draft(draft_id: int, authorization: Optional[str] = Header(default=None)):
    """Approve & send: deliver the drafted report to the customer (per the
    subscription's recipient slider), attaching the GMP invoice PDF the operator
    put on the draft. This is the single human gate in front of delivery."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        d = _get_owned_draft(db, t.id, draft_id)
        # #17: claim the draft atomically before the slow send. Lock the row,
        # re-read status inside the txn, and flip pending→sending; a second
        # concurrent approve (double-click / client retry) then sees it's no
        # longer pending and 409s instead of sending a duplicate invoice.
        locked = db.execute(
            select(ReportDraft).where(ReportDraft.id == d.id).with_for_update()
        ).scalars().first()
        if locked is None or locked.status != "pending":
            raise HTTPException(409, "draft already resolved")
        locked.status = "sending"
        db.commit()
        d = locked
        sub = _get_owned(db, t.id, d.subscription_id)
        # #4: attach the draft's manually-uploaded GMP bill for THIS send only —
        # passed through, never persisted onto the sub (persisting it made a stale
        # bill ride every future period's invoice and defeat period-correct auto-
        # attach). #3: pin the period the operator reviewed so a bill that landed
        # after review can't be sent unreviewed.
        gmp_override = bytes(d.gmp_invoice_pdf) if d.gmp_invoice_pdf is not None else None
        try:
            result = deliver_subscription(
                db, sub, t, triggered_by="approval", is_test=False, note=d.note,
                expected_period_label=d.period_label, gmp_pdf_override=gmp_override)
        except Exception:
            d.status = "pending"      # release the claim so it can be retried
            db.commit()
            raise
        if not result.get("ok"):
            d.status = "pending"      # release the claim; nothing was sent
            db.commit()
            raise HTTPException(422, result.get("error", "send failed"))
        d.status = "sent"
        d.sent_at = datetime.utcnow()
        db.commit()
        return {"ok": True, "draft": _draft_dict(d), "result": result}


@router.post("/drafts/{draft_id}/test")
def test_draft(draft_id: int, authorization: Optional[str] = Header(default=None)):
    """Send a TEST copy of this draft (the exact email + invoice the offtaker would
    get, with your edited note) to the OPERATOR — you — so you can check it before
    approving. Goes only to you; never reaches the customer; bypasses the
    utility-bill send gate since it's a self-test."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        d = _get_owned_draft(db, t.id, draft_id)
        sub = _get_owned(db, t.id, d.subscription_id)
        # Attach the draft's GMP PDF for this test send only, passed through (never
        # written onto the sub) — same per-send scoping as approve (#4).
        gmp_override = bytes(d.gmp_invoice_pdf) if d.gmp_invoice_pdf is not None else None
        result = deliver_subscription(db, sub, t, triggered_by="test",
                                      is_test=True, note=d.note,
                                      gmp_pdf_override=gmp_override)
        if not result.get("ok"):
            raise HTTPException(422, result.get("error", "test send failed"))
        return {"ok": True, "result": result}


@router.post("/drafts/{draft_id}/dismiss")
def dismiss_draft(draft_id: int, authorization: Optional[str] = Header(default=None)):
    """Discard a draft without sending."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        d = _get_owned_draft(db, t.id, draft_id)
        if d.status == "sent":
            raise HTTPException(409, "draft already sent")
        d.status = "dismissed"
        d.dismissed_at = datetime.utcnow()
        db.commit()
        return {"ok": True}


@router.post("/drafts/{draft_id}/ai-email")
def ai_email_for_draft(draft_id: int, authorization: Optional[str] = Header(default=None)):
    """Write a tailored cover email for this draft from the REAL invoice figures +
    the context that applies (budget/flat bill, annual true-up, banked-not-cashed
    credits, $0 period). Returns the text for the operator to review/edit — it does
    NOT save or send. Plain text, no invented numbers."""
    t = tenant_from_session(authorization)
    from .repro.llm import call_json, llm_available, LLMUnavailable
    if not llm_available():
        raise HTTPException(503, "AI email isn't configured (no ANTHROPIC_API_KEY).")
    with SessionLocal() as db:
        d = _get_owned_draft(db, t.id, draft_id)
        sub = db.get(BillingReportSubscription, d.subscription_id) if d.subscription_id else None
        ci = {}
        if sub is not None:
            try:
                ci = build_match(sub).computed_invoice or {}
            except Exception:  # noqa: BLE001
                ci = {}
        operator = (getattr(t, "company_name", None) or getattr(t, "operator_name", None)
                    or getattr(t, "name", None) or "your solar operator")
        # The ReportDraft row is a FROZEN snapshot — the offtaker's CURRENT name + the
        # recomputed amount live on the subscription / fresh match (this is exactly what
        # _draft_dict overlays for the card). Read d.* directly and you resurrect stale
        # values, e.g. a since-corrected name typo. Mirror the live overlay here.
        name = (getattr(sub, "customer_name", None) or d.customer_name) if sub else d.customer_name
        amt = ci.get("amount_owed")
        if amt is None:
            amt = d.amount_usd
        ctx = {
            "offtaker_name": name,
            "billing_period": d.period_label,
            "their_production_kwh": d.customer_kwh,
            "amount_due_usd": amt,
            "is_budget_bill": getattr(sub, "budget_amount_usd", None) is not None,
            "is_annual_trueup": bool(getattr(sub, "annual_trueup", False)),
            "solar_credit_rate_per_kwh": ci.get("net_rate_per_kwh"),
            "credit_banked_not_cashed": ci.get("net_rate_source") == "gmp_credit_reference",
            "zero_due": isinstance(amt, (int, float)) and abs(amt) < 0.005,
            "operator_company": operator,
            "attachments": "the invoice PDF" + (
                " and a production summary" if getattr(sub, "include_summary", True) else ""),
        }
        system = (
            "You write the short cover email a solar operator sends to an offtaker "
            "alongside their solar-credit invoice. Warm, professional, concise — 3 to 5 "
            "sentences. Use ONLY the real figures provided. Reflect whichever context "
            "applies: a flat/budget bill (do not imply it was metered this period), an "
            "annual true-up, credits that were BANKED not cashed (trued up later), or a "
            "$0 period. Open with a simple greeting, mention the billing period and the "
            "amount, note that the attachments are included, and sign off as the "
            "operator's company. Reference the figures plainly but DO NOT claim that "
            "kWh times the rate equals the total — a discount may apply, so the "
            "arithmetic won't line up. Plain text only — no subject line, no markdown, "
            "no placeholder brackets, no invented numbers."
        )
        try:
            out = call_json(
                system=system,
                user_text="Invoice context (JSON):\n" + json.dumps(ctx, default=str),
                max_tokens=600,
                schema={"type": "object",
                        "properties": {"email": {"type": "string"}},
                        "required": ["email"],
                        "additionalProperties": False})
        except LLMUnavailable:
            raise HTTPException(503, "AI email isn't configured.")
        except Exception as e:  # noqa: BLE001
            logger.warning("ai-email generation failed: %s", e)
            raise HTTPException(502, "Couldn't write the email right now — try again.")
        email = (out.get("email") or "").strip()
        if not email:
            raise HTTPException(502, "The AI returned an empty email — try again.")
        return {"ok": True, "email": email}


# ─── offtaker invoice TEMPLATE (operator's own format) ──────────────────────────
# Stage 1: upload + store the operator's invoice template per-tenant so generated
# offtaker invoices can LATER reproduce THEIR exact format. Rendering from the
# template (Stage 2) is gated behind `enabled`; storing one changes nothing about
# what is actually sent today (offtaker invoices keep using the standard PDF).

TEMPLATE_MAX_BYTES = 12 * 1024 * 1024  # 12 MB
TEMPLATE_EXT = {".pdf", ".html", ".htm", ".docx", ".doc", ".png", ".jpg", ".jpeg",
                ".xlsx", ".xls", ".xlsm"}  # Excel: we find the invoice sheet & seed HTML


async def _read_template_upload(file: UploadFile):
    data = await _read_capped(file, TEMPLATE_MAX_BYTES)
    if not data:
        raise HTTPException(400, "The uploaded file is empty")
    name = (file.filename or "template").strip()
    ext = ("." + name.rsplit(".", 1)[1].lower()) if "." in name else ""
    # Detect actual type by magic bytes and override the extension when possible.
    if data[:4] == _MAGIC_PDF:
        ext = ".pdf"
    elif data[:4] == _MAGIC_XLSX[:4]:
        if ext not in (".xlsx", ".xlsm"):
            ext = ".xlsx"   # mis-named OpenXML file — treat as xlsx
    elif data[:4] == _MAGIC_XLS:
        ext = ".xls"
    if ext not in TEMPLATE_EXT:
        raise HTTPException(400, "Upload a PDF, Word doc, HTML, or image of your invoice template")
    if ext in (".xlsx", ".xlsm", ".xls"):
        _reject_zip_bomb(data)   # #34: refuse a decompression bomb before openpyxl
    return data, name


def _template_dict(tpl) -> dict:
    return {
        "has_template": tpl is not None and (tpl.file_bytes is not None or bool(tpl.html)),
        "filename": getattr(tpl, "filename", None),
        "content_type": getattr(tpl, "content_type", None),
        "enabled": bool(getattr(tpl, "enabled", False)),
        "has_html": bool(getattr(tpl, "html", None)),
        "updated_at": tpl.updated_at.isoformat() if tpl and tpl.updated_at else None,
    }


class InvoiceTemplatePut(BaseModel):
    html: Optional[str] = None
    enabled: Optional[bool] = None


class InvoiceTemplatePreview(BaseModel):
    html: Optional[str] = None


@router.get("/invoice-template")
def get_invoice_template(authorization: Optional[str] = Header(default=None)):
    """Status + the editable token-HTML of this tenant's invoice template (seeds the
    default template when none saved, so the editor/preview work out of the box)."""
    t = tenant_from_session(authorization)
    from ..models import OfftakerInvoiceTemplate
    from .template_render import DEFAULT_TEMPLATE_HTML, AVAILABLE_TOKENS
    with SessionLocal() as db:
        tpl = db.execute(select(OfftakerInvoiceTemplate).where(
            OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
        d = _template_dict(tpl)
        d["html"] = (tpl.html if tpl and tpl.html else DEFAULT_TEMPLATE_HTML)
        d["is_default_html"] = not (tpl and tpl.html)
        d["tokens"] = AVAILABLE_TOKENS
        return {"ok": True, "template": d}


@router.put("/invoice-template")
def put_invoice_template(body: InvoiceTemplatePut,
                         authorization: Optional[str] = Header(default=None)):
    """Save the editable token-HTML + the enable toggle. Enabling means real offtaker
    invoices render from this template (Stage 2); a render failure at send time falls
    back to the standard PDF, so it can never break a send."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    from ..models import OfftakerInvoiceTemplate
    from .template_render import DEFAULT_TEMPLATE_HTML
    with SessionLocal() as db:
        tpl = db.execute(select(OfftakerInvoiceTemplate).where(
            OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
        if tpl is None:
            tpl = OfftakerInvoiceTemplate(tenant_id=t.id)
            db.add(tpl)
        if "html" in body.model_fields_set:
            tpl.html = body.html
        if "enabled" in body.model_fields_set:
            tpl.enabled = bool(body.enabled)
            # Enabling with no custom HTML seeds the default so it works immediately.
            if tpl.enabled and not tpl.html:
                tpl.html = DEFAULT_TEMPLATE_HTML
        tpl.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(tpl)
        return {"ok": True, "template": _template_dict(tpl)}


@router.post("/invoice-template/preview")
def preview_invoice_template(body: InvoiceTemplatePreview,
                             authorization: Optional[str] = Header(default=None)):
    """Render the (provided or stored) token-HTML with SAMPLE data → PDF preview."""
    t = tenant_from_session(authorization)
    # Rendering CALLER-SUPPLIED html is a compute action on operator input; keep it
    # off the shared demo session (defense-in-depth alongside the sandboxed renderer,
    # #1). Previewing the stored/default template (html is None) stays open.
    if body.html is not None:
        require_not_demo(t)
    from ..models import OfftakerInvoiceTemplate
    from .template_render import render_template_pdf, SAMPLE_CONTEXT, DEFAULT_TEMPLATE_HTML
    html = body.html
    if html is None:
        with SessionLocal() as db:
            tpl = db.execute(select(OfftakerInvoiceTemplate).where(
                OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
            html = (tpl.html if tpl and tpl.html else DEFAULT_TEMPLATE_HTML)
    try:
        pdf = render_template_pdf(html, SAMPLE_CONTEXT)
    except Exception as e:  # noqa: BLE001 — surface a readable message to the editor
        raise HTTPException(400, f"Couldn't render this template: {e}")
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=invoice_preview.pdf"})


@router.post("/invoice-template")
async def upload_invoice_template(file: UploadFile = File(...),
                                  authorization: Optional[str] = Header(default=None)):
    """Upload (or replace) the operator's own invoice template. Stored per-tenant;
    does NOT change what is sent today (render-from-template is gated, Stage 2)."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    data, name = await _read_template_upload(file)
    from ..models import OfftakerInvoiceTemplate
    with SessionLocal() as db:
        tpl = db.execute(select(OfftakerInvoiceTemplate).where(
            OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
        if tpl is None:
            tpl = OfftakerInvoiceTemplate(tenant_id=t.id)
            db.add(tpl)
        tpl.filename = name[:300]
        tpl.content_type = file.content_type or "application/octet-stream"
        tpl.file_bytes = data
        lname = name.lower()
        is_excel = (lname.endswith((".xlsx", ".xls", ".xlsm"))
                    or data[:4] == _MAGIC_XLSX[:4] or data[:4] == _MAGIC_XLS)
        if lname.endswith((".html", ".htm")):
            try:
                tpl.html = data.decode("utf-8", "replace")
            except Exception:
                pass
        elif is_excel:
            # Find the invoice sheet anywhere in the workbook and seed editable
            # token-HTML from it (Stage-1; rendering stays opt-in). Never fatal —
            # a failed extract just leaves the editor on the default template.
            try:
                from .matcher import excel_to_template_html
                sheet, html = excel_to_template_html(data)
                if html:
                    tpl.html = html
                    logger.info("invoice template: seeded HTML from Excel sheet %r", sheet)
            except Exception as e:  # noqa: BLE001
                logger.warning("invoice template Excel extract failed: %s", e)
        # A freshly uploaded template is used by DEFAULT — uploading IS the opt-in,
        # so the operator doesn't have to separately flip it on (they switch to the
        # standard format with the slider). Only when it's actually renderable
        # (xlsx pixel repro, or seeded token-HTML); otherwise leave it off.
        # BUT (#11): an .xlsx exported without cached formula values (Google Sheets /
        # LibreOffice) reads blank, so auto-enabling it would email a blank Amount
        # Due. In that case DON'T enable (turn it off if it was on) and warn — the
        # operator opens+saves it in Excel, or uses the token editor.
        warning = None
        if is_excel and _xlsx_formula_values_missing(data):
            tpl.enabled = False
            warning = ("We stored your template but couldn't read its computed values — "
                       "it looks like it was exported without cached formula results "
                       "(e.g. from Google Sheets). Open it in Excel and Save once, then "
                       "re-upload, or use the token editor. It is NOT set as your live "
                       "invoice format yet, so nothing sends blank.")
        elif is_excel or tpl.html:
            tpl.enabled = True
        tpl.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(tpl)
        out = {"ok": True, "template": _template_dict(tpl)}
        if warning:
            out["warning"] = warning
        return out


@router.get("/invoice-template/file")
def get_invoice_template_file(authorization: Optional[str] = Header(default=None)):
    """Stream the stored original template file (inline preview / download)."""
    t = tenant_from_session(authorization)
    from ..models import OfftakerInvoiceTemplate
    with SessionLocal() as db:
        tpl = db.execute(select(OfftakerInvoiceTemplate).where(
            OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
        if not tpl or not tpl.file_bytes:
            raise HTTPException(404, "No invoice template on file")
        ct = tpl.content_type or "application/octet-stream"
        fn = "".join(c for c in (tpl.filename or "template") if c.isalnum() or c in "._- ") or "template"
        return StreamingResponse(io.BytesIO(tpl.file_bytes), media_type=ct,
            headers={"Content-Disposition": "inline; filename=" + fn})


@router.get("/invoice-template/preview.pdf")
def get_invoice_template_preview_pdf(default: bool = False,
                                     authorization: Optional[str] = Header(default=None)):
    """Render OUR REPRODUCTION of the operator's template, the RIGHT way: the real
    uploaded file rendered by the real engine (xlsx/Word → Gotenberg/LibreOffice;
    PDF passthrough) — pixel-identical to their format. Deep-research finding: never
    reproduce a document by converting it to HTML (the lossy anti-pattern); fill/
    render the real artifact. Token-HTML is kept only as a last-resort fallback for
    image/HTML templates or when no renderer is configured.

    ?default=1 renders OUR DEFAULT invoice format (the standard Array Operator layout)
    with the same sample data instead of the reproduction, so the operator can compare
    their own template against our default."""
    t = tenant_from_session(authorization)
    if default:
        # OUR default format = the REAL Array Operator invoice renderer
        # (invoice.render_invoice_pdf — the same one a live send uses), rendered
        # from the operator's LATEST computable offtaker so the button shows their
        # actual latest invoice in our standard layout, not a generic token-HTML
        # sample. (Old behavior rendered DEFAULT_TEMPLATE_HTML + SAMPLE_CONTEXT —
        # a fake mock that never reflected the real invoice.) Falls back to that
        # sample only when the operator has no offtaker with a computable invoice.
        import tempfile, pathlib as _pl
        from . import invoice as _inv
        chosen = None
        with SessionLocal() as db:
            subs = db.execute(select(BillingReportSubscription).where(
                BillingReportSubscription.tenant_id == t.id,
                BillingReportSubscription.deleted_at.is_(None),
            ).order_by(BillingReportSubscription.id.desc())).scalars().all()
            for s in subs:
                try:
                    m = build_match(s)
                except Exception:  # noqa: BLE001
                    continue
                ci = m.computed_invoice or {}
                if m.matched and isinstance(ci.get("amount_owed"), (int, float)) and ci.get("amount_owed"):
                    chosen = m
                    break
        if chosen is not None:
            try:
                with tempfile.TemporaryDirectory(prefix="ao-defprev-") as tmp:
                    p = _inv.render_invoice_pdf(chosen, _pl.Path(tmp) / "default.pdf")
                    pdf = p.read_bytes()
                return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
                    headers={"Content-Disposition": "inline; filename=default_format_preview.pdf"})
            except Exception as e:  # noqa: BLE001
                logger.warning("default-format real-invoice render failed, falling back: %s", e)
        from .template_render import (render_template_pdf, SAMPLE_CONTEXT,
                                      DEFAULT_TEMPLATE_HTML)
        try:
            pdf = render_template_pdf(DEFAULT_TEMPLATE_HTML, SAMPLE_CONTEXT)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(422, f"Couldn't render default-format preview: {e}")
        return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
            headers={"Content-Disposition": "inline; filename=default_format_preview.pdf"})
    from ..models import OfftakerInvoiceTemplate
    with SessionLocal() as db:
        tpl = db.execute(select(OfftakerInvoiceTemplate).where(
            OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
        if not tpl or (not tpl.file_bytes and not tpl.html):
            raise HTTPException(404, "No invoice template on file")
        fb = bytes(tpl.file_bytes) if tpl.file_bytes else b""
        name = (tpl.filename or "template").lower()
        pdf = None
        if fb[:4] == b"%PDF":
            pdf = fb                                    # already a PDF — passthrough
        elif (fb[:4] in (b"PK\x03\x04", b"\xd0\xcf\x11\xe0")
              or name.endswith((".xlsx", ".xls", ".docx", ".doc", ".odt", ".ods"))):
            # OUR reproduction (direct-cell-write pipeline), not the raw upload — for
            # xlsx, write a sample through the same path a real send uses so the pane
            # shows what our engine produces. Falls back to a plain render otherwise.
            if fb[:4] == b"PK\x03\x04":
                try:
                    from .repro.template_repro import reproduce_template_preview
                    pdf = reproduce_template_preview(fb)
                except Exception as e:  # noqa: BLE001
                    logger.warning("template reproduction preview failed: %s", e)
            if pdf is None:
                try:
                    from .repro import render as _repro_render
                    if _repro_render.renderer_available():
                        pdf = _repro_render.render_office_to_pdf(fb, tpl.filename or "template.xlsx")
                except Exception as e:  # noqa: BLE001
                    logger.warning("template preview headless render failed: %s", e)
        if pdf is None:
            from .template_render import (render_template_pdf, SAMPLE_CONTEXT,
                                          DEFAULT_TEMPLATE_HTML)
            try:
                pdf = render_template_pdf(tpl.html or DEFAULT_TEMPLATE_HTML, SAMPLE_CONTEXT)
            except Exception as e:  # noqa: BLE001
                raise HTTPException(422, f"Couldn't render template preview: {e}")
    return StreamingResponse(io.BytesIO(pdf), media_type="application/pdf",
        headers={"Content-Disposition": "inline; filename=template_preview.pdf"})


@router.delete("/invoice-template")
def delete_invoice_template(authorization: Optional[str] = Header(default=None)):
    """Remove the tenant's invoice template."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    from ..models import OfftakerInvoiceTemplate
    with SessionLocal() as db:
        tpl = db.execute(select(OfftakerInvoiceTemplate).where(
            OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
        if tpl:
            db.delete(tpl)
            db.commit()
        return {"ok": True}


# ── Per-offtaker invoice template ────────────────────────────────────────────
# Each offtaker can have its OWN uploaded template; it OVERRIDES the tenant-wide
# default at render time (see delivery._effective_template_row). These mirror the
# tenant endpoints above, scoped to one subscription the caller owns.

def _owned_sub(db, t, sub_id: int):
    """Fetch a subscription the caller's tenant owns, or 404."""
    from ..models import BillingReportSubscription
    sub = db.get(BillingReportSubscription, sub_id)
    if not sub or sub.tenant_id != t.id:
        raise HTTPException(404, "Offtaker not found")
    return sub


def _seed_template_from_bytes(tpl, data: bytes, name: str, content_type):
    """Store an uploaded template's file + seed editable token-HTML (HTML verbatim;
    Excel → its invoice sheet's HTML). Mirrors the tenant upload path; never fatal."""
    tpl.filename = name[:300]
    tpl.content_type = content_type or "application/octet-stream"
    tpl.file_bytes = data
    lname = name.lower()
    is_excel = (lname.endswith((".xlsx", ".xls", ".xlsm"))
                or data[:4] == _MAGIC_XLSX[:4] or data[:4] == _MAGIC_XLS)
    if lname.endswith((".html", ".htm")):
        try:
            tpl.html = data.decode("utf-8", "replace")
        except Exception:  # noqa: BLE001
            pass
    elif is_excel:
        try:
            from .matcher import excel_to_template_html
            _sheet, html = excel_to_template_html(data)
            if html:
                tpl.html = html
        except Exception:  # noqa: BLE001
            logger.warning("per-offtaker template: Excel HTML seed failed", exc_info=True)


@router.get("/subscriptions/{sub_id}/invoice-template")
def get_sub_invoice_template(sub_id: int, authorization: Optional[str] = Header(default=None)):
    """This offtaker's own invoice-template status + editable token-HTML (seeds the
    default template HTML when none saved, so the editor works out of the box)."""
    t = tenant_from_session(authorization)
    from ..models import OfftakerSubscriptionTemplate
    from .template_render import DEFAULT_TEMPLATE_HTML, AVAILABLE_TOKENS
    with SessionLocal() as db:
        _owned_sub(db, t, sub_id)
        tpl = db.execute(select(OfftakerSubscriptionTemplate).where(
            OfftakerSubscriptionTemplate.subscription_id == sub_id)).scalars().first()
        d = _template_dict(tpl)
        d["html"] = (tpl.html if tpl and tpl.html else DEFAULT_TEMPLATE_HTML)
        d["is_default_html"] = not (tpl and tpl.html)
        d["tokens"] = AVAILABLE_TOKENS
        return {"ok": True, "template": d}


@router.post("/subscriptions/{sub_id}/invoice-template")
async def upload_sub_invoice_template(sub_id: int, file: UploadFile = File(...),
                                      authorization: Optional[str] = Header(default=None)):
    """Upload (or replace) THIS offtaker's own invoice template. Render-from-template
    is gated behind `enabled` (the format toggle), same as the tenant default."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    data, name = await _read_template_upload(file)
    from ..models import OfftakerSubscriptionTemplate
    with SessionLocal() as db:
        _owned_sub(db, t, sub_id)
        tpl = db.execute(select(OfftakerSubscriptionTemplate).where(
            OfftakerSubscriptionTemplate.subscription_id == sub_id)).scalars().first()
        if tpl is None:
            tpl = OfftakerSubscriptionTemplate(subscription_id=sub_id, tenant_id=t.id)
            db.add(tpl)
        _seed_template_from_bytes(tpl, data, name, file.content_type)
        db.commit()
        db.refresh(tpl)
        return {"ok": True, "template": _template_dict(tpl)}


@router.put("/subscriptions/{sub_id}/invoice-template")
def put_sub_invoice_template(sub_id: int, body: InvoiceTemplatePut,
                             authorization: Optional[str] = Header(default=None)):
    """Save THIS offtaker's editable token-HTML + the enable toggle ("Use this
    template" vs "Default format"). A render failure at send falls back to standard."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    from ..models import OfftakerSubscriptionTemplate
    with SessionLocal() as db:
        _owned_sub(db, t, sub_id)
        tpl = db.execute(select(OfftakerSubscriptionTemplate).where(
            OfftakerSubscriptionTemplate.subscription_id == sub_id)).scalars().first()
        if tpl is None:
            tpl = OfftakerSubscriptionTemplate(subscription_id=sub_id, tenant_id=t.id)
            db.add(tpl)
        if body.html is not None:
            tpl.html = body.html
        if body.enabled is not None:
            tpl.enabled = bool(body.enabled)
        db.commit()
        db.refresh(tpl)
        return {"ok": True, "template": _template_dict(tpl)}


@router.delete("/subscriptions/{sub_id}/invoice-template")
def delete_sub_invoice_template(sub_id: int, authorization: Optional[str] = Header(default=None)):
    """Remove THIS offtaker's own template (falls back to the tenant default/standard)."""
    t = tenant_from_session(authorization)
    require_not_demo(t)
    from ..models import OfftakerSubscriptionTemplate
    with SessionLocal() as db:
        _owned_sub(db, t, sub_id)
        tpl = db.execute(select(OfftakerSubscriptionTemplate).where(
            OfftakerSubscriptionTemplate.subscription_id == sub_id)).scalars().first()
        if tpl:
            db.delete(tpl)
            db.commit()
        return {"ok": True}


# ─── file library — every stored file we hold for the operator ───────────────
# A small repository for the Offtaker Invoice Generator: the operator's uploaded
# invoice template, any uploaded billing workbooks, and captured GMP utility-bill
# PDFs — newest first, so the UI features the latest upload. Each entry carries a
# `download` URL the frontend fetches (with the session bearer) and opens as a blob.

_BILLING_BASE = "/v1/array-operator/billing"


@router.get("/files")
def list_files(authorization: Optional[str] = Header(default=None)):
    """List every stored file for this operator, newest-first."""
    t = tenant_from_session(authorization)
    from ..models import OfftakerInvoiceTemplate, Bill, UtilityAccount
    files: list[dict] = []
    with SessionLocal() as db:
        tpl = db.execute(select(OfftakerInvoiceTemplate).where(
            OfftakerInvoiceTemplate.tenant_id == t.id)).scalars().first()
        if tpl and tpl.file_bytes:
            ts = tpl.updated_at or tpl.created_at
            files.append({
                "key": "template", "kind": "template",
                "name": tpl.filename or "invoice template",
                "role": "Invoice template" + (" · in use" if tpl.enabled else ""),
                "content_type": tpl.content_type,
                "size": len(tpl.file_bytes),
                "uploaded_at": ts.isoformat() if ts else None,
                "download": f"{_BILLING_BASE}/invoice-template/file",
            })
        subs = db.execute(select(BillingReportSubscription).where(
            BillingReportSubscription.tenant_id == t.id,
            BillingReportSubscription.deleted_at.is_(None))).scalars().all()
        for s in subs:
            if s.source_workbook:
                files.append({
                    "key": f"workbook-{s.id}", "kind": "workbook",
                    "name": s.source_filename or "billing workbook.xlsx",
                    "role": f"Billing workbook · {s.customer_name}",
                    "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "size": len(s.source_workbook),
                    "uploaded_at": s.updated_at.isoformat() if s.updated_at else None,
                    "download": f"{_BILLING_BASE}/files/workbook/{s.id}",
                })
        bills = db.execute(
            select(Bill, UtilityAccount.nickname)
            .join(UtilityAccount, Bill.account_id == UtilityAccount.id)
            .where(Bill.tenant_id == t.id, Bill.pdf_bytes.isnot(None))
            .order_by(Bill.bill_date.desc().nullslast()).limit(24)
        ).all()
        for b, nick in bills:
            per = b.bill_date.strftime("%Y-%m") if b.bill_date else "bill"
            label = nick or f"GMP {b.account_id}"
            ts = b.pulled_at or b.bill_date
            files.append({
                "key": f"gmp-{b.id}", "kind": "gmp_bill",
                "name": f"GMP bill — {label} {per}.pdf",
                "role": f"GMP utility bill · {label}",
                "content_type": b.pdf_content_type or "application/pdf",
                "size": len(b.pdf_bytes) if b.pdf_bytes else 0,
                "uploaded_at": ts.isoformat() if ts else None,
                "download": f"{_BILLING_BASE}/files/gmp-bill/{b.id}",
            })
    files.sort(key=lambda f: f.get("uploaded_at") or "", reverse=True)
    return {"files": files, "count": len(files)}


@router.get("/files/workbook/{sub_id}")
def get_workbook_file(sub_id: int, authorization: Optional[str] = Header(default=None)):
    """Stream an uploaded billing workbook (inline view / download)."""
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        s = _get_owned(db, t.id, sub_id)
        if not s.source_workbook:
            raise HTTPException(404, "No workbook on file")
        fn = "".join(c for c in (s.source_filename or "workbook.xlsx")
                     if c.isalnum() or c in "._- ") or "workbook.xlsx"
        return StreamingResponse(io.BytesIO(s.source_workbook),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "inline; filename=" + fn})


@router.get("/files/gmp-bill/{bill_id}")
def get_gmp_bill_file(bill_id: int, authorization: Optional[str] = Header(default=None)):
    """Stream a captured GMP utility-bill PDF (inline view / download)."""
    t = tenant_from_session(authorization)
    from ..models import Bill
    with SessionLocal() as db:
        b = db.get(Bill, bill_id)
        if not b or b.tenant_id != t.id or not b.pdf_bytes:
            raise HTTPException(404, "No GMP bill PDF on file")
        return StreamingResponse(io.BytesIO(b.pdf_bytes),
            media_type=b.pdf_content_type or "application/pdf",
            headers={"Content-Disposition": "inline; filename=gmp_bill.pdf"})


# ─── Offtaker invoice email — MASS TEMPLATE studio (Anna-scale ask) ──────────
# The letter at the top of every offtaker invoice email is a tenant-wide
# merge-tag template (same engine + studio UX as the NEPOOL report-email
# customizer; offtaker tag set). A per-draft note still overrides the letter
# for that one send. Endpoints mirror api/account.py's suite so the studio
# frontend contract is identical.

class _OfftakerTemplateBody(BaseModel):
    subject_template: Optional[str] = None
    body_template: Optional[str] = None
    signoff: Optional[str] = None


class _OfftakerChatBody(BaseModel):
    messages: list[dict]
    current_body: str
    current_subject: Optional[str] = None


def _sample_offtaker_ctx(db, t) -> tuple[dict, Optional[str], Optional[str]]:
    """(merge_ctx, sample_name, sample_email) for the studio preview, using the
    tenant's FIRST enabled offtaker with an email + their REAL current invoice
    figures (build_match), so the preview shows true numbers like the NEPOOL
    studio previews with a real client. Falls back to canned sample values."""
    from ..email_templates import build_offtaker_context
    from .delivery import _attachments_line

    sub = db.execute(
        select(BillingReportSubscription).where(
            BillingReportSubscription.tenant_id == t.id,
            BillingReportSubscription.enabled == True,  # noqa: E712
            BillingReportSubscription.deleted_at.is_(None),
            BillingReportSubscription.client_email.is_not(None),
        ).order_by(BillingReportSubscription.customer_name.asc()).limit(1)
    ).scalars().first()

    name, email = "Sample Offtaker", None
    period, ps, pe = "the latest period", "", ""
    kwh_s, amount_s, inv_no = "1,265 kWh", "$190.20", "2026-06"
    has_summary = False
    if sub is not None:
        name, email = sub.customer_name, sub.client_email
        has_summary = sub.include_summary is True
        try:
            ci = (build_match(sub).computed_invoice or {})
            if ci.get("period_start") and ci.get("period_end"):
                ps, pe = str(ci["period_start"]), str(ci["period_end"])
                period = f"{ps} → {pe}"
            if ci.get("kwh") is not None:
                kwh_s = f"{ci['kwh']:,.0f} kWh"
            if isinstance(ci.get("amount_owed"), (int, float)):
                amount_s = f"${ci['amount_owed']:,.2f}"
            if ci.get("invoice_number"):
                inv_no = str(ci["invoice_number"])
        except Exception:  # noqa: BLE001 — canned values are fine for a preview
            logger.exception("sample offtaker preview build failed")

    tenant_name = t.company_name or t.operator_name or t.name or "Your Company"
    ctx = build_offtaker_context(
        offtaker_name=name,
        tenant_name=tenant_name,
        tenant_email=(t.contact_email or ""),
        period=period, period_start=ps, period_end=pe,
        kwh=kwh_s, amount=amount_s, invoice_number=inv_no,
        attachments_line=_attachments_line(True, has_summary),
        signoff_template=t.email_signoff,
        tenant_signoff_name=(t.send_from_name or t.operator_name),
    )
    return ctx, (sub.customer_name if sub else None), email


@router.get("/email-template")
def get_offtaker_email_template(authorization: Optional[str] = Header(default=None)):
    """The tenant's offtaker invoice email template, with resolved defaults."""
    from ..email_templates import (
        DEFAULT_OFFTAKER_SUBJECT_TEMPLATE, DEFAULT_OFFTAKER_BODY_TEMPLATE,
        DEFAULT_SIGNOFF, OFFTAKER_ALLOWED_MERGE_TAGS,
    )
    from ..models import Tenant
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        t = db.get(Tenant, t.id)
        sample = db.execute(
            select(BillingReportSubscription).where(
                BillingReportSubscription.tenant_id == t.id,
                BillingReportSubscription.enabled == True,  # noqa: E712
                BillingReportSubscription.deleted_at.is_(None),
                BillingReportSubscription.client_email.is_not(None),
            ).order_by(BillingReportSubscription.customer_name.asc()).limit(1)
        ).scalars().first()
        return {
            "subject_template": t.offtaker_email_subject_template
                                or DEFAULT_OFFTAKER_SUBJECT_TEMPLATE,
            "body_template": t.offtaker_email_body_template
                             or DEFAULT_OFFTAKER_BODY_TEMPLATE,
            "signoff": t.email_signoff or DEFAULT_SIGNOFF,
            "is_default_subject": t.offtaker_email_subject_template is None,
            "is_default_body": t.offtaker_email_body_template is None,
            "is_default_signoff": t.email_signoff is None,
            "from_email": t.send_from_email or t.contact_email,
            "available_tokens": sorted(OFFTAKER_ALLOWED_MERGE_TAGS),
            "has_client_with_email": sample is not None,
            "sample_client_email": sample.client_email if sample else None,
        }


@router.post("/email-template/preview")
def preview_offtaker_email_template(body: _OfftakerTemplateBody,
                                    authorization: Optional[str] = Header(default=None)):
    """Render the proposed template with a REAL sample offtaker's figures."""
    from ..email_templates import (
        DEFAULT_OFFTAKER_SUBJECT_TEMPLATE, DEFAULT_OFFTAKER_BODY_TEMPLATE,
        DEFAULT_SIGNOFF, render_merge,
    )
    from ..models import Tenant
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        t = db.get(Tenant, t.id)
        stored_subject = t.offtaker_email_subject_template
        stored_body = t.offtaker_email_body_template
        # Request body overrides stored; stored overrides default — same
        # resolution as the send path, so the preview IS the send.
        signoff_t = (body.signoff or "").strip() or t.email_signoff or DEFAULT_SIGNOFF
        t.email_signoff = signoff_t     # in-memory only, feeds ctx below
        ctx, sample_name, _ = _sample_offtaker_ctx(db, t)
        db.rollback()                   # never persist the preview signoff
    subj_t = (body.subject_template or "").strip() or stored_subject \
        or DEFAULT_OFFTAKER_SUBJECT_TEMPLATE
    body_t = (body.body_template or "").strip() or stored_body \
        or DEFAULT_OFFTAKER_BODY_TEMPLATE
    return {
        "subject_rendered": render_merge(subj_t, ctx),
        "body_rendered": render_merge(body_t, ctx),
        "sample_client": sample_name or "Sample Offtaker",
    }


@router.post("/email-template/chat")
def chat_offtaker_email_template(body: _OfftakerChatBody,
                                 authorization: Optional[str] = Header(default=None)):
    """AI assistant: regenerate the offtaker template from the conversation."""
    import os as _os
    from ..email_templates import (
        regenerate_template_via_ai, _OFFTAKER_TEMPLATE_SYSTEM_PROMPT,
        OFFTAKER_ALLOWED_MERGE_TAGS, DEFAULT_OFFTAKER_SUBJECT_TEMPLATE,
    )
    tenant_from_session(authorization)
    api_key = _os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(503, "AI assistant not configured — set ANTHROPIC_API_KEY")
    recent = body.messages[-10:]
    for m in recent:
        if m.get("role") not in ("user", "assistant") or not isinstance(m.get("content"), str):
            raise HTTPException(400, "Each message must have role user|assistant and string content")
    current_subject = (body.current_subject or "").strip() or DEFAULT_OFFTAKER_SUBJECT_TEMPLATE
    try:
        result = regenerate_template_via_ai(
            current_body=body.current_body,
            current_subject=current_subject,
            messages=recent,
            api_key=api_key,
            system_prompt=_OFFTAKER_TEMPLATE_SYSTEM_PROMPT,
            allowed_tags=OFFTAKER_ALLOWED_MERGE_TAGS,
        )
    except Exception as exc:  # noqa: BLE001
        logger.exception("Offtaker template AI regen failed")
        raise HTTPException(502, f"AI request failed: {exc}") from exc
    return {
        "assistant_reply": result["reply"],
        "proposed_body": result["body"],
        "proposed_subject": result["subject"],
    }


@router.put("/email-template")
def save_offtaker_email_template(body: _OfftakerTemplateBody,
                                 authorization: Optional[str] = Header(default=None)):
    """Persist the tenant's offtaker email template (null/blank → default)."""
    from ..email_templates import unknown_tags, OFFTAKER_ALLOWED_MERGE_TAGS
    from ..models import Tenant
    t = tenant_from_session(authorization)
    require_not_demo(t)
    for field, value in [("subject_template", body.subject_template),
                         ("body_template", body.body_template)]:
        if value:
            bad = unknown_tags(value, OFFTAKER_ALLOWED_MERGE_TAGS)
            if bad:
                listed = ", ".join("{{" + tag + "}}" for tag in sorted(bad))
                raise HTTPException(422, f"Unknown merge tags in {field}: {listed}")
    with SessionLocal() as db:
        t = db.get(Tenant, t.id)
        if body.subject_template is not None:
            t.offtaker_email_subject_template = (body.subject_template or "").strip() or None
        if body.body_template is not None:
            t.offtaker_email_body_template = (body.body_template or "").strip() or None
        db.commit()
        return {"ok": True,
                "subject_template": t.offtaker_email_subject_template,
                "body_template": t.offtaker_email_body_template}


@router.put("/email-template/signoff")
def save_offtaker_email_signoff(body: _OfftakerTemplateBody,
                                authorization: Optional[str] = Header(default=None)):
    """Persist the SHARED operator sign-off (same block the NEPOOL report email
    uses — one operator identity across products). null/blank → default."""
    from ..models import Tenant
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        t = db.get(Tenant, t.id)
        t.email_signoff = (body.signoff or "").strip() or None
        db.commit()
        return {"ok": True, "signoff": t.email_signoff}


@router.post("/email-template/test-send")
def test_send_offtaker_email_template(body: _OfftakerTemplateBody,
                                      authorization: Optional[str] = Header(default=None)):
    """Send a [TEST] of the proposed template (real sample figures) to the operator."""
    from ..email_templates import (
        DEFAULT_OFFTAKER_SUBJECT_TEMPLATE, DEFAULT_OFFTAKER_BODY_TEMPLATE,
        DEFAULT_SIGNOFF, render_merge, html_to_text,
    )
    from ..models import Tenant
    from ..notify import _send_via_resend
    from ..email_skin import render_email_skin, render_email_skin_text
    t = tenant_from_session(authorization)
    require_not_demo(t)
    to_email = (t.contact_email or "").strip()
    if not to_email:
        raise HTTPException(422, "Add an email address to your account first.")
    with SessionLocal() as db:
        t = db.get(Tenant, t.id)
        stored_subject = t.offtaker_email_subject_template
        stored_body = t.offtaker_email_body_template
        signoff_t = (body.signoff or "").strip() or t.email_signoff or DEFAULT_SIGNOFF
        t.email_signoff = signoff_t
        ctx, sample_name, _ = _sample_offtaker_ctx(db, t)
        operator = t.company_name or t.operator_name or t.name or "your solar provider"
        db.rollback()
    subj_t = (body.subject_template or "").strip() or stored_subject \
        or DEFAULT_OFFTAKER_SUBJECT_TEMPLATE
    body_t = (body.body_template or "").strip() or stored_body \
        or DEFAULT_OFFTAKER_BODY_TEMPLATE
    subject = render_merge(subj_t, ctx)
    letter = render_merge(body_t, ctx)
    html = render_email_skin(
        preheader="Test of your offtaker invoice email template.",
        headline="Your solar credit invoice",
        intro_line=(sample_name or "sample offtaker"),
        body_html=letter,
        footer_line=f"Solar credit invoice from {operator}.  ·  Questions? just reply to this email.",
        wordmark=operator, product="array_operator")
    sent = _send_via_resend(
        to=to_email, subject=f"[TEST] {subject}", html=html,
        text=render_email_skin_text(
            headline="Your solar credit invoice",
            intro_line=(sample_name or "sample offtaker"),
            body_text=html_to_text(letter), wordmark=operator,
            product="array_operator"),
        product="array_operator")
    if not sent:
        raise HTTPException(502, "Email delivery failed — check your Resend configuration.")
    return {"ok": True, "sent_to": to_email}


@router.post("/email-template/reset")
def reset_offtaker_email_template(authorization: Optional[str] = Header(default=None)):
    """Revert the offtaker subject/body to the built-in defaults. The shared
    sign-off is left alone (the NEPOOL report email uses it too)."""
    from ..models import Tenant
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        t = db.get(Tenant, t.id)
        t.offtaker_email_subject_template = None
        t.offtaker_email_body_template = None
        db.commit()
    return {"ok": True}


# ─── Send pipeline (Ford 2026-07-03: "what's in the pipeline, what's gonna
# fire") — the flow dashboard over the offtaker list. Everything here is a
# cheap aggregate over columns the send path already stamps; nothing rebuilds
# an invoice at read time (~60s at 800 offtakers). ──────────────────────────

def _next_month_first(now: datetime) -> datetime:
    y, m = now.year, now.month
    return datetime(y + (1 if m == 12 else 0), 1 if m == 12 else m + 1, 1, 9, 0)


def _next_quarter_first(now: datetime) -> datetime:
    for m in (1, 4, 7, 10):
        cand = datetime(now.year, m, 1, 9, 0)
        if cand > now:
            return cand
    return datetime(now.year + 1, 1, 1, 9, 0)


@router.get("/send-pipeline")
def send_pipeline(authorization: Optional[str] = Header(default=None)):
    """The month-organized send-pipeline roll-up:
      last      — the most recent DELIVERED period: count + $ + when it ran;
      inflight  — pending drafts awaiting approval + subs still waiting (no
                  send for the period and no draft yet — typically holds
                  waiting on their utility bill);
      next_monthly / next_quarterly — when the scheduler fires next and how
                  many subs it touches, split auto-send vs draft-for-approval;
      paused    — the tenant's pause switch."""
    from ..models import Tenant
    t = tenant_from_session(authorization)
    with SessionLocal() as db:
        tenant = db.get(Tenant, t.id)
        rows = db.execute(
            select(BillingReportSubscription.cadence,
                   BillingReportSubscription.delivery_mode,
                   BillingReportSubscription.last_sent_period_end,
                   BillingReportSubscription.last_sent_amount_usd,
                   BillingReportSubscription.last_sent_at)
            .where(BillingReportSubscription.tenant_id == t.id,
                   BillingReportSubscription.deleted_at.is_(None),
                   BillingReportSubscription.enabled == True)  # noqa: E712
        ).all()
        # Split the pending drafts by delivery_mode — an auto-send offtaker's draft
        # is NOT "awaiting approval" (it sends itself on the run); only an approval
        # offtaker's draft needs the operator. Counting them together mislabeled the
        # whole set "awaiting your approval" (Ford 2026-07-07: "why does it say 805
        # to approve").
        pend_rows = db.execute(
            select(BillingReportSubscription.delivery_mode,
                   func.count(ReportDraft.id))
            .join(BillingReportSubscription,
                  ReportDraft.subscription_id == BillingReportSubscription.id)
            .where(ReportDraft.tenant_id == t.id,
                   ReportDraft.status == "pending",
                   BillingReportSubscription.deleted_at.is_(None),
                   BillingReportSubscription.enabled == True)  # noqa: E712
            .group_by(BillingReportSubscription.delivery_mode)
        ).all()
        pending_auto = sum(int(c) for m, c in pend_rows if (m or "approval") == "auto")
        pending_approval = sum(int(c) for m, c in pend_rows if (m or "approval") != "auto")
        pending = pending_auto + pending_approval

    total = len(rows)
    last_period = max((r.last_sent_period_end for r in rows
                       if r.last_sent_period_end), default=None)
    delivered = 0
    dollars = 0.0
    last_run_at = None
    for r in rows:
        if last_period and r.last_sent_period_end == last_period:
            delivered += 1
            if r.last_sent_amount_usd:
                dollars += float(r.last_sent_amount_usd)
        if r.last_sent_at and (last_run_at is None or r.last_sent_at > last_run_at):
            last_run_at = r.last_sent_at
    waiting = max(0, total - delivered - int(pending))

    # The tenant's dominant posture — drives the "Approve to send ⟷ Auto-send" mode
    # slider (Ford 2026-07-07). Per-offtaker overrides still count in the split.
    auto_all = sum(1 for r in rows if (r.delivery_mode or "approval") == "auto")
    approval_all = total - auto_all
    default_mode = "auto" if auto_all > approval_all else "approval"

    now = datetime.utcnow()

    def _split(cadence: str) -> dict:
        subs = [r for r in rows if (r.cadence or "monthly") == cadence]
        auto = sum(1 for r in subs if (r.delivery_mode or "approval") == "auto")
        return {"scheduled": len(subs), "auto": auto,
                "approval": len(subs) - auto}

    return {
        "ok": True,
        "total_enabled": total,
        "last": {
            "period_end": last_period,
            "period_month": (last_period or "")[:7] or None,
            "delivered": delivered,
            "dollars": round(dollars, 2),
            "last_run_at": last_run_at.isoformat() if last_run_at else None,
        },
        "inflight": {"pending_drafts": int(pending),
                     "pending_auto": int(pending_auto),
                     "pending_approval": int(pending_approval),
                     "waiting": waiting},
        "default_delivery_mode": default_mode,
        "mode_split": {"auto": auto_all, "approval": approval_all},
        "next_monthly": {"fires_at": _next_month_first(now).isoformat(),
                         **_split("monthly")},
        "next_quarterly": {"fires_at": _next_quarter_first(now).isoformat(),
                           **_split("quarterly")},
        "paused": bool(getattr(tenant, "sending_paused", False)),
    }


class _PauseBody(BaseModel):
    paused: bool


@router.patch("/sending-paused")
def set_sending_paused(body: _PauseBody,
                       authorization: Optional[str] = Header(default=None)):
    """The pipeline pause switch: True halts the SCHEDULER's billing runs for
    this tenant (no auto sends, no auto drafts). Manual sends and draft
    approvals still work — pause stops the machine, not the operator."""
    from ..models import Tenant
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        t = db.get(Tenant, t.id)
        t.sending_paused = bool(body.paused)
        db.commit()
        return {"ok": True, "paused": t.sending_paused}


class _BulkModeBody(BaseModel):
    mode: str


@router.post("/subscriptions/bulk-delivery-mode")
def bulk_delivery_mode(body: _BulkModeBody,
                       authorization: Optional[str] = Header(default=None)):
    """Flip EVERY enabled offtaker's delivery mode at once (the pipeline's
    'Auto-send all' / 'Draft all' controls). Per-offtaker edits afterwards
    still override — this is a bulk starting point, not a lock."""
    mode = (body.mode or "").strip()
    if mode not in VALID_DELIVERY:
        raise HTTPException(422, f"mode must be one of {sorted(VALID_DELIVERY)}")
    t = tenant_from_session(authorization)
    require_not_demo(t)
    with SessionLocal() as db:
        subs = db.execute(
            select(BillingReportSubscription)
            .where(BillingReportSubscription.tenant_id == t.id,
                   BillingReportSubscription.deleted_at.is_(None),
                   BillingReportSubscription.enabled == True)  # noqa: E712
        ).scalars().all()
        changed = 0
        for s in subs:
            if (s.delivery_mode or "approval") != mode:
                s.delivery_mode = mode
                changed += 1
        db.commit()
        auto = sum(1 for s in subs if (s.delivery_mode or "approval") == "auto")
        return {"ok": True, "mode": mode, "changed": changed,
                "auto": auto, "approval": len(subs) - auto}


# ─── Draft all (Ford 2026-07-04: "it should automatically draft the latest for
# each offtaker") — the pipeline's "Draft all" doesn't just flip a setting, it
# GENERATES each enabled offtaker's latest-period invoice into the review inbox.
# Runs in a background thread (a draft per offtaker rebuilds the invoice, ~60s
# at 800) with live progress; draft_subscription is idempotent per period and
# honestly HOLDS an offtaker whose bill hasn't settled (never a fabricated
# draft). ────────────────────────────────────────────────────────────────────
_bulk_draft_lock = _threading.Lock()
_bulk_draft: dict[str, dict] = {}   # tenant_id -> {total, done, drafted, held, running}


@router.post("/subscriptions/bulk-draft")
def bulk_draft(keep_mode: bool = Query(default=False),
               authorization: Optional[str] = Header(default=None)):
    """Draft the latest invoice for every enabled offtaker into the review inbox.

    `keep_mode` (the CONTINUOUS auto-draft path, Ford 2026-07-07: "all offtaker
    invoices automatically draft ... and continuously poll") leaves each offtaker's
    delivery_mode ALONE — so a poll that runs every few minutes never silently flips
    an auto-send offtaker back to approval. The default (False) keeps the legacy
    "Draft all for review" semantics (everyone → approval)."""
    from .delivery import draft_subscription
    from ..models import Tenant
    t = tenant_from_session(authorization)
    require_not_demo(t)
    tid = t.id
    with _bulk_draft_lock:
        cur = _bulk_draft.get(tid)
        if cur and cur.get("running"):
            return {"ok": True, "already_running": True, **cur}
    # Collect the ids to draft (one cheap query). Only the legacy manual path also
    # forces review mode; the auto path preserves each offtaker's send mode.
    with SessionLocal() as db:
        subs = db.execute(
            select(BillingReportSubscription)
            .where(BillingReportSubscription.tenant_id == tid,
                   BillingReportSubscription.deleted_at.is_(None),
                   BillingReportSubscription.enabled == True)  # noqa: E712
        ).scalars().all()
        ids = []
        for s in subs:
            if not keep_mode and (s.delivery_mode or "approval") != "approval":
                s.delivery_mode = "approval"
            ids.append(s.id)
        db.commit()
    with _bulk_draft_lock:
        _bulk_draft[tid] = {"total": len(ids), "done": 0, "drafted": 0,
                            "held": 0, "running": True}

    def _run():
        drafted = held = done = 0
        try:
            for sid in ids:
                try:
                    with SessionLocal() as db:
                        sub = db.get(BillingReportSubscription, sid)
                        tenant = db.get(Tenant, tid)
                        if sub and tenant:
                            r = draft_subscription(db, sub, tenant,
                                                   triggered_by="bulk-draft")
                            if r.get("ok"):
                                drafted += 1
                            elif r.get("skipped"):
                                held += 1
                except Exception:  # noqa: BLE001
                    logger.exception("bulk-draft sub %s failed", sid)
                done += 1
                with _bulk_draft_lock:
                    p = _bulk_draft.get(tid)
                    if p:
                        p["done"], p["drafted"], p["held"] = done, drafted, held
        finally:
            with _bulk_draft_lock:
                p = _bulk_draft.get(tid)
                if p:
                    p["running"] = False

    _threading.Thread(target=_run, daemon=True, name="bulk-draft").start()
    return {"ok": True, "started": True, "total": len(ids)}


@router.get("/subscriptions/bulk-draft-status")
def bulk_draft_status(authorization: Optional[str] = Header(default=None)):
    t = tenant_from_session(authorization)
    with _bulk_draft_lock:
        p = _bulk_draft.get(t.id)
        return {"ok": True, **(p or {"total": 0, "done": 0, "drafted": 0,
                                     "held": 0, "running": False})}
