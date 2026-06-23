"""
Invoice writer — reproduce the customer's OWN uploaded spreadsheet, populated.

THE THREE SPREADSHEET SYSTEMS (don't confuse them — see routes.py for the map):
  1. api/writers/gmcs_writer.py  — WRITES the NEPOOL-GIS GMCS *filing* workbook
                                   (the NEPOOL Operator product output). NOT a
                                   customer invoice.
  2. api/billing/matcher.py      — READS an uploaded customer *billing* workbook
                                   (the HCT family) into a BillingMatch.
  3. api/billing/invoice_writer.py (THIS FILE) — WRITES the customer's invoice
                                   back into THEIR OWN workbook format.

What this does, and why it's done this way
------------------------------------------
When an Array Operator onboards a customer by uploading that customer's existing
billing spreadsheet, we keep the ORIGINAL file bytes on the subscription
(`sub.source_workbook`) plus the parsed structure (`sub.parsed_map`, a
`BillingMatch.to_dict()`). To produce an invoice that looks EXACTLY like the one
the customer already recognizes, we do NOT regenerate a workbook from scratch
(that would throw away the customer's styling, formulas, merged cells, headers,
and their bespoke "Template" invoice sheet). Instead we:

  * `openpyxl.load_workbook(BytesIO(sub.source_workbook))` — preserving every
    sheet, style, number-format, merged cell, and formula untouched;
  * append (or, if it's already the latest row, refresh in place) ONE month row
    in the data-ledger sheet, at the exact columns the matcher recognized,
    copying the previous data row's full cell style + number-format so the new
    row is visually identical;
  * for the customer-kWh share we materialize a CONCRETE value (= whole-array
    kWh × allocation %) so the figure is correct and auditable even without a
    recalculation engine; the sheet's own downstream formulas (Tariff+Adder,
    Value, Bill, Savings, running totals) are translated forward from the prior
    row so the customer's Excel keeps computing them exactly as before;
  * bump the "New Row #" pointer cell that the Template invoice sheet reads
    through `INDIRECT(ADDRESS(pointer, col, …, "<ledger sheet>"))`, so the whole
    invoice refreshes to the new period with zero changes to the Template sheet.

Handles all three billing models the matcher detects (percent_of_array,
fixed_budget, flat_rate): the per-row fixed/flat amount column is simply carried
forward (it repeats monthly), while the metered columns are recomputed.

NEVER fabricates generation: the period data comes from the real parsed workbook
(`build_match`/`compute_invoice`) or an explicitly supplied period. If there is
no generation for the period, `populate_invoice_workbook` raises
`InvoiceWriterError` rather than inventing a row.

Manual (typed-in) customers have NO `source_workbook` — there is no "own format"
to reproduce — so for them we fall back to the standard generated invoice
(`api/billing/invoice.render_invoice_xlsx`).
"""
from __future__ import annotations

import io
import logging
from copy import copy
from datetime import date, datetime
from typing import Any, Optional, Union

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from .matcher import BillingMatch, Period

logger = logging.getLogger(__name__)


class InvoiceWriterError(Exception):
    """Raised when the stored workbook can't be populated for the period."""


# ─── public entry point ─────────────────────────────────────────────────────

def populate_invoice_workbook(sub, period_data: Optional[Union[Period, dict]] = None,
                              *, field_map_override: Optional[dict] = None) -> bytes:
    """Return .xlsx bytes of the customer's OWN workbook, populated for a period.

    `sub` is a BillingReportSubscription. `period_data` selects the month to
    write; it may be:
      * None       — use the latest period parsed from the stored workbook;
      * a Period   — explicit period to write;
      * a dict     — {month, period_start|start, period_end|end, array_kwh,
                      customer_kwh (optional), tariff, adder}.

    Workbook customers get their original file loaded + populated (preserving
    all styling/formulas/Template). Manual customers (no `source_workbook`) fall
    back to the standard generated invoice .xlsx.

    Raises InvoiceWriterError when there is no real generation to write (we never
    invent a row).
    """
    if not getattr(sub, "source_workbook", None):
        return _fallback_standard_invoice(sub, period_data)

    # Rebuild the match for the field map, allocation %, billing model, and the
    # real parsed periods. (build_match re-parses the stored bytes.)
    from .delivery import build_match
    match = build_match(sub)

    parsed = sub.parsed_map or {}
    data_sheet = parsed.get("data_sheet") or match.data_sheet
    # field_map_override lets the repro refine loop re-fill with an AI-corrected
    # column map when the verify guard found the numbers landed wrong.
    field_map: dict[str, int] = (field_map_override
                                 or parsed.get("field_map") or match.field_map or {})
    if not data_sheet or "month" not in field_map:
        raise InvoiceWriterError(
            "stored workbook structure is incomplete (no ledger sheet / column map)"
        )

    period = _resolve_period(match, period_data)
    if period is None:
        raise InvoiceWriterError("no billing period to write")
    if period.array_kwh is None and period.customer_kwh is None:
        raise InvoiceWriterError(
            "no generation data for this period — refusing to invent an invoice row"
        )

    # Load the ORIGINAL workbook with formulas intact (NOT data_only).
    try:
        wb = load_workbook(io.BytesIO(bytes(sub.source_workbook)))
    except Exception as e:  # noqa: BLE001
        raise InvoiceWriterError(f"could not open stored workbook: {e}") from e
    if data_sheet not in wb.sheetnames:
        raise InvoiceWriterError(f"ledger sheet '{data_sheet}' missing from workbook")
    ws = wb[data_sheet]

    pct = match.allocation_pct
    new_row = _write_ledger_row(ws, field_map, period, pct)

    # Refresh every Template invoice sheet's "New Row #" pointer to the row we
    # just wrote, so the bound INDIRECT() cells reflect this period.
    _refresh_template_pointers(wb, new_row)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ─── period resolution ──────────────────────────────────────────────────────

def _resolve_period(match: BillingMatch, period_data) -> Optional[Period]:
    if period_data is None:
        return match.latest_period
    if isinstance(period_data, Period):
        return period_data
    if isinstance(period_data, dict):
        d = period_data
        return Period(
            month=d.get("month"),
            start=_as_date(d.get("period_start") or d.get("start")),
            end=_as_date(d.get("period_end") or d.get("end")),
            array_kwh=_as_float(d.get("array_kwh")),
            customer_kwh=_as_float(d.get("customer_kwh")),
            tariff=_as_float(d.get("tariff")),
            adder=_as_float(d.get("adder")),
        )
    raise InvoiceWriterError(f"unsupported period_data type: {type(period_data)!r}")


def _as_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _as_float(v: Any) -> Optional[float]:
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("$", "").strip())
    except ValueError:
        return None


# ─── ledger geometry ────────────────────────────────────────────────────────

def _find_header_row(ws, month_col0: int) -> int:
    """1-indexed row of the ledger header (the row whose month column says
    'Month'). Falls back to 1 if not found in the first 15 rows."""
    for r in range(1, 16):
        v = ws.cell(row=r, column=month_col0 + 1).value
        if v is not None and str(v).strip().lower().startswith("month"):
            return r
    return 1


def _find_last_data_row(ws, header_row: int, month_col0: int,
                        start_col0: Optional[int]) -> int:
    """1-indexed row of the last populated data row. A data row has a month
    label or a start date. Stops after a run of blank rows so trailing
    helper/footnote cells far below don't get mistaken for data."""
    last = header_row
    blanks = 0
    for r in range(header_row + 1, ws.max_row + 1):
        month = ws.cell(row=r, column=month_col0 + 1).value
        start = ws.cell(row=r, column=start_col0 + 1).value if start_col0 is not None else None
        if (month not in (None, "")) or (start not in (None, "")):
            last = r
            blanks = 0
        else:
            blanks += 1
            if blanks >= 6:
                break
    return last


def _row_extent(ws, row: int) -> int:
    """Rightmost column (1-indexed) carrying a value in `row` — the width to
    clone styling/formulas across."""
    extent = 1
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=c).value not in (None, ""):
            extent = c
    return extent


def _copy_style(src, dst) -> None:
    """Clone visual style + number-format from one cell to another."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format


def _write_ledger_row(ws, field_map: dict[str, int], period: Period,
                      pct: Optional[float]) -> int:
    """Append (or refresh-in-place) the period's row in the ledger sheet.

    Returns the 1-indexed Excel row that was written. If the period matches the
    existing last data row (same end date / month label) we overwrite it in
    place so re-previewing the current period never duplicates a row; otherwise
    we append a brand-new row below it.
    """
    month_col0 = field_map["month"]
    start_col0 = field_map.get("start")
    end_col0 = field_map.get("end")
    array_col0 = field_map.get("array_kwh")
    cust_col0 = field_map.get("customer_kwh")
    tariff_col0 = field_map.get("tariff")
    adder_col0 = field_map.get("adder")

    header_row = _find_header_row(ws, month_col0)
    last_row = _find_last_data_row(ws, header_row, month_col0, start_col0)
    if last_row <= header_row:
        raise InvoiceWriterError("ledger has no existing data row to model the new row on")

    # Decide append vs. in-place refresh of the latest period.
    target_row = last_row + 1
    if _same_period(ws, last_row, end_col0, month_col0, period):
        target_row = last_row

    prev_row = last_row
    extent = _row_extent(ws, prev_row)

    # The customer's share, materialized as a concrete number (= whole × pct)
    # so the figure is correct/auditable without a recalc engine. Prefer an
    # explicitly supplied customer_kwh (e.g. the real parsed value on refresh).
    customer_kwh = period.customer_kwh
    if customer_kwh is None and period.array_kwh is not None and pct:
        customer_kwh = round(period.array_kwh * pct)

    # Columns we set as literal INPUTS (everything else is cloned/translated).
    literal_inputs: dict[int, Any] = {}
    if month_col0 is not None:
        literal_inputs[month_col0] = period.month
    if start_col0 is not None:
        literal_inputs[start_col0] = period.start
    if end_col0 is not None:
        literal_inputs[end_col0] = period.end
    if array_col0 is not None:
        literal_inputs[array_col0] = period.array_kwh
    if tariff_col0 is not None and period.tariff is not None:
        literal_inputs[tariff_col0] = period.tariff
    if adder_col0 is not None and period.adder is not None:
        literal_inputs[adder_col0] = period.adder

    for c in range(1, extent + 1):
        col0 = c - 1
        src = ws.cell(row=prev_row, column=c)
        dst = ws.cell(row=target_row, column=c)
        _copy_style(src, dst)

        if col0 == cust_col0:
            # Materialize the customer share as a real value.
            dst.value = customer_kwh
        elif col0 in literal_inputs:
            val = literal_inputs[col0]
            # Keep the prior literal (e.g. tariff/adder) if no fresh value given.
            dst.value = val if val is not None else _plain_value(src)
        elif isinstance(src.value, str) and src.value.startswith("="):
            # Translate the customer's own formula forward, preserving their
            # exact metered math (Tariff+Adder, Value, Bill, Savings, totals).
            origin = f"{get_column_letter(c)}{prev_row}"
            target = f"{get_column_letter(c)}{target_row}"
            dst.value = Translator(src.value, origin=origin).translate_formula(target)
        else:
            # Non-formula auxiliary cell (e.g. a fixed/flat amount that repeats
            # each month, or a text helper): carry it forward verbatim.
            dst.value = _plain_value(src)

    return target_row


def _plain_value(cell) -> Any:
    """The cell's literal value (ArrayFormula objects collapse to None — the
    ledger data rows never use them; only the Template sheet does)."""
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        return None
    return v


def _same_period(ws, row: int, end_col0: Optional[int], month_col0: int,
                 period: Period) -> bool:
    """Is `row` already the period we're about to write? Matches on end date
    when both have one, else on the month label."""
    if end_col0 is not None and period.end is not None:
        existing = _as_date(ws.cell(row=row, column=end_col0 + 1).value)
        if existing is not None:
            return existing == period.end
    if period.month:
        existing_m = ws.cell(row=row, column=month_col0 + 1).value
        return str(existing_m).strip().lower() == str(period.month).strip().lower()
    return False


# ─── Template invoice-sheet binding ─────────────────────────────────────────

def _refresh_template_pointers(wb, new_row: int) -> None:
    """Point every Template sheet's 'New Row #' cell at the row we just wrote.

    The HCT Template invoice sheets read the live period through
    `INDIRECT(ADDRESS(<pointer>, <col>, …, "<ledger>"))`, where <pointer> is the
    cell to the right of a 'New Row #'/'NEW ROW #' label. Bumping that one cell
    re-binds the whole invoice to the new period without touching any other
    Template cell.
    """
    for ws in wb.worksheets:
        if "template" not in (ws.title or "").lower():
            continue
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and ("new row" in v.lower() or "row #" in v.lower()):
                    ws.cell(row=cell.row, column=cell.column + 1).value = new_row


# ─── manual-customer fallback ───────────────────────────────────────────────

def _fallback_standard_invoice(sub, period_data) -> bytes:
    """Manual customers have no 'own format' — render the standard invoice .xlsx
    from the synthesized BillingMatch and return its bytes."""
    import pathlib
    import tempfile
    from .delivery import build_match
    from . import invoice as invoice_mod

    match = build_match(sub)
    period = _resolve_period(match, period_data)
    with tempfile.TemporaryDirectory(prefix="ao-inv-") as tmp:
        out = pathlib.Path(tmp) / "invoice.xlsx"
        invoice_mod.render_invoice_xlsx(match, out, period=period)
        return out.read_bytes()
