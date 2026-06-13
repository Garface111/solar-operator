"""
NEPOOL Operator — AI spreadsheet ingest (Mega-Vector V4).

A NEPOOL stamping-agent tenant already keeps a master roster of every operator
they report for: operator name, array name, NEPOOL-GIS ID, sometimes a GMP
account number. Hand-entering 50 arrays is a non-starter. This module collapses
that to: drop the spreadsheet → confirm a preview table → one-click commit.

Two endpoints (both auth'd as the calling tenant):

  POST /v1/ingest/preview   multipart .xlsx/.xls/.csv  →  parsed rows (NOT saved)
  POST /v1/ingest/commit    edited JSON rows           →  find_or_create + summary

Parsing strategy: flatten the sheet to plain text, hand it to a cheap LLM
(Anthropic Haiku preferred, OpenAI fallback) with a strict-JSON extraction
prompt. If no LLM key is configured OR the call fails, we fall back to a
column-name heuristic parser so the feature never hard-fails.

NOTE (deliberate deviation from the V4 brief): we do NOT add pandas. A CSV is
already plain text and openpyxl already covers .xlsx, so pandas would add a
heavy numpy-backed dependency to the Railway image for zero benefit here. CSVs
are parsed with the stdlib `csv` module.
"""
from __future__ import annotations

import difflib
import io
import os
import csv
import json
import logging
import re
from typing import Optional

import httpx
from fastapi import APIRouter, Header, HTTPException, UploadFile, File
from pydantic import BaseModel
from sqlalchemy import select

from .db import SessionLocal
from .models import Client, Array, UtilityAccount
from .account import tenant_from_session, require_not_demo, require_not_demo
from .import_examples import EXAMPLE_GMCS_STYLE, EXAMPLE_RESIDENTIAL_PORTFOLIO

logger = logging.getLogger(__name__)

router = APIRouter()

# Cap how much text we send to the LLM — a 50-array roster is a few KB; this
# guards against someone dropping a 10k-row export and running up a bill.
MAX_TEXT_CHARS = 60_000
# Cap rows committed in one shot (preview can show more, but refuse a runaway).
MAX_COMMIT_ROWS = 500

# Sonnet 4.5 for the richer hierarchical extraction (was Haiku for the flat one).
ANTHROPIC_MODEL = os.getenv("INGEST_LLM_MODEL", "claude-sonnet-4-5")
OPENAI_MODEL = os.getenv("INGEST_OPENAI_MODEL", "gpt-4o-mini")

# Build few-shot examples to embed in the prompt (compact JSON — LLM reads it fine).
_EX1 = json.dumps({"operators": [EXAMPLE_GMCS_STYLE]}, separators=(",", ":"))
_EX2 = json.dumps({"operators": [EXAMPLE_RESIDENTIAL_PORTFOLIO]}, separators=(",", ":"))

EXTRACTION_PROMPT = (
    "You are a smart data extractor reading a solar operator's client roster spreadsheet.\n\n"
    "TASK: Extract every client, their utility logins, accounts, and arrays into this EXACT nested JSON hierarchy:\n"
    '{"operators":[{"name":"<operator>","clients":[{"name":"<client>","logins":[{"utility":"gmp|vec|null","login_email":"<email or null>","accounts":[{"account_number":"<# or null>","arrays":[{"name":"<array>","nepool_gis_id":"<digits or null>","notes":"<notes or null>","confidence":0.95}]}]}]}]}]}\n\n'
    "SCANNING RULES — follow every one:\n"
    "- Scan EVERY sheet, every row, every column. Do not stop at the first sheet.\n"
    "- NEPOOL-GIS IDs are 4-6 digit numbers. They may appear: inline as 'Array Name (12345)', in a separate column labeled NEPOOL/GIS/Asset ID, in footnotes, or in sheet titles.\n"
    "- Account numbers may be labeled: Account #, GMP #, VEC #, Meter #, Acct, Account Number, etc.\n"
    "- If data is split across sheets (array names on sheet 1, NEPOOL IDs on sheet 2), join them by array name.\n"
    "- If no client/login structure is visible, group all arrays under one client using the operator name as client name, with login_email: null.\n"
    "- If an array has no NEPOOL ID, set nepool_gis_id to null — do NOT skip the array.\n"
    "- If a client has multiple utility logins (e.g. personal + spouse + business), list each as a separate login object.\n\n"
    f"EXAMPLE 1 (community solar, multiple clients with multiple accounts):\n{_EX1}\n\n"
    f"EXAMPLE 2 (residential portfolio, one client per login):\n{_EX2}\n\n"
    "- For each array, add a 'confidence' field (0.0–1.0): your certainty the data is correctly extracted. "
    "Use 0.95+ for clearly labeled values, 0.7–0.94 for minor ambiguity, below 0.7 for guessed or inferred values.\n\n"
    "Return ONLY valid JSON. No markdown fences, no prose. Set unknown fields to null."
)

FIELDS = ("operator_name", "array_name", "nepool_gis_id", "gmp_account_number", "notes")

# Sentinel strings that unambiguously mean "no NEPOOL ID supplied".
_NEPOOL_SENTINELS = frozenset({'', '-', 'n/a', 'na', 'tbd', 'none', '0'})
# 4–6 digit canonical form (extraction prompt says 4-6; CLAUDE.md says 5-6 in prod,
# but 4-digit historical IDs exist in some exports).
_NEPOOL_RE = re.compile(r'^\d{4,6}$')


def clean_text(s: object) -> str:
    """Normalise user-supplied cell text from real-world spreadsheets.

    Handles the full zoo of encoding/typographic garbage that arrives when
    operators copy-paste from PDFs, web portals, or decade-old Excel files:
    smart quotes, NBSP, em/en dashes, zero-width spaces, BOM marks, in-cell
    newlines, and Python byte-string repr leaking into cell values.

    Called at the _normalize boundary so all three ingest paths (GMCS,
    LLM, heuristic) are covered by a single implementation.
    """
    if not isinstance(s, str):
        s = str(s)
    # BOM (U+FEFF) — common at the start of cells in CSV→XLSX conversions
    s = s.lstrip('﻿')
    # Smart / curly quotes → plain ASCII
    s = (s
         .replace('‘', "'").replace('’', "'")
         .replace('“', '"').replace('”', '"'))
    # Em dash (U+2014) / en dash (U+2013) → hyphen
    s = s.replace('—', '-').replace('–', '-')
    # Non-breaking space (U+00A0) → regular space
    s = s.replace(' ', ' ')
    # Zero-width space (U+200B) → removed entirely
    s = s.replace('​', '')
    # Normalise in-cell newlines to a single space (CRLF first, then stragglers)
    s = s.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
    # Python byte-string repr that snuck into a cell: b'Chester' → Chester
    if len(s) >= 4 and s[:2] == "b'" and s[-1] == "'":
        s = s[2:-1]
    elif len(s) >= 4 and s[:2] == 'b"' and s[-1] == '"':
        s = s[2:-1]
    return s.strip()


def _coerce_nepool(raw: Optional[str]) -> tuple[Optional[str], Optional[str]]:
    """Coerce a raw cell value to a canonical NEPOOL GIS ID string.

    Returns (canonical_id | None, error_message | None).
    - (value, None)  → valid ID
    - (None, None)   → legitimately blank/sentinel (no error to show)
    - (None, error)  → parse failure; surface to operator as nepool_parse_error
    """
    if raw is None:
        return None, None
    s = clean_text(raw).lstrip("'")  # strip Excel text-coercion apostrophe
    if s.lower() in _NEPOOL_SENTINELS:
        return None, None
    # Formula errors (#REF!, #VALUE!, …) and formula-injection strings
    if s.startswith('=') or s.startswith('#'):
        return None, f"invalid NEPOOL value {s!r} — fix the source cell"
    # Strip thousands-separator commas only; keep semicolons so "53984; 53985"
    # fails the digit-only check and surfaces a parse error instead of silently
    # taking just the first number.
    cleaned = s.replace(',', '')
    # Numeric conversion handles "53984.0", "5.3984e4", etc.
    try:
        as_float = float(cleaned)
        if as_float != int(as_float):
            return None, f"NEPOOL value {raw!r} has a non-integer fractional part"
        if as_float < 0:
            return None, f"NEPOOL value {raw!r} is negative"
        cleaned = str(int(as_float))
    except ValueError:
        pass  # not a float — fall through to digit-string check
    if not _NEPOOL_RE.match(cleaned):
        return None, f"NEPOOL value {raw!r} is not a 4–6 digit number after cleaning"
    return cleaned, None


# ─── schemas ──────────────────────────────────────────────────────────────

class IngestRow(BaseModel):
    operator_name: Optional[str] = None
    array_name: Optional[str] = None
    nepool_gis_id: Optional[str] = None
    gmp_account_number: Optional[str] = None
    notes: Optional[str] = None
    # "skip" = skip this row entirely; "overwrite"/"new" = proceed normally.
    # Defaults to "new" so existing callers that omit the field keep working.
    collision_action: Optional[str] = "new"


class CommitBody(BaseModel):
    arrays: list[IngestRow]
    # When set, ALL rows are pinned to this Client regardless of the
    # row's operator_name. Used by the per-client "Import arrays into
    # this client" button on the dashboard: the operator already picked
    # the client, so we don't need to do any name-matching/auto-create
    # on the operator_name column.
    force_client_id: Optional[int] = None
    # When True, intra-file NEPOOL duplicates are allowed through anyway.
    # Default False = HARD BLOCK at commit. The preview surfaces the
    # warning; the operator must consciously choose to accept the risk
    # by re-sending with this flag. Belt-and-suspenders to the preview
    # warning, because RECs are revenue and duplicate-NEPOOL Arrays
    # silently poison production attribution downstream.
    allow_intrafile_nepool_duplicates: bool = False


# ─── file → plain text ─────────────────────────────────────────────────────

def _xlsx_to_text(data: bytes) -> str:
    """Flatten ALL worksheets to tab-separated rows with sheet separators.

    Sheets are concatenated with a `--- Sheet: <name> ---` header so the LLM
    can see rosters that span multiple tabs. Previously only the first sheet
    was read, silently dropping every subsequent tab."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for ws in wb.worksheets:
            lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                cells = ["" if c is None else str(c) for c in row]
                if any(c.strip() for c in cells):
                    lines.append("\t".join(cells))
            if lines:
                parts.append(f"--- Sheet: {ws.title} ---\n" + "\n".join(lines))
        return "\n\n".join(parts)
    finally:
        wb.close()


# ─── GMCS-shape detection and extraction ──────────────────────────────────

# Matches "<Array Name> (<NEPOOL-GIS ID>)" — the A1 title written by gmcs_writer.py.
# NEPOOL ID part is optional: arrays without an assigned ID (e.g. "Pittsfield")
# still have a valid GMCS sheet — the ID simply hasn't been entered yet.
_GMCS_A1_RE = re.compile(r"^(.+?)\s*\((\d{2,6})\)\s*$")


def _detect_gmcs_shape(wb) -> bool:
    """True if every non-summary sheet looks like GMCS writer output.

    Conservative: any mismatch → False so we never silently parse a real
    roster as GMCS.  Requires at least one matching sheet.

    NOTE: A1 may be just the array name ("Pittsfield") with no NEPOOL ID
    when the ID hasn't been assigned yet.  The row-5 header check is the
    definitive signal; we only require A1 to be non-empty."""
    checked = 0
    for ws in wb.worksheets:
        title_lower = (ws.title or "").lower()
        if title_lower in ("summary", "notes"):
            continue
        val = str(ws.cell(row=1, column=1).value or "").strip()
        if not val:
            return False
        h_a = str(ws.cell(row=5, column=1).value or "").strip().lower()
        h_b = str(ws.cell(row=5, column=2).value or "").strip().lower()
        # Row-5 header must look like "Quarter" / "Generation (MWh)"
        if "quarter" not in h_a or ("generation" not in h_b and "mwh" not in h_b):
            return False
        checked += 1
    return checked > 0


def _extract_from_gmcs(wb) -> list[dict]:
    """One row per sheet for a GMCS-format workbook.

    operator_name is always None here — the import preview surfaces a single
    global 'Owner / operator' field so the user fills it in once.

    Sheets whose A1 cell is just an array name (no NEPOOL ID in parentheses,
    e.g. "Pittsfield") are included with nepool_gis_id=None so they appear in
    the preview and can receive an ID from a different source."""
    rows: list[dict] = []
    for ws in wb.worksheets:
        if (ws.title or "").lower() in ("summary", "notes"):
            continue
        val = str(ws.cell(row=1, column=1).value or "").strip()
        if not val:
            continue
        m = _GMCS_A1_RE.match(val)
        rows.append({
            "operator_name": None,
            "array_name": m.group(1).strip() if m else val,
            "nepool_gis_id": m.group(2) if m else None,
            "gmp_account_number": None,
            "notes": None,
        })
    return rows


def _csv_to_text(data: bytes) -> str:
    """Normalize a CSV to tab-separated rows (CSV is already text)."""
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = ["\t".join(cell.strip() for cell in row) for row in reader if any(c.strip() for c in row)]
    return "\n".join(lines)


def _file_to_text(filename: str, data: bytes) -> str:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = _csv_to_text(data)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        # .xls (legacy BIFF) isn't supported by openpyxl; surface a clear error
        # rather than a cryptic parse failure.
        if name.endswith(".xls"):
            try:
                text = _xlsx_to_text(data)
            except Exception as e:  # noqa: BLE001
                raise HTTPException(
                    400,
                    "Legacy .xls files aren't supported — re-save as .xlsx or .csv "
                    "and try again.",
                ) from e
        else:
            text = _xlsx_to_text(data)
    else:
        raise HTTPException(400, "Upload a .xlsx or .csv file")
    return text[:MAX_TEXT_CHARS]


# ─── LLM extraction ────────────────────────────────────────────────────────

def _extract_json_block(raw: str) -> dict:
    """Pull the first {...} JSON object out of an LLM response, tolerating
    markdown fences or stray prose around it."""
    raw = raw.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```[a-zA-Z]*\n?", "", raw).rstrip("`").strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        pass
    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        return json.loads(raw[start : end + 1])
    raise ValueError("no JSON object in LLM response")


def _call_anthropic(text: str, api_key: str) -> dict:
    """Call Anthropic and return the full {operators: [...]} hierarchy dict."""
    resp = httpx.post(
        "https://api.anthropic.com/v1/messages",
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": ANTHROPIC_MODEL,
            "max_tokens": 8192,
            "messages": [
                {
                    "role": "user",
                    "content": f"{EXTRACTION_PROMPT}\n\nHere is the roster:\n\n{text}",
                },
                # Prefill forces the model straight into JSON.
                {"role": "assistant", "content": '{"operators":'},
            ],
        },
        timeout=60.0,
    )
    resp.raise_for_status()
    body = resp.json()
    content = "".join(b.get("text", "") for b in body.get("content", []))
    return _extract_json_block('{"operators":' + content)


def _call_openai(text: str, api_key: str) -> dict:
    """Call OpenAI and return the full {operators: [...]} hierarchy dict."""
    resp = httpx.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "content-type": "application/json"},
        json={
            "model": OPENAI_MODEL,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": EXTRACTION_PROMPT},
                {"role": "user", "content": f"Here is the roster:\n\n{text}"},
            ],
        },
        timeout=60.0,
    )
    resp.raise_for_status()
    body = resp.json()
    content = body["choices"][0]["message"]["content"]
    return _extract_json_block(content)


def _llm_extract_hierarchical(text: str) -> Optional[dict]:
    """Try Anthropic then OpenAI; return the full {operators: [...]} dict or None.

    Exported so nepool_assign can reuse the same hierarchical extraction."""
    anthropic_key = os.getenv("ANTHROPIC_API_KEY")
    openai_key = os.getenv("OPENAI_API_KEY")
    if anthropic_key:
        try:
            return _call_anthropic(text, anthropic_key)
        except Exception as e:  # noqa: BLE001
            logger.warning("Anthropic hierarchical extraction failed: %s", e)
    if openai_key:
        try:
            return _call_openai(text, openai_key)
        except Exception as e:  # noqa: BLE001
            logger.warning("OpenAI hierarchical extraction failed: %s", e)
    return None


def _flatten_hierarchy_to_rows(data: dict) -> list[dict]:
    """Flatten {operators: [...]} hierarchy to IngestRow-compatible flat dicts."""
    rows: list[dict] = []
    for operator in data.get("operators", []):
        for client in operator.get("clients", []):
            client_name = (client.get("name") or "").strip() or None
            for login in client.get("logins", []):
                utility = (login.get("utility") or "").strip().lower() or None
                for account in login.get("accounts", []):
                    acct_raw = (account.get("account_number") or "").strip() or None
                    # gmp_account_number is only meaningful for GMP logins.
                    gmp_acct = acct_raw if utility in ("gmp", None) else None
                    for array in account.get("arrays", []):
                        array_name = (array.get("name") or "").strip() or None
                        if not array_name:
                            continue
                        # Extract per-row confidence (added to prompt schema).
                        conf_raw = array.get("confidence")
                        try:
                            conf: Optional[float] = float(conf_raw) if conf_raw is not None else None
                            if conf is not None:
                                conf = max(0.0, min(1.0, conf))
                        except (TypeError, ValueError):
                            conf = None
                        rows.append({
                            "operator_name": client_name,
                            "array_name": array_name,
                            "nepool_gis_id": (array.get("nepool_gis_id") or None),
                            "gmp_account_number": gmp_acct,
                            "notes": (array.get("notes") or None),
                            "_confidence": conf,
                        })
    return rows


def _flatten_hierarchy_to_pairs(data: dict) -> list[dict]:
    """Flatten {operators: [...]} hierarchy to (array_name, nepool_gis_id) pairs.

    Exported for nepool_assign to reuse."""
    pairs: list[dict] = []
    for operator in data.get("operators", []):
        for client in operator.get("clients", []):
            for login in client.get("logins", []):
                for account in login.get("accounts", []):
                    for array in account.get("arrays", []):
                        name = (array.get("name") or "").strip()
                        gis = (array.get("nepool_gis_id") or "").strip()
                        if name and gis:
                            pairs.append({"array_name": name, "nepool_gis_id": gis})
    return pairs


def _count_logins(data: dict) -> int:
    """Count total utility logins across all clients in the hierarchy."""
    count = 0
    for operator in data.get("operators", []):
        for client in operator.get("clients", []):
            count += len(client.get("logins", []))
    return count


def _count_clients(data: dict) -> int:
    """Count total clients across all operators in the hierarchy."""
    count = 0
    for operator in data.get("operators", []):
        count += len(operator.get("clients", []))
    return count


def _llm_extract(text: str) -> Optional[list[dict]]:
    """Thin wrapper used by older callers: hierarchical extract → flat rows."""
    result = _llm_extract_hierarchical(text)
    if result is None:
        return None
    return _flatten_hierarchy_to_rows(result)


# ─── heuristic fallback parser ─────────────────────────────────────────────

# Column-header keyword → our canonical field. First match wins per column.
_HEURISTIC_MAP = [
    ("nepool_gis_id", ("nepool", "gis", "asset id", "asset_id")),
    ("gmp_account_number", ("gmp", "account", "acct")),
    ("operator_name", ("operator", "owner", "company", "client", "customer")),
    ("array_name", ("array", "installation", "site", "system", "project", "name")),
    ("notes", ("note", "comment", "remark")),
]


def _heuristic_extract(text: str) -> list[dict]:
    """Best-effort parse when no LLM is available: match column headers to
    fields by keyword, then read each row into that mapping."""
    rows = [ln.split("\t") for ln in text.splitlines() if ln.strip()]
    if not rows:
        return []
    header = [h.strip().lower() for h in rows[0]]
    # Map each column index → canonical field.
    col_field: dict[int, str] = {}
    used: set[str] = set()
    for idx, col in enumerate(header):
        for field, keywords in _HEURISTIC_MAP:
            if field in used:
                continue
            if any(k in col for k in keywords):
                col_field[idx] = field
                used.add(field)
                break

    out: list[dict] = []
    for raw in rows[1:]:
        entry = {f: None for f in FIELDS}
        any_value = False
        for idx, val in enumerate(raw):
            field = col_field.get(idx)
            v = val.strip()
            if field and v:
                entry[field] = v
                any_value = True
        # If headers were unrecognizable, fall back to positional guessing:
        # operator, array, nepool — the most common roster shape.
        if not col_field and len(raw) >= 2:
            entry["operator_name"] = (raw[0] or "").strip() or None
            entry["array_name"] = (raw[1] or "").strip() or None
            if len(raw) >= 3:
                m = re.search(r"\b\d{4,6}\b", raw[2])
                entry["nepool_gis_id"] = m.group(0) if m else (raw[2].strip() or None)
            any_value = any(entry.values())
        if any_value:
            out.append(entry)
    return out


# Internal keys that must survive _normalize so the preview and commit
# endpoints can access them after normalization.
_PASSTHROUGH_KEYS = frozenset({"_confidence", "collision_action"})


def _normalize(rows: list[dict]) -> list[dict]:
    """Coerce LLM/heuristic output into clean {FIELDS} dicts.

    Applies clean_text() to all string fields and _coerce_nepool() to the
    nepool_gis_id field. If NEPOOL coercion fails, the row is kept (not
    silently dropped) and a nepool_parse_error key is added so the preview
    UI can surface it to the operator.
    """
    clean: list[dict] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        entry: dict = {}
        for f in FIELDS:
            v = r.get(f)
            if v is None:
                entry[f] = None
            elif f == "nepool_gis_id":
                raw_str = v if isinstance(v, str) else str(v)
                coerced, err = _coerce_nepool(raw_str)
                entry[f] = coerced
                if err:
                    entry["nepool_parse_error"] = err
            else:
                s = clean_text(v)
                entry[f] = s or None
        # Preserve passthrough keys so callers can read them post-normalization.
        for key in _PASSTHROUGH_KEYS:
            if key in r:
                entry[key] = r[key]
        # Drop rows with nothing meaningful (no operator AND no array).
        if entry["operator_name"] or entry["array_name"]:
            clean.append(entry)
    return clean


def _find_intrafile_nepool_duplicates(rows: list[dict]) -> dict[str, list[int]]:
    """Find NEPOOL GIS IDs that appear on more than one row of the SAME upload.

    Returns {canonical_nepool_id: [row_index, ...]} containing only IDs that
    occur 2+ times. Rows whose nepool_gis_id is None (blank or parse-error) are
    ignored — a missing ID is not a duplicate.

    Why this is a hard flag, not a silent merge: each preview row is an *array*
    (name + NEPOOL ID + account), not an additive quantity. Two rows claiming the
    same NEPOOL ID with different array names is exactly the ambiguous case where
    we cannot know which name is correct, so we must not auto-merge. Worse, if the
    operator commits both, find_or_create makes TWO Array records sharing one
    NEPOOL ID; downstream production/billing keyed by NEPOOL (nepool_map) then
    collapses to one, silently misattributing kWh — and kWh drives RECs and
    customer revenue. Surfacing the duplicate at preview is the human gate that
    stops that before commit.
    """
    seen: dict[str, list[int]] = {}
    for idx, row in enumerate(rows):
        nid = (row.get("nepool_gis_id") or "").strip()
        if nid:
            seen.setdefault(nid, []).append(idx)
    return {nid: idxs for nid, idxs in seen.items() if len(idxs) > 1}


def _fuzzy_match_client(name: str, clients: list) -> Optional[dict]:
    """Fuzzy-match an operator name against existing Client ORM objects.

    Returns a client_match dict or None. Exact match takes priority."""
    if not name:
        return None
    name_lower = name.strip().lower()
    # Exact match first.
    for c in clients:
        if c.name.strip().lower() == name_lower:
            return {"client_id": c.id, "client_name": c.name, "match_kind": "exact"}
    # Fuzzy: SequenceMatcher >= 0.85 is the threshold specified in the brief.
    best_ratio = 0.0
    best_client = None
    for c in clients:
        ratio = difflib.SequenceMatcher(None, name_lower, c.name.strip().lower()).ratio()
        if ratio >= 0.85 and ratio > best_ratio:
            best_ratio = ratio
            best_client = c
    if best_client is not None:
        return {
            "client_id": best_client.id,
            "client_name": best_client.name,
            "match_kind": "fuzzy",
        }
    return None


# ─── endpoints ─────────────────────────────────────────────────────────────

@router.post("/v1/ingest/preview")
async def ingest_preview(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(default=None),
):
    """Parse an uploaded roster and return extracted rows AS A PREVIEW.
    Nothing is written to the database here."""
    t = tenant_from_session(authorization)

    data = await file.read()
    if not data:
        raise HTTPException(400, "The uploaded file is empty")

    source = "llm"
    rows: Optional[list[dict]] = None
    imported_logins = 0
    imported_clients = 0

    # GMCS-shape detection: load the workbook once and check before falling
    # back to the text/LLM path.  Only applicable for .xlsx.
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            try:
                if _detect_gmcs_shape(wb):
                    rows = _extract_from_gmcs(wb)
                    source = "gmcs_shape"
            finally:
                wb.close()
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(400, f"Couldn't open that Excel file: {exc}") from exc

    if rows is None:
        text = _file_to_text(file.filename or "", data)
        if not text.strip():
            raise HTTPException(400, "Couldn't read any rows from that file")
        hierarchy = _llm_extract_hierarchical(text)
        if hierarchy is not None:
            rows = _flatten_hierarchy_to_rows(hierarchy)
            imported_logins = _count_logins(hierarchy)
            imported_clients = _count_clients(hierarchy)
            source = "llm"
        else:
            rows = _heuristic_extract(text)
            source = "heuristic"

    arrays = _normalize(rows)

    # ── Provenance + collision computation ─────────────────────────────────
    # Single DB session: pull clients + arrays for both the old collision
    # field and the new per-row provenance / NEPOOL collision data.
    from sqlalchemy import and_ as _and
    with SessionLocal() as db:
        existing_client_objs = db.execute(
            select(Client).where(
                Client.tenant_id == t.id,
                Client.deleted_at.is_(None),
            )
        ).scalars().all()

        # Join arrays → clients so we can include client_name in nepool_collision.
        arr_with_clients = db.execute(
            select(Array, Client.name.label("c_name"))
            .outerjoin(
                Client,
                _and(Array.client_id == Client.id, Client.deleted_at.is_(None)),
            )
            .where(
                Array.tenant_id == t.id,
                Array.deleted_at.is_(None),
            )
        ).all()

    # Build lookup sets / maps.
    existing_clients_lower = {c.name.strip().lower() for c in existing_client_objs}
    existing_arrays_lower = {arr.name.strip().lower() for arr, _ in arr_with_clients}
    nepool_map: dict[str, dict] = {}
    for arr, c_name in arr_with_clients:
        nid = (arr.nepool_gis_id or "").strip()
        if nid:
            nepool_map[nid] = {
                "existing_array_id": arr.id,
                "existing_array_name": arr.name,
                "existing_client_name": c_name or "Unknown",
            }

    # Intra-file (within this upload) duplicate NEPOOL IDs. Independent of the
    # DB-level nepool_map collision below: two rows in the SAME file can both be
    # brand-new to the account yet collide with each other.
    intrafile_dup_map = _find_intrafile_nepool_duplicates(arrays)

    nepool_collision_count = 0
    client_fuzzy_count = 0
    warnings: list[dict] = []

    for idx, row in enumerate(arrays):
        # ── Old collision field (backwards compat) ──────────────────────────
        client_hit = bool(
            row.get("operator_name")
            and (row["operator_name"] or "").strip().lower() in existing_clients_lower
        )
        array_hit = bool(
            row.get("array_name")
            and (row["array_name"] or "").strip().lower() in existing_arrays_lower
        )
        if client_hit and array_hit:
            row["collision"] = "both"
        elif client_hit:
            row["collision"] = "client"
        elif array_hit:
            row["collision"] = "array"
        else:
            row["collision"] = None

        # ── Per-row confidence (LLM only) ───────────────────────────────────
        row_confidence: Optional[float] = None
        if source == "llm":
            raw_conf = row.pop("_confidence", None)
            if raw_conf is not None:
                row_confidence = float(raw_conf)
            else:
                row_confidence = 0.85  # backfill for rows from old prompt schema
        else:
            row.pop("_confidence", None)  # remove internal key regardless

        # ── Client match (fuzzy) ────────────────────────────────────────────
        client_match = None
        if row.get("operator_name"):
            client_match = _fuzzy_match_client(row["operator_name"], existing_client_objs)
            if client_match and client_match["match_kind"] == "fuzzy":
                client_fuzzy_count += 1

        # ── NEPOOL collision (against existing DB rows) ─────────────────────
        nepool_collision = None
        nid = (row.get("nepool_gis_id") or "").strip()
        if nid and nid in nepool_map:
            nepool_collision = nepool_map[nid]
            nepool_collision_count += 1

        # ── Intra-file NEPOOL duplicate (against other rows in THIS upload) ──
        intrafile_duplicate = None
        if nid and nid in intrafile_dup_map:
            others = [i for i in intrafile_dup_map[nid] if i != idx]
            intrafile_duplicate = {
                "nepool_gis_id": nid,
                "duplicate_row_indices": others,
                "count": len(intrafile_dup_map[nid]),
            }

        row["provenance"] = {
            "source": source,
            "confidence": row_confidence,
            "client_match": client_match,
            "nepool_collision": nepool_collision,
            "intrafile_nepool_duplicate": intrafile_duplicate,
        }

    # ── Top-level warnings ──────────────────────────────────────────────────
    if not arrays:
        warnings.append({
            "kind": "empty_file",
            "count": 0,
            "message": (
                "We couldn't find any arrays in this file. Try a different file, "
                "paste rows manually, or contact admin@solaroperator.org if this looks wrong."
            ),
        })
    else:
        if source == "llm":
            low_conf = sum(
                1 for r in arrays
                if (r.get("provenance") or {}).get("confidence") is not None
                and r["provenance"]["confidence"] < 0.85
            )
            if low_conf > 0:
                warnings.append({
                    "kind": "low_confidence_rows",
                    "count": low_conf,
                    "message": (
                        f"{low_conf} row{'s' if low_conf != 1 else ''} had low AI confidence "
                        "— review carefully before saving."
                    ),
                })
        if client_fuzzy_count > 0:
            warnings.append({
                "kind": "client_collision",
                "count": client_fuzzy_count,
                "message": (
                    f"{client_fuzzy_count} row{'s' if client_fuzzy_count != 1 else ''} "
                    "have client names similar (but not identical) to existing clients "
                    "— review highlighted rows."
                ),
            })
        if nepool_collision_count > 0:
            warnings.append({
                "kind": "nepool_collision",
                "count": nepool_collision_count,
                "message": (
                    f"{nepool_collision_count} NEPOOL ID{'s' if nepool_collision_count != 1 else ''} "
                    "already exist in your account — choose Skip, Overwrite, or New per row."
                ),
            })
        if intrafile_dup_map:
            dup_ids = len(intrafile_dup_map)
            dup_rows = sum(len(idxs) for idxs in intrafile_dup_map.values())
            warnings.append({
                "kind": "intrafile_nepool_duplicate",
                "count": dup_rows,
                # Expose the offending IDs so the UI can point straight at them.
                "nepool_ids": sorted(intrafile_dup_map.keys()),
                "message": (
                    f"{dup_rows} rows share {dup_ids} NEPOOL ID"
                    f"{'s' if dup_ids != 1 else ''} within this file "
                    f"({', '.join(sorted(intrafile_dup_map.keys()))}). "
                    "Each NEPOOL ID identifies one array — consolidate or correct "
                    "the duplicated rows before saving so production isn't split "
                    "across two arrays."
                ),
            })

    return {
        "ok": True,
        "source": source,
        "count": len(arrays),
        "arrays": arrays,
        "imported_logins": imported_logins,
        "imported_clients": imported_clients,
        "warnings": warnings,
    }


@router.post("/v1/ingest/commit")
def ingest_commit(
    body: CommitBody,
    authorization: Optional[str] = Header(default=None),
):
    """Persist the (user-confirmed, possibly edited) rows. For each row:
      find_or_create Client by operator_name → Array by name → UtilityAccount.
    Returns counts of newly created records."""
    t = tenant_from_session(authorization)
    require_not_demo(t)

    rows = _normalize([r.model_dump() for r in body.arrays])
    if not rows:
        raise HTTPException(400, "Nothing to import")
    if len(rows) > MAX_COMMIT_ROWS:
        raise HTTPException(
            400, f"Too many rows ({len(rows)}). Import at most {MAX_COMMIT_ROWS} at a time."
        )

    # HARD BLOCK: intra-file NEPOOL duplicates poison production attribution
    # downstream. Refuse the commit unless the caller has explicitly accepted
    # the risk via allow_intrafile_nepool_duplicates=True. Same detector the
    # preview uses, so any warning the operator saw is now enforced.
    if not body.allow_intrafile_nepool_duplicates:
        intrafile_dups = _find_intrafile_nepool_duplicates(rows)
        if intrafile_dups:
            n_ids = len(intrafile_dups)
            n_rows = sum(len(idxs) for idxs in intrafile_dups.values())
            sample = next(iter(intrafile_dups.keys()))
            raise HTTPException(
                status_code=409,
                detail={
                    "error": "intrafile-nepool-duplicates",
                    "message": (
                        f"{n_rows} rows share {n_ids} duplicated NEPOOL ID(s) "
                        f"(e.g. \"{sample}\"). Duplicates poison production "
                        f"attribution downstream. Resolve in the source file, "
                        f"or re-commit with allow_intrafile_nepool_duplicates=true "
                        f"to override."
                    ),
                    "duplicate_nepool_ids": list(intrafile_dups.keys()),
                    "duplicate_row_count": n_rows,
                },
            )

    clients_created = arrays_created = accounts_created = skipped_count = 0
    # Within-batch caches so two rows for the same operator don't both try to
    # create the client (and so we don't re-query each row).
    client_cache: dict[str, Client] = {}

    with SessionLocal() as db:
        # If force_client_id is set, pre-load and validate the target Client
        # once; every row will pin to it instead of routing on operator_name.
        forced_client: Optional[Client] = None
        if body.force_client_id is not None:
            forced_client = db.execute(
                select(Client).where(
                    Client.tenant_id == t.id,
                    Client.id == body.force_client_id,
                    Client.deleted_at.is_(None),
                )
            ).scalar_one_or_none()
            if forced_client is None:
                raise HTTPException(404, f"Client {body.force_client_id} not found")

        def find_or_create_client(name: str) -> Client:
            nonlocal clients_created
            # force_client_id short-circuit: ignore the row's operator_name,
            # everything lands on the pre-selected Client.
            if forced_client is not None:
                return forced_client
            key = name.lower()
            if key in client_cache:
                return client_cache[key]
            c = db.execute(
                select(Client).where(Client.tenant_id == t.id, Client.name == name)
            ).scalar_one_or_none()
            if c is None:
                c = Client(tenant_id=t.id, name=name, active=True)
                db.add(c)
                db.flush()
                clients_created += 1
            client_cache[key] = c
            return c

        for r in rows:
            # collision_action="skip" means the user explicitly chose to skip
            # this row (e.g. for a NEPOOL collision they don't want to resolve).
            if (r.get("collision_action") or "new") == "skip":
                skipped_count += 1
                continue

            operator = r["operator_name"] or "Unassigned"
            array_name = r["array_name"]
            if not array_name:
                # No array name → nothing concrete to create; skip.
                continue

            client = find_or_create_client(operator)

            # Array unique on (tenant_id, name). If it already exists, reuse it
            # and backfill NEPOOL / notes if they were blank. Exclude soft-deleted
            # rows so they don't block re-import after a bulk delete.
            arr = db.execute(
                select(Array).where(
                    Array.tenant_id == t.id,
                    Array.name == array_name,
                    Array.deleted_at.is_(None),
                )
            ).scalar_one_or_none()
            if arr is None:
                arr = Array(
                    tenant_id=t.id,
                    client_id=client.id,
                    name=array_name,
                    nepool_gis_id=r["nepool_gis_id"],
                    notes=r["notes"],
                )
                db.add(arr)
                db.flush()
                arrays_created += 1
            else:
                if not arr.client_id:
                    arr.client_id = client.id
                if not arr.nepool_gis_id and r["nepool_gis_id"]:
                    arr.nepool_gis_id = r["nepool_gis_id"]
                if not arr.notes and r["notes"]:
                    arr.notes = r["notes"]

            acct_num = (r["gmp_account_number"] or "").strip()
            if acct_num:
                # UtilityAccount unique on (tenant_id, provider, account_number).
                existing = db.execute(
                    select(UtilityAccount).where(
                        UtilityAccount.tenant_id == t.id,
                        UtilityAccount.provider == "gmp",
                        UtilityAccount.account_number == acct_num,
                    )
                ).scalar_one_or_none()
                if existing is None:
                    db.add(UtilityAccount(
                        tenant_id=t.id,
                        array_id=arr.id,
                        provider="gmp",
                        account_number=acct_num,
                        nickname=array_name,
                    ))
                    accounts_created += 1
                elif existing.array_id is None:
                    existing.array_id = arr.id

        db.commit()

    return {
        "ok": True,
        "clients_created": clients_created,
        "arrays_created": arrays_created,
        "accounts_created": accounts_created,
        "skipped_count": skipped_count,
    }
