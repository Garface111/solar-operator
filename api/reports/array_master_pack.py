"""Per-array Master Data Pack — mega spreadsheet from utility bills + daily gen.

Paul Bozuwa (2026-07): years of GMP bills are the settlement spine; owners need
a downloadable master workbook per array (and a zip of all arrays) so they can
live in their own tools (iCloud Drive via Mail/Files, Excel, etc.).

Sheets:
  Meta     — array identity, accounts, coverage, sources, generated-at
  Bills    — every absorbed utility bill (period, kWh, $, status)
  Monthly  — calendar months (daily wins where present, else bill pro-rate)
  Daily    — full daily series with source labels
  YoY      — calendar years + trailing-12-mo + vs prior year

Never fabricates rows. Empty arrays still get a valid workbook that says so.
"""
from __future__ import annotations

import io
import re
import zipfile
from collections import defaultdict
from datetime import date, datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from ..bill_attribution import distribute_kwh_by_calendar_day
from ..db import SessionLocal
from ..models import Array, Bill, DailyGeneration, Tenant, UtilityAccount, local_today

_GREEN = "047857"
_HDR_FILL = PatternFill("solid", fgColor=_GREEN)
_HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
_TITLE = Font(bold=True, size=14, color="1F4E2A")
_SECTION = Font(bold=True, size=12, color="1F4E2A")
_CELL = Font(size=10)
_BOLD = Font(bold=True, size=10)
_MUTED = Font(size=9, italic=True, color="666666")
_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_RIGHT = Alignment(horizontal="right")
_CENTER = Alignment(horizontal="center")
_MONTH = (
    "", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
)


def _safe_name(name: str, max_len: int = 40) -> str:
    s = re.sub(r"[^\w\s\-\.]", "", (name or "array").strip())
    s = re.sub(r"\s+", "-", s).strip("-._") or "array"
    return s[:max_len]


def _d(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _hdr_row(ws, row: int, labels: list[str]) -> None:
    for col, lab in enumerate(labels, 1):
        c = ws.cell(row, col, lab)
        c.font = _HDR_FONT
        c.fill = _HDR_FILL
        c.alignment = _CENTER
        c.border = _THIN


def _autosize(ws, max_width: int = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is None:
                continue
            width = min(max_width, max(width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _load_array_pack(db, tenant_id: str, array_id: int) -> Optional[dict]:
    """Pull all source rows for one array. None if array missing / wrong tenant."""
    arr = db.get(Array, array_id)
    if arr is None or arr.tenant_id != tenant_id or arr.deleted_at is not None:
        return None

    accounts = db.execute(
        select(UtilityAccount).where(
            UtilityAccount.array_id == array_id,
            UtilityAccount.deleted_at.is_(None),
        )
    ).scalars().all()
    acct_ids = [ua.id for ua in accounts]
    acct_by_id = {ua.id: ua for ua in accounts}

    bills = []
    if acct_ids:
        bills = list(db.execute(
            select(Bill).where(Bill.account_id.in_(acct_ids))
        ).scalars().all())
        # Sort in Python so SQLite/Postgres both handle null period_end the same.
        bills.sort(key=lambda b: (
            _d(b.period_end) or date.min,
            b.id or 0,
        ))

    daily_rows = db.execute(
        select(DailyGeneration)
        .where(DailyGeneration.array_id == array_id)
        .order_by(DailyGeneration.day.asc())
    ).scalars().all()

    # Monthly merge: daily wins for a month; bill fill otherwise
    daily_by_ym: dict[tuple[int, int], float] = defaultdict(float)
    daily_by_year: dict[int, float] = defaultdict(float)
    daily_by_day: list[dict] = []
    for r in daily_rows:
        if r.day is None or r.kwh is None:
            continue
        k = float(r.kwh)
        daily_by_ym[(r.day.year, r.day.month)] += k
        daily_by_year[r.day.year] += k
        daily_by_day.append({
            "day": r.day,
            "kwh": k,
            "source": r.source or "unknown",
        })

    bill_by_ym: dict[tuple[int, int], float] = defaultdict(float)
    bill_rows: list[dict] = []
    for b in bills:
        ua = acct_by_id.get(b.account_id)
        pstart = _d(b.period_start)
        pend = _d(b.period_end)
        bill_rows.append({
            "bill_id": b.id,
            "provider": (ua.provider if ua else None) or "",
            "account_number": (ua.account_number if ua else None) or "",
            "period_start": pstart,
            "period_end": pend,
            "billing_days": b.billing_days,
            "kwh_generated": b.kwh_generated,
            "kwh_sent_to_grid": b.kwh_sent_to_grid,
            "kwh_consumed": b.kwh_consumed,
            "kwh_gross_generated": b.kwh_gross_generated,
            "total_cost": b.total_cost,
            "net_credit": b.net_credit,
            "solar_credit_usd": b.solar_credit_usd,
            "avg_rate_cents_kwh": b.avg_rate_cents_kwh,
            "supplier": b.supplier,
            "document_number": b.document_number,
            "parse_status": b.parse_status,
            "is_net_metered": b.is_net_metered,
        })
        for ym, kwh in distribute_kwh_by_calendar_day(b).items():
            bill_by_ym[ym] += kwh

    monthly: list[dict] = []
    all_ym = set(daily_by_ym.keys()) | set(bill_by_ym.keys())
    for ym in sorted(all_ym):
        if ym in daily_by_ym:
            monthly.append({
                "year": ym[0], "month": ym[1],
                "kwh": round(daily_by_ym[ym], 3),
                "source": "daily",
            })
        else:
            monthly.append({
                "year": ym[0], "month": ym[1],
                "kwh": round(bill_by_ym[ym], 3),
                "source": "bill",
            })

    # YoY + TTM
    today = local_today()
    by_year: dict[int, float] = defaultdict(float)
    for m in monthly:
        by_year[m["year"]] += m["kwh"]

    ttm = 0.0
    for off in range(12):
        yy, mm = today.year, today.month - off
        while mm <= 0:
            mm += 12
            yy -= 1
        # prefer monthly merged value
        hit = next((x for x in monthly if x["year"] == yy and x["month"] == mm), None)
        if hit:
            ttm += hit["kwh"]
        else:
            ttm += daily_by_ym.get((yy, mm), 0.0) or bill_by_ym.get((yy, mm), 0.0)

    prior_year = today.year - 1
    prior_kwh = by_year.get(prior_year, 0.0)
    ttm_vs_prior = None
    if prior_kwh > 0 and ttm > 0:
        ttm_vs_prior = round(100.0 * (ttm - prior_kwh) / prior_kwh, 1)

    return {
        "array": arr,
        "accounts": accounts,
        "bills": bill_rows,
        "daily": daily_by_day,
        "monthly": monthly,
        "by_year": dict(sorted(by_year.items())),
        "ttm_kwh": round(ttm, 3),
        "prior_year": prior_year,
        "prior_year_kwh": round(prior_kwh, 3) if prior_kwh else None,
        "ttm_vs_prior_year_pct": ttm_vs_prior,
        "today": today,
        "generated_at": datetime.now(timezone.utc),
    }


def build_array_master_workbook(
    tenant_id: str,
    array_id: int,
    *,
    db=None,
) -> Optional[bytes]:
    """Return xlsx bytes for one array, or None if not found / wrong tenant."""
    own_db = db is None
    if own_db:
        db = SessionLocal()
    try:
        pack = _load_array_pack(db, tenant_id, array_id)
        if pack is None:
            return None
        return _render_workbook(pack)
    finally:
        if own_db:
            db.close()


def _render_workbook(pack: dict) -> bytes:
    arr = pack["array"]
    wb = Workbook()

    # ── Meta ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Meta"
    ws["A1"] = "Array Master Data Pack"
    ws["A1"].font = _TITLE
    ws.merge_cells("A1:B1")
    rows = [
        ("Array name", arr.name or ""),
        ("Array id", arr.id),
        ("Tenant id", arr.tenant_id),
        ("Fuel", arr.fuel_type or "solar"),
        ("NEPOOL GIS id", arr.nepool_gis_id or ""),
        ("Generated at (UTC)", pack["generated_at"].strftime("%Y-%m-%d %H:%M UTC")),
        ("Bills on file", len(pack["bills"])),
        ("Daily rows", len(pack["daily"])),
        ("Monthly months", len(pack["monthly"])),
        ("Trailing 12 mo (kWh)", pack["ttm_kwh"]),
        (f"Full {pack['prior_year']} (kWh)", pack["prior_year_kwh"] or "—"),
        ("TTM vs prior year", (
            f"{pack['ttm_vs_prior_year_pct']:+.1f}%"
            if pack["ttm_vs_prior_year_pct"] is not None else "—"
        )),
        ("", ""),
        ("Utility accounts", ""),
    ]
    r = 3
    for lab, val in rows:
        ws.cell(r, 1, lab).font = _BOLD if lab else _CELL
        ws.cell(r, 2, val).font = _CELL
        r += 1
    for ua in pack["accounts"]:
        ws.cell(r, 1, f"  {(ua.provider or '').upper()}").font = _CELL
        ws.cell(r, 2, ua.account_number or ua.nickname or f"id={ua.id}").font = _CELL
        r += 1
    if not pack["accounts"]:
        ws.cell(r, 1, "  (none linked)").font = _MUTED
        r += 1
    r += 1
    ws.cell(r, 1, "Sources").font = _SECTION
    r += 1
    for line in (
        "Bills sheet = absorbed utility bills (GMP/VEC/… settlement spine).",
        "Daily sheet = DailyGeneration rows (meter/vendor/extension), source-labeled.",
        "Monthly sheet = daily wins for a month; bill pro-rate fills months with no daily.",
        "YoY sheet = calendar years + trailing 12 months for fair year-over-year compare.",
        "Nothing is invented — empty sheets mean no data of that kind yet.",
    ):
        ws.cell(r, 1, line).font = _MUTED
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
    _autosize(ws, 48)

    # ── Bills ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Bills")
    _hdr_row(ws, 1, [
        "Period start", "Period end", "Billing days", "Provider", "Account",
        "kWh generated", "kWh to grid", "kWh consumed", "kWh gross gen",
        "Total cost $", "Net credit $", "Solar credit $", "Avg ¢/kWh",
        "Supplier", "Document #", "Net metered", "Parse status", "Bill id",
    ])
    for i, b in enumerate(pack["bills"], 2):
        vals = [
            b["period_start"], b["period_end"], b["billing_days"],
            b["provider"], b["account_number"],
            b["kwh_generated"], b["kwh_sent_to_grid"], b["kwh_consumed"],
            b["kwh_gross_generated"],
            b["total_cost"], b["net_credit"], b["solar_credit_usd"],
            b["avg_rate_cents_kwh"],
            b["supplier"], b["document_number"],
            b["is_net_metered"], b["parse_status"], b["bill_id"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            c.font = _CELL
            c.border = _THIN
            if isinstance(v, float):
                c.number_format = "#,##0.00"
                c.alignment = _RIGHT
            elif isinstance(v, int) and col not in (18,):
                c.number_format = "#,##0"
                c.alignment = _RIGHT
            elif isinstance(v, date):
                c.number_format = "YYYY-MM-DD"
    if not pack["bills"]:
        ws.cell(2, 1, "No utility bills absorbed for this array yet.").font = _MUTED
    _autosize(ws)

    # ── Monthly ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Monthly")
    _hdr_row(ws, 1, ["Year", "Month", "Month label", "kWh", "Source"])
    for i, m in enumerate(pack["monthly"], 2):
        vals = [
            m["year"], m["month"],
            f"{_MONTH[m['month']]} {m['year']}",
            m["kwh"], m["source"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(i, col, v)
            c.font = _CELL
            c.border = _THIN
            if col == 4:
                c.number_format = "#,##0.0"
                c.alignment = _RIGHT
    if not pack["monthly"]:
        ws.cell(2, 1, "No monthly production yet.").font = _MUTED
    _autosize(ws)

    # ── Daily ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Daily")
    _hdr_row(ws, 1, ["Date", "kWh", "Source"])
    for i, d in enumerate(pack["daily"], 2):
        ws.cell(i, 1, d["day"]).font = _CELL
        ws.cell(i, 1).number_format = "YYYY-MM-DD"
        ws.cell(i, 1).border = _THIN
        c = ws.cell(i, 2, d["kwh"])
        c.font = _CELL
        c.number_format = "#,##0.00"
        c.alignment = _RIGHT
        c.border = _THIN
        ws.cell(i, 3, d["source"]).font = _CELL
        ws.cell(i, 3).border = _THIN
    if not pack["daily"]:
        ws.cell(2, 1, "No daily generation rows yet.").font = _MUTED
    _autosize(ws)

    # ── YoY ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("YoY")
    _hdr_row(ws, 1, ["Period", "kWh", "Notes"])
    r = 2
    for year, kwh in pack["by_year"].items():
        note = "YTD (incomplete)" if year == pack["today"].year else "Full calendar year"
        ws.cell(r, 1, str(year)).font = _CELL
        c = ws.cell(r, 2, round(kwh, 1))
        c.font = _CELL
        c.number_format = "#,##0.0"
        c.alignment = _RIGHT
        ws.cell(r, 3, note).font = _MUTED
        for col in range(1, 4):
            ws.cell(r, col).border = _THIN
        r += 1
    # Trailing 12
    ws.cell(r, 1, "Trailing 12 mo").font = _BOLD
    c = ws.cell(r, 2, pack["ttm_kwh"])
    c.font = _BOLD
    c.number_format = "#,##0.0"
    c.alignment = _RIGHT
    ws.cell(r, 3, "Last 12 calendar months — fair compare to a full prior year").font = _MUTED
    for col in range(1, 4):
        ws.cell(r, col).border = _THIN
    r += 1
    if pack["prior_year_kwh"] is not None:
        ws.cell(r, 1, f"Full {pack['prior_year']}").font = _CELL
        c = ws.cell(r, 2, pack["prior_year_kwh"])
        c.font = _CELL
        c.number_format = "#,##0.0"
        c.alignment = _RIGHT
        note = "Prior full year"
        if pack["ttm_vs_prior_year_pct"] is not None:
            note += f" · TTM vs this: {pack['ttm_vs_prior_year_pct']:+.1f}%"
        ws.cell(r, 3, note).font = _MUTED
        for col in range(1, 4):
            ws.cell(r, col).border = _THIN
    if not pack["by_year"] and not pack["ttm_kwh"]:
        ws.cell(2, 1, "No yearly production yet.").font = _MUTED
    _autosize(ws, 56)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def array_master_filename(array_name: str, array_id: int,
                          when: Optional[date] = None) -> str:
    when = when or local_today()
    return f"{_safe_name(array_name)}-{array_id}-master-data-{when.isoformat()}.xlsx"


def build_fleet_master_zip(tenant_id: str, *, db=None) -> tuple[bytes, int]:
    """Zip of one master pack per live array. Returns (zip_bytes, array_count)."""
    own_db = db is None
    if own_db:
        db = SessionLocal()
    try:
        arrays = db.execute(
            select(Array).where(
                Array.tenant_id == tenant_id,
                Array.deleted_at.is_(None),
            ).order_by(Array.name)
        ).scalars().all()
        buf = io.BytesIO()
        n = 0
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for arr in arrays:
                blob = build_array_master_workbook(tenant_id, arr.id, db=db)
                if not blob:
                    continue
                zf.writestr(array_master_filename(arr.name or "array", arr.id), blob)
                n += 1
        return buf.getvalue(), n
    finally:
        if own_db:
            db.close()
