"""All-time aggregated fleet report — Excel + PDF.

`build_fleet_report(tenant, fmt)` aggregates the operator's WHOLE fleet,
ALL-TIME, reading live from the DB each call. Because it queries current rows
every time it is generated, the report ALWAYS reflects the newest absorbed
month automatically — it is generated on demand, never a frozen snapshot.

Data sources (real rows only — never fabricated):
  - DailyGeneration: per-array daily kWh (authoritative production history).
  - Bill: absorbed utility bills (generation, consumption, cost/credit) joined
    to arrays through UtilityAccount.array_id.

Aggregation rules mirror the rest of the app to avoid double-counting:
  - For any (array, year, month) that has DailyGeneration rows, that month's
    generation comes EXCLUSIVELY from DailyGeneration.
  - Months with no DailyGeneration fall back to Bill.kwh_generated
    (pro-rated across calendar days like the GMCS writer).
  - Consumption + credit value are summed from Bill rows (fleet totals).

An empty fleet yields a VALID report that says "no data yet" — not invented
totals.
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime, timezone
from typing import Optional

from sqlalchemy import select

from ..bill_attribution import distribute_kwh_by_calendar_day
from ..db import SessionLocal
from ..models import Array, Bill, Tenant, UtilityAccount, DailyGeneration

# Theme colors (match the app: primary-600 green, cream surfaces).
_GREEN = "047857"
_GREEN_DARK = "064E3B"
_CREAM = "FAF7F0"
_CREAM_BORDER = "E8E2D9"
_GREY = "666666"

_MONTH_NAMES = [
    "", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


# ── data shape ────────────────────────────────────────────────────────────


class FleetAggregate:
    """A computed, all-time fleet aggregate. Pure data — no formatting."""

    def __init__(self) -> None:
        self.company_name: str = ""
        self.generated_at: datetime = datetime.now(timezone.utc)
        # generation kWh
        self.by_year: dict[int, float] = {}                      # year -> kWh
        self.by_month: dict[tuple[int, int], float] = {}         # (y, m) -> kWh
        self.by_array: list[dict] = []                           # per-array rows
        self.total_generated_kwh: float = 0.0
        # bill-derived fleet totals (may be absent if no bills)
        self.total_consumed_kwh: float = 0.0
        self.total_cost: float = 0.0          # sum of bill total_cost ($)
        self.total_credit: float = 0.0        # sum of bill net_credit ($)
        self.has_consumption: bool = False
        self.has_cost: bool = False
        # coverage
        self.date_min: Optional[date] = None
        self.date_max: Optional[date] = None
        self.array_count: int = 0
        self.bill_count: int = 0

    @property
    def has_data(self) -> bool:
        return self.total_generated_kwh > 0 or self.bill_count > 0


def _bump_range(agg: FleetAggregate, d: date) -> None:
    if agg.date_min is None or d < agg.date_min:
        agg.date_min = d
    if agg.date_max is None or d > agg.date_max:
        agg.date_max = d


def aggregate_fleet(tenant: Tenant) -> FleetAggregate:
    """Read the DB live and aggregate the tenant's whole fleet, all-time.

    Generation is computed per (array, year, month): DailyGeneration wins for
    any month it covers, else Bill.kwh_generated (calendar-day pro-rated).
    """
    agg = FleetAggregate()
    agg.company_name = (
        tenant.company_name or tenant.name or tenant.operator_name or "Your fleet"
    )

    with SessionLocal() as db:
        arrays = db.execute(
            select(Array).where(
                Array.tenant_id == tenant.id,
                Array.deleted_at.is_(None),
            ).order_by(Array.name)
        ).scalars().all()
        arrays_by_id = {a.id: a for a in arrays}
        array_ids = list(arrays_by_id.keys())
        agg.array_count = len(array_ids)

        # ── 1. DailyGeneration: per (array, year, month) kWh ──────────────
        # daily_by_array[array_id][(y, m)] = kWh
        daily_by_array: dict[int, dict[tuple[int, int], float]] = defaultdict(
            lambda: defaultdict(float))
        if array_ids:
            rows = db.execute(
                select(
                    DailyGeneration.array_id,
                    DailyGeneration.day,
                    DailyGeneration.kwh,
                ).where(DailyGeneration.array_id.in_(array_ids))
            ).all()
            for aid, day, kwh in rows:
                if day is None or kwh is None:
                    continue
                daily_by_array[aid][(day.year, day.month)] += float(kwh)
                _bump_range(agg, day)

        # ── 2. Bills joined to arrays via UtilityAccount.array_id ─────────
        # account_id -> array_id
        acct_to_array: dict[int, int] = {}
        if array_ids:
            accts = db.execute(
                select(UtilityAccount.id, UtilityAccount.array_id).where(
                    UtilityAccount.array_id.in_(array_ids)
                )
            ).all()
            acct_to_array = {aid: arr for aid, arr in accts if arr is not None}

        # bill_by_array[array_id][(y, m)] = kWh generated (from bills)
        bill_gen_by_array: dict[int, dict[tuple[int, int], float]] = defaultdict(
            lambda: defaultdict(float))
        if acct_to_array:
            bills = db.execute(
                select(Bill).where(Bill.account_id.in_(list(acct_to_array.keys())))
            ).scalars().all()
            for b in bills:
                aid = acct_to_array.get(b.account_id)
                if aid is None:
                    continue
                agg.bill_count += 1
                # Fleet consumption / cost / credit totals (bill-level).
                if b.kwh_consumed is not None:
                    agg.total_consumed_kwh += float(b.kwh_consumed)
                    agg.has_consumption = True
                if b.total_cost is not None:
                    agg.total_cost += float(b.total_cost)
                    agg.has_cost = True
                if b.net_credit is not None:
                    agg.total_credit += float(b.net_credit)
                    agg.has_cost = True
                # Per-month generation, pro-rated across calendar days.
                for (year, month), kwh in distribute_kwh_by_calendar_day(b).items():
                    bill_gen_by_array[aid][(year, month)] += kwh
                    _bump_range(agg, date(year, month, 1))

        # ── 3. Merge: DailyGeneration wins per covered month ──────────────
        # per_array_month[array_id][(y, m)] = chosen kWh
        for aid in array_ids:
            merged: dict[tuple[int, int], float] = {}
            daily = daily_by_array.get(aid, {})
            billed = bill_gen_by_array.get(aid, {})
            for ym, kwh in daily.items():
                merged[ym] = kwh
            for ym, kwh in billed.items():
                if ym not in merged:  # only when no DailyGeneration that month
                    merged[ym] = kwh

            arr = arrays_by_id[aid]
            arr_total = sum(merged.values())
            # roll up into fleet year/month buckets
            for (year, month), kwh in merged.items():
                agg.by_year[year] = agg.by_year.get(year, 0.0) + kwh
                agg.by_month[(year, month)] = (
                    agg.by_month.get((year, month), 0.0) + kwh
                )
            agg.total_generated_kwh += arr_total
            agg.by_array.append({
                "array_id": aid,
                "name": arr.name,
                "nepool_gis_id": arr.nepool_gis_id,
                "kwh": arr_total,
                "months": len(merged),
            })

    # Sort per-array breakdown by generation, biggest first.
    agg.by_array.sort(key=lambda r: r["kwh"], reverse=True)
    return agg


# ── formatting helpers ────────────────────────────────────────────────────


def _date_range_label(agg: FleetAggregate) -> str:
    if agg.date_min and agg.date_max:
        return (f"{agg.date_min.strftime('%b %Y')} – "
                f"{agg.date_max.strftime('%b %Y')}")
    return "—"


def _kwh(v: float) -> str:
    return f"{v:,.0f}"


def _mwh(v: float) -> str:
    return f"{v / 1000.0:,.2f}"


# ── Excel writer ──────────────────────────────────────────────────────────


def _build_xlsx(agg: FleetAggregate) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    sh = wb.active
    sh.title = "Fleet Report"

    title_font = Font(bold=True, size=16, color=_GREEN_DARK)
    sub_font = Font(size=10, color=_GREY)
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor=_GREEN_DARK)
    section_font = Font(bold=True, size=12, color=_GREEN_DARK)
    total_font = Font(bold=True, size=11, color="1F2937")
    total_fill = PatternFill("solid", fgColor=_CREAM)
    cell_font = Font(size=10, color="1F2937")
    thin = Side(style="thin", color=_CREAM_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")
    left = Alignment(horizontal="left")
    center = Alignment(horizontal="center")

    for col, width in (("A", 28), ("B", 18), ("C", 18), ("D", 16), ("E", 14)):
        sh.column_dimensions[col].width = width

    r = 1
    sh.cell(r, 1, f"All-Time Fleet Report — {agg.company_name}").font = title_font
    sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    gen = agg.generated_at.strftime("%Y-%m-%d %H:%M UTC")
    sh.cell(r, 1,
            f"Generated {gen}  ·  Coverage: {_date_range_label(agg)}  ·  "
            f"{agg.array_count} array(s)  ·  {agg.bill_count} bill(s)").font = sub_font
    sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 2

    if not agg.has_data:
        c = sh.cell(r, 1, "No fleet data absorbed yet.")
        c.font = Font(bold=True, size=12, color=_GREY)
        sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        sh.cell(r, 1,
                "Connect a utility login or upload daily generation and your "
                "all-time totals will appear here automatically.").font = sub_font
        sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Fleet summary block ──
    sh.cell(r, 1, "Fleet summary").font = section_font
    r += 1
    summary_rows = [
        ("Total generation (kWh)", _kwh(agg.total_generated_kwh)),
        ("Total generation (MWh)", _mwh(agg.total_generated_kwh)),
    ]
    if agg.has_consumption:
        summary_rows.append(("Total consumption (kWh)", _kwh(agg.total_consumed_kwh)))
    if agg.has_cost:
        # Negative net cost == net credit earned.
        net = agg.total_cost
        if net < 0:
            summary_rows.append(("Net credit value", f"${abs(net):,.2f}"))
        else:
            summary_rows.append(("Net billed", f"${net:,.2f}"))
        if agg.total_credit:
            summary_rows.append(("Net-metering credit", f"${agg.total_credit:,.2f}"))
    for label, val in summary_rows:
        lc = sh.cell(r, 1, label); lc.font = cell_font; lc.border = border
        vc = sh.cell(r, 2, val); vc.font = total_font; vc.alignment = right
        vc.border = border
        r += 1
    r += 1

    # ── Generation by year ──
    sh.cell(r, 1, "Generation by year").font = section_font
    r += 1
    for col, label in enumerate(["Year", "Generation (kWh)", "Generation (MWh)"], 1):
        c = sh.cell(r, col, label)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
    r += 1
    for year in sorted(agg.by_year):
        kwh = agg.by_year[year]
        sh.cell(r, 1, year).font = cell_font
        sh.cell(r, 1).border = border
        c2 = sh.cell(r, 2, round(kwh, 1)); c2.alignment = right; c2.font = cell_font
        c2.number_format = "#,##0"; c2.border = border
        c3 = sh.cell(r, 3, round(kwh / 1000.0, 3)); c3.alignment = right
        c3.font = cell_font; c3.number_format = "#,##0.000"; c3.border = border
        r += 1
    # year total
    tc = sh.cell(r, 1, "Total"); tc.font = total_font; tc.fill = total_fill
    tc.border = border
    t2 = sh.cell(r, 2, round(agg.total_generated_kwh, 1)); t2.alignment = right
    t2.font = total_font; t2.fill = total_fill; t2.number_format = "#,##0"
    t2.border = border
    t3 = sh.cell(r, 3, round(agg.total_generated_kwh / 1000.0, 3)); t3.alignment = right
    t3.font = total_font; t3.fill = total_fill; t3.number_format = "#,##0.000"
    t3.border = border
    r += 2

    # ── Generation by month ──
    sh.cell(r, 1, "Generation by month").font = section_font
    r += 1
    for col, label in enumerate(["Month", "Generation (kWh)", "Generation (MWh)"], 1):
        c = sh.cell(r, col, label)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
    r += 1
    for (year, month) in sorted(agg.by_month):
        kwh = agg.by_month[(year, month)]
        label = f"{_MONTH_NAMES[month]} {year}"
        sh.cell(r, 1, label).font = cell_font
        sh.cell(r, 1).border = border
        c2 = sh.cell(r, 2, round(kwh, 1)); c2.alignment = right; c2.font = cell_font
        c2.number_format = "#,##0"; c2.border = border
        c3 = sh.cell(r, 3, round(kwh / 1000.0, 3)); c3.alignment = right
        c3.font = cell_font; c3.number_format = "#,##0.000"; c3.border = border
        r += 1
    r += 1

    # ── Per-array breakdown ──
    sh.cell(r, 1, "Per-array breakdown (all-time)").font = section_font
    r += 1
    for col, label in enumerate(
        ["Array", "NEPOOL-GIS", "Generation (kWh)", "Generation (MWh)", "Months"], 1
    ):
        c = sh.cell(r, col, label)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = border
    r += 1
    for row in agg.by_array:
        sh.cell(r, 1, row["name"]).font = cell_font
        sh.cell(r, 1).border = border
        nc = sh.cell(r, 2, row["nepool_gis_id"] or "—"); nc.font = cell_font
        nc.alignment = center; nc.border = border
        c3 = sh.cell(r, 3, round(row["kwh"], 1)); c3.alignment = right
        c3.font = cell_font; c3.number_format = "#,##0"; c3.border = border
        c4 = sh.cell(r, 4, round(row["kwh"] / 1000.0, 3)); c4.alignment = right
        c4.font = cell_font; c4.number_format = "#,##0.000"; c4.border = border
        c5 = sh.cell(r, 5, row["months"]); c5.alignment = center; c5.font = cell_font
        c5.border = border
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── PDF writer ────────────────────────────────────────────────────────────


def _build_pdf(agg: FleetAggregate) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)

    GREEN = colors.HexColor(f"#{_GREEN}")
    GREEN_DARK = colors.HexColor(f"#{_GREEN_DARK}")
    CREAM = colors.HexColor(f"#{_CREAM}")
    BORDER = colors.HexColor(f"#{_CREAM_BORDER}")

    styles = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=styles["Title"], textColor=GREEN_DARK,
                       fontSize=18, spaceAfter=4)
    sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9,
                         textColor=colors.HexColor(f"#{_GREY}"))
    sec = ParagraphStyle("sec", parent=styles["Heading2"], textColor=GREEN_DARK,
                         fontSize=13, spaceBefore=14, spaceAfter=6)
    note = ParagraphStyle("note", parent=styles["Normal"], fontSize=8.5,
                          textColor=colors.HexColor(f"#{_GREY}"))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.7 * inch,
                            bottomMargin=0.7 * inch, leftMargin=0.7 * inch,
                            rightMargin=0.7 * inch,
                            title=f"Fleet Report — {agg.company_name}")
    gen = agg.generated_at.strftime("%Y-%m-%d %H:%M UTC")
    story = [
        Paragraph(f"All-Time Fleet Report — {agg.company_name}", h),
        Paragraph(
            f"Generated {gen} &nbsp;·&nbsp; Coverage: {_date_range_label(agg)} "
            f"&nbsp;·&nbsp; {agg.array_count} array(s) &nbsp;·&nbsp; "
            f"{agg.bill_count} bill(s)", sub),
        Spacer(1, 10),
    ]

    if not agg.has_data:
        story.append(Paragraph("No fleet data absorbed yet.", sec))
        story.append(Paragraph(
            "Connect a utility login or upload daily generation and your "
            "all-time totals will appear here automatically.", note))
        doc.build(story)
        return buf.getvalue()

    def _hdr_style(extra=None):
        base = [
            ("BACKGROUND", (0, 0), (-1, 0), GREEN_DARK),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CREAM]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        if extra:
            base.extend(extra)
        return TableStyle(base)

    # ── Fleet summary ──
    story.append(Paragraph("Fleet summary", sec))
    srows = [
        ["Total generation", f"{_kwh(agg.total_generated_kwh)} kWh "
                             f"({_mwh(agg.total_generated_kwh)} MWh)"],
    ]
    if agg.has_consumption:
        srows.append(["Total consumption", f"{_kwh(agg.total_consumed_kwh)} kWh"])
    if agg.has_cost:
        net = agg.total_cost
        if net < 0:
            srows.append(["Net credit value", f"${abs(net):,.2f}"])
        else:
            srows.append(["Net billed", f"${net:,.2f}"])
        if agg.total_credit:
            srows.append(["Net-metering credit", f"${agg.total_credit:,.2f}"])
    st = Table(srows, colWidths=[2.4 * inch, 4.6 * inch])
    st.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#555555")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(st)

    # ── Generation by year ──
    story.append(Paragraph("Generation by year", sec))
    yrows = [["Year", "Generation (kWh)", "Generation (MWh)"]]
    for year in sorted(agg.by_year):
        kwh = agg.by_year[year]
        yrows.append([str(year), _kwh(kwh), _mwh(kwh)])
    yrows.append(["Total", _kwh(agg.total_generated_kwh),
                  _mwh(agg.total_generated_kwh)])
    yt = Table(yrows, colWidths=[2.0 * inch, 2.5 * inch, 2.5 * inch])
    yt.setStyle(_hdr_style([
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), CREAM),
    ]))
    story.append(yt)

    # ── Generation by month ──
    story.append(Paragraph("Generation by month", sec))
    mrows = [["Month", "Generation (kWh)", "Generation (MWh)"]]
    for (year, month) in sorted(agg.by_month):
        kwh = agg.by_month[(year, month)]
        mrows.append([f"{_MONTH_NAMES[month]} {year}", _kwh(kwh), _mwh(kwh)])
    mt = Table(mrows, colWidths=[2.0 * inch, 2.5 * inch, 2.5 * inch], repeatRows=1)
    mt.setStyle(_hdr_style())
    story.append(mt)

    # ── Per-array breakdown ──
    story.append(Paragraph("Per-array breakdown (all-time)", sec))
    arows = [["Array", "NEPOOL-GIS", "Gen (kWh)", "Gen (MWh)", "Months"]]
    for row in agg.by_array:
        arows.append([
            row["name"], row["nepool_gis_id"] or "—",
            _kwh(row["kwh"]), _mwh(row["kwh"]), str(row["months"]),
        ])
    at = Table(arows, colWidths=[2.3 * inch, 1.2 * inch, 1.3 * inch,
                                 1.2 * inch, 0.9 * inch], repeatRows=1)
    at.setStyle(_hdr_style([("ALIGN", (1, 1), (1, -1), "CENTER")]))
    story.append(at)

    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "All figures are aggregated live from your absorbed generation and "
        "billing records when this report is generated, so they always reflect "
        "the latest month. Months covered by daily generation use that data; "
        "other months use utility-bill generation.", note))

    doc.build(story)
    return buf.getvalue()


# ── public entrypoint ─────────────────────────────────────────────────────


def build_fleet_report(tenant: Tenant, fmt: str = "xlsx") -> bytes:
    """Build the all-time aggregated fleet report for `tenant`.

    Reads the DB live each call, so the output always reflects the newest
    absorbed data. `fmt` is 'xlsx' or 'pdf'. Returns the file bytes.
    """
    fmt = (fmt or "xlsx").lower()
    if fmt not in ("xlsx", "pdf"):
        raise ValueError(f"unsupported fmt {fmt!r} (expected 'xlsx' or 'pdf')")
    agg = aggregate_fleet(tenant)
    if fmt == "pdf":
        return _build_pdf(agg)
    return _build_xlsx(agg)


def report_filename(tenant: Tenant, fmt: str = "xlsx") -> str:
    """Suggested download filename, e.g. FleetReport-Acme-AllTime.xlsx."""
    company = (tenant.company_name or tenant.name or tenant.operator_name
               or "Fleet")
    safe = "".join(ch if ch.isalnum() else "-" for ch in company).strip("-")
    safe = "-".join(p for p in safe.split("-") if p) or "Fleet"
    return f"FleetReport-{safe}-AllTime.{fmt}"
