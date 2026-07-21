"""
NEPOOL Operator — workbook delivery.

Single source of truth for "build workbook + email it." Always regenerates
from solar.db, writes to a temp file, attaches via Resend, cleans up.

Phase-1 expansion (June 2026):
  Reports are now generated PER CLIENT, not per tenant. A tenant with N
  clients gets N separate emails (one workbook per client, each delivered
  to that client's own contact_email + cc_emails).

Used by:
  - /v1/account/send-report (customer self-serve; sends ALL clients)
  - /admin/tenants/{tid}/deliver (ops force-send all clients)
  - /admin/clients/{cid}/deliver (ops force-send one client)
  - scheduler.deliver_scheduled_reports (weekly/monthly/quarterly cron)
"""
from __future__ import annotations
import logging
import pathlib
import tempfile
from datetime import date
from typing import Optional

from sqlalchemy import select, func

from .db import SessionLocal
from .models import Tenant, Client, Array, now
from .writers import build_workbook
from .notify import send_workbook_email, send_internal_alert
from .email_templates import build_context, render_email, resolve_from_header
from .report_arrays import not_vendor_only

logger = logging.getLogger(__name__)


def _recipients_for_client(client: Client, tenant: Tenant,
                           override_to: Optional[str]) -> list[str]:
    """Resolve who should receive this client's report.
    Order of preference: override > client.contact_email > tenant.contact_email.
    Always dedupes and includes client.cc_emails when present.

    Falling back to tenant.contact_email is only safe when the client genuinely
    represents the operator themselves (e.g. solo-operator pilot pattern). For
    auto-created or partially-set-up clients, the caller should prefer the
    `to_me` send_mode or filter such clients out — otherwise the operator gets
    an extra copy per missing-contact client they never asked for.
    """
    if override_to:
        return [override_to]
    primary = client.contact_email or tenant.contact_email
    addrs: list[str] = []
    if primary:
        addrs.append(primary.strip())
    if client.cc_emails:
        for x in client.cc_emails.split(","):
            x = x.strip()
            if x and x not in addrs:
                addrs.append(x)
    return addrs


# Report sends that are PREVIEWS / internal test sends — never a billable output.
# Everything else (scheduled auto-sends "sched-*", self-serve "send now", "resend")
# is a real client send.
_NON_BILLABLE_TRIGGERS = {"sample", "ops"}


def _report_quarter(reference_date) -> tuple[int, int, str]:
    """The (year, quarter, "<year>-Q<quarter>") this report/output covers — the
    headline rolling quarter (last complete quarter relative to the reference)."""
    from .writers.gmcs_writer import _rolling_quarters, default_reporting_reference_date
    ref = reference_date if reference_date is not None \
        else default_reporting_reference_date(date.today())
    y, q = _rolling_quarters(ref)[-1]
    return y, q, f"{y}-Q{q}"


def record_genreport_output(tenant_id: str, client_id: int, *,
                            reference_date=None, first_source: str = "download",
                            require_data: bool = True) -> int:
    """Record the $15 billable OUTPUT for each ARRAY this client reports — THE FOLD.

    Ford's model: building + previewing is free; the $15 fires on the FIRST real
    OUTPUT — a report SEND or a DOWNLOAD of the deliverable — then unlimited that
    quarter. The UNIT IS THE ARRAY, not the client (Ford 2026-07-16): one output of a
    5-array client = 5 rows = $75.

    We bill exactly the arrays that RENDER in the workbook (gmcs_writer.
    reported_array_ids: non-excluded, non-deleted, actually producing in the window) —
    so an empty/preview output bills nothing, and a force-hidden or non-producing array
    is never charged. IDEMPOTENT per (tenant, array, quarter): the DB unique constraint
    makes any later output covering the same array+quarter a no-op (insert-or-ignore,
    catching the IntegrityError like the capture upsert); a DIFFERENT quarter is fresh.

    Only records for a REAL client (live, active, this tenant) in a reports-world,
    non-demo tenant. Returns the COUNT of NEW ledger rows (0 = nothing newly billable;
    still falsy, so existing truthiness call sites keep working). INERT money-wise
    (just ledger rows; the Stripe push is a separate, price-gated job). Never raises —
    a billing-ledger hiccup must not fail (or unsend) the actual output.

    `require_data` is retained for call-site compatibility: array-level producing-ness
    is now inherent to reported_array_ids, so this is always effectively enforced.
    """
    try:
        from sqlalchemy.exc import IntegrityError
        from .models import GenReportCharge
        from .pricing_ao_genreports import PRICE_CENTS
        from .writers.gmcs_writer import reported_array_ids

        _y, _q, quarter = _report_quarter(reference_date)
        with SessionLocal() as db:
            tenant = db.get(Tenant, tenant_id)
            # Never bill the shared demo tenant or a dead account. NOTE: we do NOT
            # require tenant.generation_reports here — generation reports are
            # enabled for EVERY operator (Ford, Jul 2026), and TAKING A REAL OUTPUT
            # *is* the engagement that bills. Gating on the marker would mean an
            # operator who never explicitly enrolled could report all quarter for
            # free. Enrollment instead FOLLOWS the output (auto-enroll below).
            if tenant is None or getattr(tenant, "is_demo", False):
                return 0
            if not (tenant.active
                    or getattr(tenant, "subscription_status", None)
                    in ("comped", "trialing")):
                return 0
            client = db.get(Client, client_id)
            if (client is None or client.tenant_id != tenant_id
                    or not client.active or client.deleted_at is not None):
                return 0

        # The arrays that actually render in this client's workbook = the billing
        # units. Empty (no producing arrays / preview) → nothing extracted, no charge.
        array_ids = reported_array_ids(client_id, reference_date=reference_date)
        if not array_ids:
            return 0

        new_rows = 0
        for arr_id in array_ids:
            with SessionLocal() as db:
                # Idempotency (fast path): already billed this array+quarter?
                already = db.execute(
                    select(GenReportCharge.id).where(
                        GenReportCharge.tenant_id == tenant_id,
                        GenReportCharge.array_id == arr_id,
                        GenReportCharge.quarter == quarter,
                    ).limit(1)
                ).first()
                if already:
                    continue
                db.add(GenReportCharge(
                    tenant_id=tenant_id, array_id=arr_id, client_id=client_id,
                    quarter=quarter, amount_cents=PRICE_CENTS,
                    first_source=first_source,
                ))
                try:
                    db.commit()
                    new_rows += 1
                except IntegrityError:
                    # A concurrent output inserted the same (tenant, array, quarter)
                    # first — the unique constraint held; treat as already charged.
                    db.rollback()

        # AUTO-ENROLL on first real engagement: an operator who has actually
        # reported an array IS a generation-reports customer, so turn their
        # reports world on (scheduled sends for the clients they enrol, digests,
        # the operator directory). Enrollment follows the output rather than
        # gating it — see the note above. Idempotent + never fatal.
        if new_rows:
            try:
                with SessionLocal() as db:
                    t = db.get(Tenant, tenant_id)
                    if t is not None and not getattr(t, "generation_reports", False):
                        t.generation_reports = True
                        db.commit()
            except Exception:  # noqa: BLE001
                logger.warning("genreport auto-enroll failed for %s", tenant_id,
                               exc_info=True)
        return new_rows
    except Exception:  # noqa: BLE001 — a billing-ledger failure must never break output
        logger.warning("genreport output ledger write failed for client %s (%s)",
                       client_id, tenant_id, exc_info=True)
        return 0


def record_genreport_directory(tenant_id: str, *, reference_date=None,
                               client_ids=None, first_source: str = "directory") -> int:
    """Record a $15 OUTPUT for EVERY ARRAY the tenant's active report clients report
    this quarter — the value an ALL-CLIENTS directory download extracts.

    Downloading the directory delivers every client's arrays at once, so each reported
    array is a billable output for the quarter (idempotent per array — a re-download
    charges nothing new). Optional `client_ids` scopes it to a subset (e.g. a picker).
    Returns the count of NEW ARRAY rows. Never raises.
    """
    try:
        with SessionLocal() as db:
            q = select(Client.id).where(
                Client.tenant_id == tenant_id,
                Client.active == True,  # noqa: E712
                Client.deleted_at.is_(None),
            )
            if client_ids:
                q = q.where(Client.id.in_(list(client_ids)))
            cids = [r[0] for r in db.execute(q).all()]
        n = 0
        for cid in cids:
            n += record_genreport_output(tenant_id, cid, reference_date=reference_date,
                                         first_source=first_source)
        return n
    except Exception:  # noqa: BLE001 — a directory-billing hiccup must not fail download
        logger.warning("genreport directory ledger write failed for tenant %s",
                       tenant_id, exc_info=True)
        return 0


def deliver_for_client(client_id: int, *, year: Optional[int] = None,
                       override_to: Optional[str] = None,
                       triggered_by: str = "manual",
                       subject_prefix: str = "",
                       skip_if_empty: bool = False,
                       reference_date: Optional[date] = None) -> dict:
    """Build & email the workbook for ONE client.

    skip_if_empty=True makes an automatic/scheduled send refuse to mail a
    workbook that would render with NO generation data (a client with arrays but
    no bills/daily readings, or an empty onboarding stub). The scheduler passes
    this so the cron never emails a blank report — the same judgment an operator
    applies by hand when they only send reports that have real numbers. Explicit
    operator sends (the dashboard buttons) leave it False so a forced send always
    goes through.

    reference_date: when set (from a dashboard quarter picker), the rolling
    workbook window ends at the complete quarter before this date — same
    semantics as GET /report.xlsx?quarter=. Email merge tags ({{quarter}},
    {{period_start}}, {{period_end}}) use the same reference.
    """
    with SessionLocal() as db:
        client = db.get(Client, client_id)
        if not client:
            raise ValueError(f"Client not found: {client_id}")
        tenant = db.get(Tenant, client.tenant_id)
        if not tenant:
            raise ValueError(f"Tenant not found for client {client_id}")
        is_active = (tenant.active or tenant.subscription_status in ("comped", "trialing")) and client.active
        client_name = client.name
        # Split-name (Jun 2026): company_name on From: header & body
        # ({{tenant_name}}); operator_name on signoff.
        tenant_name = tenant.company_name or tenant.name
        operator_name = tenant.operator_name or tenant.company_name or tenant.name
        tenant_id = tenant.id
        tenant_email = (tenant.contact_email or "").strip()
        # V2 email customization snapshot (session closes below).
        # Legacy: if no send_mode is set but cc_on_reports was true, treat as to_both.
        _raw_mode = (tenant.send_mode or "").strip()
        if not _raw_mode:
            send_mode = "to_both" if bool(tenant.cc_on_reports) else "to_client"
        else:
            send_mode = _raw_mode
        # Use the tenant's explicitly-set send_from_email if configured; otherwise
        # fall back to their contact_email so reports appear to come FROM them,
        # not from admin@solaroperator.org.  send_workbook_email will retry with
        # the platform default + Reply-To if Resend rejects an unverified domain.
        effective_from_email = tenant.send_from_email or tenant.contact_email
        send_from_name = tenant.send_from_name
        from_header = resolve_from_header(
            effective_from_email, send_from_name, tenant_name)
        # Permanent templates + any active Energy Agent schedule/one-shot override
        try:
            from . import email_copy_overrides as _eco
            _resolved = _eco.resolve_templates(
                db, tenant, "generation_report",
                scope_kind="client", scope_id=str(client_id) if client_id else None,
            )
            subject_template = _resolved.get("subject_template")
            body_template = _eco.apply_body_append(
                _resolved.get("body_template"), _resolved.get("body_append"),
                html=True,
            )
            signoff_template = _resolved.get("signoff") or tenant.email_signoff
            _email_copy_override_id = _resolved.get("override_id")
        except Exception:  # noqa: BLE001 — never block a report send on override bugs
            logger.exception("email_copy_override resolve failed; using permanent templates")
            subject_template = tenant.email_subject_template
            body_template = tenant.email_body_template
            signoff_template = tenant.email_signoff
            _email_copy_override_id = None
        # Provisional count for the {{arrays_count}} merge tag (not soft-
        # deleted, not excluded). After the workbook is built this is replaced
        # by the attachment's real sheet count, since the writers also omit
        # non-producing arrays; this DB count is the fallback if that fails.
        arrays_count = db.execute(
            select(func.count()).select_from(Array)
            .where(
                Array.client_id == client_id,
                Array.deleted_at.is_(None),
                Array.excluded.is_(False),
                not_vendor_only(),
            )
        ).scalar() or 0

    if not is_active and triggered_by != "ops":
        return {"ok": False, "reason": "tenant or client inactive",
                "client_id": client_id, "client_name": client_name,
                "recipient": "", "tenant": tenant_id}

    # Automatic/scheduled sends must never email a blank workbook. If this
    # client would render with no generation data in the reporting window, skip
    # it and surface why — instead of mailing a zero-filled report to a client
    # (the exact "bogus report" failure). Explicit operator sends pass
    # skip_if_empty=False so a deliberate force-send always goes through.
    if skip_if_empty:
        from .writers.gmcs_writer import report_has_data
        if not report_has_data(client_id):
            logger.info(
                "skip empty report: client %s (%s) tenant %s — no generation "
                "data in window, not auto-sending",
                client_id, client_name, tenant_id,
            )
            return {"ok": False, "reason": "no generation data — skipped",
                    "client_id": client_id, "client_name": client_name,
                    "recipient": "", "tenant": tenant_id, "skipped_empty": True}

    # Resolve recipients, honoring send_mode. An explicit override_to (ops
    # force-send) always wins and ignores send_mode.
    if override_to:
        recipients = [override_to]
    elif send_mode == "to_me":
        # Tenant wants the workbook themselves to forward under their own name.
        # Client + client CCs are intentionally NOT contacted.
        recipients = [tenant_email] if tenant_email else []
    elif send_mode == "to_both":
        # Client gets their copy; tenant gets a separate email below.
        recipients = _recipients_for_client(client, tenant, override_to)
    else:
        # send_mode == "to_client" — only fan out if the client has its own
        # contact email, UNLESS cc_on_reports is on (legacy solo-operator mode where
        # the tenant IS the client and tenant.contact_email is the correct recipient).
        # Without cc_on_reports, silently routing to tenant.contact_email would spam
        # a NEPOOL-agent operator once per under-configured client.
        if not (client.contact_email or "").strip() and not bool(tenant.cc_on_reports):
            return {"ok": False, "reason": "no recipient email on file",
                    "client_id": client_id, "client_name": client_name,
                    "recipient": "", "tenant": tenant_id}
        recipients = _recipients_for_client(client, tenant, override_to)
    if not recipients:
        return {"ok": False, "reason": "no recipient email on file",
                "client_id": client_id, "client_name": client_name,
                "recipient": "", "tenant": tenant_id}

    safe_client = client_name.replace(" ", "_").replace("/", "-")
    with tempfile.TemporaryDirectory(prefix=f"so-deliver-c{client_id}-") as tmpdir:
        # Neutral filename — "GMCS" is the solar report's name, but a wind/hydro
        # client gets a fuel-correct REC workbook, so don't brand it solar.
        out_path = pathlib.Path(tmpdir) / f"{safe_client}-report.xlsx"
        try:
            path = build_workbook(
                client_id=client_id, year=year, out_path=out_path,
                reference_date=reference_date,
            )
        except Exception as e:
            logger.exception("Workbook build failed for client %s", client_id)
            send_internal_alert(
                "Workbook generation failed",
                f"Client: {client_id} ({client_name})\n"
                f"Tenant: {tenant_id} ({tenant_name})\n"
                f"Triggered by: {triggered_by}\nError: {e}",
            )
            return {"ok": False, "client_id": client_id,
                    "client_name": client_name, "recipient": "",
                    "reason": "report generation failed", "error": str(e)}

        # The writers omit non-producing arrays, so the DB array count can
        # overstate what the client sees. The attachment's real sheet list is
        # the ground truth for {{arrays_count}}.
        try:
            from openpyxl import load_workbook
            _wb = load_workbook(path, read_only=True)
            _sheets = [s for s in _wb.sheetnames if s != "(no data)"]
            _wb.close()
            arrays_count = len(_sheets)
        except Exception:
            logger.warning("could not recount arrays from workbook %s", path,
                           exc_info=True)

        # Render the email from the tenant's templates (or built-in defaults).
        # Merge tags resolve against this client's real name / array count /
        # headline quarter; see api/email_templates.py.
        # Email period tags use the same reference_date as the workbook so a
        # selected quarter (e.g. Q1-2026) is honest in both attachment + body.
        email_ref = reference_date
        if email_ref is None:
            from .writers.gmcs_writer import default_reporting_reference_date
            email_ref = default_reporting_reference_date(date.today())
        ctx = build_context(
            client_name=client_name, tenant_name=tenant_name,
            arrays_count=arrays_count, tenant_email=tenant_email,
            ref=email_ref,
            signoff_template=signoff_template,
            tenant_signoff_name=send_from_name or operator_name,
        )
        subject, html, text = render_email(
            subject_template=subject_template,
            body_template=body_template, ctx=ctx,
        )
        if subject_prefix:
            subject = f"{subject_prefix}{subject}"
        # Put the headline quarter in the attachment name so a client who saves
        # several reports can tell them apart (Ford 2026-07-16). Same quarter the
        # workbook + email body use (email_ref), e.g. GMCS_Only-Q1-2026-report.xlsx.
        from .writers.gmcs_writer import _rolling_quarters
        _ly, _lq = _rolling_quarters(email_ref)[-1]
        filename = f"{safe_client}-Q{_lq}-{_ly}-report.xlsx"
        # Send to primary; cc the extras. from_header carries the tenant's
        # "send as me" address (send_workbook_email falls back to the platform
        # default if Resend rejects an unverified domain).
        primary, *cc = recipients
        sent = send_workbook_email(
            to=primary, subject=subject, html=html, text=text,
            workbook_path=str(path), filename=filename, from_addr=from_header,
        )
        # If there are CCs, send a copy to each (Resend's SDK doesn't
        # always expose CC; explicit sends are safer for delivery tracking).
        for extra in cc:
            send_workbook_email(
                to=extra, subject=f"[copy] {subject}", html=html, text=text,
                workbook_path=str(path), filename=filename, from_addr=from_header,
            )

        # Operator copy: send a separate email to the operator when:
        #   (a) send_mode is "to_both", OR
        #   (b) the legacy cc_on_reports flag is on (additive — works with any send_mode).
        # Skip when the operator address is already in recipients, and on override sends.
        if (tenant_email and not override_to and tenant_email not in recipients
                and (send_mode == "to_both" or bool(tenant.cc_on_reports))):
            send_workbook_email(
                to=tenant_email, subject=f"[copy] {subject}", html=html, text=text,
                workbook_path=str(path), filename=filename, from_addr=from_header,
            )

    if sent and not override_to:
        with SessionLocal() as db:
            c = db.get(Client, client_id)
            if c:
                c.last_delivery_at = now()
            t = db.get(Tenant, tenant_id)
            if t:
                t.last_delivery_at = now()
            # One-shot / max_sends email-copy overrides expire after a real send.
            try:
                from . import email_copy_overrides as _eco
                _eco.record_send(db, locals().get("_email_copy_override_id"))
            except Exception:  # noqa: BLE001
                logger.exception("email_copy_override record_send failed")
            db.commit()

    # Metered billing ($15/client/quarter): a report actually went out for this
    # client, so record the billable OUTPUT for the quarter the workbook covers
    # (email_ref — same quarter as the filename/body). Idempotent per client+quarter
    # (a send is free if a download already billed the quarter, and vice-versa), and
    # inert until the metered price is minted; previews/ops sends never charge.
    if sent and triggered_by not in _NON_BILLABLE_TRIGGERS:
        record_genreport_output(
            tenant_id, client_id, reference_date=email_ref, first_source="send")

    return {
        "ok": True,
        "client_id": client_id,
        "client_name": client_name,
        "tenant": tenant_id,
        "email_sent": sent,
        "recipient": recipients[0] if recipients else "",
        "recipients": recipients,
        "triggered_by": triggered_by,
        "email_copy_override_id": locals().get("_email_copy_override_id"),
    }


def deliver_operator_directory(
    tenant_id: str,
    *,
    client_ids: Optional[list[int]] = None,
    reference_date: Optional[date] = None,
    triggered_by: str = "manual",
) -> dict:
    """Email the operator a single NEPOOL-GIS directory workbook covering
    every (selected) client's arrays — same GMCS form as client reports, all
    sheets in one file for bulk upload to the NEPOOL site.

    Always goes ONLY to the tenant's contact_email (never to clients).
    Best-effort: failures are logged and returned; callers should not fail
    client delivery because the directory email missed.
    """
    from .report_eligibility import tenant_has_report_clients, tenant_in_reports_world

    with SessionLocal() as db:
        tenant = db.get(Tenant, tenant_id)
        if not tenant:
            return {"ok": False, "reason": "tenant not found", "tenant": tenant_id}
        # THE FOLD: the directory goes to any tenant whose generation-reports
        # world is live (every NEPOOL tenant; an AO tenant once the fold
        # migration flips generation_reports) and who has live report clients.
        # NOT bare client-presence — AO capture auto-creates Client rows, so
        # that alone can't identify a reports tenant. Demo never emails anyone.
        if getattr(tenant, "is_demo", False):
            return {"ok": False, "reason": "demo tenant", "tenant": tenant_id}
        if not tenant_in_reports_world(tenant):
            return {"ok": False, "reason": "generation reports not enabled",
                    "tenant": tenant_id}
        if not tenant_has_report_clients(db, tenant_id):
            return {"ok": False, "reason": "no report clients", "tenant": tenant_id}
        tenant_email = (tenant.contact_email or "").strip()
        tenant_name = tenant.company_name or tenant.name or "Operator"
        operator_name = tenant.operator_name or tenant_name
        if not tenant_email:
            return {"ok": False, "reason": "no operator email on file",
                    "tenant": tenant_id}

    from .writers.gmcs_writer import (
        build_directory_workbook,
        default_reporting_reference_date,
    )
    from .email_templates import quarter_context

    ref = reference_date if reference_date is not None \
        else default_reporting_reference_date(date.today())
    qc = quarter_context(ref)
    quarter_label = qc.get("quarter") or "current window"

    try:
        with tempfile.TemporaryDirectory(prefix=f"so-dir-{tenant_id[:8]}-") as tmpdir:
            out = pathlib.Path(tmpdir) / "NEPOOL-directory.xlsx"
            path = build_directory_workbook(
                tenant_id,
                client_ids=client_ids,
                reference_date=ref,
                out_path=out,
            )
            # Count real array sheets (exclude empty stub)
            sheet_count = 0
            try:
                from openpyxl import load_workbook
                _wb = load_workbook(path, read_only=True)
                sheet_count = sum(1 for s in _wb.sheetnames if s != "(no data)")
                _wb.close()
            except Exception:
                sheet_count = 0

            if sheet_count == 0:
                return {
                    "ok": False,
                    "reason": "no generation data in directory window",
                    "tenant": tenant_id,
                    "recipient": tenant_email,
                    "sheet_count": 0,
                }

            subject = (
                f"Your NEPOOL-GIS directory — {quarter_label} "
                f"({sheet_count} array{'s' if sheet_count != 1 else ''})"
            )
            html = (
                f"<p>Hi {operator_name},</p>"
                f"<p>Client reports for <b>{quarter_label}</b> have been processed. "
                f"Attached is your <b>NEPOOL-GIS directory</b> — one sheet per array "
                f"across all your clients, in the same NEPOOL report form, so you can "
                f"upload the whole book to the NEPOOL-GIS site.</p>"
                f"<p><b>{sheet_count}</b> array sheet"
                f"{'s' if sheet_count != 1 else ''} included "
                f"(arrays with no generation in the window are omitted).</p>"
                f"<p>Each sheet title is <code>Client — Array</code>.</p>"
            )
            text = (
                f"Hi {operator_name},\n\n"
                f"Client reports for {quarter_label} have been processed. "
                f"Attached is your NEPOOL-GIS directory ({sheet_count} arrays) "
                f"for upload to the NEPOOL-GIS site.\n"
            )
            filename = f"NEPOOL-directory-{quarter_label.replace(' ', '-')}.xlsx"
            sent = send_workbook_email(
                to=tenant_email, subject=subject, html=html, text=text,
                workbook_path=str(path), filename=filename,
            )
            return {
                "ok": bool(sent),
                "tenant": tenant_id,
                "recipient": tenant_email,
                "sheet_count": sheet_count,
                "email_sent": bool(sent),
                "triggered_by": triggered_by,
                "quarter": quarter_label,
            }
    except Exception as e:
        logger.exception("Operator directory failed for tenant %s", tenant_id)
        return {
            "ok": False,
            "reason": str(e),
            "tenant": tenant_id,
            "recipient": tenant_email,
        }


def deliver_for_tenant(tenant_id: str, *, year: Optional[int] = None,
                       override_to: Optional[str] = None,
                       triggered_by: str = "manual",
                       skip_if_empty: bool = True,
                       reference_date: Optional[date] = None,
                       send_directory: bool = True) -> dict:
    """Build & email a workbook for EVERY active client under a tenant.

    Returns an aggregate result with one entry per client. This is the
    function called by:
      - the customer-facing "send report now" button (sends all clients)
      - the scheduler cron
      - the ops admin force-send

    skip_if_empty defaults True for this BULK fan-out: a "send all" must never
    blast blank workbooks to clients that have no generation data. Skipped
    clients come back in the result (reason "no generation data — skipped") so
    the UI can tell the operator. An ops force-send passing override_to still
    delivers per-client because override_to wins; set skip_if_empty=False to
    force-send blanks deliberately.

    After client fan-out, the operator also receives a NEPOOL-GIS directory
    workbook (all clients' arrays, one sheet each) for upload to the NEPOOL
    site — unless send_directory=False.
    """
    with SessionLocal() as db:
        tenant = db.get(Tenant, tenant_id)
        if not tenant:
            raise ValueError(f"Tenant not found: {tenant_id}")
        is_active = tenant.active or tenant.subscription_status in (
            "comped", "trialing")
        if not is_active and triggered_by != "ops":
            return {"ok": False, "reason": "tenant not active",
                    "tenant": tenant_id, "results": []}
        clients = db.execute(
            select(Client).where(Client.tenant_id == tenant_id,
                                 Client.active == True)  # noqa: E712
        ).scalars().all()
        client_ids = [c.id for c in clients]

    results = []
    # An override_to send (ops force-send to a single address) is a deliberate
    # request to deliver every client regardless of data — don't skip those.
    effective_skip = skip_if_empty and override_to is None
    for cid in client_ids:
        try:
            results.append(deliver_for_client(
                cid, year=year, override_to=override_to,
                triggered_by=triggered_by, skip_if_empty=effective_skip,
                reference_date=reference_date))
        except Exception as e:
            logger.exception("Delivery failed for client %s", cid)
            results.append({"ok": False, "client_id": cid,
                            "client_name": None, "recipient": "",
                            "reason": "unexpected error", "error": str(e)})
    ok_count = sum(1 for r in results if r.get("ok"))
    skipped_empty = [r.get("client_id") for r in results if r.get("skipped_empty")]

    directory = None
    # Send the operator directory whenever at least one client report was
    # generated successfully (or even if all skipped — still useful if some
    # arrays have data under other clients). Always try when send_directory.
    if send_directory:
        try:
            delivered_ids = [
                r.get("client_id") for r in results
                if r.get("ok") and r.get("client_id") is not None
            ]
            # Prefer arrays from clients we actually mailed; fall back to all.
            dir_ids = delivered_ids if delivered_ids else client_ids
            directory = deliver_operator_directory(
                tenant_id,
                client_ids=dir_ids or None,
                reference_date=reference_date,
                triggered_by=f"{triggered_by}-directory",
            )
        except Exception as e:  # noqa: BLE001
            logger.warning("directory send failed for %s: %s", tenant_id, e)
            directory = {"ok": False, "reason": str(e), "tenant": tenant_id}

    return {
        "ok": ok_count > 0,
        "tenant": tenant_id,
        "client_count": len(client_ids),
        "delivered": ok_count,
        "skipped_empty": skipped_empty,
        "results": results,
        "triggered_by": triggered_by,
        "directory": directory,
    }
