"""Array Operator — OPERATOR/TENANT-level generation-spreadsheet tracker.

Mounted under /v1/array-operator/tracker.

THE HEADLINE FEATURE — the AUTO-MAINTAINED PER-ARRAY WORKBOOK. For a tenant we
build, fresh on every fetch/download, a spreadsheet with:

  * one ROW per BILLING MONTH (YYYY-MM, chronological), and
  * one COLUMN per ARRAY (header = the array's name), plus a trailing `Total`,

each cell = that array's MEASURED generation kWh for that month. Because it is
recomputed from current generation data on access, a newly-added array simply
appears as a new column and a new month simply appears as a new row — the
"auto-append a row per month" behaviour with none of a cron's fragility, and it
is always correct.

DATA HONESTY: the per-array monthly sums are MEASURED only — `bill_prorate`
(a monthly utility bill smeared flat across its days) is an estimate and is
EXCLUDED from every cell, the same rule jobs/usage_report.tenant_period_kwh and
the dashboard use. Soft-deleted / excluded arrays are left out too.

This surface is ADDITIVE: it only READS generation data (DailyGeneration). It
never reads or writes any billing computation, so it cannot affect the live
invoice/billing path.

The operator may ALSO upload their own existing spreadsheet as an OPTIONAL
OVERRIDE — we detect its structure ("our magic") with the SAME deterministic,
offline header heuristic the per-sub tracker uses (api.billing.sheet_tracker),
store it on the tenant, and serve it instead of the auto sheet until removed.
The auto sheet is the primary/default; the upload is a fallback for operators
who want their own layout.

Design notes (kept consistent with the per-sub tracker):
  * AUTH — `tenant_from_session` + `require_not_demo`, the same bearer used
    everywhere else. Mutating routes refuse the shared demo tenant.
  * GATE — the whole feature is gated behind SPREADSHEET_TRACKER_ENABLED (the
    same flag as the per-sub tracker). When OFF, GET returns {enabled:false}
    (the UI hides the card) and the mutating routes 404, so a half-built state
    can never disturb the live page. When ON, GET returns enabled:true for a
    signed-in non-demo operator (the card shows by default) with
    has_sheet:false until one is uploaded.
  * RESPONSE SHAPE — IDENTICAL to the per-sub tracker's `_tracker_status_dict`
    so the existing frontend renderer is reusable verbatim.
  * STORAGE — the 4 nullable Tenant.tracker_* columns (additive migration); the
    xlsx bytes + detected mapping + original filename + updated_at.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Header, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select

from .db import SessionLocal
from .models import Array, DailyGeneration, Tenant
from .account import tenant_from_session, require_not_demo

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/v1/array-operator/tracker",
                   tags=["array-operator-tracker"])

# Mirror the per-sub tracker's limits / type guards (api/billing/routes.py).
MAX_UPLOAD_BYTES = 8 * 1024 * 1024  # 8 MB — these sheets are tens of KB
_MAGIC_XLSX = b"PK\x03\x04"          # ZIP / OpenXML (.xlsx, .xlsm, …)
_MAGIC_XLS = b"\xd0\xcf\x11\xe0"     # OLE2 compound doc (.xls, …)
_XLSX_MEDIA = ("application/vnd.openxmlformats-officedocument"
               ".spreadsheetml.sheet")

# Estimate provenance to EXCLUDE from measured cells (data-honesty rule shared
# with jobs/usage_report.tenant_period_kwh + the dashboard). NULL source (legacy
# rows) is kept via coalesce, so only an explicit bill_prorate is dropped.
_ESTIMATE_SOURCE = "bill_prorate"


# ─── the auto-maintained per-array workbook ──────────────────────────────────

def _array_month_grid(db, tenant_id: str) -> dict:
    """Compute the tenant's per-array × per-month MEASURED-kWh grid, fresh from
    DailyGeneration. Returns a dict:

        {
          "arrays":  [(array_id, "Name"), ...]   # column order, by name then id
          "periods": ["2026-03", "2026-04", ...] # row order, chronological
          "cells":   {(array_id, "YYYY-MM"): kwh_float},
        }

    HONESTY: excludes source='bill_prorate' (an estimate) and soft-deleted /
    excluded arrays — identical to the billed-quantity rule used elsewhere. An
    array with zero measured rows still gets a column (so the operator sees it
    exists); a month appears as a row only if SOME array measured kWh in it.
    """
    # Arrays that should appear as columns — live, non-excluded, this tenant.
    # Ordered by name (case-insensitive) then id so columns are stable and a
    # newly-added array slots in deterministically.
    arr_rows = db.execute(
        select(Array.id, Array.name)
        .where(Array.tenant_id == tenant_id,
               Array.deleted_at.is_(None),
               Array.excluded.is_(False))
    ).all()
    arrays = sorted(
        ((aid, (name or f"Array {aid}")) for aid, name in arr_rows),
        key=lambda t: ((t[1] or "").lower(), t[0]),
    )
    array_ids = [aid for aid, _ in arrays]
    out = {"arrays": arrays, "periods": [], "cells": {}}
    if not array_ids:
        return out

    # SUM(kwh) GROUP BY array, calendar-month — measured rows only. We bucket by
    # YYYY-MM from the row's `day` (calendar month = the period the rest of the
    # product labels invoices/reports with). func.strftime is sqlite; for other
    # backends we fall back to Python bucketing below.
    cells: dict[tuple[int, str], float] = {}
    periods: set[str] = set()
    rows = db.execute(
        select(DailyGeneration.array_id, DailyGeneration.day, DailyGeneration.kwh)
        .where(DailyGeneration.array_id.in_(array_ids),
               func.coalesce(DailyGeneration.source, "") != _ESTIMATE_SOURCE)
    ).all()
    for aid, day, kwh in rows:
        if day is None:
            continue
        period = f"{day.year:04d}-{day.month:02d}"
        periods.add(period)
        cells[(aid, period)] = cells.get((aid, period), 0.0) + float(kwh or 0.0)

    out["periods"] = sorted(periods)
    out["cells"] = cells
    return out


def _auto_headers(grid: dict) -> list[str]:
    """Header row for the auto workbook: Period, <array names…>, Total."""
    return ["Period"] + [name for _, name in grid["arrays"]] + ["Total"]


def _build_auto_workbook(grid: dict) -> bytes:
    """Render the per-array × per-month grid to a styled xlsx and return bytes.
    First column Period, one column per array, trailing Total (row sum). Each
    cell is that array's measured kWh for that month (rounded to 1 dp); blank
    when an array had no measured generation that month."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Generation by array"

    headers = _auto_headers(grid)
    ws.append(headers)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E5F")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    arrays = grid["arrays"]
    for period in grid["periods"]:
        row = [period]
        total = 0.0
        for aid, _ in arrays:
            v = grid["cells"].get((aid, period))
            if v is None:
                row.append(None)
            else:
                rv = round(float(v), 1)
                row.append(rv)
                total += rv
        row.append(round(total, 1))
        ws.append(row)

    # Column widths: Period + names sized to header, kWh columns right-aligned.
    ws.column_dimensions["A"].width = 12
    for idx, name in enumerate(h for h in headers[1:]):
        col_letter = ws.cell(row=1, column=idx + 2).column_letter
        ws.column_dimensions[col_letter].width = max(12, min(28, len(name) + 3))
    ws.freeze_panes = "B2"  # keep Period column + header visible while scrolling

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _auto_status_dict(grid: dict, updated_at: datetime) -> dict:
    """Tracker card state describing the AUTO sheet. Backward-compatible with the
    per-sub `_tracker_status_dict` shape the frontend already consumes
    (enabled / has_sheet / headers / columns / filename / last_period /
    updated_at / warnings), plus an `auto:true` flag and a `data_rows` count so
    the UI can say "auto-generated, N months"."""
    headers = _auto_headers(grid)
    periods = grid["periods"]
    n_arrays = len(grid["arrays"])
    warnings: list[str] = []
    if n_arrays == 0:
        warnings.append("No arrays yet — add an array (or connect the "
                        "extension) and its column will appear automatically.")
    elif not periods:
        warnings.append("No measured generation yet — months will appear as "
                        "soon as generation data lands.")
    return {
        "enabled": True,
        "auto": True,
        "has_sheet": True,  # we always generate it
        "filename": "generation_by_array.xlsx",
        # columns: name -> 0-based index, mirroring the upload detector's map so
        # the existing renderer (which reads columns) keeps working.
        "columns": {h: i for i, h in enumerate(headers)},
        "headers": headers,
        "header_row": 1,
        "sheet": "Generation by array",
        "data_rows": len(periods),
        "last_period": periods[-1] if periods else None,
        "updated_at": updated_at.isoformat() + "Z",
        "warnings": warnings,
    }


def _status_dict(t: Tenant) -> dict:
    """Tenant-level tracker card state. Byte-for-byte the SAME shape the per-sub
    tracker returns (`_tracker_status_dict`) so the frontend renderer is reused.
    Honest about whether a sheet is attached + what we detected."""
    m = getattr(t, "tracker_map", None) or {}
    has = bool(getattr(t, "tracker_workbook", None)) and bool(m.get("ok"))
    up = getattr(t, "tracker_updated_at", None)
    return {
        "enabled": True,
        "has_sheet": has,
        "filename": getattr(t, "tracker_filename", None),
        "columns": m.get("columns") if has else None,
        "headers": m.get("headers") if has else None,
        "header_row": m.get("header_row") if has else None,
        "sheet": m.get("sheet") if has else None,
        "data_rows": m.get("data_rows") if has else None,
        "last_period": m.get("last_period") if has else None,
        "updated_at": up.isoformat() + "Z" if up else None,
        "warnings": m.get("warnings") or [],
    }


def _get_tenant(db, tenant_id: str) -> Tenant:
    t = db.get(Tenant, tenant_id)
    if t is None:
        raise HTTPException(404, "Tenant not found")
    return t


def _has_uploaded_override(t: Tenant) -> bool:
    """True when the operator uploaded their OWN sheet (the optional override).
    When present it takes precedence over the auto sheet."""
    m = getattr(t, "tracker_map", None) or {}
    return bool(getattr(t, "tracker_workbook", None)) and bool(m.get("ok"))


@router.get("")
def tracker_status(authorization: Optional[str] = Header(default=None)):
    """Operator-level tracker state (drives the card). Returns {enabled:false}
    when the feature flag is off so the UI hides. Otherwise the card reflects
    the AUTO-MAINTAINED per-array sheet (auto:true, has_sheet:true, headers =
    [Period, <array names…>, Total], data_rows = #months) — UNLESS the operator
    uploaded their own sheet, in which case that override's detected mapping is
    reflected instead (auto:false)."""
    from .billing.sheet_tracker import tracker_enabled
    t = tenant_from_session(authorization)
    if not tracker_enabled():
        return {"ok": True, "tracker": {"enabled": False}}
    with SessionLocal() as db:
        ten = _get_tenant(db, t.id)
        if _has_uploaded_override(ten):
            tr = _status_dict(ten)
            tr["auto"] = False
            return {"ok": True, "tracker": tr}
        grid = _array_month_grid(db, ten.id)
        return {"ok": True,
                "tracker": _auto_status_dict(grid, datetime.utcnow())}


@router.post("")
async def tracker_upload(file: UploadFile = File(...),
                         authorization: Optional[str] = Header(default=None)):
    """Upload the operator's existing generation spreadsheet (XLSX or CSV) as an
    OPTIONAL OVERRIDE of the auto-maintained per-array sheet. We detect its
    structure ('our magic'), normalize to xlsx, and store it on the tenant; once
    present it is served instead of the auto sheet (until DELETEd). Returns the
    detected mapping for review — same shape as GET (auto:false)."""
    from .billing.sheet_tracker import tracker_enabled, ingest_upload
    t = tenant_from_session(authorization)
    require_not_demo(t)
    if not tracker_enabled():
        raise HTTPException(404, "Spreadsheet tracker is not enabled.")
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Empty file.")
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "File too large (8 MB max).")
    name = file.filename or "generation.xlsx"
    is_x = raw[:4] in (_MAGIC_XLSX, _MAGIC_XLS)
    is_csv = name.lower().endswith(".csv") or (not is_x)
    if not is_x and not is_csv:
        raise HTTPException(415, "Upload an .xlsx or .csv generation sheet.")
    res = ingest_upload(raw, name)
    if not res.get("ok"):
        warn = "; ".join(res.get("warnings") or []) or "Couldn't read that sheet."
        raise HTTPException(422, warn)
    with SessionLocal() as db:
        ten = _get_tenant(db, t.id)
        ten.tracker_workbook = res["workbook"]
        ten.tracker_filename = name
        ten.tracker_map = res["mapping"]
        ten.tracker_updated_at = datetime.utcnow()
        db.add(ten)
        db.commit()
        db.refresh(ten)
        tr = _status_dict(ten)
        tr["auto"] = False
        return {"ok": True, "tracker": tr}


@router.get("/download")
def tracker_download(authorization: Optional[str] = Header(default=None)):
    """Stream the per-array generation workbook. By default this is the
    AUTO-MAINTAINED sheet, GENERATED FRESH from current generation data on every
    download (column per array, row per month, trailing Total) — so it is always
    current. If the operator uploaded their own override sheet, that file is
    streamed instead (exactly as stored)."""
    from .billing.sheet_tracker import tracker_enabled
    t = tenant_from_session(authorization)
    if not tracker_enabled():
        raise HTTPException(404, "Spreadsheet tracker is not enabled.")
    with SessionLocal() as db:
        ten = _get_tenant(db, t.id)
        if _has_uploaded_override(ten):
            blob = bytes(ten.tracker_workbook)
            base = getattr(ten, "tracker_filename", None) or "generation.xlsx"
            if base.lower().endswith(".csv"):
                base = base[:-4] + ".xlsx"
            elif not base.lower().endswith(".xlsx"):
                base = base + ".xlsx"
        else:
            grid = _array_month_grid(db, ten.id)
            blob = _build_auto_workbook(grid)
            tid = (ten.id or "tenant").replace("/", "_")
            base = f"generation_by_array_{tid}.xlsx"
    return StreamingResponse(
        io.BytesIO(blob), media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{base}"'})


@router.delete("")
def tracker_remove(authorization: Optional[str] = Header(default=None)):
    """Detach the operator's BYO sheet. Returns {enabled:true, has_sheet:false}."""
    from .billing.sheet_tracker import tracker_enabled
    t = tenant_from_session(authorization)
    require_not_demo(t)
    if not tracker_enabled():
        raise HTTPException(404, "Spreadsheet tracker is not enabled.")
    with SessionLocal() as db:
        ten = _get_tenant(db, t.id)
        ten.tracker_workbook = None
        ten.tracker_filename = None
        ten.tracker_map = None
        ten.tracker_updated_at = datetime.utcnow()
        db.add(ten)
        db.commit()
        db.refresh(ten)
        return {"ok": True, "tracker": _status_dict(ten)}
