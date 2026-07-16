#!/usr/bin/env python3
"""Migrate ONE NEPOOL tenant's generation-reports world onto a target Array
Operator tenant (THE FOLD, Phase 1 tooling — Phase 4 actually runs it).

What moves
----------
  * Client rows (re-pointed: tenant_id -> target; ids stay stable so
    Array.client_id / ReportDelivery.client_id links survive untouched)
  * client<->array links, per-array NEPOOL fields (nepool_gis_id, fuel_type,
    cert_registry)
  * tenant report settings: report_frequency / send_mode / cc_on_reports /
    email_subject_template / email_body_template / email_signoff are copied
    wholesale (NEPOOL-report-only semantics; AO offtaker emails use the
    separate offtaker_* columns). send_from_email / send_from_name are copied
    ONLY when the target hasn't set its own (they're shared with AO emails).
  * ReportDelivery history rows (re-pointed to the target tenant)
  * the target gains ``generation_reports=True`` — the explicit marker that
    makes an Array Operator tenant eligible for scheduled sends / digests /
    the operator directory (see api/report_eligibility.py for why this is a
    marker and not clients+cadence inference)

Array mapping (the hard part — "sibling arrays" from dual capture)
------------------------------------------------------------------
For every LIVE source array linked to a moved client, in order:
  1. exact name match (case-insensitive) against the target's LIVE arrays
     -> LINK: target array adopts the client link + NEPOOL fields; if the case
        differs the target array is RENAMED to the source's exact name so the
        report sheets stay byte-identical. The source sibling is DETACHED
        (client_id -> NULL) so the moved client can never double-count.
  2. utility account-number match ((provider, account_number) of the source
     array's accounts vs accounts linked to target live arrays)
     -> LINK + RENAME the target array to the source array's name (the name
        NEPOOL-GIS knows). Same detach rule.
  3. no confident match -> MOVE (create-on-target by re-pointing the source
     Array row wholesale: the array, its UtilityAccounts, their Bills and its
     DailyGeneration rows all move to the target tenant, so the generation
     linkage carries with it). This is the DEFAULT for unmapped arrays and is
     stated explicitly in the plan output.

  * TWO candidate matches are NEVER guessed between: the plan lists the
    ambiguity and execute/verify mode ABORTS until a human resolves it.
  * A source array with excluded=True that maps onto a target sibling is NOT
    linked (writers skip excluded arrays, so report parity holds either way;
    silently setting excluded on a live AO array would change AO billing).
    It is listed loudly for Ford's eyeball.
  * A LINK match onto a target array that is ALREADY client-linked aborts —
    UNLESS the operator explicitly claims it (below). Never inference.

Claiming capture-created client links (the Bruce-pair shape)
-------------------------------------------------------------
AO capture auto-creates a Client per utility-login holder and links every
captured array to it (e.g. Bruce's AO tenant: ONE "Green Mountain Community
Solar" capture client holding all 28 arrays). The GMP-side siblings of the
real report arrays are therefore client-LINKED, which the guard above
refuses. Two EXPLICIT operator choices unlock that:

  * --claim-linked-from <client_id>   (repeatable) A LINK/LINK-ACCOUNT match
        whose target array currently belongs to one of these client ids is
        ALLOWED: the array is re-pointed to the migrated client (same rename
        + NEPOOL-field carry + source-sibling detach semantics). A linked
        target belonging to any OTHER client still aborts. The plan prints
        "CLAIM (from client N)" per array.
  * --deactivate-client <client_id>   (repeatable, target-tenant client)
        Sets active=False in the same transaction — retires the capture
        client from the reports iteration. REFUSED if, after the claims, the
        client would still hold live report arrays (any remaining array with
        a nepool_gis_id); only a client whose remaining arrays are plain
        capture arrays may be deactivated. The plan lists it with its
        remaining-array count.

Safety
------
  * DRY-RUN by default: prints the full mapping plan, writes NOTHING.
  * --report          print the plan (same as dry-run; explicit alias).
  * --verify          run the ENTIRE migration inside one DB transaction,
                      build every moved client's report workbook from the
                      migrated target state, byte-compare against the workbook
                      generated from the source tenant pre-migration, then
                      ROLL BACK (the mutate-verify-rollback "flush proof").
                      xlsx files embed creation timestamps in docProps/core.xml;
                      when only that member differs the result is reported as
                      "content_identical" and, failing that, sheet content is
                      compared cell-by-cell via openpyxl.
                      ALSO asserts the post-migration reports iteration: the
                      target's ACTIVE clients must be exactly the migrated
                      (source-active) client set — a still-active capture
                      client fails verify — and every claimed array must sit
                      on its migrated client.
  * --execute         apply + COMMIT. Refuses while any ambiguity/conflict is
                      open. Combine with --verify to require a green byte-diff
                      in the same run before committing.
  * --yes-prod        additionally required for --execute against a Postgres
                      DATABASE_URL (Phase 4 will pass it; nothing in Phase 1
                      should).

Usage
-----
  python -m scripts.migrate_nepool_tenant --source ten_SRC --target ten_TGT
  python -m scripts.migrate_nepool_tenant --source ten_SRC --target ten_TGT --report
  python -m scripts.migrate_nepool_tenant --source ten_SRC --target ten_TGT --verify
  # the Bruce-pair shape: claim the GMP siblings off the capture client and
  # retire it from the reports iteration, all verified in one transaction:
  python -m scripts.migrate_nepool_tenant --source ten_SRC --target ten_TGT \
      --claim-linked-from 1557 --deactivate-client 1557 --verify
  python -m scripts.migrate_nepool_tenant --source ten_SRC --target ten_TGT \
      --verify --execute --yes-prod        # Phase 4 only
"""
from __future__ import annotations

import argparse
import io
import pathlib
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

from sqlalchemy import select, update, func

from api.db import SessionLocal, engine
from api.models import (Tenant, Client, Array, UtilityAccount, Bill,
                        DailyGeneration, ReportDelivery)

# Tenant settings copied wholesale — NEPOOL-report semantics only (AO offtaker
# invoice emails render from the separate offtaker_* template columns).
HARD_COPY_SETTINGS = ("report_frequency", "send_mode", "cc_on_reports",
                      "email_subject_template", "email_body_template",
                      "email_signoff")
# Shared with AO's own emails — copied only when the target hasn't set them.
SOFT_COPY_SETTINGS = ("send_from_email", "send_from_name")
ARRAY_NEPOOL_FIELDS = ("nepool_gis_id", "fuel_type", "cert_registry")


class MigrationBlocked(RuntimeError):
    """Raised in execute/verify mode when the plan has open ambiguities or
    conflicts — the tool never guesses."""


@dataclass
class ArrayAction:
    source_id: int
    source_name: str
    action: str                    # link-name | link-account | move | skip-excluded
    target_id: Optional[int] = None
    target_name: Optional[str] = None
    rename_to: Optional[str] = None
    fields: dict = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)
    # For link actions: the migrated client this array will belong to (the
    # source array's client_id — stable through the move).
    client_id: Optional[int] = None
    # Set when the target array was client-linked and the operator explicitly
    # claimed it via --claim-linked-from: the client id it is taken FROM.
    claimed_from: Optional[int] = None


@dataclass
class Plan:
    source_id: str
    target_id: str
    clients: list[dict] = field(default_factory=list)       # {id, name, active, arrays}
    arrays: list[ArrayAction] = field(default_factory=list)
    settings: list[dict] = field(default_factory=list)      # {field, source, target, action}
    deactivate: list[dict] = field(default_factory=list)    # {id, name, remaining_arrays}
    report_deliveries: int = 0
    ambiguities: list[str] = field(default_factory=list)
    conflicts: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def blockers(self) -> list[str]:
        return self.ambiguities + self.conflicts


# ───────────────────────────── plan building (read-only) ───────────────────────

def build_plan(db, source: Tenant, target: Tenant,
               allow_products: bool = False,
               claim_from: Optional[set[int]] = None,
               deactivate_clients: Optional[set[int]] = None) -> Plan:
    claim_from = set(claim_from or ())
    deactivate_clients = set(deactivate_clients or ())
    plan = Plan(source_id=source.id, target_id=target.id)

    if source.id == target.id:
        plan.conflicts.append("source and target are the same tenant")
        return plan
    if getattr(source, "is_demo", False) or getattr(target, "is_demo", False):
        plan.conflicts.append("refusing to touch a demo tenant")
    if not allow_products:
        if getattr(source, "product", "nepool") != "nepool":
            plan.conflicts.append(
                f"source {source.id} product={source.product!r} — expected 'nepool' "
                f"(--allow-products to override)")
        if getattr(target, "product", "nepool") != "array_operator":
            plan.conflicts.append(
                f"target {target.id} product={target.product!r} — expected "
                f"'array_operator' (--allow-products to override)")

    # Clients: every non-deleted source client moves (ids kept).
    src_clients = db.execute(
        select(Client).where(Client.tenant_id == source.id,
                             Client.deleted_at.is_(None))
        .order_by(Client.id)
    ).scalars().all()
    if not src_clients:
        plan.conflicts.append(f"source {source.id} has no live Client rows — nothing to migrate")
    moved_client_ids = {c.id for c in src_clients}
    # Live-array count per client so the plan can flag zero-array clients
    # (they move as-is — benign, but the operator should see it).
    src_client_array_counts = dict(db.execute(
        select(Array.client_id, func.count())
        .where(Array.tenant_id == source.id, Array.deleted_at.is_(None),
               Array.client_id.in_(moved_client_ids or {-1}))
        .group_by(Array.client_id)
    ).all()) if moved_client_ids else {}
    plan.clients = [{"id": c.id, "name": c.name, "active": bool(c.active),
                     "arrays": int(src_client_array_counts.get(c.id, 0))}
                    for c in src_clients]

    # uq_client_per_tenant spans soft-deleted rows too — any same-name client
    # already on the target blocks the re-point.
    tgt_client_names = {
        (n or "") for (n,) in db.execute(
            select(Client.name).where(Client.tenant_id == target.id)).all()
    }
    for c in src_clients:
        if c.name in tgt_client_names:
            plan.conflicts.append(
                f"client name collision on target: {c.name!r} (source client {c.id})")

    # Source arrays in scope: LIVE + linked to a moved client.
    src_arrays = db.execute(
        select(Array).where(Array.tenant_id == source.id,
                            Array.deleted_at.is_(None))
        .order_by(Array.id)
    ).scalars().all()
    in_scope = [a for a in src_arrays if a.client_id in moved_client_ids]
    for a in src_arrays:
        if a.client_id not in moved_client_ids:
            plan.notes.append(
                f"source array {a.id} {a.name!r} has no moved client — left untouched")

    tgt_arrays = db.execute(
        select(Array).where(Array.tenant_id == target.id,
                            Array.deleted_at.is_(None))
        .order_by(Array.id)
    ).scalars().all()
    tgt_by_lower: dict[str, list[Array]] = {}
    for a in tgt_arrays:
        tgt_by_lower.setdefault((a.name or "").strip().lower(), []).append(a)

    # (provider, account_number) -> [target live arrays] for account matching.
    tgt_accounts = db.execute(
        select(UtilityAccount).where(
            UtilityAccount.tenant_id == target.id,
            UtilityAccount.deleted_at.is_(None),
            UtilityAccount.array_id.isnot(None))
    ).scalars().all()
    tgt_arr_by_id = {a.id: a for a in tgt_arrays}
    tgt_by_acct: dict[tuple[str, str], list[Array]] = {}
    for ua in tgt_accounts:
        arr = tgt_arr_by_id.get(ua.array_id)
        if arr is not None:
            tgt_by_acct.setdefault((ua.provider, ua.account_number), []).append(arr)

    # ALL target account numbers (incl. soft-deleted/unlinked) — the
    # uq_account_per_tenant constraint spans them, so a MOVE collides with any.
    all_tgt_acct_keys = {
        (p, n) for (p, n) in db.execute(
            select(UtilityAccount.provider, UtilityAccount.account_number)
            .where(UtilityAccount.tenant_id == target.id)).all()
    }

    claimed_target_ids: set[int] = set()
    for a in in_scope:
        src_accts = db.execute(
            select(UtilityAccount).where(UtilityAccount.array_id == a.id)
        ).scalars().all()
        fields = {f: getattr(a, f) for f in ARRAY_NEPOOL_FIELDS}

        name_matches = tgt_by_lower.get((a.name or "").strip().lower(), [])
        acct_matches: list[Array] = []
        if not name_matches:
            seen: dict[int, Array] = {}
            for ua in src_accts:
                if ua.deleted_at is not None:
                    continue
                for cand in tgt_by_acct.get((ua.provider, ua.account_number), []):
                    seen[cand.id] = cand
            acct_matches = list(seen.values())

        matches = name_matches or acct_matches
        how = "link-name" if name_matches else "link-account"

        if len(matches) > 1:
            plan.ambiguities.append(
                f"array {a.id} {a.name!r}: {len(matches)} target candidates "
                f"({', '.join(repr(m.name) for m in matches)}) — will not guess")
            continue

        if len(matches) == 1:
            tgt = matches[0]
            if a.excluded:
                act = ArrayAction(a.id, a.name, "skip-excluded",
                                  target_id=tgt.id, target_name=tgt.name)
                act.notes.append(
                    "source array is excluded=True — NOT linked to the target "
                    "sibling (writers omit excluded arrays either way, so report "
                    "parity holds; carrying the exclusion flag would silently "
                    "change AO billing). Source sibling detached from the moved "
                    "client. NEEDS FORD'S EYEBALL.")
                plan.arrays.append(act)
                continue
            if tgt.id in claimed_target_ids:
                plan.ambiguities.append(
                    f"array {a.id} {a.name!r}: target {tgt.id} {tgt.name!r} already "
                    f"claimed by another source array — will not guess")
                continue
            claimed_from_client: Optional[int] = None
            if tgt.client_id is not None:
                if tgt.client_id in claim_from:
                    # EXPLICIT operator choice (--claim-linked-from): take this
                    # array from the named capture client. Never inference.
                    claimed_from_client = tgt.client_id
                else:
                    plan.conflicts.append(
                        f"target array {tgt.id} {tgt.name!r} already linked to client "
                        f"{tgt.client_id} — target has its own reports world "
                        f"(--claim-linked-from {tgt.client_id} to claim it explicitly)")
                    continue
            claimed_target_ids.add(tgt.id)
            act = ArrayAction(a.id, a.name, how, target_id=tgt.id,
                              target_name=tgt.name, fields=fields,
                              client_id=a.client_id,
                              claimed_from=claimed_from_client)
            if claimed_from_client is not None:
                act.notes.append(
                    f"CLAIMED from client {claimed_from_client} "
                    f"(--claim-linked-from) — re-pointed to the migrated client")
            if tgt.name != a.name:
                act.rename_to = a.name
                act.notes.append(
                    f"target renamed {tgt.name!r} -> {a.name!r} (the name on the "
                    f"NEPOOL reports) so sheets stay byte-identical")
            act.notes.append("source sibling array detached (client_id -> NULL) "
                             "so the moved client never double-counts")
            plan.arrays.append(act)
            continue

        # No confident match -> MOVE the source array wholesale (default).
        act = ArrayAction(a.id, a.name, "move", fields=fields)
        act.notes.append(
            "no target match — array is CREATED ON TARGET by re-pointing the "
            "source row (accounts + bills + daily generation move with it)")
        clash = (a.name or "").strip().lower()
        if clash in tgt_by_lower:
            # only reachable if the name-match was ambiguous above; belt anyway
            plan.conflicts.append(
                f"move array {a.id} {a.name!r} would collide with a live target "
                f"array name")
        for ua in src_accts:
            if (ua.provider, ua.account_number) in all_tgt_acct_keys:
                plan.conflicts.append(
                    f"move array {a.id} {a.name!r}: target already has a "
                    f"({ua.provider}, {ua.account_number}) utility account "
                    f"(uq_account_per_tenant) — resolve by hand")
        plan.arrays.append(act)

    # Explicit capture-client retirement (--deactivate-client). Only a
    # TARGET-tenant client, and only when — after the claims above — it would
    # hold no live report arrays (nepool_gis_id set). Plain capture arrays may
    # remain; deactivating never touches arrays, just the reports iteration.
    claimed_by_client: dict[int, set[int]] = {}
    for act in plan.arrays:
        if act.claimed_from is not None and act.target_id is not None:
            claimed_by_client.setdefault(act.claimed_from, set()).add(act.target_id)
    unused_claims = claim_from - set(claimed_by_client)
    for cid in sorted(unused_claims):
        plan.notes.append(
            f"--claim-linked-from {cid}: no linked match was claimed from it")
    for cid in sorted(deactivate_clients):
        c = db.get(Client, cid)
        if c is None or c.tenant_id != target.id or c.deleted_at is not None:
            plan.conflicts.append(
                f"--deactivate-client {cid}: not a live client on target {target.id}")
            continue
        live_arrays = db.execute(
            select(Array).where(Array.client_id == cid,
                                Array.deleted_at.is_(None))
        ).scalars().all()
        claimed_ids = claimed_by_client.get(cid, set())
        remaining = [a for a in live_arrays if a.id not in claimed_ids]
        report_left = [a for a in remaining if getattr(a, "nepool_gis_id", None)]
        if report_left:
            plan.conflicts.append(
                f"--deactivate-client {cid} {c.name!r}: would still hold "
                f"{len(report_left)} live REPORT array(s) after the claims "
                f"({', '.join(repr(a.name) for a in report_left[:5])}) — refused; "
                f"only plain capture arrays may remain on a deactivated client")
            continue
        plan.deactivate.append({"id": cid, "name": c.name,
                                "remaining_arrays": len(remaining)})

    # Tenant-level settings diff.
    for f in HARD_COPY_SETTINGS:
        sv, tv = getattr(source, f), getattr(target, f)
        plan.settings.append({"field": f, "source": sv, "target": tv,
                              "action": "copy" if sv != tv else "already-equal"})
    for f in SOFT_COPY_SETTINGS:
        sv, tv = getattr(source, f), getattr(target, f)
        if tv not in (None, "") and sv not in (None, "") and sv != tv:
            action = "keep-target (already set — source value NOT copied)"
        elif tv in (None, "") and sv not in (None, ""):
            action = "copy"
        else:
            action = "already-equal" if sv == tv else "keep-target"
        plan.settings.append({"field": f, "source": sv, "target": tv,
                              "action": action})

    plan.report_deliveries = db.execute(
        select(func.count()).select_from(ReportDelivery)
        .where(ReportDelivery.tenant_id == source.id)
    ).scalar() or 0

    return plan


# ───────────────────────────── plan application (writes) ───────────────────────

def apply_plan(db, plan: Plan) -> None:
    """Apply the plan's mutations on `db`. Flushes; NEVER commits — the caller
    owns the transaction (verify rolls back, execute commits)."""
    if plan.blockers:
        raise MigrationBlocked("; ".join(plan.blockers))

    source = db.get(Tenant, plan.source_id)
    target = db.get(Tenant, plan.target_id)

    # 1. Clients re-point (ids stable).
    for c in plan.clients:
        client = db.get(Client, c["id"])
        client.tenant_id = target.id

    # 2. Arrays.
    for act in plan.arrays:
        src_arr = db.get(Array, act.source_id)
        if act.action in ("link-name", "link-account"):
            tgt_arr = db.get(Array, act.target_id)
            tgt_arr.client_id = src_arr.client_id
            for f, v in act.fields.items():
                setattr(tgt_arr, f, v)
            if act.rename_to:
                tgt_arr.name = act.rename_to
            src_arr.client_id = None          # detach the source sibling
        elif act.action == "skip-excluded":
            src_arr.client_id = None          # detach; never link the sibling
        elif act.action == "move":
            src_arr.tenant_id = target.id     # client_id already points at the
                                              # moved client (same id)
            acct_ids = [ua.id for ua in db.execute(
                select(UtilityAccount).where(UtilityAccount.array_id == src_arr.id)
            ).scalars().all()]
            if acct_ids:
                db.execute(update(UtilityAccount)
                           .where(UtilityAccount.id.in_(acct_ids))
                           .values(tenant_id=target.id))
                db.execute(update(Bill)
                           .where(Bill.account_id.in_(acct_ids))
                           .values(tenant_id=target.id))
            db.execute(update(DailyGeneration)
                       .where(DailyGeneration.array_id == src_arr.id)
                       .values(tenant_id=target.id))

    # 3. Tenant settings + the reports-world marker. generation_reports is what
    # makes the migrated AO tenant ELIGIBLE for scheduled sends / digests /
    # the operator directory (api/report_eligibility.tenant_in_reports_world)
    # — an explicit marker, because AO capture auto-creates Client rows and
    # bare client-presence can't identify a reports tenant.
    target.generation_reports = True
    for s in plan.settings:
        if s["action"] == "copy":
            setattr(target, s["field"], getattr(source, s["field"]))

    # 4. ReportDelivery history.
    db.execute(update(ReportDelivery)
               .where(ReportDelivery.tenant_id == source.id)
               .values(tenant_id=target.id))

    # 5. Explicit capture-client retirement (validated in build_plan).
    for d in plan.deactivate:
        c = db.get(Client, d["id"])
        c.active = False

    db.flush()


# ───────────────────────── workbook oracle (verify mode) ───────────────────────

class _NonClosingSession:
    """Hands the writers OUR session while refusing to let them end the
    transaction — the verify oracle's whole point is that every read happens
    inside the one uncommitted transaction we roll back at the end. The report
    writers are read-only (they never commit); commit degrades to flush as a
    belt anyway."""

    def __init__(self, real):
        object.__setattr__(self, "_real", real)

    def __getattr__(self, name):
        return getattr(object.__getattribute__(self, "_real"), name)

    def close(self):        # noqa: D102
        pass

    def commit(self):       # noqa: D102
        object.__getattribute__(self, "_real").flush()

    def rollback(self):     # noqa: D102
        pass

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


def _patch_writer_sessions(shared) -> list:
    """Point every module-level SessionLocal on the workbook path at OUR
    session (each writer module froze its own reference at import time).
    Returns the originals for _unpatch."""
    import api.db as db_mod
    from api.writers import registry as w_registry
    from api.writers import gmcs_writer as w_gmcs
    from api.writers import rec_writer as w_rec
    mods = [db_mod, w_registry, w_gmcs, w_rec]
    saved = [(m, m.SessionLocal) for m in mods]
    for m in mods:
        m.SessionLocal = lambda: _NonClosingSession(shared)
    return saved


def _unpatch_writer_sessions(saved: list) -> None:
    for m, orig in saved:
        m.SessionLocal = orig


def _build_bytes(client_id: int, out_dir: pathlib.Path, tag: str,
                 reference_date: Optional[date]) -> bytes:
    from api.writers import build_workbook
    out = out_dir / f"client-{client_id}-{tag}.xlsx"
    path = build_workbook(client_id=client_id, out_path=out,
                          reference_date=reference_date)
    return pathlib.Path(path).read_bytes()


def compare_workbooks(pre: bytes, post: bytes) -> tuple[str, str]:
    """Returns (level, detail).

    identical          — raw bytes equal
    content_identical  — every zip member byte-equal except docProps/core.xml
                         (openpyxl stamps created/modified timestamps there)
    semantically_equal — zip members differ but sheet names + every cell value
                         match via openpyxl
    different          — real content divergence (detail lists the first few)
    """
    if pre == post:
        return "identical", "byte-identical"
    za = zipfile.ZipFile(io.BytesIO(pre))
    zb = zipfile.ZipFile(io.BytesIO(post))
    names_a, names_b = set(za.namelist()), set(zb.namelist())
    member_diffs: list[str]
    if names_a == names_b:
        member_diffs = [n for n in sorted(names_a)
                        if n != "docProps/core.xml" and za.read(n) != zb.read(n)]
        if not member_diffs:
            return ("content_identical",
                    "byte-identical except docProps/core.xml (xlsx metadata timestamps)")
    else:
        member_diffs = [f"only-in-one: {n}" for n in sorted(names_a ^ names_b)]

    sem = _semantic_diff(pre, post)
    if not sem:
        return ("semantically_equal",
                f"zip members differ ({member_diffs[:4]}) but sheet names and "
                f"every cell value match")
    return "different", "; ".join(sem[:10])


def _semantic_diff(pre: bytes, post: bytes) -> list[str]:
    from openpyxl import load_workbook
    out: list[str] = []
    wa = load_workbook(io.BytesIO(pre))
    wb = load_workbook(io.BytesIO(post))
    if wa.sheetnames != wb.sheetnames:
        out.append(f"sheetnames {wa.sheetnames} != {wb.sheetnames}")
        return out
    for name in wa.sheetnames:
        sa, sb = wa[name], wb[name]
        rows_a = [[c.value for c in row] for row in sa.iter_rows()]
        rows_b = [[c.value for c in row] for row in sb.iter_rows()]
        if len(rows_a) != len(rows_b):
            out.append(f"[{name}] row count {len(rows_a)} != {len(rows_b)}")
            continue
        for i, (ra, rb) in enumerate(zip(rows_a, rows_b), start=1):
            if ra != rb:
                out.append(f"[{name}] row {i}: {ra} != {rb}")
                if len(out) >= 10:
                    return out
    return out


def run_verify(source_id: str, target_id: str, *,
               reference_date: Optional[date] = None,
               out_dir: Optional[pathlib.Path] = None,
               allow_products: bool = False,
               claim_from: Optional[set[int]] = None,
               deactivate_clients: Optional[set[int]] = None) -> dict:
    """The byte-diff oracle: migrate INSIDE one transaction, build every moved
    client's workbook from the migrated target state, compare against the
    source-tenant workbooks built pre-migration, ROLL BACK. Returns
    {client_id: {"level":…, "detail":…}, "_plan": Plan, "_post_state": {…}}.
    Never commits.

    "_post_state" asserts the post-migration REPORTS ITERATION: the target's
    active clients must be exactly the migrated (source-active) client set —
    a capture client left active fails this — and every claimed array must
    sit on its migrated client."""
    tmp_ctx = None
    if out_dir is None:
        tmp_ctx = tempfile.TemporaryDirectory(prefix="so-migrate-verify-")
        out_dir = pathlib.Path(tmp_ctx.name)
    out_dir.mkdir(parents=True, exist_ok=True)

    session = SessionLocal()
    saved = _patch_writer_sessions(session)
    results: dict = {}
    try:
        source = session.get(Tenant, source_id)
        target = session.get(Tenant, target_id)
        if source is None or target is None:
            raise MigrationBlocked("source or target tenant not found")
        plan = build_plan(session, source, target, allow_products=allow_products,
                          claim_from=claim_from,
                          deactivate_clients=deactivate_clients)
        results["_plan"] = plan
        if plan.blockers:
            raise MigrationBlocked("; ".join(plan.blockers))

        active_clients = [c for c in plan.clients if c["active"]]
        pre = {c["id"]: _build_bytes(c["id"], out_dir, "pre-source", reference_date)
               for c in active_clients}

        apply_plan(session, plan)
        session.flush()
        session.expire_all()   # safe post-flush; forces post-builds to re-read

        for c in active_clients:
            post = _build_bytes(c["id"], out_dir, "post-target", reference_date)
            level, detail = compare_workbooks(pre[c["id"]], post)
            results[c["id"]] = {"client": c["name"], "level": level,
                                "detail": detail}

        # Post-state assertions: what the reports iteration will actually see.
        problems: list[str] = []
        expected_active = {c["id"] for c in active_clients}
        actual_active = {cid for (cid,) in session.execute(
            select(Client.id).where(Client.tenant_id == target_id,
                                    Client.active == True,   # noqa: E712
                                    Client.deleted_at.is_(None))).all()}
        if actual_active != expected_active:
            extra = sorted(actual_active - expected_active)
            missing = sorted(expected_active - actual_active)
            if extra:
                problems.append(
                    f"target still iterates non-migrated active client(s) {extra} "
                    f"— deactivate them (--deactivate-client) or explain why not")
            if missing:
                problems.append(f"migrated active client(s) {missing} missing "
                                f"from the target's active set")
        for act in plan.arrays:
            if act.claimed_from is not None:
                arr = session.get(Array, act.target_id)
                if arr.client_id != act.client_id:
                    problems.append(
                        f"claimed array {act.target_id} {act.source_name!r} sits on "
                        f"client {arr.client_id}, expected migrated client "
                        f"{act.client_id}")
        results["_post_state"] = {
            "ok": not problems,
            "detail": ("; ".join(problems) if problems else
                       f"target active clients == migrated set "
                       f"({sorted(expected_active)}); "
                       f"{sum(1 for a in plan.arrays if a.claimed_from is not None)} "
                       f"claimed array(s) on their migrated clients"),
        }
        return results
    finally:
        session.rollback()     # the whole migration evaporates — flush proof
        _unpatch_writer_sessions(saved)
        session.close()
        if tmp_ctx is not None:
            tmp_ctx.cleanup()


def run_execute(source_id: str, target_id: str, *,
                allow_products: bool = False,
                claim_from: Optional[set[int]] = None,
                deactivate_clients: Optional[set[int]] = None) -> Plan:
    """Apply + COMMIT. Raises MigrationBlocked instead of guessing."""
    with SessionLocal() as session:
        source = session.get(Tenant, source_id)
        target = session.get(Tenant, target_id)
        if source is None or target is None:
            raise MigrationBlocked("source or target tenant not found")
        plan = build_plan(session, source, target, allow_products=allow_products,
                          claim_from=claim_from,
                          deactivate_clients=deactivate_clients)
        apply_plan(session, plan)    # raises on blockers, flushes
        session.commit()
        return plan


# ─────────────────────────────────── CLI ───────────────────────────────────────

def _print_plan(plan: Plan) -> None:
    print(f"\n=== MIGRATION PLAN  {plan.source_id}  ->  {plan.target_id} ===")
    print(f"\nClients to move ({len(plan.clients)}) — re-pointed, ids stable:")
    for c in plan.clients:
        flags = "" if c["active"] else "  (inactive)"
        if c.get("arrays", None) == 0:
            flags += "  (no arrays — moves as-is)"
        print(f"  - [{c['id']}] {c['name']}{flags}")
    print(f"\nArrays ({len(plan.arrays)}):")
    for a in plan.arrays:
        verb = a.action.upper()
        if a.claimed_from is not None:
            verb = f"CLAIM (from client {a.claimed_from})"
        line = f"  - [{a.source_id}] {a.source_name!r}: {verb}"
        if a.target_id is not None:
            line += f" -> target [{a.target_id}] {a.target_name!r}"
        if a.rename_to:
            line += f" (rename -> {a.rename_to!r})"
        print(line)
        if a.fields and a.action.startswith("link"):
            carried = {k: v for k, v in a.fields.items() if v not in (None, "")}
            if carried:
                print(f"      carries: {carried}")
        for n in a.notes:
            print(f"      · {n}")
    print("\nTenant settings:")
    for s in plan.settings:
        print(f"  - {s['field']}: {s['action']}"
              + (f"  ({s['source']!r} -> target)" if s["action"] == "copy" else ""))
    if plan.deactivate:
        print(f"\nClients to DEACTIVATE on target ({len(plan.deactivate)}) — "
              f"retired from the reports iteration, arrays untouched:")
        for d in plan.deactivate:
            print(f"  - [{d['id']}] {d['name']}  "
                  f"({d['remaining_arrays']} plain capture array(s) remain on it)")
    print(f"\nReportDelivery history rows to re-point: {plan.report_deliveries}")
    print("Target tenant gains generation_reports=True — this is what turns on "
          "scheduled sends,\ndigests and the operator directory for the migrated "
          "AO tenant (report_eligibility).")
    for n in plan.notes:
        print(f"  note: {n}")
    if plan.ambiguities:
        print("\n⚠ AMBIGUITIES (execute will ABORT — resolve by hand, never guessed):")
        for a in plan.ambiguities:
            print(f"  ! {a}")
    if plan.conflicts:
        print("\n✗ CONFLICTS (execute will ABORT):")
        for c in plan.conflicts:
            print(f"  ! {c}")
    if not plan.blockers:
        print("\nNo ambiguities or conflicts — plan is executable.")


def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--source", required=True, help="NEPOOL tenant id (ten_…)")
    ap.add_argument("--target", required=True, help="Array Operator tenant id (ten_…)")
    ap.add_argument("--report", action="store_true",
                    help="print the mapping plan (dry-run default does too)")
    ap.add_argument("--verify", action="store_true",
                    help="in-transaction migrate + per-client workbook byte-diff + ROLLBACK")
    ap.add_argument("--execute", action="store_true",
                    help="apply and COMMIT (dry-run without it)")
    ap.add_argument("--yes-prod", action="store_true",
                    help="required for --execute against a Postgres DATABASE_URL")
    ap.add_argument("--allow-products", action="store_true",
                    help="skip the source=nepool / target=array_operator sanity check")
    ap.add_argument("--claim-linked-from", action="append", type=int, default=[],
                    metavar="CLIENT_ID",
                    help="repeatable: allow LINK matches onto target arrays currently "
                         "belonging to this client id (re-pointed to the migrated "
                         "client); any OTHER linked target still aborts")
    ap.add_argument("--deactivate-client", action="append", type=int, default=[],
                    metavar="CLIENT_ID",
                    help="repeatable: set this TARGET-tenant client active=False in the "
                         "same transaction; refused if it would still hold live report "
                         "arrays (nepool_gis_id) after the claims")
    ap.add_argument("--quarter", default=None,
                    help="pin the report window, e.g. Q1-2026 (default: the writers' own default)")
    ap.add_argument("--keep-workbooks", default=None,
                    help="directory to keep the pre/post verify workbooks in")
    args = ap.parse_args(argv)
    claim_from = set(args.claim_linked_from)
    deactivate_clients = set(args.deactivate_client)

    reference_date = None
    if args.quarter:
        import re as _re
        m = _re.fullmatch(r"[Qq]([1-4])-(\d{4})", args.quarter.strip())
        if not m:
            print(f"bad --quarter {args.quarter!r} (want e.g. Q1-2026)")
            return 2
        q, y = int(m.group(1)), int(m.group(2))
        # reference INSIDE the quarter AFTER the wanted last quarter → window
        # ends on the named quarter (same semantics as ?quarter=).
        nq_y, nq = (y + 1, 1) if q == 4 else (y, q + 1)
        reference_date = date(nq_y, (nq - 1) * 3 + 1, 1)

    is_postgres = engine.dialect.name.startswith("postgres")
    if args.execute and is_postgres and not args.yes_prod:
        print("REFUSED: --execute against a Postgres database needs --yes-prod "
              "(Phase 4 only — Phase 1 never mutates prod).")
        return 2

    # Always show the plan first.
    with SessionLocal() as db:
        source = db.get(Tenant, args.source)
        target = db.get(Tenant, args.target)
        if source is None or target is None:
            print(f"tenant not found: {'source ' + args.source if source is None else ''}"
                  f"{'target ' + args.target if target is None else ''}")
            return 2
        plan = build_plan(db, source, target, allow_products=args.allow_products,
                          claim_from=claim_from,
                          deactivate_clients=deactivate_clients)
    _print_plan(plan)

    if args.verify:
        out_dir = pathlib.Path(args.keep_workbooks) if args.keep_workbooks else None
        print("\n=== VERIFY (in-transaction migrate → workbook byte-diff → ROLLBACK) ===")
        try:
            results = run_verify(args.source, args.target,
                                 reference_date=reference_date, out_dir=out_dir,
                                 allow_products=args.allow_products,
                                 claim_from=claim_from,
                                 deactivate_clients=deactivate_clients)
        except MigrationBlocked as e:
            print(f"VERIFY BLOCKED: {e}")
            return 3
        ok = True
        for cid, r in results.items():
            if cid in ("_plan", "_post_state"):
                continue
            good = r["level"] in ("identical", "content_identical")
            ok = ok and good
            mark = "✓" if good else "✗"
            print(f"  {mark} client [{cid}] {r['client']}: {r['level']} — {r['detail']}")
        post = results.get("_post_state") or {"ok": False, "detail": "missing"}
        ok = ok and post["ok"]
        print(f"  {'✓' if post['ok'] else '✗'} reports iteration post-state: {post['detail']}")
        print(f"VERIFY {'PASSED' if ok else 'FAILED'} (all changes rolled back)")
        if not ok and args.execute:
            print("REFUSED: --execute skipped because verify failed.")
            return 3
        if not ok:
            return 1

    if args.execute:
        if plan.blockers:
            print("REFUSED: open ambiguities/conflicts — resolve them first.")
            return 3
        try:
            run_execute(args.source, args.target,
                        allow_products=args.allow_products,
                        claim_from=claim_from,
                        deactivate_clients=deactivate_clients)
        except MigrationBlocked as e:
            print(f"EXECUTE BLOCKED: {e}")
            return 3
        print("\nEXECUTED + COMMITTED. Re-run with --report to see the (now empty) "
              "source state, and byte-verify the target with a fresh workbook build.")
    elif not args.verify:
        print("\nDRY-RUN — nothing was written. Add --verify for the in-transaction "
              "byte-diff proof, --execute to apply.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
