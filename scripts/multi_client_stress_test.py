"""Multi-client scheduled-delivery proof.

Creates a synthetic tenant with 3 clients on different report cadences,
seeds 18 months of bills, runs the real scheduler delivery path, and asserts:
  1. No cross-client data leakage in generated workbooks
  2. Each client fans out to the correct scheduler run (quarterly vs monthly)
  3. APScheduler fires a one-shot date job at T+2 minutes (--no-oneshot skips)

Run on Railway:
  railway ssh "cd /app && python -m scripts.multi_client_stress_test"

Dry-run (no emails, proves seeding + workbook generation):
  railway ssh "cd /app && python -m scripts.multi_client_stress_test --dry-run"

WARNING: _deliver_clients_with_frequency runs against ALL active clients in prod,
not just the synthetic tenant. It will trigger deliveries for all other quarterly/
monthly clients (including Bruce). Run on a quiet day or accept the extra emails.
"""
from __future__ import annotations

import argparse
import random
import secrets
import sys
import tempfile
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

from sqlalchemy import text, select
from api.db import SessionLocal, engine, init_db
from api.models import Tenant, Client, Array, UtilityAccount, Bill
from api.writers.gmcs_writer import build_workbook as _build_workbook
from api.scheduler import _deliver_clients_with_frequency


# ─── constants ────────────────────────────────────────────────────────────────

TID = "ten_stress_mc_v1"
TENANT_EMAIL = "ford.genereaux@gmail.com"   # to_both copies land here

# Subject template with [STRESS-TEST] prefix so Ford can filter his inbox.
# Uses the same merge-tag syntax as api/email_templates.py DEFAULT_SUBJECT_TEMPLATE.
SUBJECT_TPL = (
    "[STRESS-TEST] {{client_name}} — generation report "
    "({{period_start}} to {{period_end}})"
)

# 18 months of bills — chronological, same window as stress_test_gmcs.py
MONTHS_18 = [
    (2024, 7), (2024, 8), (2024, 9), (2024, 10), (2024, 11), (2024, 12),
    (2025, 1), (2025, 2), (2025, 3), (2025, 4), (2025, 5), (2025, 6),
    (2025, 7), (2025, 8), (2025, 9), (2025, 10), (2025, 11), (2025, 12),
]

# VT capacity factor (same as stress_test_gmcs.py)
MONTHLY_CF_PCT = {
    1: 7.5, 2: 9.0, 3: 13.0, 4: 15.5, 5: 17.0, 6: 17.5,
    7: 18.0, 8: 17.0, 9: 14.0, 10: 10.5, 11: 6.5, 12: 5.5,
}

# Three clients.  Each entry:
#   name, frequency (None = inherit), contact_email, arrays list
#   Each array: (name, nepool_id, kw_capacity, num_sub_meters)
CLIENTS_DEF = [
    {
        "name": "Quarterly Co",
        "frequency": "quarterly",
        "email": "ford.genereaux+stress-q@gmail.com",
        "arrays": [
            ("Addison Solar Farm",  "AS-20001", 250, 1),
            ("Bristol Ridge Array", "BR-20002", 150, 2),
            ("Cornwall Fields",     "CF-20003", 400, 3),
            ("Duxbury Heights",     "DH-20004", 200, 1),
        ],
    },
    {
        "name": "Monthly Co",
        "frequency": "monthly",
        "email": "ford.genereaux+stress-m@gmail.com",
        "arrays": [
            ("Eden Prairie Solar", "EP-20005", 300, 1),
            ("Fairfax Commons",    "FC-20006", 175, 2),
            ("Greensboro Array",   "GA-20007", 500, 1),
        ],
    },
    {
        "name": "Inherit Co",
        "frequency": None,  # inherits tenant quarterly
        "email": "ford.genereaux+stress-i@gmail.com",
        "arrays": [
            ("Hardwick Solar",       "HS-20008", 100, 1),
            ("Irasburg Community",   "IC-20009", 350, 2),
            ("Johnson Hill Array",   "JH-20010", 225, 1),
            ("Kirby Meadow Solar",   "KM-20011", 450, 3),
            ("Lincoln Ridge Farm",   "LR-20012", 150, 1),
        ],
    },
]


# ─── helpers ──────────────────────────────────────────────────────────────────

def _kwh(kw: int, year: int, month: int, rng: random.Random) -> int:
    days = 31 if month in (1, 3, 5, 7, 8, 10, 12) else (30 if month != 2 else 28)
    cf = MONTHLY_CF_PCT[month] / 100.0
    jitter = 1 + rng.uniform(-0.08, 0.08)
    return max(1, int(round(kw * cf * days * 24 * jitter)))


def _hard_delete(tid: str) -> None:
    """Hard-delete tenant + all dependent rows (FK-safe ordering)."""
    with engine.begin() as conn:
        conn.execute(text(
            "DELETE FROM bills WHERE account_id IN "
            "(SELECT id FROM utility_accounts WHERE tenant_id = :t)"
        ), {"t": tid})
        conn.execute(text("DELETE FROM utility_accounts WHERE tenant_id = :t"), {"t": tid})
        conn.execute(text("DELETE FROM utility_sessions WHERE tenant_id = :t"), {"t": tid})
        conn.execute(text("DELETE FROM login_tokens WHERE tenant_id = :t"), {"t": tid})
        conn.execute(text("DELETE FROM arrays WHERE tenant_id = :t"), {"t": tid})
        conn.execute(text("DELETE FROM clients WHERE tenant_id = :t"), {"t": tid})
        try:
            conn.execute(text("DELETE FROM tenant_templates WHERE tenant_id = :t"), {"t": tid})
        except Exception:
            pass
        try:
            conn.execute(text("DELETE FROM delete_history WHERE tenant_id = :t"), {"t": tid})
        except Exception:
            pass
        conn.execute(text("DELETE FROM tenants WHERE id = :t"), {"t": tid})


def _sep(title: str = "") -> None:
    line = "=" * 60
    if title:
        print(f"\n{line}\n{title}\n{line}")
    else:
        print(line)


# ─── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Multi-client scheduled-delivery proof — prod acceptance test")
    ap.add_argument("--keep-tenant", action="store_true",
                    help="Skip cleanup so you can inspect the synthetic rows")
    ap.add_argument("--no-oneshot", action="store_true",
                    help="Skip APScheduler one-shot proof (use when running locally)")
    ap.add_argument("--bills-months", type=int, default=18,
                    help="Months of bills to seed (default 18)")
    ap.add_argument("--dry-run", action="store_true",
                    help="Seed + workbook assertion only — no emails, no scheduler calls")
    ap.add_argument("--client-count", type=int, default=None,
                    help="Override total client count for capacity testing (implies --dry-run --no-oneshot)")
    args = ap.parse_args()

    # --client-count implies dry-run and no-oneshot (capacity test mode)
    if args.client_count is not None:
        args.dry_run = True
        args.no_oneshot = True

    months = MONTHS_18[:args.bills_months]
    rng = random.Random(20260604)  # deterministic — same seed → same numbers

    # Expand CLIENTS_DEF to the requested client count by cycling through
    # synthetic client templates.
    clients_def = CLIENTS_DEF
    if args.client_count is not None and args.client_count > 0:
        vt_towns = [
            "Addison", "Barnard", "Bethel", "Bolton", "Brandon", "Bridport",
            "Bristol", "Burlington", "Cabot", "Calais", "Cambridge", "Charlotte",
            "Chelsea", "Chester", "Colchester", "Cornwall", "Craftsbury",
            "Danville", "Derby", "Duxbury", "East Montpelier", "Eden",
            "Fairfax", "Fairlee", "Ferrisburgh", "Franklin", "Georgia",
            "Glover", "Goshen", "Granville", "Greensboro", "Groton",
            "Hardwick", "Hinesburg", "Huntington", "Hyde Park", "Irasburg",
            "Isle La Motte", "Jamaica", "Johnson", "Kirby", "Leicester",
            "Lincoln", "Londonderry", "Ludlow", "Lyndon", "Marshfield",
            "Middlebury", "Milton", "Montpelier", "Morrisville", "Newark",
            "Newbury", "Newfane", "Newport", "Northfield", "Norwich",
            "Orwell", "Panton", "Pawlet", "Peacham", "Pittsfield",
            "Plainfield", "Plymouth", "Pomfret", "Poultney", "Proctor",
            "Randolph", "Richmond", "Rochester", "Royalton", "Rupert",
            "Rutland", "Ryegate", "Saint Albans", "Salisbury", "Shoreham",
            "Springfield", "Starksboro", "Stowe", "Sudbury", "Swanton",
            "Thetford", "Tinmouth", "Topsham", "Townshend", "Troy",
            "Tunbridge", "Underhill", "Vergennes", "Vernon", "Waitsfield",
            "Wallingford", "Warren", "Waterbury", "Weathersfield", "Wells",
            "West Rutland", "Westminster", "Whiting", "Williston", "Wilmington",
            "Wolcott", "Woodstock", "Worcester",
        ]
        clients_def = []
        arrays_per_client_cycle = [3, 4, 2, 5, 3, 4]  # cycle for variety
        for i in range(args.client_count):
            town = vt_towns[i % len(vt_towns)]
            n_arrays = arrays_per_client_cycle[i % len(arrays_per_client_cycle)]
            freq = ["quarterly", "monthly", None][i % 3]
            arrays = []
            for j in range(n_arrays):
                array_town = vt_towns[(i * 7 + j) % len(vt_towns)]
                nepool_id = f"SC{i:04d}{j}"
                kw = 100 + ((i * 13 + j * 7) % 400)
                meters = 1 if j % 3 != 2 else 2
                arrays.append((f"{array_town} Array C{i+1}-{j+1}", nepool_id, kw, meters))
            clients_def.append({
                "name": f"{town} Solar Operator {i+1}",
                "frequency": freq,
                "email": f"ford.genereaux+stress-{i}@gmail.com",
                "arrays": arrays,
            })

    init_db()

    # ── Step 1: Seed ─────────────────────────────────────────────────────────
    _sep("Step 1 — Seeding synthetic tenant")

    # Wipe any leftover from a prior run before opening the main session.
    with SessionLocal() as probe:
        found = probe.query(Tenant).filter_by(id=TID).first()
    if found:
        print(f"  Prior run found ({TID}), wiping first...")
        _hard_delete(TID)

    client_meta: list[dict] = []   # {id, name, frequency, array_ids, n_arrays, email}
    n_accounts = 0
    n_bills = 0
    t_seed_start = time.monotonic()

    with SessionLocal() as db:
        tenant = Tenant(
            id=TID,
            name="Multi-Client Stress Test Co",
            contact_email=TENANT_EMAIL,
            tenant_key="sol_stress_mc_" + secrets.token_hex(8),
            plan="comped",
            active=True,
            subscription_status="comped",
            report_frequency="quarterly",   # Inherit Co will inherit this
            send_mode="to_both",            # Ford gets a copy of every delivery
            email_subject_template=SUBJECT_TPL,
        )
        db.add(tenant)
        db.flush()

        for cdef in clients_def:
            client = Client(
                tenant_id=TID,
                name=cdef["name"],
                contact_email=cdef["email"],
                report_frequency=cdef["frequency"],
                active=True,
            )
            db.add(client)
            db.flush()
            cid = client.id
            arr_ids: list[int] = []

            for (aname, nepool, kw, meters) in cdef["arrays"]:
                arr = Array(
                    tenant_id=TID,
                    client_id=cid,
                    name=aname,
                    nepool_gis_id=nepool,
                    region="Vermont",
                    bill_offset_months=1,
                )
                db.add(arr)
                db.flush()
                arr_ids.append(arr.id)

                per_meter = kw // meters
                rem = kw - per_meter * meters
                for mi in range(meters):
                    meter_kw = per_meter + (rem if mi == 0 else 0)
                    acct_num = f"{nepool.replace('-', '')}{mi:02d}"
                    nick = aname if meters == 1 else f"{aname} M{mi + 1}"
                    acc = UtilityAccount(
                        tenant_id=TID,
                        array_id=arr.id,
                        provider="gmp",
                        account_number=acct_num,
                        nickname=nick,
                    )
                    db.add(acc)
                    db.flush()
                    n_accounts += 1

                    for (y, mo) in months:
                        days_in_mo = (31 if mo in (1, 3, 5, 7, 8, 10, 12)
                                      else 30 if mo != 2 else 28)
                        db.add(Bill(
                            tenant_id=TID,
                            account_id=acc.id,
                            bill_date=datetime(y, mo, days_in_mo),
                            period_start=datetime(y, mo, 1),
                            period_end=datetime(y, mo, days_in_mo),
                            billing_days=days_in_mo,
                            kwh_generated=_kwh(meter_kw, y, mo, rng),
                            document_number=f"ST-{acc.id}-{y}-{mo:02d}",
                            parse_status="parsed",
                        ))
                        n_bills += 1

            client_meta.append({
                "id": cid,
                "name": cdef["name"],
                "frequency": cdef["frequency"],
                "n_arrays": len(cdef["arrays"]),
                "array_ids": arr_ids,
                "email": cdef["email"],
                "array_names": [a[0] for a in cdef["arrays"]],
            })

        db.commit()

    t_seed_elapsed = time.monotonic() - t_seed_start
    n_arrays_total = sum(c["n_arrays"] for c in client_meta)
    print(f"  Tenant:   {TID}")
    print(f"  Clients:  {len(client_meta)}")
    if len(client_meta) <= 10:
        for cm in client_meta:
            print(f"    {cm['name']}: {cm['n_arrays']} arrays")
    print(f"  Arrays:   {n_arrays_total}  Accounts: {n_accounts}  Bills: {n_bills}")

    # ── Assert bill counts before delivery ────────────────────────────────────
    bills_ok = True
    with SessionLocal() as db:
        for cm in client_meta:
            for aid in cm["array_ids"]:
                acct_ids = [
                    a.id for a in
                    db.query(UtilityAccount).filter_by(array_id=aid).all()
                ]
                n = db.query(Bill).filter(Bill.account_id.in_(acct_ids)).count()
                expected = len(acct_ids) * len(months)
                if n != expected:
                    print(f"  FAIL: array {aid} has {n} bills, expected {expected}")
                    bills_ok = False
    print(f"  Bill count pre-check: {'PASS' if bills_ok else 'FAIL'} "
          f"({n_bills} rows across {n_accounts} accounts × {len(months)} months)")

    if not bills_ok:
        print("\nAborting: bill count mismatch — check seeding logic.")
        if not args.keep_tenant:
            _hard_delete(TID)
        sys.exit(1)

    # ── Dry-run exit ──────────────────────────────────────────────────────────
    if args.dry_run:
        # Still build workbooks to prove the writer path works.
        _sep("Dry-run — workbook generation only (no emails)")
        t_wb_start = time.monotonic()
        total_sheets = 0
        total_kb = 0
        errors = 0
        with tempfile.TemporaryDirectory(prefix="so-stress-mc-dry-") as tmpdir:
            from openpyxl import load_workbook as _load_wb
            for cm in client_meta:
                out = Path(tmpdir) / f"{cm['name'].replace(' ', '_')[:60]}.xlsx"
                try:
                    wb_path = _build_workbook(client_id=cm["id"], out_path=out)
                    wb = _load_wb(str(wb_path))
                    size_kb = wb_path.stat().st_size // 1024
                    total_sheets += len(wb.sheetnames)
                    total_kb += size_kb
                    if len(client_meta) <= 10:
                        print(f"  {cm['name']}: {len(wb.sheetnames)} sheets, {size_kb}KB — OK")
                except Exception as e:
                    errors += 1
                    print(f"  FAIL {cm['name']}: {e}")
        t_wb_elapsed = time.monotonic() - t_wb_start
        print(f"  Workbook generation: {len(client_meta)} clients, "
              f"{n_arrays_total} arrays, {total_kb}KB total, "
              f"{t_wb_elapsed:.1f}s ({t_wb_elapsed/len(client_meta)*1000:.0f}ms/client) "
              f"— {'PASS' if errors == 0 else f'{errors} ERRORS'}")
        if not args.keep_tenant:
            _hard_delete(TID)
            print("  Cleanup: done")
        else:
            print(f"  Cleanup: SKIPPED (--keep-tenant). Tenant: {TID}")
        print(f"\nDry-run complete. Seeding: {t_seed_elapsed:.1f}s, Generation: {t_wb_elapsed:.1f}s")
        sys.exit(0 if errors == 0 else 1)

    # ── Step 2: Scheduler delivery path ───────────────────────────────────────
    _sep("Step 2 — _deliver_clients_with_frequency (direct call)")
    print("  NOTE: this runs against ALL active clients in prod, not just this tenant.")
    print("        Deliveries to other quarterly/monthly clients will also fire.")

    q_result = _deliver_clients_with_frequency("quarterly")
    print(f"  QUARTERLY result: sent={q_result['sent']} failed={q_result['failed']}")

    m_result = _deliver_clients_with_frequency("monthly")
    print(f"  MONTHLY result:   sent={m_result['sent']} failed={m_result['failed']}")

    # ── Step 3: Assert scheduler selected the right clients ───────────────────
    _sep("Step 3 — Scheduler client-selection assertions")

    cid_q = client_meta[0]["id"]   # Quarterly Co
    cid_m = client_meta[1]["id"]   # Monthly Co
    cid_i = client_meta[2]["id"]   # Inherit Co (inherits quarterly)

    q_sent = set(q_result["sent"])
    m_sent = set(m_result["sent"])

    quarterly_pass = (
        cid_q in q_sent      # Quarterly Co must be in quarterly run
        and cid_i in q_sent  # Inherit Co must inherit tenant quarterly
        and cid_m not in q_sent  # Monthly Co must NOT appear in quarterly
    )
    monthly_pass = (
        cid_m in m_sent      # Monthly Co must be in monthly run
        and cid_q not in m_sent  # Quarterly Co must NOT appear in monthly
        and cid_i not in m_sent  # Inherit Co must NOT appear in monthly
    )

    q_names_found = [cm["name"] for cm in client_meta if cm["id"] in q_sent]
    m_names_found = [cm["name"] for cm in client_meta if cm["id"] in m_sent]

    print(f"  Quarterly fan-out: {'PASS' if quarterly_pass else 'FAIL'} "
          f"— stress clients in result: {q_names_found}")
    print(f"  Monthly fan-out:   {'PASS' if monthly_pass else 'FAIL'} "
          f"— stress clients in result: {m_names_found}")
    if q_result["failed"]:
        print(f"  WARN quarterly failures: {q_result['failed']}")
    if m_result["failed"]:
        print(f"  WARN monthly failures: {m_result['failed']}")

    # ── Step 4: Workbook leakage assertion ────────────────────────────────────
    _sep("Step 4 — Workbook sheet-leakage check")

    leakage_pass = True
    with tempfile.TemporaryDirectory(prefix="so-stress-mc-") as tmpdir:
        from openpyxl import load_workbook as _load_wb

        # All array names per client, for cross-client leak detection
        all_array_names_by_cid = {cm["id"]: set(cm["array_names"]) for cm in client_meta}
        all_other_names_by_cid = {
            cm["id"]: set().union(*(
                v for k, v in all_array_names_by_cid.items() if k != cm["id"]
            ))
            for cm in client_meta
        }

        for cm in client_meta:
            out = Path(tmpdir) / f"{cm['name'].replace(' ', '_')}.xlsx"
            wb_path = _build_workbook(client_id=cm["id"], out_path=out)

            assert wb_path.stat().st_size > 0, f"Empty workbook: {wb_path}"
            wb = _load_wb(str(wb_path))
            sheets = set(wb.sheetnames)

            # Sheet count == array count
            if len(sheets) != cm["n_arrays"]:
                print(f"  FAIL {cm['name']}: expected {cm['n_arrays']} sheets, "
                      f"got {len(sheets)}: {sheets}")
                leakage_pass = False

            # No sheet from another client's array
            # Sheet names are truncated array names (≤31 chars), so check substrings
            for sheet in sheets:
                for other_name in all_other_names_by_cid[cm["id"]]:
                    # Match on first 31 chars (Excel sheet name limit)
                    if sheet.startswith(other_name[:31]) or other_name.startswith(sheet):
                        print(f"  FAIL {cm['name']}: leaked sheet '{sheet}' "
                              f"belongs to another client (matches '{other_name}')")
                        leakage_pass = False

            # All expected arrays present (check by checking sheet names vs array names)
            for aname in cm["array_names"]:
                expected_sheet = aname[:31]
                if expected_sheet not in sheets:
                    print(f"  FAIL {cm['name']}: missing sheet for array '{aname}'")
                    leakage_pass = False

            print(f"  {cm['name']}: {len(sheets)} sheets — "
                  f"{'OK' if len(sheets) == cm['n_arrays'] else 'WRONG COUNT'}")

    print(f"  Sheet leakage check: {'PASS' if leakage_pass else 'FAIL'}")

    # ── Step 5: APScheduler one-shot proof ────────────────────────────────────
    oneshot_pass: bool | None = None
    if not args.no_oneshot:
        _sep("Step 5 — APScheduler one-shot proof (T+2 min)")

        from api.scheduler import scheduler as _sched

        oneshot_result: list[dict] = []
        fired_event = threading.Event()

        def _oneshot_job() -> None:
            r = _deliver_clients_with_frequency("quarterly")
            oneshot_result.append(r)
            fired_event.set()

        run_at = datetime.utcnow() + timedelta(minutes=2)
        print(f"  Registering one-shot job for {run_at.strftime('%H:%M:%S')} UTC")

        if not _sched.running:
            _sched.start()

        _sched.add_job(
            _oneshot_job,
            "date",
            run_date=run_at,
            id="stress-test-oneshot",
            replace_existing=True,
        )
        print(f"  Job queued. Waiting up to 180s for it to fire...")

        fired = fired_event.wait(timeout=180)
        if _sched.running:
            _sched.shutdown(wait=False)

        if fired and oneshot_result:
            os_r = oneshot_result[0]
            os_sent = set(os_r["sent"])
            oneshot_pass = (
                cid_q in os_sent
                and cid_i in os_sent
                and cid_m not in os_sent
            )
            print(f"  One-shot fired: sent={os_r['sent']} failed={os_r['failed']}")
            print(f"  APScheduler one-shot: {'PASS' if oneshot_pass else 'FAIL'}")
        else:
            oneshot_pass = False
            print("  FAIL: one-shot did not fire within 180s")
    else:
        print("\n  (Step 5 skipped — --no-oneshot)")

    # ── Step 6: Cleanup ───────────────────────────────────────────────────────
    _sep("Step 6 — Cleanup")
    cleanup_ok = False
    if args.keep_tenant:
        print(f"  SKIPPED (--keep-tenant). Inspect tenant: {TID}")
    else:
        try:
            _hard_delete(TID)
            cleanup_ok = True
            print(f"  Tenant {TID} and all dependent rows deleted.")
        except Exception as e:
            print(f"\n!!! CLEANUP FAILED: {e}")
            print(f"!!! Hand-delete this tenant: {TID}")

    # ── Final report ──────────────────────────────────────────────────────────
    print()
    _sep("=== Multi-client scheduled-delivery proof ===")
    print()
    print(f"Synthetic tenant:    {TID} (cleaned up: {cleanup_ok})")
    print(f"Clients seeded:      3 "
          f"(Quarterly Co: {client_meta[0]['n_arrays']} arrays, "
          f"Monthly Co: {client_meta[1]['n_arrays']}, "
          f"Inherit Co: {client_meta[2]['n_arrays']})")
    print(f"Bills seeded:        {n_bills} rows "
          f"({n_accounts} accounts × {len(months)} months)")
    print()

    q_v = "PASS" if quarterly_pass else "FAIL"
    m_v = "PASS" if monthly_pass else "FAIL"
    l_v = "PASS" if leakage_pass else "FAIL"

    print(f"Quarterly fan-out:   2 clients expected — "
          f"{', '.join(q_names_found) or '(none)'} — {q_v}")
    print(f"Monthly fan-out:     1 client expected — "
          f"{', '.join(m_names_found) or '(none)'} — {m_v}")
    print(f"Sheet leakage check: {l_v} — each client workbook contains only its arrays")

    # Each successful delivery: 1 to client email + 1 to tenant (to_both) = 2 per client
    direct_clients_sent = len(q_result["sent"]) + len(m_result["sent"])
    print(f"Emails from direct calls: ~{direct_clients_sent * 2} "
          f"({direct_clients_sent} clients × 2 addresses via to_both)")
    print(f"  — to_both copies land in: {TENANT_EMAIL}")

    if oneshot_pass is not None:
        os_v = "PASS" if oneshot_pass else "FAIL"
        print(f"One-shot APScheduler:  {os_v} (job fired at T+2min, "
              f"delivered {len(oneshot_result[0]['sent']) if oneshot_result else 0} more clients)")
    else:
        print("One-shot APScheduler:  SKIPPED (--no-oneshot)")

    print(f"Cleanup:             {'done' if cleanup_ok else 'FAILED or skipped'}")
    print()

    checks = [quarterly_pass, monthly_pass, leakage_pass]
    if oneshot_pass is not None:
        checks.append(oneshot_pass)
    verdict = "PASS" if all(checks) else "FAIL"
    print(f"Verdict: {verdict}")
    sys.exit(0 if verdict == "PASS" else 1)


if __name__ == "__main__":
    main()
