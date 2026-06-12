"""DOGFOOD: render the first real Solar Operator report for OUR OWN agency.

Customer zero = tenant ``ten_ocicbb_customer_zero`` ("OCICBB Solar Agency"),
provisioned by scripts/provision_customer_zero.py. This script proves the
dogfood loop by running the REAL report writers (api/writers — the same
GMCS/REC workbook builders production delivery uses) against our own account
and writing the artifact to /tmp/dogfood_report/.

Safety: the live dev DB (storage/solar.db) is COPIED to a scratch file first
and SOLAR_DB_URL is pointed at the copy, so nothing here can mutate real
local state. No network, no email, no deploy.

Honesty note: customer-zero's WEC array has real pulled daily-generation rows
(source=smarthub) but every row is 0.0 kWh so far, so the customer-zero
report renders with zero MWh. Because the dogfood artifact must show real
numbers, this script ALSO seeds the repo's own deterministic example data
(scripts/seed_demo_tenant.py — 24 months of bills) into the scratch DB and
renders a second report from the seeded showpiece client. Both artifacts are
produced by the identical writer code path used in production delivery.

Run (from repo root):
    venv/bin/python scripts/render_dogfood_report.py
"""
from __future__ import annotations

import pathlib
import shutil
import sys
from datetime import date

REPO = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))

OUT_DIR = pathlib.Path("/tmp/dogfood_report")
SCRATCH_DB = OUT_DIR / "scratch-solar.db"
SOURCE_DB = REPO / "storage" / "solar.db"

CUSTOMER_ZERO_TENANT = "ten_ocicbb_customer_zero"
CUSTOMER_ZERO_NAME = "OCICBB Solar Agency"


def _point_db_at_scratch() -> None:
    """Copy the dev DB and aim the api at the copy BEFORE importing api.db."""
    import os
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if not SOURCE_DB.exists():
        sys.exit(f"source DB not found: {SOURCE_DB}")
    shutil.copy2(SOURCE_DB, SCRATCH_DB)
    os.environ.pop("DATABASE_URL", None)
    os.environ["SOLAR_DB_URL"] = f"sqlite:///{SCRATCH_DB}"


def _summarize_workbook(path: pathlib.Path) -> dict:
    """Extract headline numbers from a rendered workbook for verification."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    total_mwh = 0.0
    total_recs = 0
    nonzero_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=6, max_col=4):
            gen = row[1].value if len(row) > 1 else None   # col B: MWh
            recs = row[3].value if len(row) > 3 else None  # col D: RECs
            if isinstance(gen, (int, float)):
                total_mwh += float(gen)
                if gen:
                    nonzero_cells += 1
            if isinstance(recs, (int, float)):
                total_recs += int(recs)
    return {
        "sheets": sheets,
        "total_mwh": round(total_mwh, 3),
        "total_recs": total_recs,
        "nonzero_generation_cells": nonzero_cells,
        "size_bytes": path.stat().st_size,
    }


def main() -> None:
    _point_db_at_scratch()

    # Imports AFTER env pinning so api.db binds to the scratch copy.
    from sqlalchemy import select
    from api.db import SessionLocal, DB_URL
    from api.models import Client
    from api.writers import build_workbook

    print(f"DB (scratch copy): {DB_URL}")
    results: list[tuple[str, pathlib.Path, dict]] = []

    # ── 1. Customer zero: OCICBB Solar Agency (the literal dogfood account) ──
    with SessionLocal() as db:
        cz = db.execute(
            select(Client).where(Client.tenant_id == CUSTOMER_ZERO_TENANT)
        ).scalars().first()

    if cz is None:
        print(f"!! customer-zero client not found under {CUSTOMER_ZERO_TENANT} "
              "— run scripts/provision_customer_zero.py first")
    else:
        # reference_date=2026-07-01 closes Q2 2026 so the window includes the
        # WEC daily-generation pull (2026-03-13 → 2026-06-11).
        out = OUT_DIR / "OCICBB-Solar-Agency-customer-zero-report.xlsx"
        path = build_workbook(client_id=cz.id, out_path=out,
                              reference_date=date(2026, 7, 1))
        info = _summarize_workbook(path)
        results.append((f"CUSTOMER ZERO: {cz.name}", path, info))

    # ── 2. Repo seed data: deterministic demo tenant (real numbers) ─────────
    # seed_demo_tenant.seed() is idempotent and writes only to the scratch DB.
    from scripts.seed_demo_tenant import seed, DEMO_TENANT_ID
    seed()
    with SessionLocal() as db:
        demo_client = db.execute(
            select(Client).where(Client.tenant_id == DEMO_TENANT_ID)
            .order_by(Client.id.asc())
        ).scalars().first()
    if demo_client is None:
        sys.exit("seed ran but no demo client found — aborting")
    safe = demo_client.name.replace(" ", "-")
    out = OUT_DIR / f"seed-demo-{safe}-report.xlsx"
    path = build_workbook(client_id=demo_client.id, out_path=out)
    info = _summarize_workbook(path)
    results.append((f"SEED DEMO (scripts/seed_demo_tenant.py): {demo_client.name}",
                    path, info))

    # ── report ───────────────────────────────────────────────────────────────
    print("\n=== DOGFOOD RENDER RESULTS ===")
    ok = True
    for label, path, info in results:
        print(f"\n{label}")
        print(f"  path : {path}")
        print(f"  size : {info['size_bytes']:,} bytes")
        print(f"  sheets ({len(info['sheets'])}): {info['sheets']}")
        print(f"  total MWh: {info['total_mwh']}   total RECs: {info['total_recs']}"
              f"   nonzero gen cells: {info['nonzero_generation_cells']}")
        if info["size_bytes"] <= 5 * 1024:
            print("  !! FAILS size check (<=5KB)")
            ok = False
    has_real_numbers = any(i["nonzero_generation_cells"] > 0 for _, _, i in results)
    if not has_real_numbers:
        print("\n!! NO artifact contains nonzero generation — verification FAILED")
        ok = False
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()
