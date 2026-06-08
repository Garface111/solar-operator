"""
Scrape full 2026 monthly kWh from GMP via the per-account CSV download endpoint.

Endpoint discovered June 2026:
  GET https://api.greenmountainpower.com/api/v2/usage/{accountNumber}/download
      ?startDate=YYYY-MM-DDTHH:MM:SS-04:00
      &endDate=YYYY-MM-DDTHH:MM:SS-04:00
      &format=csv

Returns 15-minute interval generation data per ServiceAgreement.
CSV columns: ServiceAgreement, IntervalStart, IntervalEnd, Quantity, UnitOfMeasure
IntervalStart format: YYYY-MM-DD-HH:MM:SS

NOTE: Some accounts ignore the date filter and dump all available history.
We always client-side filter by IntervalStart to the requested window.

Output: GMS/2026 Monthly kWh - GMP CSV Scrape.xlsx
"""
import csv
import io
import sqlite3
import sys
import time
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DB = "/home/fordface/solar-operator/storage/solar.db"
OUT_XLSX = "/mnt/c/Users/fordg/Desktop/GMS/2026 Monthly kWh - GMP CSV Scrape.xlsx"

# Map utility_accounts.nickname → master sheet array name
NICK_TO_ARR = {
    "Tannery Brook": "Tannery Brook",
    "Chester": "Chester",
    "Timberworks": "Timberworks",
    "Waterford": "Waterford",
    "Londonderry": "Londonderry",
    "Starlake North": "Starlake",
    "Starlake South": "Starlake",
    "Starlake Center": "Starlake",
}
REGIONS = {
    "Tannery Brook": "north", "Chester": "south", "Timberworks": "north",
    "Waterford": "north", "Londonderry": "south", "Starlake": "central",
}
ARRAYS = ["Tannery Brook", "Chester", "Timberworks", "Waterford",
          "Londonderry", "Starlake"]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def load_token():
    con = sqlite3.connect(DB)
    row = con.execute(
        "SELECT api_token FROM utility_sessions ORDER BY id DESC LIMIT 1"
    ).fetchone()
    con.close()
    return row[0]


def load_accounts():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT account_number, nickname FROM utility_accounts ORDER BY id"
    ).fetchall()
    con.close()
    return [(r["account_number"], r["nickname"]) for r in rows]


def fetch_csv(account: str, token: str, year: int = 2026,
              max_retries: int = 4) -> str:
    """Fetch CSV for an account for the entire year. Returns raw CSV text."""
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://greenmountainpower.com",
        "Referer": "https://greenmountainpower.com/",
        "GMP-Source": "web",
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0 Safari/537.36"),
    }
    url = f"https://api.greenmountainpower.com/api/v2/usage/{account}/download"
    params = {
        "startDate": f"{year}-01-01T00:00:00-04:00",
        "endDate":   f"{year+1}-01-01T00:00:00-04:00",
        "format":    "csv",
    }
    for attempt in range(max_retries):
        try:
            with httpx.Client(timeout=240, headers=headers) as c:
                r = c.get(url, params=params)
            if r.status_code == 200 and len(r.text) > 200:
                return r.text
            print(f"  [{account}] attempt {attempt+1} status={r.status_code} "
                  f"len={len(r.text)}", file=sys.stderr)
        except Exception as e:
            print(f"  [{account}] attempt {attempt+1} err: {e}", file=sys.stderr)
        time.sleep(3 * (attempt + 1))
    return ""


def parse_csv_by_month(csv_text: str, year: int = 2026) -> tuple[dict, dict]:
    """Return ({month: kwh}, {month: interval_count}) summing all
    generation-bearing rows within the target year only."""
    by_month_kwh = defaultdict(float)
    by_month_n = defaultdict(int)
    reader = csv.DictReader(io.StringIO(csv_text))
    for row in reader:
        ds = row.get("IntervalStart", "")
        if len(ds) < 10 or not ds.startswith(f"{year}-"):
            continue
        try:
            month = int(ds[5:7])
            q = float(row.get("Quantity", 0))
        except (ValueError, TypeError):
            continue
        by_month_kwh[month] += q
        by_month_n[month] += 1
    return dict(by_month_kwh), dict(by_month_n)


def main():
    token = load_token()
    accounts = load_accounts()

    print(f"\n=== GMP CSV Scrape — {len(accounts)} accounts ===\n")
    # array_name → month → kwh
    per_array = defaultdict(lambda: defaultdict(float))
    per_account_log = []  # for the detail sheet

    for acct_num, nickname in accounts:
        array = NICK_TO_ARR.get(nickname)
        if array is None:
            print(f"  SKIP {nickname} (unmapped)")
            continue
        print(f"  Fetching {nickname:<20} (#{acct_num}) ...", flush=True)
        csv_text = fetch_csv(acct_num, token)
        if not csv_text:
            print(f"    FAILED — no data returned")
            per_account_log.append({
                "nickname": nickname, "account": acct_num, "array": array,
                "status": "FAILED", "months": {},
            })
            continue

        by_month, by_count = parse_csv_by_month(csv_text)
        total = sum(by_month.values())
        intervals = sum(by_count.values())
        print(f"    → {intervals:>5,} intervals, {total:>10,.0f} kWh in 2026")
        for m, kwh in by_month.items():
            per_array[array][m] += kwh
        per_account_log.append({
            "nickname": nickname, "account": acct_num, "array": array,
            "status": "OK", "months": by_month, "intervals": intervals,
            "total": total,
        })

    # Build output workbook
    print(f"\nWriting {OUT_XLSX} ...")
    wb = Workbook()
    sh = wb.active
    sh.title = "2026 kWh Monthly"

    HDR = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill("solid", fgColor="2E6B3A")
    TOTAL_FILL = PatternFill("solid", fgColor="EEF3EC")
    TOTAL_FONT = Font(bold=True)
    BORDER = Border(*[Side(style="thin", color="C8D4C4")] * 4)

    sh["A1"] = "Green Mountain Community Solar — 2026 Monthly kWh"
    sh["A1"].font = Font(bold=True, size=14, color="2E6B3A")
    sh.merge_cells("A1:O1")
    sh["A2"] = (f"Source: GMP CSV API (15-minute interval generation, summed) · "
                f"Generated {datetime.now():%B %d, %Y %I:%M %p}")
    sh["A2"].font = Font(italic=True, size=10, color="666666")
    sh.merge_cells("A2:O2")

    sh["A4"] = "Array"; sh["B4"] = "Region"
    for i, m in enumerate(MONTHS):
        sh.cell(4, 3 + i, m)
    sh.cell(4, 15, "YTD")
    for col in range(1, 16):
        c = sh.cell(4, col)
        c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    totals = [0.0] * 12
    for r_idx, arr in enumerate(ARRAYS):
        row = 5 + r_idx
        sh.cell(row, 1, arr).font = Font(bold=True)
        sh.cell(row, 2, REGIONS[arr])
        ytd = 0.0
        for m in range(1, 13):
            v = per_array[arr].get(m, 0)
            if v:
                sh.cell(row, 2 + m, round(v))
                ytd += v
                totals[m - 1] += v
            cell = sh.cell(row, 2 + m)
            cell.number_format = "#,##0"
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
        c = sh.cell(row, 15, round(ytd))
        c.number_format = "#,##0"; c.font = Font(bold=True); c.border = BORDER
        c.alignment = Alignment(horizontal="right")
        for cc in range(1, 16):
            sh.cell(row, cc).border = BORDER

    trow = 5 + len(ARRAYS)
    sh.cell(trow, 1, "TOTAL").font = TOTAL_FONT
    for m in range(12):
        c = sh.cell(trow, 3 + m, round(totals[m]))
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL
        c.number_format = "#,##0"; c.border = BORDER
        c.alignment = Alignment(horizontal="right")
    ytd_total = sum(totals)
    yc = sh.cell(trow, 15, round(ytd_total))
    yc.font = TOTAL_FONT; yc.fill = TOTAL_FILL
    yc.number_format = "#,##0"; yc.border = BORDER
    yc.alignment = Alignment(horizontal="right")
    sh.cell(trow, 1).fill = TOTAL_FILL; sh.cell(trow, 1).border = BORDER
    sh.cell(trow, 2).fill = TOTAL_FILL; sh.cell(trow, 2).border = BORDER

    sh.column_dimensions["A"].width = 18
    sh.column_dimensions["B"].width = 10
    for i in range(3, 16):
        sh.column_dimensions[get_column_letter(i)].width = 10

    # Detail sheet — per-account, per-month
    sh2 = wb.create_sheet("Per-Account Detail")
    sh2["A1"] = "Per-account monthly kWh (Starlake sub-accounts shown individually)"
    sh2["A1"].font = Font(bold=True, size=12, color="2E6B3A")
    sh2.merge_cells("A1:O1")
    sh2.cell(3, 1, "Nickname"); sh2.cell(3, 2, "Account #")
    sh2.cell(3, 3, "Array"); sh2.cell(3, 4, "Status")
    for i, m in enumerate(MONTHS):
        sh2.cell(3, 5 + i, m)
    sh2.cell(3, 17, "YTD")
    sh2.cell(3, 18, "Intervals")
    for col in range(1, 19):
        c = sh2.cell(3, col)
        c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for i, log in enumerate(per_account_log):
        row = 4 + i
        sh2.cell(row, 1, log["nickname"])
        sh2.cell(row, 2, log["account"])
        sh2.cell(row, 3, log["array"])
        sh2.cell(row, 4, log["status"])
        if log["status"] == "OK":
            ytd_acct = 0
            for m in range(1, 13):
                v = log["months"].get(m, 0)
                if v:
                    sh2.cell(row, 4 + m, round(v))
                    sh2.cell(row, 4 + m).number_format = "#,##0"
                    ytd_acct += v
            sh2.cell(row, 17, round(ytd_acct)).number_format = "#,##0"
            sh2.cell(row, 18, log.get("intervals", 0))
        for cc in range(1, 19):
            sh2.cell(row, cc).border = BORDER
            if cc >= 5:
                sh2.cell(row, cc).alignment = Alignment(horizontal="right")

    for col, w in [(1, 18), (2, 12), (3, 14), (4, 10)]:
        sh2.column_dimensions[get_column_letter(col)].width = w
    for col in range(5, 19):
        sh2.column_dimensions[get_column_letter(col)].width = 9

    Path(OUT_XLSX).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}\n")

    # Console summary
    print(f"\n=== 2026 SUMMARY ===")
    print(f"YTD total: {ytd_total:,.0f} kWh\n")
    print(f"  {'Array':<16}  " + "  ".join(f"{m:>7}" for m in MONTHS) + f"  {'YTD':>9}")
    for arr in ARRAYS:
        cells = [f"{per_array[arr].get(m, 0):>7,.0f}" for m in range(1, 13)]
        ytd = sum(per_array[arr].values())
        print(f"  {arr:<16}  " + "  ".join(cells) + f"  {ytd:>9,.0f}")
    cells = [f"{totals[m-1]:>7,.0f}" for m in range(1, 13)]
    print(f"  {'TOTAL':<16}  " + "  ".join(cells) + f"  {ytd_total:>9,.0f}")


if __name__ == "__main__":
    main()
