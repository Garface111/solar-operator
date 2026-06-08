"""
Scrape 2026 monthly kWh from GMP using the JSON bills endpoint.

  GET https://api.greenmountainpower.com/api/v2/accounts/{acct}/bills
  → returns ALL historical bills as JSON, each with a 'KWH GENERATE'
    segmentLineItem holding total kWh for that billing period.

Bruce's month-attribution rule:
  - 5 arrays: bill represents PRIOR month → use segment.startDate.month
  - Starlake: bill represents SAME month → use bill.billDate.month
"""
import json
import sqlite3
import sys
import time
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DB = "/home/fordface/solar-operator/storage/solar.db"
OUT_XLSX = "/mnt/c/Users/fordg/Desktop/GMS/2026 Monthly kWh - GMP CSV Scrape.xlsx"
YEAR = 2026

NICK_TO_ARR = {
    "Tannery Brook": "Tannery Brook", "Chester": "Chester",
    "Timberworks": "Timberworks", "Waterford": "Waterford",
    "Londonderry": "Londonderry",
    "Starlake North": "Starlake", "Starlake South": "Starlake",
    "Starlake Center": "Starlake",
}
REGIONS = {"Tannery Brook": "north", "Chester": "south", "Timberworks": "north",
           "Waterford": "north", "Londonderry": "south", "Starlake": "central"}
ARRAYS = ["Tannery Brook", "Chester", "Timberworks", "Waterford",
          "Londonderry", "Starlake"]
MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def load_token():
    con = sqlite3.connect(DB)
    tok = con.execute(
        "SELECT api_token FROM utility_sessions ORDER BY id DESC LIMIT 1"
    ).fetchone()[0]
    con.close()
    return tok


def load_accounts():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        "SELECT account_number, nickname FROM utility_accounts ORDER BY id"
    ).fetchall()
    con.close()
    return [(r["account_number"], r["nickname"]) for r in rows]


def fetch_bills_json(account: str, token: str):
    """Returns list of bill dicts for an account, or [] on failure."""
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://greenmountainpower.com",
        "Referer": "https://greenmountainpower.com/",
        "GMP-Source": "web",
        "User-Agent": "Mozilla/5.0",
    }
    url = f"https://api.greenmountainpower.com/api/v2/accounts/{account}/bills"
    for attempt in range(3):
        try:
            with httpx.Client(timeout=60, headers=headers) as c:
                r = c.get(url)
            if r.status_code == 200:
                return r.json()
            print(f"  [{account}] HTTP {r.status_code}", file=sys.stderr)
        except Exception as e:
            print(f"  [{account}] err: {e}", file=sys.stderr)
        time.sleep(2)
    return []


def extract_kwh_generated(bill: dict) -> float | None:
    """Find the largest non-zero 'KWH GENERATE' line item value. Each bill has
    a placeholder 0.0 GENERATE row plus the real total; some have duplicate
    rows (one for generation, one for the solar incentive credit, both same
    value). Return the max — both duplicates give the same answer."""
    best = 0.0
    for seg in bill.get("billSegments", []):
        for li in seg.get("segmentLineItems", []):
            if (li.get("unitOfMeasure") == "KWH"
                    and li.get("unitCode") == "GENERATE"):
                v = float(li.get("unitCount") or 0)
                if v > best:
                    best = v
    return best if best > 0 else None


def segment_start_month(bill: dict) -> int | None:
    """Return the month (1-12) of the bill's usage period start."""
    for seg in bill.get("billSegments", []):
        sd = seg.get("segmentCalcs", [{}])[0].get("startDate")
        if sd:
            try:
                return datetime.fromisoformat(sd).month
            except Exception:
                pass
    return None


def main():
    token = load_token()
    accounts = load_accounts()
    print(f"\n=== GMP JSON Bills Scrape — {len(accounts)} accounts, year={YEAR} ===\n")

    per_array = defaultdict(lambda: defaultdict(float))
    per_account_log = []
    t_total = time.time()

    for acct_num, nickname in accounts:
        array = NICK_TO_ARR.get(nickname)
        if array is None:
            print(f"  SKIP {nickname} (unmapped)")
            continue
        t0 = time.time()
        bills = fetch_bills_json(acct_num, token)
        if not bills:
            per_account_log.append(
                {"nickname": nickname, "account": acct_num, "array": array,
                 "status": "FAILED", "months": {}, "bill_count": 0})
            print(f"  {nickname:<20} FAILED")
            continue

        # Filter: include any bill whose ATTRIBUTED month falls in YEAR.
        # This means we'll grab some bills with billDate in late YEAR-1 (their
        # production is Jan YEAR) and skip the first bill of YEAR+1 if it
        # represents Dec YEAR's production (we want it — but bill_date.year
        # will be YEAR+1, not YEAR, so we need to scan an extended window).
        log_months = defaultdict(float)
        bills_used = 0
        bill_details = []
        for b in bills:
            bd_str = b.get("billDate", "")
            if not bd_str:
                continue
            try:
                bd = datetime.fromisoformat(bd_str)
            except ValueError:
                continue
            # Skim adjacent years so we don't miss the bill that represents
            # December YEAR (its billDate is in YEAR+1) or January YEAR
            # (bill arrives in early YEAR but represents Dec YEAR-1 — skipped
            # because target_month would be 12 in the wrong year).
            if bd.year not in (YEAR, YEAR + 1):
                continue
            kwh = extract_kwh_generated(b)
            if kwh is None or kwh <= 0:
                continue

            seg_month = segment_start_month(b)
            # Apply Bruce's rule
            if array == "Starlake":
                target_month = bd.month
                target_year = bd.year
            else:
                target_month = seg_month or bd.month
                # segment_start_month gives the month; we need the year too.
                # Reconstruct from segment startDate.
                tgt_year = bd.year
                for seg in b.get("billSegments", []):
                    sd = seg.get("segmentCalcs", [{}])[0].get("startDate")
                    if sd:
                        try:
                            tgt_year = datetime.fromisoformat(sd).year
                            break
                        except Exception:
                            pass
                target_year = tgt_year

            # Only count bills attributed to YEAR
            if target_year != YEAR:
                continue

            per_array[array][target_month] += kwh
            log_months[target_month] += kwh
            bills_used += 1
            bill_details.append({
                "bill_date": bd_str,
                "period_start": (b.get("billSegments", [{}])[0]
                                   .get("segmentCalcs", [{}])[0].get("startDate")),
                "period_end":   (b.get("billSegments", [{}])[0]
                                   .get("segmentCalcs", [{}])[0].get("endDate")),
                "kwh": kwh,
                "attributed_to": target_month,
            })

        elapsed = time.time() - t0
        total = sum(log_months.values())
        print(f"  {nickname:<20}  {bills_used} bills  {total:>10,.0f} kWh   ({elapsed:.2f}s)")
        per_account_log.append({
            "nickname": nickname, "account": acct_num, "array": array,
            "status": "OK", "months": dict(log_months),
            "bill_count": bills_used, "bills": bill_details,
        })

    print(f"\nTotal scrape time: {time.time() - t_total:.2f}s")

    # ─── Build workbook ─────────────────────────────────────────
    print(f"\nWriting {OUT_XLSX} ...")
    wb = Workbook()
    sh = wb.active
    sh.title = f"{YEAR} kWh Monthly"

    HDR = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill("solid", fgColor="2E6B3A")
    TOTAL_FILL = PatternFill("solid", fgColor="EEF3EC")
    TOTAL_FONT = Font(bold=True)
    BORDER = Border(*[Side(style="thin", color="C8D4C4")] * 4)

    sh["A1"] = "Green Mountain Community Solar — 2026 Monthly kWh"
    sh["A1"].font = Font(bold=True, size=14, color="2E6B3A")
    sh.merge_cells("A1:O1")
    sh["A2"] = (f"Source: GMP JSON bills API (KWH GENERATE line item per bill) · "
                f"Generated {datetime.now():%B %d, %Y %I:%M %p}")
    sh["A2"].font = Font(italic=True, size=10, color="666666")
    sh.merge_cells("A2:O2")
    sh["A3"] = ("Month attribution: 5 arrays use bill segment startDate.month · "
                "Starlake uses billDate.month (per GMP meter-read offset rule)")
    sh["A3"].font = Font(italic=True, size=9, color="888888")
    sh.merge_cells("A3:O3")

    sh.cell(5, 1, "Array"); sh.cell(5, 2, "Region")
    for i, m in enumerate(MONTHS):
        sh.cell(5, 3 + i, m)
    sh.cell(5, 15, "YTD")
    for col in range(1, 16):
        c = sh.cell(5, col)
        c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    totals = [0.0] * 12
    for r_idx, arr in enumerate(ARRAYS):
        row = 6 + r_idx
        sh.cell(row, 1, arr).font = Font(bold=True)
        sh.cell(row, 2, REGIONS[arr])
        ytd = 0.0
        for m in range(1, 13):
            v = per_array[arr].get(m, 0)
            if v:
                sh.cell(row, 2 + m, round(v))
                ytd += v
                totals[m - 1] += v
            cell = sh.cell(row, 2 + m)
            cell.number_format = "#,##0"
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")
        c = sh.cell(row, 15, round(ytd))
        c.number_format = "#,##0"; c.font = Font(bold=True); c.border = BORDER
        c.alignment = Alignment(horizontal="right")
        for cc in (1, 2):
            sh.cell(row, cc).border = BORDER

    trow = 6 + len(ARRAYS)
    sh.cell(trow, 1, "TOTAL").font = TOTAL_FONT
    for m in range(12):
        c = sh.cell(trow, 3 + m, round(totals[m]))
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL
        c.number_format = "#,##0"; c.border = BORDER
        c.alignment = Alignment(horizontal="right")
    yc = sh.cell(trow, 15, round(sum(totals)))
    yc.font = TOTAL_FONT; yc.fill = TOTAL_FILL
    yc.number_format = "#,##0"; yc.border = BORDER
    yc.alignment = Alignment(horizontal="right")
    for cc in (1, 2):
        sh.cell(trow, cc).fill = TOTAL_FILL
        sh.cell(trow, cc).border = BORDER

    sh.column_dimensions["A"].width = 18
    sh.column_dimensions["B"].width = 10
    for i in range(3, 16):
        sh.column_dimensions[get_column_letter(i)].width = 10

    # ─── Per-account sheet ───────────────────────────────────────
    sh2 = wb.create_sheet("Per-Account Detail")
    sh2["A1"] = "Per-account monthly kWh (Starlake sub-accounts shown individually)"
    sh2["A1"].font = Font(bold=True, size=12, color="2E6B3A")
    sh2.merge_cells("A1:R1")
    sh2.cell(3, 1, "Nickname"); sh2.cell(3, 2, "Account #")
    sh2.cell(3, 3, "Array"); sh2.cell(3, 4, "Bills")
    for i, m in enumerate(MONTHS):
        sh2.cell(3, 5 + i, m)
    sh2.cell(3, 17, "YTD")
    for col in range(1, 18):
        c = sh2.cell(3, col)
        c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for i, log in enumerate(per_account_log):
        row = 4 + i
        sh2.cell(row, 1, log["nickname"])
        sh2.cell(row, 2, log["account"])
        sh2.cell(row, 3, log["array"])
        sh2.cell(row, 4, log["bill_count"])
        ytd = 0
        for m in range(1, 13):
            v = log["months"].get(m, 0)
            if v:
                sh2.cell(row, 4 + m, round(v)).number_format = "#,##0"
                ytd += v
        sh2.cell(row, 17, round(ytd)).number_format = "#,##0"
        for cc in range(1, 18):
            sh2.cell(row, cc).border = BORDER
            if cc >= 5:
                sh2.cell(row, cc).alignment = Alignment(horizontal="right")

    for col, w in [(1, 18), (2, 12), (3, 14), (4, 7)]:
        sh2.column_dimensions[get_column_letter(col)].width = w
    for col in range(5, 18):
        sh2.column_dimensions[get_column_letter(col)].width = 9

    # ─── Bill-level audit sheet ──────────────────────────────────
    sh3 = wb.create_sheet("Bill Audit")
    sh3["A1"] = "Bill-level data (one row per GMP bill in 2026)"
    sh3["A1"].font = Font(bold=True, size=12, color="2E6B3A")
    sh3.merge_cells("A1:G1")
    headers = ["Nickname", "Array", "Bill Date", "Period Start",
               "Period End", "kWh Generated", "Attributed to"]
    for i, h in enumerate(headers):
        c = sh3.cell(3, i + 1, h)
        c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    r = 4
    for log in per_account_log:
        for b in log.get("bills", []):
            sh3.cell(r, 1, log["nickname"])
            sh3.cell(r, 2, log["array"])
            sh3.cell(r, 3, b["bill_date"])
            sh3.cell(r, 4, b["period_start"] or "")
            sh3.cell(r, 5, b["period_end"] or "")
            sh3.cell(r, 6, round(b["kwh"])).number_format = "#,##0"
            sh3.cell(r, 7, MONTHS[b["attributed_to"] - 1])
            for cc in range(1, 8):
                sh3.cell(r, cc).border = BORDER
            r += 1
    for col, w in [(1, 18), (2, 14), (3, 12), (4, 12), (5, 12), (6, 14), (7, 12)]:
        sh3.column_dimensions[get_column_letter(col)].width = w

    Path(OUT_XLSX).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Saved: {OUT_XLSX}\n")

    # Console summary
    print(f"=== {YEAR} SUMMARY ===")
    print(f"YTD total: {sum(totals):,.0f} kWh\n")
    print(f"  {'Array':<16}  " + "  ".join(f"{m:>7}" for m in MONTHS) + f"  {'YTD':>9}")
    for arr in ARRAYS:
        cells = [f"{per_array[arr].get(m, 0):>7,.0f}" for m in range(1, 13)]
        ytd = sum(per_array[arr].values())
        print(f"  {arr:<16}  " + "  ".join(cells) + f"  {ytd:>9,.0f}")
    cells = [f"{totals[m - 1]:>7,.0f}" for m in range(1, 13)]
    print(f"  {'TOTAL':<16}  " + "  ".join(cells) + f"  {sum(totals):>9,.0f}")


if __name__ == "__main__":
    main()
