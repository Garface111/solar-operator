#!/usr/bin/env python3
"""Solar Operator — Trust Distribution Budget workbook.
Honest accounting: documented/receipt-backed costs kept separate from
clearly-flagged forward-looking estimates. Reconciles to the $25,000 ask.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/home/fordface/solar-operator/SolarOperator_Trust_Distribution_Budget.xlsx"

# ---- palette ---------------------------------------------------------------
DARK   = "1F4D2B"   # deep green title band
MID    = "2F9E44"   # header fill
LIGHT  = "EAF3EC"   # alt row
GOLD   = "B8860B"
DOCFILL= "DDEFD9"   # documented (solid) rows
ESTFILL= "FBF4E0"   # estimate rows
WHITE  = "FFFFFF"
GREY   = "6B6B6B"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def money(c): 
    c.number_format = '$#,##0.00'
def pct(c):
    c.number_format = '0%'

def title_band(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1,1,text)
    t.font = Font(bold=True, size=15, color=WHITE)
    t.fill = PatternFill("solid", fgColor=DARK)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(2,1,sub)
    s.font = Font(italic=True, size=9, color=GREY)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 26

def header_row(ws, r, headers):
    for j,h in enumerate(headers, start=1):
        c = ws.cell(r, j, h)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=MID)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        c.border = border
    ws.row_dimensions[r].height = 24

wb = openpyxl.Workbook()

# ===========================================================================
# SHEET 1 — Distribution Request Summary
# ===========================================================================
ws = wb.active
ws.title = "Request Summary"
ncols = 4
title_band(ws, "Solar Operator — Trust Distribution Request",
           "Prepared for trustee review · One-time $25,000 distribution · Figures as of June 2026 · "
           "STATUS column distinguishes receipt-backed (Documented) costs from forward Estimates.", ncols)

r = 4
ws.cell(r,1,"Requested distribution").font = Font(bold=True, size=11)
v = ws.cell(r,3,25000); money(v); v.font = Font(bold=True, size=12, color=DARK)
r += 2

header_row(ws, r, ["Category", "Basis", "Amount", "Status"]); r += 1
summary = [
    ("AI compute — development to date", "Provider console / receipts", 11846.70, "Documented"),
    ("AI compute — remaining build push", "Forward estimate", 7000.00, "Estimate"),
    ("Hosting & infrastructure (≈12 mo)", "Railway, email, auth, data integrations", 2400.00, "Estimate"),
    ("Listing & compliance", "Chrome Web Store, registration, misc.", 750.00, "Estimate"),
    ("Contingency / buffer to round to ask", "Variance on estimates above", 3003.30, "Estimate"),
]
start = r
for name, basis, amt, status in summary:
    fill = DOCFILL if status == "Documented" else ESTFILL
    ws.cell(r,1,name)
    ws.cell(r,2,basis)
    money(ws.cell(r,3,amt))
    st = ws.cell(r,4,status)
    st.font = Font(bold=True, color=(DARK if status=="Documented" else GOLD), size=9)
    for j in range(1,5):
        cc = ws.cell(r,j); cc.fill = PatternFill("solid", fgColor=fill); cc.border = border
        cc.alignment = Alignment(vertical="center", indent=1, wrap_text=(j==2))
    r += 1
# total
ws.cell(r,1,"TOTAL — distribution requested").font = Font(bold=True, size=11)
tv = ws.cell(r,3, f"=SUM(C{start}:C{r-1})"); money(tv); tv.font = Font(bold=True, size=11, color=WHITE)
for j in range(1,5):
    ws.cell(r,j).fill = PatternFill("solid", fgColor=DARK)
    ws.cell(r,j).font = Font(bold=True, color=WHITE)
    ws.cell(r,j).border = border
r += 2

note = ("Honesty note: $11,846.70 is documented and receipt-backed from the model-provider console. "
        "The four Estimate lines are forward projections, not invoices — they are Ford's good-faith "
        "planning numbers and may shift. Itemized forward estimates total $10,150.00; the $3,003.30 "
        "buffer absorbs estimate variance and rounds the request to the $25,000 figure. Backup detail "
        "is on the 'Costs to Date' and 'Use of Funds' tabs.")
ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=ncols)
nc = ws.cell(r,1,note)
nc.font = Font(italic=True, size=9, color=GREY)
nc.alignment = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 34
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 14
ws.freeze_panes = "A4"

# ===========================================================================
# SHEET 2 — Costs to Date (Documented)
# ===========================================================================
ws2 = wb.create_sheet("Costs to Date")
ncols = 4
title_band(ws2, "Costs to Date — Documented",
           "Receipt-backed development-phase spending already incurred. "
           "Rows marked 'Confirm' are real costs Ford should verify exact figures for before sending.", ncols)
r = 4
header_row(ws2, r, ["Item", "Source of figure", "Amount", "Status"]); r += 1
to_date = [
    ("AI / model-API compute (Stage 1 development)", "Provider console export — exact", 11846.70, "Documented"),
    ("Domain — solaroperator.org", "Registrar receipt", None, "Confirm"),
    ("Railway hosting — spend to date", "Railway billing dashboard", None, "Confirm"),
    ("Chrome Web Store developer fee (one-time)", "$5.00 standard", 5.00, "Confirm"),
    ("Stripe processing", "Per-transaction % — no upfront cost", 0.00, "Documented"),
]
start = r
for name, src, amt, status in to_date:
    fill = DOCFILL if status == "Documented" else ESTFILL
    ws2.cell(r,1,name); ws2.cell(r,2,src)
    if amt is None:
        c = ws2.cell(r,3,"— fill in —"); c.font = Font(italic=True, color="B00020")
    else:
        money(ws2.cell(r,3,amt))
    st = ws2.cell(r,4,status); st.font = Font(bold=True, size=9,
        color=(DARK if status=="Documented" else "B00020"))
    for j in range(1,5):
        cc = ws2.cell(r,j); cc.fill = PatternFill("solid", fgColor=fill); cc.border = border
        cc.alignment = Alignment(vertical="center", indent=1, wrap_text=(j in (1,2)))
    r += 1
ws2.cell(r,1,"Documented subtotal (excl. 'fill in' rows)").font = Font(bold=True)
sv = ws2.cell(r,3, f"=SUM(C{start}:C{r-1})"); money(sv); sv.font = Font(bold=True, color=DARK)
for j in range(1,5):
    ws2.cell(r,j).fill = PatternFill("solid", fgColor=LIGHT); ws2.cell(r,j).border = border
r += 2
ws2.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=ncols)
nc = ws2.cell(r,1,"The $11,846.70 compute figure is the headline documented cost and the reason the "
    "request is timed now. The 'Confirm' rows are small; fill in exact registrar/Railway amounts from "
    "your receipts before sending so every number on the sheet is defensible.")
nc.font = Font(italic=True, size=9, color=GREY); nc.alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 34
ws2.column_dimensions['C'].width = 15
ws2.column_dimensions['D'].width = 13
ws2.freeze_panes = "A5"

# ===========================================================================
# SHEET 3 — Use of Funds (Forward Budget)
# ===========================================================================
ws3 = wb.create_sheet("Use of Funds")
ncols = 4
title_band(ws3, "Use of Funds — Remaining Build Phase (Estimates)",
           "Forward-looking, good-faith estimates for closing out the build phase. "
           "Not invoices. These are planning figures and may shift.", ncols)
r = 4
header_row(ws3, r, ["Line item", "What it covers", "Estimate", "Basis / assumption"]); r += 1
fwd = [
    ("AI compute — remaining dev push", "Agent-swarm work to add utilities for v1 (New England) coverage",
     7000.00, "Below Stage-1 actual; utility onboarding is the remaining work"),
    ("Hosting & infrastructure (~12 mo)", "Railway containers, email, auth, data integrations",
     2400.00, "≈ $200/mo run-rate × 12"),
    ("Listing & compliance", "Chrome Web Store, business registration, misc. compliance",
     750.00, "Round allowance — low-cost items"),
    ("Contingency / buffer", "Estimate variance; rounds request to $25,000 ask",
     3003.30, "Plug between itemized estimates and the round ask"),
]
start = r
for name, cov, amt, basis in fwd:
    ws3.cell(r,1,name); ws3.cell(r,2,cov); money(ws3.cell(r,3,amt)); ws3.cell(r,4,basis)
    for j in range(1,5):
        cc = ws3.cell(r,j); cc.fill = PatternFill("solid", fgColor=ESTFILL); cc.border = border
        cc.alignment = Alignment(vertical="center", indent=1, wrap_text=(j in (2,4)))
    r += 1
ws3.cell(r,1,"Forward budget subtotal").font = Font(bold=True)
sv = ws3.cell(r,3, f"=SUM(C{start}:C{r-1})"); money(sv); sv.font = Font(bold=True, color=GOLD)
for j in range(1,5):
    ws3.cell(r,j).fill = PatternFill("solid", fgColor=LIGHT); ws3.cell(r,j).border = border
r += 2
ws3.merge_cells(start_row=r, start_column=1, end_row=r+2, end_column=ncols)
nc = ws3.cell(r,1,"Commitment: this is a one-time disbursement that closes out the build phase. "
    "Development-phase compute is front-loaded; once the product settles, ongoing costs drop to a "
    "fraction of these figures and live Stripe billing is already in place to cover them.")
nc.font = Font(italic=True, size=9, color=GREY); nc.alignment = Alignment(wrap_text=True, vertical="top")
ws3.column_dimensions['A'].width = 34
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 42
ws3.freeze_panes = "A5"

# ===========================================================================
# SHEET 4 — Path to Self-Sufficiency (illustrative)
# ===========================================================================
ws4 = wb.create_sheet("Path to Self-Sufficiency")
ncols = 4
title_band(ws4, "Path to Self-Sufficiency — Illustrative",
           "ILLUSTRATIVE unit economics, not a forecast or a promise. Shows how live pricing covers "
           "ongoing run-rate. Pricing is real; customer counts are scenarios.", ncols)
r = 4
ws4.cell(r,1,"Live pricing (graduated, per array / per quarter)").font = Font(bold=True, size=10)
r += 1
header_row(ws4, r, ["Array band", "Price per array", "Notes", ""]); r += 1
pricing = [
    ("Arrays 1–50", 15.00, "First band"),
    ("Arrays 51–100", 13.50, "Volume step"),
    ("Arrays 101–150", 12.00, "Volume step"),
    ("Arrays 151+", 10.50, "Volume step"),
]
for band, price, notes in pricing:
    ws4.cell(r,1,band); money(ws4.cell(r,2,price)); ws4.cell(r,3,notes)
    for j in range(1,4):
        cc = ws4.cell(r,j); cc.fill = PatternFill("solid", fgColor=LIGHT); cc.border = border
        cc.alignment = Alignment(vertical="center", indent=1)
    r += 1
r += 1
ws4.cell(r,1,"Illustrative break-even vs. settled run-rate").font = Font(bold=True, size=10)
r += 1
ws4.cell(r,1,"Assumed settled ongoing cost (~$200/mo hosting + light compute)").font = Font(size=9, italic=True)
br = ws4.cell(r,3,200); money(br)
r += 1
header_row(ws4, r, ["Scenario (billed arrays)", "Implied quarterly revenue", "≈ Monthly", "Covers run-rate?"]); r += 1
# illustrative: revenue at first-band $15 for simplicity, clearly labeled
scenarios = [(20,), (50,), (120,), (300,)]
def quarterly_rev(n):
    # graduated calc
    bands = [(50,15.0),(50,13.5),(50,12.0),(10**9,10.5)]
    rem = n; tot = 0.0
    for cnt, pr in bands:
        take = min(rem, cnt); tot += take*pr; rem -= take
        if rem <= 0: break
    return tot
start = r
for (n,) in scenarios:
    q = quarterly_rev(n); m = q/3.0
    ws4.cell(r,1,f"{n} arrays"); money(ws4.cell(r,2,q)); money(ws4.cell(r,3,m))
    covers = "Yes" if m >= 200 else "Approaching"
    cc = ws4.cell(r,4,covers); cc.font = Font(bold=True, color=(DARK if covers=="Yes" else GOLD), size=9)
    for j in range(1,5):
        c2 = ws4.cell(r,j); c2.fill = PatternFill("solid", fgColor=ESTFILL); c2.border = border
        c2.alignment = Alignment(vertical="center", indent=1)
    r += 1
r += 1
ws4.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=ncols)
nc = ws4.cell(r,1,"Read this as orders of magnitude, not a forecast. Pricing is live and real; the array "
    "counts are scenarios to show the shape of the economics. Quarterly revenue uses the graduated bands. "
    "The point: ongoing run-rate after the build phase is modest, and live billing is already wired to "
    "cover it — which is why this is a one-time request, not the start of a pattern.")
nc.font = Font(italic=True, size=9, color=GREY); nc.alignment = Alignment(wrap_text=True, vertical="top")
ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 26
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 18

wb.save(OUT)
print("WROTE", OUT)
print("Sheets:", wb.sheetnames)
